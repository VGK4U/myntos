"""
Staff Progress/Performance Dashboard API Endpoints (DC Protocol Compliant)
Aggregates data from multiple sources for staff performance tracking

Key Features:
- Tasks: Planned/Completed/Pending/Overdue counts
- KRA: Daily instance tracking with status
- Attendance: Login/Logout times, working hours, break hours
- Leads: For Sales/CRM departments - handled/new/won/lost counts
- Service Tickets: For Service departments - ticket metrics
- Travel: Journey distance and time for field staff

Access Control:
- Staff: See only own data
- Managers/Key Leadership: See downline team members via filter

Created: Jan 2026
DC Protocol: Write-Verify-Validate at all levels
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, func, case, text, Date as SQLDate
from typing import Optional, List
import time
import json
import os
from datetime import datetime, date, timedelta
from decimal import Decimal
import pytz
import logging
import io

from app.core.database import get_db
from app.models.staff import StaffEmployee, StaffRole
from app.models.staff_tasks import StaffTask, StaffTaskAssignee, StaffTaskPhase, StaffTaskActivityLog, StaffDayPlan, StaffDayPlanItem
from app.models.staff_attendance import StaffAttendance, StaffAttendanceBreak
from app.models.staff_kra import StaffKRAAssignment, StaffKRADailyInstance
from app.models.staff_timesheet import StaffTimesheetEntry
from app.models.staff_field_work import StaffFieldWorkSession
from app.models.staff_attendance_sheet import StaffAttendanceSheet
from app.models.crm import CRMLead, CRMLeadTransaction
from app.models.ticket import ServiceTicket
from app.api.v1.endpoints.staff_auth import get_current_staff_user
from app.utils.staff_hierarchy import HIDDEN_FROM_TEAM_CODES

router = APIRouter()
logger = logging.getLogger(__name__)

# DC_RANKING_CACHE: Dual-layer cache — in-memory (per-worker, fast) +
# file-based (/tmp, shared across workers, survives deployment restarts).
# After a publish both workers start cold; file cache prevents the 150-query
# cold-start hang that previously caused socket hang-ups on the first request.
_RANKING_CACHE: dict = {}
_RANKING_CACHE_TTL = 3600  # 60 minutes (DC Protocol: IST-stable, ranking scores change slowly)
_RANKING_CACHE_FILE = '/tmp/mnr_ranking_scores.json'


def _ranking_cache_key_str(cache_key: tuple) -> str:
    return json.dumps(list(cache_key))


def _read_ranking_file_cache(cache_key: tuple):
    """Read from shared file cache. Returns scores dict {str(emp_id): score} or None on miss/expiry.
    DC Protocol: Old format was a list — treat as None and recompute in new dict format."""
    try:
        with open(_RANKING_CACHE_FILE, 'r') as f:
            data = json.load(f)
        key_str = _ranking_cache_key_str(cache_key)
        entry = data.get(key_str)
        if entry and (time.time() - entry.get('ts', 0)) < _RANKING_CACHE_TTL:
            scores = entry['scores']
            if isinstance(scores, dict):
                return scores
    except Exception:
        pass
    return None


def _write_ranking_file_cache(cache_key: tuple, scores: dict):
    """Write to shared file cache atomically. Safe for concurrent workers."""
    try:
        key_str = _ranking_cache_key_str(cache_key)
        try:
            with open(_RANKING_CACHE_FILE, 'r') as f:
                data = json.load(f)
        except Exception:
            data = {}
        data[key_str] = {'scores': scores, 'ts': time.time()}
        tmp_path = _RANKING_CACHE_FILE + '.tmp'
        with open(tmp_path, 'w') as f:
            json.dump(data, f)
        os.replace(tmp_path, _RANKING_CACHE_FILE)  # atomic on POSIX
    except Exception:
        pass  # file cache is best-effort; in-memory cache still applies

IST = pytz.timezone('Asia/Kolkata')


def get_indian_date():
    """Get current date in Indian timezone (IST)"""
    return datetime.now(IST).date()


def get_indian_time():
    """Get current time in Indian timezone (IST)"""
    return datetime.now(IST).replace(tzinfo=None)


def get_downline_employee_ids(db: Session, manager_id: int, recursive: bool = True) -> List[int]:
    """
    Get all employee IDs reporting to a manager
    DC Protocol: Uses reporting_manager_id for hierarchy
    
    Args:
        db: Database session
        manager_id: Manager's employee ID
        recursive: If True, get full tree (all levels). If False, only direct reports.
    
    Returns:
        List of employee IDs in downline
    """
    if not recursive:
        direct_reports = db.query(StaffEmployee.id).filter(
            StaffEmployee.reporting_manager_id == manager_id,
            StaffEmployee.status == 'active'
        ).all()
        return [r.id for r in direct_reports]
    
    all_downline = []
    queue = [manager_id]
    visited = set()
    
    while queue:
        current_id = queue.pop(0)
        if current_id in visited:
            continue
        visited.add(current_id)
        
        direct_reports = db.query(StaffEmployee.id).filter(
            StaffEmployee.reporting_manager_id == current_id,
            StaffEmployee.status == 'active'
        ).all()
        
        for report in direct_reports:
            if report.id not in visited:
                all_downline.append(report.id)
                queue.append(report.id)
    
    return all_downline


def is_manager_or_leadership(employee: StaffEmployee) -> bool:
    """Check if employee has manager/leadership role.
    DC Protocol: role_code is authoritative — role NAME check uses exact match only.
    'Senior Executive' / 'Junior Executive' must NOT qualify."""
    if not employee.role:
        return False
    role_code = (employee.role.role_code if hasattr(employee.role, 'role_code') else str(employee.role)).lower().strip()
    role_name = (employee.role.role_name if hasattr(employee.role, 'role_name') else '').upper().strip()
    leadership_role_codes = {'vgk4u', 'vgk4u_supreme', 'key_leadership', 'manager', 'team_leader', 'hr_manager', 'ea', 'executive_admin'}
    leadership_role_names = {'VGK4U', 'VGK4U SUPREME', 'KEY LEADERSHIP', 'MANAGER', 'TEAM LEADER', 'HR MANAGER', 'EA', 'EXECUTIVE ADMIN'}
    return role_code in leadership_role_codes or 'vgk4u' in role_code or role_name in leadership_role_names


def is_key_leadership(employee: StaffEmployee) -> bool:
    """Check if employee has Key Leadership or higher role (for executive dashboard access).
    DC Protocol: EXACT match only — 'Senior Executive' must NOT pass.
    Qualifying: VGK4U, VGK4U Supreme, Key Leadership, EA, Executive Admin."""
    if not employee.role:
        return False
    role_code = (employee.role.role_code if hasattr(employee.role, 'role_code') else str(employee.role)).lower().strip()
    role_name = (employee.role.role_name if hasattr(employee.role, 'role_name') else '').upper().strip()
    export_role_codes = {'vgk4u', 'vgk4u_supreme', 'key_leadership', 'ea', 'executive_admin'}
    export_role_names = {'VGK4U', 'VGK4U SUPREME', 'KEY LEADERSHIP', 'EA', 'EXECUTIVE ADMIN'}
    return role_code in export_role_codes or 'vgk4u' in role_code or role_name in export_role_names


def has_unrestricted_access(employee: StaffEmployee) -> bool:
    """Check if employee has unrestricted cross-company access (VGK4U Supreme, EA, Key Leadership).
    DC Protocol: role_code is authoritative. Exact match only — no substring checks."""
    if not employee.role:
        return False
    role_code = (employee.role.role_code if hasattr(employee.role, 'role_code') else '').lower().strip()
    role_name = (employee.role.role_name if hasattr(employee.role, 'role_name') else '').upper().strip()
    unrestricted_codes = {'vgk4u', 'vgk4u_supreme', 'key_leadership', 'ea', 'executive_admin'}
    unrestricted_names = {'VGK4U', 'VGK4U SUPREME', 'KEY LEADERSHIP', 'EA', 'EXECUTIVE ADMIN'}
    return role_code in unrestricted_codes or 'vgk4u' in role_code or role_name in unrestricted_names


@router.get("/summary", summary="Get staff progress summary")
def get_progress_summary(
    target_date: Optional[date] = Query(None, description="Date for progress (defaults to today)"),
    date_from: Optional[date] = Query(None, description="Start date for range query"),
    date_to: Optional[date] = Query(None, description="End date for range query"),
    employee_id: Optional[int] = Query(None, description="Employee ID (managers only)"),
    include_aggregates: bool = Query(True, description="Include MTD and rolling-30 aggregates (set false for today-tab fast load)"),
    current_user: StaffEmployee = Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    """
    DC Protocol: Get comprehensive staff progress summary
    
    Note: This endpoint is intentionally synchronous to ensure proper SQLAlchemy
    session handling. FastAPI runs sync endpoints in a thread pool automatically.
    
    - Staff: Own data only
    - Managers: Can filter by downline employee_id
    - Key Leadership: Can view any employee
    
    Returns aggregated data from:
    - Tasks (Planned/Completed/Pending/Overdue)
    - KRA (Daily instances)
    - Attendance (Login/Logout/Working Hours/Breaks)
    - Leads (if applicable department)
    - Service Tickets (if applicable department)
    - Travel/Journeys
    """
    import traceback
    try:
        # DC Protocol: shorter timeout for range queries (more expensive) to avoid pool exhaustion
        _timeout_ms = '18000' if (date_from and date_to) else '22000'
        db.execute(text(f"SET LOCAL statement_timeout = '{_timeout_ms}'"))
        logger.info(f"[DC-PROGRESS] Starting summary for user {current_user.id}")
        today = get_indian_date()
        
        if target_date is None:
            target_date = today
        
        if date_from and date_to:
            query_start = date_from
            query_end = date_to
            is_range_query = True
        else:
            query_start = target_date
            query_end = target_date
            is_range_query = False
        
        target_employee_id = current_user.id
        target_employee = current_user
        
        if employee_id and employee_id != current_user.id:
            if is_manager_or_leadership(current_user):
                if has_unrestricted_access(current_user):
                    target_employee = db.query(StaffEmployee).options(
                        joinedload(StaffEmployee.department),
                        joinedload(StaffEmployee.role)
                    ).filter(
                        StaffEmployee.id == employee_id
                    ).first()
                else:
                    downline_ids = get_downline_employee_ids(db, current_user.id, recursive=True)
                    if employee_id not in downline_ids:
                        raise HTTPException(status_code=403, detail="Employee not in your downline")
                    target_employee = db.query(StaffEmployee).options(
                        joinedload(StaffEmployee.department),
                        joinedload(StaffEmployee.role)
                    ).filter(
                        StaffEmployee.id == employee_id,
                        StaffEmployee.base_company_id == current_user.base_company_id
                    ).first()
                
                if not target_employee:
                    raise HTTPException(status_code=404, detail="Employee not found")
                target_employee_id = employee_id
            else:
                raise HTTPException(status_code=403, detail="Only managers can view other employees")
        
        dept_code = (target_employee.department.department_code or '').lower() if target_employee.department else ''
        dept_name_lower = (target_employee.department.name or '').lower() if target_employee.department else ''
        is_sales_dept = 'sales' in dept_name_lower or 'crm' in dept_name_lower
        is_service_dept = 'service' in dept_name_lower
        is_procurement_dept = 'procurement' in dept_name_lower or 'purchase' in dept_name_lower or 'store' in dept_name_lower
        is_accounts_dept = 'account' in dept_name_lower or ('hr' in dept_name_lower and 'service' not in dept_name_lower)
        dept_type = 'sales' if is_sales_dept else ('service' if is_service_dept else ('procurement' if is_procurement_dept else 'other'))

        # DC Protocol: For managers in non-dept roles (e.g. Management / Key Leadership),
        # detect departments of their direct reports so the dept performance section shows.
        also_manages_depts: list = []
        if dept_type == 'other' and is_manager_or_leadership(current_user):
            _report_ids = get_downline_employee_ids(db, target_employee_id, recursive=False)
            if _report_ids:
                from app.models.staff import StaffDepartment
                _rep_depts = db.query(StaffEmployee.department_id).filter(
                    StaffEmployee.id.in_(_report_ids),
                    StaffEmployee.status == 'active'
                ).all()
                _rep_dept_ids = [r[0] for r in _rep_depts if r[0]]
                if _rep_dept_ids:
                    _dept_names = db.query(StaffDepartment.name).filter(
                        StaffDepartment.id.in_(_rep_dept_ids)
                    ).all()
                    for _dn in _dept_names:
                        _dl = (_dn[0] or '').lower()
                        if ('sales' in _dl or 'crm' in _dl) and 'sales' not in also_manages_depts:
                            also_manages_depts.append('sales')
                        elif 'service' in _dl and 'service' not in also_manages_depts:
                            also_manages_depts.append('service')
                        elif ('procurement' in _dl or 'purchase' in _dl) and 'procurement' not in also_manages_depts:
                            also_manages_depts.append('procurement')

        mtd_start = today.replace(day=1)
        rolling_start = today - timedelta(days=29)
        
        logger.info(f"[DC-PROGRESS] Fetching daily summaries for employee {target_employee_id}")
        tasks_data = get_tasks_summary(db, target_employee_id, query_start, query_end)
        kra_data = get_kra_summary(db, target_employee_id, query_start, query_end)
        attendance_data = get_attendance_summary(db, target_employee_id, query_start, query_end)
        timesheet_data = get_timesheet_summary(db, target_employee_id, query_start, query_end)
        leads_data = get_leads_summary(db, target_employee_id, query_start, query_end)
        service_data = get_service_tickets_summary(db, target_employee_id, query_start, query_end)
        travel_data = get_travel_summary(db, target_employee_id, query_start, query_end)
        day_plan_data = get_day_plan_summary(db, target_employee_id, query_start, query_end)
        day_progress_data = get_day_progress_breakdown(db, target_employee_id, target_date) if not is_range_query else None
        calls_range_data = get_calls_range_summary(db, target_employee_id, query_start, query_end) if is_range_query else None
        procurement_range_data = get_procurement_range_summary(db, target_employee_id, query_start, query_end) if is_range_query and is_procurement_dept else None

        # ── Department KPI (single-day queries only) ───────────────────────────
        dept_kpi: dict = {}
        if not is_range_query:
            if is_sales_dept:
                from app.models.call_tracking import StaffCallLog
                call_date_str = target_date.isoformat()
                _day_logs = db.query(StaffCallLog).filter(
                    StaffCallLog.staff_id == target_employee_id,
                    StaffCallLog.call_date == call_date_str
                ).all()
                def _tt_fmt(s):
                    s = int(s or 0)
                    h, m = s // 3600, (s % 3600) // 60
                    return f"{h}h {m}m"
                talk_secs = sum(l.duration_seconds or 0 for l in _day_logs)
                leads_talk_secs = sum(l.duration_seconds or 0 for l in _day_logs if l.matched_lead_id)
                other_talk_secs = talk_secs - leads_talk_secs
                talk_h = talk_secs // 3600
                talk_m = (talk_secs % 3600) // 60
                # DC_OVERDUE_FIX: Use full OR-based assignment filter (matches Auto Dialer logic).
                # Previously only checked telecaller_id and field_staff_id — now includes
                # handler_id (via emp_code) and primary_owner_id to avoid undercount.
                # Mirrors _crm_assignment_filter() defined in crm.py; cannot import to avoid
                # circular deps. Keep these two in sync if the assignment logic ever changes.
                _target_emp_code = target_employee.emp_code
                crm_overdue = int(db.query(func.count(CRMLead.id)).filter(
                    or_(
                        and_(CRMLead.handler_type == 'staff', CRMLead.handler_id == _target_emp_code),
                        CRMLead.telecaller_id == target_employee_id,
                        CRMLead.field_staff_id == target_employee_id,
                        and_(CRMLead.primary_owner_type == 'staff', CRMLead.primary_owner_id == target_employee_id),
                    ),
                    func.date(CRMLead.next_followup_date) < target_date,
                    CRMLead.status.notin_(['won', 'lost', 'dropped', 'completed'])
                ).scalar() or 0)
                # DC Protocol (Apr 2026): Revenue In Hand = current balance on all assigned won leads
                _target_emp_code2 = target_employee.emp_code
                _rev_in_hand = float(db.query(func.sum(CRMLead.deal_value_balance)).filter(
                    or_(
                        and_(CRMLead.handler_type == 'staff', CRMLead.handler_id == _target_emp_code2),
                        CRMLead.telecaller_id == target_employee_id,
                        CRMLead.field_staff_id == target_employee_id,
                        and_(CRMLead.primary_owner_type == 'staff', CRMLead.primary_owner_id == target_employee_id),
                    ),
                    CRMLead.status == 'won',
                    CRMLead.deal_value_balance > 0,
                ).scalar() or 0)
                dept_kpi = {
                    'talk_time_secs': talk_secs,
                    'talk_time_formatted': f"{talk_h}h {talk_m}m",
                    'leads_talk_secs': leads_talk_secs,
                    'leads_talk_formatted': _tt_fmt(leads_talk_secs),
                    'other_talk_secs': other_talk_secs,
                    'other_talk_formatted': _tt_fmt(other_talk_secs),
                    'leads_handled_today': int(leads_data['summary'].get('handled_today', 0)),
                    'leads_new_today': int(leads_data['summary'].get('new_created', 0)),
                    'self_leads_new_today': int(leads_data['summary'].get('self_leads', {}).get('new', 0)),
                    'company_leads_new_today': int(leads_data['summary'].get('new', 0)),
                    'leads_won_today': int(leads_data['summary'].get('won', 0)),
                    'leads_lost_today': int(leads_data['summary'].get('lost', 0)),
                    'deal_value_today': float(leads_data['summary'].get('deal_value_updated', 0)),
                    'deal_value_received_today': float(leads_data['summary'].get('deal_value_received', 0)),
                    'lost_deal_value_today': float(leads_data['summary'].get('lost_revenue', 0)),
                    'deal_value_in_hand': _rev_in_hand,
                    'overdue_leads': crm_overdue,
                }
            elif is_service_dept:
                svc_sum = service_data['summary']
                total_closed = int(svc_sum.get('resolved', 0))
                within_tat = int(svc_sum.get('within_tat', 0))
                out_of_tat = int(svc_sum.get('out_of_tat', 0))
                tat_pct = round((within_tat / total_closed) * 100) if total_closed > 0 else 0
                dept_kpi = {
                    'tickets_handled': int(svc_sum.get('new', 0)),
                    'tickets_resolved': total_closed,
                    'within_tat_count': within_tat,
                    'within_tat_pct': tat_pct,
                    'above_tat_count': out_of_tat,
                }
            elif is_procurement_dept:
                dept_kpi = get_procurement_kpi_summary(db, target_employee_id, target_date)

        # ── Accounts / HR data (shown in addition to standard sections) ────────
        accounts_data: dict = {}
        if is_accounts_dept or is_manager_or_leadership(current_user):
            try:
                _co_id = target_employee.base_company_id
                _qdate = query_start if not is_range_query else query_start
                # Team attendance: count present today across company
                _att_row = db.execute(text(
                    "SELECT COUNT(*) as total_employees, "
                    " SUM(CASE WHEN sa.clock_in IS NOT NULL THEN 1 ELSE 0 END) as present_count "
                    " FROM staff_employees se "
                    " LEFT JOIN staff_attendance sa ON sa.employee_id = se.id AND sa.date = :qdate "
                    " WHERE se.base_company_id = :co_id AND se.status = 'active'"
                ), {'co_id': _co_id, 'qdate': _qdate}).fetchone()
                _total_emp = int(_att_row[0] or 0) if _att_row else 0
                _present = int(_att_row[1] or 0) if _att_row else 0
                # Income entries summary for period
                _ie_rows = db.execute(text(
                    "SELECT status, COUNT(*) as cnt, COALESCE(SUM(amount),0) as total "
                    " FROM income_entries "
                    " WHERE company_id = :co_id AND income_date BETWEEN :df AND :dt "
                    " GROUP BY status"
                ), {'co_id': _co_id, 'df': query_start, 'dt': query_end}).fetchall()
                _ie_by_status = {r[0].upper(): {'count': int(r[1]), 'amount': float(r[2])} for r in _ie_rows}
                _ie_pending = _ie_by_status.get('PENDING', {'count': 0, 'amount': 0.0})
                _ie_confirmed = _ie_by_status.get('CONFIRMED', {'count': 0, 'amount': 0.0})
                # Employee fund ledger — self balance (last balance entry)
                _efl_row = db.execute(text(
                    "SELECT balance FROM employee_fund_ledger "
                    " WHERE employee_id = :eid ORDER BY id DESC LIMIT 1"
                ), {'eid': target_employee_id}).fetchone()
                _self_balance = float(_efl_row[0]) if _efl_row else None
                # Company account ledger — latest balance
                _cal_row = db.execute(text(
                    "SELECT balance FROM company_account_ledger "
                    " WHERE company_id = :co_id ORDER BY id DESC LIMIT 1"
                ), {'co_id': _co_id}).fetchone()
                _company_balance = float(_cal_row[0]) if _cal_row else None
                # Count employees using company money (have company_account_ledger entries created by them)
                _co_users_row = db.execute(text(
                    "SELECT COUNT(DISTINCT created_by_id) FROM company_account_ledger "
                    " WHERE company_id = :co_id AND transaction_date BETWEEN :df AND :dt"
                ), {'co_id': _co_id, 'df': query_start, 'dt': query_end}).fetchone()
                _co_money_users = int(_co_users_row[0] or 0) if _co_users_row else 0
                accounts_data = {
                    'team_present': _present,
                    'team_total_employees': _total_emp,
                    'team_absent': _total_emp - _present,
                    'income_entries': {
                        'pending_count': _ie_pending['count'],
                        'pending_amount': _ie_pending['amount'],
                        'confirmed_count': _ie_confirmed['count'],
                        'confirmed_amount': _ie_confirmed['amount'],
                        'total_count': _ie_pending['count'] + _ie_confirmed['count'],
                        'total_amount': _ie_pending['amount'] + _ie_confirmed['amount'],
                    },
                    'fund_ledger': {
                        'self_balance': _self_balance,
                        'has_self_ledger': _self_balance is not None,
                    },
                    'company_ledger': {
                        'balance': _company_balance,
                        'users_using_company_money': _co_money_users,
                    },
                }
            except Exception as _ae:
                logger.warning(f"[DC-PROGRESS] accounts_data failed (non-fatal): {_ae}")
                accounts_data = {}

        # ── PO / PR performance (procurement dept only) ────────────────────────
        po_pr_data: dict = {}
        if is_procurement_dept:
            try:
                _co_id2 = target_employee.base_company_id
                # Procurement requests from marketplace_procurement_requests (MTD range, company-wide)
                _pr_df = mtd_start if not is_range_query else query_start
                _pr_rows = db.execute(text(
                    "SELECT status, COUNT(*) as cnt "
                    " FROM marketplace_procurement_requests "
                    " WHERE company_id = :co_id "
                    "   AND DATE(created_at) BETWEEN :df AND :dt "
                    " GROUP BY status"
                ), {'co_id': _co_id2, 'df': _pr_df, 'dt': query_end}).fetchall()
                _pr_by_status = {r[0]: {'count': int(r[1])} for r in _pr_rows}
                _pr_approved = sum(v['count'] for k, v in _pr_by_status.items() if k in ('received',))
                _pr_pending = sum(v['count'] for k, v in _pr_by_status.items() if k in ('pending', 'confirmed', 'ordered'))
                _pr_total = sum(v['count'] for v in _pr_by_status.values())
                # PO summary from marketplace_purchase_orders (MTD range for better coverage)
                _po_df = mtd_start if not is_range_query else query_start
                _po_rows = db.execute(text(
                    "SELECT status, COUNT(*) as cnt, COALESCE(SUM(total_value),0) as total_val "
                    " FROM marketplace_purchase_orders "
                    " WHERE company_id = :co_id "
                    "   AND DATE(created_at) BETWEEN :df AND :dt "
                    " GROUP BY status"
                ), {'co_id': _co_id2, 'df': _po_df, 'dt': query_end}).fetchall()
                _po_by_status = {r[0]: {'count': int(r[1]), 'total_val': float(r[2])} for r in _po_rows}
                _po_confirmed = _po_by_status.get('confirmed', {'count': 0, 'total_val': 0.0})
                _po_dispatched = _po_by_status.get('dispatched', {'count': 0, 'total_val': 0.0})
                _po_total = sum(v['count'] for v in _po_by_status.values())
                _po_total_val = sum(v['total_val'] for v in _po_by_status.values())
                po_pr_data = {
                    'procurement_requests': {
                        'total': _pr_total,
                        'approved': _pr_approved,
                        'pending': _pr_pending,
                        'by_status': _pr_by_status,
                    },
                    'purchase_orders': {
                        'total': _po_total,
                        'confirmed': _po_confirmed['count'],
                        'dispatched': _po_dispatched['count'],
                        'total_value': _po_total_val,
                    },
                }
            except Exception as _pe:
                logger.warning(f"[DC-PROGRESS] po_pr_data failed (non-fatal): {_pe}")
                po_pr_data = {}

        if include_aggregates and not is_range_query:
            logger.info(f"[DC-PROGRESS] Fetching MTD summaries")
            mtd_tasks = get_tasks_summary(db, target_employee_id, mtd_start, today)
            mtd_kra = get_kra_summary(db, target_employee_id, mtd_start, today)
            mtd_leads = get_leads_summary(db, target_employee_id, mtd_start, today)
            mtd_service = get_service_tickets_summary(db, target_employee_id, mtd_start, today)
            mtd_timesheet = get_timesheet_summary(db, target_employee_id, mtd_start, today)

            working_days = max(1, sum(1 for i in range((today - mtd_start).days + 1)
                                      if (mtd_start + timedelta(days=i)).weekday() < 5))

            mtd_dept_kpi: dict = {}
            if is_sales_dept:
                from app.models.call_tracking import StaffCallLog as _CL
                _mtd_talk_secs = int(db.query(func.sum(_CL.duration_seconds)).filter(
                    _CL.staff_id == target_employee_id,
                    _CL.call_date.between(str(mtd_start), str(today))
                ).scalar() or 0)
                _mtd_h = _mtd_talk_secs // 3600
                _mtd_m = (_mtd_talk_secs % 3600) // 60
                _avg_secs = _mtd_talk_secs // working_days
                # DC Protocol (Apr 2026): Rev In Hand = current balance on all assigned won leads (snapshot)
                _mtd_emp_code = target_employee.emp_code
                _mtd_rev_in_hand = float(db.query(func.sum(CRMLead.deal_value_balance)).filter(
                    or_(
                        and_(CRMLead.handler_type == 'staff', CRMLead.handler_id == _mtd_emp_code),
                        CRMLead.telecaller_id == target_employee_id,
                        CRMLead.field_staff_id == target_employee_id,
                        and_(CRMLead.primary_owner_type == 'staff', CRMLead.primary_owner_id == target_employee_id),
                    ),
                    CRMLead.status == 'won',
                    CRMLead.deal_value_balance > 0,
                ).scalar() or 0)
                mtd_dept_kpi = {
                    'talk_time_secs': _mtd_talk_secs,
                    'talk_time_formatted': f"{_mtd_h}h {_mtd_m}m",
                    'talk_time_avg_formatted': f"{_avg_secs//3600}h {(_avg_secs%3600)//60}m",
                    'leads_new': int(mtd_leads.get('summary', {}).get('new_created', 0)),
                    'self_leads_new': int(mtd_leads.get('summary', {}).get('self_leads', {}).get('new', 0)),
                    'leads_handled': int(mtd_leads.get('summary', {}).get('handled_today', 0)),
                    'leads_won': int(mtd_leads.get('summary', {}).get('won', 0)),
                    'leads_lost': int(mtd_leads.get('summary', {}).get('lost', 0)),
                    'deal_value_closed': float(mtd_leads.get('summary', {}).get('deal_value_updated', 0)),
                    'deal_value_received': float(mtd_leads.get('summary', {}).get('deal_value_received', 0)),
                    'revenue': float(mtd_leads.get('summary', {}).get('revenue', 0)),
                    'lost_deal_value': float(mtd_leads.get('summary', {}).get('lost_revenue', 0)),
                    'deal_value_in_hand': _mtd_rev_in_hand,
                }
            elif is_service_dept:
                _svc = mtd_service.get('summary', {})
                _svc_resolved = int(_svc.get('resolved', 0))
                _svc_wtat = int(_svc.get('within_tat', 0))
                mtd_dept_kpi = {
                    'tickets_new': int(_svc.get('new', 0)),
                    'tickets_resolved': _svc_resolved,
                    'within_tat': _svc_wtat,
                    'out_of_tat': int(_svc.get('out_of_tat', 0)),
                    'pending': int(_svc.get('pending', 0)),
                    'tat_pct': round((_svc_wtat / _svc_resolved) * 100) if _svc_resolved > 0 else 0,
                }
            elif is_procurement_dept:
                try:
                    _proc_mtd = get_procurement_range_summary(db, target_employee_id, mtd_start, today)
                    _proc_s = _proc_mtd.get('summary', {})
                    mtd_dept_kpi = {
                        'received_count': int(_proc_s.get('received_count', 0)),
                        'received_value': float(_proc_s.get('received_value', 0)),
                        'pending_count': int(_proc_s.get('pending_count', 0)),
                        'above_tat': int(_proc_s.get('above_tat', 0)),
                        'completion_pct': int(_proc_s.get('completion_pct', 0)),
                    }
                except Exception as _pm:
                    logger.warning(f"[DC-PROGRESS] procurement MTD failed (non-fatal): {_pm}")
                    mtd_dept_kpi = {'received_count': 0, 'received_value': 0.0, 'pending_count': 0, 'above_tat': 0, 'completion_pct': 0}

            logger.info(f"[DC-PROGRESS] Fetching Rolling 30-day summaries")
            rolling_tasks = get_tasks_summary(db, target_employee_id, rolling_start, today)
            rolling_kra = get_kra_summary(db, target_employee_id, rolling_start, today)
            rolling_leads = get_leads_summary(db, target_employee_id, rolling_start, today)
            rolling_service = get_service_tickets_summary(db, target_employee_id, rolling_start, today)
        else:
            logger.info(f"[DC-PROGRESS] Skipping MTD/rolling aggregates (include_aggregates=false)")
            mtd_tasks = mtd_kra = mtd_leads = mtd_service = mtd_timesheet = {}
            rolling_tasks = rolling_kra = rolling_leads = rolling_service = {}
            working_days = 1
            mtd_dept_kpi = {}

        logger.info(f"[DC-PROGRESS] All summaries fetched successfully")

        _mtd_ts_sum = mtd_timesheet.get("summary", {}) if include_aggregates else {}
        _mtd_ts_min = int(_mtd_ts_sum.get("total_minutes", 0))
        _mtd_ts_sub_min = int(_mtd_ts_sum.get("submitted_minutes", 0))
        _mtd_ts_apr_min = int(_mtd_ts_sum.get("approved_minutes", 0))
        _mtd_ts_avg_min = _mtd_ts_min // working_days if include_aggregates else 0
        _mtd_ts_sub_avg_min = _mtd_ts_sub_min // working_days if include_aggregates else 0
        _mtd_ts_apr_avg_min = _mtd_ts_apr_min // working_days if include_aggregates else 0
        overall_data = {
            "mtd": {
                "period": f"{mtd_start.isoformat()} to {today.isoformat()}",
                "label": "Month to Date",
                "working_days": working_days if include_aggregates else 0,
                "tasks": mtd_tasks.get("summary", {}),
                "kra": mtd_kra.get("summary", {}),
                "leads": mtd_leads.get("summary", {}),
                "service_tickets": mtd_service.get("summary", {}),
                "timesheet": {
                    "total_minutes": _mtd_ts_min,
                    "total_hours": _mtd_ts_sum.get("total_hours", "0h 0m"),
                    "submitted_hours": _mtd_ts_sum.get("submitted_hours", "0h 0m"),
                    "approved_hours": _mtd_ts_sum.get("approved_hours", "0h 0m"),
                    "avg_minutes_per_day": _mtd_ts_avg_min,
                    "avg_hours_per_day": f"{_mtd_ts_avg_min//60}h {_mtd_ts_avg_min%60}m",
                    "submitted_avg_per_day": f"{_mtd_ts_sub_avg_min//60}h {_mtd_ts_sub_avg_min%60}m",
                    "approved_avg_per_day": f"{_mtd_ts_apr_avg_min//60}h {_mtd_ts_apr_avg_min%60}m",
                },
                "dept_kpi": mtd_dept_kpi,
            },
            "rolling_30": {
                "period": f"{rolling_start.isoformat()} to {today.isoformat()}",
                "label": "Rolling 30 Days",
                "tasks": rolling_tasks.get("summary", {}),
                "kra": rolling_kra.get("summary", {}),
                "leads": rolling_leads.get("summary", {}),
                "service_tickets": rolling_service.get("summary", {})
            }
        }
        
        downline_options = None
        if is_manager_or_leadership(current_user):
            # DC Protocol: VGK4U Supreme gets ALL active employees including self; others get downline
            if has_unrestricted_access(current_user):
                _rc = (current_user.role.role_code if current_user.role else '').lower()
                _is_vgk = _rc in ('vgk4u', 'vgk4u_supreme')
                downline_options = _get_all_employees_options(db, current_user.id, current_user.base_company_id, is_vgk_supreme=_is_vgk)
            else:
                downline_options = _get_downline_options(db, current_user.id)
        
        logger.info(f"[DC-PROGRESS] Building response")

        # DC Protocol: Team Performance block — for both day and range views.
        # Show when: (a) viewing own page and user has downline, OR (b) admin viewer looking at a manager's page.
        team_performance_data = None
        _tp_source_opts = None  # the employee options list to iterate for members_perf
        _tp_ids: list = []

        if target_employee_id == current_user.id and downline_options:
            # Own page: use current user's downline_options
            _tp_ids = [o['id'] for o in downline_options if o.get('id') != current_user.id]
            _tp_source_opts = [o for o in downline_options if o.get('id') != current_user.id]
        elif target_employee_id != current_user.id and has_unrestricted_access(current_user):
            # Admin viewing another employee's page — show that employee's team if they have a downline
            _target_downline = _get_downline_options(db, target_employee_id)
            if _target_downline:
                _tp_ids = [o['id'] for o in _target_downline]
                _tp_source_opts = _target_downline

        if _tp_ids and _tp_source_opts is not None:
            try:
                _tp_kra  = _batch_kra_summary(db, _tp_ids, query_start, query_end)
                _tp_att  = _batch_attendance_summary(db, _tp_ids, query_start, query_end)
                _tp_dp   = _batch_day_plan_summary(db, _tp_ids, query_start, query_end)
                _tp_ts: dict = {}
                _ts_rows_tp = db.execute(
                    text(
                        "SELECT employee_id, COUNT(DISTINCT date) FROM staff_timesheet_entries "
                        "WHERE employee_id = ANY(:ids) AND date BETWEEN :df AND :dt GROUP BY employee_id"
                    ), {'ids': _tp_ids, 'df': query_start, 'dt': query_end}
                ).fetchall()
                for _r in _ts_rows_tp:
                    _tp_ts[_r[0]] = int(_r[1] or 0)

                # DC Protocol: leads & service tickets for dept_score per member
                _tp_leads = _batch_leads_summary(db, _tp_ids, query_start, query_end)
                _tp_svc_rows = db.execute(
                    text(
                        "SELECT service_technician_id,"
                        " COUNT(*) AS total,"
                        " SUM(CASE WHEN status IN ('Resolved','Closed') THEN 1 ELSE 0 END) AS resolved"
                        " FROM service_ticket"
                        " WHERE service_technician_id = ANY(:ids)"
                        "   AND DATE(created_date) BETWEEN :df AND :dt"
                        " GROUP BY service_technician_id"
                    ), {'ids': _tp_ids, 'df': query_start, 'dt': query_end}
                ).fetchall()
                _tp_svc = {r[0]: {'total': int(r[1] or 0), 'resolved': int(r[2] or 0)} for r in _tp_svc_rows}

                # Member dept lookup for dept_score computation
                _mem_dept_map: dict = {}
                _mem_rows = db.query(StaffEmployee.id, StaffEmployee.department_id).filter(
                    StaffEmployee.id.in_(_tp_ids)
                ).all()
                if _mem_rows:
                    _mem_dept_ids = list({r[1] for r in _mem_rows if r[1]})
                    from app.models.staff import StaffDepartment as _StaffDept
                    _dept_name_rows = db.query(_StaffDept.id, _StaffDept.name).filter(
                        _StaffDept.id.in_(_mem_dept_ids)
                    ).all()
                    _dept_name_lkp = {r[0]: (r[1] or '').lower() for r in _dept_name_rows}
                    for r in _mem_rows:
                        _mem_dept_map[r[0]] = _dept_name_lkp.get(r[1], '')

                _tp_working_days = (query_end - query_start).days + 1 or 1

                members_perf = []
                for _opt in _tp_source_opts:
                    _eid = _opt.get('id')
                    _k  = _tp_kra.get(_eid, {})
                    _a  = _tp_att.get(_eid, {})
                    _dp = _tp_dp.get(_eid, {})
                    _ts_days = _tp_ts.get(_eid, 0)
                    _kra_pct = round(_k.get('completed', 0) / _k.get('effective_total', 1) * 100) if _k.get('effective_total', 0) > 0 else 0
                    _att_pct = round(_a.get('days_present', 0) / _tp_working_days * 100) if _a.get('days_present') is not None else 0
                    _dp_pct  = _dp.get('pct', 0) if _dp.get('has_plan') else 0
                    _ts_pct  = round(_ts_days / _tp_working_days * 100)
                    # Dept score per member based on their department
                    _mdept = _mem_dept_map.get(_eid, '')
                    _dept_score = 0
                    if 'sales' in _mdept or 'crm' in _mdept:
                        _ld = _tp_leads.get(_eid, {})
                        _dept_score = round((_ld.get('won', 0) / _ld.get('total', 1)) * 100) if _ld.get('total', 0) > 0 else 0
                    elif 'service' in _mdept:
                        _sv = _tp_svc.get(_eid, {'total': 0, 'resolved': 0})
                        _dept_score = round((_sv['resolved'] / _sv['total']) * 100) if _sv['total'] > 0 else 0
                    _ld_detail = _tp_leads.get(_eid, {})
                    _sv_detail = _tp_svc.get(_eid, {'total': 0, 'resolved': 0})
                    members_perf.append({
                        'id': _eid,
                        'name': _opt.get('full_name', ''),
                        'emp_code': _opt.get('emp_code', ''),
                        'dept': _opt.get('department', ''),
                        'kra_pct': _kra_pct,
                        'att_pct': _att_pct,
                        'plan_pct': _dp_pct,
                        'ts_pct': _ts_pct,
                        'dept_score': _dept_score,
                        'kra_done': _k.get('completed', 0),
                        'kra_total': _k.get('effective_total', 0),
                        'att_days': _a.get('days_present', 0),
                        'dept_type': _mdept,
                        'leads_new': int(_ld_detail.get('total', 0)),
                        'leads_won': int(_ld_detail.get('won', 0)),
                        'leads_lost': int(_ld_detail.get('lost', 0)),
                        'deal_value': float(_ld_detail.get('revenue', 0) if _ld_detail else 0),
                        'svc_total': int(_sv_detail.get('total', 0)),
                        'svc_resolved': int(_sv_detail.get('resolved', 0)),
                    })

                avg = lambda key: round(sum(m[key] for m in members_perf) / len(members_perf)) if members_perf else 0

                # W9: Departmental breakdown — group members by dept, compute avg per dept
                _dept_groups: dict = {}
                for _m in members_perf:
                    _d = (_m.get('dept') or 'Unassigned').strip() or 'Unassigned'
                    _dept_groups.setdefault(_d, []).append(_m)
                dept_breakdown = {
                    _dept: {
                        'member_count': len(_mems),
                        'avg_kra_pct': round(sum(x['kra_pct'] for x in _mems) / len(_mems)) if _mems else 0,
                        'avg_att_pct': round(sum(x['att_pct'] for x in _mems) / len(_mems)) if _mems else 0,
                        'avg_plan_pct': round(sum(x['plan_pct'] for x in _mems) / len(_mems)) if _mems else 0,
                        'avg_ts_pct': round(sum(x['ts_pct'] for x in _mems) / len(_mems)) if _mems else 0,
                        'avg_dept_score': round(sum(x['dept_score'] for x in _mems) / len(_mems)) if _mems else 0,
                    }
                    for _dept, _mems in _dept_groups.items()
                }

                # W8/W9: Section averages for this manager's team
                _avg_ts_minutes = 0
                if _tp_ts and members_perf:
                    _avg_ts_minutes = round(
                        sum(_tp_ts.get(_eid, 0) * 60 for _eid in _tp_ids) / len(_tp_ids)
                    )

                team_performance_data = {
                    'team_size': len(_tp_ids),
                    'working_days': _tp_working_days,
                    'summary': {
                        'avg_kra_pct': avg('kra_pct'),
                        'avg_att_pct': avg('att_pct'),
                        'avg_plan_pct': avg('plan_pct'),
                        'avg_ts_pct': avg('ts_pct'),
                        'avg_dept_score': avg('dept_score'),
                    },
                    # W9: Section-level team averages (for all managers with direct reports)
                    'section_avgs': {
                        'kra_avg_pct': avg('kra_pct'),
                        'attendance_avg_pct': avg('att_pct'),
                        'tasks_avg_pct': avg('plan_pct'),
                        'ts_avg_pct': avg('ts_pct'),
                        'ts_avg_minutes': _avg_ts_minutes,
                        'ts_avg_hours': f"{_avg_ts_minutes // 60}h {_avg_ts_minutes % 60}m",
                        'dept_avg_score': avg('dept_score'),
                    },
                    # W9: Per-department averages (departmental metrics)
                    'dept_breakdown': dept_breakdown,
                    'members': members_perf,
                }
            except Exception as _te:
                logger.warning(f"[DC-PROGRESS] team_performance failed (non-fatal): {_te}")
                import traceback as _tb
                logger.warning(f"[DC-PROGRESS] team_performance traceback: {_tb.format_exc()}")

        # DC Protocol: Include permissions and downline_options inside data for mobile compatibility
        return {
            "success": True,
            "data": {
                "employee": {
                    "id": target_employee.id,
                    "emp_code": target_employee.emp_code,
                    "full_name": target_employee.full_name,
                    "role": target_employee.role.role_name if target_employee.role else None,
                    "department": target_employee.department.name if target_employee.department else None,
                    "department_code": dept_code
                },
                "query": {
                    "date": target_date.isoformat(),
                    "date_from": query_start.isoformat() if is_range_query else None,
                    "date_to": query_end.isoformat() if is_range_query else None,
                    "is_range": is_range_query
                },
                "tasks": tasks_data,
                "kra": kra_data,
                "attendance": attendance_data,
                "timesheet": timesheet_data,
                "leads": leads_data,
                "service_tickets": service_data,
                "travel": travel_data,
                "day_plan": day_plan_data,
                "day_progress": day_progress_data,
                "overall": overall_data,
                "permissions": {
                    "can_view_team": is_manager_or_leadership(current_user),
                    "can_export": True,
                    "downline_count": len(downline_options) if downline_options else 0
                },
                "downline_options": downline_options,
                "dept_type": dept_type,
                "also_manages_depts": also_manages_depts,
                "dept_kpi": dept_kpi,
                "calls": calls_range_data,
                "procurement": procurement_range_data,
                "team_performance": team_performance_data,
                "accounts_data": accounts_data,
                "po_pr_data": po_pr_data,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        _err_str = str(e)
        logger.error(f"[DC-PROGRESS] CRASH: {_err_str}")
        logger.error(f"[DC-PROGRESS] Traceback: {traceback.format_exc()}")
        # DC Protocol: return 408 on statement_timeout so proxy stops retrying immediately
        if 'statement timeout' in _err_str.lower() or 'QueryCanceled' in _err_str or 'canceling statement' in _err_str.lower():
            try:
                db.rollback()
            except Exception:
                pass
            raise HTTPException(status_code=408, detail="Progress query timed out — try a shorter date range or try again shortly")
        raise HTTPException(status_code=500, detail=f"Progress summary error: {_err_str}")


@router.get("/team-overview", summary="Get team overview for all downline employees")
def get_team_overview(
    target_date: Optional[date] = Query(None, description="Date for progress (defaults to today)"),
    date_from: Optional[date] = Query(None, description="Start date for range query"),
    date_to: Optional[date] = Query(None, description="End date for range query"),
    current_user: StaffEmployee = Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    import traceback
    try:
        if not is_manager_or_leadership(current_user):
            raise HTTPException(status_code=403, detail="Only managers can view team overview")

        today = get_indian_date()
        if target_date is None:
            target_date = today

        if date_from and date_to:
            query_start = date_from
            query_end = date_to
        else:
            query_start = target_date
            query_end = target_date

        hidden_codes = HIDDEN_FROM_TEAM_CODES or []

        if has_unrestricted_access(current_user):
            q = db.query(StaffEmployee).options(
                joinedload(StaffEmployee.department),
                joinedload(StaffEmployee.role)
            ).filter(
                StaffEmployee.status == 'active',
                StaffEmployee.id != current_user.id
            )
            if hidden_codes:
                q = q.filter(~StaffEmployee.emp_code.in_(hidden_codes))
            if current_user.base_company_id:
                q = q.filter(StaffEmployee.base_company_id == current_user.base_company_id)
            downline_employees = q.order_by(StaffEmployee.full_name).limit(200).all()
        else:
            downline_ids = get_downline_employee_ids(db, current_user.id, recursive=True)
            if not downline_ids:
                return {"success": True, "data": {"members": [], "summary": {}}}
            q2 = db.query(StaffEmployee).options(
                joinedload(StaffEmployee.department),
                joinedload(StaffEmployee.role)
            ).filter(
                StaffEmployee.id.in_(downline_ids)
            )
            if hidden_codes:
                q2 = q2.filter(~StaffEmployee.emp_code.in_(hidden_codes))
            if current_user.base_company_id:
                q2 = q2.filter(StaffEmployee.base_company_id == current_user.base_company_id)
            downline_employees = q2.order_by(StaffEmployee.full_name).all()

        emp_ids = [e.id for e in downline_employees]
        if not emp_ids:
            return {"success": True, "data": {"members": [], "summary": {}}}

        tasks_by_emp = _batch_tasks_summary(db, emp_ids, query_start, query_end)
        kra_by_emp = _batch_kra_summary(db, emp_ids, query_start, query_end)
        attendance_by_emp = _batch_attendance_summary(db, emp_ids, query_start, query_end)
        dayplan_by_emp = _batch_day_plan_summary(db, emp_ids, query_start, query_end)
        leads_by_emp = _batch_leads_summary(db, emp_ids, query_start, query_end)

        team_total_tasks = 0
        team_completed_tasks = 0
        team_total_kra = 0
        team_completed_kra = 0
        team_present = 0
        team_plans_created = 0

        members = []
        for emp in downline_employees:
            eid = emp.id
            t = tasks_by_emp.get(eid, {})
            k = kra_by_emp.get(eid, {})
            a = attendance_by_emp.get(eid, {})
            dp = dayplan_by_emp.get(eid, {})
            l = leads_by_emp.get(eid, {})

            tasks_planned = t.get('planned', 0)
            tasks_completed = t.get('completed', 0)
            tasks_pending = t.get('pending', 0)
            tasks_overdue = t.get('overdue', 0)
            kra_total = k.get('total', 0)
            kra_effective = k.get('effective_total', float(kra_total))
            kra_completed = k.get('completed', 0)
            kra_pending = k.get('pending', 0)

            team_total_tasks += tasks_planned
            team_completed_tasks += tasks_completed
            team_total_kra += kra_effective
            team_completed_kra += kra_completed
            att_status = a.get('hr_attendance') or a.get('status', 'absent')
            if att_status in ('present', 'half_day', 'on_duty'):
                team_present += 1
            if dp.get('has_plan'):
                team_plans_created += 1

            members.append({
                "id": emp.id,
                "emp_code": emp.emp_code,
                "full_name": emp.full_name,
                "role": emp.role.role_name if emp.role else None,
                "department": emp.department.name if emp.department else None,
                "tasks": {
                    "planned": tasks_planned,
                    "completed": tasks_completed,
                    "pending": tasks_pending,
                    "overdue": tasks_overdue
                },
                "kra": {
                    "total": kra_total,
                    "effective_total": kra_effective,
                    "completed": kra_completed,
                    "pending": kra_pending
                },
                "attendance": {
                    "status": a.get('status', 'absent'),
                    "hr_attendance": a.get('hr_attendance'),
                    "hr_approval": a.get('hr_approval', 'na')
                },
                "day_plan": {
                    "has_plan": dp.get('has_plan', False),
                    "status": dp.get('status', 'not_created'),
                    "total_items": dp.get('total_items', 0),
                    "completed_items": dp.get('completed_items', 0),
                    "is_finalized": dp.get('is_finalized', False)
                },
                "leads": {
                    "total": l.get('total', 0),
                    "won": l.get('won', 0),
                    "revenue": float(l.get('revenue', 0))
                }
            })

        summary = {
            "total_members": len(members),
            "present": team_present,
            "absent": len(members) - team_present,
            "plans_created": team_plans_created,
            "plans_pending": len(members) - team_plans_created,
            "total_tasks": team_total_tasks,
            "completed_tasks": team_completed_tasks,
            "tasks_completion_pct": round((team_completed_tasks / team_total_tasks * 100) if team_total_tasks > 0 else 0),
            "total_kra": team_total_kra,
            "completed_kra": team_completed_kra,
            "kra_completion_pct": round((team_completed_kra / team_total_kra * 100) if team_total_kra > 0 else 0)
        }

        return {
            "success": True,
            "data": {
                "query": {
                    "date": target_date.isoformat(),
                    "date_from": query_start.isoformat(),
                    "date_to": query_end.isoformat()
                },
                "members": members,
                "summary": summary
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[DC-TEAM-OVERVIEW] CRASH: {e}")
        logger.error(f"[DC-TEAM-OVERVIEW] Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Team overview error: {str(e)}")


def _batch_tasks_summary(db: Session, emp_ids: List[int], date_from: date, date_to: date) -> dict:
    from sqlalchemy import literal_column
    results = {}
    primary_tasks = db.query(
        StaffTask.primary_assignee_id.label('emp_id'),
        StaffTask.status,
        StaffTask.due_date,
        func.count().label('cnt')
    ).filter(
        StaffTask.primary_assignee_id.in_(emp_ids),
        or_(
            func.date(StaffTask.due_date).between(date_from, date_to),
            and_(
                func.date(StaffTask.created_at).between(date_from, date_to),
                StaffTask.due_date.is_(None)
            )
        )
    ).group_by(StaffTask.primary_assignee_id, StaffTask.status, StaffTask.due_date).all()

    today = get_indian_date()
    for row in primary_tasks:
        eid = row.emp_id
        if eid not in results:
            results[eid] = {'planned': 0, 'completed': 0, 'pending': 0, 'overdue': 0}
        results[eid]['planned'] += row.cnt
        if row.status == 'completed':
            results[eid]['completed'] += row.cnt
        elif row.status != 'cancelled':
            due_val = _safe_date(row.due_date) if row.due_date else None
            if due_val and due_val < today:
                results[eid]['overdue'] += row.cnt
            else:
                results[eid]['pending'] += row.cnt
    return results


def _batch_kra_summary(db: Session, emp_ids: List[int], date_from: date, date_to: date) -> dict:
    """
    DC Protocol: Batch KRA summary with leave/holiday/Sunday exclusion.
    Sundays are auto-excluded. HR-marked leaves/holidays and approved leave requests
    are excluded per employee. Half-days count as 0.5 in the effective denominator.
    Fully dynamic — retroactive attendance changes reflected on every call.
    """
    # Fetch individual instances (needed for per-employee leave filtering)
    rows = db.query(
        StaffKRADailyInstance.employee_id,
        StaffKRADailyInstance.instance_date,
        StaffKRADailyInstance.completion_status,
    ).filter(
        StaffKRADailyInstance.employee_id.in_(emp_ids),
        StaffKRADailyInstance.instance_date.between(date_from, date_to)
    ).all()

    if not rows:
        return {}

    # Build leave/half-day maps — includes auto-Sundays for all employees
    from app.utils.leave_utils import get_employee_nonworking_data
    leave_map, half_day_map = get_employee_nonworking_data(db, emp_ids, date_from, date_to)

    results = {}
    for row in rows:
        eid = row.employee_id
        inst_date = row.instance_date.date() if hasattr(row.instance_date, 'date') and callable(row.instance_date.date) else row.instance_date

        # Skip fully non-working days (Sunday, holiday, full leave)
        if inst_date in leave_map.get(eid, set()):
            continue

        if eid not in results:
            results[eid] = {'total': 0, 'effective_total': 0.0, 'completed': 0, 'pending': 0}

        weight = 0.5 if inst_date in half_day_map.get(eid, set()) else 1.0
        results[eid]['total'] += 1
        results[eid]['effective_total'] += weight
        if row.completion_status == 'completed':
            results[eid]['completed'] += 1
        elif row.completion_status in ('pending', 'in_progress'):
            results[eid]['pending'] += 1

    return results


def _batch_attendance_summary(db: Session, emp_ids: List[int], date_from: date, date_to: date) -> dict:
    att_rows = db.query(
        StaffAttendance.employee_id,
        func.count().label('days_present')
    ).filter(
        StaffAttendance.employee_id.in_(emp_ids),
        func.date(StaffAttendance.clock_in).between(date_from, date_to),
        StaffAttendance.status == 'present'
    ).group_by(StaffAttendance.employee_id).all()

    results = {}
    present_set = set()
    for row in att_rows:
        present_set.add(row.employee_id)
        results[row.employee_id] = {'status': 'present', 'days_present': row.days_present}

    hr_rows = db.query(
        StaffAttendanceSheet.employee_id,
        StaffAttendanceSheet.attendance_status,
        StaffAttendanceSheet.approval_status,
        StaffAttendanceSheet.date
    ).filter(
        StaffAttendanceSheet.employee_id.in_(emp_ids),
        StaffAttendanceSheet.date.between(date_from, date_to)
    ).order_by(StaffAttendanceSheet.date.desc()).all()

    seen_hr = set()
    for row in hr_rows:
        eid = row.employee_id
        if eid in seen_hr:
            continue
        seen_hr.add(eid)
        if eid not in results:
            results[eid] = {'status': 'absent'}
        att_val = row.attendance_status.value if row.attendance_status else None
        results[eid]['hr_attendance'] = att_val
        results[eid]['hr_approval'] = row.approval_status.value if row.approval_status else 'pending'
        if att_val and att_val in ('present', 'half_day', 'on_duty') and eid not in present_set:
            results[eid]['status'] = att_val

    return results


def _batch_day_plan_summary(db: Session, emp_ids: List[int], date_from: date, date_to: date) -> dict:
    plans = db.query(StaffDayPlan).options(
        joinedload(StaffDayPlan.items)
    ).filter(
        StaffDayPlan.employee_id.in_(emp_ids),
        StaffDayPlan.plan_date.between(date_from, date_to)
    ).all()

    results = {}
    for plan in plans:
        eid = plan.employee_id
        items = plan.items or []
        completed = sum(1 for i in items if i.eod_status == 'delivered')
        if eid not in results:
            results[eid] = {
                'has_plan': True,
                'status': plan.status,
                'total_items': len(items),
                'completed_items': completed,
                'is_finalized': plan.finalized_at is not None
            }
        else:
            results[eid]['total_items'] += len(items)
            results[eid]['completed_items'] += completed
    return results


def _batch_leads_summary(db: Session, emp_ids: List[int], date_from: date, date_to: date) -> dict:
    rows = db.query(
        CRMLead.primary_owner_id.label('emp_id'),
        CRMLead.status,
        func.count().label('cnt'),
        func.coalesce(func.sum(CRMLead.deal_value_total), 0).label('revenue')
    ).filter(
        CRMLead.primary_owner_type == 'staff',
        CRMLead.primary_owner_id.in_(emp_ids),
        func.date(CRMLead.created_at).between(date_from, date_to)
    ).group_by(CRMLead.primary_owner_id, CRMLead.status).all()

    results = {}
    for row in rows:
        eid = row.emp_id
        if eid not in results:
            results[eid] = {'total': 0, 'won': 0, 'revenue': 0}
        results[eid]['total'] += row.cnt
        if row.status == 'won':
            results[eid]['won'] += row.cnt
            results[eid]['revenue'] += float(row.revenue or 0)
    return results


def _safe_date(dt):
    """DC Protocol: Safely extract date from date or datetime object"""
    if dt is None:
        return None
    return dt.date() if hasattr(dt, 'date') and callable(getattr(dt, 'date', None)) else dt


def _get_downline_options(db: Session, manager_id: int) -> Optional[List[dict]]:
    """Get downline employee options for team filter dropdown (DC Protocol)"""
    downline_ids = get_downline_employee_ids(db, manager_id, recursive=True)
    if not downline_ids:
        return None
    
    hidden_codes = HIDDEN_FROM_TEAM_CODES or []
    q = db.query(StaffEmployee).filter(
        StaffEmployee.id.in_(downline_ids)
    )
    if hidden_codes:
        q = q.filter(~StaffEmployee.emp_code.in_(hidden_codes))
    downline_employees = q.order_by(StaffEmployee.full_name).all()
    
    return [
        {
            "id": emp.id,
            "emp_code": emp.emp_code,
            "full_name": emp.full_name,
            "role": emp.role.role_name if emp.role else None,
            "department": emp.department.name if emp.department else None
        }
        for emp in downline_employees
    ]


def _get_all_employees_options(db: Session, current_user_id: int, company_id: Optional[int], is_vgk_supreme: bool = False) -> Optional[List[dict]]:
    """
    DC Protocol: Get ALL active employees for VGK4U Supreme and Key Leadership
    - VGK4U Supreme: sees ALL employees across ALL companies including themselves
    - Key Leadership: sees ALL employees across ALL companies (excludes HIDDEN_FROM_TEAM_CODES)
    - HIDDEN_FROM_TEAM_CODES (MR10001) is only hidden from non-VGK-Supreme users
    """
    query = db.query(StaffEmployee).options(
        joinedload(StaffEmployee.department),
        joinedload(StaffEmployee.role)
    ).filter(
        StaffEmployee.status == 'active',
    )

    if not is_vgk_supreme:
        # Non-VGK-Supreme: exclude self and hidden codes
        query = query.filter(StaffEmployee.id != current_user_id)
        hidden_codes = HIDDEN_FROM_TEAM_CODES or []
        if hidden_codes:
            query = query.filter(~StaffEmployee.emp_code.in_(hidden_codes))
    # VGK Supreme: no exclusions — sees all staff including themselves

    # No company_id filter — cross-company access for leadership
    all_employees = query.order_by(StaffEmployee.full_name).limit(500).all()
    
    if not all_employees:
        return None
    
    return [
        {
            "id": emp.id,
            "emp_code": emp.emp_code,
            "full_name": emp.full_name,
            "role": emp.role.role_name if emp.role else None,
            "department": emp.department.name if emp.department else None
        }
        for emp in all_employees
    ]


def get_day_plan_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get Day Plan status for the date range from StaffDayPlan/StaffDayPlanItem"""
    plans = db.query(StaffDayPlan).options(
        joinedload(StaffDayPlan.items)
    ).filter(
        StaffDayPlan.employee_id == employee_id,
        StaffDayPlan.plan_date.between(date_from, date_to)
    ).order_by(StaffDayPlan.plan_date.desc()).all()

    if not plans:
        from datetime import timedelta
        empty_dw = []
        cur = date_from
        while cur <= date_to:
            empty_dw.append({"date": cur.isoformat(), "was_planned": False, "is_finished": False,
                             "tasks_total": 0, "tasks_done": 0, "tasks_pending": 0, "pct_done": 0,
                             "status": "not_created"})
            cur += timedelta(days=1)
        return {
            "has_plan": False, "status": "not_created",
            "total_items": 0, "completed_items": 0, "in_progress_items": 0,
            "pending_items": 0, "completion_percent": 0,
            "is_finalized": False, "finalized_at": None,
            "plans": [], "datewise": empty_dw,
        }

    total_items = 0
    completed_items = 0
    in_progress_items = 0
    pending_items = 0
    finalized_count = 0

    plan_records = []
    for plan in plans:
        items = plan.items or []
        p_completed = sum(1 for i in items if i.eod_status == 'delivered')
        p_in_progress = sum(1 for i in items if i.eod_status == 'in_progress')
        p_pending = len(items) - p_completed - p_in_progress

        total_items += len(items)
        completed_items += p_completed
        in_progress_items += p_in_progress
        pending_items += p_pending

        if plan.finalized_at:
            finalized_count += 1

        plan_records.append({
            "date": plan.plan_date.isoformat(),
            "status": plan.status,
            "items": len(items),
            "completed": p_completed,
            "in_progress": p_in_progress,
            "pending": p_pending,
            "is_finalized": plan.finalized_at is not None,
            "finalized_at": plan.finalized_at.isoformat() if plan.finalized_at else None,
            "task_items": [
                {
                    "title": (item.task.title if item.task else (item.phase.phase_title if item.phase else "Unknown")),
                    "item_type": item.item_type,
                    "eod_status": item.eod_status or "pending",
                    "eod_progress": item.eod_progress or 0,
                    "time_spent_minutes": item.time_spent_minutes or 0,
                    "priority": item.task.priority if item.task else None,
                    "notes": item.eod_notes or "",
                    "is_carried_forward": item.is_carried_forward,
                }
                for item in items
            ]
        })

    latest = plans[0]
    # Build full date-range datewise (include NO-plan days)
    from datetime import timedelta
    plan_by_date = {r["date"]: r for r in plan_records}
    full_datewise = []
    cur = date_from
    while cur <= date_to:
        d = cur.isoformat()
        rec = plan_by_date.get(d)
        if rec:
            items_total = rec["items"]
            done = rec["completed"]
            pct = round(done / items_total * 100) if items_total > 0 else 0
            full_datewise.append({
                "date": d, "was_planned": True,
                "is_finished": rec["is_finalized"],
                "tasks_total": items_total, "tasks_done": done,
                "tasks_pending": rec["pending"],
                "pct_done": pct,
                "status": rec["status"],
                "task_items": rec.get("task_items", []),
            })
        else:
            full_datewise.append({
                "date": d, "was_planned": False,
                "is_finished": False,
                "tasks_total": 0, "tasks_done": 0, "tasks_pending": 0,
                "pct_done": 0, "status": "not_created",
            })
        cur += timedelta(days=1)

    overall_pct = round(completed_items / total_items * 100) if total_items > 0 else 0
    return {
        "has_plan": True,
        "status": latest.status,
        "total_items": total_items,
        "completed_items": completed_items,
        "in_progress_items": in_progress_items,
        "pending_items": pending_items,
        "completion_percent": overall_pct,
        "is_finalized": latest.finalized_at is not None,
        "finalized_at": latest.finalized_at.isoformat() if latest.finalized_at else None,
        "plans": plan_records,
        "datewise": full_datewise,
    }


def get_day_progress_breakdown(db: Session, employee_id: int, target_date: date) -> dict:
    """Get day planner/closure breakdown for a single employee on a specific date.
    Returns fields matching the web Day Progress tab for mobile parity."""
    plan = db.query(StaffDayPlan).options(
        joinedload(StaffDayPlan.items)
    ).filter(
        StaffDayPlan.employee_id == employee_id,
        StaffDayPlan.plan_date == target_date
    ).first()

    overall_task_ids = set()
    primary_tasks = db.query(StaffTask.id).filter(
        StaffTask.primary_assignee_id == employee_id,
        StaffTask.status.in_(['pending', 'in_progress', 'on_hold', 'under_review']),
        StaffTask.is_deleted == False
    ).all()
    for r in primary_tasks:
        overall_task_ids.add(('task', r.id))

    secondary_tasks = db.query(StaffTask.id).join(
        StaffTaskAssignee, StaffTaskAssignee.task_id == StaffTask.id
    ).filter(
        StaffTaskAssignee.employee_id == employee_id,
        StaffTask.status.in_(['pending', 'in_progress', 'on_hold', 'under_review']),
        StaffTask.is_deleted == False
    ).all()
    for r in secondary_tasks:
        overall_task_ids.add(('task', r.id))

    phase_ids = db.query(StaffTaskPhase.id).filter(
        StaffTaskPhase.phase_assignee_id == employee_id,
        StaffTaskPhase.phase_status.in_(['pending', 'in_progress', 'on_hold']),
        StaffTaskPhase.is_deleted == False
    ).all()
    for r in phase_ids:
        overall_task_ids.add(('phase', r.id))

    overall_count = len(overall_task_ids)
    planned_count = len(plan.items) if plan else 0
    overall_pending = max(0, overall_count - planned_count)

    if not plan:
        planner_status = "na" if overall_count == 0 else "pending"
    elif planned_count > 0:
        planner_status = "done"
    else:
        planner_status = "pending"

    total_planned_eod = 0
    closed_count = 0
    left_count = 0
    eod_filled = 0
    plan_task_ids = set()
    closure_status = "na"

    if plan and plan.items:
        total_planned_eod = len(plan.items)
        for item in plan.items:
            effective_status = item.eod_status if item.eod_status else item.planned_status
            if effective_status == 'completed':
                closed_count += 1
            elif item.eod_status is None and effective_status != 'completed':
                left_count += 1
            if item.eod_status is not None:
                eod_filled += 1
            if item.task_id:
                plan_task_ids.add(item.task_id)

        if plan.finalized_at:
            closure_status = "completed"
        else:
            closure_status = "pending"

    worked_count = 0
    if plan_task_ids:
        target_start = datetime.combine(target_date, datetime.min.time())
        target_end = datetime.combine(target_date, datetime.max.time())
        activity_rows = db.query(StaffTaskActivityLog.task_id).filter(
            StaffTaskActivityLog.task_id.in_(list(plan_task_ids)),
            StaffTaskActivityLog.created_at >= target_start,
            StaffTaskActivityLog.created_at <= target_end,
            StaffTaskActivityLog.action.in_([
                'status_changed', 'status_change', 'progress_update',
                'completed', 'updated', 'phase_status_change'
            ])
        ).distinct().all()
        worked_count = len(activity_rows)

    worked_combined = max(eod_filled, worked_count)

    return {
        "day_planner": planner_status,
        "planner_overall": overall_count,
        "planner_overall_pending": overall_pending,
        "planner_overall_planned": planned_count,
        "day_closure": closure_status,
        "closure_planned": total_planned_eod,
        "closure_closed": closed_count,
        "closure_left": left_count,
        "closure_worked": worked_combined
    }


def _get_tat_days(db, module_key: str, default: int = 3) -> int:
    """Read TAT threshold from staff_tat_config table (editable by EA/Key Leadership)."""
    try:
        row = db.execute(
            text("SELECT tat_days FROM staff_tat_config WHERE module_key = :k"),
            {"k": module_key}
        ).fetchone()
        return int(row[0]) if row else default
    except Exception:
        return default


def get_procurement_kpi_summary(db: Session, employee_id: int, target_date: date) -> dict:
    """Get procurement KPI summary for a single date (FTD).
    W4: expanded with new_today, completed_today, completed_pct, within_tat, out_of_tat.
    TAT threshold read from staff_tat_config (default 3 days, editable by EA/KL).
    """
    from app.models.marketplace import MarketplaceProcurementRequest, MarketplacePOItem
    from app.models.staff import StaffEmployee

    target_employee = db.query(StaffEmployee).filter(StaffEmployee.id == employee_id).first()
    company_id = target_employee.base_company_id or 1 if target_employee else 1
    tat_days = _get_tat_days(db, 'procurement_pr', default=3)

    # PRs created today
    new_today = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        func.date(MarketplaceProcurementRequest.created_at) == target_date
    ).scalar() or 0)

    # received_today: count + sum(unit_price * received_qty)
    received_query = db.query(
        func.count(MarketplaceProcurementRequest.id),
        func.sum(MarketplacePOItem.unit_final_price * MarketplaceProcurementRequest.received_qty)
    ).join(
        MarketplacePOItem, MarketplaceProcurementRequest.po_item_id == MarketplacePOItem.id, isouter=True
    ).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status == 'received',
        func.date(MarketplaceProcurementRequest.updated_at) == target_date
    ).first()

    received_cnt = int(received_query[0] or 0)
    received_val = float(received_query[1] or 0)

    pending_cnt = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status.in_(['pending', 'confirmed', 'ordered'])
    ).scalar() or 0)

    above_tat_cnt = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status.in_(['pending', 'confirmed', 'ordered']),
        func.date(MarketplaceProcurementRequest.created_at) < target_date - timedelta(days=tat_days)
    ).scalar() or 0)

    # W4: Within TAT today — received today where (received_at - created_at) <= tat_days
    within_tat_today = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status == 'received',
        func.date(MarketplaceProcurementRequest.updated_at) == target_date,
        func.date(MarketplaceProcurementRequest.updated_at) <= func.date(MarketplaceProcurementRequest.created_at) + timedelta(days=tat_days)
    ).scalar() or 0)
    out_of_tat_today = max(0, received_cnt - within_tat_today)
    completed_pct = round((received_cnt / (received_cnt + pending_cnt) * 100) if (received_cnt + pending_cnt) > 0 else 0)

    return {
        'new_today': new_today,
        'received_today_count': received_cnt,
        'completed_today': received_cnt,
        'received_today_value': received_val,
        'pending_count': pending_cnt,
        'above_tat_count': above_tat_cnt,
        'within_tat_today': within_tat_today,
        'out_of_tat_today': out_of_tat_today,
        'completed_pct': completed_pct,
        'tat_days': tat_days,
    }


def get_procurement_range_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get procurement/PO summary for date range.
    W4: expanded with new_count, completed (received), within_tat, out_of_tat, completion_pct.
    TAT threshold from staff_tat_config.
    """
    from app.models.marketplace import MarketplaceProcurementRequest, MarketplacePOItem
    from app.models.staff import StaffEmployee
    from datetime import timedelta

    emp = db.query(StaffEmployee).filter(StaffEmployee.id == employee_id).first()
    company_id = emp.base_company_id if emp else 1
    tat_days = _get_tat_days(db, 'procurement_pr', default=3)

    # PRs created in the date range
    new_count = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        func.date(MarketplaceProcurementRequest.created_at).between(date_from, date_to)
    ).scalar() or 0)

    received_q = db.query(
        func.count(MarketplaceProcurementRequest.id),
        func.sum(MarketplacePOItem.unit_final_price * MarketplaceProcurementRequest.received_qty)
    ).join(
        MarketplacePOItem, MarketplaceProcurementRequest.po_item_id == MarketplacePOItem.id, isouter=True
    ).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status == 'received',
        func.date(MarketplaceProcurementRequest.updated_at).between(date_from, date_to)
    ).first()

    received_count = int(received_q[0] or 0)
    received_value = float(received_q[1] or 0)

    pending_count = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status.in_(['pending', 'confirmed', 'ordered'])
    ).scalar() or 0)

    above_tat = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status.in_(['pending', 'confirmed', 'ordered']),
        func.date(MarketplaceProcurementRequest.created_at) < date_from - timedelta(days=tat_days)
    ).scalar() or 0)

    # W4: Within TAT (received in range where received_at - created_at <= tat_days)
    within_tat_count = int(db.query(func.count(MarketplaceProcurementRequest.id)).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status == 'received',
        func.date(MarketplaceProcurementRequest.updated_at).between(date_from, date_to),
        func.date(MarketplaceProcurementRequest.updated_at) <= func.date(MarketplaceProcurementRequest.created_at) + timedelta(days=tat_days)
    ).scalar() or 0)
    out_of_tat_count = max(0, received_count - within_tat_count)

    ordered_q = db.query(
        func.count(MarketplaceProcurementRequest.id),
        func.sum(MarketplacePOItem.unit_final_price * MarketplaceProcurementRequest.received_qty)
    ).join(
        MarketplacePOItem, MarketplaceProcurementRequest.po_item_id == MarketplacePOItem.id, isouter=True
    ).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status.in_(['confirmed', 'ordered']),
        func.date(MarketplaceProcurementRequest.updated_at).between(date_from, date_to)
    ).first()
    ordered_count = int(ordered_q[0] or 0)
    ordered_value = float(ordered_q[1] or 0)

    dw_q = db.query(
        func.date(MarketplaceProcurementRequest.updated_at).label('day'),
        func.count(MarketplaceProcurementRequest.id).label('received_cnt'),
        func.sum(MarketplacePOItem.unit_final_price * MarketplaceProcurementRequest.received_qty).label('received_val')
    ).join(
        MarketplacePOItem, MarketplaceProcurementRequest.po_item_id == MarketplacePOItem.id, isouter=True
    ).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        MarketplaceProcurementRequest.status == 'received',
        func.date(MarketplaceProcurementRequest.updated_at).between(date_from, date_to)
    ).group_by(func.date(MarketplaceProcurementRequest.updated_at)).all()

    dw_by_day = {str(r.day): {'count': int(r.received_cnt or 0), 'value': float(r.received_val or 0)} for r in dw_q}

    # W4: daily new PRs count
    dw_new_q = db.query(
        func.date(MarketplaceProcurementRequest.created_at).label('day'),
        func.count(MarketplaceProcurementRequest.id).label('cnt')
    ).filter(
        MarketplaceProcurementRequest.company_id == company_id,
        func.date(MarketplaceProcurementRequest.created_at).between(date_from, date_to)
    ).group_by(func.date(MarketplaceProcurementRequest.created_at)).all()
    dw_new_by_day = {str(r.day): int(r.cnt or 0) for r in dw_new_q}

    datewise = []
    current_date = date_from
    while current_date <= date_to:
        d = current_date.isoformat()
        entry = dw_by_day.get(d, {'count': 0, 'value': 0.0})
        datewise.append({
            "date": d,
            "new_count": dw_new_by_day.get(d, 0),
            "received_count": entry['count'],
            "received_value": entry['value'],
        })
        current_date += timedelta(days=1)

    total_pr = new_count if new_count > 0 else (received_count + pending_count)
    completion_pct = round((received_count / total_pr) * 100) if total_pr > 0 else 0

    return {
        "summary": {
            "new_count": new_count,
            "received_count": received_count,
            "completed_count": received_count,
            "received_value": received_value,
            "ordered_count": ordered_count,
            "ordered_value": ordered_value,
            "pending_count": pending_count,
            "above_tat": above_tat,
            "within_tat_count": within_tat_count,
            "out_of_tat_count": out_of_tat_count,
            "total_pr": total_pr,
            "completion_pct": completion_pct,
            "tat_days": tat_days,
        },
        "datewise": datewise
    }


def get_tasks_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get tasks summary for date range"""
    base_query = db.query(StaffTask).filter(
        or_(
            StaffTask.primary_assignee_id == employee_id,
            StaffTask.id.in_(
                db.query(StaffTaskAssignee.task_id).filter(
                    StaffTaskAssignee.employee_id == employee_id
                )
            )
        )
    )
    
    if date_from == date_to:
        date_filter = or_(
            func.date(StaffTask.due_date) == date_from,
            and_(
                func.date(StaffTask.created_at) == date_from,
                StaffTask.due_date.is_(None)
            )
        )
    else:
        date_filter = or_(
            func.date(StaffTask.due_date).between(date_from, date_to),
            and_(
                func.date(StaffTask.created_at).between(date_from, date_to),
                StaffTask.due_date.is_(None)
            )
        )
    
    tasks = base_query.options(
        joinedload(StaffTask.primary_assignee),
        joinedload(StaffTask.secondary_assignees).joinedload(StaffTaskAssignee.employee)
    ).filter(date_filter).all()

    # DC: Build non-working dates for this employee — excludes Sundays, holidays, leaves
    from app.utils.leave_utils import get_employee_nonworking_data as _get_nw_tasks
    _tsk_leave_map, _tsk_half_map = _get_nw_tasks(db, [employee_id], date_from, date_to)
    _tsk_nw_dates = _tsk_leave_map.get(employee_id, set())

    today = get_indian_date()
    planned = 0
    completed = 0
    pending = 0
    overdue = 0

    activities = []

    for task in tasks:
        # Skip tasks whose due date falls on a non-working day
        due_d = _safe_date(task.due_date) if task.due_date else _safe_date(task.created_at)
        if due_d and due_d in _tsk_nw_dates:
            continue

        planned += 1

        is_overdue = False
        if task.due_date:
            due_date_val = task.due_date.date() if hasattr(task.due_date, 'date') else task.due_date
            if due_date_val < today and task.status not in ['completed', 'cancelled']:
                is_overdue = True
                overdue += 1
        
        if task.status == 'completed':
            completed += 1
        elif task.status not in ['cancelled']:
            pending += 1
        
        assignees = []
        if task.primary_assignee:
            assignees.append({
                "name": task.primary_assignee.full_name,
                "emp_code": task.primary_assignee.emp_code,
                "role": "primary"
            })
        if task.secondary_assignees:
            for sa in task.secondary_assignees:
                if sa.employee:
                    assignees.append({
                        "name": sa.employee.full_name,
                        "emp_code": sa.employee.emp_code,
                        "role": sa.role or "secondary"
                    })
        
        activities.append({
            "id": task.id,
            "task_code": task.task_code,
            "title": task.title,
            "status": task.status,
            "priority": task.priority,
            "due_date": task.due_date.isoformat() if task.due_date else None,
            "is_overdue": is_overdue,
            "progress_percent": task.progress or 0,
            "remarks": task.completion_notes or "",
            "assignees": assignees,
            "total_assignees": len(assignees)
        })
    
    total_actionable = pending + overdue + completed
    completion_percent = round((completed / total_actionable * 100) if total_actionable > 0 else 0)
    
    from datetime import timedelta
    # DC: Build non-working day map for datewise display (Sunday + leave + holiday)
    from app.utils.leave_utils import get_employee_nonworking_data as _get_nw
    _task_leave_map, _task_half_map = _get_nw(db, [employee_id], date_from, date_to)
    _task_leave_dates = _task_leave_map.get(employee_id, set())
    _task_half_dates = _task_half_map.get(employee_id, set())

    datewise = []
    current_date = date_from
    while current_date <= date_to:
        is_non_working = current_date in _task_leave_dates
        datewise.append({
            "date": current_date.isoformat(),
            "is_non_working": is_non_working,
            "is_half_day": current_date in _task_half_dates,
            **({"planned": 0, "completed": 0, "pending": 0, "overdue": 0} if is_non_working else {
                "planned": len([t for t in tasks if (t.due_date and _safe_date(t.due_date) == current_date) or
                                (not t.due_date and t.created_at and _safe_date(t.created_at) == current_date)]),
                "completed": sum(1 for t in tasks if t.status == 'completed' and (
                    (t.due_date and _safe_date(t.due_date) == current_date) or
                    (not t.due_date and t.created_at and _safe_date(t.created_at) == current_date))),
                "pending": sum(1 for t in tasks if t.status not in ['completed', 'cancelled'] and (
                    (t.due_date and _safe_date(t.due_date) == current_date) or
                    (not t.due_date and t.created_at and _safe_date(t.created_at) == current_date))),
                "overdue": sum(1 for t in tasks if t.due_date and _safe_date(t.due_date) == current_date and
                               _safe_date(t.due_date) < today and t.status not in ['completed', 'cancelled']),
            })
        })
        current_date += timedelta(days=1)
    
    return {
        "summary": {
            "planned": planned,
            "completed": completed,
            "pending": pending,
            "overdue": overdue,
            "completion_percent": completion_percent
        },
        "activities": activities[:20],
        "datewise": datewise
    }


def get_kra_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """
    Get KRA summary for date range.
    DC Protocol: Sundays, holidays, and leave days are automatically excluded from
    all counts and the datewise breakdown. Half-days are prorated at 0.5 weight.
    This is fully dynamic — retroactive HR attendance marking is picked up on every call.
    """
    from app.models.staff_kra import StaffKRATemplate
    instances = db.query(StaffKRADailyInstance).join(
        StaffKRAAssignment,
        StaffKRADailyInstance.kra_assignment_id == StaffKRAAssignment.id
    ).options(
        joinedload(StaffKRADailyInstance.kra_template)
    ).filter(
        StaffKRAAssignment.employee_id == employee_id,
        StaffKRADailyInstance.instance_date.between(date_from, date_to)
    ).all()

    # DC: Build non-working day maps — fully dynamic from live DB
    from app.utils.leave_utils import get_employee_nonworking_data
    leave_map, half_day_map = get_employee_nonworking_data(db, [employee_id], date_from, date_to)
    _leave_dates = leave_map.get(employee_id, set())
    _half_dates = half_day_map.get(employee_id, set())

    def _idate(inst):
        d = inst.instance_date
        return d.date() if hasattr(d, 'date') and callable(d.date) else d

    # Filter out instances on fully non-working days (Sunday, holiday, full leave)
    instances = [i for i in instances if _idate(i) not in _leave_dates]

    total = len(instances)
    # Half-day instances count as 0.5 toward the effective denominator
    effective_total = sum(0.5 if _idate(i) in _half_dates else 1.0 for i in instances)
    completed = sum(1 for i in instances if i.completion_status == 'completed')
    pending = sum(1 for i in instances if i.completion_status in ['pending', 'in_progress'])
    missed = sum(1 for i in instances if i.completion_status in ['skipped', 'na'])

    from app.api.v1.endpoints.staff_kra import _check_kra_delayed
    kra_items = []
    for instance in instances:
        tpl = instance.kra_template
        kra_items.append({
            "id": instance.id,
            "kra_name": tpl.title if tpl else "Unknown",
            "date": instance.instance_date.isoformat(),
            "status": instance.completion_status,
            "is_half_day": _idate(instance) in _half_dates,
            "completion_time": instance.completed_at.isoformat() if instance.completed_at else None,
            "target_time": tpl.target_time.strftime("%H:%M") if tpl and hasattr(tpl, 'target_time') and tpl.target_time else None,
            "completed_at": instance.completed_at.isoformat() if instance.completed_at else None,
            "is_delayed": _check_kra_delayed(instance, tpl),
            "remarks": instance.staff_notes
        })

    completion_percent = round((completed / effective_total * 100) if effective_total > 0 else 0)

    from datetime import timedelta
    datewise = []
    current_date = date_from
    while current_date <= date_to:
        # DC: Skip non-working days in datewise breakdown
        if current_date in _leave_dates:
            datewise.append({
                "date": current_date.isoformat(),
                "total": 0,
                "completed": 0,
                "pending": 0,
                "missed": 0,
                "pct": 0,
                "is_non_working": True,
                "kra_items": []
            })
            current_date += timedelta(days=1)
            continue
        day_instances = [i for i in instances if _idate(i) == current_date]
        day_completed = sum(1 for i in day_instances if i.completion_status == 'completed')
        day_pending = sum(1 for i in day_instances if i.completion_status in ['pending', 'in_progress'])
        day_missed = sum(1 for i in day_instances if i.completion_status in ['skipped', 'na'])
        day_effective = sum(0.5 if _idate(i) in _half_dates else 1.0 for i in day_instances)
        day_kra_items = []
        for inst in day_instances:
            tpl = inst.kra_template
            day_kra_items.append({
                "kra_name": tpl.title if tpl else "Unknown",
                "status": inst.completion_status,
                "is_half_day": _idate(inst) in _half_dates,
                "target_time": tpl.target_time.strftime("%H:%M") if tpl and hasattr(tpl, 'target_time') and tpl.target_time else None,
                "completed_at": inst.completed_at.isoformat() if inst.completed_at else None,
                "is_delayed": _check_kra_delayed(inst, tpl),
                "remarks": inst.staff_notes
            })
        datewise.append({
            "date": current_date.isoformat(),
            "total": len(day_instances),
            "effective_total": day_effective,
            "completed": day_completed,
            "pending": day_pending,
            "missed": day_missed,
            "is_non_working": False,
            "is_half_day": current_date in _half_dates,
            "pct": round(day_completed / day_effective * 100) if day_effective > 0 else 0,
            "kra_items": day_kra_items
        })
        current_date += timedelta(days=1)

    # W6: Count delayed KRA instances across the range
    delayed_count = sum(1 for i in kra_items if i.get("is_delayed"))

    return {
        "summary": {
            "total": total,
            "effective_total": effective_total,
            "completed": completed,
            "pending": pending,
            "missed": missed,
            "completion_percent": completion_percent,
            "delayed_count": delayed_count,
        },
        "items": kra_items,
        "datewise": datewise
    }


def get_attendance_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get attendance summary for date range"""
    attendances = db.query(StaffAttendance).filter(
        StaffAttendance.employee_id == employee_id,
        StaffAttendance.date.between(date_from, date_to)
    ).order_by(StaffAttendance.date.desc()).all()

    # [DC-LOGIN-TIME-FALLBACK-001] For new employees with no GPS punch, fetch portal last_login
    from app.models.staff import StaffEmployee as _SE
    _emp = db.query(_SE).filter(_SE.id == employee_id).first()
    _portal_login_str = None
    if _emp and _emp.last_login:
        from datetime import datetime as _dt, timezone as _tz
        _ist_offset = 5.5 * 3600
        _ll = _emp.last_login
        _ll_ist = _ll + __import__('datetime').timedelta(seconds=int(_ist_offset))
        if _ll_ist.date() == date.today():
            _portal_login_str = _ll_ist.strftime("%I:%M %p")
    
    total_worked_minutes = 0
    total_break_minutes = 0
    days_present = 0
    
    records = []
    for att in attendances:
        # W5: Compute worked_minutes live from timestamps when DB value is null/zero
        _brk = att.break_minutes or 0
        _worked = att.worked_minutes or 0
        if _worked == 0 and att.clock_in and att.clock_out:
            _delta_secs = (att.clock_out - att.clock_in).total_seconds()
            _worked = max(0, int(_delta_secs / 60) - _brk)
        # W5: Treat any record with clock_in as at minimum 'present' when status is null/blank
        _status = att.status
        if not _status and att.clock_in:
            _status = 'present'

        if _status == 'present':
            days_present += 1
        total_worked_minutes += _worked
        total_break_minutes += _brk

        records.append({
            "date": att.date.isoformat(),
            "clock_in": att.clock_in.strftime("%I:%M %p") if att.clock_in else None,
            "clock_out": att.clock_out.strftime("%I:%M %p") if att.clock_out else None,
            "worked_minutes": _worked,
            "break_minutes": _brk,
            "status": _status,
            "location_mode": att.location_mode
        })
    
    worked_hours = total_worked_minutes // 60
    worked_mins = total_worked_minutes % 60
    break_hours = total_break_minutes // 60
    break_mins = total_break_minutes % 60
    
    avg_per_day_minutes = total_worked_minutes // days_present if days_present > 0 else 0
    avg_hours = avg_per_day_minutes // 60
    avg_mins = avg_per_day_minutes % 60
    
    latest = records[0] if records else None
    
    from datetime import timedelta
    datewise = []
    current_date = date_from
    while current_date <= date_to:
        day_att = next((a for a in attendances if a.date == current_date), None)
        is_sunday = current_date.weekday() == 6  # DC Protocol: Sunday = non-working day
        if day_att:
            # W5: Live compute worked_minutes if DB value is null/0 but timestamps exist
            _dbrk = day_att.break_minutes or 0
            _dwrk = day_att.worked_minutes or 0
            if _dwrk == 0 and day_att.clock_in and day_att.clock_out:
                _dwrk = max(0, int((day_att.clock_out - day_att.clock_in).total_seconds() / 60) - _dbrk)
            # W5: Treat clock_in as at minimum present when status is null
            _dstatus = day_att.status
            if not _dstatus and day_att.clock_in:
                _dstatus = 'present'
            datewise.append({
                "date": current_date.isoformat(),
                "is_working_day": True,
                "status": _dstatus,
                "worked_minutes": _dwrk,
                "break_minutes": _dbrk,
                "clock_in": day_att.clock_in.strftime("%I:%M %p") if day_att.clock_in else None,
                "clock_out": day_att.clock_out.strftime("%I:%M %p") if day_att.clock_out else None,
            })
        elif is_sunday:
            datewise.append({
                "date": current_date.isoformat(),
                "is_working_day": False,
                "status": "sunday",
                "worked_minutes": 0,
                "break_minutes": 0,
                "clock_in": None,
                "clock_out": None,
            })
        else:
            datewise.append({
                "date": current_date.isoformat(),
                "is_working_day": True,
                "status": "absent",
                "worked_minutes": 0,
                "break_minutes": 0,
                "clock_in": None,
                "clock_out": None,
            })
        current_date += timedelta(days=1)
    
    hr_sheet_rows = db.query(StaffAttendanceSheet).filter(
        StaffAttendanceSheet.employee_id == employee_id,
        StaffAttendanceSheet.date.between(date_from, date_to)
    ).order_by(StaffAttendanceSheet.date.desc()).all()

    hr_sheet_records = []
    hr_lookup = {}
    for sheet in hr_sheet_rows:
        hr_sheet_records.append({
            "date": sheet.date.isoformat(),
            "attendance_status": sheet.attendance_status.value if sheet.attendance_status else None,
            "approval_status": sheet.approval_status.value if sheet.approval_status else "pending",
            "marked_hours": float(sheet.marked_hours or 0),
            "approved_hours": float(sheet.approved_hours) if sheet.approved_hours is not None else None
        })
        hr_lookup[sheet.date.isoformat()] = sheet

    _hr_present_statuses = {'present', 'half_day', 'on_duty', 'work_from_home'}
    _hr_working_override = {'present', 'on_duty', 'work_from_home'}
    for d in datewise:
        sheet = hr_lookup.get(d["date"])
        hr_status = sheet.attendance_status.value if sheet and sheet.attendance_status else None
        d["final_attendance"] = hr_status
        d["sheet_approved_hours"] = float(sheet.approved_hours) if sheet and sheet.approved_hours is not None else None
        if not d.get("is_working_day", True):
            # Tentatively non-working (e.g., Sunday with no clock-in) — check if HR overrides
            if hr_status and hr_status in (_hr_working_override | {'half_day'}):
                d["is_working_day"] = True
                d["status"] = hr_status
                if sheet and sheet.approved_hours:
                    d["worked_minutes"] = int(float(sheet.approved_hours) * 60)
        else:
            # Normal working day: apply HR enrichment
            if d["status"] == "absent" and hr_status and hr_status in _hr_present_statuses:
                d["status"] = hr_status
                if sheet and sheet.approved_hours:
                    d["worked_minutes"] = int(float(sheet.approved_hours) * 60)
            # WO mark on any day: flip to non-working
            if hr_status == 'weekend':
                d["is_working_day"] = False
                d["status"] = "weekend"

    # Recount days_present after incorporating HR data
    days_present = sum(
        1 for d in datewise
        if d.get("is_working_day", True) and d["status"] in _hr_present_statuses
    )
    # Recompute total worked minutes after HR correction
    total_worked_minutes = sum(d["worked_minutes"] for d in datewise)
    worked_hours = total_worked_minutes // 60
    worked_mins = total_worked_minutes % 60
    avg_per_day_minutes = total_worked_minutes // days_present if days_present > 0 else 0
    avg_hours = avg_per_day_minutes // 60
    avg_mins = avg_per_day_minutes % 60

    latest_sheet = hr_sheet_rows[0] if hr_sheet_rows else None
    hr_attendance_status = None
    hr_approval_status = "na"
    if latest_sheet:
        hr_attendance_status = latest_sheet.attendance_status.value if latest_sheet.attendance_status else None
        hr_approval_status = latest_sheet.approval_status.value if latest_sheet.approval_status else "pending"

    return {
        "summary": {
            "days_present": days_present,
            "total_worked_hours": f"{worked_hours}h {worked_mins}m",
            "total_worked_minutes": total_worked_minutes,
            "total_break_hours": f"{break_hours}h {break_mins}m",
            "total_break_minutes": total_break_minutes,
            "avg_worked_per_day": f"{avg_hours}h {avg_mins}m"
        },
        "today": {
            # [DC-LOGIN-TIME-FALLBACK-001] Fall back to portal login time when no GPS punch exists
            "clock_in": (latest["clock_in"] if latest else None) or _portal_login_str,
            "clock_out": latest["clock_out"] if latest else None,
            "worked": f"{(latest['worked_minutes'] or 0) // 60}h {(latest['worked_minutes'] or 0) % 60}m" if latest else "0h 0m",
            "breaks": f"{(latest['break_minutes'] or 0) // 60}h {(latest['break_minutes'] or 0) % 60}m" if latest else "0h 0m",
            "status": latest["status"] if latest else "absent",
            "hr_attendance": hr_attendance_status,
            "hr_approval": hr_approval_status
        },
        "records": records,
        "hr_sheet_records": hr_sheet_records,
        "datewise": datewise
    }


def get_timesheet_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get timesheet entries for date range"""
    entries = db.query(StaffTimesheetEntry).filter(
        StaffTimesheetEntry.employee_id == employee_id,
        StaffTimesheetEntry.date.between(date_from, date_to)
    ).order_by(StaffTimesheetEntry.date.desc(), StaffTimesheetEntry.start_time.desc()).all()
    
    total_minutes = sum(e.duration_minutes or 0 for e in entries)
    total_hours = total_minutes // 60
    total_mins = total_minutes % 60
    
    records = []
    for entry in entries:
        records.append({
            "id": entry.id,
            "date": entry.date.isoformat(),
            "start_time": entry.start_time.strftime("%I:%M %p") if entry.start_time else None,
            "end_time": entry.end_time.strftime("%I:%M %p") if entry.end_time else None,
            "duration_minutes": entry.duration_minutes or 0,
            "activity": entry.entry_type,
            "description": entry.comments,
            "approval_status": entry.status
        })
    
    from datetime import timedelta
    datewise = []
    current_date = date_from
    while current_date <= date_to:
        day_entries = [e for e in entries if e.date == current_date]
        day_minutes = sum(e.duration_minutes or 0 for e in day_entries)
        day_hrs = day_minutes // 60
        day_mins = day_minutes % 60
        kra_min = sum(e.duration_minutes or 0 for e in day_entries if e.entry_type == 'kra')
        act_min = sum(e.duration_minutes or 0 for e in day_entries if e.entry_type in ('task', 'lead', 'journey'))
        oth_min = sum(e.duration_minutes or 0 for e in day_entries if e.entry_type == 'others')
        sub_min = sum(e.duration_minutes or 0 for e in day_entries if e.status in ('submitted', 'approved'))
        apr_min = sum(e.duration_minutes or 0 for e in day_entries if e.status == 'approved')
        datewise.append({
            "date": current_date.isoformat(),
            "entries_count": len(day_entries),
            "total_minutes": day_minutes,
            "total_hours": f"{day_hrs}h {day_mins}m",
            "kra_minutes": kra_min,
            "activity_minutes": act_min,
            "other_minutes": oth_min,
            "submitted_minutes": sub_min,
            "approved_minutes": apr_min,
            "has_submitted": any(e.status in ('submitted', 'approved') for e in day_entries),
        })
        current_date += timedelta(days=1)

    total_submitted = sum(e.duration_minutes or 0 for e in entries if e.status in ('submitted','approved'))
    total_approved = sum(e.duration_minutes or 0 for e in entries if e.status == 'approved')
    total_kra_min = sum(e.duration_minutes or 0 for e in entries if e.entry_type == 'kra')
    total_act_min = sum(e.duration_minutes or 0 for e in entries if e.entry_type in ('task','lead','journey'))
    total_oth_min = sum(e.duration_minutes or 0 for e in entries if e.entry_type == 'others')

    return {
        "summary": {
            "count": len(entries),
            "total_minutes": total_minutes,
            "total_hours": f"{total_hours}h {total_mins}m",
            "submitted_minutes": total_submitted,
            "submitted_hours": f"{total_submitted//60}h {total_submitted%60}m",
            "approved_minutes": total_approved,
            "approved_hours": f"{total_approved//60}h {total_approved%60}m",
            "kra_minutes": total_kra_min,
            "activity_minutes": total_act_min,
            "other_minutes": total_oth_min,
        },
        "entries": records,
        "datewise": datewise
    }


def get_leads_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get CRM leads summary for sales/crm departments with Company/Self breakdown
    
    DC Protocol: Optimized with GROUP BY aggregation to avoid N+1 queries
    """
    from datetime import timedelta
    
    base_filter = or_(
        CRMLead.telecaller_id == employee_id,
        CRMLead.field_staff_id == employee_id
    )
    base_query = db.query(CRMLead).filter(base_filter)
    
    # DC Protocol (Feb 2026): Self Leads = leads with source='Self Lead'
    # Company Leads = (total assigned leads) - (self leads) = all other sources
    from app.models.crm import SELF_LEAD_SOURCE_NAME
    self_filter = CRMLead.source == SELF_LEAD_SOURCE_NAME
    
    total_assigned = base_query.count()
    date_filter = func.date(CRMLead.created_at).between(date_from, date_to)
    new_leads = base_query.filter(date_filter).count()
    
    contacted_filter = func.date(CRMLead.last_contact_date).between(date_from, date_to)
    handled_today = base_query.filter(contacted_filter).count()
    
    won_filter = and_(
        CRMLead.status == 'won',
        func.date(CRMLead.actual_close_date).between(date_from, date_to)
    )
    won_leads = base_query.filter(won_filter).count()
    
    lost_filter = and_(
        CRMLead.status == 'lost',
        func.date(CRMLead.updated_at).between(date_from, date_to)
    )
    lost_leads = base_query.filter(lost_filter).count()

    # DC Protocol (Apr 2026): Lost revenue = deal_value_total of leads marked lost in range
    lost_revenue = db.query(func.sum(CRMLead.deal_value_total)).filter(
        base_filter,
        CRMLead.status == 'lost',
        CRMLead.deal_value_total > 0,
        func.date(CRMLead.updated_at).between(date_from, date_to)
    ).scalar() or 0
    
    revenue = db.query(func.sum(CRMLead.deal_value_total)).filter(
        base_filter,
        CRMLead.status == 'won',
        func.date(CRMLead.actual_close_date).between(date_from, date_to)
    ).scalar() or 0

    # Deal value logged/updated in range (any lead where deal_value_total > 0 and updated_at in range)
    deal_value_updated = db.query(func.sum(CRMLead.deal_value_total)).filter(
        base_filter,
        CRMLead.deal_value_total > 0,
        func.date(CRMLead.updated_at).between(date_from, date_to)
    ).scalar() or 0
    # Transaction value received (deal_value_received updated in range)
    deal_value_received_range = db.query(func.sum(CRMLead.deal_value_received)).filter(
        base_filter,
        CRMLead.deal_value_received > 0,
        func.date(CRMLead.updated_at).between(date_from, date_to)
    ).scalar() or 0
    # Confirmed final value (locked payout base) — sum for all leads with CFV locked
    confirmed_final_value_sum = db.query(func.sum(CRMLead.confirmed_final_value)).filter(
        base_filter,
        CRMLead.confirmed_final_value != None,
    ).scalar() or 0
    
    # Self leads counts (source = 'Self Lead')
    self_total = base_query.filter(self_filter).count()
    self_new = base_query.filter(self_filter, date_filter).count()
    self_won = base_query.filter(self_filter, won_filter).count()
    self_lost = base_query.filter(self_filter, lost_filter).count()
    self_revenue = db.query(func.sum(CRMLead.deal_value_total)).filter(
        base_filter, self_filter, CRMLead.status == 'won',
        func.date(CRMLead.actual_close_date).between(date_from, date_to)
    ).scalar() or 0
    
    # DC Protocol (Feb 2026): Company Leads = Total assigned - Self Leads (subtraction approach)
    # This ensures company_leads + self_leads = total_assigned
    company_total = total_assigned - self_total
    company_new = new_leads - self_new
    company_won = won_leads - self_won
    company_lost = lost_leads - self_lost
    company_revenue = float(revenue) - float(self_revenue)
    
    # For datewise calculations, use filter-based approach
    is_self_created = case(
        (self_filter, 1),
        else_=0
    )
    # Company filter = NOT self (for datewise GROUP BY)
    company_filter = or_(
        CRMLead.source != SELF_LEAD_SOURCE_NAME,
        CRMLead.source.is_(None)
    )
    is_company_created = case(
        (company_filter, 1),
        else_=0
    )
    
    new_stats = db.query(
        func.date(CRMLead.created_at).label('day'),
        func.sum(is_self_created).label('self_new'),
        func.sum(is_company_created).label('company_new')
    ).filter(
        base_filter,
        func.date(CRMLead.created_at).between(date_from, date_to)
    ).group_by(func.date(CRMLead.created_at)).all()
    
    won_stats = db.query(
        func.date(CRMLead.actual_close_date).label('day'),
        func.sum(case((self_filter, 1), else_=0)).label('self_won'),
        func.sum(case((company_filter, 1), else_=0)).label('company_won'),
        func.sum(case((self_filter, CRMLead.deal_value_total), else_=0)).label('self_revenue'),
        func.sum(case((company_filter, CRMLead.deal_value_total), else_=0)).label('company_revenue')
    ).filter(
        base_filter,
        CRMLead.status == 'won',
        func.date(CRMLead.actual_close_date).between(date_from, date_to)
    ).group_by(func.date(CRMLead.actual_close_date)).all()
    
    lost_stats = db.query(
        func.date(CRMLead.updated_at).label('day'),
        func.sum(case((self_filter, 1), else_=0)).label('self_lost'),
        func.sum(case((company_filter, 1), else_=0)).label('company_lost')
    ).filter(
        base_filter,
        CRMLead.status == 'lost',
        func.date(CRMLead.updated_at).between(date_from, date_to)
    ).group_by(func.date(CRMLead.updated_at)).all()
    
    new_by_day = {row.day: {'self_new': int(row.self_new or 0), 'company_new': int(row.company_new or 0)} for row in new_stats}
    won_by_day = {row.day: {
        'self_won': int(row.self_won or 0), 
        'company_won': int(row.company_won or 0),
        'self_revenue': float(row.self_revenue or 0),
        'company_revenue': float(row.company_revenue or 0)
    } for row in won_stats}
    lost_by_day = {row.day: {'self_lost': int(row.self_lost or 0), 'company_lost': int(row.company_lost or 0)} for row in lost_stats}
    
    datewise = []
    current_date = date_from
    while current_date <= date_to:
        new_data = new_by_day.get(current_date, {'self_new': 0, 'company_new': 0})
        won_data = won_by_day.get(current_date, {'self_won': 0, 'company_won': 0, 'self_revenue': 0.0, 'company_revenue': 0.0})
        lost_data = lost_by_day.get(current_date, {'self_lost': 0, 'company_lost': 0})
        
        datewise.append({
            "date": current_date.isoformat(),
            "company_new": new_data['company_new'],
            "company_won": won_data['company_won'],
            "company_lost": lost_data['company_lost'],
            "company_revenue": won_data['company_revenue'],
            "self_new": new_data['self_new'],
            "self_won": won_data['self_won'],
            "self_lost": lost_data['self_lost'],
            "self_revenue": won_data['self_revenue']
        })
        current_date += timedelta(days=1)

    # Day-wise deal value updated (sum of deal_value_total where lead updated in range)
    _dvu_rows = db.query(
        func.date(CRMLead.updated_at).label('day'),
        func.sum(CRMLead.deal_value_total).label('dv_total')
    ).filter(
        base_filter,
        CRMLead.deal_value_total > 0,
        func.date(CRMLead.updated_at).between(date_from, date_to)
    ).group_by(func.date(CRMLead.updated_at)).all()
    _dvu_by_day = {str(r.day): float(r.dv_total or 0) for r in _dvu_rows}
    # Backfill deal_value_updated into datewise
    for _dw in datewise:
        _dw['deal_value_updated'] = _dvu_by_day.get(_dw['date'], 0.0)

    # Day-wise CRM transaction amounts
    _txn_rows = db.query(
        func.date(CRMLeadTransaction.transaction_date).label('day'),
        func.count(CRMLeadTransaction.id).label('cnt'),
        func.sum(CRMLeadTransaction.amount).label('amt')
    ).filter(
        CRMLeadTransaction.collected_by_id == employee_id,
        func.date(CRMLeadTransaction.transaction_date).between(date_from, date_to),
        CRMLeadTransaction.validation_status != 'rejected'
    ).group_by(func.date(CRMLeadTransaction.transaction_date)).all()
    _txn_by_day = {str(r.day): {'count': int(r.cnt or 0), 'amount': float(r.amt or 0)} for r in _txn_rows}
    _crm_txn_dw = []
    _cur2 = date_from
    while _cur2 <= date_to:
        _d2 = _cur2.isoformat()
        _t = _txn_by_day.get(_d2, {'count': 0, 'amount': 0.0})
        _crm_txn_dw.append({"date": _d2, "transaction_count": _t['count'], "amount": _t['amount']})
        _cur2 += timedelta(days=1)

    return {
        "summary": {
            "total_assigned": total_assigned,
            "new_created": new_leads,
            "handled_today": handled_today,
            "won": won_leads,
            "lost": lost_leads,
            "revenue": float(revenue),
            "lost_revenue": float(lost_revenue),
            "deal_value_updated": float(deal_value_updated),
            "deal_value_received": float(deal_value_received_range),
            "confirmed_final_value_sum": float(confirmed_final_value_sum),
            "company_leads": {
                "total": company_total,
                "new": company_new,
                "won": company_won,
                "lost": company_lost,
                "revenue": float(company_revenue)
            },
            "self_leads": {
                "total": self_total,
                "new": self_new,
                "won": self_won,
                "lost": self_lost,
                "revenue": float(self_revenue)
            }
        },
        "datewise": datewise,
        "crm_txn_datewise": _crm_txn_dw
    }


def get_calls_range_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Day-wise call log stats (totals + per-day breakdown) for date range reports."""
    from app.models.call_tracking import StaffCallLog
    from datetime import timedelta
    rows = db.query(
        StaffCallLog.call_date,
        func.count(StaffCallLog.id).label('total_calls'),
        func.sum(StaffCallLog.duration_seconds).label('total_seconds')
    ).filter(
        StaffCallLog.staff_id == employee_id,
        StaffCallLog.call_date.between(str(date_from), str(date_to))
    ).group_by(StaffCallLog.call_date).all()
    daily_map = {str(r.call_date): {'total_calls': int(r.total_calls or 0), 'total_seconds': int(r.total_seconds or 0)} for r in rows}
    total_calls = sum(v['total_calls'] for v in daily_map.values())
    total_seconds = sum(v['total_seconds'] for v in daily_map.values())
    total_h, total_m = total_seconds // 3600, (total_seconds % 3600) // 60
    datewise = []
    cur = date_from
    while cur <= date_to:
        d = cur.isoformat()
        day = daily_map.get(d, {'total_calls': 0, 'total_seconds': 0})
        secs = day['total_seconds']
        h, m = secs // 3600, (secs % 3600) // 60
        datewise.append({"date": d, "total_calls": day['total_calls'], "total_seconds": secs, "formatted": f"{h}h {m}m" if secs > 0 else "0h 0m"})
        cur += timedelta(days=1)
    return {
        "summary": {"total_calls": total_calls, "total_seconds": total_seconds, "total_talk_time": f"{total_h}h {total_m}m"},
        "datewise": datewise,
    }


def get_service_tickets_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get service tickets summary for service departments
    
    DC Protocol: Optimized with GROUP BY aggregation to avoid N+1 queries
    """
    from datetime import timedelta
    
    base_filter = or_(
        ServiceTicket.service_manager_id == employee_id,
        ServiceTicket.service_technician_id == employee_id
    )
    base_query = db.query(ServiceTicket).filter(base_filter)
    
    total = base_query.count()
    
    new_filter = func.date(ServiceTicket.created_date).between(date_from, date_to)
    new_tickets = base_query.filter(new_filter).count()
    
    old_tickets = total - new_tickets
    
    resolved_filter = and_(
        ServiceTicket.status == 'Closed',
        func.date(ServiceTicket.closed_date).between(date_from, date_to)
    )
    resolved = base_query.filter(resolved_filter).count()
    
    pending = base_query.filter(
        ServiceTicket.sub_status.in_(['new', 'acknowledged', 'diagnosing', 'awaiting_spares', 'procurement_in_progress', 'ready_for_work', 'work_complete', 'closed'])
    ).count()
    
    today = get_indian_date()
    overdue = base_query.filter(
        func.date(ServiceTicket.tat_due_at) < today,
        ServiceTicket.status.notin_(['Closed', 'Resolved'])
    ).count()
    
    awaiting_spares = base_query.filter(
        ServiceTicket.sub_status == 'awaiting_spares'
    ).count()
    
    new_by_day_query = db.query(
        func.date(ServiceTicket.created_date).label('day'),
        func.count(ServiceTicket.id).label('count')
    ).filter(
        base_filter,
        func.date(ServiceTicket.created_date).between(date_from, date_to)
    ).group_by(func.date(ServiceTicket.created_date)).all()
    
    resolved_by_day_query = db.query(
        func.date(ServiceTicket.closed_date).label('day'),
        func.count(ServiceTicket.id).label('count')
    ).filter(
        base_filter,
        ServiceTicket.status == 'Closed',
        func.date(ServiceTicket.closed_date).between(date_from, date_to)
    ).group_by(func.date(ServiceTicket.closed_date)).all()
    
    # TAT compliance: tickets closed where closed_date <= tat_due_at
    within_tat_q = db.query(ServiceTicket).filter(
        base_filter,
        ServiceTicket.status == 'Closed',
        func.date(ServiceTicket.closed_date).between(date_from, date_to),
        ServiceTicket.tat_due_at.isnot(None),
        ServiceTicket.closed_date <= ServiceTicket.tat_due_at
    ).count()
    out_of_tat = resolved - within_tat_q if resolved > within_tat_q else 0

    within_tat_day_q = db.query(
        func.date(ServiceTicket.closed_date).label('day'),
        func.count(ServiceTicket.id).label('count')
    ).filter(
        base_filter,
        ServiceTicket.status == 'Closed',
        func.date(ServiceTicket.closed_date).between(date_from, date_to),
        ServiceTicket.tat_due_at.isnot(None),
        ServiceTicket.closed_date <= ServiceTicket.tat_due_at
    ).group_by(func.date(ServiceTicket.closed_date)).all()
    within_tat_by_day = {row.day: row.count for row in within_tat_day_q}

    new_by_day = {row.day: row.count for row in new_by_day_query}
    resolved_by_day = {row.day: row.count for row in resolved_by_day_query}

    datewise = []
    current_date = date_from
    while current_date <= date_to:
        res = resolved_by_day.get(current_date, 0)
        wtat = within_tat_by_day.get(current_date, 0)
        datewise.append({
            "date": current_date.isoformat(),
            "new": new_by_day.get(current_date, 0),
            "resolved": res,
            "within_tat": wtat,
            "out_of_tat": max(0, res - wtat),
        })
        current_date += timedelta(days=1)

    # W3: FTD resolved count (for when function is called as FTD scope = today only)
    ftd_resolved = resolved  # tickets closed in this date range
    ftd_pending_active = pending  # globally active/pending tickets
    ftd_resolved_pct = round((ftd_resolved / new_tickets * 100) if new_tickets > 0 else 0)
    resolved_pct = round((resolved / new_tickets * 100) if new_tickets > 0 else 0)

    return {
        "summary": {
            "total": total,
            "old": old_tickets,
            "new": new_tickets,
            "resolved": resolved,
            "resolved_pct": resolved_pct,
            "pending": pending,
            "overdue": overdue,
            "awaiting_spares": awaiting_spares,
            "within_tat": within_tat_q,
            "out_of_tat": out_of_tat,
            "ftd_resolved": ftd_resolved,
            "ftd_pending": ftd_pending_active,
            "ftd_resolved_pct": ftd_resolved_pct,
        },
        "datewise": datewise
    }


def get_travel_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get travel/journey summary from field work sessions"""
    sessions = db.query(StaffFieldWorkSession).filter(
        StaffFieldWorkSession.employee_id == employee_id,
        func.date(StaffFieldWorkSession.session_start).between(date_from, date_to)
    ).all()
    
    total_km = sum(float(s.total_km or 0) for s in sessions)
    total_minutes = 0
    
    for session in sessions:
        if session.session_start and session.session_end:
            delta = session.session_end - session.session_start
            total_minutes += delta.total_seconds() / 60
    
    hours = int(total_minutes // 60)
    mins = int(total_minutes % 60)
    
    session_records = []
    for s in sessions:
        session_date = s.session_start.date() if s.session_start else None
        duration_mins = 0
        if s.session_start and s.session_end:
            duration_mins = int((s.session_end - s.session_start).total_seconds() / 60)
        session_records.append({
            "date": session_date.isoformat() if session_date else None,
            "sessions": 1,
            "distance_km": float(s.total_km or 0),
            "duration": f"{duration_mins // 60}h {duration_mins % 60}m",
            "transport_mode": s.transport_mode or 'N/A'
        })
    
    from datetime import timedelta
    datewise = []
    current_date = date_from
    while current_date <= date_to:
        day_sessions = [s for s in sessions if s.session_start and s.session_start.date() == current_date]
        day_km = sum(float(s.total_km or 0) for s in day_sessions)
        day_mins = 0
        for s in day_sessions:
            if s.session_start and s.session_end:
                day_mins += int((s.session_end - s.session_start).total_seconds() / 60)
        datewise.append({
            "date": current_date.isoformat(),
            "sessions": len(day_sessions),
            "km_travelled": round(day_km, 2),
            "duration_minutes": day_mins,
            "duration": f"{day_mins // 60}h {day_mins % 60}m"
        })
        current_date += timedelta(days=1)
    
    return {
        "summary": {
            "total_km": round(total_km, 2),
            "total_time": f"{hours}h {mins}m",
            "total_minutes": int(total_minutes),
            "session_count": len(sessions)
        },
        "sessions": session_records,
        "datewise": datewise
    }


def get_calls_summary(db: Session, employee_id: int, date_from: date, date_to: date) -> dict:
    """Get call log summary from staff_call_logs for the date range.
    call_date is stored as a VARCHAR 'YYYY-MM-DD' string.
    staff_call_logs uses staff_id = employee_id.
    """
    from app.models.call_tracking import StaffCallLog
    from datetime import timedelta

    date_from_str = date_from.isoformat()
    date_to_str = date_to.isoformat()

    logs = db.query(StaffCallLog).filter(
        StaffCallLog.staff_id == employee_id,
        StaffCallLog.call_date >= date_from_str,
        StaffCallLog.call_date <= date_to_str
    ).all()

    total_calls = len(logs)
    outgoing = sum(1 for l in logs if l.call_type == 'OUTGOING')
    incoming = sum(1 for l in logs if l.call_type == 'INCOMING')
    missed = sum(1 for l in logs if l.call_type == 'MISSED')
    rejected = sum(1 for l in logs if l.call_type == 'REJECTED')
    total_seconds = sum(l.duration_seconds or 0 for l in logs)
    talk_seconds = sum(l.duration_seconds or 0 for l in logs if l.call_type in ('OUTGOING', 'INCOMING'))

    def _fmt(secs):
        h = secs // 3600
        m = (secs % 3600) // 60
        s = secs % 60
        return f"{h}h {m}m {s}s" if h else f"{m}m {s}s"

    datewise = []
    current_date = date_from
    while current_date <= date_to:
        ds = current_date.isoformat()
        day_logs = [l for l in logs if l.call_date == ds]
        day_sec = sum(l.duration_seconds or 0 for l in day_logs)
        day_talk = sum(l.duration_seconds or 0 for l in day_logs if l.call_type in ('OUTGOING', 'INCOMING'))
        datewise.append({
            "date": ds,
            "total": len(day_logs),
            "outgoing": sum(1 for l in day_logs if l.call_type == 'OUTGOING'),
            "incoming": sum(1 for l in day_logs if l.call_type == 'INCOMING'),
            "missed": sum(1 for l in day_logs if l.call_type == 'MISSED'),
            "rejected": sum(1 for l in day_logs if l.call_type == 'REJECTED'),
            "total_duration": _fmt(day_sec),
            "talk_time": _fmt(day_talk),
            "talk_seconds": day_talk,
        })
        current_date += timedelta(days=1)

    return {
        "summary": {
            "total_calls": total_calls,
            "outgoing": outgoing,
            "incoming": incoming,
            "missed": missed,
            "rejected": rejected,
            "total_duration": _fmt(total_seconds),
            "talk_time": _fmt(talk_seconds),
            "talk_seconds": talk_seconds,
        },
        "datewise": datewise
    }


@router.get("/team-members", summary="Get downline team members for filter dropdown")
async def get_team_members(
    current_user: StaffEmployee = Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    """
    DC Protocol: Get list of downline team members for managers
    Used to populate the employee filter dropdown
    """
    if not is_manager_or_leadership(current_user):
        return {
            "success": True,
            "data": [],
            "message": "No team members (not a manager)"
        }
    
    downline_ids = get_downline_employee_ids(db, current_user.id, recursive=True)
    
    if not downline_ids:
        return {
            "success": True,
            "data": [],
            "message": "No direct reports found"
        }
    
    hidden_codes = HIDDEN_FROM_TEAM_CODES or []

    employees = db.query(StaffEmployee).filter(
        StaffEmployee.id.in_(downline_ids),
        StaffEmployee.status == 'active'
    ).order_by(StaffEmployee.full_name).all()
    
    return {
        "success": True,
        "data": [
            {
                "id": emp.id,
                "emp_code": emp.emp_code,
                "full_name": emp.full_name,
                "role": emp.role.role_name if emp.role else None,
                "department": emp.department.name if emp.department else None
            }
            for emp in employees
            if emp.emp_code not in hidden_codes
        ]
    }




@router.get("/team-daily-compliance", summary="Day-wise team presence and compliance — Key Leadership/EA only")
def get_team_daily_compliance(
    date_from: date = Query(..., description="Start date"),
    date_to: date = Query(..., description="End date"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user),
):
    """
    DC Protocol: Returns day-wise team compliance for Key Leadership and EA roles.
    For each day in [date_from, date_to]:
      - total_employees: active downline count
      - present: HR attendance = present / half_day / on_duty
      - kra_complied: employees who completed all applicable KRAs for that day
      - plan_complied: employees who submitted/finalized their day planner
      - timesheet_complied: employees with at least one timesheet entry
      - overall_complied: employees who met ALL three (KRA + plan + timesheet)
    Sundays are marked as non-working and skipped from counts.
    Dynamic — retroactive HR marking automatically reflected on each API call.
    """
    if not is_key_leadership(current_user) and not has_unrestricted_access(current_user):
        raise HTTPException(status_code=403, detail="Key Leadership or EA access required")

    if (date_to - date_from).days > 90:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 90 days")

    # Get all active downline employees (or all org for VGK Supreme)
    if has_unrestricted_access(current_user):
        hidden_codes = HIDDEN_FROM_TEAM_CODES or []
        all_emp_q = db.query(StaffEmployee).options(
            joinedload(StaffEmployee.role),
            joinedload(StaffEmployee.department),
        ).filter(StaffEmployee.status == 'active')
        if hidden_codes:
            all_emp_q = all_emp_q.filter(~StaffEmployee.emp_code.in_(hidden_codes))
        all_employees = all_emp_q.all()
    else:
        downline_ids = get_downline_employee_ids(db, current_user.id, recursive=True)
        hidden_codes = HIDDEN_FROM_TEAM_CODES or []
        all_employees = db.query(StaffEmployee).options(
            joinedload(StaffEmployee.role),
            joinedload(StaffEmployee.department),
        ).filter(
            StaffEmployee.id.in_(downline_ids),
            StaffEmployee.status == 'active',
        ).all()
        if hidden_codes:
            all_employees = [e for e in all_employees if e.emp_code not in hidden_codes]

    emp_ids = [e.id for e in all_employees]
    total_employees = len(emp_ids)

    if not emp_ids:
        return {"success": True, "total_employees": 0, "days": []}

    # ── HR Attendance (presence per day) ────────────────────────────────────
    from app.models.staff_attendance_sheet import StaffAttendanceSheet
    PRESENT_STATUSES = {'present', 'half_day', 'on_duty'}
    att_rows = db.query(
        StaffAttendanceSheet.employee_id,
        StaffAttendanceSheet.date,
        StaffAttendanceSheet.attendance_status,
    ).filter(
        StaffAttendanceSheet.employee_id.in_(emp_ids),
        StaffAttendanceSheet.date.between(date_from, date_to),
    ).all()

    # present_by_day[date] = set of employee_ids
    from collections import defaultdict
    present_by_day: dict = defaultdict(set)
    wo_by_day: dict = defaultdict(set)
    for row in att_rows:
        val = row.attendance_status.value if hasattr(row.attendance_status, 'value') else str(row.attendance_status)
        if val in PRESENT_STATUSES:
            present_by_day[row.date].add(row.employee_id)
        elif val == 'weekend':
            wo_by_day[row.date].add(row.employee_id)

    # ── KRA Instances — build day-wise: complied = all pending KRAs completed ──
    kra_rows = db.query(
        StaffKRADailyInstance.employee_id,
        StaffKRADailyInstance.instance_date,
        StaffKRADailyInstance.completion_status,
    ).filter(
        StaffKRADailyInstance.employee_id.in_(emp_ids),
        StaffKRADailyInstance.instance_date.between(date_from, date_to),
    ).all()

    # kra_by_emp_day[(emp_id, date)] = {'total': N, 'completed': M}
    kra_by_emp_day: dict = defaultdict(lambda: {'total': 0, 'completed': 0})
    for row in kra_rows:
        inst_d = row.instance_date.date() if hasattr(row.instance_date, 'date') and callable(row.instance_date.date) else row.instance_date
        key = (row.employee_id, inst_d)
        kra_by_emp_day[key]['total'] += 1
        if row.completion_status == 'completed':
            kra_by_emp_day[key]['completed'] += 1

    # ── Day Plans — submitted/finalized per day ──────────────────────────────
    plan_rows = db.query(
        StaffDayPlan.employee_id,
        StaffDayPlan.plan_date,
        StaffDayPlan.status,
        StaffDayPlan.finalized_at,
    ).filter(
        StaffDayPlan.employee_id.in_(emp_ids),
        StaffDayPlan.plan_date.between(date_from, date_to),
    ).all()

    # plan_done_by_day[date] = set of employee_ids who finalized/closed plan
    plan_done_by_day: dict = defaultdict(set)
    for row in plan_rows:
        pd = row.plan_date.date() if hasattr(row.plan_date, 'date') and callable(row.plan_date.date) else row.plan_date
        if row.finalized_at or row.status in ('closed', 'submitted', 'approved'):
            plan_done_by_day[pd].add(row.employee_id)

    # ── Timesheet — at least one entry per day ───────────────────────────────
    ts_rows = db.query(
        StaffTimesheetEntry.employee_id,
        StaffTimesheetEntry.date,
    ).filter(
        StaffTimesheetEntry.employee_id.in_(emp_ids),
        StaffTimesheetEntry.date.between(date_from, date_to),
    ).all()

    ts_by_day: dict = defaultdict(set)
    for row in ts_rows:
        ts_d = row.date.date() if hasattr(row.date, 'date') and callable(row.date.date) else row.date
        ts_by_day[ts_d].add(row.employee_id)

    # ── Build day-wise result ────────────────────────────────────────────────
    days = []
    current_date = date_from
    while current_date <= date_to:
        is_sunday = (current_date.weekday() == 6)
        day_names_map = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        day_name = day_names_map[current_date.weekday()]

        _present_count = len(present_by_day.get(current_date, set()))
        if is_sunday and _present_count == 0:
            days.append({
                "date": current_date.isoformat(),
                "day": day_name,
                "is_sunday": True,
                "is_working_day": False,
                "total_employees": total_employees,
                "present": 0,
                "kra_complied": 0,
                "plan_complied": 0,
                "timesheet_complied": 0,
                "overall_complied": 0,
                "present_pct": 0,
                "kra_pct": 0,
                "plan_pct": 0,
                "ts_pct": 0,
                "overall_pct": 0,
            })
            current_date += timedelta(days=1)
            continue

        # Present employees on this day
        present_emps = present_by_day.get(current_date, set())
        present_count = len(present_emps)

        # KRA complied: employee has KRA instances AND all are completed
        kra_complied = 0
        plan_complied = len(plan_done_by_day.get(current_date, set()))
        ts_complied = len(ts_by_day.get(current_date, set()))
        overall_complied = 0

        for eid in emp_ids:
            kd = kra_by_emp_day.get((eid, current_date))
            if kd and kd['total'] > 0 and kd['completed'] >= kd['total']:
                kra_complied += 1

        # Overall: present + KRA + plan + timesheet all done
        plan_done_set = plan_done_by_day.get(current_date, set())
        ts_done_set = ts_by_day.get(current_date, set())
        for eid in emp_ids:
            kd = kra_by_emp_day.get((eid, current_date))
            kra_ok = kd and kd['total'] > 0 and kd['completed'] >= kd['total']
            plan_ok = eid in plan_done_set
            ts_ok = eid in ts_done_set
            if kra_ok and plan_ok and ts_ok:
                overall_complied += 1

        def _pct(n, d):
            return round(n / d * 100) if d > 0 else 0

        days.append({
            "date": current_date.isoformat(),
            "day": day_name,
            "is_sunday": False,
            "is_working_day": True,
            "total_employees": total_employees,
            "present": present_count,
            "kra_complied": kra_complied,
            "plan_complied": plan_complied,
            "timesheet_complied": ts_complied,
            "overall_complied": overall_complied,
            "present_pct": _pct(present_count, total_employees),
            "kra_pct": _pct(kra_complied, total_employees),
            "plan_pct": _pct(plan_complied, total_employees),
            "ts_pct": _pct(ts_complied, total_employees),
            "overall_pct": _pct(overall_complied, total_employees),
        })
        current_date += timedelta(days=1)

    working_days = [d for d in days if d.get('is_working_day')]
    avg_present_pct = round(sum(d['present_pct'] for d in working_days) / len(working_days)) if working_days else 0
    avg_kra_pct = round(sum(d['kra_pct'] for d in working_days) / len(working_days)) if working_days else 0
    avg_plan_pct = round(sum(d['plan_pct'] for d in working_days) / len(working_days)) if working_days else 0
    avg_overall_pct = round(sum(d['overall_pct'] for d in working_days) / len(working_days)) if working_days else 0

    return {
        "success": True,
        "total_employees": total_employees,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "summary": {
            "avg_present_pct": avg_present_pct,
            "avg_kra_pct": avg_kra_pct,
            "avg_plan_pct": avg_plan_pct,
            "avg_overall_pct": avg_overall_pct,
        },
        "days": days,
    }


@router.get("/team-compliance", summary="Team A/B/C compliance — KL/EA only")
def get_team_compliance(
    target_date: date = Query(default=None, description="Date for FTD (defaults to today)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user),
):
    """
    W2: Team compliance grouped by team_tag (Team A/B/C).
    Compliance = Present + KRA completed + Timesheet submitted + Day Closure done.
    Returns FTD (target_date) and MTD (1st of month to target_date) per team.
    Accessible to Key Leadership, EA, VGK4U admin.
    """
    if not is_key_leadership(current_user) and not has_unrestricted_access(current_user):
        raise HTTPException(status_code=403, detail="Key Leadership or EA access required")

    today = get_indian_date()
    ftd = target_date or today
    mtd_start = ftd.replace(day=1)

    # All active employees with a team_tag
    all_emps = db.query(StaffEmployee).filter(
        StaffEmployee.status == 'active',
        StaffEmployee.team_tag.isnot(None)
    ).all()

    if not all_emps:
        return {"success": True, "date": ftd.isoformat(), "teams": [], "no_teams": True}

    emp_ids = [e.id for e in all_emps]
    team_map = {e.id: (e.team_tag or 'unassigned') for e in all_emps}

    # ── Helper: compliance per employee per date range ───────────────────────
    from collections import defaultdict

    def _compliance_data(d_from, d_to):
        # Attendance (HR-marked or self-clocked)
        _PRESENT = {'present', 'half_day', 'on_duty', 'work_from_home'}
        att_hr = db.query(
            StaffAttendanceSheet.employee_id,
            StaffAttendanceSheet.date,
            StaffAttendanceSheet.attendance_status,
        ).filter(
            StaffAttendanceSheet.employee_id.in_(emp_ids),
            StaffAttendanceSheet.date.between(d_from, d_to)
        ).all()
        # Also check self clock-in
        att_self = db.query(
            StaffAttendance.employee_id,
            StaffAttendance.date,
        ).filter(
            StaffAttendance.employee_id.in_(emp_ids),
            StaffAttendance.date.between(d_from, d_to),
            StaffAttendance.clock_in.isnot(None)
        ).all()
        present_set: dict = defaultdict(set)
        for r in att_hr:
            val = r.attendance_status.value if hasattr(r.attendance_status, 'value') else str(r.attendance_status or '')
            if val in _PRESENT:
                present_set[r.date].add(r.employee_id)
        for r in att_self:
            present_set[r.date].add(r.employee_id)

        # KRA (all instances completed for that day)
        kra_rows = db.query(
            StaffKRADailyInstance.employee_id,
            StaffKRADailyInstance.instance_date,
            StaffKRADailyInstance.completion_status,
        ).filter(
            StaffKRADailyInstance.employee_id.in_(emp_ids),
            StaffKRADailyInstance.instance_date.between(d_from, d_to),
        ).all()
        kra_map = defaultdict(lambda: {'total': 0, 'completed': 0})
        for r in kra_rows:
            _d = r.instance_date.date() if hasattr(r.instance_date, 'date') and callable(r.instance_date.date) else r.instance_date
            kra_map[(r.employee_id, _d)]['total'] += 1
            if r.completion_status == 'completed':
                kra_map[(r.employee_id, _d)]['completed'] += 1

        # Timesheet (at least one entry per day)
        ts_rows = db.query(
            StaffTimesheetEntry.employee_id,
            StaffTimesheetEntry.date,
        ).filter(
            StaffTimesheetEntry.employee_id.in_(emp_ids),
            StaffTimesheetEntry.date.between(d_from, d_to),
        ).all()
        ts_set: dict = defaultdict(set)
        for r in ts_rows:
            _d = r.date.date() if hasattr(r.date, 'date') and callable(r.date.date) else r.date
            ts_set[_d].add(r.employee_id)

        # Day Closure (StaffDayPlan.finalized_at is not None)
        plan_rows = db.query(
            StaffDayPlan.employee_id,
            StaffDayPlan.plan_date,
            StaffDayPlan.finalized_at,
        ).filter(
            StaffDayPlan.employee_id.in_(emp_ids),
            StaffDayPlan.plan_date.between(d_from, d_to),
            StaffDayPlan.finalized_at.isnot(None)
        ).all()
        closure_set: dict = defaultdict(set)
        for r in plan_rows:
            _d = r.plan_date.date() if hasattr(r.plan_date, 'date') and callable(r.plan_date.date) else r.plan_date
            closure_set[_d].add(r.employee_id)

        # Build per-employee compliance across the date range
        compliant_set = set()
        present_any = set()
        kra_ok_set = set()
        ts_ok_set = set()
        closure_ok_set = set()

        _cur = d_from
        while _cur <= d_to:
            if _cur.weekday() == 6 and not present_set.get(_cur):  # Skip Sundays with no attendance
                _cur += timedelta(days=1)
                continue
            _pres_day = present_set.get(_cur, set())
            _ts_day = ts_set.get(_cur, set())
            _cl_day = closure_set.get(_cur, set())
            for eid in emp_ids:
                _pok = eid in _pres_day
                _kok = False
                _kd = kra_map.get((eid, _cur))
                if _kd and _kd['total'] > 0 and _kd['completed'] >= _kd['total']:
                    _kok = True
                _tok = eid in _ts_day
                _dok = eid in _cl_day
                if _pok:
                    present_any.add(eid)
                if _kok:
                    kra_ok_set.add(eid)
                if _tok:
                    ts_ok_set.add(eid)
                if _dok:
                    closure_ok_set.add(eid)
                if _pok and _kok and _tok and _dok:
                    compliant_set.add(eid)
            _cur += timedelta(days=1)

        return {
            'compliant': compliant_set,
            'present': present_any,
            'kra_ok': kra_ok_set,
            'ts_ok': ts_ok_set,
            'closure_ok': closure_ok_set,
        }

    # FTD compliance
    ftd_data = _compliance_data(ftd, ftd)
    # MTD compliance
    mtd_data = _compliance_data(mtd_start, ftd)

    # Group by team
    teams_raw: dict = defaultdict(list)
    for e in all_emps:
        teams_raw[e.team_tag or 'unassigned'].append(e.id)

    def _pct(n, d):
        return round(n / d * 100) if d > 0 else 0

    teams_out = []
    for tag in sorted(teams_raw.keys()):
        t_ids = teams_raw[tag]
        t_size = len(t_ids)
        # FTD
        ftd_compliant = sum(1 for eid in t_ids if eid in ftd_data['compliant'])
        ftd_present = sum(1 for eid in t_ids if eid in ftd_data['present'])
        ftd_kra = sum(1 for eid in t_ids if eid in ftd_data['kra_ok'])
        ftd_ts = sum(1 for eid in t_ids if eid in ftd_data['ts_ok'])
        ftd_closure = sum(1 for eid in t_ids if eid in ftd_data['closure_ok'])
        # MTD
        mtd_compliant = sum(1 for eid in t_ids if eid in mtd_data['compliant'])
        mtd_present = sum(1 for eid in t_ids if eid in mtd_data['present'])
        mtd_kra = sum(1 for eid in t_ids if eid in mtd_data['kra_ok'])
        mtd_ts = sum(1 for eid in t_ids if eid in mtd_data['ts_ok'])
        mtd_closure = sum(1 for eid in t_ids if eid in mtd_data['closure_ok'])

        label_map = {'team_a': 'Team A', 'team_b': 'Team B', 'team_c': 'Team C'}
        teams_out.append({
            'team_tag': tag,
            'team_label': label_map.get(tag, tag.replace('_', ' ').title()),
            'member_count': t_size,
            'ftd': {
                'present': ftd_present,
                'present_pct': _pct(ftd_present, t_size),
                'kra_complied': ftd_kra,
                'kra_pct': _pct(ftd_kra, t_size),
                'timesheet_complied': ftd_ts,
                'ts_pct': _pct(ftd_ts, t_size),
                'closure_complied': ftd_closure,
                'closure_pct': _pct(ftd_closure, t_size),
                'fully_compliant': ftd_compliant,
                'compliance_pct': _pct(ftd_compliant, t_size),
            },
            'mtd': {
                'present': mtd_present,
                'present_pct': _pct(mtd_present, t_size),
                'kra_complied': mtd_kra,
                'kra_pct': _pct(mtd_kra, t_size),
                'timesheet_complied': mtd_ts,
                'ts_pct': _pct(mtd_ts, t_size),
                'closure_complied': mtd_closure,
                'closure_pct': _pct(mtd_closure, t_size),
                'fully_compliant': mtd_compliant,
                'compliance_pct': _pct(mtd_compliant, t_size),
            },
        })

    return {
        "success": True,
        "date": ftd.isoformat(),
        "mtd_start": mtd_start.isoformat(),
        "total_tagged_employees": len(emp_ids),
        "teams": teams_out,
    }


@router.put("/tat-settings", summary="Update TAT threshold days — EA/Key Leadership only")
def update_tat_settings(
    module_key: str = Query(..., description="Module: procurement_po or procurement_pr"),
    tat_days: int = Query(..., ge=1, le=90, description="TAT threshold in days"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user),
):
    """
    W4: Allow EA and Key Leadership (VGK mentor) to update the TAT threshold days
    for procurement modules. Default is 3 days. Changes take effect immediately.
    """
    _rc = (current_user.role.role_code if current_user.role else '').lower()
    if _rc not in ('vgk4u', 'key_leadership', 'ea'):
        raise HTTPException(status_code=403, detail="Only EA or Key Leadership can update TAT settings")

    valid_keys = ['procurement_po', 'procurement_pr']
    if module_key not in valid_keys:
        raise HTTPException(status_code=400, detail=f"module_key must be one of {valid_keys}")

    try:
        db.execute(
            text("""
                INSERT INTO staff_tat_config (module_key, tat_days, updated_by, updated_at)
                VALUES (:k, :d, :u, NOW())
                ON CONFLICT (module_key) DO UPDATE
                SET tat_days = :d, updated_by = :u, updated_at = NOW()
            """),
            {"k": module_key, "d": tat_days, "u": current_user.id}
        )
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update TAT settings: {e}")

    return {
        "success": True,
        "module_key": module_key,
        "tat_days": tat_days,
        "updated_by": current_user.full_name,
        "message": f"TAT threshold for {module_key} updated to {tat_days} days"
    }


@router.get("/ranking", summary="Get employee ranking vs all active staff")
def get_employee_ranking(
    date_from: date = Query(...),
    date_to: date = Query(...),
    employee_id: Optional[int] = Query(None),
    current_user: StaffEmployee = Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    """DC Protocol: Compute weighted performance score and rank employee vs all active staff.
    Weights: Attendance 7.5%, Day Plan 7.5%, KRA 10%, Timesheet/Activities 15%, Dept 60%.
    If a section is N/A, its weight is redistributed proportionally to applicable sections.
    """
    # Limit to max 30 days
    delta = (date_to - date_from).days
    if delta > 30:
        date_from = date_to - timedelta(days=29)

    target_id = current_user.id
    if employee_id and employee_id != current_user.id:
        if is_manager_or_leadership(current_user) or has_unrestricted_access(current_user):
            target_id = employee_id
        else:
            raise HTTPException(status_code=403, detail="Not authorised")

    target_emp = db.query(StaffEmployee).options(
        joinedload(StaffEmployee.department)
    ).filter(StaffEmployee.id == target_id).first()
    if not target_emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    dept_name = (target_emp.department.name or '').lower() if target_emp.department else ''
    is_sales = 'sales' in dept_name
    is_service = 'service' in dept_name
    has_dept_section = is_sales or is_service

    # --- Score calculator for a single employee ---
    def _compute_score(eid: int, dept_name_lower: str):
        emp_is_sales = 'sales' in dept_name_lower
        emp_is_service = 'service' in dept_name_lower
        emp_has_dept = emp_is_sales or emp_is_service

        # Attendance score
        att = get_attendance_summary(db, eid, date_from, date_to)
        total_days = (date_to - date_from).days + 1
        present = att['summary']['days_present']
        att_score = round((present / total_days) * 100) if total_days > 0 else 0

        # Day Plan score
        dp = get_day_plan_summary(db, eid, date_from, date_to)
        dp_pct = dp.get('completion_percent', 0) if dp.get('has_plan') else 0
        dp_score = dp_pct

        # KRA score
        kra = get_kra_summary(db, eid, date_from, date_to)
        kra_score = kra['summary']['completion_percent']

        # Timesheet score: % days with submitted entries
        ts = get_timesheet_summary(db, eid, date_from, date_to)
        days_submitted = sum(1 for d in ts['datewise'] if d.get('has_submitted'))
        ts_score = round((days_submitted / total_days) * 100) if total_days > 0 else 0

        # Dept score
        dept_score = 0
        if emp_is_sales:
            leads = get_leads_summary(db, eid, date_from, date_to)
            # Score: won leads ratio (new leads → won)
            new_l = leads['summary']['new_created'] or 0
            won_l = leads['summary']['won'] or 0
            dept_score = round((won_l / new_l) * 100) if new_l > 0 else (100 if won_l > 0 else 0)
        elif emp_is_service:
            svc = get_service_tickets_summary(db, eid, date_from, date_to)
            resolved = svc['summary']['resolved'] or 0
            total_t = svc['summary']['new'] or 0
            within = svc['summary'].get('within_tat', 0)
            dept_score_base = round((resolved / total_t) * 100) if total_t > 0 else 0
            tat_bonus = round((within / resolved) * 20) if resolved > 0 else 0
            dept_score = min(100, dept_score_base + tat_bonus)

        # Apply weights with N/A redistribution
        base_weights = {
            'attendance': 5.0, 'day_plan': 10.0,
            'kra': 15.0, 'timesheet': 10.0,
            'dept': 60.0 if emp_has_dept else 0.0,
        }
        scores = {
            'attendance': att_score, 'day_plan': dp_score,
            'kra': kra_score, 'timesheet': ts_score,
            'dept': dept_score,
        }
        # Redistribute dept weight if N/A
        if not emp_has_dept:
            leftover = 60.0
            applicable_keys = ['attendance', 'day_plan', 'kra', 'timesheet']
            total_base = sum(base_weights[k] for k in applicable_keys)
            for k in applicable_keys:
                base_weights[k] += leftover * (base_weights[k] / total_base)

        total_weight = sum(base_weights.values())
        weighted = sum(scores[k] * base_weights[k] / total_weight for k in scores)
        return round(weighted, 1), {k: scores[k] for k in scores}

    # DC_RANKING_CACHE: Check dual-layer cache first — in-memory (fastest, per-worker) then
    # shared file cache (cross-worker, survives restarts). Cache stores a dict {str(emp_id): score}
    # so the target employee's score can be looked up without running _compute_score on every hit.
    cache_key = (target_emp.base_company_id, date_from.isoformat(), date_to.isoformat())

    # Layer 1: in-memory cache
    cached = _RANKING_CACHE.get(cache_key)
    if cached and (time.time() - cached['ts']) < _RANKING_CACHE_TTL:
        scores_map = cached['scores']
    else:
        # Layer 2: shared file cache
        scores_map = _read_ranking_file_cache(cache_key)
        if scores_map is None:
            # Cache miss on both layers — compute all active employees
            all_emps = db.query(StaffEmployee).options(
                joinedload(StaffEmployee.department)
            ).filter(
                StaffEmployee.status == 'active',
                StaffEmployee.base_company_id == target_emp.base_company_id,
                StaffEmployee.emp_code.isnot(None),
            ).all()

            try:
                db.execute(text("SET LOCAL statement_timeout = '25000'"))
            except Exception:
                pass

            scores_map = {}
            for emp in all_emps:
                d = (emp.department.name or '').lower() if emp.department else ''
                try:
                    s, _ = _compute_score(emp.id, d)
                    scores_map[str(emp.id)] = s
                except Exception:
                    scores_map[str(emp.id)] = 0.0

            _write_ranking_file_cache(cache_key, scores_map)

        _RANKING_CACHE[cache_key] = {'scores': scores_map, 'ts': time.time()}

    # Use cached target score when available — avoids 4-5 DB queries on every cache-hit request
    str_target = str(target_id)
    if str_target in scores_map:
        target_score = scores_map[str_target]
        target_breakdown = {}  # breakdown not stored in cache; only available on cold start
    else:
        # Employee not yet in cache (new hire activated after last cache write)
        target_score, target_breakdown = _compute_score(target_id, dept_name)
        scores_map[str_target] = target_score
        _RANKING_CACHE[cache_key]['scores'] = scores_map
        _write_ranking_file_cache(cache_key, scores_map)

    all_scores = list(scores_map.values())
    total_active = len(all_scores)
    # DC Protocol: competition ranking — count how many employees scored STRICTLY higher
    rank = sum(1 for s in all_scores if s > target_score) + 1
    percentile = round(((total_active - rank) / total_active) * 100) if total_active > 0 else 0
    is_ranked = target_score > 0.0

    return {
        "success": True,
        "data": {
            "rank": rank,
            "is_ranked": is_ranked,
            "total_active_staff": total_active,
            "overall_score": target_score,
            "percentile": percentile,
            "score_breakdown": target_breakdown,
            "weights": {
                "attendance": 5.0, "day_plan": 10.0,
                "kra": 15.0, "timesheet": 10.0, "dept": 60.0,
                "dept_applicable": has_dept_section,
            },
            "dept_type": "sales" if is_sales else ("service" if is_service else "other"),
            "date_range": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        }
    }


@router.get("/executive-dashboard", summary="Executive Performance Dashboard — Key Leadership only")
def get_executive_dashboard(
    date_from: date = Query(..., description="Start date"),
    date_to: date = Query(..., description="End date"),
    department_id: Optional[int] = Query(None, description="Filter by department"),
    company_id: Optional[int] = Query(None, description="Filter by company (blank = all companies)"),
    include_self: bool = Query(False, description="Include current user (self) in table"),
    include_supreme: bool = Query(False, description="[deprecated] alias for include_self"),
    my_team_only: bool = Query(False, description="Restrict to viewer's direct downline hierarchy (non-VGK4U only)"),
    current_user: StaffEmployee = Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    """DC Protocol: Batch performance dashboard for all active staff across all companies.
    Gate: Key Leadership / EA / Executive Admin only.
    VGK4U/VGK4U Supreme: unrestricted (see all employees).
    Other key leadership: see all by default; use my_team_only=true to restrict to downline."""
    if not is_key_leadership(current_user):
        raise HTTPException(status_code=403, detail="Executive access required")
    _viewer_is_unrestricted = has_unrestricted_access(current_user)

    if (date_to - date_from).days > 30:
        date_from = date_to - timedelta(days=29)
    total_days = (date_to - date_from).days + 1
    # DC Protocol: working days are driven by HR WO markings, not calendar Sundays.
    # Per-employee working_days are computed below from att_map['wo_days'].

    # ── Employee scope ─────────────────────────────────────────────────────────
    from app.models.staff import StaffDepartment
    from app.models.staff_attendance import StaffAttendance
    from app.models.staff_tasks import StaffDayPlan
    from app.models.staff_kra import StaffKRADailyInstance
    from app.models.staff_timesheet import StaffTimesheetEntry
    from app.models.ticket import ServiceTicket
    from app.models.staff_accounts import AssociatedCompany

    emp_q = (
        db.query(StaffEmployee)
        .options(joinedload(StaffEmployee.department), joinedload(StaffEmployee.role))
        .filter(
            StaffEmployee.status == 'active',
            StaffEmployee.emp_code.isnot(None),
        )
    )
    # Optional company filter — no filter = show all companies
    if company_id:
        emp_q = emp_q.filter(StaffEmployee.base_company_id == company_id)
    if department_id:
        emp_q = emp_q.filter(StaffEmployee.department_id == department_id)
    # Exclude VGK Supreme by default; include_self=True reveals them (VGK Supreme checkbox)
    if not include_self:
        _hidden = HIDDEN_FROM_TEAM_CODES or []
        if _hidden:
            emp_q = emp_q.filter(~StaffEmployee.emp_code.in_(_hidden))

    # DC Protocol: my_team_only restricts to viewer's downline hierarchy
    # VGK4U/VGK4U Supreme are always unrestricted — this param is ignored for them
    _downline_ids_for_filter: Optional[List[int]] = None
    if my_team_only and not _viewer_is_unrestricted:
        _downline_ids_for_filter = get_downline_employee_ids(db, current_user.id, recursive=True)
        if _downline_ids_for_filter:
            emp_q = emp_q.filter(StaffEmployee.id.in_(_downline_ids_for_filter))
        else:
            # viewer has no downline — return empty result set
            return {
                "success": True, "employees": [], "weights": {}, "config": {},
                "viewer_unrestricted": False, "my_team_only_applied": True,
                "message": "No downline employees found under your reporting hierarchy."
            }

    all_emps = emp_q.all()

    all_ids = [e.id for e in all_emps]
    _emp_name_map = {e.id: e.full_name for e in all_emps}
    _missing_mgr_ids = {e.reporting_manager_id for e in all_emps if e.reporting_manager_id and e.reporting_manager_id not in _emp_name_map}
    if _missing_mgr_ids:
        _mgr_rows = db.query(StaffEmployee.id, StaffEmployee.full_name).filter(StaffEmployee.id.in_(_missing_mgr_ids)).all()
        for _mr in _mgr_rows:
            _emp_name_map[_mr.id] = _mr.full_name

    if not all_ids:
        return {"success": True, "data": [], "total": 0,
                "departments": [], "weights": {},
                "date_range": {"from": date_from.isoformat(), "to": date_to.isoformat()}}

    # ── Batch: Attendance ─────────────────────────────────────────────────────
    # days_present: HR-marked only (present=1, half_day=0.5, absent=0)
    hr_att_rows = db.query(
        StaffAttendanceSheet.employee_id,
        func.sum(case(
            (StaffAttendanceSheet.attendance_status == 'present', 1),
            (StaffAttendanceSheet.attendance_status == 'half_day', 0.5),
            else_=0
        )).label('days_present'),
        func.sum(case(
            (StaffAttendanceSheet.attendance_status == 'weekend', 1),
            else_=0
        )).label('wo_days'),
    ).filter(
        StaffAttendanceSheet.employee_id.in_(all_ids),
        StaffAttendanceSheet.date.between(date_from, date_to),
    ).group_by(StaffAttendanceSheet.employee_id).all()

    wm_rows = db.query(
        StaffAttendance.employee_id,
        func.sum(func.coalesce(StaffAttendance.worked_minutes, 0)).label('worked_minutes'),
    ).filter(
        StaffAttendance.employee_id.in_(all_ids),
        StaffAttendance.date.between(date_from, date_to),
    ).group_by(StaffAttendance.employee_id).all()
    wm_map = {r.employee_id: int(r.worked_minutes or 0) for r in wm_rows}

    att_map = {r.employee_id: {'days_present': float(r.days_present or 0),
                                'worked_minutes': wm_map.get(r.employee_id, 0),
                                'wo_days': int(r.wo_days or 0)}
               for r in hr_att_rows}
    for eid, wm in wm_map.items():
        if eid not in att_map:
            att_map[eid] = {'days_present': 0.0, 'worked_minutes': wm, 'wo_days': 0}

    # ── Batch: Day Plan (live JOIN — avoids stale denormalized total_completed) ─
    # DC Protocol: Count delivered items live from staff_day_plan_items instead of using
    # the denormalized total_completed column on StaffDayPlan (which is NOT updated on EOD edits)
    _dp_raw = db.execute(
        text(
            "SELECT p.employee_id,"
            " COUNT(i.id) AS live_planned,"
            " COALESCE(SUM(CASE"
            "   WHEN i.eod_status = 'delivered' THEN 1"
            "   WHEN i.eod_status IS NULL AND i.planned_status = 'completed' THEN 1"
            "   ELSE 0 END), 0) AS live_completed"
            " FROM staff_day_plans p"
            " LEFT JOIN staff_day_plan_items i ON i.day_plan_id = p.id"
            " WHERE p.employee_id = ANY(:ids)"
            "   AND p.plan_date BETWEEN :df AND :dt"
            " GROUP BY p.employee_id"
        ),
        {'ids': all_ids, 'df': date_from, 'dt': date_to}
    ).fetchall()
    dp_map = {
        r[0]: {
            'has_plan': True,
            'total_planned': int(r[1] or 0),
            'total_completed': int(r[2] or 0),
            'pct': round((r[2] / r[1] * 100) if r[1] else 0),
        }
        for r in _dp_raw
    }

    # ── Batch: KRA — uses leave/Sunday-aware helper so Sundays & leave days are excluded ──
    _raw_kra_map = _batch_kra_summary(db, all_ids, date_from, date_to)
    kra_map = {eid: {
        'total': v.get('effective_total', v.get('total', 0)),  # use effective_total (leave-exempt)
        'completed': v.get('completed', 0),
    } for eid, v in _raw_kra_map.items()}

    # ── Batch: Timesheet ──────────────────────────────────────────────────────
    ts_rows = db.query(
        StaffTimesheetEntry.employee_id,
        func.count(func.distinct(StaffTimesheetEntry.date)).label('days_submitted'),
        func.sum(StaffTimesheetEntry.duration_minutes).label('total_minutes'),
        func.sum(case((StaffTimesheetEntry.status == 'approved', StaffTimesheetEntry.duration_minutes), else_=0)).label('approved_minutes'),
    ).filter(
        StaffTimesheetEntry.employee_id.in_(all_ids),
        StaffTimesheetEntry.status.in_(['submitted', 'approved']),
        StaffTimesheetEntry.date.between(date_from, date_to),
    ).group_by(StaffTimesheetEntry.employee_id).all()
    ts_map = {r.employee_id: {
        'days_submitted': int(r.days_submitted or 0),
        'total_minutes': int(r.total_minutes or 0),
        'approved_minutes': int(r.approved_minutes or 0),
    } for r in ts_rows}

    # ── Batch: Leads (Sales) ──────────────────────────────────────────────────
    lead_rows = db.query(
        CRMLead.telecaller_id,
        func.count(CRMLead.id).label('new_leads'),
        func.sum(case((CRMLead.status == 'won', 1), else_=0)).label('won'),
        func.sum(case((CRMLead.status == 'won',
                       func.coalesce(CRMLead.deal_value_received, 0)), else_=0)).label('dv_received'),
        func.sum(case((CRMLead.confirmed_final_value != None,
                       func.coalesce(CRMLead.confirmed_final_value, 0)), else_=0)).label('cfv_sum'),
    ).filter(
        CRMLead.telecaller_id.in_(all_ids),
        func.date(CRMLead.created_at).between(date_from, date_to),
    ).group_by(CRMLead.telecaller_id).all()
    leads_map = {r.telecaller_id: {
        'new_leads': int(r.new_leads or 0), 'won': int(r.won or 0),
        'dv_received': float(r.dv_received or 0),
        'cfv_sum': float(r.cfv_sum or 0),
    } for r in lead_rows}

    # ── Batch: Service Tickets ────────────────────────────────────────────────
    svc_rows = db.query(
        ServiceTicket.service_technician_id,
        func.count(ServiceTicket.id).label('total'),
        func.sum(case((ServiceTicket.status.in_(['Resolved', 'Closed']), 1), else_=0)).label('resolved'),
        func.sum(case((ServiceTicket.sla_status == 'Within SLA', 1), else_=0)).label('within_tat_count'),
    ).filter(
        ServiceTicket.service_technician_id.in_(all_ids),
        func.date(ServiceTicket.created_date).between(date_from, date_to),
    ).group_by(ServiceTicket.service_technician_id).all()
    svc_map = {r.service_technician_id: {
        'total': int(r.total or 0),
        'resolved': int(r.resolved or 0),
        'within_tat_count': int(r.within_tat_count or 0),
    } for r in svc_rows}

    # ── Batch: CRM Transaction Revenue ────────────────────────────────────────
    txn_rows = db.query(
        CRMLeadTransaction.collected_by_id,
        func.coalesce(func.sum(CRMLeadTransaction.amount), 0).label('txn_revenue'),
    ).filter(
        CRMLeadTransaction.collected_by_id.in_(all_ids),
        func.date(CRMLeadTransaction.transaction_date).between(date_from, date_to),
        CRMLeadTransaction.validation_status != 'rejected'
    ).group_by(CRMLeadTransaction.collected_by_id).all()
    txn_map = {r.collected_by_id: float(r.txn_revenue or 0) for r in txn_rows}

    # ── Batch: Call Talk Time (Sales) ─────────────────────────────────────────
    from app.models.call_tracking import StaffCallLog
    call_rows = db.query(
        StaffCallLog.staff_id,
        func.sum(StaffCallLog.duration_seconds).label('total_seconds'),
    ).filter(
        StaffCallLog.staff_id.in_(all_ids),
        StaffCallLog.call_date.between(str(date_from), str(date_to)),
    ).group_by(StaffCallLog.staff_id).all()
    call_map = {r.staff_id: int(r.total_seconds or 0) for r in call_rows}

    # ── Batch: PO Handling (store_manager_id, 48h TAT) ────────────────────────
    from sqlalchemy import text as _sqlt
    _po_rows = db.execute(_sqlt("""
        SELECT store_manager_id,
               COUNT(*) AS total,
               SUM(CASE WHEN status IN ('payment_received','dispatched','completed')
                         AND completed_at IS NOT NULL
                         AND EXTRACT(EPOCH FROM (completed_at - created_at)) <= 172800
                    THEN 1 ELSE 0 END) AS within_tat
        FROM marketplace_purchase_orders
        WHERE store_manager_id = ANY(:ids)
          AND status NOT IN ('cancelled','hold')
          AND created_at::date BETWEEN :df AND :dt
        GROUP BY store_manager_id
    """), {'ids': all_ids, 'df': date_from, 'dt': date_to}).fetchall()
    po_map = {r[0]: {'total': int(r[1] or 0), 'within_tat': int(r[2] or 0)} for r in _po_rows}

    # ── Batch: Procurement Handling (48h quality_confirmed/added_to_stock) ────
    _prh_rows = db.execute(_sqlt("""
        SELECT store_manager_id,
               COUNT(*) AS total,
               SUM(CASE WHEN status IN ('quality_confirmed','added_to_stock')
                         AND EXTRACT(EPOCH FROM (updated_at - created_at)) <= 172800
                    THEN 1 ELSE 0 END) AS within_tat
        FROM marketplace_procurement_requests
        WHERE store_manager_id = ANY(:ids)
          AND status NOT IN ('cancelled','hold')
          AND created_at::date BETWEEN :df AND :dt
        GROUP BY store_manager_id
    """), {'ids': all_ids, 'df': date_from, 'dt': date_to}).fetchall()
    prh_map = {r[0]: {'total': int(r[1] or 0), 'within_tat': int(r[2] or 0)} for r in _prh_rows}

    # ── In-memory hierarchy for manager revenue ───────────────────────────────
    reporting = {e.id: e.reporting_manager_id for e in all_emps}
    _dl_cache: dict = {}

    def _downline(mgr_id: int, _visited: set = None) -> list:
        if _visited is None:
            _visited = set()
        if mgr_id in _dl_cache:
            return _dl_cache[mgr_id]
        _visited.add(mgr_id)
        direct = [eid for eid, mid in reporting.items() if mid == mgr_id and eid not in _visited]
        ids = list(direct)
        for d in direct:
            ids += _downline(d, _visited.copy())
        _dl_cache[mgr_id] = ids
        return ids

    # ── Load dynamic KPI config from staff_performance_config ────────────────
    from sqlalchemy import text as _sa_text
    _cfg_rows = db.execute(_sa_text(
        "SELECT kpi_code, is_enabled, target_value, weightage_pct, sub_config "
        "FROM staff_performance_config ORDER BY id"
    )).fetchall()
    global_cfg = {
        r[0]: {
            'is_enabled': bool(r[1]),
            'target_value': float(r[2] or 1),
            'weightage_pct': float(r[3] or 0),
            'sub_config': r[4] or {},
        }
        for r in _cfg_rows
    }

    # ── Load per-employee targets (month-specific → default fallback) ─────────
    _rpt_month = date_from.month
    _rpt_year  = date_from.year
    _tgt_rows  = db.execute(_sa_text("""
        WITH ranked AS (
            SELECT employee_id, kpi_code, target_value, is_enabled, weightage_pct,
                   ROW_NUMBER() OVER (
                       PARTITION BY employee_id, kpi_code
                       ORDER BY CASE WHEN month = :m AND year = :y THEN 0 ELSE 1 END, id DESC
                   ) AS rn
            FROM staff_performance_employee_kpi
            WHERE employee_id = ANY(:ids)
              AND ((month = :m AND year = :y) OR (month = 0 AND year = 0))
        )
        SELECT employee_id, kpi_code, target_value, is_enabled, weightage_pct
        FROM ranked WHERE rn = 1
    """), {'ids': all_ids, 'm': _rpt_month, 'y': _rpt_year}).fetchall()
    emp_targets: dict = {}
    for _r in _tgt_rows:
        emp_targets.setdefault(_r[0], {})[_r[1]] = {
            'target_value':  float(_r[2]) if _r[2] is not None else None,
            'is_enabled':    _r[3],
            'weightage_pct': float(_r[4]) if _r[4] is not None else None,
        }

    BONUS_CAP = 120.0

    def _eff_cfg(kpi_code: str, emp_id: int):
        """Returns (is_enabled, target, weightage, sub_config) with per-employee override."""
        g   = global_cfg.get(kpi_code, {})
        row = emp_targets.get(emp_id, {}).get(kpi_code)
        is_en = row['is_enabled'] if (row and row['is_enabled'] is not None) else g.get('is_enabled', False)
        tgt   = float(row['target_value'] if row and row.get('target_value') is not None else g.get('target_value', 1)) or 1.0
        wt    = float(row['weightage_pct'] if row and row.get('weightage_pct') is not None else g.get('weightage_pct', 0))
        sub   = g.get('sub_config') or {}
        return is_en, tgt, wt, sub

    # ── Score + assemble each employee ───────────────────────────────────────
    results = []

    for emp in all_emps:
        dept_lower = (emp.department.name or '').lower() if emp.department else ''
        is_sales   = 'sales'   in dept_lower
        is_service = 'service' in dept_lower

        att  = att_map.get(emp.id, {'days_present': 0, 'worked_minutes': 0, 'wo_days': 0})
        _ts  = ts_map.get(emp.id, {'days_submitted': 0, 'total_minutes': 0, 'approved_minutes': 0})
        _dp  = dp_map.get(emp.id, {'has_plan': False, 'pct': 0})
        _kra = kra_map.get(emp.id, {'total': 0, 'completed': 0})
        _svc = svc_map.get(emp.id, {'total': 0, 'resolved': 0, 'within_tat_count': 0})
        _talk_secs = call_map.get(emp.id, 0)
        _svc_total = _svc.get('total', 0)
        _days_pres = att['days_present']
        _wo_days = att.get('wo_days', 0)
        _working_days_emp = max(1, total_days - _wo_days)

        # Raw actuals used both for per-KPI scoring and legacy response fields
        # DC Protocol: working days per employee = total days minus HR-marked WO days
        att_pct  = round((_days_pres / _working_days_emp) * 100, 1) if _working_days_emp > 0 else 0.0
        dp_pct   = float(_dp['pct']) if _dp.get('has_plan') else 0.0
        kra_pct  = round((_kra['completed'] / _kra['total']) * 100, 1) if _kra['total'] > 0 else 0.0
        ts_pct   = round((_ts.get('days_submitted', 0) / _working_days_emp) * 100, 1) if _working_days_emp > 0 else 0.0
        talk_min = round(_talk_secs / 60 / max(_days_pres, 1), 1)
        crm_rev  = txn_map.get(emp.id, 0.0)

        # Dept score (legacy UI field)
        dept_score = 0.0
        if is_sales:
            ld = leads_map.get(emp.id, {'new_leads': 0, 'won': 0})
            dept_score = round((ld['won'] / ld['new_leads']) * 100, 1) if ld['new_leads'] > 0 else (100.0 if ld['won'] > 0 else 0.0)
        elif is_service:
            dept_score = round((_svc.get('resolved', 0) / _svc_total) * 100, 1) if _svc_total > 0 else 0.0

        # ── Dynamic KPI scoring (config + per-employee targets) ───────────────
        kpi_scores: dict = {}
        overall  = 0.0
        total_w  = 0.0

        for kpi_code in global_cfg:
            is_en, tgt, wt, sub = _eff_cfg(kpi_code, emp.id)
            if not is_en:
                continue

            if   kpi_code == 'attendance':      actual = att_pct
            elif kpi_code == 'day_plan':        actual = dp_pct
            elif kpi_code == 'kra':             actual = kra_pct
            elif kpi_code == 'timesheet':       actual = ts_pct
            elif kpi_code == 'talk_time':       actual = talk_min
            elif kpi_code == 'crm_revenue':     actual = crm_rev
            elif kpi_code == 'service_tickets':
                min_t = int(sub.get('min_tickets', 0))
                tat_t = float(sub.get('tat_pct', 0))
                if _svc_total < min_t:
                    actual = 0.0
                else:
                    res_pct = round((_svc.get('resolved', 0) / _svc_total) * 100, 1) if _svc_total > 0 else 0.0
                    tat_pct = round((_svc.get('within_tat_count', 0) / _svc_total) * 100, 1) if _svc_total > 0 else 0.0
                    actual  = (res_pct + tat_pct) / 2 if tat_t > 0 else res_pct
            else:
                actual = 0.0  # procurement — no batch data in exec dashboard

            raw_score = min((actual / tgt) * 100, BONUS_CAP) if tgt > 0 else 0.0
            score     = round(raw_score, 1)
            _emp_row  = emp_targets.get(emp.id, {}).get(kpi_code)
            kpi_scores[kpi_code] = {
                'actual': actual, 'target': tgt, 'weightage': wt,
                'score': score, 'bonus': score > 100,
                'is_custom_target': _emp_row is not None and _emp_row.get('target_value') is not None,
                'is_custom_weight': _emp_row is not None and _emp_row.get('weightage_pct') is not None,
            }
            total_w += wt

        if total_w > 0:
            for ks in kpi_scores.values():
                overall += ks['score'] * (ks['weightage'] / total_w)

        # Revenue hierarchy
        dl_ids       = _downline(emp.id)
        team_revenue = sum(leads_map.get(did, {}).get('dv_received', 0) for did in dl_ids)
        own_revenue  = leads_map.get(emp.id, {}).get('dv_received', 0)
        own_cfv      = leads_map.get(emp.id, {}).get('cfv_sum', 0)
        team_cfv     = sum(leads_map.get(did, {}).get('cfv_sum', 0) for did in dl_ids)
        _tat_count   = _svc.get('within_tat_count', 0)
        _ts_appr     = _ts.get('approved_minutes', 0)
        _po          = po_map.get(emp.id, {'total': 0, 'within_tat': 0})
        _prh         = prh_map.get(emp.id, {'total': 0, 'within_tat': 0})

        results.append({
            "employee_id": emp.id,
            "emp_code": emp.emp_code or '',
            "name": emp.full_name or '',
            "department": emp.department.name if emp.department else '—',
            "department_id": emp.department_id,
            "dept_type": 'sales' if is_sales else ('service' if is_service else 'other'),
            "role": emp.role.role_name if emp.role else '—',
            "reporting_manager_id": emp.reporting_manager_id,
            "reporting_manager_name": _emp_name_map.get(emp.reporting_manager_id, '—') if emp.reporting_manager_id else '—',
            "overall_score": round(overall, 1),
            "score_breakdown": {
                'attendance': round(att_pct), 'day_plan': round(dp_pct),
                'kra': round(kra_pct), 'timesheet': round(ts_pct), 'dept': round(dept_score),
            },
            "kpi_scores": kpi_scores,
            "att_score":  round(att_pct),
            "dp_score":   round(dp_pct),
            "kra_score":  round(kra_pct),
            "ts_score":   round(ts_pct),
            "dept_score": round(dept_score),
            "total_worked_minutes": att['worked_minutes'],
            "days_present": _days_pres,
            "avg_approved_minutes_per_day": (_ts_appr // _days_pres) if _days_pres > 0 else 0,
            "own_deal_value_received": own_revenue,
            "team_deal_value_received": team_revenue,
            "own_confirmed_final_value": own_cfv,
            "team_confirmed_final_value": team_cfv,
            "tickets_total": _svc_total,
            "tickets_resolved": _svc.get('resolved', 0),
            "within_tat_count": _tat_count,
            "within_tat_pct": round(_tat_count / _svc_total * 100) if _svc_total > 0 else 0,
            "talk_time_seconds": _talk_secs,
            "avg_talk_secs_per_day": (_talk_secs // _days_pres) if _days_pres > 0 else 0,
            "leads_new": leads_map.get(emp.id, {}).get('new_leads', 0),
            "leads_won": leads_map.get(emp.id, {}).get('won', 0),
            "crm_txn_revenue": crm_rev,
            "po_total": _po['total'],
            "po_within_tat": _po['within_tat'],
            "po_tat_pct": round(_po['within_tat'] / _po['total'] * 100, 1) if _po['total'] > 0 else None,
            "prh_total": _prh['total'],
            "prh_within_tat": _prh['within_tat'],
            "prh_tat_pct": round(_prh['within_tat'] / _prh['total'] * 100, 1) if _prh['total'] > 0 else None,
            "total_days": total_days,
        })

    results.sort(key=lambda x: x['overall_score'], reverse=True)
    for i, r in enumerate(results):
        r['rank'] = i + 1

    # Department list for filter UI
    dept_ids = list({e.department_id for e in all_emps if e.department_id})
    dept_objs = db.query(StaffDepartment).filter(StaffDepartment.id.in_(dept_ids)).all() if dept_ids else []

    # Company list for filter UI (all companies that have active staff)
    co_ids = list({e.base_company_id for e in all_emps if e.base_company_id})
    co_objs = db.query(AssociatedCompany).filter(AssociatedCompany.id.in_(co_ids)).all() if co_ids else []

    return {
        "success": True,
        "data": results,
        "total": len(results),
        "departments": [{"id": d.id, "name": d.name} for d in dept_objs],
        "companies": [{"id": c.id, "name": c.company_name, "code": c.company_code} for c in co_objs],
        "date_range": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        "weights": {k: v['weightage_pct'] for k, v in global_cfg.items() if v['is_enabled']},
        "config": {k: {'target': v['target_value'], 'weightage': v['weightage_pct'], 'enabled': v['is_enabled']}
                   for k, v in global_cfg.items()},
        "viewer_unrestricted": _viewer_is_unrestricted,
        "my_team_only_applied": my_team_only and not _viewer_is_unrestricted,
    }

@router.get("/export", summary="Export progress report (all roles)")
async def export_progress_report(
    target_date: Optional[date] = Query(None, description="Date for progress"),
    date_from: Optional[date] = Query(None, description="Start date for range"),
    date_to: Optional[date] = Query(None, description="End date for range"),
    employee_id: Optional[int] = Query(None, description="Employee ID"),
    format: str = Query("excel", description="Export format: excel or pdf"),
    current_user: StaffEmployee = Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    """
    DC Protocol: Export progress report as Excel or PDF
    Access: All staff — non-leadership restricted to own data only
    """
    if not is_key_leadership(current_user) and employee_id and employee_id != current_user.id:
        employee_id = current_user.id

    logger.info(f"[DC-EXPORT] {current_user.emp_code} requested {format} export for emp={employee_id or current_user.id} date_from={date_from} date_to={date_to}")

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    def _write_sheet(wb, sheet_name, data_list):
        if not data_list:
            return
        ws = wb.create_sheet(sheet_name)
        if isinstance(data_list[0], dict):
            headers = list(data_list[0].keys())
            ws.append(headers)
            for item in data_list:
                ws.append([item.get(h) for h in headers])
        else:
            for row in data_list:
                ws.append(row)

    today = get_indian_date()
    if target_date is None:
        target_date = today

    target_employee_id = employee_id or current_user.id
    target_employee = db.query(StaffEmployee).filter(
        StaffEmployee.id == target_employee_id
    ).first()

    if not target_employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    query_start = date_from or target_date
    query_end = date_to or target_date

    # DC Protocol: per-statement timeout — prevents export blocking a worker indefinitely.
    # 30 s is generous for a 30-day range across 4 sub-queries.
    try:
        db.execute(text("SET LOCAL statement_timeout = '30000'"))
    except Exception:
        pass

    try:
        tasks_data = get_tasks_summary(db, target_employee_id, query_start, query_end)
        kra_data = get_kra_summary(db, target_employee_id, query_start, query_end)
        attendance_data = get_attendance_summary(db, target_employee_id, query_start, query_end)
        travel_data = get_travel_summary(db, target_employee_id, query_start, query_end)
    except Exception as _qe:
        _qe_str = str(_qe)
        if 'statement timeout' in _qe_str.lower() or 'canceling statement' in _qe_str.lower() or 'QueryCanceled' in _qe_str:
            logger.warning(f"[DC-EXPORT] Query timeout for emp={target_employee_id} range={query_start}..{query_end}")
            raise HTTPException(status_code=408, detail="Export query timed out — please try a shorter date range (e.g. 2 weeks) or retry shortly.")
        raise

    if format == "excel":
        output = BytesIO()

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        ws_summary.append(["Metric", "Value"])
        summary_rows = [
            ("Tasks Planned", tasks_data["summary"]["planned"]),
            ("Tasks Completed", tasks_data["summary"]["completed"]),
            ("Tasks Pending", tasks_data["summary"]["pending"]),
            ("Tasks Overdue", tasks_data["summary"]["overdue"]),
            ("KRA Total", kra_data["summary"]["total"]),
            ("KRA Completed", kra_data["summary"]["completed"]),
            ("KRA Pending", kra_data["summary"]["pending"]),
            ("KRA Missed", kra_data["summary"]["missed"]),
            ("Days Present", attendance_data["summary"]["days_present"]),
            ("Total Working Hours", attendance_data["summary"]["total_worked_hours"]),
            ("Total Break Hours", attendance_data["summary"]["total_break_hours"]),
            ("Travel Distance (km)", travel_data["summary"]["total_km"]),
            ("Travel Time", travel_data["summary"]["total_time"]),
        ]
        for row in summary_rows:
            ws_summary.append(list(row))

        _write_sheet(wb, 'Tasks', tasks_data.get("activities", []))
        _write_sheet(wb, 'KRA', kra_data.get("items", []))
        _write_sheet(wb, 'Attendance', attendance_data.get("records", []))

        wb.save(output)
        output.seek(0)

        filename = f"progress_report_{target_employee.emp_code}_{query_start.isoformat()}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "pdf":
        from io import BytesIO as _BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        output = _BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=5)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=10)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=11, spaceBefore=8, spaceAfter=4, textColor=colors.HexColor('#6B21A8'))
        
        elements = []
        
        elements.append(Paragraph("Staff Progress Report", title_style))
        
        emp_info = f"<b>{target_employee.full_name}</b> ({target_employee.emp_code})"
        if target_employee.department:
            emp_info += f" - {target_employee.department.name}"
        if target_employee.role:
            emp_info += f" | {target_employee.role.role_name}"
        elements.append(Paragraph(emp_info, subtitle_style))
        
        date_range_text = f"Date: {query_start.strftime('%d %b %Y')}"
        if query_start != query_end:
            date_range_text = f"Period: {query_start.strftime('%d %b %Y')} to {query_end.strftime('%d %b %Y')}"
        export_info = f"{date_range_text} | Exported: {datetime.now(IST).strftime('%d %b %Y, %I:%M %p')}"
        elements.append(Paragraph(export_info, subtitle_style))
        elements.append(Spacer(1, 5*mm))
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B21A8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F8F8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ])
        
        elements.append(Paragraph("Tasks", section_style))
        tasks_table_data = [
            ['Planned', 'Completed', 'Pending', 'Overdue', '% Done'],
            [
                str(tasks_data["summary"]["planned"]),
                str(tasks_data["summary"]["completed"]),
                str(tasks_data["summary"]["pending"]),
                str(tasks_data["summary"]["overdue"]),
                f'{tasks_data["summary"]["completion_percent"]}%'
            ]
        ]
        t = Table(tasks_table_data, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)
        
        elements.append(Paragraph("KRA", section_style))
        kra_table_data = [
            ['Total', 'Completed', 'Pending', 'Missed', '% Done'],
            [
                str(kra_data["summary"]["total"]),
                str(kra_data["summary"]["completed"]),
                str(kra_data["summary"]["pending"]),
                str(kra_data["summary"]["missed"]),
                f'{kra_data["summary"]["completion_percent"]}%'
            ]
        ]
        t = Table(kra_table_data, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)
        
        elements.append(Paragraph("Time & Attendance", section_style))
        att_table_data = [
            ['Days Present', 'Working Hours', 'Break Hours', 'Avg/Day'],
            [
                str(attendance_data["summary"]["days_present"]),
                attendance_data["summary"]["total_worked_hours"],
                attendance_data["summary"]["total_break_hours"],
                attendance_data["summary"]["avg_worked_per_day"]
            ]
        ]
        t = Table(att_table_data, colWidths=[43.75*mm]*4)
        t.setStyle(table_style)
        elements.append(t)
        
        elements.append(Paragraph("Travel/Journey", section_style))
        travel_table_data = [
            ['Sessions', 'Total Distance', 'Total Time'],
            [
                str(travel_data["summary"]["session_count"]),
                f'{travel_data["summary"]["total_km"]} km',
                travel_data["summary"]["total_time"]
            ]
        ]
        t = Table(travel_table_data, colWidths=[58.33*mm]*3)
        t.setStyle(table_style)
        elements.append(t)
        
        leads_data = get_leads_summary(db, target_employee_id, query_start, query_end)
        timesheet_data = get_timesheet_summary(db, target_employee_id, query_start, query_end)
        
        elements.append(Paragraph("Leads (Sales) - Overall", section_style))
        leads_table_data = [
            ['Total Assigned', 'New Created', 'Won', 'Lost', 'Revenue'],
            [
                str(leads_data["summary"]["total_assigned"]),
                str(leads_data["summary"]["new_created"]),
                str(leads_data["summary"]["won"]),
                str(leads_data["summary"]["lost"]),
                f'₹{leads_data["summary"]["revenue"]:,.0f}'
            ]
        ]
        t = Table(leads_table_data, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)
        
        company_leads = leads_data["summary"].get("company_leads", {})
        self_leads = leads_data["summary"].get("self_leads", {})
        
        elements.append(Paragraph("Leads - Company Assigned", section_style))
        company_leads_table = [
            ['Total', 'New', 'Won', 'Lost', 'Revenue'],
            [
                str(company_leads.get("total", 0)),
                str(company_leads.get("new", 0)),
                str(company_leads.get("won", 0)),
                str(company_leads.get("lost", 0)),
                f'₹{company_leads.get("revenue", 0):,.0f}'
            ]
        ]
        t = Table(company_leads_table, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)
        
        elements.append(Paragraph("Leads - Self Created", section_style))
        self_leads_table = [
            ['Total', 'New', 'Won', 'Lost', 'Revenue'],
            [
                str(self_leads.get("total", 0)),
                str(self_leads.get("new", 0)),
                str(self_leads.get("won", 0)),
                str(self_leads.get("lost", 0)),
                f'₹{self_leads.get("revenue", 0):,.0f}'
            ]
        ]
        t = Table(self_leads_table, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)
        
        service_data = get_service_tickets_summary(db, target_employee_id, query_start, query_end)
        elements.append(Paragraph("Service Tickets", section_style))
        svc_table_data = [
            ['Total', 'New', 'Resolved', 'Pending', 'Overdue'],
            [
                str(service_data["summary"]["total"]),
                str(service_data["summary"]["new"]),
                str(service_data["summary"]["resolved"]),
                str(service_data["summary"]["pending"]),
                str(service_data["summary"]["overdue"])
            ]
        ]
        t = Table(svc_table_data, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)
        
        _ts_tot_min = timesheet_data["summary"]["total_minutes"] or 0
        elements.append(Paragraph("Timesheet", section_style))
        ts_table_data = [
            ['Entries', 'Total Hours'],
            [
                str(timesheet_data["summary"]["count"]),
                f"{_ts_tot_min//60}h {_ts_tot_min%60}m"
            ]
        ]
        t = Table(ts_table_data, colWidths=[87.5*mm]*2)
        t.setStyle(table_style)
        elements.append(t)

        calls_data = get_calls_summary(db, target_employee_id, query_start, query_end)
        elements.append(Paragraph("Calls / Talk Time", section_style))
        calls_table_data = [
            ['Total Calls', 'Outgoing', 'Incoming', 'Missed', 'Talk Time'],
            [
                str(calls_data["summary"]["total_calls"]),
                str(calls_data["summary"]["outgoing"]),
                str(calls_data["summary"]["incoming"]),
                str(calls_data["summary"]["missed"]),
                calls_data["summary"]["talk_time"]
            ]
        ]
        t = Table(calls_table_data, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)

        _exp_dept_name = (target_employee.department.name or '').lower() if target_employee.department else ''
        _exp_is_sales = 'sales' in _exp_dept_name or 'crm' in _exp_dept_name
        if _exp_is_sales:
            try:
                from datetime import datetime as _dt
                _df_dt = _dt.combine(query_start, _dt.min.time())
                _dt_dt = _dt.combine(query_end, _dt.max.time().replace(microsecond=0))
                _user_ref_str = str(target_employee_id)
                _dial_overall = db.execute(text("""
                    SELECT
                        COUNT(*) as total_dials,
                        COUNT(CASE WHEN call_outcome = 'answered' THEN 1 END) as answered,
                        COUNT(CASE WHEN call_outcome = 'skip' THEN 1 END) as skipped,
                        COUNT(CASE WHEN call_outcome = 'no_answer' THEN 1 END) as no_answer,
                        COALESCE(SUM(duration_seconds), 0) as total_talk_seconds,
                        COALESCE(AVG(NULLIF(duration_seconds, 0)), 0) as avg_duration_seconds
                    FROM crm_dialer_attempts
                    WHERE user_ref = :ur AND created_at BETWEEN :df AND :dt
                """), {"ur": _user_ref_str, "df": _df_dt, "dt": _dt_dt}).fetchone()
                _dial_sess = db.execute(text("""
                    SELECT COALESCE(SUM(
                        EXTRACT(EPOCH FROM (COALESCE(closed_at, last_active_at, NOW()::timestamp) - started_at))
                    ), 0)::int as total_session_seconds
                    FROM crm_dialer_sessions
                    WHERE user_ref = :ur AND started_at BETWEEN :df AND :dt
                """), {"ur": _user_ref_str, "df": _df_dt, "dt": _dt_dt}).fetchone()

                def _fmt_sec(s):
                    s = int(s or 0)
                    h = s // 3600; m = (s % 3600) // 60; sec = s % 60
                    if h > 0: return f"{h}h {m:02d}m"
                    if m > 0: return f"{m}m {sec}s"
                    return f"{sec}s" if sec > 0 else "0s"

                _td = int(_dial_overall[0] or 0)
                _ta = int(_dial_overall[1] or 0)
                _ts = int(_dial_overall[2] or 0)
                _tna = int(_dial_overall[3] or 0)
                _ttk = int(_dial_overall[4] or 0)
                _sess_s = int(_dial_sess[0] or 0) if _dial_sess else 0
                _idle_s = max(0, _sess_s - _ttk)
                _days = max(1, (query_end - query_start).days + 1)
                _avg_talk = _fmt_sec(_ttk // _days) if _ttk > 0 else '—'

                elements.append(Paragraph("CRM Auto Dialer", section_style))
                dial_table_data = [
                    ['Total Dials', 'Connected', 'Skipped', 'No Answer', 'Talk Time', 'Avg/Day'],
                    [str(_td), str(_ta), str(_ts), str(_tna), _fmt_sec(_ttk), _avg_talk]
                ]
                t = Table(dial_table_data, colWidths=[29.16*mm]*6)
                t.setStyle(table_style)
                elements.append(t)
                if _sess_s > 0:
                    sess_table_data = [
                        ['Session Time', 'Talk Time', 'Idle/Paused'],
                        [_fmt_sec(_sess_s), _fmt_sec(_ttk), _fmt_sec(_idle_s)]
                    ]
                    t2 = Table(sess_table_data, colWidths=[58.33*mm]*3)
                    t2.setStyle(table_style)
                    elements.append(t2)
            except Exception as _de:
                logger.warning(f"[DC-EXPORT] Dialer section failed (non-fatal): {_de}")

        day_plan_data = get_day_plan_summary(db, target_employee_id, query_start, query_end)
        elements.append(Paragraph("Day Plan", section_style))
        dp_table_data = [
            ['Days with Plan', 'Planned', 'Completed', 'Pending', '% Done'],
            [
                str(len(day_plan_data.get("plans", []))),
                str(day_plan_data.get("total_items", 0)),
                str(day_plan_data.get("completed_items", 0)),
                str(day_plan_data.get("pending_items", 0)),
                f'{day_plan_data.get("completion_percent", 0)}%'
            ]
        ]
        t = Table(dp_table_data, colWidths=[35*mm]*5)
        t.setStyle(table_style)
        elements.append(t)

        is_date_range = query_start != query_end
        if is_date_range:
            elements.append(Spacer(1, 5*mm))
            elements.append(Paragraph("Date-wise Breakdown", title_style))
            
            if tasks_data.get("datewise"):
                elements.append(Paragraph("Tasks - Daily", section_style))
                tasks_dw = [['Date', 'Planned', 'Completed', 'Pending', 'Overdue', 'Overall\nPending', 'Overall\nOverdue']]
                _cum_pend = 0
                _cum_over = 0
                for dw in tasks_data["datewise"]:
                    _cum_pend += dw["pending"]
                    _cum_over += dw["overdue"]
                    tasks_dw.append([dw["date"], str(dw["planned"]), str(dw["completed"]), str(dw["pending"]), str(dw["overdue"]), str(_cum_pend), str(_cum_over)])
                t = Table(tasks_dw, colWidths=[27*mm, 24*mm, 24*mm, 22*mm, 22*mm, 24*mm, 24*mm])
                t.setStyle(table_style)
                elements.append(t)
            
            if kra_data.get("datewise"):
                elements.append(Paragraph("KRA - Daily", section_style))
                kra_dw = [['Date', 'Total', 'Completed', 'Pending', 'Missed']]
                for dw in kra_data["datewise"]:
                    kra_dw.append([dw["date"], str(dw["total"]), str(dw["completed"]), str(dw["pending"]), str(dw["missed"])])
                t = Table(kra_dw, colWidths=[35*mm]*5)
                t.setStyle(table_style)
                elements.append(t)
            
            if attendance_data.get("datewise"):
                elements.append(Paragraph("Attendance - Daily", section_style))
                att_dw = [['Date', 'Status', 'Worked', 'Break', 'Clock In', 'Clock Out']]
                for dw in attendance_data["datewise"]:
                    wm = dw["worked_minutes"] or 0
                    bm = dw["break_minutes"] or 0
                    att_dw.append([
                        dw["date"], dw["status"],
                        f"{wm//60}h {wm%60}m",
                        f"{bm//60}h {bm%60}m",
                        dw.get("clock_in") or "-",
                        dw.get("clock_out") or "-"
                    ])
                t = Table(att_dw, colWidths=[29*mm]*6)
                t.setStyle(table_style)
                elements.append(t)
            
            if timesheet_data.get("datewise"):
                elements.append(Paragraph("Timesheet - Daily", section_style))
                ts_dw = [['Date', 'Entries', 'Total Hours']]
                for dw in timesheet_data["datewise"]:
                    tm = dw["total_minutes"] or 0
                    ts_dw.append([dw["date"], str(dw["entries_count"]), f"{tm//60}h {tm%60}m"])
                t = Table(ts_dw, colWidths=[58.33*mm]*3)
                t.setStyle(table_style)
                elements.append(t)
            
            if leads_data.get("datewise"):
                elements.append(Paragraph("Leads (Company) - Daily", section_style))
                leads_c_dw = [['Date', 'New', 'Won', 'Lost', 'Revenue']]
                for dw in leads_data["datewise"]:
                    leads_c_dw.append([dw["date"], str(dw["company_new"]), str(dw["company_won"]), str(dw["company_lost"]), f'₹{dw["company_revenue"]:,.0f}'])
                t = Table(leads_c_dw, colWidths=[35*mm]*5)
                t.setStyle(table_style)
                elements.append(t)
                
                elements.append(Paragraph("Leads (Self) - Daily", section_style))
                leads_s_dw = [['Date', 'New', 'Won', 'Lost', 'Revenue']]
                for dw in leads_data["datewise"]:
                    leads_s_dw.append([dw["date"], str(dw["self_new"]), str(dw["self_won"]), str(dw["self_lost"]), f'₹{dw["self_revenue"]:,.0f}'])
                t = Table(leads_s_dw, colWidths=[35*mm]*5)
                t.setStyle(table_style)
                elements.append(t)
            
            if service_data.get("datewise"):
                elements.append(Paragraph("Service Tickets - Daily", section_style))
                svc_dw = [['Date', 'New', 'Resolved']]
                for dw in service_data["datewise"]:
                    svc_dw.append([dw["date"], str(dw["new"]), str(dw["resolved"])])
                t = Table(svc_dw, colWidths=[58.33*mm]*3)
                t.setStyle(table_style)
                elements.append(t)
            
            if travel_data.get("datewise"):
                elements.append(Paragraph("Travel - Daily", section_style))
                travel_dw = [['Date', 'Sessions', 'KM', 'Duration']]
                for dw in travel_data["datewise"]:
                    travel_dw.append([dw["date"], str(dw["sessions"]), str(dw["km_travelled"]), dw["duration"]])
                t = Table(travel_dw, colWidths=[43.75*mm]*4)
                t.setStyle(table_style)
                elements.append(t)

            if calls_data.get("datewise"):
                elements.append(Paragraph("Calls - Daily", section_style))
                calls_dw = [['Date', 'Total', 'Outgoing', 'Incoming', 'Missed', 'Talk Time']]
                for dw in calls_data["datewise"]:
                    calls_dw.append([dw["date"], str(dw["total"]), str(dw["outgoing"]), str(dw["incoming"]), str(dw["missed"]), dw["talk_time"]])
                t = Table(calls_dw, colWidths=[28*mm, 23*mm, 23*mm, 23*mm, 23*mm, 30*mm])
                t.setStyle(table_style)
                elements.append(t)

            if day_plan_data.get("datewise"):
                elements.append(Paragraph("Day Plan - Daily", section_style))
                dp_dw = [['Date', 'Tasks', 'Done', 'Pending', '% Done']]
                for dw in day_plan_data["datewise"]:
                    dp_dw.append([dw["date"], str(dw.get("tasks_total", 0)), str(dw.get("tasks_done", 0)), str(dw.get("tasks_pending", 0)), f'{dw.get("pct_done", 0)}%'])
                t = Table(dp_dw, colWidths=[35*mm]*5)
                t.setStyle(table_style)
                elements.append(t)

        elements.append(Spacer(1, 8*mm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER, textColor=colors.grey)
        elements.append(Paragraph("Generated by Mynt Real Staff Portal | Confidential", footer_style))
        
        doc.build(elements)
        output.seek(0)
        
        filename = f"progress_report_{target_employee.emp_code}_{query_start.isoformat()}.pdf"
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'pdf' or 'excel'.")
