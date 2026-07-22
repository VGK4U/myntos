"""
Staff Financial Management System API Endpoints - DC_SFMS_001
DC Protocol Compliant - JWT-based authentication with VGK/EA RBAC

Created: Dec 06, 2025
DC Protocol: Write-Verify-Validate at all levels

VGK and EA have equal full permissions on all Accounts module features.
"""

from fastapi import APIRouter, Depends, HTTPException, Query, Path, status, File, Form, UploadFile, Body, Request
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List
from decimal import Decimal
from datetime import date

from app.core.database import get_db
from app.models.staff import StaffEmployee
from app.models.staff_accounts import PurchaseInvoiceUpload, PurchaseInvoiceLineItem, HSNMaster, VendorMaster, StockItemMaster, StockItemImage, RevenueCategory, IncomeEntry, AssociatedCompany, EmployeeFundLedger, EmployeeFundTransfer, ExpenseEntry, CashflowRegister, FundAllocation, SolarVendorLedger, StockLedger
from app.api.v1.endpoints.staff_auth import get_current_staff_user
from app.services.staff_accounts_service import (
    AssociatedCompanyService,
    CompanySegmentService,
    IncomeSourceTypeService,
    PricingConfigurationService,
    HSNMasterService,
    VendorMasterService,
    StockItemMasterService,
    IncomeEntryService,
    ExpenseMainCategoryService,
    ExpenseSubCategoryService,
    AccountsRBACError,
    AccountsValidationError,
    AccountsNotFoundError,
    AccountsDuplicateError,
    validate_accounts_access,
    create_fund_allocation,
    confirm_fund_allocation,
    settle_fund_allocation,
    cancel_fund_allocation,
    list_fund_allocations,
    get_fund_allocation,
    create_expense_entry,
    submit_expense_entry,
    approve_expense_entry,
    list_expense_entries,
    get_expense_entry,
    mark_expense_paid,
    compute_balance_sheet,
    get_balance_sheet_trend,
    get_dashboard_alerts,
    get_balance_sheet,
    AccountsCreditService,
    BOMService,
    ManufacturingService,
    ACCOUNTS_ALLOWED_ROLES,
    is_accounts_allowed_role,
    is_accounts_allowed_employee,
    ServiceCenterGivenOutService,
    log_accounts_audit
)
from app.services.sfms_seed import run_sfms_seed, seed_mynt_real_llp, seed_default_income_sources
from app.services.sfms_cache_service import sfms_cache
from app.schemas.staff_accounts import (
    AssociatedCompanyCreate,
    AssociatedCompanyUpdate,
    AssociatedCompanyResponse,
    AssociatedCompanyListResponse,
    CompanySegmentCreate,
    CompanySegmentUpdate,
    CompanySegmentResponse,
    CompanySegmentListResponse,
    IncomeSourceTypeCreate,
    IncomeSourceTypeUpdate,
    IncomeSourceTypeResponse,
    IncomeSourceTypeListResponse,
    PricingConfigurationCreate,
    PricingConfigurationUpdate,
    PricingConfigurationResponse,
    PricingConfigurationListResponse,
    PriceCalculationRequest,
    PriceCalculationPreview,
    HSNMasterCreate,
    HSNMasterUpdate,
    HSNMasterResponse,
    HSNMasterListResponse,
    GSTCalculationRequest,
    GSTCalculationResponse,
    VendorMasterCreate,
    VendorMasterUpdate,
    VendorMasterResponse,
    VendorMasterListResponse,
    StockItemMasterCreate,
    StockItemMasterUpdate,
    StockItemMasterResponse,
    StockItemMasterListResponse,
    StockItemBulkUploadResult,
    StockItemBulkUploadError,
    StockItemImageResponse,
    StockItemImageUploadResponse,
    StockItemImagesListResponse,
    IncomeEntryCreate,
    IncomeEntryUpdate,
    IncomeEntryStatusUpdate,
    IncomeEntryResponse,
    IncomeEntryListResponse,
    ExpenseMainCategoryCreate,
    ExpenseMainCategoryUpdate,
    ExpenseMainCategoryResponse,
    ExpenseMainCategorySimpleResponse,
    ExpenseMainCategoryListResponse,
    ExpenseSubCategoryCreate,
    ExpenseSubCategoryUpdate,
    ExpenseSubCategoryResponse,
    ExpenseSubCategoryListResponse,
    SuccessResponse,
    ErrorResponse,
    FundAllocationCreate,
    FundAllocationUpdate,
    FundAllocationConfirm,
    FundAllocationSettle,
    FundAllocationCancel,
    FundAllocationResponse,
    FundAllocationListResponse,
    ExpenseEntryCreate,
    ExpenseEntryUpdate,
    ExpenseEntrySubmit,
    ExpenseEntryApprove,
    ExpenseEntryMarkPaid,
    ExpenseEntryResponse,
    ExpenseEntryListResponse,
    BalanceSheetSummaryResponse,
    BalanceSheetComputeRequest,
    BalanceSheetDashboardResponse,
    RecordVendorPaymentRequest,
    RecordCustomerReceiptRequest,
    PayablesListResponse,
    ReceivablesListResponse,
    AgingSummaryResponse,
    PaymentTransactionListResponse,
    CreditDashboardResponse,
    CreditAgingBucket,
    BOMCreate,
    BOMUpdate,
    BOMResponse,
    BOMListResponse,
    BOMApprovalRequest,
    BOMCopyRequest,
    BOMLineItemCreate,
    BOMLineItemUpdate,
    BOMLineItemResponse,
    ManufacturingOrderCreate,
    ManufacturingOrderUpdate,
    ManufacturingOrderResponse,
    ManufacturingOrderListResponse,
    ManufacturingStartRequest,
    ManufacturingCompleteRequest,
    ManufacturingCancelRequest,
    ManufacturingUpdateRequest,
    ManufacturingConsumptionRequest,
    ManufacturingStockCheckResponse,
    # Purchase Intake Schemas (DC_INTAKE_001 - Jan 2026)
    PurchaseIntakeBatchReceive,
    PurchaseIntakeBatchQCSubmit,
    PurchaseIntakeBatchApprove,
    PurchaseIntakeBatchReject,
    RevenueCategoryCreate,
    RevenueCategoryUpdate,
    RevenueCategoryResponse,
)

router = APIRouter(prefix="/staff/accounts", tags=["Staff Accounts"])


import threading as _threading
import logging as _logging
_sa_logger = _logging.getLogger(__name__)

def _bg_refresh_mkt_qty():
    """DC-STOCK-MKT-003: Fire-and-forget background thread — refreshes marketplace_spares.available_qty
    from stock_ledger after any purchase or sale. Keeps ecom/marketplace quantities live.
    Non-blocking — caller returns immediately."""
    def _run():
        from app.core.database import SessionLocal
        from app.services.marketplace_sync import run_stock_sync
        db = SessionLocal()
        try:
            run_stock_sync(db, company_id=1, triggered_by='post-transaction')
            _sa_logger.info('[DC-STOCK-MKT-003] Marketplace qty refreshed after transaction')
        except Exception as _e:
            _sa_logger.warning(f'[DC-STOCK-MKT-003] Marketplace qty refresh failed (non-fatal): {_e}')
        finally:
            db.close()
    _threading.Thread(target=_run, daemon=True, name='mkt-qty-refresh').start()


def handle_accounts_error(e: Exception) -> JSONResponse:
    """Centralized error handling for accounts endpoints"""
    if isinstance(e, AccountsRBACError):
        return JSONResponse(
            status_code=status.HTTP_403_FORBIDDEN,
            content={"success": False, "message": str(e), "error_code": "RBAC_DENIED"}
        )
    elif isinstance(e, AccountsValidationError):
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": str(e), "error_code": "VALIDATION_ERROR"}
        )
    elif isinstance(e, AccountsNotFoundError):
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": str(e), "error_code": "NOT_FOUND"}
        )
    elif isinstance(e, AccountsDuplicateError):
        return JSONResponse(
            status_code=status.HTTP_409_CONFLICT,
            content={"success": False, "message": str(e), "error_code": "DUPLICATE"}
        )
    else:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Internal error: {str(e)}", "error_code": "INTERNAL_ERROR"}
        )


# ==================== ASSOCIATED COMPANIES ENDPOINTS ====================

@router.post("/companies", response_model=AssociatedCompanyResponse)
async def create_company(
    data: AssociatedCompanyCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new associated company (tenant)
    DC_SAAS_CONSOLE_001 (May 2026): Tenant creation is restricted to VGK4U Supreme only.
    """
    # DC_SAAS_CONSOLE_001: Supreme-only guard — tenant onboarding is a SaaS-level
    # action and must never be available to regular staff.
    _staff_type = (getattr(current_user, "staff_type", "") or "").strip()
    if _staff_type != "VGK4U Supreme":
        return JSONResponse(
            status_code=403,
            content={
                "success": False,
                "message": "Only VGK4U Supreme administrators can create new tenants.",
                "error_code": "TENANT_CREATE_FORBIDDEN",
            },
        )
    try:
        company = AssociatedCompanyService.create_company(db, data, current_user)
        return JSONResponse(content={
            "success": True,
            "message": f"Company '{company.company_name}' created successfully",
            "company": AssociatedCompanyResponse.model_validate(company).model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/companies", response_model=AssociatedCompanyListResponse)
async def list_companies(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status_filter: Optional[str] = Query(None, description="Filter by status: ACTIVE, INACTIVE, ALL. Default (None or ACTIVE) returns active companies only."),
    search: Optional[str] = Query(None, description="Search by name, code, or city"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List all associated companies with pagination and filters
    DC: VGK/EA access only
    """
    try:
        companies, total = AssociatedCompanyService.list_companies(
            db, current_user, page, page_size, status_filter, search
        )
        return JSONResponse(content={
            "success": True,
            "companies": [AssociatedCompanyResponse.model_validate(c).model_dump(mode='json') for c in companies],
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/companies/{company_id}", response_model=AssociatedCompanyResponse)
async def get_company(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single associated company by ID
    DC: VGK/EA access only
    """
    try:
        company = AssociatedCompanyService.get_company(db, company_id, current_user)
        return JSONResponse(content={
            "success": True,
            "company": AssociatedCompanyResponse.model_validate(company).model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/companies/{company_id}", response_model=AssociatedCompanyResponse)
async def update_company(
    company_id: int,
    data: AssociatedCompanyUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an associated company
    DC: VGK/EA access only
    """
    try:
        company = AssociatedCompanyService.update_company(db, company_id, data, current_user)
        return JSONResponse(content={
            "success": True,
            "message": f"Company '{company.company_name}' updated successfully",
            "company": AssociatedCompanyResponse.model_validate(company).model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/companies/{company_id}")
async def delete_company(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft delete (deactivate) an associated company
    DC: VGK/EA access only. Cannot delete book keeper.
    """
    try:
        AssociatedCompanyService.delete_company(db, company_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Company deactivated successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/companies/{company_id}/transfer-summary")
async def get_company_transfer_summary(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-CO-TRANSFER-001: Return count of transferable data linked to a company."""
    try:
        summary = AssociatedCompanyService.get_company_transfer_summary(db, company_id, current_user)
        return JSONResponse(content={"success": True, "summary": summary})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/companies/{company_id}/transfer-to/{target_id}")
async def transfer_company_data(
    company_id: int,
    target_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-CO-TRANSFER-001: Transfer all data from company to target, then deactivate source."""
    try:
        result = AssociatedCompanyService.transfer_company_data(db, company_id, target_id, current_user)
        return JSONResponse(content={"success": True, "message": "Data transferred and company deactivated.", **result})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/companies/{company_id}/set-book-keeper")
async def set_book_keeper(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Set a company as the book keeper (Mynt Real LLP)
    DC: VGK/EA access only. Only one company can be book keeper.
    """
    try:
        company = AssociatedCompanyService.set_book_keeper(db, company_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": f"'{company.company_name}' is now the book keeper",
            "company": AssociatedCompanyResponse.model_validate(company).model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== COMPANY BANK ACCOUNTS (DC-BANK-001) ====================

@router.get("/companies/{company_id}/bank-accounts")
async def list_company_bank_accounts(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List all active bank accounts for a company. DC-BANK-001."""
    try:
        from app.models.staff_accounts import CompanyBankAccount as _CBA
        rows = db.query(_CBA).filter(
            _CBA.company_id == company_id,
            _CBA.is_active == True
        ).order_by(_CBA.is_primary.desc(), _CBA.id).all()
        return JSONResponse(content={"bank_accounts": [r.to_dict() for r in rows]})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/companies/{company_id}/bank-accounts")
async def create_company_bank_account(
    company_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Add a bank account to a company. DC-BANK-001. VGK/EA/Accounts access."""
    try:
        from app.models.staff_accounts import CompanyBankAccount as _CBA
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        bank_name = (payload.get('bank_name') or '').strip()
        account_number = (payload.get('account_number') or '').strip()
        if not bank_name or not account_number:
            raise HTTPException(status_code=400, detail="bank_name and account_number are required")
        is_primary = bool(payload.get('is_primary', False))
        if is_primary:
            db.query(_CBA).filter(_CBA.company_id == company_id, _CBA.is_active == True).update({'is_primary': False})
        ba = _CBA(
            company_id=company_id,
            bank_name=bank_name,
            branch=(payload.get('branch') or '').strip() or None,
            account_number=account_number,
            ifsc_code=(payload.get('ifsc_code') or '').strip() or None,
            account_type=(payload.get('account_type') or 'CURRENT').upper(),
            is_primary=is_primary,
            notes=(payload.get('notes') or '').strip() or None,
            created_by_id=current_user.id,
            updated_by_id=current_user.id,
        )
        db.add(ba)
        db.commit()
        db.refresh(ba)
        return JSONResponse(content={"success": True, "bank_account": ba.to_dict()}, status_code=201)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/companies/{company_id}/bank-accounts/{ba_id}")
async def update_company_bank_account(
    company_id: int,
    ba_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update a company bank account. DC-BANK-001."""
    try:
        from app.models.staff_accounts import CompanyBankAccount as _CBA
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        ba = db.query(_CBA).filter(_CBA.id == ba_id, _CBA.company_id == company_id).first()
        if not ba:
            raise HTTPException(status_code=404, detail="Bank account not found")
        for field in ('bank_name', 'branch', 'account_number', 'ifsc_code', 'notes'):
            if field in payload:
                setattr(ba, field, (payload[field] or '').strip() or None)
        if 'account_type' in payload:
            ba.account_type = (payload['account_type'] or 'CURRENT').upper()
        if 'is_primary' in payload and payload['is_primary']:
            db.query(_CBA).filter(_CBA.company_id == company_id, _CBA.is_active == True, _CBA.id != ba_id).update({'is_primary': False})
            ba.is_primary = True
        ba.updated_by_id = current_user.id
        db.commit()
        db.refresh(ba)
        return JSONResponse(content={"success": True, "bank_account": ba.to_dict()})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/companies/{company_id}/bank-accounts/{ba_id}")
async def deactivate_company_bank_account(
    company_id: int,
    ba_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Deactivate (soft-delete) a company bank account. DC-BANK-001."""
    try:
        from app.models.staff_accounts import CompanyBankAccount as _CBA
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        ba = db.query(_CBA).filter(_CBA.id == ba_id, _CBA.company_id == company_id).first()
        if not ba:
            raise HTTPException(status_code=404, detail="Bank account not found")
        ba.is_active = False
        ba.updated_by_id = current_user.id
        db.commit()
        return JSONResponse(content={"success": True, "message": "Bank account deactivated"})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== COMPANY SEGMENTS ENDPOINTS ====================

@router.post("/segments", response_model=CompanySegmentResponse)
async def create_segment(
    data: CompanySegmentCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new company segment
    DC: VGK/EA access only
    """
    try:
        segment = CompanySegmentService.create_segment(db, data, current_user)
        
        response = CompanySegmentResponse.model_validate(segment).model_dump(mode='json')
        
        return JSONResponse(content={
            "success": True,
            "message": f"Segment '{segment.segment_name}' created successfully",
            "segment": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/segments", response_model=CompanySegmentListResponse)
async def list_segments(
    company_id: Optional[int] = Query(None, description="Filter by company ID"),
    status_filter: Optional[str] = Query(None, description="Filter by status: ACTIVE, INACTIVE"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List company segments with filters
    DC: VGK/EA access only
    """
    try:
        segments, total = CompanySegmentService.list_segments(
            db, current_user, company_id, status_filter
        )
        
        from app.models.staff_accounts import AssociatedCompany
        company_map = {}
        if segments:
            company_ids = list(set(s.company_id for s in segments))
            companies = db.query(AssociatedCompany).filter(
                AssociatedCompany.id.in_(company_ids)
            ).all()
            company_map = {c.id: c.company_name for c in companies}
        
        segment_list = []
        for s in segments:
            seg_dict = CompanySegmentResponse.model_validate(s).model_dump(mode='json')
            seg_dict['company_name'] = company_map.get(s.company_id)
            segment_list.append(seg_dict)
        
        return JSONResponse(content={
            "success": True,
            "segments": segment_list,
            "total": total
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/segments/{segment_id}", response_model=CompanySegmentResponse)
async def get_segment(
    segment_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single company segment by ID
    DC: VGK/EA access only
    """
    try:
        segment = CompanySegmentService.get_segment(db, segment_id, current_user)
        
        from app.models.staff_accounts import AssociatedCompany
        company = db.query(AssociatedCompany).filter(
            AssociatedCompany.id == segment.company_id
        ).first()
        
        seg_dict = CompanySegmentResponse.model_validate(segment).model_dump(mode='json')
        seg_dict['company_name'] = company.company_name if company else None
        
        return JSONResponse(content={
            "success": True,
            "segment": seg_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/segments/{segment_id}", response_model=CompanySegmentResponse)
async def update_segment(
    segment_id: int,
    data: CompanySegmentUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update a company segment
    DC: VGK/EA access only
    """
    try:
        segment = CompanySegmentService.update_segment(db, segment_id, data, current_user)
        
        from app.models.staff_accounts import AssociatedCompany
        company = db.query(AssociatedCompany).filter(
            AssociatedCompany.id == segment.company_id
        ).first()
        
        seg_dict = CompanySegmentResponse.model_validate(segment).model_dump(mode='json')
        seg_dict['company_name'] = company.company_name if company else None
        
        return JSONResponse(content={
            "success": True,
            "message": f"Segment '{segment.segment_name}' updated successfully",
            "segment": seg_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== REVENUE CATEGORY ENDPOINTS ====================

@router.get("/revenue-categories")
async def list_revenue_categories(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    active_only: bool = Query(True, description="Only active categories"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    validate_accounts_access(current_user)
    from app.models.signup_category import SignupCategory
    query = db.query(SignupCategory)
    if company_id:
        query = query.filter(SignupCategory.company_id == company_id)
    if active_only:
        query = query.filter(SignupCategory.is_active == True)
    categories = query.order_by(SignupCategory.company_id, SignupCategory.display_order).all()

    company_ids = list(set(c.company_id for c in categories))
    companies = db.query(AssociatedCompany).filter(AssociatedCompany.id.in_(company_ids)).all() if company_ids else []
    company_map = {c.id: c.company_name for c in companies}

    result = []
    for cat in categories:
        d = cat.to_dict()
        d['category_name'] = cat.name
        d['category_code'] = cat.slug
        d['company_name'] = company_map.get(cat.company_id)
        result.append(d)

    return JSONResponse(content={"success": True, "categories": result, "total": len(result), "source": "signup_categories"})


@router.post("/revenue-categories")
async def create_revenue_category(
    data: RevenueCategoryCreate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    validate_accounts_access(current_user)
    return JSONResponse(
        status_code=410,
        content={
            "success": False,
            "message": "Revenue categories are now managed via Configuration > Signup Categories. Please use that page to add new categories.",
            "redirect": "/staff/signup-categories"
        }
    )


@router.put("/revenue-categories/{category_id}")
async def update_revenue_category(
    category_id: int,
    data: RevenueCategoryUpdate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    validate_accounts_access(current_user)
    return JSONResponse(
        status_code=410,
        content={
            "success": False,
            "message": "Revenue categories are now managed via Configuration > Signup Categories. Please use that page to edit categories.",
            "redirect": "/staff/signup-categories"
        }
    )


@router.get("/revenue-categories/{category_id}")
async def get_revenue_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    validate_accounts_access(current_user)
    from app.models.signup_category import SignupCategory
    cat = db.query(SignupCategory).filter(SignupCategory.id == category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    result = cat.to_dict()
    result['category_name'] = cat.name
    result['category_code'] = cat.slug
    company = db.query(AssociatedCompany).filter(AssociatedCompany.id == cat.company_id).first()
    result['company_name'] = company.company_name if company else None
    return JSONResponse(content={"success": True, "category": result})


# ==================== INCOME SOURCE TYPES ENDPOINTS ====================

@router.post("/income-sources", response_model=IncomeSourceTypeResponse)
async def create_income_source(
    data: IncomeSourceTypeCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new income source type
    DC: VGK/EA access only
    """
    try:
        source = IncomeSourceTypeService.create_income_source(db, data, current_user)
        
        from app.models.staff_accounts import AssociatedCompany
        company = db.query(AssociatedCompany).filter(
            AssociatedCompany.id == source.company_id
        ).first()
        
        source_dict = IncomeSourceTypeResponse.model_validate(source).model_dump(mode='json')
        source_dict['company_name'] = company.company_name if company else None
        
        return JSONResponse(content={
            "success": True,
            "message": f"Income source '{source.source_name}' created successfully",
            "income_source": source_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-sources", response_model=IncomeSourceTypeListResponse)
async def list_income_sources(
    company_id: Optional[int] = Query(None, description="Filter by company ID"),
    category: Optional[str] = Query(None, description="Filter by category: SALES, SERVICE, MNR_PAYMENT, OTHER"),
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List income source types with filters
    DC: VGK/EA access only
    """
    try:
        sources, total = IncomeSourceTypeService.list_income_sources(
            db, current_user, is_active=is_active
        )
        
        source_list = []
        for s in sources:
            src_dict = IncomeSourceTypeResponse.model_validate(s).model_dump(mode='json')
            source_list.append(src_dict)
        
        return JSONResponse(content={
            "success": True,
            "income_source_types": source_list,
            "total": total
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/income-sources/{source_id}", response_model=IncomeSourceTypeResponse)
async def update_income_source(
    source_id: int,
    data: IncomeSourceTypeUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an income source type
    DC: VGK/EA access only
    """
    try:
        source = IncomeSourceTypeService.update_income_source(db, source_id, data, current_user)
        
        from app.models.staff_accounts import AssociatedCompany
        company = db.query(AssociatedCompany).filter(
            AssociatedCompany.id == source.company_id
        ).first()
        
        source_dict = IncomeSourceTypeResponse.model_validate(source).model_dump(mode='json')
        source_dict['company_name'] = company.company_name if company else None
        
        return JSONResponse(content={
            "success": True,
            "message": f"Income source '{source.source_name}' updated successfully",
            "income_source": source_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== PRICING CONFIGURATION ENDPOINTS ====================

@router.get("/pricing-config/{company_id}", response_model=PricingConfigurationResponse)
async def get_pricing_config(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get pricing configuration for a company (creates default if not exists)
    DC: VGK/EA access only
    """
    try:
        config = PricingConfigurationService.get_or_create_config(db, company_id, current_user)
        
        from app.models.staff_accounts import AssociatedCompany
        company = db.query(AssociatedCompany).filter(
            AssociatedCompany.id == config.company_id
        ).first()
        
        config_dict = PricingConfigurationResponse.model_validate(config).model_dump(mode='json')
        config_dict['company_name'] = company.company_name if company else None
        
        return JSONResponse(content={
            "success": True,
            "pricing_config": config_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/pricing-config/{config_id}", response_model=PricingConfigurationResponse)
async def update_pricing_config(
    config_id: int,
    data: PricingConfigurationUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update pricing configuration
    DC: VGK/EA access only. Validates markup constraints.
    """
    try:
        config = PricingConfigurationService.update_config(db, config_id, data, current_user)
        
        from app.models.staff_accounts import AssociatedCompany
        company = db.query(AssociatedCompany).filter(
            AssociatedCompany.id == config.company_id
        ).first()
        
        config_dict = PricingConfigurationResponse.model_validate(config).model_dump(mode='json')
        config_dict['company_name'] = company.company_name if company else None
        
        return JSONResponse(content={
            "success": True,
            "message": "Pricing configuration updated successfully",
            "pricing_config": config_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/pricing-config/calculate", response_model=PriceCalculationPreview)
async def calculate_price(
    data: PriceCalculationRequest,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Calculate price preview with markup and incentive
    DC: VGK/EA access only. Shows derived prices before saving.
    """
    try:
        result = PricingConfigurationService.calculate_price(db, data, current_user)
        return JSONResponse(content={
            "success": True,
            "calculation": result.model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== HSN MASTER ENDPOINTS ====================

@router.post("/hsn", response_model=dict)
async def create_hsn(
    data: HSNMasterCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new HSN/SAC code entry with GST rates
    DC_HSN_001: VGK4U Supreme, EA, and Accounts team have equal access
    """
    try:
        hsn = HSNMasterService.create_hsn(db, data, current_user)
        response = HSNMasterResponse.model_validate(hsn).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"HSN code '{hsn.hsn_code}' created successfully",
            "hsn": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/hsn", response_model=dict)
async def list_hsn_codes(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=200, description="Items per page"),
    search: Optional[str] = Query(None, description="Search by HSN code or description"),
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List HSN/SAC codes with pagination and filtering
    DC_HSN_001: VGK4U Supreme, EA, and Accounts team have equal access
    """
    try:
        skip = (page - 1) * page_size
        hsn_codes, total = HSNMasterService.list_hsn_codes(
            db,
            skip=skip,
            limit=page_size,
            search=search,
            is_active=is_active
        )
        
        response_list = [
            HSNMasterResponse.model_validate(hsn).model_dump(mode='json')
            for hsn in hsn_codes
        ]
        
        return JSONResponse(content={
            "success": True,
            "hsn_codes": response_list,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/hsn/{hsn_id}", response_model=dict)
async def get_hsn(
    hsn_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get HSN/SAC code by ID
    DC_HSN_001: VGK4U Supreme, EA, and Accounts team have equal access
    """
    try:
        hsn = HSNMasterService.get_hsn_by_id(db, hsn_id)
        if not hsn:
            raise AccountsNotFoundError(f"HSN code with ID {hsn_id} not found")
        
        response = HSNMasterResponse.model_validate(hsn).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "hsn": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/hsn/code/{hsn_code}", response_model=dict)
async def get_hsn_by_code(
    hsn_code: str,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get HSN/SAC code by code string
    DC_HSN_001: VGK4U Supreme, EA, and Accounts team have equal access
    """
    try:
        hsn = HSNMasterService.get_hsn_by_code(db, hsn_code)
        if not hsn:
            raise AccountsNotFoundError(f"HSN code '{hsn_code}' not found")
        
        response = HSNMasterResponse.model_validate(hsn).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "hsn": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/hsn/{hsn_id}", response_model=dict)
async def update_hsn(
    hsn_id: int,
    data: HSNMasterUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an existing HSN/SAC code
    DC_HSN_001: VGK4U Supreme, EA, and Accounts team have equal access
    """
    try:
        hsn = HSNMasterService.update_hsn(db, hsn_id, data, current_user)
        response = HSNMasterResponse.model_validate(hsn).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"HSN code '{hsn.hsn_code}' updated successfully",
            "hsn": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/hsn/{hsn_id}")
async def delete_hsn(
    hsn_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft delete (deactivate) an HSN/SAC code
    DC_HSN_001: VGK4U Supreme, EA, and Accounts team have equal access
    """
    try:
        HSNMasterService.delete_hsn(db, hsn_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "HSN code deactivated successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/hsn/calculate-gst", response_model=dict)
async def calculate_gst(
    data: GSTCalculationRequest,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Calculate GST based on HSN code and transaction type
    DC_GST_CALC_001: Determines intra-state (CGST+SGST) vs inter-state (IGST)
    """
    try:
        result = HSNMasterService.calculate_gst(db, data)
        return JSONResponse(content={
            "success": True,
            "calculation": result.model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== PURCHASE INVOICE UPLOAD ENDPOINTS (DC_PURCHASE_001) ====================

@router.post("/purchase-uploads")
async def create_purchase_upload(
    company_id: int = Form(...),
    file: UploadFile = File(...),
    vendor_id: Optional[int] = Form(None),
    vendor_invoice_no: Optional[str] = Form(None),
    vendor_invoice_date: Optional[str] = Form(None),
    is_credit_purchase: bool = Form(False),
    credit_days: int = Form(0),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Upload a purchase invoice file (PDF, Image, Excel, CSV)
    DC_PURCHASE_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        import os
        import uuid
        from datetime import datetime
        
        file_ext = file.filename.split('.')[-1].lower() if '.' in file.filename else 'bin'
        file_type = PurchaseInvoiceUploadService._detect_file_type(file.filename)
        
        from app.services.universal_upload_service import UniversalUploadService
        
        upload_result = await UniversalUploadService.handle_upload(
            file=file,
            table_name='purchase_invoice_uploads',
            record_id=0,
            uploaded_by_id=current_user.id,
            uploaded_by_type='staff',
            storage_dir='purchase_invoices',
            db=db,
            emp_code=current_user.emp_code
        )
        
        file_path = upload_result['file_path']
        
        invoice_date = None
        if vendor_invoice_date:
            try:
                invoice_date = datetime.strptime(vendor_invoice_date, '%Y-%m-%d').date()
            except:
                pass
        
        upload = PurchaseInvoiceUploadService.create_upload(
            db=db,
            company_id=company_id,
            file_path=file_path,
            file_name=file.filename,
            file_type=file_type,
            employee=current_user,
            vendor_id=vendor_id,
            vendor_invoice_no=vendor_invoice_no,
            vendor_invoice_date=invoice_date,
            is_credit_purchase=is_credit_purchase,
            credit_days=credit_days
        )
        
        extraction_message = ""
        try:
            upload = PurchaseInvoiceUploadService.extract_invoice_data(
                db=db,
                upload_id=upload.id,
                employee=current_user
            )
            extraction_message = f" Data extracted successfully ({upload.extraction_method})."
        except Exception as ext_err:
            extraction_message = f" Extraction note: {str(ext_err)[:100]}"
        
        return JSONResponse(content={
            "success": True,
            "message": f"Invoice uploaded successfully: {upload.upload_number}.{extraction_message}",
            "upload": upload.to_dict()
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/manual")
async def create_manual_purchase_invoice(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a purchase invoice manually without file upload
    DC_PURCHASE_002: Manual entry with full line item support
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        from datetime import datetime
        
        company_id = data.get('company_id')
        vendor_id = data.get('vendor_id')
        vendor_invoice_no = data.get('vendor_invoice_no')
        vendor_invoice_date_str = data.get('vendor_invoice_date')
        
        if not company_id:
            raise AccountsValidationError("Company ID is required")
        if not vendor_id:
            raise AccountsValidationError("Vendor ID is required")
        if not vendor_invoice_no:
            raise AccountsValidationError("Vendor invoice number is required")
        if not vendor_invoice_date_str:
            raise AccountsValidationError("Vendor invoice date is required")
        
        invoice_date = None
        if vendor_invoice_date_str:
            try:
                invoice_date = datetime.strptime(vendor_invoice_date_str, '%Y-%m-%d').date()
            except:
                raise AccountsValidationError("Invalid invoice date format. Use YYYY-MM-DD")
        
        due_date = None
        if data.get('due_date'):
            try:
                due_date = datetime.strptime(data['due_date'], '%Y-%m-%d').date()
            except:
                pass
        
        upload = PurchaseInvoiceUploadService.create_manual_upload(
            db=db,
            company_id=company_id,
            vendor_id=vendor_id,
            vendor_invoice_no=vendor_invoice_no,
            vendor_invoice_date=invoice_date,
            due_date=due_date,
            is_credit_purchase=data.get('is_credit_purchase', False),
            credit_days=data.get('credit_days', 0),
            segment_id=data.get('segment_id'),
            review_notes=data.get('review_notes'),
            line_items=data.get('line_items', []),
            employee=current_user,
            is_igst=data.get('is_igst'),
            document_type=data.get('document_type', 'invoice'),
            return_reference=data.get('return_reference'),
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Manual invoice created: {upload.upload_number}",
            "upload": upload.to_dict()
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-uploads")
async def list_purchase_uploads(
    company_id: Optional[int] = Query(None),
    vendor_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List purchase invoice uploads with filtering
    DC_PURCHASE_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        uploads, total = PurchaseInvoiceUploadService.list_uploads(
            db=db,
            employee=current_user,
            company_id=company_id,
            vendor_id=vendor_id,
            status=status,
            date_from=date_from,
            date_to=date_to,
            page=page,
            page_size=page_size
        )
        
        stats = PurchaseInvoiceUploadService.get_stats(
            db=db,
            company_id=company_id,
            vendor_id=vendor_id,
            date_from=date_from,
            date_to=date_to
        )
        
        return JSONResponse(content={
            "success": True,
            "uploads": [u.to_dict() for u in uploads],
            "total": total,
            "page": page,
            "page_size": page_size,
            "stats": stats
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/check-duplicate")
async def check_purchase_duplicate(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Check for duplicate purchase invoice uploads
    DC_PURCHASE_001: Duplicate detection by vendor + invoice_no + date
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        from datetime import datetime
        
        company_id = data.get('company_id')
        vendor_id = data.get('vendor_id')
        vendor_invoice_no = data.get('vendor_invoice_no')
        vendor_invoice_date_str = data.get('vendor_invoice_date')
        exclude_upload_id = data.get('exclude_upload_id')
        
        vendor_invoice_date = None
        if vendor_invoice_date_str:
            try:
                vendor_invoice_date = datetime.strptime(vendor_invoice_date_str, '%Y-%m-%d').date()
            except:
                pass
        
        result = PurchaseInvoiceUploadService.check_duplicate(
            db=db,
            company_id=company_id,
            vendor_id=vendor_id,
            vendor_invoice_no=vendor_invoice_no,
            vendor_invoice_date=vendor_invoice_date,
            exclude_upload_id=exclude_upload_id
        )
        
        return JSONResponse(content={
            "success": True,
            **result
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-uploads/{upload_id}")
async def get_purchase_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a purchase invoice upload with line items
    DC_PURCHASE_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        upload = PurchaseInvoiceUploadService.get_upload(db, upload_id, current_user)
        
        upload_dict = upload.to_dict()
        upload_dict['line_items'] = [item.to_dict() for item in upload.line_items] if upload.line_items else []
        upload_dict['extracted_data'] = upload.extracted_data
        
        return JSONResponse(content={
            "success": True,
            "upload": upload_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/re-extract")
async def re_extract_purchase_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Re-extract data from an existing purchase invoice upload
    DC_PURCHASE_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        upload = PurchaseInvoiceUploadService.extract_invoice_data(
            db=db,
            upload_id=upload_id,
            employee=current_user
        )
        
        upload_dict = upload.to_dict()
        
        return JSONResponse(content={
            "success": True,
            "message": f"Data re-extracted successfully ({upload.extraction_method})",
            "upload": upload_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/purchase-uploads/{upload_id}")
async def update_purchase_upload(
    upload_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update a purchase invoice upload - for editing extracted data
    DC_PURCHASE_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        upload = PurchaseInvoiceUploadService.update_upload(
            db=db,
            upload_id=upload_id,
            employee=current_user,
            **data
        )
        
        upload_dict = upload.to_dict()
        upload_dict['line_items'] = [item.to_dict() for item in upload.line_items] if upload.line_items else []
        
        return JSONResponse(content={
            "success": True,
            "message": "Upload updated successfully",
            "upload": upload_dict
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/confirm")
async def confirm_purchase_upload(
    upload_id: int,
    data: dict = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Confirm a purchase invoice - creates vendor transaction and updates Stock/AP/Ledger
    DC_PURCHASE_001: Transaction atomicity for all updates
    WVV Protocol: Write-Verify-Validate
    DC Protocol: Blocks confirmation if matching_status.all_resolved is False
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        data = data or {}
        
        upload = db.query(PurchaseInvoiceUpload).filter(PurchaseInvoiceUpload.id == upload_id).first()
        if not upload:
            raise AccountsValidationError("Invoice upload not found")
        
        if upload.extracted_data and 'matching_status' in upload.extracted_data:
            from app.services.staff_accounts_service import recalculate_matching_status
            extracted = recalculate_matching_status(upload.extracted_data.copy())
            matching_status = extracted['matching_status']
            if not matching_status.get('all_resolved', False):
                unresolved_count = matching_status.get('unresolved_count', 0)
                raise AccountsValidationError(
                    f"Cannot confirm: {unresolved_count} unresolved item(s) need attention. "
                    "Please resolve all vendor, stock item, and HSN matches before confirming."
                )
        
        upload = PurchaseInvoiceUploadService.confirm_upload(
            db=db,
            upload_id=upload_id,
            employee=current_user,
            create_vendor_transaction=data.get('create_vendor_transaction', True),
            update_stock_ledger=data.get('update_stock_ledger', True),
            update_accounts_payable=data.get('update_accounts_payable', True),
            update_party_ledger=data.get('update_party_ledger', True),
            confirmation_notes=data.get('confirmation_notes')
        )
        
        # DC-STOCK-MKT-003: refresh marketplace qty from stock_ledger in background
        _bg_refresh_mkt_qty()
        return JSONResponse(content={
            "success": True,
            "message": f"Purchase invoice confirmed: {upload.upload_number}",
            "upload": upload.to_dict(),
            "vendor_transaction_id": upload.vendor_transaction_id
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/reject")
async def reject_purchase_upload(
    upload_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Reject a purchase invoice upload
    DC_PURCHASE_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        rejection_reason = data.get('rejection_reason', 'No reason provided')
        
        upload = PurchaseInvoiceUploadService.reject_upload(
            db=db,
            upload_id=upload_id,
            employee=current_user,
            rejection_reason=rejection_reason
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Purchase invoice rejected: {upload.upload_number}",
            "upload": upload.to_dict()
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/void")
async def void_purchase_upload(
    upload_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-VOID-001: Void a CONFIRMED purchase invoice with full ledger reversal.
    Reverses stock, AP schedules, party ledger, vendor transaction, account ledger.
    Blocks if payment already recorded. Sets status → VOIDED.
    Only VGK/EA/Accounts roles allowed.
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService

        reason = data.get('reason') or data.get('void_reason', '')
        if not reason or len(reason.strip()) < 5:
            return JSONResponse(
                status_code=400,
                content={"success": False, "detail": "Void reason required (min 5 characters)"}
            )

        upload = PurchaseInvoiceUploadService.void_upload(
            db=db,
            upload_id=upload_id,
            employee=current_user,
            reason=reason
        )

        return JSONResponse(content={
            "success": True,
            "message": f"Purchase invoice {upload.upload_number} voided and all entries reversed",
            "upload": upload.to_dict()
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/purchase-uploads/{upload_id}")
async def delete_purchase_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Delete a purchase invoice upload (only non-confirmed invoices)
    DC_PURCHASE_001: VGK/EA/Accounts access only
    WVV Protocol: Hard delete with audit trail
    """
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        result = PurchaseInvoiceUploadService.delete_upload(
            db=db,
            upload_id=upload_id,
            employee=current_user
        )
        
        return JSONResponse(content={
            "success": True,
            "message": result.get('message', 'Invoice deleted successfully'),
            "upload_number": result.get('upload_number')
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-uploads/{upload_id}/pdf")
async def download_purchase_invoice_pdf(
    upload_id: int,
    company_id: int = Query(..., description="Company ID for data segregation — DC Protocol"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_PURCHASE_PDF_002: Generate and download PDF for a purchase document.
    CONFIRMED → PURCHASE INVOICE (tax_invoice mode).
    Non-confirmed → PURCHASE ORDER (purchase_order mode, auto-selected by service).
    Seller = Vendor, Buyer = Company. Courier/transport as separate line items.
    WVV Protocol: Read-only — no data mutation.
    """
    from fastapi.responses import Response
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        from app.models.staff_accounts import PurchaseInvoiceUpload
        pdf_bytes = PurchaseInvoiceUploadService.generate_pdf(db, upload_id, company_id)
        upload = db.query(PurchaseInvoiceUpload).filter_by(id=upload_id).first()
        inv_num = (upload.vendor_invoice_no or upload.upload_number) if upload else f"PIU-{upload_id}"
        _doc_label = 'purchase-invoice' if (upload and upload.status == 'CONFIRMED') else 'purchase-order'
        filename = f"{inv_num.replace('/', '-')}-{_doc_label}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== QUICK CREATE ENDPOINTS FOR INVOICE REVIEW ====================

@router.post("/purchase-uploads/{upload_id}/quick-create-vendor")
async def quick_create_vendor_for_invoice(
    upload_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Quick create vendor from invoice review and link to upload
    DC_PURCHASE_001: Creates vendor and updates upload.vendor_id
    """
    try:
        from app.schemas.staff_accounts import VendorMasterCreate
        from app.services.staff_accounts_service import PurchaseInvoiceUploadService
        
        upload = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.id == upload_id
        ).first()
        if not upload:
            raise AccountsValidationError("Invoice upload not found")
        
        vendor_data = VendorMasterCreate(
            vendor_name=data.get('vendor_name'),
            gst_number=data.get('gst_number'),
            pan_number=data.get('pan_number'),
            vendor_type=data.get('vendor_type', 'SUPPLIER'),
            address_line1=data.get('address_line1'),
            city=data.get('city'),
            state=data.get('state'),
            pincode=data.get('pincode'),
            contact_phone=data.get('contact_phone'),
            contact_email=data.get('contact_email'),
            is_active=True
        )
        
        vendor = VendorMasterService.create_vendor(db, vendor_data, current_user)
        
        upload.vendor_id = vendor.id
        if upload.extracted_data:
            from sqlalchemy.orm.attributes import flag_modified
            from app.services.staff_accounts_service import recalculate_matching_status
            extracted = upload.extracted_data.copy()
            if 'matching_status' in extracted:
                extracted['matching_status']['vendor']['matched'] = True
                extracted['matching_status']['vendor']['matched_id'] = vendor.id
                extracted['matching_status']['vendor']['matched_name'] = vendor.vendor_name
                extracted['matching_status']['vendor']['needs_creation'] = False
            extracted = recalculate_matching_status(extracted)
            upload.extracted_data = extracted
            flag_modified(upload, 'extracted_data')
        
        db.commit()
        
        response = VendorMasterResponse.model_validate(vendor).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Vendor '{vendor.vendor_name}' created and linked to invoice",
            "vendor": response,
            "vendor_id": vendor.id
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/quick-create-stock-item")
async def quick_create_stock_item_for_invoice(
    upload_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Quick create stock item from invoice review and link to line item
    DC_PURCHASE_001: Creates stock item and updates line item
    """
    try:
        from app.schemas.staff_accounts import StockItemMasterCreate
        
        upload = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.id == upload_id
        ).first()
        if not upload:
            raise AccountsValidationError("Invoice upload not found")
        
        line_number = data.get('line_number')
        
        hsn_id = data.get('hsn_id')
        if not hsn_id and data.get('hsn_code'):
            hsn = db.query(HSNMaster).filter(
                HSNMaster.hsn_code == data.get('hsn_code')
            ).first()
            if hsn:
                hsn_id = hsn.id
        
        applicable_companies = data.get('applicable_companies', [upload.company_id])
        if upload.company_id not in applicable_companies:
            applicable_companies.append(upload.company_id)
        
        item_name = data.get('item_name', 'New Item')
        item_code = data.get('item_code')
        if not item_code:
            import uuid
            item_code = f"ITEM_{str(uuid.uuid4())[:8].upper()}"
        
        stock_data = StockItemMasterCreate(
            item_name=item_name,
            item_code=item_code,
            item_category=data.get('item_category', data.get('category', 'PRODUCT')),
            unit_of_measure=data.get('unit_of_measure', 'PCS'),
            hsn_id=hsn_id,
            purchase_rate=Decimal(str(data.get('purchase_rate', 0))),
            selling_rate=Decimal(str(data.get('selling_rate', 0))),
            applicable_companies=applicable_companies,
            is_active=True
        )
        
        stock_item = StockItemMasterService.create_stock_item(db, stock_data, current_user)
        
        if line_number:
            line_item = db.query(PurchaseInvoiceLineItem).filter(
                PurchaseInvoiceLineItem.upload_id == upload_id,
                PurchaseInvoiceLineItem.line_number == line_number
            ).first()
            if line_item:
                line_item.item_id = stock_item.id
                if hsn_id:
                    line_item.hsn_id = hsn_id
        
        if upload.extracted_data:
            from sqlalchemy.orm.attributes import flag_modified
            from app.services.staff_accounts_service import recalculate_matching_status
            extracted = upload.extracted_data.copy()
            if 'matching_status' in extracted:
                for item_status in extracted['matching_status'].get('line_items', []):
                    if item_status.get('line_number') == line_number:
                        item_status['stock_item']['matched'] = True
                        item_status['stock_item']['matched_id'] = stock_item.id
                        item_status['stock_item']['matched_name'] = stock_item.item_name
                        item_status['stock_item']['needs_creation'] = False
                        break
            extracted = recalculate_matching_status(extracted)
            upload.extracted_data = extracted
            flag_modified(upload, 'extracted_data')
        
        db.commit()
        
        response = StockItemMasterResponse.model_validate(stock_item).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Stock item '{stock_item.item_name}' created and linked",
            "stock_item": response,
            "stock_item_id": stock_item.id
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/quick-create-hsn")
async def quick_create_hsn_for_invoice(
    upload_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Quick create HSN code from invoice review and link to line item
    DC_PURCHASE_001: Creates HSN and updates line item
    """
    try:
        from app.schemas.staff_accounts import HSNMasterCreate
        
        upload = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.id == upload_id
        ).first()
        if not upload:
            raise AccountsValidationError("Invoice upload not found")
        
        line_number = data.get('line_number')
        
        hsn_data = HSNMasterCreate(
            hsn_code=data.get('hsn_code'),
            description=data.get('description', f"HSN {data.get('hsn_code')}"),
            cgst_rate=data.get('cgst_rate', data.get('gst_rate', 0) / 2),
            sgst_rate=data.get('sgst_rate', data.get('gst_rate', 0) / 2),
            igst_rate=data.get('igst_rate', data.get('gst_rate', 0)),
            is_active=True
        )
        
        hsn = HSNMasterService.create_hsn(db, hsn_data, current_user)
        
        if line_number:
            line_item = db.query(PurchaseInvoiceLineItem).filter(
                PurchaseInvoiceLineItem.upload_id == upload_id,
                PurchaseInvoiceLineItem.line_number == line_number
            ).first()
            if line_item:
                line_item.hsn_id = hsn.id
                line_item.hsn_code = hsn.hsn_code
                line_item.gst_rate = hsn.cgst_rate + hsn.sgst_rate
        
        if upload.extracted_data:
            from sqlalchemy.orm.attributes import flag_modified
            from app.services.staff_accounts_service import recalculate_matching_status
            extracted = upload.extracted_data.copy()
            if 'matching_status' in extracted:
                for item_status in extracted['matching_status'].get('line_items', []):
                    if item_status.get('line_number') == line_number:
                        item_status['hsn']['matched'] = True
                        item_status['hsn']['matched_id'] = hsn.id
                        item_status['hsn']['matched_code'] = hsn.hsn_code
                        item_status['hsn']['needs_creation'] = False
                        break
            extracted = recalculate_matching_status(extracted)
            upload.extracted_data = extracted
            flag_modified(upload, 'extracted_data')
        
        db.commit()
        
        response = HSNMasterResponse.model_validate(hsn).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"HSN '{hsn.hsn_code}' created and linked",
            "hsn": response,
            "hsn_id": hsn.id
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-uploads/{upload_id}/link-line-item")
async def link_line_item_to_stock(
    upload_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Link an existing stock item to a line item
    DC_PURCHASE_001: Updates line item with existing stock item
    """
    try:
        upload = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.id == upload_id
        ).first()
        if not upload:
            raise AccountsValidationError("Invoice upload not found")
        
        line_number = data.get('line_number')
        stock_item_id = data.get('stock_item_id')
        
        if not line_number or not stock_item_id:
            raise AccountsValidationError("line_number and stock_item_id are required")
        
        stock_item = db.query(StockItemMaster).filter(
            StockItemMaster.id == stock_item_id
        ).first()
        if not stock_item:
            raise AccountsValidationError("Stock item not found")
        
        line_item = db.query(PurchaseInvoiceLineItem).filter(
            PurchaseInvoiceLineItem.upload_id == upload_id,
            PurchaseInvoiceLineItem.line_number == line_number
        ).first()
        
        if line_item:
            line_item.item_id = stock_item.id
            if stock_item.hsn_id:
                hsn = db.query(HSNMaster).filter(HSNMaster.id == stock_item.hsn_id).first()
                if hsn:
                    line_item.hsn_id = hsn.id
                    line_item.hsn_code = hsn.hsn_code
                    line_item.gst_rate = hsn.cgst_rate + hsn.sgst_rate
        
        if upload.extracted_data:
            extracted = upload.extracted_data.copy()
            if 'matching_status' in extracted:
                for item_status in extracted['matching_status'].get('line_items', []):
                    if item_status.get('line_number') == line_number:
                        if not item_status['stock_item']['matched']:
                            item_status['stock_item']['matched'] = True
                            item_status['stock_item']['matched_id'] = stock_item.id
                            item_status['stock_item']['matched_name'] = stock_item.item_name
                            item_status['stock_item']['needs_creation'] = False
                            if extracted['matching_status']['unresolved_count'] > 0:
                                extracted['matching_status']['unresolved_count'] -= 1
                        break
                extracted['matching_status']['all_resolved'] = extracted['matching_status']['unresolved_count'] == 0
            upload.extracted_data = extracted
        
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": f"Line item linked to '{stock_item.item_name}'",
            "stock_item_id": stock_item.id,
            "stock_item_name": stock_item.item_name
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== VENDOR MASTER ENDPOINTS ====================

@router.post("/vendors", response_model=VendorMasterResponse)
async def create_vendor(
    data: VendorMasterCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new vendor with GST/PAN validation
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor = VendorMasterService.create_vendor(db, data, current_user)
        response = VendorMasterResponse.model_validate(vendor).model_dump(mode='json')
        
        await sfms_cache.invalidate_vendors()
        
        return JSONResponse(content={
            "success": True,
            "message": f"Vendor '{vendor.vendor_name}' created successfully",
            "vendor": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendors", response_model=VendorMasterListResponse)
async def list_vendors(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=500, description="Items per page"),
    vendor_type: Optional[str] = Query(None, description="Filter by type: PRODUCT, SERVICE, BOTH"),
    is_active: Optional[bool] = Query(True, description="Filter by active status (default: active only)"),
    include_inactive: bool = Query(False, description="Include inactive vendors"),
    search: Optional[str] = Query(None, description="Search by name, code, GST, city"),
    company_ids: Optional[str] = Query(None, description="DC_MULTICOMPANY_001: Comma-separated company IDs to filter by applicable_companies"),
    skip_cache: bool = Query(False, description="Skip Redis cache for fresh data"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List vendors with filtering and pagination. Defaults to active vendors only.
    DC_SFMS_001: VGK/EA/Accounts access only
    Phase 3: Redis caching for search results (3-minute TTL)
    """
    try:
        use_cache = (
            not skip_cache and 
            is_active is True and 
            not include_inactive and
            not company_ids
        )
        
        if use_cache:
            cached_result = await sfms_cache.get_vendors_cached(
                search or "", page, page_size, vendor_type
            )
            if cached_result:
                import logging
                logging.getLogger(__name__).info(f"[SFMS-CACHE] HIT vendors search={search}")
                return JSONResponse(content=cached_result)
        
        # DC_MULTICOMPANY_001: Parse comma-separated company_ids for vendor filtering
        parsed_company_ids = None
        if company_ids:
            try:
                parsed_company_ids = [int(c.strip()) for c in company_ids.split(',') if c.strip().isdigit()]
            except Exception:
                parsed_company_ids = None
        vendors, total = VendorMasterService.list_vendors(
            db, current_user, page, page_size, vendor_type, is_active, include_inactive, search,
            company_ids=parsed_company_ids
        )
        
        vendor_list = [
            VendorMasterResponse.model_validate(v).model_dump(mode='json')
            for v in vendors
        ]
        
        response_data = {
            "success": True,
            "vendors": vendor_list,
            "total": total,
            "page": page,
            "page_size": page_size
        }
        
        if use_cache:
            await sfms_cache.set_vendors_cached(
                search or "", page, page_size, vendor_type, response_data
            )
            import logging
            logging.getLogger(__name__).info(f"[SFMS-CACHE] SET vendors search={search}")
        
        return JSONResponse(content=response_data)
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendors/{vendor_id}", response_model=VendorMasterResponse)
async def get_vendor(
    vendor_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single vendor by ID
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor = VendorMasterService.get_vendor(db, vendor_id, current_user)
        response = VendorMasterResponse.model_validate(vendor).model_dump(mode='json')
        # DC_VENDOR_OB_OVERRIDE_001: Enrich response with current OB from AccountLedgerMaster
        try:
            from app.models.staff_accounts import AccountLedgerMaster as _ALM_GET
            _alm = db.query(_ALM_GET).filter(
                _ALM_GET.account_code == vendor.vendor_code,
                _ALM_GET.account_type == 'PARTY'
            ).first()
            if _alm:
                response['opening_balance'] = float(_alm.opening_balance or 0)
                response['opening_balance_type'] = _alm.opening_balance_type or 'CREDIT'
                response['opening_balance_date'] = str(_alm.opening_balance_date) if _alm.opening_balance_date else None
            # DC_VENDOR_OB_PERCO_001: Return per-company OB values (exclude zeroed rows)
            _all_alm = db.query(_ALM_GET).filter(
                _ALM_GET.account_code == vendor.vendor_code,
                _ALM_GET.account_type == 'PARTY',
                _ALM_GET.opening_balance > 0   # skip rows cleared by DC-OB-REPLACE-001
            ).all()
            response['opening_balances'] = [
                {'company_id': _a.company_id,
                 'amount': float(_a.opening_balance or 0),
                 'type': _a.opening_balance_type or 'CREDIT',
                 'date': str(_a.opening_balance_date) if _a.opening_balance_date else None}
                for _a in _all_alm
            ]
        except Exception:
            pass
        return JSONResponse(content={
            "success": True,
            "vendor": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendors/code/{vendor_code}", response_model=VendorMasterResponse)
async def get_vendor_by_code(
    vendor_code: str,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single vendor by code
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor = VendorMasterService.get_vendor_by_code(db, vendor_code, current_user)
        response = VendorMasterResponse.model_validate(vendor).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "vendor": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/vendors/{vendor_id}", response_model=VendorMasterResponse)
async def update_vendor(
    vendor_id: int,
    data: VendorMasterUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an existing vendor
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor = VendorMasterService.update_vendor(db, vendor_id, data, current_user)
        response = VendorMasterResponse.model_validate(vendor).model_dump(mode='json')
        
        await sfms_cache.invalidate_vendors()
        
        return JSONResponse(content={
            "success": True,
            "message": f"Vendor '{vendor.vendor_name}' updated successfully",
            "vendor": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/vendors/{vendor_id}")
async def delete_vendor(
    vendor_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft delete (deactivate) a vendor
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        VendorMasterService.delete_vendor(db, vendor_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Vendor deactivated successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendors/{vendor_id}/products")
async def set_vendor_products(
    vendor_id: int,
    product_ids: List[int] = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Set products associated with a vendor (replaces existing associations)
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        associations = VendorMasterService.set_vendor_products(
            db, vendor_id, product_ids, current_user
        )
        return JSONResponse(content={
            "success": True,
            "message": f"Set {len(associations)} products for vendor",
            "count": len(associations)
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendors/{vendor_id}/products")
async def get_vendor_products(
    vendor_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get products associated with a vendor
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        products = VendorMasterService.get_vendor_products(db, vendor_id, current_user)
        return JSONResponse(content={
            "success": True,
            "products": products
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendors/{vendor_id}/purchase-history-products")
async def get_vendor_purchase_history_products(
    vendor_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Return distinct stock items that appear in purchase invoices for this vendor.
    DC_VENDOR_PHIST_001: used by Products tab auto-populate.
    """
    try:
        rows = db.execute(
            text("""
                SELECT DISTINCT
                    li.item_id,
                    COALESCE(sm.item_code, li.item_code) AS item_code,
                    COALESCE(sm.item_name, li.item_description) AS item_name
                FROM purchase_invoice_line_items li
                JOIN purchase_invoice_uploads pu ON pu.id = li.upload_id
                LEFT JOIN stock_item_master sm ON sm.id = li.item_id
                WHERE pu.vendor_id = :vid
                  AND li.item_id IS NOT NULL
                ORDER BY item_name
            """),
            {"vid": vendor_id}
        ).fetchall()
        products = [
            {"item_id": r[0], "item_code": r[1] or "", "item_name": r[2] or ""}
            for r in rows
        ]
        return JSONResponse(content={"success": True, "products": products, "count": len(products)})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendors/{vendor_id}/scanner")
async def upload_vendor_scanner(
    vendor_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Upload payment scanner QR code image for a vendor
    DC_SFMS_001: VGK/EA/Accounts access only
    Security: MIME validation, size limits, sanitized filenames
    """
    try:
        import os
        import uuid
        import magic
        
        MAX_FILE_SIZE = 2 * 1024 * 1024
        ALLOWED_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
        MIME_TO_EXT = {'image/png': 'png', 'image/jpeg': 'jpg', 'image/gif': 'gif', 'image/webp': 'webp'}
        
        content = await file.read()
        
        if len(content) > MAX_FILE_SIZE:
            return JSONResponse(content={
                "success": False,
                "message": f"File too large. Maximum size: 2MB"
            }, status_code=400)
        
        detected_mime = magic.from_buffer(content, mime=True)
        if detected_mime not in ALLOWED_MIME_TYPES:
            return JSONResponse(content={
                "success": False,
                "message": f"Invalid file type. Allowed: PNG, JPG, GIF, WEBP"
            }, status_code=400)
        
        ext = MIME_TO_EXT.get(detected_mime, 'png')
        
        from app.services.object_storage import storage_service
        
        safe_filename = f"scanner_{vendor_id}_{uuid.uuid4().hex}.{ext}"
        storage_path = f"vendor_scanners/{safe_filename}"
        
        success = storage_service.upload_file(storage_path, content)
        if not success:
            return JSONResponse(content={
                "success": False,
                "message": "Failed to upload scanner image to storage"
            }, status_code=500)
        
        scanner_path = f"/storage/{storage_path}"
        vendor = VendorMasterService.update_payment_scanner(
            db, vendor_id, scanner_path, current_user
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Payment scanner uploaded successfully",
            "scanner_path": scanner_path
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendors/{vendor_id}/stamp")
async def upload_vendor_stamp(
    vendor_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Upload company stamp image for a solar vendor.
    DC_SFMS_001: Returns /storage/... path stored in vendor.stamp_image_url.
    Security: MIME validation, 2 MB size limit, sanitized filenames.
    """
    try:
        import uuid
        import magic

        MAX_FILE_SIZE = 2 * 1024 * 1024
        ALLOWED_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
        MIME_TO_EXT = {'image/png': 'png', 'image/jpeg': 'jpg', 'image/gif': 'gif', 'image/webp': 'webp'}

        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            return JSONResponse(content={"success": False, "message": "File too large. Maximum size: 2MB"}, status_code=400)

        detected_mime = magic.from_buffer(content, mime=True)
        if detected_mime not in ALLOWED_MIME_TYPES:
            return JSONResponse(content={"success": False, "message": "Invalid file type. Allowed: PNG, JPG, GIF, WEBP"}, status_code=400)

        ext = MIME_TO_EXT.get(detected_mime, 'png')
        from app.services.object_storage import storage_service
        safe_filename = f"stamp_{vendor_id}_{uuid.uuid4().hex}.{ext}"
        storage_path = f"vendor_stamps/{safe_filename}"
        success = storage_service.upload_file(storage_path, content)
        if not success:
            return JSONResponse(content={"success": False, "message": "Failed to upload stamp image"}, status_code=500)

        image_url = f"/storage/{storage_path}"
        return JSONResponse(content={"success": True, "image_url": image_url})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendors/{vendor_id}/signature")
async def upload_vendor_signature(
    vendor_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Upload technician signature image for a solar vendor.
    DC_SFMS_001: Returns /storage/... path stored in vendor.tech_signature_url.
    Security: MIME validation, 2 MB size limit, sanitized filenames.
    """
    try:
        import uuid
        import magic

        MAX_FILE_SIZE = 2 * 1024 * 1024
        ALLOWED_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
        MIME_TO_EXT = {'image/png': 'png', 'image/jpeg': 'jpg', 'image/gif': 'gif', 'image/webp': 'webp'}

        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            return JSONResponse(content={"success": False, "message": "File too large. Maximum size: 2MB"}, status_code=400)

        detected_mime = magic.from_buffer(content, mime=True)
        if detected_mime not in ALLOWED_MIME_TYPES:
            return JSONResponse(content={"success": False, "message": "Invalid file type. Allowed: PNG, JPG, GIF, WEBP"}, status_code=400)

        ext = MIME_TO_EXT.get(detected_mime, 'png')
        from app.services.object_storage import storage_service
        safe_filename = f"signature_{vendor_id}_{uuid.uuid4().hex}.{ext}"
        storage_path = f"vendor_signatures/{safe_filename}"
        success = storage_service.upload_file(storage_path, content)
        if not success:
            return JSONResponse(content={"success": False, "message": "Failed to upload signature image"}, status_code=500)

        image_url = f"/storage/{storage_path}"
        return JSONResponse(content={"success": True, "image_url": image_url})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendors/{vendor_id}/rep-signature")
async def upload_vendor_rep_signature(
    vendor_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Upload authorized representative signature image for a solar vendor.
    DC_SFMS_001: Returns /storage/... path stored in vendor.rep_signature_url.
    This signature appears on all "Vendor Signature with Stamp" blocks in solar documents.
    Security: MIME validation, 2 MB size limit, sanitized filenames.
    """
    try:
        import uuid
        import magic

        MAX_FILE_SIZE = 2 * 1024 * 1024
        ALLOWED_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
        MIME_TO_EXT = {'image/png': 'png', 'image/jpeg': 'jpg', 'image/gif': 'gif', 'image/webp': 'webp'}

        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            return JSONResponse(content={"success": False, "message": "File too large. Maximum size: 2MB"}, status_code=400)

        detected_mime = magic.from_buffer(content, mime=True)
        if detected_mime not in ALLOWED_MIME_TYPES:
            return JSONResponse(content={"success": False, "message": "Invalid file type. Allowed: PNG, JPG, GIF, WEBP"}, status_code=400)

        ext = MIME_TO_EXT.get(detected_mime, 'png')
        from app.services.object_storage import storage_service
        safe_filename = f"rep_sig_{vendor_id}_{uuid.uuid4().hex}.{ext}"
        storage_path = f"vendor_rep_signatures/{safe_filename}"
        success = storage_service.upload_file(storage_path, content)
        if not success:
            return JSONResponse(content={"success": False, "message": "Failed to upload representative signature image"}, status_code=500)

        image_url = f"/storage/{storage_path}"
        return JSONResponse(content={"success": True, "image_url": image_url})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/pincode/{pincode}")
@router.get("/pincode-lookup/{pincode}")
async def lookup_pincode(
    pincode: str,
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Lookup Indian PIN code to auto-fill city and state
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        from app.services.staff_accounts_service import PinCodeLookupService
        
        result = PinCodeLookupService.lookup_pincode(pincode)
        return JSONResponse(content=result)
    except Exception as e:
        return handle_accounts_error(e)


# ==================== STOCK ITEM MASTER ENDPOINTS ====================

@router.post("/stock-items", response_model=StockItemMasterResponse)
async def create_stock_item(
    data: StockItemMasterCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new stock item with HSN validation
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        item = StockItemMasterService.create_stock_item(db, data, current_user)
        response = StockItemMasterResponse.model_validate(item).model_dump(mode='json')
        
        if item.applicable_companies:
            for company_id in item.applicable_companies:
                await sfms_cache.invalidate_stock_items(company_id)
        
        return JSONResponse(content={
            "success": True,
            "message": f"Stock item '{item.item_name}' created successfully",
            "stock_item": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-items/template")
async def download_stock_items_template(
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Download Excel template for bulk stock item upload.
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import StreamingResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Items"
        
        headers = [
            "item_code*", "item_name*", "item_category*", "applicable_companies",
            "unit_of_measure*", "description", "brand", "model_compat",
            "specification", "size", "colors",
            "hsn_code", "default_gst_rate", "reorder_level", "purchase_rate", "selling_rate",
            "image_url"
        ]
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 18
        
        sample_data = [
            "STK001", "EV Battery 48V", "PRODUCT", "1,2",
            "UNIT", "High capacity lithium battery", "Luminous", "Hero Splendor / Honda Activa",
            "48V 60Ah Li-ion", "Large",
            "Black,Silver", "85044090", "18", "10", "25000", "32000",
            "https://drive.google.com/file/d/xxx/view"
        ]
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
        
        ws_info = wb.create_sheet("Instructions")
        instructions = [
            ("Stock Items Bulk Upload Template", None),
            ("", None),
            ("REQUIRED FIELDS (marked with *):", None),
            ("item_code*", "Unique code for the item (e.g., STK001, EV-BAT-001)"),
            ("item_name*", "Name of the stock item"),
            ("item_category*", "One of: PRODUCT, RAW_MATERIAL, CONSUMABLE, SPARE_PART, ACCESSORY"),
            ("unit_of_measure*", "One of: PCS, KG, LTR, MTR, SET, BOX, PACK, PAIR, UNIT"),
            ("", None),
            ("OPTIONAL FIELDS:", None),
            ("applicable_companies", "Comma-separated company IDs (e.g., 1,2,3)"),
            ("description", "Item description text"),
            ("brand", "Brand / manufacturer name (e.g., Luminous, TVS, Honda)"),
            ("model_compat", "Compatible with / Model — also accepts 'compatible with', 'model' columns from Google Sheet"),
            ("specification", "Technical specs — also accepts 'spec' column"),
            ("size", "Size/dimensions (e.g., Large, 10x20cm)"),
            ("colors", "Comma-separated colors — also accepts 'color' column (e.g., Red,Blue,Green)"),
            ("hsn_code", "4-8 digit HSN/SAC code"),
            ("default_gst_rate", "GST rate percentage (default: 18)"),
            ("reorder_level", "Minimum stock level for reorder alert (default: 0)"),
            ("purchase_rate", "Purchase price (default: 0)"),
            ("selling_rate", "Selling price (default: 0)"),
            ("image_url", "Google Drive share link or direct image URL (will be downloaded and stored)"),
            ("", None),
            ("IMAGE URL NOTES:", None),
            ("", "For Google Drive: Use shareable link (File > Share > Anyone with link)"),
            ("", "Supported formats: JPEG, PNG, GIF, WebP (max 5MB)"),
            ("", "Images are automatically compressed for faster loading"),
        ]
        for row, (field, desc) in enumerate(instructions, 1):
            ws_info.cell(row=row, column=1, value=field)
            if desc:
                ws_info.cell(row=row, column=2, value=desc)
        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=stock_items_template.xlsx"}
        )
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-items/bulk-upload", response_model=StockItemBulkUploadResult)
async def bulk_upload_stock_items(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Bulk upload stock items from Excel file.
    DC_SFMS_001: VGK/EA/Accounts access only
    Returns detailed error list for correction.
    """
    try:
        import openpyxl
        from io import BytesIO
        from decimal import Decimal, InvalidOperation

        def _notna(v):
            if v is None:
                return False
            return str(v).strip().lower() not in ('', 'none', 'nan', 'nat')

        def _get_col(row, *keys):
            """Flexible column lookup — tries each key in order, returns first non-empty value."""
            for k in keys:
                v = row.get(k)
                if _notna(v):
                    return str(v).strip()
            return None

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise AccountsValidationError("Please upload an Excel file (.xlsx or .xls)")

        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            raise AccountsValidationError("Excel file is empty")
        headers = [str(h).replace('*', '').strip().lower() if h is not None else '' for h in raw_headers]

        required_columns = ['item_code', 'item_name', 'item_category', 'unit_of_measure']
        missing_cols = [c for c in required_columns if c not in headers]
        if missing_cols:
            raise AccountsValidationError(f"Missing required columns: {', '.join(missing_cols)}")

        # DC-STOCK-CAT-001: No fixed category list — any non-empty 2-50 char string is valid
        valid_uoms = ['PCS', 'KG', 'LTR', 'MTR', 'SET', 'BOX', 'PACK', 'PAIR', 'UNIT']

        from app.models.staff_accounts import StockItemMaster

        errors = []
        created_items = []
        batch_codes = set()

        existing_codes = set(
            code[0].upper() for code in
            db.query(StockItemMaster.item_code).all()
        )

        row_num = 1
        for row_num, raw_row in enumerate(rows_iter, start=2):
            if all(v is None for v in raw_row):
                continue
            row = dict(zip(headers, raw_row))
            row_errors = []
            original_data = row

            item_code = str(row.get('item_code', '')).strip().upper() if _notna(row.get('item_code')) else ''
            item_name = str(row.get('item_name', '')).strip() if _notna(row.get('item_name')) else ''
            item_category = str(row.get('item_category', 'PRODUCT')).strip().upper() if _notna(row.get('item_category')) else 'PRODUCT'
            unit_of_measure = str(row.get('unit_of_measure', 'PCS')).strip().upper() if _notna(row.get('unit_of_measure')) else 'PCS'

            if not item_code:
                row_errors.append("Item code is required")
            elif not item_code.replace('_', '').replace('-', '').isalnum():
                row_errors.append("Item code must be alphanumeric with underscores/hyphens only")
            elif item_code in existing_codes:
                row_errors.append(f"Item code '{item_code}' already exists in database")
            elif item_code in batch_codes:
                row_errors.append(f"Item code '{item_code}' is duplicated in this upload")

            if not item_name:
                row_errors.append("Item name is required")
            elif len(item_name) < 2:
                row_errors.append("Item name must be at least 2 characters")

            # DC-STOCK-CAT-001: Allow any category string (2-50 chars)
            if not item_category or len(item_category) < 2:
                row_errors.append("Item category must be at least 2 characters")

            if unit_of_measure not in valid_uoms:
                row_errors.append(f"Invalid unit. Must be one of: {', '.join(valid_uoms)}")

            hsn_code = None
            if _notna(row.get('hsn_code')):
                hsn_code = str(row.get('hsn_code')).strip()
                if hsn_code and (not hsn_code.isdigit() or len(hsn_code) < 4 or len(hsn_code) > 8):
                    row_errors.append("HSN code must be 4-8 digits")

            default_gst_rate = Decimal("18.00")
            if _notna(row.get('default_gst_rate')):
                try:
                    default_gst_rate = Decimal(str(row.get('default_gst_rate')))
                    if default_gst_rate < 0 or default_gst_rate > 100:
                        row_errors.append("GST rate must be between 0 and 100")
                except (InvalidOperation, ValueError):
                    row_errors.append("Invalid GST rate value")

            reorder_level = 0
            if _notna(row.get('reorder_level')):
                try:
                    reorder_level = int(float(row.get('reorder_level')))
                    if reorder_level < 0:
                        row_errors.append("Reorder level cannot be negative")
                except ValueError:
                    row_errors.append("Invalid reorder level value")

            purchase_rate = Decimal("0.00")
            if _notna(row.get('purchase_rate')):
                try:
                    purchase_rate = Decimal(str(row.get('purchase_rate')))
                    if purchase_rate < 0:
                        row_errors.append("Purchase rate cannot be negative")
                except (InvalidOperation, ValueError):
                    row_errors.append("Invalid purchase rate value")

            selling_rate = Decimal("0.00")
            if _notna(row.get('selling_rate')):
                try:
                    selling_rate = Decimal(str(row.get('selling_rate')))
                    if selling_rate < 0:
                        row_errors.append("Selling rate cannot be negative")
                except (InvalidOperation, ValueError):
                    row_errors.append("Invalid selling rate value")

            applicable_companies = []
            if _notna(row.get('applicable_companies')):
                try:
                    company_str = str(row.get('applicable_companies')).strip()
                    if company_str:
                        applicable_companies = [int(c.strip()) for c in company_str.split(',') if c.strip()]
                except ValueError:
                    row_errors.append("Invalid company IDs format. Use comma-separated numbers.")

            # Brand — accept 'brand' column
            brand = _get_col(row, 'brand')

            # Compatible With / Model — accept many aliases matching Google Sheet headers
            _compat = _get_col(row, 'compatible with / model', 'compatible with/model',
                               'compatible_with_model', 'model_compat')
            _compat_part = _get_col(row, 'compatible with', 'compatible_with', 'compatible')
            _model_part = _get_col(row, 'model', 'model name', 'model_name')
            if _compat:
                model_compat = _compat
            elif _compat_part or _model_part:
                model_compat = ' / '.join(p for p in [_compat_part, _model_part] if p) or None
            else:
                model_compat = None

            # Colors — accept 'colors' or 'color'
            colors = None
            color_str = _get_col(row, 'colors', 'color')
            if color_str:
                colors = [c.strip() for c in color_str.split(',') if c.strip()]

            # Image URL — accept 'image_url' or Google Sheet's 'image drive url'
            image_url = _get_col(row, 'image_url', 'image drive url', 'image_drive_url',
                                 'image url', 'image link', 'drive url')

            if row_errors:
                errors.append(StockItemBulkUploadError(
                    row_number=row_num,
                    item_code=item_code if item_code else None,
                    item_name=item_name if item_name else None,
                    errors=row_errors,
                    original_data={k: (str(v) if _notna(v) else None) for k, v in original_data.items()}
                ))
            else:
                new_item = StockItemMaster(
                    item_code=item_code,
                    item_name=item_name,
                    item_category=item_category,
                    unit_of_measure=unit_of_measure,
                    description=_get_col(row, 'description'),
                    brand=brand,
                    model_compat=model_compat,
                    specification=_get_col(row, 'specification', 'spec', 'specifications'),
                    size=_get_col(row, 'size', 'dimensions'),
                    colors=colors,
                    hsn_code=hsn_code,
                    default_gst_rate=default_gst_rate,
                    reorder_level=reorder_level,
                    purchase_rate=purchase_rate,
                    selling_rate=selling_rate,
                    applicable_companies=applicable_companies,
                    is_active=True,
                    created_by_id=current_user.id
                )
                db.add(new_item)
                db.flush()
                
                if image_url:
                    try:
                        import requests
                        from app.services.universal_upload_service import UniversalUploadService

                        # Google Drive FOLDER link → store as folder reference only
                        if 'drive.google.com/drive/folders' in image_url or ('drive.google.com' in image_url and '/drive/folders/' in image_url):
                            stock_image = StockItemImage(
                                stock_item_id=new_item.id,
                                original_path=image_url,
                                file_name='google_drive_folder',
                                is_primary=False,
                                display_order=0,
                                source_type='folder_link',
                                source_url=image_url,
                                uploaded_by_id=current_user.id
                            )
                            db.add(stock_image)
                        else:
                            # Direct image URL or Google Drive file link
                            download_url = image_url
                            if 'drive.google.com' in image_url and '/file/d/' in image_url:
                                file_id = image_url.split('/file/d/')[1].split('/')[0]
                                download_url = f'https://drive.google.com/uc?export=download&id={file_id}'

                            response = requests.get(download_url, timeout=30, stream=True, allow_redirects=True)
                            if response.status_code == 200:
                                content = response.content
                                if len(content) <= 5 * 1024 * 1024:
                                    content_type = response.headers.get('content-type', 'image/jpeg')
                                    if content_type.startswith('image/') or 'octet-stream' in content_type:
                                        filename = f"{item_code}_image.jpg"

                                        upload_service = UniversalUploadService(db, current_user)
                                        result = upload_service.upload_file(
                                            file_data=content,
                                            original_filename=filename,
                                            content_type='image/jpeg',
                                            entity_id=new_item.id,
                                            segment="STOCK",
                                            description=f"Bulk upload image for {item_code}"
                                        )

                                        if result.get('success'):
                                            stock_image = StockItemImage(
                                                stock_item_id=new_item.id,
                                                original_path=result.get('file_path', ''),
                                                compressed_path=result.get('compressed_path'),
                                                thumbnail_path=result.get('thumbnail_path'),
                                                file_name=filename,
                                                file_size=len(content),
                                                compressed_size=result.get('compressed_size'),
                                                mime_type='image/jpeg',
                                                is_primary=True,
                                                display_order=0,
                                                source_type='url',
                                                source_url=image_url,
                                                uploaded_by_id=current_user.id
                                            )
                                            db.add(stock_image)
                    except Exception as img_err:
                        import logging
                        logging.getLogger(__name__).warning(f"Failed to download image for {item_code}: {str(img_err)}")
                
                batch_codes.add(item_code)
                created_items.append(item_code)
        
        if created_items:
            db.commit()
            await sfms_cache.invalidate_stock_items(None)
        
        result = StockItemBulkUploadResult(
            success=len(errors) == 0,
            total_rows=max(0, row_num - 1),
            created_count=len(created_items),
            error_count=len(errors),
            errors=[e.model_dump() for e in errors],
            created_items=created_items
        )
        
        return JSONResponse(content=result.model_dump())
        
    except AccountsValidationError as e:
        return handle_accounts_error(e)
    except Exception as e:
        db.rollback()
        import logging
        logging.getLogger(__name__).error(f"Bulk upload error: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": f"Upload failed: {str(e)}"}
        )


@router.get("/stock-items", response_model=StockItemMasterListResponse)
async def list_stock_items(
    company_id: Optional[int] = Query(None, description="Filter by company (mandatory company-wise segregation)"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=2000, description="Items per page"),
    item_category: Optional[str] = Query(None, description="Filter by category: PRODUCT, RAW_MATERIAL, CONSUMABLE, SPARE_PART, ACCESSORY"),
    is_active: Optional[bool] = Query(True, description="Filter by active status (default: active only)"),
    include_inactive: bool = Query(False, description="Include inactive items"),
    search: Optional[str] = Query(None, description="Search by code, name, HSN, description"),
    specification: Optional[str] = Query(None, description="Filter by specification text"),
    colors: Optional[str] = Query(None, description="Filter by color (comma-separated for multiple)"),
    stock_level: Optional[str] = Query(None, description="Filter by stock level: in_stock, out_of_stock, below_reorder, below_zero"),
    include_summary: bool = Query(True, description="Include summary metrics"),
    skip_cache: bool = Query(False, description="Skip Redis cache for fresh data"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List stock items with filtering and pagination. Defaults to active items only.
    DC_SFMS_001: VGK/EA/Accounts access only
    DC_STOCK_002: Company-wise filtering
    DC_STOCK_SUMMARY: Summary metrics for dashboard
    DC_STOCK_LEVEL_FILTER_001: Stock level filter (in_stock, out_of_stock, below_reorder, below_zero)
    Phase 3: Redis caching for search results (3-minute TTL)
    """
    try:
        # DC_CACHE_FIX_001 (Apr 2026): Removed company_id is not None restriction
        # so all-companies view is also cached. Summary requests still bypass cache
        # (they now run in 2 queries, not N+1, so they are fast regardless).
        use_cache = (
            not skip_cache and
            is_active is True and
            not include_inactive and
            item_category is None and
            specification is None and
            colors is None and
            stock_level is None and
            not include_summary
        )
        
        if use_cache:
            cached_result = await sfms_cache.get_stock_items_cached(
                company_id, search or "", page, page_size
            )
            if cached_result:
                import logging
                logging.getLogger(__name__).info(f"[SFMS-CACHE] HIT stock_items company={company_id} search={search}")
                return JSONResponse(content=cached_result)
        
        items, total = StockItemMasterService.list_stock_items(
            db, current_user, page, page_size, item_category, is_active, include_inactive, search, company_id, specification, colors, stock_level
        )
        
        item_list = []
        
        item_ids = [item.id for item in items]
        images_by_item = {}
        if item_ids:
            all_images = db.query(StockItemImage).filter(
                StockItemImage.stock_item_id.in_(item_ids)
            ).order_by(StockItemImage.is_primary.desc(), StockItemImage.display_order).all()
            for img in all_images:
                if img.stock_item_id not in images_by_item:
                    images_by_item[img.stock_item_id] = []
                images_by_item[img.stock_item_id].append({
                    'id': img.id,
                    'original_path': img.original_path,
                    'compressed_path': img.compressed_path,
                    'thumbnail_path': img.thumbnail_path,
                    'file_name': img.file_name,
                    'is_primary': img.is_primary,
                    'source_type': img.source_type or 'upload',
                    'source_url': img.source_url,
                })
        
        # DC_STOCK_INOUT_001: Batch-fetch qty_in / qty_out totals for all items in one query
        from app.models.staff_accounts import StockLedger as _SL_list
        from sqlalchemy import func as _func_list
        _inout_q = db.query(
            _SL_list.item_id,
            _func_list.coalesce(_func_list.sum(_SL_list.quantity_in), 0).label('total_in'),
            _func_list.coalesce(_func_list.sum(_SL_list.quantity_out), 0).label('total_out')
        ).filter(_SL_list.item_id.in_(item_ids))
        if company_id:
            _inout_q = _inout_q.filter(_SL_list.company_id == company_id)
        _inout_map = {r.item_id: (float(r.total_in), float(r.total_out)) for r in _inout_q.group_by(_SL_list.item_id).all()}

        # DC_STOCK_PRICING_BATCH_001 (Apr 2026): Batch purchase rates — replaces per-item N+1 call
        from decimal import Decimal as _Dec
        _price_map = {}  # item_id -> (rate: Decimal, is_average: bool)
        _avg_needed_ids = [item.id for item in items if not (item.purchase_rate and item.purchase_rate > 0)]
        if _avg_needed_ids:
            _pr_q = db.query(
                _SL_list.item_id,
                _func_list.sum(_SL_list.quantity_in * _SL_list.unit_rate).label('tv'),
                _func_list.sum(_SL_list.quantity_in).label('tq'),
            ).filter(
                _SL_list.item_id.in_(_avg_needed_ids),
                _SL_list.entry_type == 'PURCHASE',
                _SL_list.quantity_in > 0,
            )
            if company_id:
                _pr_q = _pr_q.filter(_SL_list.company_id == company_id)
            for _pr in _pr_q.group_by(_SL_list.item_id).all():
                if _pr.tq and _pr.tq > 0:
                    _price_map[_pr.item_id] = (
                        (_Dec(str(_pr.tv)) / _Dec(str(_pr.tq))).quantize(_Dec('0.01')), True
                    )

        # DC_STOCK_MKT_BATCH_001 (Apr 2026): Batch marketplace qty — replaces per-item N+1
        _mkt_skus = [item.marketplace_sku for item in items if item.marketplace_sku]
        _mkt_qty_map = {}
        if _mkt_skus:
            from app.models.marketplace import MarketspareItem as _MktBatch
            for _mr in db.query(_MktBatch.sku, _MktBatch.available_qty).filter(
                _MktBatch.sku.in_(_mkt_skus)
            ).all():
                _mkt_qty_map[_mr.sku] = _mr.available_qty

        for item in items:
            item_data = StockItemMasterResponse.model_validate(item).model_dump(mode='json')
            
            item_images = images_by_item.get(item.id, [])
            item_data['images'] = item_images
            # Skip folder_link entries for actual image display
            displayable = [img for img in item_images if img.get('source_type') != 'folder_link']
            primary = next((img for img in displayable if img['is_primary']), displayable[0] if displayable else None)
            item_data['primary_image'] = (primary.get('compressed_path') or primary.get('original_path')) if primary else None
            # Expose folder reference for display
            folder_links = [img for img in item_images if img.get('source_type') == 'folder_link']
            item_data['has_folder_link'] = bool(folder_links)
            item_data['folder_url'] = folder_links[0].get('source_url') if folder_links else None
            
            _in, _out = _inout_map.get(item.id, (0.0, 0.0))
            item_data['total_qty_in'] = _in
            item_data['total_qty_out'] = _out
            # DC_STOCK_BALANCE_001: derive balance from batch totals for consistency across all-companies view
            item_data['current_stock'] = round(_in - _out, 4)
            
            # DC_STOCK_PRICING_BATCH_001: use pre-fetched batch pricing map
            if item.purchase_rate and item.purchase_rate > 0:
                _eff_rate = _Dec(str(item.purchase_rate))
                _is_avg = False
            else:
                _eff_rate, _is_avg = _price_map.get(item.id, (_Dec('0'), False))
            item_data['effective_purchase_rate'] = float(_eff_rate)
            item_data['is_average_price'] = _is_avg

            if item.selling_rate and item.selling_rate > 0:
                item_data['effective_selling_rate'] = float(item.selling_rate)
            else:
                suggested = StockItemMasterService.calculate_selling_rate(_eff_rate)
                item_data['effective_selling_rate'] = float(suggested)
                item_data['is_suggested_price'] = True

            # DC_STOCK_MKT_BATCH_001: use pre-fetched marketplace qty map
            item_data['marketplace_available_qty'] = (
                _mkt_qty_map.get(item.marketplace_sku) if item.marketplace_sku else None
            )

            item_list.append(item_data)
        
        response_data = {
            "success": True,
            "stock_items": item_list,
            "total": total,
            "page": page,
            "page_size": page_size
        }
        
        if include_summary:
            summary = StockItemMasterService.get_stock_items_summary(db, current_user, company_id)
            response_data["summary"] = summary
        
        if use_cache:
            await sfms_cache.set_stock_items_cached(
                company_id, search or "", page, page_size, response_data
            )
            import logging
            logging.getLogger(__name__).info(f"[SFMS-CACHE] SET stock_items company={company_id} search={search}")
        
        return JSONResponse(content=response_data)
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-items/generate-code")
async def generate_stock_item_code(
    category: str = Query("PRODUCT", description="Item category for code generation"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Preview the next auto-generated item code for a given category.
    DC_SFMS_001: Used by the frontend to pre-fill the read-only code field.
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    code = StockItemMasterService.generate_item_code(db, category)
    return JSONResponse(content={"success": True, "item_code": code})


@router.get("/stock-items/categories")
async def list_stock_item_categories(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Return all distinct categories currently in use, merged with the standard list.
    DC_SFMS_001: Allows dynamic category selection including custom values.
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    standard = ['PRODUCT', 'RAW_MATERIAL', 'CONSUMABLE', 'SPARE_PART', 'ACCESSORY', 'BATTERIES']
    from sqlalchemy import distinct as _distinct
    db_cats = [r[0] for r in db.query(_distinct(StockItemMaster.item_category)).filter(
        StockItemMaster.item_category.isnot(None)
    ).all()]
    merged = sorted(set(standard) | set(db_cats))
    return JSONResponse(content={"success": True, "categories": merged})


@router.get("/stock-items/{item_id}", response_model=StockItemMasterResponse)
async def get_stock_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single stock item by ID
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        item = StockItemMasterService.get_stock_item(db, item_id, current_user)
        response = StockItemMasterResponse.model_validate(item).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "stock_item": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-items/code/{item_code}", response_model=StockItemMasterResponse)
async def get_stock_item_by_code(
    item_code: str,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single stock item by code
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        item = StockItemMasterService.get_stock_item_by_code(db, item_code, current_user)
        response = StockItemMasterResponse.model_validate(item).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "stock_item": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/stock-items/{item_id}", response_model=StockItemMasterResponse)
async def update_stock_item(
    item_id: int,
    data: StockItemMasterUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an existing stock item
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        item = StockItemMasterService.update_stock_item(db, item_id, data, current_user)
        response = StockItemMasterResponse.model_validate(item).model_dump(mode='json')
        
        if item.applicable_companies:
            for company_id in item.applicable_companies:
                await sfms_cache.invalidate_stock_items(company_id)
        
        return JSONResponse(content={
            "success": True,
            "message": f"Stock item '{item.item_name}' updated successfully",
            "stock_item": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/stock-items/{item_id}")
async def delete_stock_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft delete (deactivate) a stock item
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        StockItemMasterService.delete_stock_item(db, item_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Stock item deactivated successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-items/{item_id}/opening-balance")
async def add_opening_balance(
    item_id: int,
    company_id: int = Query(..., description="Company ID for the opening balance"),
    quantity: float = Query(..., ge=0.01, description="Opening quantity"),
    total_value: float = Query(..., ge=0.01, description="Total value of opening stock"),
    balance_date: Optional[str] = Query(None, description="Opening balance date (YYYY-MM-DD), defaults to today"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add opening balance for a stock item in a specific company.
    Creates a OPENING_BALANCE transaction in stock ledger.
    One-time entry per item per company - duplicate check enforced.
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        from datetime import datetime, date as date_type
        from sqlalchemy import func
        
        item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Stock item not found")
        
        company = db.query(AssociatedCompany).filter(AssociatedCompany.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")
        
        existing_opening = db.query(StockLedger).filter(
            StockLedger.stock_item_id == item_id,
            StockLedger.company_id == company_id,
            StockLedger.transaction_type == 'OPENING_BALANCE'
        ).first()
        
        if existing_opening:
            raise HTTPException(
                status_code=400, 
                detail=f"Opening balance already exists for this item in {company.company_name}. Created on {existing_opening.transaction_date.strftime('%Y-%m-%d')}."
            )
        
        if balance_date:
            try:
                txn_date = datetime.strptime(balance_date, '%Y-%m-%d').date()
            except ValueError:
                txn_date = date_type.today()
        else:
            txn_date = date_type.today()
        
        unit_rate = round(total_value / quantity, 2) if quantity > 0 else 0
        
        opening_entry = StockLedger(
            stock_item_id=item_id,
            company_id=company_id,
            transaction_type='OPENING_BALANCE',
            transaction_date=txn_date,
            quantity_in=quantity,
            quantity_out=0,
            rate=unit_rate,
            total_value=round(total_value, 2),
            reference_number=f"OB-{item.item_code}-{company_id}-{txn_date.strftime('%Y%m%d')}",
            narration=f"Opening balance entry for {item.item_name}",
            created_by_id=current_user.id
        )
        
        db.add(opening_entry)
        db.commit()
        db.refresh(opening_entry)

        # DC_STOCK_MKTLINK_001: Propagate updated qty to marketplace if linked
        try:
            StockItemMasterService.propagate_stock_qty_to_marketplace(db, item_id)
        except Exception:
            pass
        
        return JSONResponse(content={
            "success": True,
            "message": f"Opening balance added successfully for {item.item_name} in {company.company_name}",
            "opening_balance": {
                "id": opening_entry.id,
                "stock_item_id": item_id,
                "company_id": company_id,
                "quantity": quantity,
                "unit_rate": unit_rate,
                "total_value": total_value,
                "balance_date": txn_date.isoformat()
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.get("/stock-items/{item_id}/opening-balance")
async def get_opening_balance(
    item_id: int,
    company_id: Optional[int] = Query(None, description="Filter by company"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get opening balance entries for a stock item.
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        query = db.query(StockLedger).filter(
            StockLedger.stock_item_id == item_id,
            StockLedger.transaction_type == 'OPENING_BALANCE'
        )
        
        if company_id:
            query = query.filter(StockLedger.company_id == company_id)
        
        entries = query.all()
        
        results = []
        for entry in entries:
            company = db.query(AssociatedCompany).filter(AssociatedCompany.id == entry.company_id).first()
            results.append({
                "id": entry.id,
                "company_id": entry.company_id,
                "company_name": company.company_name if company else "Unknown",
                "quantity": float(entry.quantity_in or 0),
                "unit_rate": float(entry.rate or 0),
                "total_value": float(entry.total_value or 0),
                "balance_date": entry.transaction_date.isoformat() if entry.transaction_date else None,
                "created_at": entry.created_at.isoformat() if entry.created_at else None
            })
        
        return JSONResponse(content={
            "success": True,
            "opening_balances": results
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-items/{item_id}/adjust")
async def adjust_stock_item(
    item_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-STOCK-ADJ-001: Manual stock adjustment using delta qty (+/-).
    Creates an ADJUSTMENT entry in the stock ledger.
    """
    from datetime import date as date_type
    from decimal import Decimal as _Dec
    from app.services.stock_service import append_stock_ledger

    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")

    company_id = payload.get('company_id')
    delta_qty  = payload.get('delta_qty')
    rate       = payload.get('rate')
    reason     = payload.get('reason')

    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if delta_qty is None or float(delta_qty) == 0:
        raise HTTPException(status_code=400, detail="delta_qty must be a non-zero number")

    delta_qty = float(delta_qty)

    item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found")

    company = db.query(AssociatedCompany).filter(AssociatedCompany.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    try:
        unit_rate = _Dec(str(rate)) if rate and float(rate) > 0 else _Dec('0')
        qty_in    = _Dec(str(abs(delta_qty))) if delta_qty > 0 else _Dec('0')
        qty_out   = _Dec(str(abs(delta_qty))) if delta_qty < 0 else _Dec('0')
        ref_num   = f"ADJ-{item.item_code}-{company_id}-{date_type.today().strftime('%Y%m%d')}-{current_user.id}"
        narration = reason or f"Manual stock adjustment by {current_user.full_name or current_user.username}"

        entry = append_stock_ledger(
            db,
            item_id=item_id,
            company_id=int(company_id),
            entry_type='ADJUSTMENT',
            reference_type='ADJUSTMENT',
            reference_id=current_user.id,
            reference_number=ref_num,
            quantity_in=qty_in,
            quantity_out=qty_out,
            unit_rate=unit_rate,
            narration=narration,
            txn_date=date_type.today(),
            updated_by_id=current_user.id,
        )
        db.commit()

        return JSONResponse(content={
            "success": True,
            "message": f"Stock adjusted by {'+' if delta_qty > 0 else ''}{delta_qty} for {item.item_name} in {company.company_name}",
            "entry_id": entry.id,
            "new_balance": float(entry.balance_qty),
            "reference_number": ref_num,
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.post("/stock-items/{item_id}/marketplace-link")
async def link_stock_item_to_marketplace(
    item_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_STOCK_MKTLINK_001: Link a stock item to a marketplace SKU.
    Sets marketplace_sku on the stock item and immediately propagates
    selling_rate → dealer_price and net stock qty → available_qty.
    Locks those fields from Google Sheet sync.
    """
    try:
        marketplace_sku = (data.get('marketplace_sku') or '').strip()
        if not marketplace_sku:
            return JSONResponse(status_code=400, content={"success": False, "message": "marketplace_sku is required"})

        item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
        if not item:
            return JSONResponse(status_code=404, content={"success": False, "message": "Stock item not found"})

        # Verify the marketplace SKU exists
        from app.models.marketplace import MarketspareItem
        mkt = db.query(MarketspareItem).filter(MarketspareItem.sku == marketplace_sku).first()
        if not mkt:
            return JSONResponse(status_code=404, content={
                "success": False,
                "message": f"Marketplace SKU '{marketplace_sku}' not found. Verify the SKU and try again."
            })

        item.marketplace_sku = marketplace_sku
        db.commit()
        db.refresh(item)

        # Immediately propagate price + qty
        propagated = StockItemMasterService._propagate_stock_to_marketplace(db, item)

        return JSONResponse(content={
            "success": True,
            "message": f"Linked to marketplace SKU '{marketplace_sku}' — price and qty propagated.",
            "marketplace_sku": marketplace_sku,
            "propagated": propagated,
            "dealer_price": float(item.selling_rate or 0),
            "marketplace_name": mkt.name
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.delete("/stock-items/{item_id}/marketplace-link")
async def unlink_stock_item_from_marketplace(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_STOCK_MKTLINK_001: Remove the marketplace link from a stock item.
    Clears marketplace_sku, removes dealer_price + available_qty from override_fields.
    The marketplace item will resume syncing from Google Sheet on next sync.
    """
    try:
        item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
        if not item:
            return JSONResponse(status_code=404, content={"success": False, "message": "Stock item not found"})

        old_sku = item.marketplace_sku
        if not old_sku:
            return JSONResponse(content={"success": True, "message": "No marketplace link to remove."})

        # Remove override lock from the marketplace item
        from app.models.marketplace import MarketspareItem
        mkt = db.query(MarketspareItem).filter(MarketspareItem.sku == old_sku).first()
        if mkt:
            protected = set(mkt.override_fields or [])
            protected.discard('dealer_price')
            protected.discard('available_qty')
            mkt.override_fields = list(protected)
            mkt.manually_overridden = len(protected) > 0

        item.marketplace_sku = None
        db.commit()

        return JSONResponse(content={
            "success": True,
            "message": f"Marketplace link to '{old_sku}' removed. Sheet sync will resume on next run."
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.get("/stock-items/{item_id}/vendors")
async def get_vendors_for_stock_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get all vendors that supply a specific stock item
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendors = VendorMasterService.get_vendors_for_product(db, item_id, current_user)
        return JSONResponse(content={
            "success": True,
            "vendors": vendors
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== STOCK ITEM IMAGE ENDPOINTS ====================

@router.post("/stock-items/{item_id}/images")
async def upload_stock_item_image(
    item_id: int,
    file: UploadFile = File(...),
    is_primary: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Upload an image for a stock item
    DC_STOCK_004: Multiple images per stock item with compression to <100KB
    Supports 5MB upload limit, compressed for frontend display
    """
    from app.services.universal_upload_service import UniversalUploadService
    
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        print(f"[DC-RBAC] Stock Image Upload - Access Denied: User role '{user_role}' not in allowed roles {ACCOUNTS_ALLOWED_ROLES}")
        raise HTTPException(status_code=403, detail=f"Access denied. Your role '{user_role}' is not authorized. Only VGK4U, EA, and Accounts roles can access this feature.")
    
    stock_item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
    if not stock_item:
        raise HTTPException(status_code=404, detail=f"Stock item with ID {item_id} not found")
    
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 5MB limit")
    
    allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"File type {file.content_type} not allowed. Use JPEG, PNG, GIF, or WebP")
    
    try:
        upload_service = UniversalUploadService(db, current_user)
        result = upload_service.upload_file(
            file_data=content,
            original_filename=file.filename,
            content_type=file.content_type,
            entity_id=item_id,
            segment="STOCK",
            description=f"Stock item image for {stock_item.item_code}"
        )
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error', 'Upload failed'))
        
        if is_primary:
            db.query(StockItemImage).filter(
                StockItemImage.stock_item_id == item_id,
                StockItemImage.is_primary == True
            ).update({"is_primary": False})
        
        max_order = db.query(StockItemImage).filter(
            StockItemImage.stock_item_id == item_id
        ).count()
        
        stock_image = StockItemImage(
            stock_item_id=item_id,
            original_path=result.get('file_path', ''),
            compressed_path=result.get('compressed_path'),
            thumbnail_path=result.get('thumbnail_path'),
            file_name=file.filename,
            file_size=len(content),
            compressed_size=result.get('compressed_size'),
            mime_type=file.content_type,
            is_primary=is_primary or max_order == 0,
            display_order=max_order,
            source_type="upload",
            uploaded_by_id=current_user.id
        )
        db.add(stock_image)
        db.commit()
        db.refresh(stock_image)
        
        return JSONResponse(content={
            "success": True,
            "message": "Image uploaded successfully",
            "image": StockItemImageResponse.model_validate(stock_image).model_dump(mode='json')
        })
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stock-items/{item_id}/images")
async def list_stock_item_images(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get all images for a stock item
    DC_STOCK_004: Returns images ordered by display_order with primary first
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    stock_item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
    if not stock_item:
        raise HTTPException(status_code=404, detail=f"Stock item with ID {item_id} not found")
    
    images = db.query(StockItemImage).filter(
        StockItemImage.stock_item_id == item_id
    ).order_by(StockItemImage.is_primary.desc(), StockItemImage.display_order).all()
    
    return JSONResponse(content={
        "success": True,
        "stock_item_id": item_id,
        "images": [StockItemImageResponse.model_validate(img).model_dump(mode='json') for img in images],
        "total": len(images)
    })


@router.delete("/stock-items/{item_id}/images/{image_id}")
async def delete_stock_item_image(
    item_id: int,
    image_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Delete a stock item image
    DC_STOCK_004: Removes image record (files remain for audit trail)
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    image = db.query(StockItemImage).filter(
        StockItemImage.id == image_id,
        StockItemImage.stock_item_id == item_id
    ).first()
    
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")
    
    was_primary = image.is_primary
    db.delete(image)
    db.commit()
    
    if was_primary:
        next_image = db.query(StockItemImage).filter(
            StockItemImage.stock_item_id == item_id
        ).order_by(StockItemImage.display_order).first()
        if next_image:
            next_image.is_primary = True
            db.commit()
    
    return JSONResponse(content={
        "success": True,
        "message": "Image deleted successfully"
    })


@router.put("/stock-items/{item_id}/images/{image_id}/primary")
async def set_primary_stock_item_image(
    item_id: int,
    image_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Set an image as the primary image for a stock item
    DC_STOCK_004: Only one primary image per stock item
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    image = db.query(StockItemImage).filter(
        StockItemImage.id == image_id,
        StockItemImage.stock_item_id == item_id
    ).first()
    
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")
    
    db.query(StockItemImage).filter(
        StockItemImage.stock_item_id == item_id,
        StockItemImage.is_primary == True
    ).update({"is_primary": False})
    
    image.is_primary = True
    db.commit()
    
    return JSONResponse(content={
        "success": True,
        "message": "Primary image set successfully"
    })


@router.post("/stock-items/{item_id}/images/from-url")
async def add_stock_item_image_from_url(
    item_id: int,
    image_url: str = Form(...),
    is_primary: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add a stock item image from URL (for bulk upload support)
    DC_STOCK_004: Downloads image from URL, compresses, and stores with dual evidence
    """
    import requests
    from app.services.universal_upload_service import UniversalUploadService
    
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    stock_item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
    if not stock_item:
        raise HTTPException(status_code=404, detail=f"Stock item with ID {item_id} not found")
    
    try:
        response = requests.get(image_url, timeout=30, stream=True)
        response.raise_for_status()
        
        content = response.content
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Image from URL exceeds 5MB limit")
        
        content_type = response.headers.get('content-type', 'image/jpeg')
        if not content_type.startswith('image/'):
            content_type = 'image/jpeg'
        
        filename = image_url.split('/')[-1].split('?')[0] or f"image_{item_id}.jpg"
        
        upload_service = UniversalUploadService(db, current_user)
        result = upload_service.upload_file(
            file_data=content,
            original_filename=filename,
            content_type=content_type,
            entity_id=item_id,
            segment="STOCK",
            description=f"Stock item image from URL for {stock_item.item_code}"
        )
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error', 'Upload failed'))
        
        if is_primary:
            db.query(StockItemImage).filter(
                StockItemImage.stock_item_id == item_id,
                StockItemImage.is_primary == True
            ).update({"is_primary": False})
        
        max_order = db.query(StockItemImage).filter(
            StockItemImage.stock_item_id == item_id
        ).count()
        
        stock_image = StockItemImage(
            stock_item_id=item_id,
            original_path=result.get('file_path', ''),
            compressed_path=result.get('compressed_path'),
            thumbnail_path=result.get('thumbnail_path'),
            file_name=filename,
            file_size=len(content),
            compressed_size=result.get('compressed_size'),
            mime_type=content_type,
            is_primary=is_primary or max_order == 0,
            display_order=max_order,
            source_type="url",
            source_url=image_url,
            uploaded_by_id=current_user.id
        )
        db.add(stock_image)
        db.commit()
        db.refresh(stock_image)
        
        return JSONResponse(content={
            "success": True,
            "message": "Image added from URL successfully",
            "image": StockItemImageResponse.model_validate(stock_image).model_dump(mode='json')
        })
    except requests.RequestException as e:
        raise HTTPException(status_code=400, detail=f"Failed to fetch image from URL: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/stock-items/{item_id}/image-link")
async def upsert_stock_item_image_link(
    item_id: int,
    source_url: str = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add or replace the Google Drive folder link for a stock item.
    DC_STOCK_004: Upsert — any existing folder_link record is replaced.
    Passing an empty source_url removes the folder link entirely.
    Only Google Drive links are accepted.
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")

    stock_item = db.query(StockItemMaster).filter(StockItemMaster.id == item_id).first()
    if not stock_item:
        raise HTTPException(status_code=404, detail=f"Stock item {item_id} not found")

    # Remove any existing folder_link record for this item (upsert semantics)
    db.query(StockItemImage).filter(
        StockItemImage.stock_item_id == item_id,
        StockItemImage.source_type == 'folder_link'
    ).delete(synchronize_session=False)
    db.flush()

    url = (source_url or '').strip()
    if not url:
        db.commit()
        return JSONResponse(content={"success": True, "message": "Image link removed", "image": None})

    if 'drive.google.com' not in url:
        db.rollback()
        raise HTTPException(status_code=400, detail="Only Google Drive links are accepted")

    # Determine display order and primary flag after the delete
    remaining_count = db.query(StockItemImage).filter(
        StockItemImage.stock_item_id == item_id
    ).count()

    link_image = StockItemImage(
        stock_item_id=item_id,
        original_path=url,
        file_name=f"gdrive_link_{item_id}",
        is_primary=(remaining_count == 0),
        display_order=remaining_count,
        source_type='folder_link',
        source_url=url,
        uploaded_by_id=current_user.id
    )
    db.add(link_image)
    db.commit()
    db.refresh(link_image)

    return JSONResponse(content={
        "success": True,
        "message": "Image link saved",
        "image": StockItemImageResponse.model_validate(link_image).model_dump(mode='json')
    })


# ==================== INCOME ENTRY ENDPOINTS ====================

@router.post("/income-entries", response_model=IncomeEntryResponse)
async def create_income_entry(
    data: IncomeEntryCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new income entry
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        entry = IncomeEntryService.create_income_entry(db, data, current_user)
        response = IncomeEntryResponse.model_validate(entry).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Income entry '{entry.entry_number}' created successfully",
            "income_entry": response
        })
    except ValueError as ve:
        # DC-IE-DEDUP-001: Service raises ValueError("DUPLICATE_ENTRY:entry_num|user_msg")
        msg = str(ve)
        if msg.startswith('DUPLICATE_ENTRY:'):
            parts = msg.split('|', 1)
            dup_ref = parts[0].replace('DUPLICATE_ENTRY:', '').strip()
            user_msg = parts[1].strip() if len(parts) > 1 else msg
            return JSONResponse(
                status_code=409,
                content={
                    "success": False,
                    "message": user_msg,
                    "error_code": "DUPLICATE",
                    "duplicate_entry": dup_ref
                }
            )
        return handle_accounts_error(ve)
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries", response_model=IncomeEntryListResponse)
async def list_income_entries(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=500, description="Items per page"),
    limit: Optional[int] = Query(None, ge=1, le=500, description="Alias for page_size"),
    company_id: Optional[int] = Query(None, description="Filter by company"),
    income_source_id: Optional[int] = Query(None, description="Filter by income source type"),
    status: Optional[str] = Query(None, description="Filter by status: PENDING, CONFIRMED, EXCEPTION_TALLY, ADJUSTMENT, TALLY_DONE"),
    payment_mode: Optional[str] = Query(None, description="Filter by payment mode"),
    transaction_type: Optional[str] = Query(None, description="Filter by transaction type"),
    date_from: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    search: Optional[str] = Query(None, description="Search by entry number, payer name, reference"),
    sort_by: Optional[str] = Query(None, description="Sort field"),
    sort_order: Optional[str] = Query("desc", description="Sort order: asc/desc"),
    source: Optional[str] = Query(None, description="Filter by source: crm/manual"),
    city: Optional[str] = Query(None, description="Filter by city"),
    state: Optional[str] = Query(None, description="Filter by state"),
    collected_by: Optional[int] = Query(None, description="Filter by collected_by employee ID"),
    lead_owner: Optional[int] = Query(None, description="Filter by lead owner employee ID"),
    revenue_category_id: Optional[int] = Query(None, description="Filter by revenue category ID"),
    category_name: Optional[str] = Query(None, description="DC-IE-CAT-NAME-001: Filter by category name across all companies"),
    destination_employee_id: Optional[int] = Query(None, description="DC-IE-DEST-FILTER-001: Filter by destination employee (entries where cash was routed to this staff member)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List income entries with filtering, sorting, and enriched data
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        from datetime import datetime as dt
        from app.models.staff import StaffEmployee as SE
        
        effective_page_size = limit or page_size
        parsed_date_from = dt.strptime(date_from, "%Y-%m-%d").date() if date_from else None
        parsed_date_to = dt.strptime(date_to, "%Y-%m-%d").date() if date_to else None
        
        entries, total = IncomeEntryService.list_income_entries(
            db, current_user, page, effective_page_size, company_id, income_source_id,
            status, payment_mode, parsed_date_from, parsed_date_to, search,
            transaction_type=transaction_type, source=source, city=city, state=state,
            collected_by_id=collected_by, lead_owner_id=lead_owner,
            sort_by=sort_by, sort_order=sort_order or 'desc',
            revenue_category_id=revenue_category_id,
            destination_employee_id=destination_employee_id,
            category_name=category_name,
        )
        
        cat_ids = list(set(e.revenue_category_id for e in entries if e.revenue_category_id))
        cat_map = {}
        if cat_ids:
            from app.models.signup_category import SignupCategory
            cats = db.query(SignupCategory).filter(SignupCategory.id.in_(cat_ids)).all()
            cat_map = {c.id: c.name for c in cats}

        comp_ids = list(set(
            cid for e in entries
            for cid in [e.company_id, getattr(e, 'destination_company_id', None)]
            if cid
        ))
        comp_map = {}
        comp_code_map = {}
        if comp_ids:
            comps = db.query(AssociatedCompany).filter(AssociatedCompany.id.in_(comp_ids)).all()
            comp_map = {c.id: (c.company_name or c.company_code) for c in comps}
            comp_code_map = {c.id: c.company_code for c in comps}

        from app.models.staff_accounts import IncomeSourceType as _IST
        src_ids = list(set(e.income_source_id for e in entries if e.income_source_id))
        src_map = {}
        if src_ids:
            srcs = db.query(_IST).filter(_IST.id.in_(src_ids)).all()
            src_map = {s.id: s.source_name for s in srcs}

        staff_ids = set()
        for e in entries:
            if e.lead_owner_id: staff_ids.add(e.lead_owner_id)
            if e.collected_by_id: staff_ids.add(e.collected_by_id)
            if e.confirmed_by_id: staff_ids.add(e.confirmed_by_id)
            if getattr(e, 'updated_by_id', None): staff_ids.add(e.updated_by_id)
            if getattr(e, 'destination_employee_id', None): staff_ids.add(e.destination_employee_id)
        staff_map = {}
        if staff_ids:
            staffs = db.query(SE).filter(SE.id.in_(list(staff_ids))).all()
            staff_map = {s.id: {"name": s.full_name, "emp_code": s.emp_code, "designation": getattr(s, 'designation', '')} for s in staffs}

        lead_ids = list(set(e.lead_id for e in entries if e.lead_id))
        lead_map = {}
        if lead_ids:
            from app.models.crm import CRMLead
            leads = db.query(CRMLead).filter(CRMLead.id.in_(lead_ids)).all()
            lead_map = {l.id: {"name": l.name, "city": l.city, "state": l.state, "phone": l.phone} for l in leads}

        txn_ids = list(set(e.crm_transaction_id for e in entries if e.crm_transaction_id))
        deal_code_map = {}
        if txn_ids:
            from app.models.crm import CRMLeadTransaction, CRMLeadDeal
            txns = db.query(CRMLeadTransaction).filter(CRMLeadTransaction.id.in_(txn_ids)).all()
            deal_ids = list(set(t.deal_id for t in txns if t.deal_id))
            txn_deal_map = {t.id: t.deal_id for t in txns}
            if deal_ids:
                deals = db.query(CRMLeadDeal).filter(CRMLeadDeal.id.in_(deal_ids)).all()
                deal_info = {d.id: d.deal_code for d in deals}
                deal_code_map = {tid: deal_info.get(did) for tid, did in txn_deal_map.items() if did}

        entry_list = []
        for e in entries:
            d = IncomeEntryResponse.model_validate(e).model_dump(mode='json')
            d['revenue_category_name'] = cat_map.get(e.revenue_category_id)
            d['company_name'] = comp_map.get(e.company_id)
            d['deal_code'] = deal_code_map.get(e.crm_transaction_id) if e.crm_transaction_id else None
            d['lead_name'] = lead_map.get(e.lead_id, {}).get('name') if e.lead_id else None
            d['lead_city'] = e.payer_city or (lead_map.get(e.lead_id, {}).get('city') if e.lead_id else None)
            d['lead_state'] = e.payer_state or (lead_map.get(e.lead_id, {}).get('state') if e.lead_id else None)
            d['lead_phone'] = lead_map.get(e.lead_id, {}).get('phone') if e.lead_id else None
            d['lead_owner_name'] = staff_map.get(e.lead_owner_id, {}).get('name') if e.lead_owner_id else None
            d['lead_owner_emp_code'] = staff_map.get(e.lead_owner_id, {}).get('emp_code') if e.lead_owner_id else None
            d['collected_by_name'] = staff_map.get(e.collected_by_id, {}).get('name') if e.collected_by_id else None
            d['collected_by_emp_code'] = staff_map.get(e.collected_by_id, {}).get('emp_code') if e.collected_by_id else None
            d['confirmed_by_name'] = staff_map.get(e.confirmed_by_id, {}).get('name') if e.confirmed_by_id else None
            _upd_id = getattr(e, 'updated_by_id', None)
            d['updated_by_name'] = staff_map.get(_upd_id, {}).get('name') if _upd_id else None
            d['updated_by_emp_code'] = staff_map.get(_upd_id, {}).get('emp_code') if _upd_id else None
            d['income_source_name'] = src_map.get(e.income_source_id) if e.income_source_id else None
            d['destination_employee_name'] = staff_map.get(e.destination_employee_id, {}).get('name') if getattr(e, 'destination_employee_id', None) else None
            d['destination_employee_emp_code'] = staff_map.get(e.destination_employee_id, {}).get('emp_code') if getattr(e, 'destination_employee_id', None) else None
            d['destination_company_name'] = comp_map.get(e.destination_company_id) if getattr(e, 'destination_company_id', None) else None
            entry_list.append(d)
        
        return JSONResponse(content={
            "success": True,
            "income_entries": entry_list,
            "incomes": entry_list,
            "total": total,
            "page": page,
            "page_size": effective_page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/estimations/executive-summary")
async def get_estimation_executive_summary(
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: Get executive summary of IN/OUT estimates and payments."""
    try:
        from app.services.staff_accounts_service import EstimationService
        summary = EstimationService.executive_summary(db, current_user, company_id=company_id)
        return JSONResponse(content={"success": True, "summary": summary})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/estimations/stock")
async def list_estimate_stock_movements(
    company_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 200,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: List stock_ledger entries with is_estimate=True."""
    try:
        from app.services.staff_accounts_service import EstimationService
        rows = EstimationService.list_estimate_stock(db, current_user, company_id=company_id, page=page, page_size=page_size)
        return JSONResponse(content={"success": True, "stock_movements": rows})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/estimations/out")
async def list_estimation_out_records(
    company_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 100,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: List OUT estimation planning records."""
    try:
        from app.services.staff_accounts_service import EstimationService
        rows, total = EstimationService.list_out(db, current_user, company_id=company_id, page=page, page_size=page_size)
        return JSONResponse(content={"success": True, "records": rows, "total": total})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/income-entries/estimations/out")
async def create_estimation_out_record(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: Create an OUT estimation planning record."""
    try:
        from app.services.staff_accounts_service import EstimationService
        rec = EstimationService.create_out(db, current_user, payload)
        return JSONResponse(content={"success": True, "record": rec})
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/income-entries/estimations/out/{record_id}")
async def update_estimation_out_record(
    record_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: Update an OUT estimation planning record."""
    try:
        from app.services.staff_accounts_service import EstimationService
        rec = EstimationService.update_out(db, current_user, record_id, payload)
        return JSONResponse(content={"success": True, "record": rec})
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/income-entries/estimations/out/{record_id}")
async def delete_estimation_out_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: Soft-delete an OUT estimation planning record."""
    try:
        from app.services.staff_accounts_service import EstimationService
        EstimationService.delete_out(db, current_user, record_id)
        return JSONResponse(content={"success": True, "message": "Record deleted"})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/estimations")
async def list_estimated_income_entries(
    company_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 200,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: List income entries with status=ESTIMATED."""
    try:
        from app.services.staff_accounts_service import EstimationService
        from app.models.staff_accounts import AssociatedCompany as _AC, IncomeSourceType as _IST
        entries, total = EstimationService.list_estimated_entries(
            db, current_user, company_id=company_id, page=page, page_size=page_size
        )
        comp_map = {c.id: c.company_name for c in db.query(_AC).all()}
        src_map = {s.id: s.source_name for s in db.query(_IST).filter(_IST.is_active == True).all()}
        result = []
        for e in entries:
            d = IncomeEntryResponse.model_validate(e).model_dump(mode='json')
            d['company_name'] = comp_map.get(e.company_id)
            d['income_source_name'] = src_map.get(e.income_source_id) if e.income_source_id else None
            result.append(d)
        return JSONResponse(content={"success": True, "entries": result, "total": total})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/{entry_id}/estimation-payments")
async def list_estimation_payments(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: List payments recorded against an estimated income entry."""
    try:
        from app.services.staff_accounts_service import EstimationService
        payments = EstimationService.list_payments(db, current_user, entry_id)
        return JSONResponse(content={"success": True, "payments": payments})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/income-entries/{entry_id}/estimation-payments")
async def add_estimation_payment(
    entry_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: Add a payment record against an estimated income entry."""
    try:
        from app.services.staff_accounts_service import EstimationService
        pmt = EstimationService.add_payment(db, current_user, entry_id, payload)
        return JSONResponse(content={"success": True, "payment": pmt})
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/income-entries/{entry_id}/estimation-payments/{payment_id}")
async def delete_estimation_payment(
    entry_id: int,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: Delete a payment record from an estimated income entry."""
    try:
        from app.services.staff_accounts_service import EstimationService
        EstimationService.delete_payment(db, current_user, payment_id)
        return JSONResponse(content={"success": True, "message": "Payment deleted"})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/ledger-masters/bank-accounts")
async def list_bank_ledger_accounts(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-ESTIMATIONS-001: List BANK/UPI/CASH/CHEQUE ledger master accounts for income confirmation dropdown."""
    try:
        from app.models.staff_accounts import AccountLedgerMaster as _ALM
        rows = db.query(_ALM).filter(
            _ALM.company_id == company_id,
            _ALM.account_type.in_(['BANK', 'UPI', 'CASH', 'CHEQUE']),
            _ALM.is_active == True
        ).order_by(_ALM.account_type, _ALM.account_name).all()
        result = [{'id': r.id, 'account_name': r.account_name, 'account_type': r.account_type} for r in rows]
        return JSONResponse(content={"success": True, "accounts": result})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/{entry_id}", response_model=IncomeEntryResponse)
async def get_income_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single income entry by ID
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        entry = IncomeEntryService.get_income_entry(db, entry_id, current_user)
        response = IncomeEntryResponse.model_validate(entry).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "income_entry": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/number/{entry_number}", response_model=IncomeEntryResponse)
async def get_income_entry_by_number(
    entry_number: str,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single income entry by entry number
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        entry = IncomeEntryService.get_income_entry_by_number(db, entry_number, current_user)
        response = IncomeEntryResponse.model_validate(entry).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "income_entry": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/income-entries/{entry_id}", response_model=IncomeEntryResponse)
async def update_income_entry(
    entry_id: int,
    data: IncomeEntryUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an existing income entry (only PENDING entries can be modified)
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        entry = IncomeEntryService.update_income_entry(db, entry_id, data, current_user)
        response = IncomeEntryResponse.model_validate(entry).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Income entry '{entry.entry_number}' updated successfully",
            "income_entry": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/income-entries/{entry_id}/status", response_model=IncomeEntryResponse)
async def update_income_entry_status(
    entry_id: int,
    data: IncomeEntryStatusUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update income entry status (verify/approve/reject/cancel)
    DC_SFMS_001: VGK/EA/Accounts access only
    Valid transitions: PENDING -> VERIFIED/REJECTED/CANCELLED,
                      VERIFIED -> APPROVED/REJECTED/CANCELLED,
                      APPROVED -> CANCELLED, REJECTED -> PENDING
    """
    try:
        entry = IncomeEntryService.update_income_entry_status(db, entry_id, data, current_user)
        response = IncomeEntryResponse.model_validate(entry).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Income entry '{entry.entry_number}' status updated to '{entry.status}'",
            "income_entry": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/income-entries/{entry_id}/show-in-ledger", response_model=IncomeEntryResponse)
async def toggle_income_show_in_ledger_endpoint(
    entry_id: int,
    show_in_ledger: bool = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-SHOW-IN-LEDGER-001: toggle whether this income entry posts to the transaction ledger.
    Optional flag, editable at any status.
    """
    try:
        entry = IncomeEntryService.toggle_show_in_ledger(db, entry_id, show_in_ledger, current_user)
        response = IncomeEntryResponse.model_validate(entry).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Income entry '{entry.entry_number}' show_in_ledger set to {entry.show_in_ledger}",
            "income_entry": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/solar-vendor-ledger")
async def list_solar_vendor_ledger(
    vendor_id: Optional[int] = Query(None, description="Filter by solar vendor"),
    direction: Optional[str] = Query(None, description="RECEIVED or RETURNED"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-SOLAR-VENDOR-LEDGER-001: List solar vendor ledger entries (staff)."""
    try:
        from sqlalchemy import text as _t
        from app.models.base import get_indian_time
        q = db.query(SolarVendorLedger)
        if vendor_id:
            q = q.filter(SolarVendorLedger.solar_vendor_id == vendor_id)
        if direction:
            q = q.filter(SolarVendorLedger.direction == direction.upper())
        if date_from:
            q = q.filter(SolarVendorLedger.transaction_date >= date_from)
        if date_to:
            q = q.filter(SolarVendorLedger.transaction_date <= date_to)
        total = q.count()
        rows = q.order_by(SolarVendorLedger.transaction_date.desc(), SolarVendorLedger.id.desc()) \
                .offset((page - 1) * page_size).limit(page_size).all()
        # Enrich with vendor names
        vendor_ids = list(set(r.solar_vendor_id for r in rows))
        vendors = {}
        if vendor_ids:
            for v in db.query(VendorMaster).filter(VendorMaster.id.in_(vendor_ids)).all():
                vendors[v.id] = v.vendor_name
        result = []
        for r in rows:
            result.append({
                "id": r.id,
                "solar_vendor_id": r.solar_vendor_id,
                "vendor_name": vendors.get(r.solar_vendor_id, ""),
                "income_entry_id": r.income_entry_id,
                "entry_number": r.entry_number,
                "transaction_date": r.transaction_date.isoformat() if r.transaction_date else None,
                "customer_name": r.customer_name,
                "amount": float(r.amount),
                "company_id": r.company_id,
                "direction": r.direction,
                "utr_reference": r.utr_reference,
                "payment_mode": r.payment_mode,
                "notes": r.notes,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            })
        # Compute summary across full (unfiltered-by-page) set
        all_q = db.query(SolarVendorLedger)
        if vendor_id:
            all_q = all_q.filter(SolarVendorLedger.solar_vendor_id == vendor_id)
        if direction:
            all_q = all_q.filter(SolarVendorLedger.direction == direction.upper())
        if date_from:
            all_q = all_q.filter(SolarVendorLedger.transaction_date >= date_from)
        if date_to:
            all_q = all_q.filter(SolarVendorLedger.transaction_date <= date_to)
        all_rows = all_q.all()
        total_received = sum(float(r.amount) for r in all_rows if r.direction == 'RECEIVED')
        total_returned = sum(float(r.amount) for r in all_rows if r.direction == 'RETURNED')
        summary = {
            "total_received": round(total_received, 2),
            "total_returned": round(total_returned, 2),
            "balance": round(total_received - total_returned, 2),
            "count_received": sum(1 for r in all_rows if r.direction == 'RECEIVED'),
            "count_returned": sum(1 for r in all_rows if r.direction == 'RETURNED'),
        }
        return JSONResponse(content={"success": True, "rows": result, "summary": summary, "total": total, "page": page, "page_size": page_size})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/solar-vendor-ledger/return")
async def record_solar_vendor_return(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-SOLAR-VENDOR-LEDGER-001: Record a payment returned by solar vendor to MNR."""
    try:
        from app.models.base import get_indian_time
        from decimal import Decimal
        vendor_id = payload.get("solar_vendor_id")
        if not vendor_id:
            return JSONResponse(status_code=400, content={"success": False, "detail": "solar_vendor_id required"})
        amount = Decimal(str(payload.get("amount", 0)))
        if amount <= 0:
            return JSONResponse(status_code=400, content={"success": False, "detail": "amount must be > 0"})
        txn_date_str = payload.get("transaction_date")
        if not txn_date_str:
            return JSONResponse(status_code=400, content={"success": False, "detail": "transaction_date required"})
        from datetime import date as _date
        txn_date = _date.fromisoformat(txn_date_str)
        row = SolarVendorLedger(
            solar_vendor_id=int(vendor_id),
            income_entry_id=None,
            entry_number=None,
            transaction_date=txn_date,
            customer_name=payload.get("customer_name"),
            amount=amount,
            company_id=payload.get("company_id"),
            direction='RETURNED',
            utr_reference=payload.get("utr_reference"),
            payment_mode=payload.get("payment_mode"),
            notes=payload.get("notes"),
            created_by_id=current_user.id,
            created_at=get_indian_time(),
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return JSONResponse(content={"success": True, "message": "Vendor return recorded", "id": row.id})
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/solar-vendor-ledger/{entry_id}")
async def update_solar_vendor_ledger_entry(
    entry_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-SOLAR-VENDOR-LEDGER-001: Edit date/customer/mode/UTR/notes on any SVL entry."""
    try:
        from datetime import date as _date
        row = db.query(SolarVendorLedger).filter(SolarVendorLedger.id == entry_id).first()
        if not row:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Entry not found"})
        if 'transaction_date' in payload and payload['transaction_date']:
            row.transaction_date = _date.fromisoformat(payload['transaction_date'])
        if 'customer_name' in payload:
            row.customer_name = payload['customer_name'] or None
        if 'payment_mode' in payload:
            row.payment_mode = payload['payment_mode'] or None
        if 'utr_reference' in payload:
            row.utr_reference = payload['utr_reference'] or None
        if 'notes' in payload:
            row.notes = payload['notes'] or None
        db.commit()
        return JSONResponse(content={"success": True, "message": "Entry updated"})
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/solar-vendor-ledger/{entry_id}")
async def delete_solar_vendor_ledger_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-SOLAR-VENDOR-LEDGER-001: Delete a RETURNED solar vendor ledger entry (staff only)."""
    try:
        row = db.query(SolarVendorLedger).filter(SolarVendorLedger.id == entry_id).first()
        if not row:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Entry not found"})
        if row.direction != 'RETURNED':
            return JSONResponse(status_code=400, content={"success": False, "detail": "Only RETURNED entries can be deleted"})
        db.delete(row)
        db.commit()
        return JSONResponse(content={"success": True, "message": "Entry deleted"})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/income-entries/deleted")
async def list_deleted_income_entries(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=500, description="Items per page"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List deleted income entries for audit trail."""
    try:
        entries, total = IncomeEntryService.list_deleted_income_entries(db, current_user, page, page_size)
        return JSONResponse(content={
            "success": True,
            "income_entries": entries,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/income-entries/{entry_id}")
async def delete_income_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft-delete an income entry (sets is_deleted=True, preserves audit trail).
    DC_INCOME_DELETE_001: Restricted to VGK Mentor (vgk4u) and EA roles only.
    """
    try:
        IncomeEntryService.delete_income_entry(db, entry_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Income entry deleted successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/income-entries/{entry_id}/reject")
async def reject_income_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Reject an income entry — excluded from all income calculations and staff performance KPIs.
    Also marks the linked CRM transaction as rejected.
    DC_SFMS_REJECT: Accounts role (vgk4u/ea/accounts) OR MENTOR designation
    """
    try:
        designation = (current_user.designation or "").upper()
        is_allowed = is_accounts_allowed_employee(current_user) or "MENTOR" in designation
        if not is_allowed:
            return JSONResponse(status_code=403, content={
                "success": False,
                "message": "Access denied. Only Accounts, EA, or MENTOR roles can reject income entries."
            })

        entry = db.query(IncomeEntry).filter(IncomeEntry.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "message": "Income entry not found"})

        if entry.status == "REJECTED":
            return JSONResponse(status_code=400, content={
                "success": False,
                "message": f"Entry {entry.entry_number} is already rejected."
            })

        from datetime import datetime as _dt
        from app.models.crm import CRMLeadTransaction as _CRMLeadTxn
        import pytz as _tz
        now_ist = _dt.now(_tz.timezone("Asia/Kolkata"))

        old_status = entry.status
        entry.status = "REJECTED"
        entry.rejected_by_id = current_user.id
        entry.rejected_at = now_ist
        entry.updated_at = now_ist

        crm_txn_count = 0
        if entry.crm_transaction_id:
            crm_txn = db.query(_CRMLeadTxn).filter(
                _CRMLeadTxn.id == entry.crm_transaction_id
            ).first()
            if crm_txn and crm_txn.validation_status != "rejected":
                crm_txn.validation_status = "rejected"
                crm_txn.rejection_reason = f"Income entry {entry.entry_number} rejected by {current_user.emp_code}"
                crm_txn_count = 1

        db.commit()
        db.refresh(entry)

        log_accounts_audit(
            db=db,
            employee_id=current_user.id,
            action="REJECT_INCOME_ENTRY",
            entity_type="IncomeEntry",
            entity_id=entry.id,
            old_values={"status": old_status},
            new_values={"status": "REJECTED"},
            description=f"Income entry {entry.entry_number} rejected by {current_user.emp_code}"
        )

        return JSONResponse(content={
            "success": True,
            "message": f"Income entry {entry.entry_number} rejected and excluded from all calculations.",
            "entry_number": entry.entry_number,
            "crm_transactions_updated": crm_txn_count
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


# ==================== EXPENSE MAIN CATEGORY ENDPOINTS (SFMS Staff Access) ====================

@router.post("/expense-categories/main", response_model=ExpenseMainCategorySimpleResponse)
async def create_expense_main_category(
    data: ExpenseMainCategoryCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new expense main category
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        category = ExpenseMainCategoryService.create_main_category(
            db=db,
            name=data.name,
            employee=current_user,
            description=data.description
        )
        response = ExpenseMainCategorySimpleResponse.model_validate(category).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Expense main category '{category.name}' created successfully",
            "category": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/expense-categories/main", response_model=ExpenseMainCategoryListResponse)
async def list_expense_main_categories(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    include_inactive: bool = Query(False, description="Include inactive categories"),
    include_sub_categories: bool = Query(True, description="Include nested sub-categories"),
    search: Optional[str] = Query(None, description="Search by name"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List expense main categories with optional sub-categories
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        categories, total = ExpenseMainCategoryService.list_main_categories(
            db, current_user, page, page_size, include_inactive, include_sub_categories, search
        )
        
        if include_sub_categories:
            cat_list = [
                ExpenseMainCategoryResponse.model_validate(c).model_dump(mode='json')
                for c in categories
            ]
        else:
            cat_list = [
                ExpenseMainCategorySimpleResponse.model_validate(c).model_dump(mode='json')
                for c in categories
            ]
        
        return JSONResponse(content={
            "success": True,
            "categories": cat_list,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/expense-categories/main/{category_id}", response_model=ExpenseMainCategoryResponse)
async def get_expense_main_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single expense main category by ID with its sub-categories
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        category = ExpenseMainCategoryService.get_main_category(db, category_id, current_user)
        response = ExpenseMainCategoryResponse.model_validate(category).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "category": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/expense-categories/main/{category_id}", response_model=ExpenseMainCategorySimpleResponse)
async def update_expense_main_category(
    category_id: int,
    data: ExpenseMainCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an existing expense main category
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        category = ExpenseMainCategoryService.update_main_category(
            db=db,
            category_id=category_id,
            employee=current_user,
            name=data.name,
            description=data.description,
            is_active=data.is_active
        )
        response = ExpenseMainCategorySimpleResponse.model_validate(category).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Expense main category '{category.name}' updated successfully",
            "category": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/expense-categories/main/{category_id}")
async def delete_expense_main_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft delete (deactivate) an expense main category
    DC_SFMS_001: VGK/EA/Accounts access only
    Note: Cannot deactivate if active sub-categories exist
    """
    try:
        ExpenseMainCategoryService.delete_main_category(db, category_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Expense main category deactivated successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== EXPENSE SUB CATEGORY ENDPOINTS (SFMS Staff Access) ====================

@router.post("/expense-categories/sub", response_model=ExpenseSubCategoryResponse)
async def create_expense_sub_category(
    data: ExpenseSubCategoryCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new expense sub category
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        sub_cat = ExpenseSubCategoryService.create_sub_category(
            db=db,
            name=data.name,
            main_category_id=data.main_category_id,
            employee=current_user,
            description=data.description
        )
        response = ExpenseSubCategoryResponse.model_validate(sub_cat).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Expense sub-category '{sub_cat.name}' created successfully",
            "sub_category": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/expense-categories/sub", response_model=ExpenseSubCategoryListResponse)
async def list_expense_sub_categories(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
    main_category_id: Optional[int] = Query(None, description="Filter by main category"),
    include_inactive: bool = Query(False, description="Include inactive sub-categories"),
    search: Optional[str] = Query(None, description="Search by name"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List expense sub categories with filtering
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        sub_cats, total = ExpenseSubCategoryService.list_sub_categories(
            db, current_user, page, page_size, main_category_id, include_inactive, search
        )
        
        sub_cat_list = [
            ExpenseSubCategoryResponse.model_validate(s).model_dump(mode='json')
            for s in sub_cats
        ]
        
        return JSONResponse(content={
            "success": True,
            "sub_categories": sub_cat_list,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/expense-categories/sub/{sub_category_id}", response_model=ExpenseSubCategoryResponse)
async def get_expense_sub_category(
    sub_category_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single expense sub category by ID
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        sub_cat = ExpenseSubCategoryService.get_sub_category(db, sub_category_id, current_user)
        response = ExpenseSubCategoryResponse.model_validate(sub_cat).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "sub_category": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/expense-categories/sub/{sub_category_id}", response_model=ExpenseSubCategoryResponse)
async def update_expense_sub_category(
    sub_category_id: int,
    data: ExpenseSubCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update an existing expense sub category
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        sub_cat = ExpenseSubCategoryService.update_sub_category(
            db=db,
            sub_category_id=sub_category_id,
            employee=current_user,
            name=data.name,
            description=data.description,
            is_active=data.is_active
        )
        response = ExpenseSubCategoryResponse.model_validate(sub_cat).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Expense sub-category '{sub_cat.name}' updated successfully",
            "sub_category": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/expense-categories/sub/{sub_category_id}")
async def delete_expense_sub_category(
    sub_category_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Soft delete (deactivate) an expense sub category
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        ExpenseSubCategoryService.delete_sub_category(db, sub_category_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Expense sub-category deactivated successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== SEED DATA ENDPOINTS ====================

@router.post("/seed/mynt-real-llp")
async def seed_book_keeper(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Seed Mynt Real LLP as default book keeper (idempotent)
    DC: VGK/EA access only
    """
    try:
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        
        company = seed_mynt_real_llp(db, current_user.id)
        return JSONResponse(content={
            "success": True,
            "message": f"Book keeper '{company.company_name}' is ready",
            "company": AssociatedCompanyResponse.model_validate(company).model_dump(mode='json')
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/seed/income-sources/{company_id}")
async def seed_income_sources(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Seed default income source types for a company (idempotent)
    DC: VGK/EA access only
    """
    try:
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        
        created = seed_default_income_sources(db, company_id, current_user.id)
        return JSONResponse(content={
            "success": True,
            "message": f"Created {len(created)} income source types",
            "created_count": len(created)
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== SYSTEM / SEED ENDPOINTS ====================

@router.post("/system/seed")
async def run_sfms_seed_endpoint(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Initialize SFMS with default data (Mynt Real LLP, segments, income sources, pricing)
    DC_SFMS_SEED_001: Idempotent - safe to run multiple times
    DC: VGK/EA access only
    """
    try:
        validate_accounts_access(current_user)
        
        result = run_sfms_seed(db, current_user.id)
        
        if result["success"]:
            return JSONResponse(content={
                "success": True,
                "message": "SFMS seed completed successfully",
                "data": result
            })
        else:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={
                    "success": False,
                    "message": f"SFMS seed failed: {result.get('error', 'Unknown error')}"
                }
            )
    except Exception as e:
        return handle_accounts_error(e)


# ==================== DASHBOARD / SUMMARY ENDPOINTS ====================

@router.get("/dashboard/summary")
async def get_accounts_dashboard(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get accounts dashboard summary
    DC: VGK/EA access only
    """
    try:
        from app.models.staff_accounts import (
            AssociatedCompany, CompanySegment, IncomeSourceType,
            VendorMaster, StockItemMaster
        )
        
        validate_accounts_access(current_user)
        
        total_companies = db.query(AssociatedCompany).filter(
            AssociatedCompany.is_active == True
        ).count()
        
        total_segments = db.query(CompanySegment).filter(
            CompanySegment.is_active == True
        ).count()
        
        total_income_sources = db.query(IncomeSourceType).filter(
            IncomeSourceType.is_active == True
        ).count()
        
        total_vendors = db.query(VendorMaster).filter(
            VendorMaster.is_active == True
        ).count()
        
        total_stock_items = db.query(StockItemMaster).filter(
            StockItemMaster.is_active == True
        ).count()
        
        book_keeper = db.query(AssociatedCompany).filter(
            AssociatedCompany.is_book_keeper == True
        ).first()
        
        return JSONResponse(content={
            "success": True,
            "dashboard": {
                "total_companies": total_companies,
                "total_segments": total_segments,
                "total_income_sources": total_income_sources,
                "total_vendors": total_vendors,
                "total_stock_items": total_stock_items,
                "book_keeper": book_keeper.company_name if book_keeper else None,
                "book_keeper_id": book_keeper.id if book_keeper else None
            }
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== VENDOR TRANSACTION ENDPOINTS (Phase 2.9) ====================

from app.schemas.staff_accounts import (
    VendorTransactionCreate, VendorTransactionUpdate, VendorTransactionStatusChange,
    VendorTransactionResponse, VendorTransactionSimpleResponse, VendorTransactionListResponse,
    VendorTransactionLineItemCreate, VendorTransactionLineItemResponse
)
from app.services.staff_accounts_service import VendorTransactionService


@router.post("/vendor-transactions", response_model=VendorTransactionResponse)
async def create_vendor_transaction(
    data: VendorTransactionCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new vendor transaction with nested line items
    DC_SFMS_001: VGK/EA/Accounts access only
    Transaction types: PURCHASE, PAYMENT, ADVANCE, REFUND, DEBIT_NOTE, CREDIT_NOTE
    """
    try:
        transaction = VendorTransactionService.create_transaction(db, data, current_user)
        response = VendorTransactionResponse.model_validate(transaction).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Vendor transaction '{transaction.transaction_number}' created successfully",
            "transaction": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendor-transactions", response_model=VendorTransactionListResponse)
async def list_vendor_transactions(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    company_id: Optional[int] = Query(None, description="Filter by company"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor"),
    transaction_type: Optional[str] = Query(None, description="Filter by type (PURCHASE, PAYMENT, etc.)"),
    status: Optional[str] = Query(None, description="Filter by status"),
    date_from: Optional[date] = Query(None, description="Filter by start date"),
    date_to: Optional[date] = Query(None, description="Filter by end date"),
    search: Optional[str] = Query(None, description="Search by transaction number or narration"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List vendor transactions with filters
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transactions, total = VendorTransactionService.list_transactions(
            db, current_user, page, page_size, company_id, vendor_id,
            transaction_type, status, date_from, date_to, search
        )
        
        txn_list = [
            VendorTransactionSimpleResponse.model_validate(t).model_dump(mode='json')
            for t in transactions
        ]
        
        return JSONResponse(content={
            "success": True,
            "transactions": txn_list,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendor-transactions/{transaction_id}", response_model=VendorTransactionResponse)
async def get_vendor_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a single vendor transaction by ID with all line items
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transaction = VendorTransactionService.get_transaction(db, transaction_id, current_user)
        response = VendorTransactionResponse.model_validate(transaction).model_dump(mode='json')
        # Enrich with vendor details for display
        vendor = db.query(VendorMaster).filter(VendorMaster.id == transaction.vendor_id).first()
        if vendor:
            response['vendor_name'] = vendor.vendor_name
            response['vendor_gstin'] = vendor.gst_number
            response['vendor_phone'] = vendor.phone
            response['vendor_email'] = vendor.email
        response['payment_status'] = transaction.payment_status
        response['is_credit_purchase'] = transaction.is_credit_purchase
        response['due_date'] = str(transaction.due_date) if transaction.due_date else None
        response['credit_days'] = transaction.credit_days
        return JSONResponse(content={
            "success": True,
            "transaction": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/vendor-transactions/{transaction_id}", response_model=VendorTransactionResponse)
async def update_vendor_transaction(
    transaction_id: int,
    data: VendorTransactionUpdate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update a vendor transaction
    DC_SFMS_001: VGK/EA/Accounts access only
    Note: Only limited fields can be updated after transaction is submitted
    """
    try:
        transaction = VendorTransactionService.update_transaction(db, transaction_id, data, current_user)
        response = VendorTransactionResponse.model_validate(transaction).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Vendor transaction '{transaction.transaction_number}' updated successfully",
            "transaction": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-transactions/{transaction_id}/status", response_model=VendorTransactionResponse)
async def change_vendor_transaction_status(
    transaction_id: int,
    data: VendorTransactionStatusChange,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Change vendor transaction status
    DC_SFMS_001: VGK/EA/Accounts access only
    Valid transitions: DRAFT→SUBMITTED→APPROVED→PAID/CANCELLED
    """
    try:
        transaction = VendorTransactionService.change_status(
            db, transaction_id, data.new_status, current_user, data.reason
        )
        response = VendorTransactionResponse.model_validate(transaction).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": f"Transaction status changed to {data.new_status}",
            "transaction": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-transactions/{transaction_id}/line-items", response_model=VendorTransactionLineItemResponse)
async def add_vendor_transaction_line_item(
    transaction_id: int,
    data: VendorTransactionLineItemCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add a line item to a vendor transaction (DRAFT status only)
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        line_item = VendorTransactionService.add_line_item(db, transaction_id, data, current_user)
        response = VendorTransactionLineItemResponse.model_validate(line_item).model_dump(mode='json')
        return JSONResponse(content={
            "success": True,
            "message": "Line item added successfully",
            "line_item": response
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/vendor-transactions/{transaction_id}/line-items/{line_item_id}")
async def remove_vendor_transaction_line_item(
    transaction_id: int,
    line_item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Remove a line item from a vendor transaction (DRAFT status only)
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        VendorTransactionService.remove_line_item(db, transaction_id, line_item_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Line item removed successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== SERVICE ITEMS USED ENDPOINTS (Phase 2.10) ====================

from app.schemas.staff_accounts import (
    ServiceItemUsedCreate, ServiceItemUsedResponse, ServiceItemUsedListResponse,
    ServiceItemUsedCreateResponse
)
from app.services.staff_accounts_service import ServiceItemsUsedService


@router.post("/vendor-transactions/{transaction_id}/service-items", response_model=ServiceItemUsedCreateResponse)
async def add_service_item_used(
    transaction_id: int,
    data: ServiceItemUsedCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add a service item used to a SERVICE type transaction
    DC_SFMS_001: VGK/EA/Accounts access only
    Includes automatic incentive calculation based on pricing config
    """
    try:
        service_item = ServiceItemsUsedService.add_service_item(
            db, transaction_id, data.item_id, data.quantity_used,
            data.unit_of_measure, data.custom_price, current_user
        )
        return ServiceItemUsedCreateResponse(
            success=True,
            message=f"Service item '{service_item.item_code}' added with incentive: {service_item.incentive_amount}",
            item=ServiceItemUsedResponse.model_validate(service_item)
        )
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendor-transactions/{transaction_id}/service-items", response_model=ServiceItemUsedListResponse)
async def list_service_items_used(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List all service items used in a transaction with totals
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        items, totals = ServiceItemsUsedService.list_service_items(db, transaction_id, current_user)
        item_list = [ServiceItemUsedResponse.model_validate(item) for item in items]
        return ServiceItemUsedListResponse(
            success=True,
            items=item_list,
            total=len(items),
            total_selling=totals['total_selling'],
            total_cost=totals['total_cost'],
            total_profit=totals['total_profit'],
            total_incentive=totals['total_incentive']
        )
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/vendor-transactions/{transaction_id}/service-items/{service_item_id}")
async def remove_service_item_used(
    transaction_id: int,
    service_item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Remove a service item from a transaction (before stock deduction)
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        ServiceItemsUsedService.remove_service_item(db, transaction_id, service_item_id, current_user)
        return JSONResponse(content={
            "success": True,
            "message": "Service item removed successfully"
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== STOCK LEDGER ENDPOINTS (Phase 2.11) ====================

from app.schemas.staff_accounts import (
    StockLedgerResponse, StockLedgerListResponse,
    StockBalanceResponse, StockBalanceListResponse
)
from app.services.staff_accounts_service import StockLedgerService


@router.get("/stock-ledger", response_model=StockLedgerListResponse)
async def list_stock_ledger_entries(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    company_id: Optional[int] = Query(None, description="Filter by company"),
    item_id: Optional[int] = Query(None, description="Filter by item"),
    entry_type: Optional[str] = Query(None, description="Filter by entry type"),
    date_from: Optional[date] = Query(None, description="Filter by start date"),
    date_to: Optional[date] = Query(None, description="Filter by end date"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List stock ledger entries with filters
    DC_SFMS_001: VGK/EA/Accounts access only
    DC_STOCK_LEDGER_002: Include item_name and company_name for display
    """
    try:
        entries, total = StockLedgerService.list_entries(
            db, current_user, company_id, item_id, entry_type,
            date_from, date_to, page, page_size
        )
        
        from app.models.staff_accounts import StockItemMaster, AssociatedCompany
        
        item_ids = set(e.item_id for e in entries if e.item_id)
        company_ids = set(e.company_id for e in entries if e.company_id)
        updater_ids = set(e.updated_by_id for e in entries if getattr(e, 'updated_by_id', None))
        
        items_map = {}
        if item_ids:
            items = db.query(StockItemMaster).filter(StockItemMaster.id.in_(item_ids)).all()
            items_map = {i.id: {
                'code': i.item_code,
                'name': i.item_name,
                'brand': i.brand,
                'model_compat': i.model_compat,
                'specification': i.specification,
                'size': i.size,
            } for i in items}
        
        companies_map = {}
        if company_ids:
            companies = db.query(AssociatedCompany).filter(AssociatedCompany.id.in_(company_ids)).all()
            companies_map = {c.id: c.company_name for c in companies}
        
        employees_map = {}
        if updater_ids:
            emps = db.query(StaffEmployee).filter(StaffEmployee.id.in_(updater_ids)).all()
            employees_map = {e.id: f"{e.first_name or ''} {e.last_name or ''}".strip() or e.emp_code for e in emps}
        
        entry_list = []
        for entry in entries:
            entry_data = StockLedgerResponse.model_validate(entry).model_dump(mode='json')
            item_info = items_map.get(entry.item_id, {})
            entry_data['item_code'] = item_info.get('code', f'ITEM-{entry.item_id}')
            entry_data['item_name'] = item_info.get('name', f'Item #{entry.item_id}')
            entry_data['brand'] = item_info.get('brand')
            entry_data['model_compat'] = item_info.get('model_compat')
            entry_data['specification'] = item_info.get('specification') or getattr(entry, 'specification', None)
            entry_data['size'] = item_info.get('size')
            entry_data['company_name'] = companies_map.get(entry.company_id, f'Company #{entry.company_id}')
            upd_id = getattr(entry, 'updated_by_id', None)
            entry_data['updated_by_name'] = employees_map.get(upd_id, '') if upd_id else ''
            entry_data['updated_at'] = entry.updated_at.isoformat() if getattr(entry, 'updated_at', None) else None
            entry_data['color'] = getattr(entry, 'color', None)
            entry_list.append(entry_data)
        
        return JSONResponse(content={
            "success": True,
            "entries": entry_list,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-ledger/balances", response_model=StockBalanceListResponse)
async def get_stock_balances(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get current stock balances per item
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        balances = StockLedgerService.get_stock_balances(db, current_user, company_id)
        return JSONResponse(content={
            "success": True,
            "balances": [
                {
                    "item_id": b['item_id'],
                    "item_code": b['item_code'],
                    "item_name": b['item_name'],
                    "company_id": b['company_id'],
                    "balance_qty": str(b['balance_qty']),
                    "balance_value": str(b['balance_value']),
                    "last_transaction_date": str(b['last_transaction_date']) if b['last_transaction_date'] else None
                }
                for b in balances
            ],
            "total": len(balances)
        })
    except Exception as e:
        return handle_accounts_error(e)


# ==================== PARTY LEDGER ENDPOINTS (Phase 2.12) ====================

from app.schemas.staff_accounts import (
    PartyLedgerResponse, PartyLedgerListResponse,
    PartyBalanceResponse, PartyBalanceListResponse
)
from app.services.staff_accounts_service import PartyLedgerService


@router.get("/party-ledger/summary")
async def get_party_ledger_summary(
    company_id: Optional[int] = Query(None),
    party_type: Optional[str] = Query(None, description="VENDOR|CUSTOMER|EMPLOYEE|EXTERNAL"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-PWISE-001: Party-wise consolidated summary.
    Groups party_ledger by party_name + party_type; supports date + company filters.
    """
    try:
        from app.services.staff_accounts_service import PartyLedgerService
        rows = PartyLedgerService.get_party_summary(
            db, current_user,
            company_id=company_id,
            party_type=party_type,
            date_from=date_from,
            date_to=date_to,
        )
        return JSONResponse(content={"success": True, "total": len(rows), "parties": rows})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/party-ledger", response_model=PartyLedgerListResponse)
async def list_party_ledger_entries(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=10000, description="Items per page"),
    party_type: Optional[str] = Query(None, description="Filter by party type (VENDOR, EMPLOYEE, etc.)"),
    party_id: Optional[int] = Query(None, description="Filter by party ID"),
    party_name: Optional[str] = Query(None, description="Filter by party name (fuzzy match)"),
    company_id: Optional[int] = Query(None, description="Filter by company"),
    date_from: Optional[date] = Query(None, description="Filter by start date"),
    date_to: Optional[date] = Query(None, description="Filter by end date"),
    particulars: Optional[str] = Query(None, description="Search by stock item / SKU name in particulars"),
    reference_number: Optional[str] = Query(None, description="Search by reference number (PO#, Invoice#)"),
    reference_type: Optional[str] = Query(None, description="Filter by reference type (INCOME, SALES_INVOICE, etc.)"),
    source_status: Optional[str] = Query(None, description="Comma-separated voucher statuses: CONFIRMED,MANUAL,TALLY_IMPORT,OPENING_BALANCE"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List party ledger entries with filters and summary
    DC_SFMS_001: VGK/EA/Accounts access only
    DC: Added party_name fuzzy, particulars/SKU, reference_number, reference_type filters
    DC-SOURCE-STATUS-001: source_status multi-value filter added
    """
    try:
        # DC-PL-TYPE-NORM-001c: translate frontend/party-search types to DB-canonical types
        # so the ledger filter works regardless of which synonym the caller uses.
        _PL_TYPE_API_NORM = {'STAFF': 'EMPLOYEE', 'USER': 'MNR_USER'}
        if party_type:
            party_type = _PL_TYPE_API_NORM.get(party_type.upper(), party_type.upper())
        entries, total, summary = PartyLedgerService.list_entries(
            db, current_user, party_type, party_id, company_id,
            date_from, date_to, page, page_size,
            party_name=party_name,
            particulars=particulars,
            reference_number=reference_number,
            reference_type=reference_type,
            source_status=source_status,
        )
        # DC_LEDGER_CATEGORY_001: batch-load category names for display
        _pl_sub_ids  = list({e.sub_category_id  for e in entries if getattr(e, 'sub_category_id',  None)})
        _pl_main_ids = list({e.main_category_id for e in entries if getattr(e, 'main_category_id', None)})
        _pl_sub_names: dict  = {}
        _pl_main_names: dict = {}
        if _pl_sub_ids or _pl_main_ids:
            try:
                from app.models.expense_category import ExpenseSubCategory as _ESC_PL, ExpenseMainCategory as _EMC_PL
                if _pl_sub_ids:
                    _pl_sub_names = {s.id: s.name for s in db.query(_ESC_PL).filter(_ESC_PL.id.in_(_pl_sub_ids)).all()}
                if _pl_main_ids:
                    _pl_main_names = {m.id: m.name for m in db.query(_EMC_PL).filter(_EMC_PL.id.in_(_pl_main_ids)).all()}
            except Exception:
                pass

        def _enrich_pl(entry):
            d = PartyLedgerResponse.model_validate(entry).model_dump(mode='json')
            d['main_category_id']   = getattr(entry, 'main_category_id', None)
            d['sub_category_id']    = getattr(entry, 'sub_category_id', None)
            d['main_category_name'] = _pl_main_names.get(entry.main_category_id) if getattr(entry, 'main_category_id', None) else None
            d['sub_category_name']  = _pl_sub_names.get(entry.sub_category_id)   if getattr(entry, 'sub_category_id',  None) else None
            return d

        entry_list = [_enrich_pl(entry) for entry in entries]
        return JSONResponse(content={
            "success": True,
            "entries": entry_list,
            "total": total,
            "page": page,
            "page_size": page_size,
            "opening_balance": str(summary['opening_balance']),
            "opening_balance_date": summary.get('opening_balance_date'),
            "total_debit": str(summary['total_debit']),
            "total_credit": str(summary['total_credit']),
            "closing_balance": str(summary['closing_balance'])
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/party-ledger/employees")
async def list_employees_for_ledger(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Minimal employee list for party ledger party-name search (accessible to all accounts staff)."""
    employees = db.query(StaffEmployee).filter(
        StaffEmployee.status == 'active'
    ).order_by(StaffEmployee.full_name).all()
    return JSONResponse(content={
        "success": True,
        "employees": [
            {
                "id": e.id,
                "full_name": e.full_name or '',
                "emp_code": e.emp_code or ''
            }
            for e in employees
        ]
    })


@router.get("/party-ledger/parties")
async def list_ledger_parties(
    company_id: Optional[int] = Query(None),
    party_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-PARTY-MASTER-001: Single-source party list from master tables.
    Returns vendors + partners + employees from master tables (canonical names),
    plus any legacy ledger-only entries not in master tables."""
    from app.models.staff_accounts import VendorMaster as _VM_P, PartyLedger as _PL_P
    from app.models.staff_accounts import OfficialPartner as _OP_P
    from app.models.staff import StaffEmployee as _SE_P
    from sqlalchemy import func

    parties = []  # list of dicts: {party_type, party_id, party_name}
    seen_ids = {}  # (party_type, party_id) -> True

    # 1. VENDORS from VendorMaster (canonical names, single source)
    if not party_type or party_type == 'VENDOR':
        vq = db.query(_VM_P.id, _VM_P.vendor_name).filter(_VM_P.is_active == True)
        if company_id:
            from app.models.staff_accounts import VendorStockItemAssociation as _VSIA
            # Filter vendors applicable to the company
            vq = vq.filter(
                _VM_P.applicable_companies.op('@>')(func.to_jsonb(company_id))
            )
        for vid, vname in vq.order_by(_VM_P.vendor_name).all():
            parties.append({'party_type': 'VENDOR', 'party_id': vid, 'party_name': vname})
            seen_ids[('VENDOR', vid)] = True

    # 2. PARTNERS from OfficialPartner (canonical names)
    if not party_type or party_type == 'PARTNER':
        pq = db.query(_OP_P.id, _OP_P.partner_name).filter(_OP_P.is_active == True)
        if company_id:
            pq = pq.filter(_OP_P.company_id == company_id)
        for pid, pname in pq.order_by(_OP_P.partner_name).all():
            parties.append({'party_type': 'PARTNER', 'party_id': pid, 'party_name': pname})
            seen_ids[('PARTNER', pid)] = True

    # 3. EMPLOYEES from StaffEmployee
    if not party_type or party_type == 'EMPLOYEE':
        eq = db.query(_SE_P.id, _SE_P.full_name, _SE_P.emp_code).filter(_SE_P.status == 'active')
        for eid, ename, ecode in eq.order_by(_SE_P.full_name).all():
            label = (ename or '') + (f' ({ecode})' if ecode else '')
            parties.append({'party_type': 'EMPLOYEE', 'party_id': eid, 'party_name': label})
            seen_ids[('EMPLOYEE', eid)] = True

    # 4. Ledger-only entries (CUSTOMER, EXTERNAL, custom) not in master tables
    lq = db.query(_PL_P.party_type, _PL_P.party_id, _PL_P.party_name).distinct()
    if company_id:
        lq = lq.filter(_PL_P.company_id == company_id)
    legacy_types = ['CUSTOMER', 'EXTERNAL', 'USER']
    if party_type:
        if party_type in legacy_types:
            lq = lq.filter(_PL_P.party_type == party_type)
        else:
            lq = lq.filter(_PL_P.party_type == party_type)
    else:
        lq = lq.filter(_PL_P.party_type.in_(legacy_types))
    _leg_seen: dict = {}
    for lr in lq.order_by(_PL_P.party_name).all():
        if seen_ids.get((lr.party_type, lr.party_id)):
            continue
        _dk = ((lr.party_name or '').lower(), lr.party_type)
        if _dk not in _leg_seen:
            _leg_seen[_dk] = lr
        elif (lr.party_name or '') == (lr.party_name or '').upper():
            _leg_seen[_dk] = lr
    for lr in _leg_seen.values():
        parties.append({'party_type': lr.party_type, 'party_id': lr.party_id, 'party_name': lr.party_name})

    parties.sort(key=lambda x: (x['party_type'], (x['party_name'] or '').lower()))
    return JSONResponse(content={"success": True, "parties": parties})


@router.get("/party-ledger/balances", response_model=PartyBalanceListResponse)
async def get_party_balances(
    party_type: Optional[str] = Query(None, description="Filter by party type"),
    company_id: Optional[int] = Query(None, description="Filter by company"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get current balance summary per party
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        balances = PartyLedgerService.get_party_balances(db, current_user, party_type, company_id)
        return JSONResponse(content={
            "success": True,
            "balances": [
                {
                    "party_type": b['party_type'],
                    "party_id": b['party_id'],
                    "party_name": b['party_name'],
                    "company_id": b['company_id'],
                    "total_debit": str(b['total_debit']),
                    "total_credit": str(b['total_credit']),
                    "balance": str(b['balance']),
                    "last_transaction_date": str(b['last_transaction_date']) if b['last_transaction_date'] else None
                }
                for b in balances
            ],
            "total": len(balances)
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/party-ledger", status_code=201)
async def create_manual_ledger_entry(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_SFMS_MANUAL_001: Create a manual party ledger entry.
    Accepts: party_type, party_id, party_name, company_id, transaction_date,
             entry_type (DEBIT/CREDIT), amount, reference_number, narration
    """
    try:
        from app.models.staff_accounts import PartyLedger
        from decimal import Decimal
        import datetime as _dt

        party_type = data.get('party_type', 'EXTERNAL').upper()
        party_id = int(data.get('party_id', 0))
        party_name = str(data.get('party_name', 'Manual Entry')).strip()
        company_id = int(data['company_id'])
        txn_date_str = data.get('transaction_date', str(_dt.date.today()))
        txn_date = _dt.date.fromisoformat(txn_date_str)
        entry_type = data.get('entry_type', 'DEBIT').upper()
        amount = Decimal(str(data.get('amount', 0)))
        reference_number = data.get('reference_number', '') or ''
        narration = data.get('narration', '') or ''
        voucher_type = (data.get('voucher_type') or '').strip() or None
        particulars = (data.get('particulars') or '').strip() or None
        category    = (data.get('category') or '').strip() or None

        if entry_type not in ('DEBIT', 'CREDIT'):
            return JSONResponse(status_code=400, content={"success": False, "message": "entry_type must be DEBIT or CREDIT"})
        if amount <= 0:
            return JSONResponse(status_code=400, content={"success": False, "message": "amount must be greater than 0"})

        # Compute running balance from last entry for this party
        last = db.query(PartyLedger).filter(
            PartyLedger.party_type == party_type,
            PartyLedger.party_id == party_id,
            PartyLedger.company_id == company_id
        ).order_by(PartyLedger.transaction_date.desc(), PartyLedger.id.desc()).first()

        prev_balance = Decimal(str(last.running_balance)) if last else Decimal('0')
        if entry_type == 'DEBIT':
            running_balance = prev_balance + amount
        else:
            running_balance = prev_balance - amount

        entry = PartyLedger(
            party_type=party_type,
            party_id=party_id,
            party_name=party_name,
            company_id=company_id,
            transaction_date=txn_date,
            entry_type=entry_type,
            reference_type='MANUAL',
            reference_id=0,
            reference_number=reference_number or None,
            debit_amount=amount if entry_type == 'DEBIT' else Decimal('0'),
            credit_amount=amount if entry_type == 'CREDIT' else Decimal('0'),
            running_balance=running_balance,
            narration=narration or None,
            voucher_type=voucher_type,
            particulars=particulars,
            category=category,
            updated_by_id=current_user.id
        )
        db.add(entry)
        db.commit()
        db.refresh(entry)

        return JSONResponse(status_code=201, content={
            "success": True,
            "message": "Ledger entry created successfully",
            "entry_id": entry.id
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.post("/party-ledger/tally-import")
async def import_tally_ledger(
    file: UploadFile = File(...),
    company_id: int = Form(...),
    party_type: str = Form(...),
    party_id: int = Form(...),
    party_name: str = Form(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_SFMS_TALLY_002: Import Tally XML, CSV, or Tally PDF-text ledger files.
    - PDF text format: column-layout with Date/Particulars/Vch Type/Vch No./Debit/Credit
    - XML: Tally daybook/voucher export (YYYYMMDD dates, ALLLEDGERENTRIES.LIST)
    - CSV: Date,Reference,Description,Debit,Credit (+ optional Vch Type, Particulars columns)
    Returns: imported count, party name from file, balance verification, new particulars list
    """
    import xml.etree.ElementTree as ET
    import csv
    import io
    import re
    import datetime as _dt
    from decimal import Decimal
    from app.models.staff_accounts import PartyLedger

    def parse_indian_amount(s):
        if not s:
            return Decimal('0')
        s = str(s).replace(',', '').replace('\u2014', '').strip()
        s = re.sub(r'[^\d\.]', '', s)
        if not s:
            return Decimal('0')
        try:
            return Decimal(s)
        except Exception:
            return Decimal('0')

    def parse_tally_date(s):
        s = (s or '').strip()
        for fmt in ('%d-%b-%y', '%d-%b-%Y', '%Y%m%d', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def fuzzy_name_match(n1, n2):
        a = n1.lower().strip()
        b = n2.lower().strip()
        if a == b:
            return 1.0
        if a in b or b in a:
            return 0.85
        wa = set(w for w in a.split() if len(w) > 2)
        wb = set(w for w in b.split() if len(w) > 2)
        if not wa or not wb:
            return 0.0
        return len(wa & wb) / max(len(wa), len(wb))

    try:
        content = await file.read()
        filename = (file.filename or '').lower()
        entries_data = []
        party_name_in_file = None
        date_range_in_file = None
        file_closing_balance = None
        file_debit_total = None
        file_credit_total = None

        content_head = content[:600].decode('utf-8', errors='replace')
        is_xml = filename.endswith('.xml') or '<ENVELOPE>' in content_head or '<envelope>' in content_head.lower()
        is_tally_text = (
            filename.endswith('.pdf') or filename.endswith('.txt') or
            ('Date' in content_head and 'Particulars' in content_head and 'Vch Type' in content_head)
        )

        if is_xml:
            # ── XML: Tally daybook / voucher export ──
            try:
                root = ET.fromstring(content)
            except ET.ParseError as pe:
                return JSONResponse(status_code=400, content={"success": False, "message": f"Invalid XML: {pe}"})
            vouchers = root.findall('.//VOUCHER') or root.findall('.//voucher')
            for vch in vouchers:
                date_el = vch.find('DATE') or vch.find('date')
                narr_el = vch.find('NARRATION') or vch.find('narration')
                vch_no_el = vch.find('VOUCHERNUMBER') or vch.find('vouchernumber')
                vch_type_el = vch.find('VOUCHERTYPENAME') or vch.find('VOUCHTYPENAME')
                date_str = date_el.text.strip() if date_el is not None and date_el.text else None
                if not date_str:
                    continue
                txn_date = parse_tally_date(date_str)
                if not txn_date:
                    continue
                narration = narr_el.text.strip() if narr_el is not None and narr_el.text else ''
                vch_no = vch_no_el.text.strip() if vch_no_el is not None and vch_no_el.text else ''
                vch_type = vch_type_el.text.strip() if vch_type_el is not None and vch_type_el.text else ''
                for le in (vch.findall('.//ALLLEDGERENTRIES.LIST') + vch.findall('.//LEDGERENTRIES.LIST')):
                    amt_el = le.find('AMOUNT') or le.find('amount')
                    ledger_name_el = le.find('LEDGERNAME') or le.find('ledgername')
                    if amt_el is None or not amt_el.text:
                        continue
                    try:
                        raw_amount = Decimal(amt_el.text.strip().replace(',', ''))
                    except Exception:
                        continue
                    ledger_name = ledger_name_el.text.strip() if ledger_name_el is not None and ledger_name_el.text else ''
                    if raw_amount < 0:
                        entries_data.append({
                            'txn_date': txn_date, 'reference_number': vch_no,
                            'narration': narration, 'debit': abs(raw_amount), 'credit': Decimal('0'),
                            'voucher_type': vch_type, 'particulars': ledger_name
                        })
                    elif raw_amount > 0:
                        entries_data.append({
                            'txn_date': txn_date, 'reference_number': vch_no,
                            'narration': narration, 'debit': Decimal('0'), 'credit': raw_amount,
                            'voucher_type': vch_type, 'particulars': ledger_name
                        })

        elif is_tally_text:
            # ── PDF / plain-text: Tally printed ledger with column layout ──
            if filename.endswith('.pdf'):
                # ── Binary PDF: use pdfminer.six LTTextLine spatial extraction ──
                # Avoids pdfplumber's LTLine/original_path bug entirely.
                try:
                    from pdfminer.high_level import extract_pages
                    from pdfminer.layout import LTTextBox, LTTextLine, LAParams
                    import io as _io

                    laparams = LAParams(line_margin=0.3, char_margin=2.0, boxes_flow=None)
                    pdf_lines = []  # list of {x0, x1, y_top, text}
                    for _page_layout in extract_pages(_io.BytesIO(content), laparams=laparams):
                        _ph = _page_layout.height
                        for _box in _page_layout:
                            if isinstance(_box, LTTextBox):
                                for _line in _box:
                                    if isinstance(_line, LTTextLine):
                                        _txt = _line.get_text().strip()
                                        if _txt:
                                            _x0, _y0, _x1, _y1 = _line.bbox
                                            pdf_lines.append({
                                                'x0': _x0,
                                                'x1': _x1,   # right edge — critical for right-aligned cols
                                                'y_top': _ph - _y1,
                                                'text': _txt,
                                            })

                    # Sort top-to-bottom, left-to-right
                    pdf_lines.sort(key=lambda ll: (round(ll['y_top'] / 3) * 3, ll['x0']))

                    # Find header row
                    HDR_WORDS = {'Date', 'Particulars', 'Vch Type', 'Vch No.', 'Debit', 'Credit'}
                    header_y = None
                    col_x = {}
                    for _ll in pdf_lines:
                        if _ll['text'] in HDR_WORDS:
                            _y_bucket = round(_ll['y_top'] / 5) * 5
                            _same_row = [l2 for l2 in pdf_lines
                                         if abs(round(l2['y_top'] / 5) * 5 - _y_bucket) <= 1
                                         and l2['text'] in HDR_WORDS]
                            _found = {l2['text'] for l2 in _same_row}
                            if len(_found) >= 4:
                                header_y = _ll['y_top']
                                col_x = {l2['text']: l2['x0'] for l2 in _same_row}
                                break

                    if header_y is None or 'Vch Type' not in col_x:
                        return JSONResponse(status_code=400, content={
                            "success": False,
                            "message": "Could not find Tally ledger column header (Date / Particulars / Vch Type / Vch No.) in the PDF. "
                                       "Please export the ledger as plain text (.txt) from Tally."
                        })

                    vtype_x = col_x.get('Vch Type', 237)
                    vno_x   = col_x.get('Vch No.', 370)
                    debit_x = col_x.get('Debit', 466)
                    credit_x = col_x.get('Credit', 546)

                    # Extract metadata from lines ABOVE the header
                    _prev_txt = ''
                    for _ll in pdf_lines:
                        if _ll['y_top'] >= header_y - 2:
                            break
                        _t = _ll['text']
                        if 'Ledger Account' in _t:
                            party_name_in_file = _prev_txt
                        if re.search(r'\d{1,2}-[A-Za-z]{3}-\d{2}\s+to\s+\d{1,2}-[A-Za-z]{3}-\d{2}', _t, re.IGNORECASE):
                            date_range_in_file = _t
                        _prev_txt = _t

                    # Group data lines below header by row (y-bucket of 3px)
                    data_lines = [ll for ll in pdf_lines if ll['y_top'] > header_y + 3]
                    row_groups = {}
                    for _ll in data_lines:
                        _yb = round(_ll['y_top'] / 3) * 3
                        row_groups.setdefault(_yb, []).append(_ll)

                    # Column geometry: all positions are left-edges of header words (x0).
                    # Amounts are RIGHT-aligned within each column:
                    #   Debit amounts → x0 in [debit_x - col_width, credit_x)
                    #   Credit amounts → x0 in [credit_x - col_width, balance_x)
                    #   Balance amounts → x0 >= balance_x - col_width
                    col_width  = max(credit_x - debit_x, 40)   # e.g. 546-466 = 80 pts
                    balance_x  = credit_x + col_width            # estimated Balance column start

                    def _is_amount(txt):
                        return bool(re.search(r'\d[\d,]*\.\d{2}', txt))

                    _current_date = None
                    for _yb in sorted(row_groups):
                        _row = sorted(row_groups[_yb], key=lambda ll: ll['x0'])

                        # ── Step 1: split items into text-label items and numeric-amount items ──
                        _label_items  = [ll for ll in _row if not _is_amount(ll['text'])]
                        _amount_items = sorted(
                            [ll for ll in _row if _is_amount(ll['text'])],
                            key=lambda ll: ll['x0']
                        )

                        # ── Step 2: classify text labels by left-edge ──
                        _date_part_txt = ''
                        _vtype_txt = ''
                        _vno_txt   = ''
                        for _ll in _label_items:
                            _x0 = _ll['x0']
                            _t  = _ll['text']
                            if _x0 < vtype_x - 5:
                                _date_part_txt = _t
                            elif _x0 < vno_x - 5:
                                _vtype_txt = _t
                            elif _x0 < debit_x - 5:
                                _vno_txt = _t
                            # else: spurious column-header repeat in data zone — ignore

                        # ── Step 3: classify amounts by left-edge vs column boundaries ──
                        # Debit:   x0 < credit_x   (fits within debit column)
                        # Credit:  credit_x <= x0 < balance_x (fits within credit column)
                        # Balance: x0 >= balance_x - col_width (rightmost column) → SKIP
                        _debit_str  = ''
                        _credit_str = ''
                        for _ai in _amount_items:
                            _x0 = _ai['x0']
                            _t  = _ai['text']
                            if _x0 >= balance_x - 5:
                                pass  # Balance column — skip
                            elif _x0 >= credit_x - 5:
                                _credit_str = _t  # Credit column
                            elif _x0 >= debit_x - col_width - 5:
                                _debit_str = _t   # Debit column
                            # else: page numbers / small leading text — ignore

                        # Skip the totals summary row (both columns have amounts, no text label)
                        if _debit_str and _credit_str and not _vtype_txt and not _vno_txt:
                            _db_t = parse_indian_amount(re.findall(r'[\d,]+\.\d{2}', _debit_str)[0]) if re.findall(r'[\d,]+\.\d{2}', _debit_str) else None
                            _cr_t = parse_indian_amount(re.findall(r'[\d,]+\.\d{2}', _credit_str)[0]) if re.findall(r'[\d,]+\.\d{2}', _credit_str) else None
                            if _db_t:
                                file_debit_total = _db_t
                            if _cr_t:
                                file_credit_total = _cr_t
                            continue

                        # Closing Balance row: grab final amount
                        if 'Closing Balance' in _date_part_txt:
                            _amts = re.findall(r'[\d,]+\.\d{2}', _credit_str or _debit_str or _date_part_txt)
                            if _amts:
                                file_closing_balance = parse_indian_amount(_amts[-1])
                            continue

                        # Try to extract date from date_part_txt
                        _dm = re.search(r'(\d{1,2}-[A-Za-z]{3}-\d{2})', _date_part_txt, re.IGNORECASE)
                        if _dm:
                            _parsed = parse_tally_date(_dm.group(1))
                            if _parsed:
                                _current_date = _parsed

                        if _current_date is None:
                            continue

                        # Debit / credit amounts
                        _db_amts = re.findall(r'[\d,]+\.\d{2}', _debit_str)
                        _cr_amts = re.findall(r'[\d,]+\.\d{2}', _credit_str)
                        _debit  = parse_indian_amount(_db_amts[0]) if _db_amts else Decimal('0')
                        _credit = parse_indian_amount(_cr_amts[0]) if _cr_amts else Decimal('0')

                        if _debit == 0 and _credit == 0:
                            continue

                        # Strip Dr/Cr prefix from particulars and date prefix
                        _account = re.sub(r'^(Dr|Cr)\s+', '', _date_part_txt, flags=re.IGNORECASE).strip()
                        _account = re.sub(r'^\d{1,2}-[A-Za-z]{3}-\d{2}\s*', '', _account).strip()
                        _account = re.sub(r'^(Dr|Cr)\s+', '', _account, flags=re.IGNORECASE).strip()

                        entries_data.append({
                            'txn_date': _current_date,
                            'reference_number': _vno_txt,
                            'narration': _account,
                            'debit': _debit,
                            'credit': _credit,
                            'voucher_type': _vtype_txt,
                            'particulars': _account,
                        })

                except Exception as pdf_err:
                    return JSONResponse(status_code=400, content={
                        "success": False,
                        "message": f"Could not parse PDF: {pdf_err}. "
                                   "Please export the ledger as plain text (.txt) from Tally instead."
                    })

            elif not filename.endswith('.pdf'):
                # ── Plain text (.txt): column-based layout parser ──
                try:
                    content_str = content.decode('utf-8', errors='replace')
                except Exception:
                    content_str = content.decode('latin-1', errors='replace')

                lines = content_str.split('\n')

                # Find the header line: "Date  Particulars  Vch Type  Vch No.  Debit  Credit"
                header_idx = None
                for i, line in enumerate(lines):
                    if 'Date' in line and 'Particulars' in line and 'Vch Type' in line and 'Vch No' in line:
                        header_idx = i
                        break

                if header_idx is None:
                    return JSONResponse(status_code=400, content={
                        "success": False,
                        "message": "Could not find Tally ledger column header (Date / Particulars / Vch Type / Vch No.) in the file. "
                                   "Please export the ledger as PDF or plain text from Tally."
                    })

                hdr = lines[header_idx]

                def col(name):
                    idx = hdr.find(name)
                    return idx if idx >= 0 else None

                date_col  = col('Date')  or 0
                part_col  = col('Particulars') or 18
                vcht_col  = col('Vch Type') or 50
                vno_col   = col('Vch No') or 82
                db_col    = col('Debit') or 99
                cr_col    = col('Credit') or 113

                # Extract metadata from lines before the header
                for i in range(header_idx):
                    ls = lines[i].strip()
                    if not ls:
                        continue
                    if i + 1 < len(lines) and 'Ledger Account' in lines[i + 1]:
                        party_name_in_file = ls
                    if re.search(r'\d{1,2}-[A-Za-z]{3}-\d{2}\s+to\s+\d{1,2}-[A-Za-z]{3}-\d{2}', ls, re.IGNORECASE):
                        date_range_in_file = ls

                # Parse data rows
                current_date = None
                for line in lines[header_idx + 1:]:
                    stripped = line.strip()
                    if not stripped:
                        continue

                    if 'Closing Balance' in line:
                        amts = re.findall(r'[\d,]+\.\d{2}', line)
                        if amts:
                            file_closing_balance = parse_indian_amount(amts[-1])
                        continue
                    if re.match(r'^[\s\d,\.]+$', stripped):
                        amts = re.findall(r'[\d,]+\.\d{2}', line)
                        if len(amts) >= 2:
                            file_debit_total  = parse_indian_amount(amts[-2])
                            file_credit_total = parse_indian_amount(amts[-1])
                        continue

                    def extract(start, end=None):
                        if start is None or len(line) <= start:
                            return ''
                        seg = line[start:end] if end and len(line) > end else line[start:]
                        return seg.strip()

                    date_sec = extract(date_col, part_col)
                    part_sec = extract(part_col, vcht_col)
                    vcht_sec = extract(vcht_col, vno_col)
                    vno_sec  = extract(vno_col, db_col)
                    db_sec   = extract(db_col, cr_col)
                    cr_sec   = extract(cr_col)

                    dm = re.search(r'(\d{1,2}-[A-Za-z]{3}-\d{2})', date_sec, re.IGNORECASE)
                    if dm:
                        parsed = parse_tally_date(dm.group(1))
                        if parsed:
                            current_date = parsed

                    if current_date is None:
                        continue

                    db_amts = re.findall(r'[\d,]+\.\d{2}', db_sec)
                    cr_amts = re.findall(r'[\d,]+\.\d{2}', cr_sec)
                    debit  = parse_indian_amount(db_amts[0]) if db_amts else Decimal('0')
                    credit = parse_indian_amount(cr_amts[0]) if cr_amts else Decimal('0')

                    if debit == 0 and credit == 0:
                        continue

                    account_name = part_sec
                    m = re.match(r'^(Dr|Cr)\s+', part_sec, re.IGNORECASE)
                    if m:
                        account_name = part_sec[m.end():].strip()

                    entries_data.append({
                        'txn_date': current_date,
                        'reference_number': vno_sec,
                        'narration': account_name,
                        'debit': debit,
                        'credit': credit,
                        'voucher_type': vcht_sec,
                        'particulars': account_name,
                    })

        elif filename.endswith('.csv') or b'Date' in content[:100] or b'date' in content[:100]:
            # ── CSV: Date, Reference, Description, Debit, Credit (+ optional Vch Type, Particulars) ──
            text_content = content.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text_content))
            for row in reader:
                date_str = (row.get('Date') or row.get('date') or '').strip()
                if not date_str:
                    continue
                txn_date = parse_tally_date(date_str)
                if not txn_date:
                    continue
                ref  = (row.get('Reference') or row.get('reference') or row.get('Voucher No') or row.get('Vch No') or '').strip()
                desc = (row.get('Description') or row.get('description') or row.get('Narration') or row.get('narration') or '').strip()
                vch_type   = (row.get('Vch Type') or row.get('voucher_type') or row.get('Type') or '').strip()
                particulars_val = (row.get('Particulars') or row.get('particulars') or '').strip()
                debit  = parse_indian_amount((row.get('Debit')  or row.get('debit')  or '0').strip())
                credit = parse_indian_amount((row.get('Credit') or row.get('credit') or '0').strip())
                if debit > 0 or credit > 0:
                    entries_data.append({
                        'txn_date': txn_date, 'reference_number': ref,
                        'narration': desc, 'debit': debit, 'credit': credit,
                        'voucher_type': vch_type, 'particulars': particulars_val
                    })
        else:
            return JSONResponse(status_code=400, content={
                "success": False,
                "message": "Unsupported file format. Upload Tally XML, Tally PDF/text ledger, or CSV file."
            })

        if not entries_data:
            return JSONResponse(status_code=400, content={"success": False, "message": "No valid entries found in the uploaded file."})

        # Sort chronologically
        entries_data.sort(key=lambda x: x['txn_date'])

        # ─── Party name match ───
        name_match_score = 1.0
        name_deviation   = False
        if party_name_in_file:
            name_match_score = fuzzy_name_match(party_name_in_file, party_name)
            name_deviation   = name_match_score < 0.70

        # ─── Duplicate detection (by date + ref + debit + credit for TALLY_IMPORT rows) ───
        existing_keys = set()
        dup_q = db.query(PartyLedger).filter(
            PartyLedger.party_type == party_type,
            PartyLedger.party_id  == party_id,
            PartyLedger.company_id == company_id,
            PartyLedger.reference_type.in_(['TALLY_IMPORT', 'OPENING'])
        )
        # When party_id=0 all external parties share the same ID; scope by name
        if party_id == 0:
            dup_q = dup_q.filter(PartyLedger.party_name == party_name)
        for ex in dup_q.all():
            existing_keys.add((
                str(ex.transaction_date),
                str(ex.reference_number or ''),
                str(ex.debit_amount),
                str(ex.credit_amount)
            ))

        # ─── Running balance anchor ───
        bal_q = db.query(PartyLedger).filter(
            PartyLedger.party_type == party_type,
            PartyLedger.party_id  == party_id,
            PartyLedger.company_id == company_id
        )
        if party_id == 0:
            bal_q = bal_q.filter(PartyLedger.party_name == party_name)
        last = bal_q.order_by(PartyLedger.transaction_date.desc(), PartyLedger.id.desc()).first()
        running_balance = Decimal(str(last.running_balance)) if last else Decimal('0')

        created = 0
        skipped = 0
        new_particulars = set()

        for ed in entries_data:
            debit  = ed['debit']
            credit = ed['credit']
            ref_no = (ed.get('reference_number') or '').strip()

            dup_key = (str(ed['txn_date']), ref_no, str(debit), str(credit))
            if dup_key in existing_keys:
                skipped += 1
                continue

            entry_type = 'DEBIT' if debit > 0 else 'CREDIT'
            running_balance = running_balance + debit - credit

            part_val = (ed.get('particulars') or '').strip()
            narr_val = (ed.get('narration') or '').strip()
            is_opening = 'Opening Balance' in narr_val or 'Opening Balance' in part_val
            ref_type = 'OPENING' if is_opening else 'TALLY_IMPORT'

            if part_val and 'Opening Balance' not in part_val:
                new_particulars.add(part_val)

            new_entry = PartyLedger(
                party_type=party_type,
                party_id=party_id,
                party_name=party_name,
                company_id=company_id,
                transaction_date=ed['txn_date'],
                entry_type=entry_type,
                reference_type=ref_type,
                reference_id=0,
                reference_number=ref_no or None,
                debit_amount=debit,
                credit_amount=credit,
                running_balance=running_balance,
                narration=narr_val or None,
                voucher_type=(ed.get('voucher_type') or '').strip() or None,
                particulars=part_val or None,
                updated_by_id=current_user.id
            )
            db.add(new_entry)
            created += 1
            existing_keys.add(dup_key)  # prevent within-batch duplicates

        db.commit()

        # ─── Balance verification ───
        balance_verified = True
        balance_diff = None
        if file_closing_balance is not None:
            balance_diff = abs(running_balance - file_closing_balance)
            balance_verified = balance_diff < Decimal('1.00')

        msg = f"Successfully imported {created} ledger entr{'ies' if created != 1 else 'y'}"
        if skipped:
            msg += f" ({skipped} duplicate{'s' if skipped != 1 else ''} skipped)"

        return JSONResponse(content={
            "success": True,
            "message": msg,
            "imported": created,
            "skipped_duplicates": skipped,
            "closing_balance": str(running_balance),
            "party_name_in_file": party_name_in_file,
            "date_range_in_file": date_range_in_file,
            "file_closing_balance": str(file_closing_balance) if file_closing_balance is not None else None,
            "file_debit_total": str(file_debit_total) if file_debit_total is not None else None,
            "file_credit_total": str(file_credit_total) if file_credit_total is not None else None,
            "balance_verified": balance_verified,
            "balance_diff": str(balance_diff) if balance_diff is not None else None,
            "name_deviation": name_deviation,
            "name_match_score": round(name_match_score, 2),
            "new_particulars": sorted(new_particulars)[:30],
        })

    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


def _recompute_party_running_balances(db, party_type, party_id, party_name, company_id):
    """Re-sort all entries by date+id and recompute running_balance from scratch."""
    from app.models.staff_accounts import PartyLedger as _PL
    q = db.query(_PL).filter(
        _PL.party_type == party_type,
        _PL.party_id   == party_id,
        _PL.company_id == company_id,
    )
    if party_id == 0:
        q = q.filter(_PL.party_name == party_name)
    entries = q.order_by(_PL.transaction_date.asc(), _PL.id.asc()).all()
    bal = Decimal('0')
    for e in entries:
        bal = bal + Decimal(str(e.debit_amount or 0)) - Decimal(str(e.credit_amount or 0))
        e.running_balance = bal
    db.flush()


@router.patch("/party-ledger/rename-party")
async def rename_party_in_ledger(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Rename party_name (and optionally party_type) across all ledger entries where old_party_name+old_party_type matches."""
    from app.models.staff_accounts import PartyLedger
    try:
        company_id     = int(payload.get('company_id', 0))
        party_id       = int(payload.get('party_id', 0))
        party_type     = str(payload.get('party_type', '')).strip()      # old type filter
        old_party_name = str(payload.get('old_party_name', '')).strip()
        new_party_name = str(payload.get('new_party_name', '')).strip()
        new_party_type = str(payload.get('new_party_type', '')).strip()  # optional: change type too
        if not old_party_name or not new_party_name:
            return JSONResponse(status_code=400, content={"success": False, "message": "old_party_name and new_party_name are required"})
        q = db.query(PartyLedger).filter(
            PartyLedger.company_id == company_id,
            PartyLedger.party_name == old_party_name,
        )
        if party_id:
            q = q.filter(PartyLedger.party_id == party_id)
        if party_type:
            q = q.filter(PartyLedger.party_type == party_type)
        rows = q.all()
        updated = 0
        for row in rows:
            row.party_name = new_party_name
            if new_party_type:
                row.party_type = new_party_type
            updated += 1
        db.commit()
        # Recompute balances for the new combined party
        _recompute_party_running_balances(
            db,
            new_party_type if new_party_type else party_type,
            party_id,
            new_party_name,
            company_id
        )
        db.commit()
        type_msg = f" (type changed to {new_party_type})" if new_party_type and new_party_type != party_type else ""
        return JSONResponse(content={
            "success": True,
            "message": f"Updated {updated} entries: renamed '{old_party_name}' → '{new_party_name}'{type_msg}",
            "updated": updated
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.patch("/party-ledger/{entry_id}")
async def update_ledger_entry(
    entry_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update a single party ledger entry and recompute running balances."""
    from app.models.staff_accounts import PartyLedger
    try:
        entry = db.query(PartyLedger).filter(PartyLedger.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "message": "Entry not found"})

        if 'transaction_date' in payload and payload['transaction_date']:
            from datetime import date as _date
            try:
                entry.transaction_date = _date.fromisoformat(payload['transaction_date'])
            except ValueError:
                return JSONResponse(status_code=400, content={"success": False, "message": "Invalid date format"})

        if 'entry_type' in payload and payload['entry_type'] in ('DEBIT', 'CREDIT'):
            new_type = payload['entry_type']
            entry.entry_type = new_type
            # Swap debit/credit amounts according to new type
            raw_amt = Decimal(str(payload.get('amount', entry.debit_amount or entry.credit_amount or 0)))
            entry.debit_amount  = raw_amt if new_type == 'DEBIT' else Decimal('0')
            entry.credit_amount = raw_amt if new_type == 'CREDIT' else Decimal('0')
        elif 'amount' in payload:
            raw_amt = Decimal(str(payload['amount']))
            if entry.entry_type == 'DEBIT':
                entry.debit_amount  = raw_amt
                entry.credit_amount = Decimal('0')
            else:
                entry.credit_amount = raw_amt
                entry.debit_amount  = Decimal('0')

        if 'reference_number' in payload:
            entry.reference_number = payload['reference_number'] or None
        if 'voucher_type' in payload:
            entry.voucher_type = payload['voucher_type'] or None
        if 'particulars' in payload:
            entry.particulars = payload['particulars'] or None
        if 'narration' in payload:
            entry.narration = payload['narration'] or None

        # Recompute all running balances for this party after update
        _recompute_party_running_balances(
            db, entry.party_type, entry.party_id, entry.party_name, entry.company_id
        )
        db.commit()
        return JSONResponse(content={"success": True, "message": "Entry updated and balances recomputed"})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.delete("/party-ledger/{entry_id}")
async def delete_ledger_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Delete a party ledger entry and recompute running balances."""
    from app.models.staff_accounts import PartyLedger
    try:
        entry = db.query(PartyLedger).filter(PartyLedger.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "message": "Entry not found"})
        pt, pid, pn, cid = entry.party_type, entry.party_id, entry.party_name, entry.company_id
        db.delete(entry)
        db.flush()
        _recompute_party_running_balances(db, pt, pid, pn, cid)
        db.commit()
        return JSONResponse(content={"success": True, "message": "Entry deleted and balances recomputed"})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)




# ==================== VENDOR RETURN ENDPOINTS (Phase 2.13) ====================

from app.schemas.staff_accounts import (
    VendorReturnCreate, VendorReturnResolve, VendorReturnStatusChange,
    VendorReturnResponse, VendorReturnListResponse
)
from app.services.staff_accounts_service import VendorReturnService


@router.post("/vendor-returns", response_model=VendorReturnResponse, status_code=201)
async def create_vendor_return(
    data: VendorReturnCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new vendor return
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor_return = VendorReturnService.create_return(db, data, current_user)
        return VendorReturnResponse.model_validate(vendor_return)
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendor-returns", response_model=VendorReturnListResponse)
async def list_vendor_returns(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    company_id: Optional[int] = Query(None, description="Filter by company"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor"),
    status: Optional[str] = Query(None, description="Filter by status"),
    resolution_status: Optional[str] = Query(None, description="Filter by resolution status"),
    date_from: Optional[date] = Query(None, description="Filter by start date"),
    date_to: Optional[date] = Query(None, description="Filter by end date"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List vendor returns with filters
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        returns, total = VendorReturnService.list_returns(
            db, current_user, company_id, vendor_id, status,
            resolution_status, date_from, date_to, page, page_size
        )
        return_list = [
            VendorReturnResponse.model_validate(r).model_dump(mode='json')
            for r in returns
        ]
        return JSONResponse(content={
            "success": True,
            "returns": return_list,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


# ===========================================================================================
# VENDOR RETURN REQUESTS WORKFLOW (DC_INTAKE_001 - Jan 2026)
# NOTE: These /requests routes MUST be declared BEFORE /{return_id} to avoid route conflicts
# ===========================================================================================

@router.get("/vendor-returns/requests")
async def list_vendor_return_requests(
    company_id: Optional[int] = None,
    status: Optional[str] = None,
    vendor_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List vendor return requests
    DC Protocol: Company-scoped access
    """
    try:
        from app.services.staff_accounts_service import VendorReturnRequestService
        
        filters = {
            'company_id': company_id,
            'status': status,
            'vendor_id': vendor_id
        }
        
        returns, total = VendorReturnRequestService.list_returns(
            db, current_user, filters, page, page_size
        )
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "items": returns,
                "total": total,
                "page": page,
                "page_size": page_size
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/requests")
async def create_vendor_return_request(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create vendor return request from rejected intake items
    DC Protocol: Links to intake items, triggers notification
    """
    try:
        from app.services.staff_accounts_service import VendorReturnRequestService
        
        return_request = VendorReturnRequestService.create_return(
            db, current_user, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Return request created",
            "data": return_request
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/vendor-returns/requests/{return_id}")
async def get_vendor_return_request(
    return_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get single vendor return request with items"""
    try:
        from app.services.staff_accounts_service import VendorReturnRequestService
        
        return_request = VendorReturnRequestService.get_return(db, current_user, return_id)
        
        return JSONResponse(content={
            "success": True,
            "data": return_request
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/requests/{return_id}/send-notification")
async def send_return_request_notification(
    return_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Send email/WhatsApp notification to vendor
    DC Protocol: One-click trigger, tracks notification history
    """
    try:
        from app.services.staff_accounts_service import VendorReturnRequestService
        
        channel = data.get('channel', 'email')
        
        result = VendorReturnRequestService.send_notification(
            db, current_user, return_id, channel
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Notification sent via {channel}",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/requests/{return_id}/dispatch")
async def dispatch_return_request(
    return_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record return dispatch to vendor
    DC Protocol: Tracks courier/vehicle details
    """
    try:
        from app.services.staff_accounts_service import VendorReturnRequestService
        
        result = VendorReturnRequestService.record_dispatch(
            db, current_user, return_id, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Dispatch recorded",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/requests/{return_id}/complete")
async def complete_vendor_return_request(
    return_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Complete return with credit note/replacement
    DC Protocol: Updates party ledger, creates SFMS journal entry
    """
    try:
        from app.services.staff_accounts_service import VendorReturnRequestService
        
        resolution_type = data.get('resolution_type', 'credit_note')
        
        result = VendorReturnRequestService.complete_return(
            db, current_user, return_id, resolution_type, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Return completed with {resolution_type}",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ===========================================================================================
# VENDOR RETURN BY ID (Phase 2.13) - MUST be after /requests routes
# ===========================================================================================

@router.get("/vendor-returns/{return_id}", response_model=VendorReturnResponse)
async def get_vendor_return(
    return_id: int = Path(..., description="Vendor return ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a vendor return by ID
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor_return = VendorReturnService.get_return(db, return_id, current_user)
        return VendorReturnResponse.model_validate(vendor_return)
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/vendor-returns/{return_id}/status")
async def change_vendor_return_status(
    return_id: int = Path(..., description="Vendor return ID"),
    data: VendorReturnStatusChange = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Change vendor return status with workflow validation
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor_return = VendorReturnService.change_status(
            db, return_id, data.status, current_user, data.remarks
        )
        return VendorReturnResponse.model_validate(vendor_return)
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/{return_id}/resolve")
async def resolve_vendor_return(
    return_id: int = Path(..., description="Vendor return ID"),
    data: VendorReturnResolve = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Resolve a vendor return with credit note, replacement, or refund
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor_return = VendorReturnService.resolve_return(db, return_id, data, current_user)
        return VendorReturnResponse.model_validate(vendor_return)
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/{return_id}/reverse-stock")
async def reverse_vendor_return_stock(
    return_id: int = Path(..., description="Vendor return ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Reverse stock ledger entry for a vendor return
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        vendor_return = VendorReturnService.reverse_stock(db, return_id, current_user)
        return VendorReturnResponse.model_validate(vendor_return)
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/vendor-returns/requests/{return_id}/log-whatsapp")
async def log_vendor_return_whatsapp(
    return_id: int,
    phone: str = Body(..., embed=True),
    message: str = Body(..., embed=True),
    context: Optional[str] = Body(default=None, embed=True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-VENDOR-REPAIR-TRACKER-001: Log a WhatsApp send for a vendor return request (stock returns).
    Appends entry to whatsapp_log JSONB. Called when staff clicks 'Open WhatsApp' in tracker.
    """
    import json
    from sqlalchemy import text as _wl_t
    from app.models.staff_accounts import VendorReturnRequest
    from datetime import datetime

    req = db.query(VendorReturnRequest).filter(VendorReturnRequest.id == return_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Return request not found")

    existing = req.whatsapp_log or []
    existing.append({
        "sent_at": datetime.now().isoformat(),
        "sent_by": f"{current_user.first_name} {current_user.last_name}".strip(),
        "sent_by_id": current_user.id,
        "phone": phone,
        "message": message,
        "context": context or "Vendor Return Communication"
    })
    db.execute(_wl_t(
        "UPDATE vendor_return_requests SET whatsapp_log = :log, last_action_date = NOW() WHERE id = :id"
    ), {"log": json.dumps(existing), "id": return_id})
    db.commit()
    return {"success": True, "total_sends": len(existing), "whatsapp_log": existing}


# ==================== STOCK TRANSFER ENDPOINTS (Phase 2.14) ====================

from app.schemas.staff_accounts import (
    StockTransferCreate, StockTransferDispatch, StockTransferReceive,
    StockTransferStatusChange, StockTransferResponse, StockTransferListResponse
)
from app.services.staff_accounts_service import StockTransferService


@router.post("/stock-transfers", response_model=StockTransferResponse, status_code=201)
async def create_stock_transfer(
    data: StockTransferCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new stock transfer
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transfer = StockTransferService.create_transfer(db, data, current_user)
        return StockTransferResponse.model_validate(transfer)
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-transfers", response_model=StockTransferListResponse)
async def list_stock_transfers(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    from_company_id: Optional[int] = Query(None, description="Filter by source company"),
    to_company_id: Optional[int] = Query(None, description="Filter by destination company"),
    status: Optional[str] = Query(None, description="Filter by status"),
    transfer_type: Optional[str] = Query(None, description="Filter by transfer type"),
    date_from: Optional[date] = Query(None, description="Filter by start date"),
    date_to: Optional[date] = Query(None, description="Filter by end date"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List stock transfers with filters
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transfers, total = StockTransferService.list_transfers(
            db, current_user, from_company_id, to_company_id, status,
            transfer_type, date_from, date_to, page, page_size
        )
        transfer_list = [
            StockTransferResponse.model_validate(t).model_dump(mode='json')
            for t in transfers
        ]
        return JSONResponse(content={
            "success": True,
            "transfers": transfer_list,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-transfers/{transfer_id}", response_model=StockTransferResponse)
async def get_stock_transfer(
    transfer_id: int = Path(..., description="Stock transfer ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a stock transfer by ID
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transfer = StockTransferService.get_transfer(db, transfer_id, current_user)
        return StockTransferResponse.model_validate(transfer)
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-transfers/{transfer_id}/dispatch")
async def dispatch_stock_transfer(
    transfer_id: int = Path(..., description="Stock transfer ID"),
    data: StockTransferDispatch = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Dispatch a stock transfer - creates stock OUT for source company
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        remarks = data.remarks if data else None
        transfer = StockTransferService.dispatch_transfer(db, transfer_id, current_user, remarks)
        return StockTransferResponse.model_validate(transfer)
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-transfers/{transfer_id}/receive")
async def receive_stock_transfer(
    transfer_id: int = Path(..., description="Stock transfer ID"),
    data: StockTransferReceive = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Receive a stock transfer - creates stock IN for destination company
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        received_quantity = data.received_quantity if data else None
        remarks = data.remarks if data else None
        transfer = StockTransferService.receive_transfer(
            db, transfer_id, current_user, received_quantity, remarks
        )
        return StockTransferResponse.model_validate(transfer)
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/stock-transfers/{transfer_id}/status")
async def change_stock_transfer_status(
    transfer_id: int = Path(..., description="Stock transfer ID"),
    data: StockTransferStatusChange = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Change stock transfer status with workflow validation
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transfer = StockTransferService.change_status(
            db, transfer_id, data.status, current_user, data.remarks
        )
        return StockTransferResponse.model_validate(transfer)
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-transfers/{transfer_id}/cancel")
async def cancel_stock_transfer(
    transfer_id: int = Path(..., description="Stock transfer ID"),
    reason: str = Query(..., description="Cancellation reason"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Cancel a stock transfer with optional ledger reversal
    DC_SFMS_001: VGK/EA/Accounts access only
    """
    try:
        transfer = StockTransferService.cancel_transfer(db, transfer_id, current_user, reason)
        return StockTransferResponse.model_validate(transfer)
    except Exception as e:
        return handle_accounts_error(e)


# ==================== FUND ALLOCATION ENDPOINTS (Phase 2.15) ====================

@router.get("/fund-allocation-parties")
async def get_fund_allocation_parties(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC Protocol (May 2026): Dedicated endpoint for fund allocation party/recipient dropdown.
    Returns all active employees regardless of the caller's role — accounts-allowed users
    need to see the full employee list to select allocation recipients.
    Gated by is_accounts_allowed_employee so only authorised accounts staff can call it.
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Accounts access required")
    employees = (
        db.query(StaffEmployee)
        .filter(StaffEmployee.status == 'active')
        .order_by(StaffEmployee.full_name)
        .all()
    )
    return {
        "success": True,
        "employees": [
            {
                "id": e.id,
                "emp_code": e.emp_code,
                "full_name": e.full_name or f"{e.first_name or ''} {e.last_name or ''}".strip(),
                "is_active": e.status == 'active',
                "department": e.department.name if e.department else None,
                "role": e.role.role_name if e.role else None,
            }
            for e in employees
        ]
    }


@router.get("/expense-behalf-employees")
async def get_expense_behalf_employees(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_EXP_BEHALF_001: Returns employees the current user can create expenses on behalf of.
    Privileged (VGK/EA/Accounts) → all active employees.
    Reporting managers (non-privileged) → their direct reports only.
    Others → empty list.
    """
    if is_accounts_allowed_employee(current_user):
        employees = (
            db.query(StaffEmployee)
            .filter(StaffEmployee.status == 'active')
            .order_by(StaffEmployee.full_name)
            .all()
        )
    else:
        employees = (
            db.query(StaffEmployee)
            .filter(
                StaffEmployee.reporting_manager_id == current_user.id,
                StaffEmployee.status == 'active'
            )
            .order_by(StaffEmployee.full_name)
            .all()
        )
    return {
        "success": True,
        "employees": [
            {
                "id": e.id,
                "emp_code": e.emp_code,
                "full_name": e.full_name or f"{e.first_name or ''} {e.last_name or ''}".strip(),
                "department": e.department.name if e.department else None,
                "role": e.role.role_name if e.role else None,
            }
            for e in employees
        ]
    }


@router.get("/expense-consolidated")
async def get_expense_consolidated(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    from_date: Optional[date] = Query(None, description="Filter expense date from"),
    to_date: Optional[date] = Query(None, description="Filter expense date to"),
    search: Optional[str] = Query(None, description="Search by employee name or emp_code"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_EXP_CONSOLIDATED_001: Per-employee expense + fund summary table.
    Privileged (Accounts/VGK/EA/MR10001) → ALL active employees.
    Reporting managers (non-privileged) → their direct reports only.
    """
    from sqlalchemy import func as _func, case as _case, or_ as _or

    _is_privileged = is_accounts_allowed_employee(current_user)

    # Determine which employees are visible to this viewer
    emp_q = db.query(StaffEmployee).filter(StaffEmployee.status == 'active')
    if not _is_privileged:
        # Reporting managers see only their direct reports
        direct_report_ids = [
            r[0] for r in db.query(StaffEmployee.id).filter(
                StaffEmployee.reporting_manager_id == current_user.id
            ).all()
        ]
        if not direct_report_ids:
            raise HTTPException(status_code=403, detail="No team data available. Only Accounts/VGK/EA/Managers can view consolidated expenses.")
        emp_q = emp_q.filter(StaffEmployee.id.in_(direct_report_ids))

    if search:
        _s = f"%{search}%"
        emp_q = emp_q.filter(
            _or(
                StaffEmployee.full_name.ilike(_s),
                StaffEmployee.emp_code.ilike(_s)
            )
        )
    employees = emp_q.order_by(StaffEmployee.full_name).all()
    emp_ids = [e.id for e in employees]
    if not emp_ids:
        return {"success": True, "rows": [], "total": 0}

    # Aggregate fund allocations per employee
    fa_q = (
        db.query(
            FundAllocation.to_employee_id,
            _func.sum(FundAllocation.amount).label('total_allocated'),
            _func.sum(FundAllocation.balance_remaining).label('total_balance'),
            _func.sum(FundAllocation.total_expensed).label('total_expensed'),
        )
        .filter(
            FundAllocation.to_employee_id.in_(emp_ids),
            FundAllocation.status.in_(['PENDING', 'CONFIRMED', 'PARTIALLY_SETTLED', 'SETTLED'])
        )
    )
    if company_id:
        fa_q = fa_q.filter(FundAllocation.company_id == company_id)
    fa_rows = fa_q.group_by(FundAllocation.to_employee_id).all()
    fa_map = {r.to_employee_id: r for r in fa_rows}

    # DC_CONSO_TRANSFER_001: Aggregate fund transfers per employee (sent / received)
    # This ensures fund_balance correctly reflects transfers out/in — not just allocations.
    from app.models.staff_accounts import EmployeeFundTransfer as _EFT
    ft_sent_q = (
        db.query(
            _EFT.from_employee_id.label('employee_id'),
            _func.coalesce(_func.sum(_EFT.amount), 0).label('total_sent'),
        )
        .filter(
            _EFT.from_employee_id.in_(emp_ids),
            _EFT.status == 'CONFIRMED',
        )
    )
    ft_recv_q = (
        db.query(
            _EFT.to_employee_id.label('employee_id'),
            _func.coalesce(_func.sum(_EFT.amount), 0).label('total_received'),
        )
        .filter(
            _EFT.to_employee_id.in_(emp_ids),
            _EFT.status == 'CONFIRMED',
        )
    )
    if company_id:
        ft_sent_q = ft_sent_q.filter(_EFT.company_id == company_id)
        ft_recv_q = ft_recv_q.filter(_EFT.company_id == company_id)
    ft_sent_map = {r.employee_id: float(r.total_sent or 0) for r in ft_sent_q.group_by(_EFT.from_employee_id).all()}
    ft_recv_map = {r.employee_id: float(r.total_received or 0) for r in ft_recv_q.group_by(_EFT.to_employee_id).all()}

    # Aggregate expense entries per employee
    exp_q = (
        db.query(
            ExpenseEntry.created_by_id,
            _func.count().label('total_count'),
            _func.sum(_case((ExpenseEntry.status == 'DRAFT', 1), else_=0)).label('draft_count'),
            _func.sum(_case((ExpenseEntry.status == 'SUBMITTED', 1), else_=0)).label('submitted_count'),
            _func.sum(_case((ExpenseEntry.status == 'APPROVED', 1), else_=0)).label('approved_count'),
            _func.sum(_case((ExpenseEntry.status == 'REJECTED', 1), else_=0)).label('rejected_count'),
            _func.sum(_case((ExpenseEntry.is_paid == True, 1), else_=0)).label('paid_count'),
            _func.sum(ExpenseEntry.amount).label('total_amount'),
            _func.sum(_case((ExpenseEntry.status == 'DRAFT', ExpenseEntry.amount), else_=0)).label('draft_amount'),
            _func.sum(_case((ExpenseEntry.status == 'SUBMITTED', ExpenseEntry.amount), else_=0)).label('submitted_amount'),
            _func.sum(_case((ExpenseEntry.status == 'APPROVED', ExpenseEntry.amount), else_=0)).label('approved_amount'),
            _func.sum(_case((ExpenseEntry.status == 'REJECTED', ExpenseEntry.amount), else_=0)).label('rejected_amount'),
            _func.sum(_case((ExpenseEntry.is_paid == True, ExpenseEntry.amount), else_=0)).label('paid_amount'),
        )
        .filter(ExpenseEntry.created_by_id.in_(emp_ids))
    )
    if company_id:
        exp_q = exp_q.filter(ExpenseEntry.company_id == company_id)
    if from_date:
        exp_q = exp_q.filter(ExpenseEntry.expense_date >= from_date)
    if to_date:
        exp_q = exp_q.filter(ExpenseEntry.expense_date <= to_date)
    exp_rows = exp_q.group_by(ExpenseEntry.created_by_id).all()
    exp_map = {r.created_by_id: r for r in exp_rows}

    # DC_CONSO_CASH_RECV_001: Aggregate cash received by each employee as IE destination
    # Field staff receive customer payments (IE destination_type='EMPLOYEE'); show as Cash IN
    ie_recv_q = (
        db.query(
            IncomeEntry.destination_employee_id,
            _func.sum(IncomeEntry.amount).label('total_received'),
            _func.count().label('receipt_count'),
        )
        .filter(
            IncomeEntry.destination_employee_id.in_(emp_ids),
            IncomeEntry.destination_type == 'EMPLOYEE',
            IncomeEntry.reference_type != 'JOURNAL_VOUCHER',
            IncomeEntry.is_deleted == False,
        )
    )
    if company_id:
        ie_recv_q = ie_recv_q.filter(IncomeEntry.company_id == company_id)
    ie_recv_rows = ie_recv_q.group_by(IncomeEntry.destination_employee_id).all()
    ie_map = {r.destination_employee_id: r for r in ie_recv_rows}

    rows = []
    for emp in employees:
        fa = fa_map.get(emp.id)
        ex = exp_map.get(emp.id)
        ir = ie_map.get(emp.id)
        if fa is None and ex is None and ir is None:
            continue  # skip employees with zero data across all cash flows
        # DC_CONSO_TRANSFER_001: fund_balance = allocated + received_transfers - sent_transfers - approved_expenses
        total_allocated = float(fa.total_allocated or 0) if fa else 0
        approved_amount = float(ex.approved_amount or 0) if ex else 0
        sent = ft_sent_map.get(emp.id, 0.0)
        received = ft_recv_map.get(emp.id, 0.0)
        fund_balance = total_allocated + received - sent - approved_amount
        cash_received = float(ir.total_received or 0) if ir else 0
        cash_receipt_count = int(ir.receipt_count or 0) if ir else 0
        # cash_balance = cash received from customers - approved expenses (field staff ledger)
        cash_balance = cash_received - approved_amount
        rows.append({
            "employee_id": emp.id,
            "emp_code": emp.emp_code,
            "full_name": emp.full_name or "",
            "department": emp.department.name if emp.department else None,
            "role": emp.role.role_name if emp.role else None,
            "fund_allocated": total_allocated,
            "fund_transferred_out": sent,
            "fund_transferred_in": received,
            "fund_balance": fund_balance,
            "fund_used": float(fa.total_expensed or 0) if fa else 0,
            "cash_received": cash_received,
            "cash_receipt_count": cash_receipt_count,
            "cash_balance": cash_balance,
            "total_expenses": int(ex.total_count or 0) if ex else 0,
            "draft_count": int(ex.draft_count or 0) if ex else 0,
            "submitted_count": int(ex.submitted_count or 0) if ex else 0,
            "approved_count": int(ex.approved_count or 0) if ex else 0,
            "rejected_count": int(ex.rejected_count or 0) if ex else 0,
            "paid_count": int(ex.paid_count or 0) if ex else 0,
            "total_amount": float(ex.total_amount or 0) if ex else 0,
            "draft_amount": float(ex.draft_amount or 0) if ex else 0,
            "submitted_amount": float(ex.submitted_amount or 0) if ex else 0,
            "approved_amount": float(ex.approved_amount or 0) if ex else 0,
            "rejected_amount": float(ex.rejected_amount or 0) if ex else 0,
            "paid_amount": float(ex.paid_amount or 0) if ex else 0,
        })

    return {"success": True, "rows": rows, "total": len(rows)}


@router.post("/fund-allocations", response_model=FundAllocationResponse, status_code=201)
async def create_fund_allocation_endpoint(
    data: FundAllocationCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new fund allocation from accountant to employee
    DC_SFMS_001: VGK/EA/Accounts access only for creation
    """
    try:
        validate_accounts_access(current_user)
        allocation = create_fund_allocation(db, current_user, data.model_dump())
        return FundAllocationResponse.model_validate(allocation)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/fund-allocations")
async def list_fund_allocations_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    status: Optional[str] = Query(None, description="Filter by status"),
    from_employee_id: Optional[int] = Query(None, description="Filter by allocator"),
    to_employee_id: Optional[int] = Query(None, description="Filter by recipient"),
    from_date: Optional[date] = Query(None, description="Filter from date"),
    to_date: Optional[date] = Query(None, description="Filter to date"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List fund allocations with filters — enriched with employee and company names
    DC_SFMS_001: VGK/EA see all, others see only their allocations (no RBAC check - ownership enforced in service)
    """
    try:
        filters = {
            'company_id': company_id,
            'status': status,
            'from_employee_id': from_employee_id,
            'to_employee_id': to_employee_id,
            'from_date': from_date,
            'to_date': to_date,
            'page': page,
            'page_size': page_size
        }
        allocations, total = list_fund_allocations(db, current_user, filters)

        emp_ids = set()
        co_ids = set()
        for a in allocations:
            emp_ids.add(a.from_employee_id)
            emp_ids.add(a.to_employee_id)
            co_ids.add(a.company_id)

        emp_map = {}
        if emp_ids:
            from app.models.staff import StaffEmployee as _SE
            emps = db.query(_SE).filter(_SE.id.in_(list(emp_ids))).all()
            emp_map = {e.id: e.full_name for e in emps}

        co_map = {}
        if co_ids:
            from app.models.staff_accounts import AssociatedCompany as _AC
            cos = db.query(_AC).filter(_AC.id.in_(list(co_ids))).all()
            co_map = {c.id: (c.company_name if hasattr(c, 'company_name') else getattr(c, 'name', f'Company {c.id}')) for c in cos}

        result = []
        for a in allocations:
            amt = float(a.amount) if a.amount is not None else 0.0
            bal = float(a.balance_remaining) if a.balance_remaining is not None else amt
            to_name = emp_map.get(a.to_employee_id, '')
            from_name = emp_map.get(a.from_employee_id, '')
            d = {
                'id': a.id,
                'allocation_number': a.allocation_number,
                'company_id': a.company_id,
                'company_name': co_map.get(a.company_id, ''),
                'segment_id': a.segment_id,
                'from_employee_id': a.from_employee_id,
                'from_employee_name': from_name,
                'to_employee_id': a.to_employee_id,
                'to_employee_name': to_name,
                'recipient_name': to_name,
                'allocation_date': a.allocation_date.isoformat() if a.allocation_date else None,
                'amount': amt,
                'purpose': a.purpose,
                'category_id': a.category_id,
                'payment_mode': a.payment_mode,
                'payment_reference': a.payment_reference,
                'bank_account_id': a.bank_account_id,
                'status': a.status,
                'balance_remaining': bal,
                'balance_used': round(amt - bal, 2),
                'total_expensed': float(a.total_expensed) if a.total_expensed is not None else 0.0,
                'settlement_date': a.settlement_date.isoformat() if a.settlement_date else None,
                'settlement_remarks': a.settlement_remarks,
                'confirmed_by_id': a.confirmed_by_id,
                'confirmed_at': a.confirmed_at.isoformat() if a.confirmed_at else None,
                'ledger_entry_id': a.ledger_entry_id,
                'created_at': a.created_at.isoformat() if a.created_at else None,
                'updated_at': a.updated_at.isoformat() if a.updated_at else None,
            }
            result.append(d)

        total_pages = max(1, (total + page_size - 1) // page_size)
        return JSONResponse(content={
            'success': True,
            'allocations': result,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/fund-allocations/{allocation_id}", response_model=FundAllocationResponse)
async def get_fund_allocation_endpoint(
    allocation_id: int = Path(..., description="Fund allocation ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get a fund allocation by ID
    DC_SFMS_001: VGK/EA see all, others see only their allocations (ownership enforced in service)
    """
    try:
        allocation = get_fund_allocation(db, current_user, allocation_id)
        return FundAllocationResponse.model_validate(allocation)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/fund-allocations/{allocation_id}/confirm", response_model=FundAllocationResponse)
async def confirm_fund_allocation_endpoint(
    allocation_id: int = Path(..., description="Fund allocation ID"),
    data: FundAllocationConfirm = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Confirm receipt of fund allocation (recipient only)
    DC_SFMS_001: Creates party ledger credit entry (recipient ownership enforced in service)
    """
    try:
        remarks = data.confirmation_remarks if data else None
        allocation = confirm_fund_allocation(db, current_user, allocation_id, remarks)
        return FundAllocationResponse.model_validate(allocation)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/fund-allocations/{allocation_id}/settle", response_model=FundAllocationResponse)
async def settle_fund_allocation_endpoint(
    allocation_id: int = Path(..., description="Fund allocation ID"),
    data: FundAllocationSettle = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Settle a fund allocation
    DC_SFMS_001: Allocator/recipient/VGK/EA can settle (ownership enforced in service)
    """
    try:
        remarks = data.settlement_remarks if data else None
        amount = data.settlement_amount if data else None
        allocation = settle_fund_allocation(db, current_user, allocation_id, remarks, amount)
        return FundAllocationResponse.model_validate(allocation)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/fund-allocations/{allocation_id}/cancel", response_model=FundAllocationResponse)
async def cancel_fund_allocation_endpoint(
    allocation_id: int = Path(..., description="Fund allocation ID"),
    data: FundAllocationCancel = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Cancel a fund allocation with optional ledger reversal
    DC_SFMS_001: VGK/EA/Allocator access only (ownership enforced in service)
    """
    try:
        reason = data.cancellation_reason if data else "No reason provided"
        allocation = cancel_fund_allocation(db, current_user, allocation_id, reason)
        return FundAllocationResponse.model_validate(allocation)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== EXPENSE ENTRY ENDPOINTS (Phase 2.16) ====================

@router.post("/expense-entries", response_model=ExpenseEntryResponse, status_code=201)
async def create_expense_entry_endpoint(
    data: ExpenseEntryCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new expense entry
    DC_SFMS_001: Any authenticated staff can create (fund allocation balance validated in service)
    """
    try:
        expense = create_expense_entry(db, current_user, data.model_dump())
        return ExpenseEntryResponse.model_validate(expense)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/expense-entries", response_model=ExpenseEntryListResponse)
async def list_expense_entries_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    status: Optional[str] = Query(None, description="Filter by status"),
    fund_allocation_id: Optional[int] = Query(None, description="Filter by fund allocation"),
    main_category_id: Optional[int] = Query(None, description="Filter by main category"),
    from_date: Optional[date] = Query(None, description="Filter from date"),
    to_date: Optional[date] = Query(None, description="Filter to date"),
    team_view: bool = Query(False, description="Show team members' expenses (reporting managers only)"),
    employee_id: Optional[int] = Query(None, description="Filter by specific employee (team_view only)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List expense entries with filters
    DC_SFMS_001: VGK/EA see all, others see only their entries (ownership enforced in service)
    When team_view=true, reporting managers see their direct reports' expenses
    employee_id: further narrow team_view to a single employee (Accounts staff only)
    """
    try:
        filters = {
            'company_id': company_id,
            'status': status,
            'fund_allocation_id': fund_allocation_id,
            'main_category_id': main_category_id,
            'from_date': from_date,
            'to_date': to_date,
            'team_view': team_view,
            'employee_id': employee_id,
            'page': page,
            'page_size': page_size
        }
        entries, total, summary = list_expense_entries(db, current_user, filters)
        from app.schemas.staff_accounts import ExpenseEntrySummary
        return ExpenseEntryListResponse(
            success=True,
            entries=[ExpenseEntryResponse.model_validate(e) for e in entries],
            total=total,
            page=page,
            page_size=page_size,
            summary=ExpenseEntrySummary(**summary)
        )
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/expense-entries/{entry_id}", response_model=ExpenseEntryResponse)
async def get_expense_entry_endpoint(
    entry_id: int = Path(..., description="Expense entry ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get an expense entry by ID
    DC_SFMS_001: VGK/EA see all, others see only their entries (ownership enforced in service)
    """
    try:
        expense = get_expense_entry(db, current_user, entry_id)
        return ExpenseEntryResponse.model_validate(expense)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/expense-entries/{entry_id}/show-in-ledger", response_model=ExpenseEntryResponse)
async def toggle_expense_show_in_ledger_endpoint(
    entry_id: int = Path(..., description="Expense entry ID"),
    show_in_ledger: bool = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-SHOW-IN-LEDGER-001: toggle whether this expense entry posts to the transaction ledger.
    Optional flag, editable at any status.
    """
    try:
        from app.services.staff_accounts_service import toggle_expense_show_in_ledger
        expense = toggle_expense_show_in_ledger(db, current_user, entry_id, show_in_ledger)
        return ExpenseEntryResponse.model_validate(expense)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/expense-entries/{entry_id}/submit", response_model=ExpenseEntryResponse)
async def submit_expense_entry_endpoint(
    entry_id: int = Path(..., description="Expense entry ID"),
    data: ExpenseEntrySubmit = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Submit expense entry for approval
    DC_SFMS_001: Only creator can submit (ownership enforced in service)
    """
    try:
        remarks = data.submission_remarks if data else None
        expense = submit_expense_entry(db, current_user, entry_id, remarks)
        return ExpenseEntryResponse.model_validate(expense)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/expense-entries/{entry_id}/approve", response_model=ExpenseEntryResponse)
async def approve_expense_entry_endpoint(
    data: ExpenseEntryApprove,
    entry_id: int = Path(..., description="Expense entry ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Approve/Reject/Return expense entry
    DC_SFMS_001: VGK/EA only (role enforced in service), updates fund allocation if linked
    """
    try:
        expense = approve_expense_entry(db, current_user, entry_id, data.action, data.remarks)
        return ExpenseEntryResponse.model_validate(expense)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== EXPENSE ENTRY EDIT + TALLY ACTION ENDPOINTS ====================

@router.put("/expense-entries/{entry_id}/edit")
async def edit_expense_entry_endpoint(
    entry_id: int = Path(...),
    data: ExpenseEntryUpdate = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Edit an expense entry (amount, category, notes).
    Accounts/VEA roles only. Only DRAFT or SUBMITTED entries may be edited.
    """
    try:
        from app.models.staff_accounts import ExpenseEntry as _EE
        from app.models.expense_category import ExpenseMainCategory as _MC, ExpenseSubCategory as _SC

        if not is_accounts_allowed_employee(current_user):
            return JSONResponse(status_code=403, content={"success": False, "detail": "Accounts/VEA access required"})

        entry = db.query(_EE).filter(_EE.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Expense entry not found"})
        if entry.status not in ('DRAFT', 'SUBMITTED'):
            return JSONResponse(status_code=400, content={"success": False, "detail": f"Cannot edit entry with status '{entry.status}'"})

        update_fields = data.model_dump(exclude_unset=True) if data else {}
        for field, value in update_fields.items():
            if hasattr(entry, field):
                setattr(entry, field, value)
        from datetime import datetime as _dt
        entry.updated_at = _dt.utcnow()
        db.commit()
        db.refresh(entry)
        return JSONResponse(content={"success": True, "message": "Expense entry updated", "entry_id": entry.id})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/expense-entries/{entry_id}/mark-paid", response_model=ExpenseEntryResponse)
async def mark_expense_paid_endpoint(
    entry_id: int = Path(..., description="Expense entry ID"),
    data: ExpenseEntryMarkPaid = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Mark expense as paid — auto-approves (DRAFT/SUBMITTED → APPROVED + is_paid=True).
    DC Protocol: Paid = auto-approved. Finance/Accounts/VGK/EA only.
    Records payment_utr, bank_account_id, paid_at, paid_by_id.
    """
    try:
        payment_utr = data.payment_utr if data else None
        bank_account_id = data.bank_account_id if data else None
        notes = data.notes if data else None
        expense = mark_expense_paid(db, current_user, entry_id, payment_utr, bank_account_id, notes)
        return ExpenseEntryResponse.model_validate(expense)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.patch("/expense-entries/{entry_id}/tally-action")
async def expense_tally_action_endpoint(
    entry_id: int = Path(...),
    action: str = Body(..., embed=True, description="confirm | exception | tally_done"),
    notes: Optional[str] = Body(None, embed=True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Tally/accounts action on an expense entry.
    confirm   → status=APPROVED + EmployeeFundLedger debit entry created
    exception → tally_status=EXCLUDED (no ledger entry)
    tally_done → tally_status=SYNCED
    Only Accounts/VEA roles.
    """
    try:
        from app.models.staff_accounts import ExpenseEntry as _EE
        from decimal import Decimal as _Dec
        from datetime import datetime as _dt

        if not is_accounts_allowed_employee(current_user):
            return JSONResponse(status_code=403, content={"success": False, "detail": "Accounts/VEA access required"})

        entry = db.query(_EE).filter(_EE.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Expense entry not found"})

        now = _dt.utcnow()
        action = action.lower()

        if action == 'confirm':
            entry.status = 'APPROVED'
            entry.updated_at = now
            db.commit()
            db.refresh(entry)

            # Determine employee to debit: prefer fund_allocation recipient, fallback to creator
            emp_to_debit = None
            if entry.fund_allocation_id:
                alloc = db.query(FundAllocation).filter(FundAllocation.id == entry.fund_allocation_id).first()
                if alloc:
                    emp_to_debit = alloc.to_employee_id
            if not emp_to_debit:
                emp_to_debit = entry.created_by_id

            # Create EmployeeFundLedger debit entry only on APPROVED
            if emp_to_debit:
                last_ledger = db.query(EmployeeFundLedger).filter(
                    EmployeeFundLedger.employee_id == emp_to_debit
                ).order_by(EmployeeFundLedger.id.desc()).first()
                running_balance = _Dec(str(last_ledger.balance)) if last_ledger else _Dec('0')
                new_balance = running_balance - _Dec(str(entry.amount))

                ledger_entry = EmployeeFundLedger(
                    employee_id=emp_to_debit,
                    company_id=entry.company_id,
                    transaction_date=entry.expense_date,
                    entry_type='EXPENSE_MADE',
                    reference_type='EXPENSE_ENTRY',
                    reference_id=entry.id,
                    reference_number=entry.entry_number,
                    debit_amount=_Dec(str(entry.amount)),
                    credit_amount=_Dec('0'),
                    balance=new_balance,
                    narration=notes or f"Expense approved: {entry.entry_number}",
                    updated_by_id=current_user.id
                )
                db.add(ledger_entry)
                db.commit()

        elif action == 'exception':
            entry.tally_status = 'EXCLUDED'
            entry.updated_at = now
            db.commit()

        elif action == 'tally_done':
            entry.tally_status = 'SYNCED'
            entry.updated_at = now
            db.commit()

        else:
            return JSONResponse(status_code=400, content={"success": False, "detail": "Invalid action. Use confirm, exception, or tally_done"})

        return JSONResponse(content={"success": True, "message": f"Action '{action}' applied", "entry_id": entry_id})
    except Exception as e:
        return handle_accounts_error(e)


# ==================== FUND LEDGER ENDPOINTS ====================

@router.post("/fund-ledger/opening-balance")
async def add_fund_opening_balance(
    employee_id: Optional[int] = Body(None, embed=False),
    company_id: int = Body(..., embed=False),
    amount: float = Body(..., embed=False),
    balance_date: Optional[str] = Body(None, embed=False),
    narration: Optional[str] = Body(None, embed=False),
    ledger_type: str = Body("EMPLOYEE", embed=False, description="EMPLOYEE or COMPANY"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add opening balance for an employee fund ledger or company account ledger.
    Accounts/VEA roles only.
    """
    try:
        from decimal import Decimal as _Dec
        from datetime import datetime as _dt, date as _date
        from sqlalchemy import text as _text

        if not is_accounts_allowed_employee(current_user):
            return JSONResponse(status_code=403, content={"success": False, "detail": "Accounts/VEA access required"})

        txn_date = _date.today()
        if balance_date:
            txn_date = _dt.strptime(balance_date, '%Y-%m-%d').date()

        amt = _Dec(str(amount))

        if ledger_type == 'EMPLOYEE':
            if not employee_id:
                return JSONResponse(status_code=400, content={"success": False, "detail": "employee_id required for EMPLOYEE ledger"})
            emp = db.query(StaffEmployee).filter(StaffEmployee.id == employee_id).first()
            if not emp:
                return JSONResponse(status_code=404, content={"success": False, "detail": "Employee not found"})

            existing = db.query(EmployeeFundLedger).filter(
                EmployeeFundLedger.employee_id == employee_id,
                EmployeeFundLedger.entry_type == 'OPENING_BALANCE'
            ).first()
            if existing:
                existing.credit_amount = amt
                existing.balance = amt
                existing.narration = narration or "Opening balance updated"
                existing.updated_by_id = current_user.id
                db.commit()
                return JSONResponse(content={"success": True, "message": "Opening balance updated", "employee_id": employee_id})

            ledger_entry = EmployeeFundLedger(
                employee_id=employee_id,
                company_id=company_id,
                transaction_date=txn_date,
                entry_type='OPENING_BALANCE',
                reference_type='OPENING_BALANCE',
                reference_id=0,
                reference_number='OB-' + str(employee_id),
                debit_amount=_Dec('0'),
                credit_amount=amt,
                balance=amt,
                narration=narration or "Opening balance",
                updated_by_id=current_user.id
            )
            db.add(ledger_entry)
            db.commit()
            return JSONResponse(content={"success": True, "message": "Opening balance added", "employee_id": employee_id, "amount": float(amt)})

        elif ledger_type == 'COMPANY':
            existing_row = db.execute(_text("""
                SELECT id FROM company_account_ledger
                WHERE company_id = :cid AND entry_type = 'OPENING_BALANCE'
            """), {"cid": company_id}).fetchone()

            if existing_row:
                db.execute(_text("""
                    UPDATE company_account_ledger
                    SET credit_amount = :amt, balance = :amt, narration = :narr, updated_at = NOW()
                    WHERE id = :rid
                """), {"amt": float(amt), "narr": narration or "Opening balance updated", "rid": existing_row[0]})
                db.commit()
                return JSONResponse(content={"success": True, "message": "Company opening balance updated", "company_id": company_id})

            db.execute(_text("""
                INSERT INTO company_account_ledger
                (company_id, transaction_date, entry_type, reference_type, credit_amount, debit_amount, balance, narration, created_by_id, created_at, updated_at)
                VALUES (:cid, :txn_date, 'OPENING_BALANCE', 'OPENING_BALANCE', :amt, 0, :amt, :narr, :uid, NOW(), NOW())
            """), {"cid": company_id, "txn_date": txn_date, "amt": float(amt), "narr": narration or "Opening balance", "uid": current_user.id})
            db.commit()
            return JSONResponse(content={"success": True, "message": "Company opening balance added", "company_id": company_id, "amount": float(amt)})

        return JSONResponse(status_code=400, content={"success": False, "detail": "ledger_type must be EMPLOYEE or COMPANY"})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/fund-ledger/{employee_id}/balance")
async def get_employee_fund_balance(
    employee_id: int = Path(...),
    company_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get full balance breakdown for an employee:
    - total_received (all credits)
    - approved_expenses (APPROVED expense entries)
    - pending_submitted (SUBMITTED expenses not yet approved)
    - pending_draft (DRAFT expenses)
    - available_balance (ledger running balance)
    """
    try:
        from sqlalchemy import func as _func
        from decimal import Decimal as _Dec

        emp = db.query(StaffEmployee).filter(StaffEmployee.id == employee_id).first()
        if not emp:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Employee not found"})

        last_ledger = db.query(EmployeeFundLedger).filter(
            EmployeeFundLedger.employee_id == employee_id
        ).order_by(EmployeeFundLedger.id.desc()).first()
        available_balance = float(last_ledger.balance) if last_ledger else 0.0

        total_credits = db.query(_func.coalesce(_func.sum(EmployeeFundLedger.credit_amount), 0)).filter(
            EmployeeFundLedger.employee_id == employee_id
        ).scalar() or 0

        total_debits = db.query(_func.coalesce(_func.sum(EmployeeFundLedger.debit_amount), 0)).filter(
            EmployeeFundLedger.employee_id == employee_id
        ).scalar() or 0

        from app.models.staff_accounts import ExpenseEntry as _EE, FundAllocation as _FA
        from sqlalchemy import or_ as _or2

        def _expense_sum_for_employee(status_filter):
            # Expenses linked to employee via fund allocation OR directly created by them
            q_alloc = db.query(_func.coalesce(_func.sum(_EE.amount), 0)).join(
                _FA, _EE.fund_allocation_id == _FA.id, isouter=True
            ).filter(
                _or2(
                    _FA.to_employee_id == employee_id,
                    _EE.created_by_id == employee_id
                ),
                _EE.status == status_filter
            )
            if company_id:
                q_alloc = q_alloc.filter(_EE.company_id == company_id)
            return float(q_alloc.scalar() or 0)

        approved_amount = _expense_sum_for_employee('APPROVED')
        submitted_amount = _expense_sum_for_employee('SUBMITTED')
        draft_amount = _expense_sum_for_employee('DRAFT')

        emp_name = f"{getattr(emp, 'first_name', '')} {getattr(emp, 'last_name', '')}".strip()

        # DC_OB_PREFILL_001: Fetch existing opening balance entry so UI can pre-fill it
        ob_entry = db.query(EmployeeFundLedger).filter(
            EmployeeFundLedger.employee_id == employee_id,
            EmployeeFundLedger.entry_type == 'OPENING_BALANCE'
        ).first()
        opening_balance_amount = float(ob_entry.credit_amount) if ob_entry else None

        return JSONResponse(content={
            "success": True,
            "employee_id": employee_id,
            "employee_name": emp_name,
            "opening_balance": opening_balance_amount,
            "balance_summary": {
                "total_received": float(total_credits),
                "total_debited": float(total_debits),
                "available_balance": available_balance,
                "approved_expenses": approved_amount,
                "pending_submitted": submitted_amount,
                "pending_draft": draft_amount,
                "effective_balance_after_submitted": round(available_balance - submitted_amount, 2),
                "effective_balance_after_all_pending": round(available_balance - submitted_amount - draft_amount, 2)
            }
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/fund-ledger/{employee_id}/statement")
async def get_employee_fund_statement(
    employee_id: int = Path(...),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Full ledger statement for an employee with running balance.
    """
    try:
        emp = db.query(StaffEmployee).filter(StaffEmployee.id == employee_id).first()
        if not emp:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Employee not found"})

        q = db.query(EmployeeFundLedger).filter(EmployeeFundLedger.employee_id == employee_id)
        if from_date:
            q = q.filter(EmployeeFundLedger.transaction_date >= from_date)
        if to_date:
            q = q.filter(EmployeeFundLedger.transaction_date <= to_date)

        total = q.count()
        entries = q.order_by(EmployeeFundLedger.id.desc()).offset((page - 1) * page_size).limit(page_size).all()

        rows = []
        for e in entries:
            rows.append({
                "id": e.id,
                "transaction_date": str(e.transaction_date),
                "entry_type": e.entry_type,
                "reference_type": e.reference_type,
                "reference_number": e.reference_number,
                "debit_amount": float(e.debit_amount),
                "credit_amount": float(e.credit_amount),
                "balance": float(e.balance),
                "narration": e.narration
            })

        return JSONResponse(content={"success": True, "entries": rows, "total": total, "page": page, "page_size": page_size})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/fund-transfers")
async def create_fund_transfer_endpoint(
    from_employee_id: int = Body(..., embed=False),
    to_employee_id: int = Body(..., embed=False),
    company_id: int = Body(..., embed=False),
    amount: float = Body(..., embed=False),
    purpose: Optional[str] = Body(None, embed=False),
    payment_mode: Optional[str] = Body(None, embed=False),
    payment_reference: Optional[str] = Body(None, embed=False),
    transfer_date: Optional[str] = Body(None, embed=False),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Transfer funds from one employee to another.
    Creates EmployeeFundTransfer + two EmployeeFundLedger entries (debit sender, credit recipient).
    """
    try:
        from decimal import Decimal as _Dec
        from datetime import datetime as _dt, date as _date

        if from_employee_id == to_employee_id:
            return JSONResponse(status_code=400, content={"success": False, "detail": "Cannot transfer to self"})

        amt = _Dec(str(amount))
        if amt <= 0:
            return JSONResponse(status_code=400, content={"success": False, "detail": "Amount must be positive"})

        txn_date = _date.today()
        if transfer_date:
            txn_date = _dt.strptime(transfer_date, '%Y-%m-%d').date()

        from_emp = db.query(StaffEmployee).filter(StaffEmployee.id == from_employee_id).first()
        to_emp = db.query(StaffEmployee).filter(StaffEmployee.id == to_employee_id).first()
        if not from_emp or not to_emp:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Employee not found"})

        from sqlalchemy import func as _func
        transfer_count = db.query(_func.count(EmployeeFundTransfer.id)).scalar() or 0
        transfer_number = f"FT-{str(transfer_count + 1).zfill(5)}"

        transfer = EmployeeFundTransfer(
            transfer_number=transfer_number,
            from_employee_id=from_employee_id,
            to_employee_id=to_employee_id,
            company_id=company_id,
            transfer_date=txn_date,
            amount=amt,
            purpose=purpose,
            payment_mode=payment_mode,
            payment_reference=payment_reference,
            status='CONFIRMED',
            confirmed_by_id=current_user.id,
            confirmed_at=_dt.utcnow()
        )
        db.add(transfer)
        db.flush()

        # Debit sender ledger
        last_from = db.query(EmployeeFundLedger).filter(
            EmployeeFundLedger.employee_id == from_employee_id
        ).order_by(EmployeeFundLedger.id.desc()).first()
        from_balance = _Dec(str(last_from.balance)) if last_from else _Dec('0')
        from_entry = EmployeeFundLedger(
            employee_id=from_employee_id,
            company_id=company_id,
            transaction_date=txn_date,
            entry_type='TRANSFER_SENT',
            reference_type='FUND_TRANSFER',
            reference_id=transfer.id,
            reference_number=transfer_number,
            debit_amount=amt,
            credit_amount=_Dec('0'),
            balance=from_balance - amt,
            narration=purpose or f"Transfer to {getattr(to_emp, 'first_name', '')} {getattr(to_emp, 'last_name', '')}".strip(),
            updated_by_id=current_user.id
        )
        db.add(from_entry)
        db.flush()

        # Credit recipient ledger
        last_to = db.query(EmployeeFundLedger).filter(
            EmployeeFundLedger.employee_id == to_employee_id
        ).order_by(EmployeeFundLedger.id.desc()).first()
        to_balance = _Dec(str(last_to.balance)) if last_to else _Dec('0')
        to_entry = EmployeeFundLedger(
            employee_id=to_employee_id,
            company_id=company_id,
            transaction_date=txn_date,
            entry_type='TRANSFER_RECEIVED',
            reference_type='FUND_TRANSFER',
            reference_id=transfer.id,
            reference_number=transfer_number,
            debit_amount=_Dec('0'),
            credit_amount=amt,
            balance=to_balance + amt,
            narration=purpose or f"Transfer from {getattr(from_emp, 'first_name', '')} {getattr(from_emp, 'last_name', '')}".strip(),
            updated_by_id=current_user.id
        )
        db.add(to_entry)
        db.flush()

        transfer.from_ledger_entry_id = from_entry.id
        transfer.to_ledger_entry_id = to_entry.id
        db.commit()

        return JSONResponse(content={
            "success": True,
            "message": "Transfer completed",
            "transfer_number": transfer_number,
            "from_new_balance": float(from_balance - amt),
            "to_new_balance": float(to_balance + amt)
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/fund-transfers")
async def list_fund_transfers_endpoint(
    employee_id: Optional[int] = Query(None, description="Filter by from or to employee"),
    from_employee_id: Optional[int] = Query(None),
    to_employee_id: Optional[int] = Query(None),
    company_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List fund transfers with filters."""
    try:
        from sqlalchemy import or_ as _or

        q = db.query(EmployeeFundTransfer)
        if employee_id:
            q = q.filter(_or(EmployeeFundTransfer.from_employee_id == employee_id, EmployeeFundTransfer.to_employee_id == employee_id))
        if from_employee_id:
            q = q.filter(EmployeeFundTransfer.from_employee_id == from_employee_id)
        if to_employee_id:
            q = q.filter(EmployeeFundTransfer.to_employee_id == to_employee_id)
        if company_id:
            q = q.filter(EmployeeFundTransfer.company_id == company_id)
        if from_date:
            q = q.filter(EmployeeFundTransfer.transfer_date >= from_date)
        if to_date:
            q = q.filter(EmployeeFundTransfer.transfer_date <= to_date)

        total = q.count()
        transfers = q.order_by(EmployeeFundTransfer.id.desc()).offset((page - 1) * page_size).limit(page_size).all()

        emp_ids = set()
        for t in transfers:
            emp_ids.add(t.from_employee_id)
            emp_ids.add(t.to_employee_id)

        emps = db.query(StaffEmployee).filter(StaffEmployee.id.in_(emp_ids)).all()
        emp_map = {e.id: f"{getattr(e, 'first_name', '')} {getattr(e, 'last_name', '')}".strip() for e in emps}

        rows = []
        for t in transfers:
            rows.append({
                "id": t.id,
                "transfer_number": t.transfer_number,
                "from_employee_id": t.from_employee_id,
                "from_employee_name": emp_map.get(t.from_employee_id, ''),
                "to_employee_id": t.to_employee_id,
                "to_employee_name": emp_map.get(t.to_employee_id, ''),
                "transfer_date": str(t.transfer_date),
                "amount": float(t.amount),
                "purpose": t.purpose,
                "payment_mode": t.payment_mode,
                "status": t.status
            })

        return JSONResponse(content={"success": True, "transfers": rows, "total": total, "page": page, "page_size": page_size})
    except Exception as e:
        return handle_accounts_error(e)


# ==================== BALANCE SHEET DASHBOARD ENDPOINTS (Phase 2.17) ====================

@router.post("/balance-sheet/compute", response_model=BalanceSheetSummaryResponse)
async def compute_balance_sheet_endpoint(
    data: BalanceSheetComputeRequest,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Compute or recompute balance sheet summary for a company
    DC_SFMS_001: VGK/EA only - aggregates income, expenses, liabilities
    """
    try:
        summary = compute_balance_sheet(
            db, current_user,
            data.company_id, data.period_type, data.period_date, data.force_recompute
        )
        return BalanceSheetSummaryResponse.model_validate(summary)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/consolidated-overview")
async def get_consolidated_overview_endpoint(
    company_id: Optional[List[int]] = Query(None, description="Company IDs (repeat param for multiple)"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    period_type: str = Query('OVERALL', description="OVERALL, DAILY, MONTHLY, QUARTERLY, YEARLY"),
    sort_by: str = Query('company_name', description="company_name, balance, income, expense"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import get_consolidated_overview
        data = get_consolidated_overview(
            db, current_user,
            company_ids=company_id or None,
            date_from=date_from,
            date_to=date_to,
            period_type=period_type,
            sort_by=sort_by
        )
        return JSONResponse(content={'success': True, **data})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/balance-sheet/aggregate")
async def get_aggregate_balance_sheet(
    company_id: Optional[int] = Query(None, description="Filter by company (omit for all companies)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get cumulative as-on-date balance sheet, optionally for a specific company.
    No date-period restriction — aggregates all data up to today.
    """
    try:
        from app.services.staff_accounts_service import get_aggregate_balance_summary
        summary = get_aggregate_balance_summary(db, current_user, company_id)

        party_balances = []
        try:
            from app.models.staff_accounts import PartyLedger
            from sqlalchemy import func as _func
            q = db.query(
                PartyLedger.party_name,
                PartyLedger.party_type,
                _func.coalesce(_func.sum(PartyLedger.debit_amount), 0).label('total_debit'),
                _func.coalesce(_func.sum(PartyLedger.credit_amount), 0).label('total_credit'),
            )
            if company_id:
                q = q.filter(PartyLedger.company_id == company_id)
            rows = q.group_by(PartyLedger.party_name, PartyLedger.party_type).order_by(PartyLedger.party_name).all()
            party_balances = [
                {
                    'party_name': r.party_name,
                    'party_type': r.party_type,
                    'total_debit': float(r.total_debit),
                    'total_credit': float(r.total_credit),
                    'balance': float(r.total_debit - r.total_credit),
                }
                for r in rows
            ]
        except Exception:
            pass

        alerts = []
        try:
            alerts = get_dashboard_alerts(db, current_user, company_id or 0)
        except Exception:
            pass

        return JSONResponse(content={
            'success': True,
            'summary': summary,
            'party_balances': party_balances,
            'alerts': alerts,
            'trend_data': [],
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/balance-sheet/{company_id}", response_model=BalanceSheetDashboardResponse)
async def get_balance_sheet_dashboard_endpoint(
    company_id: int = Path(..., description="Company ID"),
    period_type: str = Query("MONTHLY", description="Period type: DAILY, MONTHLY, QUARTERLY, YEARLY"),
    period_date: date = Query(None, description="Period date (defaults to today)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get balance sheet dashboard with summary, trends, and alerts
    DC_SFMS_001: VGK/EA only - comprehensive financial dashboard
    """
    try:
        from datetime import date as date_type
        if not period_date:
            from app.services.staff_accounts_service import get_indian_time
            period_date = get_indian_time().date()
        
        summary = compute_balance_sheet(
            db, current_user, company_id, period_type, period_date, True
        )
        
        trends = get_balance_sheet_trend(db, current_user, company_id, period_type, 6)
        
        trend_data = []
        for t in trends:
            trend_data.append({
                "period_date": str(t.period_date),
                "total_income": str(t.total_income),
                "total_expense": str(t.total_expense),
                "net_balance": str(t.net_balance),
                "available_balance": str(t.available_balance)
            })
        
        alerts = get_dashboard_alerts(db, current_user, company_id)
        
        return BalanceSheetDashboardResponse(
            success=True,
            summary=BalanceSheetSummaryResponse.model_validate(summary),
            trend_data=trend_data,
            alerts=alerts
        )
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/balance-sheet/{company_id}/trend")
async def get_balance_sheet_trend_endpoint(
    company_id: int = Path(..., description="Company ID"),
    period_type: str = Query("MONTHLY", description="Period type"),
    periods: int = Query(6, ge=1, le=24, description="Number of periods"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get historical balance sheet trend
    DC_SFMS_001: VGK/EA only - returns last N periods for trend analysis
    """
    try:
        trends = get_balance_sheet_trend(db, current_user, company_id, period_type, periods)
        
        return {
            "success": True,
            "company_id": company_id,
            "period_type": period_type,
            "trends": [BalanceSheetSummaryResponse.model_validate(t) for t in trends]
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== ACCOUNTS PAYABLE ENDPOINTS (DC_CREDIT_001) ====================

@router.get("/payables-summary")
async def list_payables_summary_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company (0=all)"),
    as_on_date: Optional[str] = Query(None, description="Deprecated: use from_date/to_date"),
    from_date: Optional[str] = Query(None, description="Period start YYYY-MM-DD (omit = all time)"),
    to_date: Optional[str] = Query(None, description="Period end YYYY-MM-DD (omit = today)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-AP-SUMMARY-002: Payables summary from party_ledger using Party Ledger's exact formula.
    Groups by canonical vendor name (strips LLP/PVT LTD suffixes) to combine fragmented entries.
    Uses date range (from_date..to_date) matching Party Ledger period behaviour.
    OB (OPENING_BALANCE entries) treated as pre-period base; period activity = non-OB entries.
    Closing = OB_net + period_debit - period_credit  (positive = vendor owes us / debit balance).
    AP shows vendors where closing > 0 (we owe them = credit balance convention: credit > debit).
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        # Support legacy as_on_date for backwards compat (treat as to_date)
        _to = to_date or as_on_date
        date_to   = _date.fromisoformat(_to) if _to else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "vendors": [], "total_outstanding": "0.00", "company_count": 0}

        placeholders = ",".join(str(c) for c in co_ids)
        from sqlalchemy import text as _text

        # DC-AP-CANONICAL-001: canonical name strips common legal suffixes so that
        # "MYNTREAL" and "MYNTREAL LLP" merge into one row — same as Party Ledger ILIKE search.
        _canon_expr = """
            UPPER(TRIM(regexp_replace(
                pl.party_name,
                '\\s+(LLP|PVT\\.?\\s*LTD\\.?|PVT|LTD|PRIVATE LIMITED|LIMITED|INDIA|&\\s*CO\\.?)$',
                '', 'gi'
            )))
        """

        # DC-AP-OB-001: Opening Balance = SUM(credit - debit) of OPENING_BALANCE reference rows.
        # These represent the vendor's pre-period (or brought-forward) balance.
        # For a date-range view, OB = last running_balance before date_from (Party Ledger style).
        # For all-time view, OB = sum of all OPENING_BALANCE credit rows (net of any OB debits).

        # DC-AP-ALL-TYPES-001: Mirror Party Ledger "All Types" name-search behaviour.
        # When Party Ledger is searched by party_name (ILIKE), it returns rows of ANY party_type
        # (VENDOR, COMPANY, PARTNER, etc.) that match.  The closing balance therefore includes
        # credits/debits posted under the same canonical name regardless of party_type.
        # AP must do the same: group ALL party_type rows by canonical name so the total matches.
        # Only OPENING_BALANCE rows come exclusively from VENDOR-type entries (by convention),
        # so we sum OB across all types as well to be safe.

        if date_from:
            # DC-AP-PERIOD-003: Period view using same formula as Overall.
            # OB = SUM(credit-debit) for entries strictly before date_from
            #    + SUM(credit-debit) of OPENING_BALANCE entries anywhere in [date_from, date_to]
            #      (OPENING_BALANCE entries on e.g. Apr-1 are carried-forward balances, not
            #       period activity; they belong in OB regardless of whether date_from is Mar-31
            #       or Apr-1).
            # Period activity = non-OPENING_BALANCE entries in [date_from, date_to].
            # closing = ob_cr_net + period_deb - period_crd
            # Payable = closing > 0  (same sign convention as Overall branch).
            ob_rows = db.execute(_text(f"""
                SELECT
                    {_canon_expr} AS canon,
                    pl.company_id,
                    SUM(pl.credit_amount - pl.debit_amount) AS ob_cr_net
                FROM party_ledger pl
                WHERE pl.company_id IN ({placeholders})
                  AND (
                    pl.transaction_date < :date_from
                    OR (pl.reference_type = 'OPENING_BALANCE'
                        AND pl.transaction_date >= :date_from
                        AND pl.transaction_date <= :date_to)
                  )
                GROUP BY {_canon_expr}, pl.company_id
            """), {"date_from": date_from, "date_to": date_to}).fetchall()
            ob_map = {(r.canon, r.company_id): float(r.ob_cr_net) for r in ob_rows}

            # Period activity: non-OPENING_BALANCE entries in [date_from, date_to]
            period_rows = db.execute(_text(f"""
                SELECT
                    {_canon_expr} AS canon,
                    pl.company_id,
                    MAX(ac.company_name) AS company_name,
                    SUM(CASE WHEN pl.reference_type != 'OPENING_BALANCE' THEN pl.debit_amount ELSE 0 END)  AS period_deb,
                    SUM(CASE WHEN pl.reference_type != 'OPENING_BALANCE' THEN pl.credit_amount ELSE 0 END) AS period_crd
                FROM party_ledger pl
                JOIN associated_companies ac ON ac.id = pl.company_id
                WHERE pl.company_id IN ({placeholders})
                  AND pl.transaction_date >= :date_from
                  AND pl.transaction_date <= :date_to
                GROUP BY {_canon_expr}, pl.company_id
            """), {"date_from": date_from, "date_to": date_to}).fetchall()

            seen = set()
            vendor_rows = []
            for r in period_rows:
                seen.add((r.canon, r.company_id))
                ob_cr = ob_map.get((r.canon, r.company_id), 0.0)
                # closing = ob_credit_net + period_deb - period_crd
                # Payable = closing > 0 (mirrors Overall formula exactly)
                closing = ob_cr + float(r.period_deb or 0) - float(r.period_crd or 0)
                if closing > 0:
                    vendor_rows.append({
                        "vendor_name": r.canon,
                        "company_id": r.company_id,
                        "company_name": r.company_name,
                        "net_payable": round(closing, 2),
                    })
            # Vendors with OB but zero period activity in range (fully carried-forward payable)
            for (canon, co_id), ob_cr in ob_map.items():
                if (canon, co_id) not in seen and ob_cr > 0:
                    co_name = next((c.get("company_name") or c.get("name", "") for c in companies if c["id"] == co_id), "")
                    vendor_rows.append({
                        "vendor_name": canon,
                        "company_id": co_id,
                        "company_name": co_name,
                        "net_payable": round(ob_cr, 2),
                    })
        else:
            # All-time view: OB = OPENING_BALANCE entries (credit - debit); activity = all non-OB.
            # Formula mirrors Party Ledger: closing = OB_credit_net + activity_deb - activity_crd
            # Payable = closing > 0  (net credit → we owe vendor)
            # Include ALL party_types (same as Party Ledger "All Types" ILIKE name search).
            all_rows = db.execute(_text(f"""
                SELECT
                    {_canon_expr} AS canon,
                    pl.company_id,
                    MAX(ac.company_name) AS company_name,
                    SUM(CASE WHEN pl.reference_type = 'OPENING_BALANCE'
                             THEN pl.credit_amount - pl.debit_amount ELSE 0 END) AS ob_credit_net,
                    SUM(CASE WHEN pl.reference_type != 'OPENING_BALANCE'
                             THEN pl.debit_amount ELSE 0 END)  AS activity_deb,
                    SUM(CASE WHEN pl.reference_type != 'OPENING_BALANCE'
                             THEN pl.credit_amount ELSE 0 END) AS activity_crd
                FROM party_ledger pl
                JOIN associated_companies ac ON ac.id = pl.company_id
                WHERE pl.company_id IN ({placeholders})
                  AND pl.transaction_date <= :date_to
                GROUP BY {_canon_expr}, pl.company_id
            """), {"date_to": date_to}).fetchall()

            vendor_rows = []
            for r in all_rows:
                # closing_payable = OB_credit_net + activity_deb - activity_crd
                # Positive = net credit (we owe vendor) = payable
                closing_payable = (
                    float(r.ob_credit_net or 0)
                    + float(r.activity_deb or 0)
                    - float(r.activity_crd or 0)
                )
                if closing_payable > 0:
                    vendor_rows.append({
                        "vendor_name": r.canon,
                        "company_id": r.company_id,
                        "company_name": r.company_name,
                        "net_payable": round(closing_payable, 2),
                    })

        vendor_rows.sort(key=lambda x: x["net_payable"], reverse=True)
        total = round(sum(v["net_payable"] for v in vendor_rows), 2)

        return {
            "success": True,
            "vendors": vendor_rows,
            "total_outstanding": f"{total:.2f}",
            "from_date": str(date_from) if date_from else None,
            "to_date": str(date_to),
            "company_count": len(co_ids),
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/payables")
async def list_payables_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor"),
    status: Optional[str] = Query(None, description="Filter by status: PENDING, PARTIAL_PAID, OVERDUE"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=500, description="Items per page"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List accounts payable (payment pending to vendors)
    DC_CREDIT_001: VGK/EA access - Displays vendor transactions with pending payments
    """
    try:
        payables, total, total_pending, total_overdue = AccountsCreditService.get_payables_list(
            db, current_user, company_id, vendor_id, status, page, page_size
        )
        
        return {
            "success": True,
            "payables": payables,
            "total": total,
            "total_pending": str(total_pending),
            "total_overdue": str(total_overdue),
            "page": page,
            "page_size": page_size
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/payables/record-payment")
async def record_vendor_payment_endpoint(
    data: RecordVendorPaymentRequest,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record payment to vendor
    DC_CREDIT_001: WVV Protocol - Creates payment transaction and updates vendor balance
    """
    try:
        result = AccountsCreditService.record_vendor_payment(
            db=db,
            employee=current_user,
            transaction_id=data.transaction_id,
            amount=data.amount,
            payment_date=data.payment_date,
            payment_mode=data.payment_mode,
            payment_reference=data.payment_reference,
            bank_reference=data.bank_reference,
            cheque_number=data.cheque_number,
            cheque_date=data.cheque_date,
            narration=data.narration
        )
        return result
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== ACCOUNTS RECEIVABLE ENDPOINTS (DC_CREDIT_001) ====================

@router.get("/receivables-summary")
async def list_receivables_summary_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company (0=all)"),
    as_on_date: Optional[str] = Query(None, description="Deprecated: use from_date/to_date"),
    from_date: Optional[str] = Query(None, description="Period start YYYY-MM-DD (omit = all time)"),
    to_date: Optional[str] = Query(None, description="Period end YYYY-MM-DD (omit = today)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-AR-SUMMARY-002: Receivables summary grouped by customer from sales_invoices.
    Supports date range (from_date..to_date) for period filtering matching Party Ledger behaviour.
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        _to = to_date or as_on_date
        date_to   = _date.fromisoformat(_to) if _to else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "customers": [], "total_outstanding": "0.00"}

        placeholders = ",".join(str(c) for c in co_ids)

        date_filter = "AND si.invoice_date <= :date_to"
        params: dict = {"date_to": date_to}
        if date_from:
            date_filter = "AND si.invoice_date >= :date_from AND si.invoice_date <= :date_to"
            params["date_from"] = date_from

        total_row = db.execute(_text(f"""
            SELECT COALESCE(SUM(balance_due), 0)
            FROM sales_invoices si
            WHERE si.status = 'CONFIRMED' AND si.balance_due > 0
              AND si.company_id IN ({placeholders})
              {date_filter}
        """), params).scalar() or 0

        rows = db.execute(_text(f"""
            SELECT si.customer_name,
                   si.company_id,
                   ac.company_name,
                   COUNT(*) AS invoice_count,
                   SUM(si.grand_total) AS total_billed,
                   SUM(si.amount_received) AS total_received,
                   SUM(si.balance_due) AS total_outstanding
            FROM sales_invoices si
            JOIN associated_companies ac ON ac.id = si.company_id
            WHERE si.status = 'CONFIRMED' AND si.balance_due > 0
              AND si.company_id IN ({placeholders})
              {date_filter}
            GROUP BY si.customer_name, si.company_id, ac.company_name
            ORDER BY total_outstanding DESC
        """), params).fetchall()

        customers = [
            {
                "customer_name": r.customer_name,
                "company_id": r.company_id,
                "company_name": r.company_name,
                "invoice_count": r.invoice_count,
                "total_billed": round(float(r.total_billed or 0), 2),
                "total_received": round(float(r.total_received or 0), 2),
                "total_outstanding": round(float(r.total_outstanding or 0), 2),
            }
            for r in rows
        ]
        return {
            "success": True,
            "customers": customers,
            "total_outstanding": f"{round(float(total_row), 2):.2f}",
            "from_date": str(date_from) if date_from else None,
            "to_date": str(date_to),
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/receivables")
async def list_receivables_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    party_type: Optional[str] = Query(None, description="Filter by party type: CUSTOMER, USER, VENDOR"),
    status: Optional[str] = Query(None, description="Filter by status: UNPAID, PARTIAL, OVERDUE"),
    as_on_date: Optional[str] = Query(None, description="As-on date YYYY-MM-DD (matches Consolidated BS date)"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=500, description="Items per page"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List accounts receivable — DC-AR-001: now sourced from sales_invoices (same as Consolidated BS Sundry Debtors).
    """
    try:
        receivables, total, total_pending, total_overdue = AccountsCreditService.get_receivables_list(
            db, current_user, company_id, party_type, status, page, page_size, as_on_date
        )
        
        return {
            "success": True,
            "receivables": receivables,
            "total": total,
            "total_pending": str(total_pending),
            "total_overdue": str(total_overdue),
            "page": page,
            "page_size": page_size
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/receivables/record-receipt")
async def record_customer_receipt_endpoint(
    data: RecordCustomerReceiptRequest,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record receipt from customer
    DC_CREDIT_001: WVV Protocol - Creates receipt transaction and updates invoice balance
    """
    try:
        result = AccountsCreditService.record_customer_receipt(
            db=db,
            employee=current_user,
            invoice_id=data.invoice_id,
            amount=data.amount,
            receipt_date=data.receipt_date,
            payment_mode=data.payment_mode,
            payment_reference=data.payment_reference,
            bank_reference=data.bank_reference,
            narration=data.narration
        )
        return result
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== DUTIES & TAXES (GST) SUMMARY (DC-GST-SUMMARY-001) ====================

@router.get("/duties-taxes-summary")
async def duties_taxes_summary_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company (0=all)"),
    as_on_date: Optional[str] = Query(None, description="Deprecated: use to_date"),
    from_date: Optional[str] = Query(None, description="Period start YYYY-MM-DD (omit = all time)"),
    to_date: Optional[str] = Query(None, description="Period end YYYY-MM-DD (omit = today)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-GST-SUMMARY-001: Duties & Taxes (GST) summary per company.
    Net GST Liability = Output Tax (from sales_invoices) - Input Tax Credit (from purchase_invoice_uploads).
    Shows CGST, SGST, IGST breakdown. Net > 0 means we owe the government; Net < 0 means ITC credit.
    Matches the Duties & Taxes line in the Consolidated Balance Sheet.
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        _to = to_date or as_on_date
        date_to   = _date.fromisoformat(_to) if _to else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "companies": [], "total_output_gst": "0.00",
                    "total_input_gst": "0.00", "net_gst_liability": "0.00"}

        placeholders = ",".join(str(c) for c in co_ids)
        co_map = {c["id"]: c.get("company_name") or c.get("name", "") for c in companies}

        params: dict = {"date_to": date_to}
        si_from_clause = ""
        pi_from_clause = ""
        if date_from:
            si_from_clause = "AND si.invoice_date >= :date_from"
            pi_from_clause = "AND pu.vendor_invoice_date >= :date_from"
            params["date_from"] = date_from

        # Output GST — sales invoices (confirmed)
        output_rows = db.execute(_text(f"""
            SELECT si.company_id,
                COUNT(*)                                                        AS inv_count,
                COALESCE(SUM(si.taxable_amount),0)                             AS taxable,
                COALESCE(SUM(si.cgst_amount),0)                                AS cgst,
                COALESCE(SUM(si.sgst_amount),0)                                AS sgst,
                COALESCE(SUM(si.igst_amount),0)                                AS igst,
                COALESCE(SUM(si.cgst_amount+si.sgst_amount+si.igst_amount),0)  AS total_tax
            FROM sales_invoices si
            WHERE si.company_id IN ({placeholders})
              AND si.status = 'CONFIRMED'
              AND si.invoice_date <= :date_to
              {si_from_clause}
            GROUP BY si.company_id
        """), params).fetchall()

        # Input GST (ITC) — purchase invoices (confirmed)
        input_rows = db.execute(_text(f"""
            SELECT pu.company_id,
                COUNT(*)                                                         AS inv_count,
                COALESCE(SUM(pu.taxable_amount),0)                              AS taxable,
                COALESCE(SUM(pu.cgst_amount),0)                                 AS cgst,
                COALESCE(SUM(pu.sgst_amount),0)                                 AS sgst,
                COALESCE(SUM(pu.igst_amount),0)                                 AS igst,
                COALESCE(SUM(pu.cgst_amount+pu.sgst_amount+pu.igst_amount),0)   AS total_tax
            FROM purchase_invoice_uploads pu
            WHERE pu.company_id IN ({placeholders})
              AND pu.status = 'CONFIRMED'
              AND pu.vendor_invoice_date <= :date_to
              {pi_from_clause}
            GROUP BY pu.company_id
        """), params).fetchall()

        out_map = {r.company_id: r for r in output_rows}
        in_map  = {r.company_id: r for r in input_rows}

        result_cos = []
        total_out = total_in = 0.0

        for coid in co_ids:
            o = out_map.get(coid)
            i = in_map.get(coid)

            o_cgst    = float(o.cgst)      if o else 0.0
            o_sgst    = float(o.sgst)      if o else 0.0
            o_igst    = float(o.igst)      if o else 0.0
            o_total   = float(o.total_tax) if o else 0.0
            o_taxable = float(o.taxable)   if o else 0.0
            o_count   = int(o.inv_count)   if o else 0

            i_cgst    = float(i.cgst)      if i else 0.0
            i_sgst    = float(i.sgst)      if i else 0.0
            i_igst    = float(i.igst)      if i else 0.0
            i_total   = float(i.total_tax) if i else 0.0
            i_taxable = float(i.taxable)   if i else 0.0
            i_count   = int(i.inv_count)   if i else 0

            net = o_total - i_total
            total_out += o_total
            total_in  += i_total

            result_cos.append({
                "company_id":   coid,
                "company_name": co_map.get(coid, ""),
                "output": {
                    "invoice_count": o_count,
                    "taxable": f"{o_taxable:.2f}",
                    "cgst":    f"{o_cgst:.2f}",
                    "sgst":    f"{o_sgst:.2f}",
                    "igst":    f"{o_igst:.2f}",
                    "total_tax": f"{o_total:.2f}"
                },
                "input": {
                    "invoice_count": i_count,
                    "taxable": f"{i_taxable:.2f}",
                    "cgst":    f"{i_cgst:.2f}",
                    "sgst":    f"{i_sgst:.2f}",
                    "igst":    f"{i_igst:.2f}",
                    "total_tax": f"{i_total:.2f}"
                },
                "net_liability": f"{net:.2f}"
            })

        net_total = total_out - total_in
        return {
            "success": True,
            "as_on":           str(date_to),
            "from_date":       str(date_from) if date_from else None,
            "companies":       result_cos,
            "total_output_gst":   f"{total_out:.2f}",
            "total_input_gst":    f"{total_in:.2f}",
            "net_gst_liability":  f"{net_total:.2f}"
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/capital-summary")
async def capital_summary_endpoint(
    company_id: Optional[int] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    as_on_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CAPITAL-001: Capital Account summary per company per ledger.
    Shows Owner's Capital, Reserves & Surplus, Capital Account breakdown.
    Opening Balance + Period Activity (Credits - Debits) = Closing Balance.
    Matches Capital Account line in Consolidated Balance Sheet.
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    from collections import defaultdict
    validate_accounts_access(current_user)
    try:
        _to = to_date or as_on_date
        date_to   = _date.fromisoformat(_to) if _to else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "companies": [], "total_capital": "0.00"}

        placeholders = ",".join(str(c) for c in co_ids)
        co_map = {c["id"]: c.get("company_name") or c.get("name", "") for c in companies}

        # All CAPITAL ledger masters
        ledger_rows = db.execute(_text(f"""
            SELECT id, company_id, account_name, account_code, parent_group,
                   COALESCE(opening_balance, 0)            AS opening_balance,
                   COALESCE(opening_balance_type, 'CREDIT') AS opening_balance_type
            FROM account_ledger_masters
            WHERE company_id IN ({placeholders})
              AND account_type = 'CAPITAL'
              AND is_active = true
            ORDER BY company_id, account_code, account_name
        """)).fetchall()

        # Period transactions from account_ledger
        date_params: dict = {"date_to": date_to}
        date_from_clause = ""
        if date_from:
            date_from_clause = "AND al.transaction_date >= :date_from"
            date_params["date_from"] = date_from

        txn_rows = db.execute(_text(f"""
            SELECT al.company_id, al.account_name,
                   COALESCE(SUM(al.debit_amount),  0) AS total_debit,
                   COALESCE(SUM(al.credit_amount), 0) AS total_credit,
                   COUNT(*)                            AS txn_count
            FROM account_ledger al
            WHERE al.company_id IN ({placeholders})
              AND al.account_type = 'CAPITAL'
              AND al.transaction_date <= :date_to
              {date_from_clause}
            GROUP BY al.company_id, al.account_name
        """), date_params).fetchall()

        txn_map = {(r.company_id, r.account_name): r for r in txn_rows}

        co_ledgers: dict = defaultdict(list)
        for lr in ledger_rows:
            ob      = float(lr.opening_balance)
            ob_type = lr.opening_balance_type or "CREDIT"
            # Capital is credit-nature → CREDIT OB = positive, DEBIT OB = negative
            ob_signed = ob if ob_type == "CREDIT" else -ob

            txn = txn_map.get((lr.company_id, lr.account_name))
            period_debit  = float(txn.total_debit)  if txn else 0.0
            period_credit = float(txn.total_credit) if txn else 0.0
            txn_count     = int(txn.txn_count)       if txn else 0

            # Closing = OB + Credits − Debits
            closing = ob_signed + period_credit - period_debit

            co_ledgers[lr.company_id].append({
                "ledger_name":      lr.account_name,
                "account_code":     lr.account_code  or "",
                "parent_group":     lr.parent_group   or "",
                "opening_balance":  f"{ob_signed:.2f}",
                "period_debit":     f"{period_debit:.2f}",
                "period_credit":    f"{period_credit:.2f}",
                "closing_balance":  f"{closing:.2f}",
                "txn_count":        txn_count,
            })

        result_cos = []
        grand_total = 0.0
        for coid in co_ids:
            ledgers   = co_ledgers.get(coid, [])
            co_total  = sum(float(l["closing_balance"]) for l in ledgers)
            grand_total += co_total
            result_cos.append({
                "company_id":    coid,
                "company_name":  co_map.get(coid, ""),
                "ledgers":       ledgers,
                "total_capital": f"{co_total:.2f}",
            })

        return {
            "success":       True,
            "as_on":         str(date_to),
            "from_date":     str(date_from) if date_from else None,
            "companies":     result_cos,
            "total_capital": f"{grand_total:.2f}",
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/capital-transactions")
async def capital_transactions_endpoint(
    company_id: Optional[int] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    ledger_name: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CAPITAL-002: Capital Account transaction history from account_ledger.
    Returns individual debit/credit entries for all CAPITAL type accounts.
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        date_to   = _date.fromisoformat(to_date)   if to_date   else _date.today()
        date_from = _date.fromisoformat(from_date)  if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids    = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "transactions": [], "count": 0}

        placeholders = ",".join(str(c) for c in co_ids)
        co_map = {c["id"]: c.get("company_name") or c.get("name", "") for c in companies}

        params: dict = {"date_to": date_to}
        extra = ""
        if date_from:
            extra += " AND al.transaction_date >= :date_from"
            params["date_from"] = date_from
        if ledger_name:
            extra += " AND al.account_name = :ledger_name"
            params["ledger_name"] = ledger_name

        rows = db.execute(_text(f"""
            SELECT al.id, al.company_id, al.account_name, al.transaction_date,
                   al.entry_type, al.debit_amount, al.credit_amount,
                   al.reference_type, al.reference_number,
                   al.narration, al.voucher_type, al.particulars
            FROM account_ledger al
            WHERE al.company_id IN ({placeholders})
              AND al.account_type = 'CAPITAL'
              AND al.transaction_date <= :date_to
              {extra}
            ORDER BY al.transaction_date DESC, al.id DESC
            LIMIT 500
        """), params).fetchall()

        txns = [{
            "id":               r.id,
            "company_id":       r.company_id,
            "company_name":     co_map.get(r.company_id, ""),
            "account_name":     r.account_name,
            "date":             str(r.transaction_date),
            "entry_type":       r.entry_type,
            "debit":            f"{float(r.debit_amount):.2f}",
            "credit":           f"{float(r.credit_amount):.2f}",
            "reference_type":   r.reference_type   or "",
            "reference_number": r.reference_number or "",
            "narration":        r.narration         or "",
            "voucher_type":     r.voucher_type      or "",
            "particulars":      r.particulars        or "",
        } for r in rows]

        return {"success": True, "transactions": txns, "count": len(txns)}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/cash-in-hand-summary")
async def cash_in_hand_summary_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company (0=all)"),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    as_on_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CASH-001: Cash in Hand consolidated summary per company.
    Sources:
      1. income_entries (payment_mode='CASH') → cash received by company
      2. expense_entries (payment_mode='CASH') → cash paid out by company
      3. fund_allocations (balance_remaining) → unspent cash float in staff hands
      4. account_ledger (account_type='CASH') → formal double-entry (future-proof)
    Net Cash in Hand = (Cash In - Cash Out) + Fund Float + Ledger Balance
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        _to = to_date or as_on_date
        date_to   = _date.fromisoformat(_to) if _to else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids    = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "companies": [], "total_cash": "0.00",
                    "total_receipts": "0.00", "total_payments": "0.00",
                    "total_fund_float": "0.00"}

        phs = ",".join(str(c) for c in co_ids)
        co_map = {c["id"]: c.get("company_name") or c.get("name", "") for c in companies}

        p: dict = {"date_to": date_to}
        df_ie = df_ee = ""
        if date_from:
            df_ie = "AND ie.income_date >= :date_from"
            df_ee = "AND ee.expense_date >= :date_from"
            p["date_from"] = date_from

        # ── 1. Cash Receipts: income_entries with payment_mode='CASH' ──────────
        ie_rows = db.execute(_text(f"""
            SELECT company_id,
                   COALESCE(SUM(amount), 0) AS cash_in,
                   COUNT(*)                 AS cnt
            FROM income_entries ie
            WHERE company_id IN ({phs})
              AND payment_mode = 'CASH'
              AND status IN ('CONFIRMED', 'APPROVED')
              AND (is_deleted IS NULL OR is_deleted = FALSE)
              AND ie.income_date <= :date_to
              {df_ie}
            GROUP BY company_id
        """), p).fetchall()
        ie_map = {r.company_id: {"cash_in": float(r.cash_in), "cnt": int(r.cnt)} for r in ie_rows}

        # ── 2. Cash Payments: expense_entries with payment_mode='CASH' ─────────
        ee_rows = db.execute(_text(f"""
            SELECT company_id,
                   COALESCE(SUM(amount), 0) AS cash_out,
                   COUNT(*)                 AS cnt
            FROM expense_entries ee
            WHERE company_id IN ({phs})
              AND payment_mode = 'CASH'
              AND status = 'APPROVED'
              AND ee.expense_date <= :date_to
              {df_ee}
            GROUP BY company_id
        """), p).fetchall()
        ee_map = {r.company_id: {"cash_out": float(r.cash_out), "cnt": int(r.cnt)} for r in ee_rows}

        # ── 3. Staff Float: fund_allocations balance_remaining (date-independent) ─
        fa_rows = db.execute(_text(f"""
            SELECT company_id,
                   COALESCE(SUM(balance_remaining), 0) AS float_balance,
                   COALESCE(SUM(amount), 0)            AS total_allocated,
                   COUNT(*)                            AS cnt
            FROM fund_allocations
            WHERE company_id IN ({phs})
              AND status IN ('APPROVED', 'ACTIVE', 'PENDING')
              AND COALESCE(balance_remaining, 0) > 0
            GROUP BY company_id
        """)).fetchall()
        fa_map = {r.company_id: {"float": float(r.float_balance),
                                  "allocated": float(r.total_allocated),
                                  "cnt": int(r.cnt)} for r in fa_rows}

        # ── 4. Formal Cash Ledger (account_ledger CASH type, currently 0 rows) ─
        ob_rows = db.execute(_text(f"""
            SELECT company_id,
                   COALESCE(SUM(
                       CASE WHEN opening_balance_type='DEBIT' THEN opening_balance
                            ELSE -opening_balance END), 0) AS ob
            FROM account_ledger_masters
            WHERE company_id IN ({phs})
              AND account_type = 'CASH' AND is_active = TRUE
            GROUP BY company_id
        """)).fetchall()
        ob_map = {r.company_id: float(r.ob) for r in ob_rows}

        ldg_rows = db.execute(_text(f"""
            SELECT company_id,
                   COALESCE(SUM(debit_amount), 0)  AS dr,
                   COALESCE(SUM(credit_amount), 0) AS cr
            FROM account_ledger
            WHERE company_id IN ({phs})
              AND account_type = 'CASH'
              AND transaction_date <= :date_to
              {"AND transaction_date >= :date_from" if date_from else ""}
            GROUP BY company_id
        """), p).fetchall()
        ldg_map = {r.company_id: {"dr": float(r.dr), "cr": float(r.cr)} for r in ldg_rows}

        # ── Build per-company result ──────────────────────────────────────────
        result_cos = []
        grand_total = grand_receipts = grand_payments = grand_float = 0.0

        for coid in co_ids:
            ie   = ie_map.get(coid,  {"cash_in": 0.0, "cnt": 0})
            ee   = ee_map.get(coid,  {"cash_out": 0.0, "cnt": 0})
            fa   = fa_map.get(coid,  {"float": 0.0, "allocated": 0.0, "cnt": 0})
            ob   = ob_map.get(coid,  0.0)
            ldg  = ldg_map.get(coid, {"dr": 0.0, "cr": 0.0})

            cash_in      = ie["cash_in"]
            cash_out     = ee["cash_out"]
            fund_float   = fa["float"]
            ledger_net   = ob + ldg["dr"] - ldg["cr"]   # formal ledger portion

            net_cash = (cash_in - cash_out) + fund_float + ledger_net

            grand_total    += net_cash
            grand_receipts += cash_in
            grand_payments += cash_out
            grand_float    += fund_float

            result_cos.append({
                "company_id":        coid,
                "company_name":      co_map.get(coid, ""),
                "cash_receipts":     f"{cash_in:.2f}",
                "cash_receipts_cnt": ie["cnt"],
                "cash_payments":     f"{cash_out:.2f}",
                "cash_payments_cnt": ee["cnt"],
                "fund_float":        f"{fund_float:.2f}",
                "fund_float_cnt":    fa["cnt"],
                "fund_allocated":    f"{fa['allocated']:.2f}",
                "ledger_net":        f"{ledger_net:.2f}",
                "net_cash_in_hand":  f"{net_cash:.2f}",
            })

        return {
            "success":          True,
            "as_on":            str(date_to),
            "from_date":        str(date_from) if date_from else None,
            "companies":        result_cos,
            "total_cash":       f"{grand_total:.2f}",
            "total_receipts":   f"{grand_receipts:.2f}",
            "total_payments":   f"{grand_payments:.2f}",
            "total_fund_float": f"{grand_float:.2f}",
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/cash-in-hand-transactions")
async def cash_in_hand_transactions_endpoint(
    company_id: Optional[int] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CASH-002: Cash in Hand transaction history.
    UNION of:
      - income_entries (payment_mode='CASH') → RECEIPT rows
      - expense_entries (payment_mode='CASH') → PAYMENT rows
      - fund_allocations (active balance) → STAFF FLOAT rows
      - account_ledger CASH entries → LEDGER rows
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        date_to   = _date.fromisoformat(to_date)  if to_date  else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids    = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "transactions": [], "count": 0}

        phs    = ",".join(str(c) for c in co_ids)
        co_map = {c["id"]: c.get("company_name") or c.get("name", "") for c in companies}
        p: dict = {"date_to": date_to}
        df_ie = df_ee = df_ldg = ""
        if date_from:
            df_ie  = "AND ie.income_date  >= :date_from"
            df_ee  = "AND ee.expense_date >= :date_from"
            df_ldg = "AND al.transaction_date >= :date_from"
            p["date_from"] = date_from

        # Income entries (cash receipts)
        ie_rows = db.execute(_text(f"""
            SELECT ie.id, ie.company_id, ie.income_date AS txn_date,
                   'RECEIPT'           AS txn_type,
                   ie.amount           AS receipt_amt,
                   0                   AS payment_amt,
                   ie.entry_number     AS ref_number,
                   ie.payer_name       AS party,
                   ie.narration        AS narration,
                   ie.status           AS status,
                   'INCOME'            AS source
            FROM income_entries ie
            WHERE ie.company_id IN ({phs})
              AND ie.payment_mode = 'CASH'
              AND ie.status IN ('CONFIRMED', 'APPROVED')
              AND (ie.is_deleted IS NULL OR ie.is_deleted = FALSE)
              AND ie.income_date <= :date_to
              {df_ie}
        """), p).fetchall()

        # Expense entries (cash payments)
        ee_rows = db.execute(_text(f"""
            SELECT ee.id, ee.company_id, ee.expense_date AS txn_date,
                   'PAYMENT'           AS txn_type,
                   0                   AS receipt_amt,
                   ee.amount           AS payment_amt,
                   ee.entry_number     AS ref_number,
                   COALESCE(ee.vendor_name, '')  AS party,
                   ee.narration        AS narration,
                   ee.status           AS status,
                   'EXPENSE'           AS source
            FROM expense_entries ee
            WHERE ee.company_id IN ({phs})
              AND ee.payment_mode = 'CASH'
              AND ee.status = 'APPROVED'
              AND ee.expense_date <= :date_to
              {df_ee}
        """), p).fetchall()

        # Fund allocations (staff holding cash float)
        fa_rows = db.execute(_text(f"""
            SELECT fa.id, fa.company_id, fa.allocation_date AS txn_date,
                   'STAFF FLOAT'       AS txn_type,
                   fa.amount           AS receipt_amt,
                   (fa.amount - COALESCE(fa.balance_remaining,0)) AS payment_amt,
                   fa.allocation_number AS ref_number,
                   COALESCE(se.employee_name, 'Staff #' || fa.to_employee_id::text) AS party,
                   fa.purpose          AS narration,
                   fa.status           AS status,
                   'FUND_ALLOC'        AS source
            FROM fund_allocations fa
            LEFT JOIN staff_employees se ON se.id = fa.to_employee_id
            WHERE fa.company_id IN ({phs})
              AND fa.status IN ('APPROVED', 'ACTIVE', 'PENDING')
              AND COALESCE(fa.balance_remaining, 0) > 0
        """)).fetchall()

        # Formal ledger entries (currently empty but future-proof)
        ldg_rows = db.execute(_text(f"""
            SELECT al.id, al.company_id, al.transaction_date AS txn_date,
                   CASE WHEN al.debit_amount > 0 THEN 'RECEIPT' ELSE 'PAYMENT' END AS txn_type,
                   COALESCE(al.debit_amount,  0) AS receipt_amt,
                   COALESCE(al.credit_amount, 0) AS payment_amt,
                   al.reference_number   AS ref_number,
                   al.particulars        AS party,
                   al.narration          AS narration,
                   'POSTED'              AS status,
                   'LEDGER'              AS source
            FROM account_ledger al
            WHERE al.company_id IN ({phs})
              AND al.account_type = 'CASH'
              AND al.transaction_date <= :date_to
              {df_ldg}
        """), p).fetchall()

        all_rows = list(ie_rows) + list(ee_rows) + list(fa_rows) + list(ldg_rows)
        all_rows.sort(key=lambda r: str(r.txn_date), reverse=True)

        txns = [{
            "id":           r.id,
            "company_id":   r.company_id,
            "company_name": co_map.get(r.company_id, ""),
            "date":         str(r.txn_date),
            "txn_type":     r.txn_type,
            "source":       r.source,
            "receipt":      f"{float(r.receipt_amt):.2f}",
            "payment":      f"{float(r.payment_amt):.2f}",
            "ref_number":   r.ref_number or "",
            "party":        r.party      or "",
            "narration":    r.narration  or "",
            "status":       r.status     or "",
        } for r in all_rows[:500]]

        return {"success": True, "transactions": txns, "count": len(txns)}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/duties-taxes-hsn-summary")
async def duties_taxes_hsn_summary_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company (0=all)"),
    as_on_date: Optional[str] = Query(None, description="Deprecated: use to_date"),
    from_date: Optional[str] = Query(None, description="Period start YYYY-MM-DD"),
    to_date: Optional[str] = Query(None, description="Period end YYYY-MM-DD"),
    side: Optional[str] = Query("both", description="output | input | both"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-GST-HSN-001: HSN-code-wise GST breakdown for GST return filing.
    Output side: sales_invoice_line_items grouped by HSN code.
    Input side:  purchase_invoice_line_items grouped by HSN code.
    Useful for GSTR-1 (output) and GSTR-2A/2B (input) reconciliation.
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        _to = to_date or as_on_date
        date_to   = _date.fromisoformat(_to) if _to else _date.today()
        date_from = _date.fromisoformat(from_date) if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "output_hsn": [], "input_hsn": []}

        placeholders = ",".join(str(c) for c in co_ids)
        params: dict = {"date_to": date_to}
        si_from_clause = ""
        pi_from_clause = ""
        if date_from:
            si_from_clause = "AND si.invoice_date >= :date_from"
            pi_from_clause = "AND pu.vendor_invoice_date >= :date_from"
            params["date_from"] = date_from

        output_hsn = []
        if side in ("output", "both"):
            rows = db.execute(_text(f"""
                SELECT
                    COALESCE(NULLIF(TRIM(sili.hsn_code),''), '(No HSN)') AS hsn_code,
                    COUNT(DISTINCT si.id)                                AS inv_count,
                    SUM(sili.gst_rate)  / NULLIF(COUNT(*),0)            AS gst_rate,
                    COALESCE(SUM(sili.taxable_amount),0)                AS taxable,
                    COALESCE(SUM(sili.cgst_amount),0)                   AS cgst,
                    COALESCE(SUM(sili.sgst_amount),0)                   AS sgst,
                    COALESCE(SUM(sili.igst_amount),0)                   AS igst,
                    COALESCE(SUM(sili.cgst_amount+sili.sgst_amount+sili.igst_amount),0) AS total_tax
                FROM sales_invoice_line_items sili
                JOIN sales_invoices si ON si.id = sili.invoice_id
                    AND si.company_id IN ({placeholders})
                    AND si.status = 'CONFIRMED'
                    AND si.invoice_date <= :date_to
                    {si_from_clause}
                GROUP BY 1
                ORDER BY total_tax DESC
            """), params).fetchall()
            output_hsn = [
                {
                    "hsn_code":   r.hsn_code,
                    "inv_count":  int(r.inv_count),
                    "gst_rate":   f"{float(r.gst_rate or 0):.2f}",
                    "taxable":    f"{float(r.taxable):.2f}",
                    "cgst":       f"{float(r.cgst):.2f}",
                    "sgst":       f"{float(r.sgst):.2f}",
                    "igst":       f"{float(r.igst):.2f}",
                    "total_tax":  f"{float(r.total_tax):.2f}"
                } for r in rows
            ]

        input_hsn = []
        if side in ("input", "both"):
            rows = db.execute(_text(f"""
                SELECT
                    COALESCE(NULLIF(TRIM(pili.hsn_code),''), '(No HSN)') AS hsn_code,
                    COUNT(DISTINCT pu.id)                                AS inv_count,
                    SUM(pili.gst_rate) / NULLIF(COUNT(*),0)             AS gst_rate,
                    COALESCE(SUM(pili.taxable_amount),0)                AS taxable,
                    COALESCE(SUM(pili.cgst_amount),0)                   AS cgst,
                    COALESCE(SUM(pili.sgst_amount),0)                   AS sgst,
                    COALESCE(SUM(pili.igst_amount),0)                   AS igst,
                    COALESCE(SUM(pili.cgst_amount+pili.sgst_amount+pili.igst_amount),0) AS total_tax
                FROM purchase_invoice_line_items pili
                JOIN purchase_invoice_uploads pu ON pu.id = pili.upload_id
                    AND pu.company_id IN ({placeholders})
                    AND pu.status = 'CONFIRMED'
                    AND pu.vendor_invoice_date <= :date_to
                    {pi_from_clause}
                GROUP BY 1
                ORDER BY total_tax DESC
            """), params).fetchall()
            input_hsn = [
                {
                    "hsn_code":   r.hsn_code,
                    "inv_count":  int(r.inv_count),
                    "gst_rate":   f"{float(r.gst_rate or 0):.2f}",
                    "taxable":    f"{float(r.taxable):.2f}",
                    "cgst":       f"{float(r.cgst):.2f}",
                    "sgst":       f"{float(r.sgst):.2f}",
                    "igst":       f"{float(r.igst):.2f}",
                    "total_tax":  f"{float(r.total_tax):.2f}"
                } for r in rows
            ]

        return {
            "success":    True,
            "as_on":      str(date_to),
            "from_date":  str(date_from) if date_from else None,
            "output_hsn": output_hsn,
            "input_hsn":  input_hsn
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/duties-taxes-hsn-detail")
async def duties_taxes_hsn_detail_endpoint(
    hsn_code: str = Query(..., description="HSN code to drill into (use '(No HSN)' for blank)"),
    side: str = Query("output", description="output (sales) | input (purchases)"),
    company_id: Optional[int] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-GST-HSN-002: Invoice-level drill-through for a specific HSN code.
    Output: sales_invoice_line_items grouped per invoice.
    Input:  purchase_invoice_line_items grouped per invoice.
    Each row = one invoice with sum of line items for that HSN code.
    """
    from app.services.staff_accounts_service import validate_accounts_access, _get_companies_map
    from sqlalchemy import text as _text
    from datetime import date as _date
    validate_accounts_access(current_user)
    try:
        date_to   = _date.fromisoformat(to_date)   if to_date   else _date.today()
        date_from = _date.fromisoformat(from_date)  if from_date else None

        cid = company_id if company_id else 0
        companies = _get_companies_map(db, cid, current_user)
        co_ids    = [c["id"] for c in companies]
        if not co_ids:
            return {"success": True, "invoices": [], "line_items": []}

        phs    = ",".join(str(c) for c in co_ids)
        p: dict = {"date_to": date_to}
        # HSN filter: '(No HSN)' → match blank/null
        is_no_hsn = hsn_code.strip() == "(No HSN)"
        if not is_no_hsn:
            p["hsn_code"] = hsn_code.strip()

        hsn_filter_sili = "AND (COALESCE(NULLIF(TRIM(sili.hsn_code),''), '(No HSN)') = '(No HSN)')" if is_no_hsn else "AND TRIM(sili.hsn_code) = :hsn_code"
        hsn_filter_pili = "AND (COALESCE(NULLIF(TRIM(pili.hsn_code),''), '(No HSN)') = '(No HSN)')" if is_no_hsn else "AND TRIM(pili.hsn_code) = :hsn_code"

        from_clause_si = "AND si.invoice_date >= :date_from" if date_from else ""
        from_clause_pu = "AND pu.vendor_invoice_date >= :date_from" if date_from else ""
        if date_from:
            p["date_from"] = date_from

        invoices = []
        line_items = []

        if side == "output":
            # ── Per-invoice summary (one row per sales invoice) ──────────────
            inv_rows = db.execute(_text(f"""
                SELECT
                    si.id                                           AS invoice_id,
                    si.invoice_number                              AS invoice_number,
                    si.invoice_date                                AS invoice_date,
                    si.customer_name                               AS party_name,
                    COALESCE(si.customer_gstin, '')                AS gstin,
                    COUNT(sili.id)                                 AS line_count,
                    COALESCE(SUM(sili.quantity), 0)                AS total_qty,
                    COALESCE(SUM(sili.taxable_amount), 0)          AS taxable,
                    COALESCE(SUM(sili.cgst_amount), 0)             AS cgst,
                    COALESCE(SUM(sili.sgst_amount), 0)             AS sgst,
                    COALESCE(SUM(sili.igst_amount), 0)             AS igst,
                    COALESCE(SUM(sili.cgst_amount + sili.sgst_amount + sili.igst_amount), 0) AS total_tax
                FROM sales_invoice_line_items sili
                JOIN sales_invoices si ON si.id = sili.invoice_id
                    AND si.company_id IN ({phs})
                    AND si.status = 'CONFIRMED'
                    AND si.invoice_date <= :date_to
                    {from_clause_si}
                WHERE 1=1 {hsn_filter_sili}
                GROUP BY si.id, si.invoice_number, si.invoice_date, si.customer_name, si.customer_gstin
                ORDER BY si.invoice_date DESC, si.invoice_number
            """), p).fetchall()

            for r in inv_rows:
                invoices.append({
                    "invoice_id":     r.invoice_id,
                    "invoice_number": r.invoice_number or "—",
                    "invoice_date":   str(r.invoice_date),
                    "party_name":     r.party_name     or "—",
                    "gstin":          r.gstin,
                    "line_count":     int(r.line_count),
                    "total_qty":      f"{float(r.total_qty):.3f}".rstrip('0').rstrip('.'),
                    "taxable":        f"{float(r.taxable):.2f}",
                    "cgst":           f"{float(r.cgst):.2f}",
                    "sgst":           f"{float(r.sgst):.2f}",
                    "igst":           f"{float(r.igst):.2f}",
                    "total_tax":      f"{float(r.total_tax):.2f}",
                })

            # ── Individual line items across all matching invoices ────────────
            li_rows = db.execute(_text(f"""
                SELECT
                    si.invoice_number                              AS invoice_number,
                    si.invoice_date                                AS invoice_date,
                    si.customer_name                               AS party_name,
                    sili.item_code                                 AS item_code,
                    sili.item_description                          AS item_description,
                    COALESCE(sili.quantity, 0)                     AS quantity,
                    COALESCE(sili.unit_of_measure, '')             AS uom,
                    COALESCE(sili.unit_rate, 0)                    AS unit_rate,
                    COALESCE(sili.taxable_amount, 0)               AS taxable,
                    COALESCE(sili.gst_rate, 0)                     AS gst_rate,
                    COALESCE(sili.cgst_amount, 0)                  AS cgst,
                    COALESCE(sili.sgst_amount, 0)                  AS sgst,
                    COALESCE(sili.igst_amount, 0)                  AS igst,
                    COALESCE(sili.cgst_amount+sili.sgst_amount+sili.igst_amount, 0) AS total_tax,
                    COALESCE(sili.line_total, 0)                   AS line_total
                FROM sales_invoice_line_items sili
                JOIN sales_invoices si ON si.id = sili.invoice_id
                    AND si.company_id IN ({phs})
                    AND si.status = 'CONFIRMED'
                    AND si.invoice_date <= :date_to
                    {from_clause_si}
                WHERE 1=1 {hsn_filter_sili}
                ORDER BY si.invoice_date DESC, si.invoice_number, sili.line_number
                LIMIT 1000
            """), p).fetchall()

            for r in li_rows:
                line_items.append({
                    "invoice_number":   r.invoice_number   or "—",
                    "invoice_date":     str(r.invoice_date),
                    "party_name":       r.party_name       or "—",
                    "item_code":        r.item_code        or "—",
                    "item_description": r.item_description or "—",
                    "quantity":         f"{float(r.quantity):.3f}".rstrip('0').rstrip('.'),
                    "uom":              r.uom,
                    "unit_rate":        f"{float(r.unit_rate):.2f}",
                    "taxable":          f"{float(r.taxable):.2f}",
                    "gst_rate":         f"{float(r.gst_rate):.1f}",
                    "cgst":             f"{float(r.cgst):.2f}",
                    "sgst":             f"{float(r.sgst):.2f}",
                    "igst":             f"{float(r.igst):.2f}",
                    "total_tax":        f"{float(r.total_tax):.2f}",
                    "line_total":       f"{float(r.line_total):.2f}",
                })

        else:  # input / purchases
            inv_rows = db.execute(_text(f"""
                SELECT
                    pu.id                                          AS invoice_id,
                    COALESCE(pu.vendor_invoice_no, pu.id::text)    AS invoice_number,
                    pu.vendor_invoice_date                         AS invoice_date,
                    COALESCE(vm.vendor_name, 'Vendor #' || pu.vendor_id::text) AS party_name,
                    COUNT(pili.id)                                 AS line_count,
                    COALESCE(SUM(pili.quantity), 0)                AS total_qty,
                    COALESCE(SUM(pili.taxable_amount), 0)          AS taxable,
                    COALESCE(SUM(pili.cgst_amount), 0)             AS cgst,
                    COALESCE(SUM(pili.sgst_amount), 0)             AS sgst,
                    COALESCE(SUM(pili.igst_amount), 0)             AS igst,
                    COALESCE(SUM(pili.cgst_amount+pili.sgst_amount+pili.igst_amount), 0) AS total_tax
                FROM purchase_invoice_line_items pili
                JOIN purchase_invoice_uploads pu ON pu.id = pili.upload_id
                    AND pu.company_id IN ({phs})
                    AND pu.status = 'CONFIRMED'
                    AND pu.vendor_invoice_date <= :date_to
                    {from_clause_pu}
                LEFT JOIN vendor_master vm ON vm.id = pu.vendor_id
                WHERE 1=1 {hsn_filter_pili}
                GROUP BY pu.id, pu.vendor_invoice_no, pu.vendor_invoice_date, pu.vendor_id, vm.vendor_name
                ORDER BY pu.vendor_invoice_date DESC, pu.vendor_invoice_no
            """), p).fetchall()

            for r in inv_rows:
                invoices.append({
                    "invoice_id":     r.invoice_id,
                    "invoice_number": r.invoice_number or "—",
                    "invoice_date":   str(r.invoice_date) if r.invoice_date else "—",
                    "party_name":     r.party_name     or "—",
                    "gstin":          "",
                    "line_count":     int(r.line_count),
                    "total_qty":      f"{float(r.total_qty):.3f}".rstrip('0').rstrip('.'),
                    "taxable":        f"{float(r.taxable):.2f}",
                    "cgst":           f"{float(r.cgst):.2f}",
                    "sgst":           f"{float(r.sgst):.2f}",
                    "igst":           f"{float(r.igst):.2f}",
                    "total_tax":      f"{float(r.total_tax):.2f}",
                })

            li_rows = db.execute(_text(f"""
                SELECT
                    COALESCE(pu.vendor_invoice_no, pu.id::text)    AS invoice_number,
                    pu.vendor_invoice_date                         AS invoice_date,
                    COALESCE(vm.vendor_name, 'Vendor #' || pu.vendor_id::text) AS party_name,
                    pili.item_code                                 AS item_code,
                    pili.item_description                          AS item_description,
                    COALESCE(pili.quantity, 0)                     AS quantity,
                    COALESCE(pili.unit_of_measure, '')             AS uom,
                    COALESCE(pili.unit_rate, 0)                    AS unit_rate,
                    COALESCE(pili.taxable_amount, 0)               AS taxable,
                    COALESCE(pili.gst_rate, 0)                     AS gst_rate,
                    COALESCE(pili.cgst_amount, 0)                  AS cgst,
                    COALESCE(pili.sgst_amount, 0)                  AS sgst,
                    COALESCE(pili.igst_amount, 0)                  AS igst,
                    COALESCE(pili.cgst_amount+pili.sgst_amount+pili.igst_amount, 0) AS total_tax,
                    COALESCE(pili.line_total, 0)                   AS line_total
                FROM purchase_invoice_line_items pili
                JOIN purchase_invoice_uploads pu ON pu.id = pili.upload_id
                    AND pu.company_id IN ({phs})
                    AND pu.status = 'CONFIRMED'
                    AND pu.vendor_invoice_date <= :date_to
                    {from_clause_pu}
                LEFT JOIN vendor_master vm ON vm.id = pu.vendor_id
                WHERE 1=1 {hsn_filter_pili}
                ORDER BY pu.vendor_invoice_date DESC, pu.vendor_invoice_no, pili.line_number
                LIMIT 1000
            """), p).fetchall()

            for r in li_rows:
                line_items.append({
                    "invoice_number":   r.invoice_number   or "—",
                    "invoice_date":     str(r.invoice_date) if r.invoice_date else "—",
                    "party_name":       r.party_name       or "—",
                    "item_code":        r.item_code        or "—",
                    "item_description": r.item_description or "—",
                    "quantity":         f"{float(r.quantity):.3f}".rstrip('0').rstrip('.'),
                    "uom":              r.uom,
                    "unit_rate":        f"{float(r.unit_rate):.2f}",
                    "taxable":          f"{float(r.taxable):.2f}",
                    "gst_rate":         f"{float(r.gst_rate):.1f}",
                    "cgst":             f"{float(r.cgst):.2f}",
                    "sgst":             f"{float(r.sgst):.2f}",
                    "igst":             f"{float(r.igst):.2f}",
                    "total_tax":        f"{float(r.total_tax):.2f}",
                    "line_total":       f"{float(r.line_total):.2f}",
                })

        return {
            "success":    True,
            "hsn_code":   hsn_code,
            "side":       side,
            "invoices":   invoices,
            "line_items": line_items,
            "inv_count":  len(invoices),
            "li_count":   len(line_items),
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== AGING ANALYSIS ENDPOINTS (DC_CREDIT_001) ====================

@router.get("/credit/aging/{company_id}")
async def get_aging_summary_endpoint(
    company_id: int = Path(..., description="Company ID"),
    credit_type: str = Query(..., description="Credit type: PAYABLE or RECEIVABLE"),
    as_of_date: Optional[date] = Query(None, description="As-of date for aging calculation"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get aging summary with bucket breakdown
    DC_CREDIT_001: VGK/EA access - Calculates aging buckets (Current, 1-30, 31-60, 61-90, 90+ days)
    """
    try:
        result = AccountsCreditService.get_aging_summary(
            db, current_user, company_id, credit_type, as_of_date
        )
        return result
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/credit/transactions")
async def list_payment_transactions_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company"),
    transaction_type: Optional[str] = Query(None, description="Filter by type: PAYMENT_TO_VENDOR, RECEIPT_FROM_CUSTOMER"),
    start_date: Optional[date] = Query(None, description="Start date filter"),
    end_date: Optional[date] = Query(None, description="End date filter"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List payment/receipt transactions
    DC_CREDIT_001: VGK/EA access - Payment history for auditing
    """
    try:
        transactions, total, total_amount = AccountsCreditService.get_payment_transactions(
            db, current_user, company_id, transaction_type, start_date, end_date, page, page_size
        )
        
        return {
            "success": True,
            "transactions": [
                {
                    "id": t.id,
                    "transaction_number": t.transaction_number,
                    "transaction_type": t.transaction_type,
                    "company_id": t.company_id,
                    "party_type": t.party_type,
                    "party_name": t.party_name,
                    "transaction_date": str(t.transaction_date),
                    "amount": str(t.amount),
                    "payment_mode": t.payment_mode,
                    "payment_reference": t.payment_reference,
                    "status": t.status,
                    "created_at": str(t.created_at)
                }
                for t in transactions
            ],
            "total": total,
            "total_amount": str(total_amount),
            "page": page,
            "page_size": page_size
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/credit/dashboard/{company_id}")
async def get_credit_dashboard_endpoint(
    company_id: int = Path(..., description="Company ID"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get credit dashboard overview
    DC_CREDIT_001: VGK/EA access - Combined view of payables, receivables, and recent transactions
    """
    try:
        from app.services.staff_accounts_service import get_indian_time
        today = get_indian_time().date()
        
        payables_summary = AccountsCreditService.get_aging_summary(
            db, current_user, company_id, 'PAYABLE', today
        )
        
        receivables_summary = AccountsCreditService.get_aging_summary(
            db, current_user, company_id, 'RECEIVABLE', today
        )
        
        payables, _, _, _ = AccountsCreditService.get_payables_list(
            db, current_user, company_id, None, 'OVERDUE', 1, 5
        )
        
        receivables, _, _, _ = AccountsCreditService.get_receivables_list(
            db, current_user, company_id, None, 'OVERDUE', 1, 5
        )
        
        payments, _, _ = AccountsCreditService.get_payment_transactions(
            db, current_user, company_id, 'PAYMENT_TO_VENDOR', None, None, 1, 5
        )
        
        receipts, _, _ = AccountsCreditService.get_payment_transactions(
            db, current_user, company_id, 'RECEIPT_FROM_CUSTOMER', None, None, 1, 5
        )
        
        net_position = (
            receivables_summary['summary']['total_outstanding'] - 
            payables_summary['summary']['total_outstanding']
        )
        
        return {
            "success": True,
            "payables_summary": payables_summary['summary'],
            "receivables_summary": receivables_summary['summary'],
            "net_credit_position": str(net_position),
            "payables_count": payables_summary['total_transactions'],
            "receivables_count": receivables_summary['total_transactions'],
            "overdue_payables": payables,
            "overdue_receivables": receivables,
            "recent_payments": [
                {
                    "id": t.id,
                    "transaction_number": t.transaction_number,
                    "party_name": t.party_name,
                    "amount": str(t.amount),
                    "transaction_date": str(t.transaction_date),
                    "payment_mode": t.payment_mode
                }
                for t in payments
            ],
            "recent_receipts": [
                {
                    "id": t.id,
                    "transaction_number": t.transaction_number,
                    "party_name": t.party_name,
                    "amount": str(t.amount),
                    "transaction_date": str(t.transaction_date),
                    "payment_mode": t.payment_mode
                }
                for t in receipts
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== BOM ENDPOINTS (DC_BOM_001 - Dec 06, 2025) ====================

@router.post("/bom", status_code=status.HTTP_201_CREATED)
async def create_bom(
    data: BOMCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new Bill of Materials
    DC_BOM_001: VGK/EA access - Create BOM with components
    """
    try:
        line_items = [item.model_dump() for item in data.line_items]
        
        bom = BOMService.create_bom(
            db=db,
            bom_name=data.bom_name,
            company_id=data.company_id,
            finished_product_id=data.finished_product_id,
            line_items=line_items,
            employee=current_user,
            standard_qty=data.standard_qty,
            unit_of_measure=data.unit_of_measure.value if data.unit_of_measure else 'PCS',
            effective_from=data.effective_from,
            effective_to=data.effective_to,
            estimated_time_hours=data.estimated_time_hours,
            notes=data.notes
        )
        
        return _format_bom_response(bom, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/bom")
async def list_boms(
    company_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List BOMs with filters
    DC_BOM_001: VGK/EA access
    """
    try:
        boms, total = BOMService.list_boms(
            db=db,
            employee=current_user,
            company_id=company_id,
            status=status,
            search=search,
            page=page,
            page_size=page_size
        )

        return {
            "success": True,
            "boms": [_format_bom_response(b, include_lines=False) for b in boms],
            "total": total,
            "page": page,
            "page_size": page_size
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/bom/{bom_id}")
async def get_bom(
    bom_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get BOM by ID with line items
    DC_BOM_001: VGK/EA access
    """
    try:
        bom = BOMService.get_bom(db, bom_id, current_user)
        return _format_bom_response(bom, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/bom/{bom_id}")
async def update_bom(
    bom_id: int = Path(..., ge=1),
    data: BOMUpdate = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update BOM - DC Protocol Re-approval Workflow
    DC_BOM_001: VGK/EA access
    DRAFT/PENDING: Direct update
    APPROVED: Triggers re-approval (status -> PENDING_APPROVAL)
    REJECTED: Cannot edit (use copy instead)
    """
    try:
        updates = data.model_dump(exclude_unset=True) if data else {}
        change_reason = updates.pop('change_reason', None)
        if 'unit_of_measure' in updates and updates['unit_of_measure']:
            updates['unit_of_measure'] = updates['unit_of_measure'].value
        
        bom = BOMService.update_bom(db, bom_id, current_user, change_reason=change_reason, **updates)
        return _format_bom_response(bom, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/bom/{bom_id}/approval")
async def approve_bom(
    bom_id: int = Path(..., ge=1),
    data: BOMApprovalRequest = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Approve or reject BOM
    DC_BOM_001: VGK/EA access
    """
    try:
        bom = BOMService.approve_bom(
            db, bom_id, current_user, 
            data.action if data else 'approve',
            data.remarks if data else None
        )
        return _format_bom_response(bom, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/bom/{bom_id}/copy")
async def copy_bom(
    bom_id: int = Path(..., ge=1),
    data: BOMCopyRequest = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Copy BOM to create new version
    DC_BOM_001: VGK/EA access
    """
    try:
        new_bom = BOMService.copy_bom(
            db, bom_id, current_user,
            new_bom_name=data.new_bom_name if data else None,
            notes=data.notes if data else None
        )
        return _format_bom_response(new_bom, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/bom/{bom_id}/line-items")
async def add_bom_line_item(
    bom_id: int = Path(..., ge=1),
    data: BOMLineItemCreate = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add line item to BOM
    DC_BOM_001: VGK/EA access
    """
    try:
        line = BOMService.add_line_item(
            db=db,
            bom_id=bom_id,
            component_id=data.component_id,
            quantity_required=data.quantity_required,
            employee=current_user,
            unit_of_measure=data.unit_of_measure.value if data.unit_of_measure else 'PCS',
            wastage_pct=data.wastage_pct or Decimal('0'),
            sequence_order=data.sequence_order or 1,
            is_optional=data.is_optional or False,
            notes=data.notes,
            change_reason=data.change_reason
        )
        return _format_bom_line_response(line)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/bom/line-items/{line_id}")
async def update_bom_line_item(
    line_id: int = Path(..., ge=1),
    data: BOMLineItemUpdate = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update line item in BOM
    DC_COMPONENT_001: VGK/EA access
    Editing APPROVED BOMs triggers re-approval workflow
    """
    try:
        line = BOMService.update_line_item(
            db=db,
            line_id=line_id,
            employee=current_user,
            quantity_required=data.quantity_required if data else None,
            unit_of_measure=data.unit_of_measure.value if data and data.unit_of_measure else None,
            wastage_pct=data.wastage_pct if data else None,
            sequence_order=data.sequence_order if data else None,
            is_optional=data.is_optional if data else None,
            notes=data.notes if data else None,
            change_reason=data.change_reason if data else None
        )
        return _format_bom_line_response(line)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/bom/line-items/{line_id}")
async def remove_bom_line_item(
    line_id: int = Path(..., ge=1),
    reason: str = Query(None, description="Reason for removal"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Remove line item from BOM
    DC_COMPONENT_001: VGK/EA access
    Removing from APPROVED BOMs triggers re-approval workflow
    """
    try:
        BOMService.remove_line_item(db, line_id, current_user, reason=reason)
        return {"success": True, "message": "Line item removed"}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


def _format_bom_response(bom, include_lines=False):
    """Format BOM for API response"""
    from app.models.staff_accounts import StockItemMaster, AssociatedCompany
    from app.models.staff import StaffEmployee
    
    finished_product = None
    company = None
    created_by = None
    approved_by = None
    
    if hasattr(bom, 'finished_product') and bom.finished_product:
        finished_product = bom.finished_product
    if hasattr(bom, 'company') and bom.company:
        company = bom.company
    if hasattr(bom, 'created_by') and bom.created_by:
        created_by = bom.created_by
    if hasattr(bom, 'approved_by') and bom.approved_by:
        approved_by = bom.approved_by
    
    response = {
        "id": bom.id,
        "bom_code": bom.bom_code,
        "bom_name": bom.bom_name,
        "description": bom.description,
        "company_id": bom.company_id,
        "company_name": company.company_name if company else None,
        "finished_product_id": bom.finished_product_id,
        "finished_product_name": finished_product.item_name if finished_product else None,
        "finished_product_code": finished_product.item_code if finished_product else None,
        "standard_qty": str(bom.standard_qty),
        "unit_of_measure": bom.unit_of_measure,
        "version": bom.version,
        "effective_from": str(bom.effective_from) if bom.effective_from else None,
        "effective_to": str(bom.effective_to) if bom.effective_to else None,
        "status": bom.status,
        "estimated_cost": str(bom.estimated_cost or 0),
        "estimated_time_hours": str(bom.estimated_time_hours) if bom.estimated_time_hours else None,
        "notes": bom.notes,
        "is_active": bom.is_active,
        "created_by_name": f"{created_by.first_name} {created_by.last_name}" if created_by else None,
        "approved_by_name": f"{approved_by.first_name} {approved_by.last_name}" if approved_by else None,
        "approved_at": str(bom.approved_at) if bom.approved_at else None,
        "created_at": str(bom.created_at) if bom.created_at else None
    }
    
    if include_lines and hasattr(bom, 'line_items'):
        response["line_items"] = [_format_bom_line_response(line) for line in bom.line_items]
    
    return response


def _format_bom_line_response(line):
    """Format BOM Line Item for API response"""
    component = None
    if hasattr(line, 'component') and line.component:
        component = line.component
    
    return {
        "id": line.id,
        "bom_id": line.bom_id,
        "component_id": line.component_id,
        "component_name": component.item_name if component else None,
        "component_code": component.item_code if component else None,
        "quantity_required": str(line.quantity_required),
        "unit_of_measure": line.unit_of_measure,
        "wastage_pct": str(line.wastage_pct or 0),
        "unit_cost": str(line.unit_cost or 0),
        "total_cost": str(line.total_cost or 0),
        "sequence_order": line.sequence_order,
        "is_optional": line.is_optional,
        "notes": line.notes
    }


# ==================== MANUFACTURING ENDPOINTS (DC_BOM_001 - Dec 06, 2025) ====================

@router.post("/manufacturing", status_code=status.HTTP_201_CREATED)
async def create_manufacturing_order(
    data: ManufacturingOrderCreate,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new Manufacturing Order
    DC_BOM_001: VGK/EA access
    """
    try:
        order = ManufacturingService.create_order(
            db=db,
            company_id=data.company_id,
            bom_id=data.bom_id,
            planned_qty=data.planned_qty,
            employee=current_user,
            planned_start_date=data.planned_start_date,
            planned_end_date=data.planned_end_date,
            priority=data.priority.value if data.priority else 'NORMAL',
            notes=data.notes
        )
        
        return _format_manufacturing_response(order, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/manufacturing")
async def list_manufacturing_orders(
    company_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List Manufacturing Orders with filters
    DC_BOM_001: VGK/EA access
    """
    try:
        orders, total = ManufacturingService.list_orders(
            db=db,
            employee=current_user,
            company_id=company_id,
            status=status,
            priority=priority,
            start_date=start_date,
            end_date=end_date,
            page=page,
            page_size=page_size
        )
        
        return {
            "success": True,
            "orders": [_format_manufacturing_response(o, include_lines=False) for o in orders],
            "total": total,
            "page": page,
            "page_size": page_size
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/manufacturing/check-stock")
async def check_manufacturing_stock(
    bom_id: int = Query(...),
    company_id: int = Query(...),
    planned_qty: Decimal = Query(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Check stock availability for manufacturing
    DC_BOM_001: VGK/EA access - GET for read-only stock availability check
    NOTE: This route MUST be defined BEFORE /manufacturing/{order_id} to prevent FastAPI from matching "check-stock" as order_id
    """
    try:
        can_manufacture, components = ManufacturingService.check_stock_availability(
            db, bom_id, planned_qty, company_id, current_user
        )
        
        return {
            "success": True,
            "can_manufacture": can_manufacture,
            "bom_id": bom_id,
            "planned_qty": str(planned_qty),
            "components": components,
            "message": "Sufficient stock available" if can_manufacture else "Insufficient stock for some components"
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/manufacturing/{order_id}")
async def get_manufacturing_order(
    order_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get Manufacturing Order by ID with line items
    DC_BOM_001: VGK/EA access
    """
    try:
        order = ManufacturingService.get_order(db, order_id, current_user)
        return _format_manufacturing_response(order, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/manufacturing/{order_id}")
async def update_manufacturing_order(
    order_id: int = Path(..., ge=1),
    data: ManufacturingUpdateRequest = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update a Manufacturing Order - DC Protocol compliant
    DC_BOM_001: VGK/EA access
    PLANNED: Full edit allowed
    APPROVED: Edit triggers re-approval (status → PENDING_APPROVAL)
    IN_PROGRESS: Only notes can be edited (fields disabled in frontend)
    """
    try:
        order = ManufacturingService.update_order(
            db=db,
            order_id=order_id,
            employee=current_user,
            planned_qty=data.planned_qty,
            priority=data.priority,
            planned_start_date=data.planned_start_date,
            planned_end_date=data.planned_end_date,
            notes=data.notes,
            change_reason=data.change_reason
        )
        return _format_manufacturing_response(order, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/manufacturing/{order_id}/start")
async def start_manufacturing(
    order_id: int = Path(..., ge=1),
    data: ManufacturingStartRequest = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Start manufacturing - consume components from stock
    DC_BOM_001: VGK/EA access
    """
    try:
        order = ManufacturingService.start_manufacturing(
            db, order_id, current_user,
            remarks=data.remarks if data else None
        )
        return _format_manufacturing_response(order, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/manufacturing/{order_id}/complete")
async def complete_manufacturing(
    order_id: int = Path(..., ge=1),
    data: ManufacturingCompleteRequest = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Complete manufacturing - add finished goods to stock
    DC_BOM_001: VGK/EA access
    """
    try:
        order = ManufacturingService.complete_manufacturing(
            db=db,
            order_id=order_id,
            actual_qty=data.actual_qty,
            employee=current_user,
            rejected_qty=data.rejected_qty or Decimal('0'),
            remarks=data.remarks
        )
        return _format_manufacturing_response(order, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/manufacturing/{order_id}")
async def delete_manufacturing_order(
    order_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Hard-delete a Manufacturing Order — DC_BOM_DELETE_001
    Only PLANNED or CANCELLED orders can be deleted.
    PLANNED  : no stock consumed — safe to delete.
    CANCELLED: stock already reversed — safe to delete.
    COMPLETED / IN_PROGRESS: rejected (stock impact exists).
    """
    try:
        order_number = ManufacturingService.delete_order(db, order_id, current_user)
        return {"success": True, "message": f"Manufacturing Order {order_number} deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/manufacturing/{order_id}/cancel")
async def cancel_manufacturing(
    order_id: int = Path(..., ge=1),
    data: ManufacturingCancelRequest = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Cancel manufacturing order - reverse stock if started
    DC_BOM_001: VGK/EA access
    """
    try:
        order = ManufacturingService.cancel_order(
            db, order_id, current_user,
            reason=data.reason if data else "Cancelled by user"
        )
        return _format_manufacturing_response(order, include_lines=True)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== MANUFACTURING LINE ITEM ENDPOINTS (DC_COMPONENT_001) ====================

@router.post("/manufacturing/{order_id}/line-items")
async def add_manufacturing_line_item(
    order_id: int = Path(..., ge=1),
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add additional material to Manufacturing Order
    DC_COMPONENT_001: VGK/EA access
    Marks new materials as is_additional=True for tracking
    """
    try:
        line = ManufacturingService.add_line_item(
            db=db,
            order_id=order_id,
            component_id=data.get('component_id'),
            planned_qty=Decimal(str(data.get('planned_qty', 0))),
            employee=current_user,
            unit_of_measure=data.get('unit_of_measure', 'PCS'),
            notes=data.get('notes')
        )
        return _format_manufacturing_line_response(line)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/manufacturing/line-items/{line_id}")
async def update_manufacturing_line_item(
    line_id: int = Path(..., ge=1),
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Update line item in Manufacturing Order
    DC_COMPONENT_001: VGK/EA access
    Only allowed for PLANNED/IN_PROGRESS orders and PENDING line items
    """
    try:
        planned_qty = Decimal(str(data.get('planned_qty'))) if data.get('planned_qty') is not None else None
        
        line = ManufacturingService.update_line_item(
            db=db,
            line_id=line_id,
            employee=current_user,
            planned_qty=planned_qty,
            unit_of_measure=data.get('unit_of_measure'),
            notes=data.get('notes')
        )
        return _format_manufacturing_line_response(line)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/manufacturing/line-items/{line_id}")
async def remove_manufacturing_line_item(
    line_id: int = Path(..., ge=1),
    reason: str = Query(None, description="Reason for removal"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Remove additional material from Manufacturing Order
    DC_COMPONENT_001: VGK/EA access
    Only additional materials (is_additional=True) can be removed
    BOM-originated materials cannot be removed
    """
    try:
        ManufacturingService.remove_line_item(db, line_id, current_user, reason=reason)
        return {"success": True, "message": "Line item removed"}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


def _format_manufacturing_response(order, include_lines=False):
    """Format Manufacturing Order for API response"""
    bom = None
    finished_product = None
    company = None
    created_by = None
    approved_by = None
    started_by = None
    completed_by = None
    
    if hasattr(order, 'bom') and order.bom:
        bom = order.bom
    if hasattr(order, 'finished_product') and order.finished_product:
        finished_product = order.finished_product
    if hasattr(order, 'company') and order.company:
        company = order.company
    if hasattr(order, 'created_by') and order.created_by:
        created_by = order.created_by
    if hasattr(order, 'approved_by') and order.approved_by:
        approved_by = order.approved_by
    if hasattr(order, 'started_by') and order.started_by:
        started_by = order.started_by
    if hasattr(order, 'completed_by') and order.completed_by:
        completed_by = order.completed_by
    
    response = {
        "id": order.id,
        "order_number": order.order_number,
        "company_id": order.company_id,
        "company_name": company.company_name if company else None,
        "bom_id": order.bom_id,
        "bom_name": bom.bom_name if bom else None,
        "bom_code": bom.bom_code if bom else None,
        "finished_product_id": order.finished_product_id,
        "finished_product_name": finished_product.item_name if finished_product else None,
        "finished_product_code": finished_product.item_code if finished_product else None,
        "planned_qty": str(order.planned_qty),
        "actual_qty": str(order.actual_qty or 0),
        "rejected_qty": str(order.rejected_qty or 0),
        "unit_of_measure": order.unit_of_measure,
        "planned_start_date": str(order.planned_start_date) if order.planned_start_date else None,
        "planned_end_date": str(order.planned_end_date) if order.planned_end_date else None,
        "actual_start_date": str(order.actual_start_date) if order.actual_start_date else None,
        "actual_end_date": str(order.actual_end_date) if order.actual_end_date else None,
        "status": order.status,
        "priority": order.priority,
        "estimated_cost": str(order.estimated_cost or 0),
        "actual_cost": str(order.actual_cost or 0),
        "notes": order.notes,
        "remarks": order.remarks,
        "created_by_name": f"{created_by.first_name} {created_by.last_name}" if created_by else None,
        "approved_by_name": f"{approved_by.first_name} {approved_by.last_name}" if approved_by else None,
        "started_by_name": f"{started_by.first_name} {started_by.last_name}" if started_by else None,
        "completed_by_name": f"{completed_by.first_name} {completed_by.last_name}" if completed_by else None,
        "approved_at": str(order.approved_at) if order.approved_at else None,
        "started_at": str(order.started_at) if order.started_at else None,
        "completed_at": str(order.completed_at) if order.completed_at else None,
        "created_at": str(order.created_at) if order.created_at else None
    }
    
    if include_lines and hasattr(order, 'line_items'):
        response["line_items"] = [_format_manufacturing_line_response(line) for line in order.line_items]
    
    return response


def _format_manufacturing_line_response(line):
    """Format Manufacturing Order Line for API response"""
    component = None
    consumed_by = None
    added_by = None
    updated_by = None
    
    if hasattr(line, 'component') and line.component:
        component = line.component
    if hasattr(line, 'consumed_by') and line.consumed_by:
        consumed_by = line.consumed_by
    if hasattr(line, 'added_by') and line.added_by:
        added_by = line.added_by
    if hasattr(line, 'updated_by') and line.updated_by:
        updated_by = line.updated_by
    
    return {
        "id": line.id,
        "manufacturing_order_id": line.manufacturing_order_id,
        "component_id": line.component_id,
        "component_name": component.item_name if component else None,
        "component_code": component.item_code if component else None,
        "planned_qty": str(line.planned_qty),
        "actual_qty_consumed": str(line.actual_qty_consumed or 0),
        "wastage_qty": str(line.wastage_qty or 0),
        "returned_qty": str(line.returned_qty or 0),
        "unit_of_measure": line.unit_of_measure,
        "planned_cost": str(line.planned_cost or 0),
        "actual_cost": str(line.actual_cost or 0),
        "status": line.status,
        "is_additional": getattr(line, 'is_additional', False),
        "added_by_id": getattr(line, 'added_by_id', None),
        "added_by_name": f"{added_by.first_name} {added_by.last_name}" if added_by else None,
        "added_at": str(line.added_at) if getattr(line, 'added_at', None) else None,
        "updated_by_id": getattr(line, 'updated_by_id', None),
        "updated_by_name": f"{updated_by.first_name} {updated_by.last_name}" if updated_by else None,
        "consumed_at": str(line.consumed_at) if line.consumed_at else None,
        "consumed_by_name": f"{consumed_by.first_name} {consumed_by.last_name}" if consumed_by else None,
        "notes": line.notes
    }


# ==================== PROCUREMENT PLANNING ENDPOINTS (DC_PROCUREMENT_001) ====================

@router.get("/manufacturing/{order_id}/material-status")
async def get_material_status(
    order_id: int = Path(..., ge=1),
    company_id: int = Query(..., description="DC Protocol REQUIRED: Company ID for segregation validation"),
    force_refresh: bool = Query(False, description="Force refresh the material status"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get material status for a manufacturing order with detailed component breakdown
    DC_PROCUREMENT_001: VGK/EA/Accounts access with MANDATORY company segregation
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        result = ProcurementPlanningService.get_material_status(
            db, order_id, current_user, company_id=company_id, force_refresh=force_refresh
        )
        return {"success": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/manufacturing/{order_id}/material-status/refresh")
async def refresh_material_status(
    order_id: int = Path(..., ge=1),
    company_id: int = Query(..., description="DC Protocol REQUIRED: Company ID for segregation validation"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Force refresh material status for a manufacturing order
    DC_PROCUREMENT_001: VGK/EA/Accounts access with MANDATORY company segregation
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        result = ProcurementPlanningService.refresh_material_status(
            db, order_id, current_user, company_id=company_id
        )
        return {"success": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/low-stock")
async def get_low_stock_items(
    company_id: int = Query(None, description="Filter by company"),
    category: str = Query(None, description="Filter by category"),
    search: str = Query(None, description="Search by item code or name"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get stock items below reorder level with company-wise filtering
    DC_PROCUREMENT_001: VGK/EA/Accounts access
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        skip = (page - 1) * limit
        items, total = ProcurementPlanningService.get_low_stock_items(
            db, current_user,
            company_id=company_id,
            category=category,
            search=search,
            skip=skip,
            limit=limit
        )
        return {
            "success": True,
            "items": items,
            "total": total,
            "page": page,
            "pages": (total + limit - 1) // limit
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/requirements")
async def list_procurement_requirements(
    company_id: int = Query(None, description="Filter by company"),
    status: str = Query(None, description="Filter by status"),
    source_type: str = Query(None, description="Filter by source type"),
    priority: str = Query(None, description="Filter by priority"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List procurement requirements with filters
    DC_PROCUREMENT_001: VGK/EA/Accounts access
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        skip = (page - 1) * limit
        requirements, total = ProcurementPlanningService.get_procurement_requirements(
            db, current_user,
            company_id=company_id,
            status=status,
            source_type=source_type,
            priority=priority,
            skip=skip,
            limit=limit
        )
        return {
            "success": True,
            "requirements": [r.to_dict() for r in requirements],
            "total": total,
            "page": page,
            "pages": (total + limit - 1) // limit
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/requirements")
async def create_procurement_requirement(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a procurement requirement from shortage data
    DC_PROCUREMENT_001: VGK/EA/Accounts access
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        requirement = ProcurementPlanningService.create_procurement_requirement(
            db=db,
            company_id=data.get('company_id'),
            source_type=data.get('source_type', 'MANUAL'),
            employee=current_user,
            priority=data.get('priority', 'NORMAL'),
            notes=data.get('notes'),
            line_items=data.get('line_items', [])
        )
        return {
            "success": True,
            "message": f"Procurement requirement {requirement.requirement_number} created",
            "requirement": requirement.to_dict()
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/requirements/{requirement_id}/trigger")
async def trigger_procurement(
    requirement_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Mark a procurement requirement as triggered for purchase
    DC_PROCUREMENT_001: VGK/EA/Accounts access
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        requirement = ProcurementPlanningService.trigger_procurement(
            db, requirement_id, current_user
        )
        return {
            "success": True,
            "message": f"Procurement requirement {requirement.requirement_number} triggered",
            "requirement": requirement.to_dict()
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/aggregated-shortages")
async def get_aggregated_shortages(
    company_id: int = Query(None, description="Filter by company"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get aggregated material shortages from all active Manufacturing and Partner Orders
    DC_PROCUREMENT_001: VGK/EA/Accounts access
    """
    try:
        from app.services.staff_accounts_service import ProcurementPlanningService
        
        result = ProcurementPlanningService.get_aggregated_shortages(
            db, current_user, company_id=company_id
        )
        return {"success": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== SALES INVOICE ENDPOINTS ====================

@router.post("/sales-invoices")
async def create_sales_invoice(
    invoice_data: dict = Body(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Create a new sales invoice with line items"""
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        print(f"[DC-RBAC] Sales Invoice Create - Access Denied: User role '{user_role}' not in allowed roles {ACCOUNTS_ALLOWED_ROLES}")
        raise HTTPException(status_code=403, detail=f"Access denied. Your role '{user_role}' is not authorized. Only VGK4U, EA, and Accounts roles can access this feature.")
    
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        
        invoice = SalesInvoiceService.create_invoice(
            db=db,
            data=invoice_data,
            created_by_id=current_user.id
        )
        
        return {
            "success": True,
            "message": f"Sales invoice {invoice.invoice_number} created successfully",
            "invoice": invoice.to_dict()
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sales-invoices")
async def list_sales_invoices(
    request: Request,
    company_id: Optional[int] = None,
    status: str = None,
    customer_type: str = None,
    from_date: date = None,
    to_date: date = None,
    page: int = 1,
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List sales invoices — company_id optional; omit to list across all companies"""
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        print(f"[DC-RBAC] Sales Invoice List - Access Denied: User role '{user_role}' not in allowed roles {ACCOUNTS_ALLOWED_ROLES}")
        raise HTTPException(status_code=403, detail=f"Access denied. Your role '{user_role}' is not authorized. Only VGK4U, EA, and Accounts roles can access this feature.")
    
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        
        result = SalesInvoiceService.list_invoices(
            db=db,
            company_id=company_id,
            status=status,
            customer_type=customer_type,
            from_date=from_date,
            to_date=to_date,
            page=page,
            page_size=limit
        )
        
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sales-invoices/billing-companies")
async def list_sales_invoice_billing_companies(
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List all associated companies available as billing companies for sales invoices"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import AssociatedCompany
        companies = db.query(AssociatedCompany).filter(AssociatedCompany.is_active == True).order_by(AssociatedCompany.company_name).all()
        return {
            "companies": [
                {"id": c.id, "name": c.company_name, "gst_number": c.gst_number, "state": c.state}
                for c in companies
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/purchase-invoices/pending-receipt")
async def get_purchase_invoices_pending_receipt(
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-001 + DC-DISPATCH-EXTRA-001: List CONFIRMED purchase invoices with outstanding receipt qty.
    Pending qty per line = purchase_pending_line_config.pending_qty (if set) else invoice_line.quantity,
    minus already-received qty. Extra items from purchase_pending_extra_items are appended.
    """
    try:
        from app.models.staff_accounts import (
            PurchaseInvoiceUpload, PurchaseInvoiceLineItem, PurchaseIntakeItem,
            PurchasePendingLineConfig, PurchasePendingExtraItem
        )
        from sqlalchemy import func

        q = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.status == 'CONFIRMED',
            PurchaseInvoiceUpload.track_physical_receipt == True
        )
        if company_id:
            q = q.filter(PurchaseInvoiceUpload.company_id == company_id)

        invoices = q.order_by(PurchaseInvoiceUpload.vendor_invoice_date.desc()).all()

        result = []
        for inv in invoices:
            lines = db.query(PurchaseInvoiceLineItem).filter(
                PurchaseInvoiceLineItem.upload_id == inv.id
            ).all()

            # Pre-load line configs for this invoice
            line_configs = {
                cfg.invoice_line_id: float(cfg.pending_qty)
                for cfg in db.query(PurchasePendingLineConfig).filter(
                    PurchasePendingLineConfig.invoice_id == inv.id
                ).all()
            }

            line_data = []
            total_pending_qty = 0
            total_pending_value = 0.0

            for ln in lines:
                received_row = db.execute(
                    _tv_text("""
                        SELECT COALESCE(SUM(pii.received_qty), 0)
                        FROM purchase_intake_items pii
                        JOIN purchase_intake_batches pib ON pii.batch_id = pib.id
                        WHERE pii.purchase_line_id = :lid
                          AND pib.intake_status NOT IN ('CANCELLED', 'REJECTED')
                    """),
                    {"lid": ln.id}
                ).fetchone()
                received_qty = float(received_row[0]) if received_row else 0.0
                ordered_qty = float(ln.quantity or 0)
                # DC-DISPATCH-EXTRA-001: use configured pending qty as the target ceiling
                configured_qty = line_configs.get(ln.id, ordered_qty)
                remaining_qty = max(configured_qty - received_qty, 0.0)
                unit_rate = float(ln.unit_rate) if ln.unit_rate else 0.0
                pending_value = round(remaining_qty * unit_rate, 2)

                if remaining_qty > 0:
                    total_pending_qty += remaining_qty
                    total_pending_value += pending_value

                line_data.append({
                    'id': ln.id,
                    'line_number': ln.line_number,
                    'item_id': ln.item_id,
                    'item_code': ln.item_code,
                    'item_description': ln.item_description,
                    'hsn_code': ln.hsn_code,
                    'unit_of_measure': ln.unit_of_measure,
                    'unit_rate': unit_rate,
                    'ordered_qty': ordered_qty,
                    'configured_pending_qty': configured_qty,
                    'has_config': ln.id in line_configs,
                    'received_qty': received_qty,
                    'pending_qty': remaining_qty,
                    'pending_value': pending_value,
                    'gst_rate': float(ln.gst_rate) if ln.gst_rate else 0,
                })

            # DC-DISPATCH-EXTRA-001: include extra items
            extra_items_raw = db.query(PurchasePendingExtraItem).filter(
                PurchasePendingExtraItem.invoice_id == inv.id
            ).order_by(PurchasePendingExtraItem.id).all()
            extra_items = []
            for ei in extra_items_raw:
                ei_pqty = float(ei.pending_qty or 0)
                ei_recv = float(ei.received_qty or 0)
                ei_remain = max(ei_pqty - ei_recv, 0.0)
                if ei_remain > 0:
                    total_pending_qty += ei_remain
                extra_items.append({
                    'id': ei.id,
                    'item_id': ei.item_id,
                    'item_code': ei.item_code,
                    'item_description': ei.item_description,
                    'unit_of_measure': ei.unit_of_measure,
                    'pending_qty': ei_pqty,
                    'received_qty': ei_recv,
                    'remaining_qty': ei_remain,
                    'receipt_status': ei.receipt_status,
                    'notes': ei.notes,
                })

            has_outstanding = total_pending_qty > 0
            if has_outstanding or extra_items:
                vendor_name = None
                if inv.vendor_id:
                    vrow = db.execute(
                        _tv_text("SELECT vendor_name FROM vendor_master WHERE id = :vid LIMIT 1"),
                        {"vid": inv.vendor_id}
                    ).fetchone()
                    if vrow:
                        vendor_name = vrow[0]

                result.append({
                    'id': inv.id,
                    'upload_number': inv.upload_number,
                    'vendor_invoice_no': inv.vendor_invoice_no,
                    'vendor_invoice_date': inv.vendor_invoice_date.isoformat() if inv.vendor_invoice_date else None,
                    'vendor_id': inv.vendor_id,
                    'vendor_name': vendor_name,
                    'company_id': inv.company_id,
                    'grand_total': float(inv.grand_total) if inv.grand_total else 0,
                    'total_pending_qty': round(total_pending_qty, 3),
                    'total_pending_value': round(total_pending_value, 2),
                    'lines': line_data,
                    'extra_items': extra_items,
                })

        return JSONResponse(content={"success": True, "data": result, "total": len(result)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-invoices/search-for-pending")
async def search_purchase_invoices_for_pending(
    company_id: Optional[int] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-DISPATCH-003: Search CONFIRMED purchase invoices to add to Pending Receipt tab."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import PurchaseInvoiceUpload, PurchaseInvoiceLineItem, VendorMaster
        query = db.query(PurchaseInvoiceUpload).outerjoin(
            VendorMaster, VendorMaster.id == PurchaseInvoiceUpload.vendor_id
        ).filter(PurchaseInvoiceUpload.status == 'CONFIRMED')
        if company_id:
            query = query.filter(PurchaseInvoiceUpload.company_id == company_id)
        if q and q.strip():
            term = f"%{q.strip()}%"
            query = query.filter(
                (PurchaseInvoiceUpload.vendor_invoice_no.ilike(term)) |
                (PurchaseInvoiceUpload.upload_number.ilike(term)) |
                (VendorMaster.vendor_name.ilike(term))
            )
        uploads = query.order_by(PurchaseInvoiceUpload.created_at.desc()).limit(20).all()
        results = []
        for u in uploads:
            vendor = db.query(VendorMaster).filter(VendorMaster.id == u.vendor_id).first() if u.vendor_id else None
            lines = db.query(PurchaseInvoiceLineItem).filter(
                PurchaseInvoiceLineItem.upload_id == u.id
            ).order_by(PurchaseInvoiceLineItem.line_number).all()
            line_data = [
                {
                    "id": ln.id,
                    "line_number": ln.line_number,
                    "item_code": ln.item_code or "",
                    "item_description": ln.item_description or "",
                    "hsn_code": ln.hsn_code or "",
                    "unit_of_measure": ln.unit_of_measure or "PCS",
                    "quantity": float(ln.quantity or 0),
                    "unit_rate": float(ln.unit_rate or 0),
                }
                for ln in lines
            ]
            results.append({
                "id": u.id,
                "upload_number": u.upload_number or f"PIU-{u.id}",
                "vendor_invoice_no": u.vendor_invoice_no or "",
                "vendor_name": vendor.vendor_name if vendor else (u.extracted_vendor_name or ""),
                "vendor_invoice_date": str(u.vendor_invoice_date) if u.vendor_invoice_date else "",
                "total_amount": float(u.total_amount or 0),
                "track_physical_receipt": bool(u.track_physical_receipt),
                "line_items": line_data,
            })
        return JSONResponse(content={"success": True, "invoices": results})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/sales-invoices/{invoice_id}/toggle-dispatch-tracking")
async def toggle_sales_dispatch_tracking(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-003 + DC-DISPATCH-EXTRA-001: Toggle physical dispatch tracking for a sales invoice.
    Body: {enable, line_overrides: [{line_id, pending_qty}], extra_items: [{item_id, item_description, item_code, uom, pending_qty, notes}]}
    line_overrides and extra_items are replaced atomically on each enable.
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesInvoice, SalesInvoiceLineItem, SalesPendingLineConfig, SalesPendingExtraItem
        inv = db.query(SalesInvoice).filter(SalesInvoice.id == invoice_id).first()
        if not inv:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if inv.status != 'CONFIRMED':
            raise HTTPException(status_code=400, detail="Only CONFIRMED invoices can be tracked")
        enable = body.get('enable', True)
        inv.track_physical_dispatch = bool(enable)
        if enable and inv.dispatch_status == 'FULLY_DISPATCHED':
            inv.dispatch_status = 'NOT_DISPATCHED'

        if enable:
            # DC-DISPATCH-EXTRA-001: replace line configs atomically
            db.query(SalesPendingLineConfig).filter(SalesPendingLineConfig.invoice_id == invoice_id).delete()
            for ov in body.get('line_overrides', []):
                lid = ov.get('line_id')
                pqty = float(ov.get('pending_qty', 0) or 0)
                if not lid or pqty <= 0:
                    continue
                ln = db.query(SalesInvoiceLineItem).filter(
                    SalesInvoiceLineItem.id == lid,
                    SalesInvoiceLineItem.invoice_id == invoice_id
                ).first()
                if not ln:
                    continue
                inv_qty = float(ln.quantity or 0)
                if pqty > inv_qty:
                    raise HTTPException(status_code=400, detail=f"Pending qty ({pqty}) exceeds invoiced qty ({inv_qty}) for line {lid}")
                db.add(SalesPendingLineConfig(
                    company_id=inv.company_id, invoice_id=invoice_id, invoice_line_id=lid,
                    pending_qty=pqty, created_by_id=current_user.id
                ))

            # DC-DISPATCH-EXTRA-001: replace extra items atomically
            db.query(SalesPendingExtraItem).filter(SalesPendingExtraItem.invoice_id == invoice_id).delete()
            for ei in body.get('extra_items', []):
                desc = (ei.get('item_description') or '').strip()
                pqty = float(ei.get('pending_qty', 0) or 0)
                uom = (ei.get('uom') or 'PCS').strip()
                if not desc or pqty <= 0:
                    continue
                db.add(SalesPendingExtraItem(
                    company_id=inv.company_id, invoice_id=invoice_id,
                    item_id=ei.get('item_id') or None,
                    item_description=desc,
                    item_code=(ei.get('item_code') or '').strip() or None,
                    unit_of_measure=uom, pending_qty=pqty,
                    dispatched_qty=0, dispatch_status='NOT_DISPATCHED',
                    notes=(ei.get('notes') or '').strip() or None,
                    created_by_id=current_user.id
                ))

        db.commit()
        return {"success": True, "track_physical_dispatch": inv.track_physical_dispatch, "invoice_id": invoice_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.post("/purchase-invoices/{invoice_id}/toggle-receipt-tracking")
async def toggle_purchase_receipt_tracking(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-003 + DC-DISPATCH-EXTRA-001: Toggle physical receipt tracking for a purchase invoice.
    Body: {enable, line_overrides: [{line_id, pending_qty}], extra_items: [{item_id, item_description, item_code, uom, pending_qty, notes}]}
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import PurchaseInvoiceUpload, PurchaseInvoiceLineItem, PurchasePendingLineConfig, PurchasePendingExtraItem
        inv = db.query(PurchaseInvoiceUpload).filter(PurchaseInvoiceUpload.id == invoice_id).first()
        if not inv:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if inv.status != 'CONFIRMED':
            raise HTTPException(status_code=400, detail="Only CONFIRMED invoices can be tracked")
        enable = body.get('enable', True)
        inv.track_physical_receipt = bool(enable)

        if enable:
            # DC-DISPATCH-EXTRA-001: replace line configs atomically
            db.query(PurchasePendingLineConfig).filter(PurchasePendingLineConfig.invoice_id == invoice_id).delete()
            for ov in body.get('line_overrides', []):
                lid = ov.get('line_id')
                pqty = float(ov.get('pending_qty', 0) or 0)
                if not lid or pqty <= 0:
                    continue
                ln = db.query(PurchaseInvoiceLineItem).filter(
                    PurchaseInvoiceLineItem.id == lid,
                    PurchaseInvoiceLineItem.upload_id == invoice_id
                ).first()
                if not ln:
                    continue
                inv_qty = float(ln.quantity or 0)
                if pqty > inv_qty:
                    raise HTTPException(status_code=400, detail=f"Pending qty ({pqty}) exceeds ordered qty ({inv_qty}) for line {lid}")
                db.add(PurchasePendingLineConfig(
                    company_id=inv.company_id, invoice_id=invoice_id, invoice_line_id=lid,
                    pending_qty=pqty, created_by_id=current_user.id
                ))

            # DC-DISPATCH-EXTRA-001: replace extra items atomically
            db.query(PurchasePendingExtraItem).filter(PurchasePendingExtraItem.invoice_id == invoice_id).delete()
            for ei in body.get('extra_items', []):
                desc = (ei.get('item_description') or '').strip()
                pqty = float(ei.get('pending_qty', 0) or 0)
                uom = (ei.get('uom') or 'PCS').strip()
                if not desc or pqty <= 0:
                    continue
                db.add(PurchasePendingExtraItem(
                    company_id=inv.company_id, invoice_id=invoice_id,
                    item_id=ei.get('item_id') or None,
                    item_description=desc,
                    item_code=(ei.get('item_code') or '').strip() or None,
                    unit_of_measure=uom, pending_qty=pqty,
                    received_qty=0, receipt_status='NOT_RECEIVED',
                    notes=(ei.get('notes') or '').strip() or None,
                    created_by_id=current_user.id
                ))

        db.commit()
        return {"success": True, "track_physical_receipt": inv.track_physical_receipt, "invoice_id": invoice_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.get("/sales-invoices/pending-dispatch")
async def get_sales_invoices_pending_dispatch(
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-001 + DC-DISPATCH-EXTRA-001: List CONFIRMED sales invoices with outstanding dispatch qty.
    Pending qty per line = sales_pending_line_config.pending_qty (if set) else invoice_line.quantity,
    minus total dispatched. Extra items from sales_pending_extra_items are appended.
    """
    try:
        from app.models.staff_accounts import (
            SalesInvoice, SalesInvoiceLineItem, SalesDispatchRecord,
            SalesPendingLineConfig, SalesPendingExtraItem
        )

        q = db.query(SalesInvoice).filter(
            SalesInvoice.status == 'CONFIRMED',
            SalesInvoice.track_physical_dispatch == True,
            SalesInvoice.dispatch_status.in_(['NOT_DISPATCHED', 'PARTIALLY_DISPATCHED'])
        )
        if company_id:
            q = q.filter(SalesInvoice.company_id == company_id)

        invoices = q.order_by(SalesInvoice.invoice_date.desc()).all()

        result = []
        for inv in invoices:
            lines = db.query(SalesInvoiceLineItem).filter(
                SalesInvoiceLineItem.invoice_id == inv.id
            ).all()

            # Pre-load line configs for this invoice
            line_configs = {
                cfg.invoice_line_id: float(cfg.pending_qty)
                for cfg in db.query(SalesPendingLineConfig).filter(
                    SalesPendingLineConfig.invoice_id == inv.id
                ).all()
            }

            line_data = []
            total_pending_qty = 0
            total_pending_value = 0.0

            for ln in lines:
                dispatched_row = db.execute(
                    _tv_text("SELECT COALESCE(SUM(dispatched_qty),0) FROM sales_dispatch_records WHERE invoice_line_id = :lid"),
                    {"lid": ln.id}
                ).fetchone()
                dispatched_qty = float(dispatched_row[0]) if dispatched_row else 0.0
                invoiced_qty = float(ln.quantity or 0)
                # DC-DISPATCH-EXTRA-001: use configured pending qty as the target ceiling
                configured_qty = line_configs.get(ln.id, invoiced_qty)
                remaining_qty = max(configured_qty - dispatched_qty, 0.0)

                avg_cost_row = db.execute(
                    _tv_text("""
                        SELECT CASE WHEN balance_qty > 0 THEN balance_value / balance_qty ELSE 0 END
                        FROM stock_ledger
                        WHERE company_id = :cid AND item_id = :iid
                        ORDER BY transaction_date DESC, id DESC LIMIT 1
                    """),
                    {"cid": inv.company_id, "iid": ln.item_id}
                ).fetchone() if ln.item_id else None
                avg_cost = float(avg_cost_row[0]) if avg_cost_row and avg_cost_row[0] else 0.0
                pending_value = round(remaining_qty * avg_cost, 2)

                if remaining_qty > 0:
                    total_pending_qty += remaining_qty
                    total_pending_value += pending_value

                line_data.append({
                    'id': ln.id,
                    'line_number': ln.line_number,
                    'item_id': ln.item_id,
                    'item_code': ln.item_code,
                    'item_description': ln.item_description,
                    'hsn_code': ln.hsn_code,
                    'unit_of_measure': ln.unit_of_measure,
                    'sale_rate': float(ln.unit_rate) if ln.unit_rate else 0,
                    'avg_cost': avg_cost,
                    'invoiced_qty': invoiced_qty,
                    'configured_pending_qty': configured_qty,
                    'has_config': ln.id in line_configs,
                    'dispatched_qty': dispatched_qty,
                    'pending_qty': remaining_qty,
                    'pending_value': pending_value,
                })

            # DC-DISPATCH-EXTRA-001: include extra items
            extra_items_raw = db.query(SalesPendingExtraItem).filter(
                SalesPendingExtraItem.invoice_id == inv.id
            ).order_by(SalesPendingExtraItem.id).all()
            extra_items = []
            for ei in extra_items_raw:
                ei_pqty = float(ei.pending_qty or 0)
                ei_disp = float(ei.dispatched_qty or 0)
                ei_remain = max(ei_pqty - ei_disp, 0.0)
                if ei_remain > 0:
                    total_pending_qty += ei_remain
                extra_items.append({
                    'id': ei.id,
                    'item_id': ei.item_id,
                    'item_code': ei.item_code,
                    'item_description': ei.item_description,
                    'unit_of_measure': ei.unit_of_measure,
                    'pending_qty': ei_pqty,
                    'dispatched_qty': ei_disp,
                    'remaining_qty': ei_remain,
                    'dispatch_status': ei.dispatch_status,
                    'notes': ei.notes,
                })

            has_outstanding = total_pending_qty > 0
            if has_outstanding or extra_items:
                result.append({
                    'id': inv.id,
                    'invoice_number': inv.invoice_number,
                    'invoice_date': inv.invoice_date.isoformat() if inv.invoice_date else None,
                    'customer_name': inv.customer_name,
                    'customer_phone': inv.customer_phone,
                    'company_id': inv.company_id,
                    'grand_total': float(inv.grand_total) if inv.grand_total else 0,
                    'dispatch_status': inv.dispatch_status,
                    'total_pending_qty': round(total_pending_qty, 3),
                    'total_pending_value': round(total_pending_value, 2),
                    'lines': line_data,
                    'extra_items': extra_items,
                })

        return JSONResponse(content={"success": True, "data": result, "total": len(result)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/pending-dispatch/summary")
async def get_pending_dispatch_summary(
    company_id: Optional[int] = None,
    invoice_type: str = Query("purchase", description="purchase or sales"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-001: Aggregate pending summary — item-wise, day-wise, vendor/customer-wise.
    """
    try:
        from app.models.staff_accounts import (
            PurchaseInvoiceUpload, PurchaseInvoiceLineItem, PurchaseIntakeItem,
            SalesInvoice, SalesInvoiceLineItem, SalesDispatchRecord
        )

        item_map = {}
        day_map = {}
        party_map = {}

        if invoice_type == 'purchase':
            q = db.query(PurchaseInvoiceUpload).filter(
                PurchaseInvoiceUpload.status == 'CONFIRMED',
                PurchaseInvoiceUpload.track_physical_receipt == True  # DC-DISPATCH-003
            )
            if company_id:
                q = q.filter(PurchaseInvoiceUpload.company_id == company_id)
            invoices = q.all()

            for inv in invoices:
                lines = db.query(PurchaseInvoiceLineItem).filter(
                    PurchaseInvoiceLineItem.upload_id == inv.id
                ).all()
                vendor_name = None
                if inv.vendor_id:
                    vr = db.execute(
                        _tv_text("SELECT vendor_name FROM vendor_master WHERE id=:v LIMIT 1"),
                        {"v": inv.vendor_id}
                    ).fetchone()
                    if vr:
                        vendor_name = vr[0]

                for ln in lines:
                    from app.models.staff_accounts import PurchasePendingLineConfig as _PurCfg
                    _cfg = db.query(_PurCfg).filter(
                        _PurCfg.invoice_id == inv.id, _PurCfg.invoice_line_id == ln.id
                    ).first()
                    target_qty = float(_cfg.pending_qty) if _cfg is not None else float(ln.quantity or 0)
                    if target_qty <= 0:
                        continue
                    recv = db.execute(
                        _tv_text("""
                            SELECT COALESCE(SUM(pii.received_qty),0)
                            FROM purchase_intake_items pii
                            JOIN purchase_intake_batches pib ON pii.batch_id = pib.id
                            WHERE pii.purchase_line_id=:lid
                              AND pib.intake_status NOT IN ('CANCELLED','REJECTED')
                        """),
                        {"lid": ln.id}
                    ).scalar() or 0
                    pending_qty = max(target_qty - float(recv), 0)
                    if pending_qty <= 0:
                        continue
                    pval = round(pending_qty * float(ln.unit_rate or 0), 2)
                    day_key = inv.vendor_invoice_date.isoformat() if inv.vendor_invoice_date else 'Unknown'
                    item_key = ln.item_description or 'Unknown Item'
                    party_key = vendor_name or f"Vendor #{inv.vendor_id}"

                    item_map[item_key] = item_map.get(item_key, {'pending_qty': 0, 'pending_value': 0})
                    item_map[item_key]['pending_qty'] += pending_qty
                    item_map[item_key]['pending_value'] += pval

                    day_map[day_key] = day_map.get(day_key, {'invoice_count': 0, 'pending_value': 0, 'invoice_ids': set()})
                    day_map[day_key]['pending_value'] += pval
                    if inv.id not in day_map[day_key]['invoice_ids']:
                        day_map[day_key]['invoice_count'] += 1
                        day_map[day_key]['invoice_ids'].add(inv.id)

                    party_map[party_key] = party_map.get(party_key, {'pending_qty': 0, 'pending_value': 0})
                    party_map[party_key]['pending_qty'] += pending_qty
                    party_map[party_key]['pending_value'] += pval

        else:
            q = db.query(SalesInvoice).filter(
                SalesInvoice.status == 'CONFIRMED',
                SalesInvoice.track_physical_dispatch == True,  # DC-DISPATCH-003
                SalesInvoice.dispatch_status.in_(['NOT_DISPATCHED', 'PARTIALLY_DISPATCHED'])
            )
            if company_id:
                q = q.filter(SalesInvoice.company_id == company_id)
            invoices = q.all()

            for inv in invoices:
                lines = db.query(SalesInvoiceLineItem).filter(
                    SalesInvoiceLineItem.invoice_id == inv.id
                ).all()
                for ln in lines:
                    from app.models.staff_accounts import SalesPendingLineConfig as _SalCfg
                    _cfg = db.query(_SalCfg).filter(
                        _SalCfg.invoice_id == inv.id, _SalCfg.invoice_line_id == ln.id
                    ).first()
                    target_qty = float(_cfg.pending_qty) if _cfg is not None else float(ln.quantity or 0)
                    if target_qty <= 0:
                        continue
                    disp = db.execute(
                        _tv_text("SELECT COALESCE(SUM(dispatched_qty),0) FROM sales_dispatch_records WHERE invoice_line_id=:lid"),
                        {"lid": ln.id}
                    ).scalar() or 0
                    pending_qty = max(target_qty - float(disp), 0)
                    if pending_qty <= 0:
                        continue
                    pval = round(pending_qty * float(ln.unit_rate or 0), 2)

                    day_key = inv.invoice_date.isoformat() if inv.invoice_date else 'Unknown'
                    item_key = ln.item_description or 'Unknown Item'
                    party_key = inv.customer_name or 'Unknown Customer'

                    item_map[item_key] = item_map.get(item_key, {'pending_qty': 0, 'pending_value': 0})
                    item_map[item_key]['pending_qty'] += pending_qty
                    item_map[item_key]['pending_value'] += pval

                    day_map[day_key] = day_map.get(day_key, {'invoice_count': 0, 'pending_value': 0, 'invoice_ids': set()})
                    day_map[day_key]['pending_value'] += pval
                    if inv.id not in day_map[day_key]['invoice_ids']:
                        day_map[day_key]['invoice_count'] += 1
                        day_map[day_key]['invoice_ids'].add(inv.id)

                    party_map[party_key] = party_map.get(party_key, {'pending_qty': 0, 'pending_value': 0})
                    party_map[party_key]['pending_qty'] += pending_qty
                    party_map[party_key]['pending_value'] += pval

        item_wise = sorted(
            [{'item': k, 'pending_qty': round(v['pending_qty'], 3), 'pending_value': round(v['pending_value'], 2)} for k, v in item_map.items()],
            key=lambda x: x['pending_value'], reverse=True
        )
        day_wise = sorted(
            [{'date': k, 'invoice_count': v['invoice_count'], 'pending_value': round(v['pending_value'], 2)} for k, v in day_map.items()],
            key=lambda x: x['date'], reverse=True
        )
        party_wise = sorted(
            [{'party': k, 'pending_qty': round(v['pending_qty'], 3), 'pending_value': round(v['pending_value'], 2)} for k, v in party_map.items()],
            key=lambda x: x['pending_value'], reverse=True
        )

        return JSONResponse(content={
            "success": True,
            "invoice_type": invoice_type,
            "item_wise": item_wise,
            "day_wise": day_wise,
            "party_wise": party_wise,
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/sales-invoices/search-for-pending")
async def search_sales_invoices_for_pending(
    company_id: Optional[int] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-DISPATCH-003: Search CONFIRMED sales invoices to add to Pending Dispatch tab."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesInvoice, SalesInvoiceLineItem
        query = db.query(SalesInvoice).filter(SalesInvoice.status == 'CONFIRMED')
        if company_id:
            query = query.filter(SalesInvoice.company_id == company_id)
        if q and q.strip():
            term = f"%{q.strip()}%"
            query = query.filter(
                (SalesInvoice.invoice_number.ilike(term)) |
                (SalesInvoice.customer_name.ilike(term))
            )
        invoices = query.order_by(SalesInvoice.invoice_date.desc()).limit(20).all()
        results = []
        for inv in invoices:
            lines = db.query(SalesInvoiceLineItem).filter(
                SalesInvoiceLineItem.invoice_id == inv.id
            ).order_by(SalesInvoiceLineItem.line_number).all()
            line_data = [
                {
                    "id": ln.id,
                    "line_number": ln.line_number,
                    "item_code": ln.item_code or "",
                    "item_description": ln.item_description or "",
                    "unit_of_measure": ln.unit_of_measure or "PCS",
                    "quantity": float(ln.quantity or 0),
                    "unit_price": float(ln.unit_rate or 0),
                }
                for ln in lines
            ]
            results.append({
                "id": inv.id,
                "invoice_number": inv.invoice_number or "",
                "customer_name": inv.customer_name or "",
                "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
                "grand_total": float(inv.grand_total or 0),
                "track_physical_dispatch": bool(inv.track_physical_dispatch),
                "line_items": line_data,
            })
        return JSONResponse(content={"success": True, "invoices": results})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/sales-invoices/{invoice_id}")
async def get_sales_invoice(
    invoice_id: int,
    company_id: Optional[int] = Query(None, description="Company ID for data segregation verification (optional)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get sales invoice with line items — company_id optional; when provided, ownership is verified"""
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        print(f"[DC-RBAC] Sales Invoice Get - Access Denied: User role '{user_role}' not in allowed roles {ACCOUNTS_ALLOWED_ROLES}")
        raise HTTPException(status_code=403, detail=f"Access denied. Your role '{user_role}' is not authorized. Only VGK4U, EA, and Accounts roles can access this feature.")
    
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        
        invoice = SalesInvoiceService.get_invoice(db, invoice_id)
        
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        if company_id and invoice.company_id != company_id:
            raise HTTPException(status_code=403, detail="Invoice does not belong to specified company")
        
        result = invoice.to_dict()
        result['line_items'] = [li.to_dict() for li in invoice.line_items]
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sales-invoices/{invoice_id}/confirm")
async def confirm_sales_invoice(
    invoice_id: int,
    confirm_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Confirm sales invoice - updates Stock Ledger, Accounts Receivable, Party Ledger
    DC: Mandatory company_id verification for data segregation
    """
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        print(f"[DC-RBAC] Sales Invoice Confirm - Access Denied: User role '{user_role}' not in allowed roles {ACCOUNTS_ALLOWED_ROLES}")
        raise HTTPException(status_code=403, detail=f"Access denied. Your role '{user_role}' is not authorized. Only VGK4U, EA, and Accounts roles can access this feature.")
    
    company_id = confirm_data.get('company_id')
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required for data segregation")
    
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        from decimal import Decimal
        
        invoice = SalesInvoiceService.get_invoice(db, invoice_id)
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if invoice.company_id != company_id:
            raise HTTPException(status_code=403, detail="Invoice does not belong to specified company")
        
        amount_received = Decimal(str(confirm_data.get('amount_received', 0)))
        
        invoice = SalesInvoiceService.confirm_invoice(
            db=db,
            invoice_id=invoice_id,
            amount_received=amount_received,
            confirmed_by_id=current_user.id
        )

        # DC_PARTNER_STOCK_AUTOSYNC_001: deduct from partner stock when invoice has partner_id
        if invoice.partner_id:
            try:
                from app.services.partner_stock_service import auto_partner_stock_sync
                from app.models.staff_accounts import SalesInvoiceLineItem
                _line_items = db.query(SalesInvoiceLineItem).filter_by(invoice_id=invoice.id).all()
                _sync_items = [
                    {
                        "item_name": li.item_description,
                        "item_code": li.item_code or "",
                        "stock_item_id": li.item_id,
                        "qty": float(li.quantity),
                        "unit_of_measure": li.unit_of_measure or "PCS",
                        "selling_price": float(li.unit_rate) if li.unit_rate else None,
                        "hsn_code": li.hsn_code or "",
                    }
                    for li in _line_items
                ]
                auto_partner_stock_sync(
                    db=db,
                    partner_id=invoice.partner_id,
                    items=_sync_items,
                    adj_type="SALE_OUT",
                    ref_doc_type="STAFF_INVOICE",
                    ref_doc_id=invoice.id,
                    ref_doc_number=invoice.invoice_number,
                    reason=f"Auto: Sale via staff invoice {invoice.invoice_number}",
                    created_by=f"staff:{current_user.emp_code}",
                )
                db.commit()
            except Exception as _e:
                import logging
                logging.getLogger(__name__).warning(f"[AUTO_STOCK_SYNC] Staff invoice hook skipped: {_e}")

        # DC-STOCK-MKT-003: refresh marketplace qty from stock_ledger in background
        _bg_refresh_mkt_qty()
        return {
            "success": True,
            "message": f"Sales invoice {invoice.invoice_number} confirmed successfully",
            "invoice": invoice.to_dict()
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sales-invoices/{invoice_id}/cancel")
async def cancel_sales_invoice(
    invoice_id: int,
    cancel_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Cancel a sales invoice - DC: Mandatory company_id verification for data segregation"""
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        print(f"[DC-RBAC] Sales Invoice Cancel - Access Denied: User role '{user_role}' not in allowed roles {ACCOUNTS_ALLOWED_ROLES}")
        raise HTTPException(status_code=403, detail=f"Access denied. Your role '{user_role}' is not authorized. Only VGK4U, EA, and Accounts roles can access this feature.")
    
    company_id = cancel_data.get('company_id')
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required for data segregation")
    
    reason = cancel_data.get('reason') or cancel_data.get('cancellation_reason')
    if not reason or len(reason) < 5:
        raise HTTPException(status_code=400, detail="Cancellation reason required (min 5 chars)")
    
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        
        invoice = SalesInvoiceService.get_invoice(db, invoice_id)
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if invoice.company_id != company_id:
            raise HTTPException(status_code=403, detail="Invoice does not belong to specified company")
        
        invoice = SalesInvoiceService.cancel_invoice(
            db=db,
            invoice_id=invoice_id,
            reason=reason,
            cancelled_by_id=current_user.id
        )
        
        action_word = "voided" if invoice.status == 'VOIDED' else "cancelled"
        return {
            "success": True,
            "message": f"Sales invoice {invoice.invoice_number} {action_word}",
            "invoice": invoice.to_dict(),
            "status": invoice.status
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sales-invoices/{invoice_id}")
async def delete_sales_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Delete a DRAFT sales invoice permanently.
    DC_SALES_DELETE_001: Restricted to VGK4U Mentor / EA / Accounts roles.
    Only DRAFT invoices are deletable — CONFIRMED and CANCELLED are immutable.
    WVV Protocol: audit trail written before deletion.
    """
    if not is_accounts_allowed_employee(current_user):
        user_role = current_user.role.role_code if current_user.role else 'No role assigned'
        raise HTTPException(
            status_code=403,
            detail=f"Access denied. Only VGK4U Mentor, EA, and Accounts staff can delete invoices."
        )
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        result = SalesInvoiceService.delete_invoice(db, invoice_id, current_user)
        return JSONResponse(content={"success": True, "message": result['message']})
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/sales-invoices/{invoice_id}/line-items")
async def update_sales_invoice_line_items(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Replace line items on a DRAFT sales invoice and recalculate totals"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Not authorized")
    company_id = body.get('company_id')
    line_items = body.get('line_items', [])
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id required")
    header_data = body.get('header_data')
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        inv = SalesInvoiceService.update_draft_line_items(db, invoice_id, company_id, line_items, header_data=header_data)
        return {"success": True, "invoice": inv.to_dict()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sales-invoices/{invoice_id}/apply-coupon")
async def apply_sales_invoice_coupon(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Apply a coupon code to a sales invoice (pre-GST discount)"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    company_id = body.get('company_id')
    coupon_code = body.get('coupon_code', '').strip()
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if not coupon_code:
        raise HTTPException(status_code=400, detail="coupon_code is required")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        invoice = SalesInvoiceService.apply_coupon(db, invoice_id, coupon_code, company_id)
        return {"success": True, "message": f"Coupon '{coupon_code}' applied", "invoice": invoice.to_dict()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sales-invoices/{invoice_id}/coupon")
async def remove_sales_invoice_coupon(
    invoice_id: int,
    company_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Remove coupon from a sales invoice"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        invoice = SalesInvoiceService.remove_coupon(db, invoice_id, company_id)
        return {"success": True, "message": "Coupon removed", "invoice": invoice.to_dict()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sales-coupons")
async def list_sales_coupons(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List all sales coupon codes (global)"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesCouponMaster
        coupons = db.query(SalesCouponMaster).order_by(SalesCouponMaster.created_at.desc()).all()
        return {"success": True, "coupons": [c.to_dict() for c in coupons]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sales-coupons")
async def create_sales_coupon(
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Create a new sales coupon code"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesCouponMaster
        from datetime import date
        coupon_code = (body.get('coupon_code') or '').strip().upper()
        if not coupon_code:
            raise HTTPException(status_code=400, detail="coupon_code required")
        discount_pct = float(body.get('discount_percentage', body.get('discount_pct', 0)))
        if discount_pct <= 0 or discount_pct > 100:
            raise HTTPException(status_code=400, detail="discount_percentage must be between 0.01 and 100")
        existing = db.query(SalesCouponMaster).filter(
            SalesCouponMaster.coupon_code == coupon_code
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Coupon '{coupon_code}' already exists")
        valid_from_raw = body.get('valid_from')
        valid_until_raw = body.get('valid_until')
        coupon = SalesCouponMaster(
            coupon_code=coupon_code,
            discount_percentage=discount_pct,
            description=body.get('description'),
            is_active=body.get('is_active', True),
            valid_from=date.fromisoformat(valid_from_raw) if valid_from_raw else None,
            valid_until=date.fromisoformat(valid_until_raw) if valid_until_raw else None,
            max_uses=body.get('max_uses'),
        )
        db.add(coupon)
        db.commit()
        db.refresh(coupon)
        return {"success": True, "message": f"Coupon '{coupon_code}' created", "coupon": coupon.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sales-coupons/{coupon_id}")
async def delete_sales_coupon(
    coupon_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Delete a sales coupon"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesCouponMaster
        coupon = db.query(SalesCouponMaster).filter(SalesCouponMaster.id == coupon_id).first()
        if not coupon:
            raise HTTPException(status_code=404, detail="Coupon not found")
        db.delete(coupon)
        db.commit()
        return {"success": True, "message": "Coupon deleted"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sales-invoices/{invoice_id}/manual-discount")
async def set_sales_invoice_manual_discount(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Set or clear manual post-GST discount on a sales invoice"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    from decimal import Decimal
    company_id = body.get('company_id')
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    amount = Decimal(str(body.get('amount', 0)))
    note = body.get('note', '')
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        invoice = SalesInvoiceService.set_manual_discount(db, invoice_id, company_id, amount, note)
        return {"success": True, "message": "Manual discount updated", "invoice": invoice.to_dict()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sales-invoices/{invoice_id}/document-type")
async def set_sales_invoice_document_type(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Switch invoice document type between estimate and tax_invoice"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    company_id = body.get('company_id')
    document_type = body.get('document_type', 'tax_invoice')
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        invoice = SalesInvoiceService.set_document_type(db, invoice_id, company_id, document_type)
        return {"success": True, "message": f"Document type set to '{document_type}'", "invoice": invoice.to_dict()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sales-invoices/{invoice_id}/billing-company")
async def set_sales_invoice_billing_company(
    invoice_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Set the billing company for a sales invoice"""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    company_id = body.get('company_id')
    billing_company_id = body.get('billing_company_id')
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        invoice = SalesInvoiceService.set_billing_company(db, invoice_id, company_id, billing_company_id)
        return {"success": True, "message": "Billing company updated", "invoice": invoice.to_dict()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sales-invoices/{invoice_id}/payments")
async def record_invoice_payment(
    invoice_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-SALES-PAY: Record a payment transaction against a sales invoice.
    Updates invoice amount_received, balance_due, and payment_status.
    """
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")

    company_id = data.get('company_id')
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")

    amount = data.get('amount')
    if not amount or float(amount) <= 0:
        raise HTTPException(status_code=400, detail="amount must be greater than zero")

    payment_date = data.get('payment_date')
    if not payment_date:
        from datetime import date
        payment_date = str(date.today())

    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        from decimal import Decimal
        invoice, payment = SalesInvoiceService.record_payment(
            db=db,
            invoice_id=invoice_id,
            company_id=int(company_id),
            payment_date=payment_date,
            amount=Decimal(str(amount)),
            payment_mode=data.get('payment_mode', 'CASH'),
            reference_number=data.get('reference_number'),
            notes=data.get('notes'),
            created_by_id=current_user.id,
        )
        return {
            "success": True,
            "message": "Payment recorded successfully",
            "payment": payment.to_dict(),
            "invoice": invoice.to_dict(),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sales-invoices/{invoice_id}/payments")
async def list_invoice_payments(
    invoice_id: int,
    company_id: int = Query(..., description="Company ID for data segregation"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-SALES-PAY: List all payment transactions for an invoice."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        payments = SalesInvoiceService.list_payments(db, invoice_id, company_id)
        return {"payments": [p.to_dict() for p in payments]}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sales-invoices/{invoice_id}/pdf")
async def download_sales_invoice_pdf(
    invoice_id: int,
    company_id: int = Query(..., description="Company ID for data segregation"),
    mode: str = Query('tax_invoice', description="PDF mode: 'estimate' or 'tax_invoice'"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Generate and download a PDF for a sales invoice"""
    from fastapi.responses import Response
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    if mode not in ('estimate', 'tax_invoice', 'proforma_invoice'):
        raise HTTPException(status_code=400, detail="mode must be 'estimate', 'tax_invoice', or 'proforma_invoice'")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        from app.models.staff_accounts import SalesInvoice as SalesInvoiceModel
        pdf_bytes = SalesInvoiceService.generate_pdf(db, invoice_id, company_id, mode)
        invoice = db.query(SalesInvoiceModel).filter_by(id=invoice_id).first()
        inv_num = invoice.invoice_number if invoice else f"INV-{invoice_id}"
        # Determine rendered mode (service auto-overrides DRAFT → proforma_invoice)
        _rendered = 'proforma-invoice' if (mode == 'tax_invoice' and invoice and invoice.status == 'DRAFT') else \
                    ('estimate' if mode == 'estimate' else 'tax-invoice')
        filename = f"{inv_num}-{_rendered}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== INVOICE REPORTS ENDPOINTS ====================
# DC Protocol: All reports enforce mandatory company_id filtering

@router.get("/reports/sales-invoices")
async def get_sales_invoice_report(
    company_id: int = Query(..., description="Company ID - MANDATORY for data segregation"),
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status: DRAFT, CONFIRMED, CANCELLED"),
    customer_type: Optional[str] = Query(None, description="Filter by customer type"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Sales Invoice Report with HSN-wise tax breakdown
    DC Protocol: Mandatory company_id with strict date validation
    """
    from datetime import datetime
    from sqlalchemy import and_, func
    from app.models.staff_accounts import SalesInvoice, SalesInvoiceLineItem
    
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    from_date_parsed = None
    to_date_parsed = None
    
    if date_from:
        try:
            from_date_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_from format: {date_from}. Use YYYY-MM-DD")
    
    if date_to:
        try:
            to_date_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_to format: {date_to}. Use YYYY-MM-DD")
    
    try:
        query = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id)
        
        if from_date_parsed:
            query = query.filter(SalesInvoice.invoice_date >= from_date_parsed)
        if to_date_parsed:
            query = query.filter(SalesInvoice.invoice_date <= to_date_parsed)
        
        if status:
            query = query.filter(SalesInvoice.status == status.upper())
        
        if customer_type:
            query = query.filter(SalesInvoice.customer_type == customer_type.upper())
        
        total = query.count()
        invoices = query.order_by(SalesInvoice.invoice_date.desc()).offset((page - 1) * page_size).limit(page_size).all()
        
        invoice_list = []
        total_taxable = 0
        total_cgst = 0
        total_sgst = 0
        total_igst = 0
        grand_total = 0
        
        for inv in invoices:
            inv_data = inv.to_dict() if hasattr(inv, 'to_dict') else {
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'invoice_date': inv.invoice_date.isoformat() if inv.invoice_date else None,
                'customer_name': inv.customer_name,
                'customer_type': inv.customer_type,
                'status': inv.status,
                'subtotal': float(inv.subtotal or 0),
                'discount_amount': float(inv.discount_amount or 0),
                'taxable_amount': float(inv.taxable_amount or 0),
                'cgst_amount': float(inv.cgst_amount or 0),
                'sgst_amount': float(inv.sgst_amount or 0),
                'igst_amount': float(inv.igst_amount or 0),
                'total_tax': float(inv.total_tax or 0),
                'grand_total': float(inv.grand_total or 0),
                'payment_status': inv.payment_status
            }
            
            total_taxable += float(inv.taxable_amount or 0)
            total_cgst += float(inv.cgst_amount or 0)
            total_sgst += float(inv.sgst_amount or 0)
            total_igst += float(inv.igst_amount or 0)
            grand_total += float(inv.grand_total or 0)
            
            invoice_list.append(inv_data)
        
        return JSONResponse(content={
            "success": True,
            "invoices": invoice_list,
            "summary": {
                "total_invoices": total,
                "total_taxable": round(total_taxable, 2),
                "total_cgst": round(total_cgst, 2),
                "total_sgst": round(total_sgst, 2),
                "total_igst": round(total_igst, 2),
                "total_tax": round(total_cgst + total_sgst + total_igst, 2),
                "grand_total": round(grand_total, 2)
            },
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/reports/purchase-invoices")
async def get_purchase_invoice_report(
    company_id: int = Query(..., description="Company ID - MANDATORY for data segregation"),
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status: PENDING, CONFIRMED"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Purchase Invoice Report with HSN-wise tax breakdown
    DC Protocol: Mandatory company_id with strict date validation
    """
    from datetime import datetime
    
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    from_date_parsed = None
    to_date_parsed = None
    
    if date_from:
        try:
            from_date_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_from format: {date_from}. Use YYYY-MM-DD")
    
    if date_to:
        try:
            to_date_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_to format: {date_to}. Use YYYY-MM-DD")
    
    try:
        query = db.query(PurchaseInvoiceUpload).filter(PurchaseInvoiceUpload.company_id == company_id)
        
        if from_date_parsed:
            query = query.filter(PurchaseInvoiceUpload.vendor_invoice_date >= from_date_parsed)
        if to_date_parsed:
            query = query.filter(PurchaseInvoiceUpload.vendor_invoice_date <= to_date_parsed)
        
        if status:
            if status.upper() == 'CONFIRMED':
                query = query.filter(PurchaseInvoiceUpload.status == 'CONFIRMED')
            elif status.upper() == 'PENDING':
                query = query.filter(PurchaseInvoiceUpload.status.in_(['UPLOADED', 'EXTRACTED', 'REVIEWED']))
        
        if vendor_id:
            query = query.filter(PurchaseInvoiceUpload.vendor_id == vendor_id)
        
        total = query.count()
        invoices = query.order_by(PurchaseInvoiceUpload.vendor_invoice_date.desc()).offset((page - 1) * page_size).limit(page_size).all()
        
        invoice_list = []
        total_taxable = 0
        total_cgst = 0
        total_sgst = 0
        total_igst = 0
        grand_total = 0
        
        for inv in invoices:
            vendor = db.query(VendorMaster).filter(VendorMaster.id == inv.vendor_id).first() if inv.vendor_id else None
            
            taxable = float(inv.taxable_amount or inv.subtotal or 0)
            cgst = float(inv.cgst_amount or 0)
            sgst = float(inv.sgst_amount or 0)
            igst = float(inv.igst_amount or 0)
            total_amt = float(inv.grand_total or 0)
            
            extracted = inv.extracted_data or {}
            inv_data = {
                'id': inv.id,
                'invoice_number': inv.vendor_invoice_no,
                'invoice_date': inv.vendor_invoice_date.isoformat() if inv.vendor_invoice_date else None,
                'vendor_id': inv.vendor_id,
                'vendor_name': vendor.vendor_name if vendor else extracted.get('vendor_name', 'Unknown'),
                'vendor_gstin': vendor.gst_number if vendor else extracted.get('vendor_gstin', ''),
                'status': inv.status or 'PENDING',
                'taxable_amount': taxable,
                'cgst_amount': cgst,
                'sgst_amount': sgst,
                'igst_amount': igst,
                'total_tax': cgst + sgst + igst,
                'grand_total': total_amt
            }
            
            total_taxable += taxable
            total_cgst += cgst
            total_sgst += sgst
            total_igst += igst
            grand_total += total_amt
            
            invoice_list.append(inv_data)
        
        return JSONResponse(content={
            "success": True,
            "invoices": invoice_list,
            "summary": {
                "total_invoices": total,
                "total_taxable": round(total_taxable, 2),
                "total_cgst": round(total_cgst, 2),
                "total_sgst": round(total_sgst, 2),
                "total_igst": round(total_igst, 2),
                "total_tax": round(total_cgst + total_sgst + total_igst, 2),
                "grand_total": round(grand_total, 2)
            },
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        import logging
        logging.error(f"Purchase Invoice Report Error: {str(e)}")
        logging.error(traceback.format_exc())
        return handle_accounts_error(e)


@router.get("/reports/hsn-sales")
async def get_hsn_sales_report(
    company_id: int = Query(..., description="Company ID - MANDATORY for data segregation"),
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    HSN Code-wise Sales Report with tax breakdown using proper SQL joins
    DC Protocol: Mandatory company_id for data segregation
    """
    from datetime import datetime
    from sqlalchemy import func, and_, case, literal
    from app.models.staff_accounts import SalesInvoice, SalesInvoiceLineItem
    
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    from_date = None
    to_date = None
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_from format: {date_from}. Use YYYY-MM-DD")
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_to format: {date_to}. Use YYYY-MM-DD")
    
    try:
        base_query = db.query(
            func.coalesce(SalesInvoiceLineItem.hsn_code, literal('N/A')).label('hsn_code'),
            func.sum(func.coalesce(SalesInvoiceLineItem.quantity, 1)).label('quantity'),
            func.sum(func.coalesce(SalesInvoiceLineItem.taxable_amount, SalesInvoiceLineItem.gross_amount, 0)).label('taxable_value'),
            func.max(SalesInvoiceLineItem.gst_rate).label('gst_rate'),
            func.max(SalesInvoiceLineItem.item_description).label('description'),
            func.sum(func.coalesce(SalesInvoiceLineItem.cgst_amount, 0)).label('cgst_amount'),
            func.sum(func.coalesce(SalesInvoiceLineItem.sgst_amount, 0)).label('sgst_amount'),
            func.sum(func.coalesce(SalesInvoiceLineItem.igst_amount, 0)).label('igst_amount'),
            func.sum(func.coalesce(SalesInvoiceLineItem.total_tax, 0)).label('total_tax')
        ).join(
            SalesInvoice, SalesInvoiceLineItem.invoice_id == SalesInvoice.id
        ).filter(
            SalesInvoice.company_id == company_id,
            SalesInvoice.status == 'CONFIRMED'
        )
        
        if from_date:
            base_query = base_query.filter(SalesInvoice.invoice_date >= from_date)
        if to_date:
            base_query = base_query.filter(SalesInvoice.invoice_date <= to_date)
        
        results = base_query.group_by(
            func.coalesce(SalesInvoiceLineItem.hsn_code, literal('N/A'))
        ).order_by(
            func.coalesce(SalesInvoiceLineItem.hsn_code, literal('N/A'))
        ).all()
        
        hsn_list = []
        for row in results:
            gst_rate = float(row.gst_rate or 18)
            taxable = float(row.taxable_value or 0)
            cgst = float(row.cgst_amount or 0)
            sgst = float(row.sgst_amount or 0)
            igst = float(row.igst_amount or 0)
            total_tax = float(row.total_tax or 0)
            
            hsn_list.append({
                'hsn_code': row.hsn_code,
                'description': row.description or '',
                'gst_rate': gst_rate,
                'quantity': float(row.quantity or 0),
                'taxable_value': round(taxable, 2),
                'cgst_rate': gst_rate / 2,
                'cgst_amount': round(cgst, 2),
                'sgst_rate': gst_rate / 2,
                'sgst_amount': round(sgst, 2),
                'igst_rate': gst_rate,
                'igst_amount': round(igst, 2),
                'total_tax': round(total_tax, 2),
                'total_value': round(taxable + total_tax, 2)
            })
        
        summary = {
            'total_taxable': round(sum(h['taxable_value'] for h in hsn_list), 2),
            'total_cgst': round(sum(h['cgst_amount'] for h in hsn_list), 2),
            'total_sgst': round(sum(h['sgst_amount'] for h in hsn_list), 2),
            'total_igst': round(sum(h['igst_amount'] for h in hsn_list), 2),
            'total_tax': round(sum(h['total_tax'] for h in hsn_list), 2),
            'grand_total': round(sum(h['total_value'] for h in hsn_list), 2)
        }
        
        return JSONResponse(content={
            "success": True,
            "hsn_data": hsn_list,
            "summary": summary,
            "filters": {
                "company_id": company_id,
                "date_from": date_from,
                "date_to": date_to
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        import logging
        logging.error(f"HSN Sales Report Error: {str(e)}")
        logging.error(traceback.format_exc())
        return handle_accounts_error(e)


@router.get("/reports/hsn-purchases")
async def get_hsn_purchases_report(
    company_id: int = Query(..., description="Company ID - MANDATORY for data segregation"),
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    HSN Code-wise Purchase Report with tax breakdown
    DC Protocol: Mandatory company_id for data segregation with strict date validation
    """
    from datetime import datetime
    
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    
    from_date_parsed = None
    to_date_parsed = None
    
    if date_from:
        try:
            from_date_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_from format: {date_from}. Use YYYY-MM-DD")
    
    if date_to:
        try:
            to_date_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date_to format: {date_to}. Use YYYY-MM-DD")
    
    try:
        query = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.company_id == company_id,
            PurchaseInvoiceUpload.status == 'CONFIRMED'
        )
        
        if from_date_parsed:
            query = query.filter(PurchaseInvoiceUpload.vendor_invoice_date >= from_date_parsed)
        if to_date_parsed:
            query = query.filter(PurchaseInvoiceUpload.vendor_invoice_date <= to_date_parsed)
        
        invoices = query.all()
        
        hsn_data = {}
        
        for inv in invoices:
            extracted = inv.extracted_data or {}
            line_items = extracted.get('line_items', [])
            is_igst = extracted.get('is_igst', False)
            
            for item in line_items:
                hsn_code = item.get('hsn_code') or item.get('hsn') or 'N/A'
                gst_rate = float(item.get('gst_rate') or item.get('gst') or 18)
                taxable = float(item.get('amount') or item.get('taxable_value') or 0)
                tax_amount = float(item.get('tax_amount') or (taxable * gst_rate / 100))
                
                if hsn_code not in hsn_data:
                    hsn_data[hsn_code] = {
                        'hsn_code': hsn_code,
                        'description': item.get('description') or item.get('item_name') or '',
                        'gst_rate': gst_rate,
                        'quantity': 0,
                        'taxable_value': 0,
                        'cgst_rate': gst_rate / 2,
                        'cgst_amount': 0,
                        'sgst_rate': gst_rate / 2,
                        'sgst_amount': 0,
                        'igst_rate': gst_rate,
                        'igst_amount': 0,
                        'total_tax': 0,
                        'total_value': 0
                    }
                
                hsn_data[hsn_code]['quantity'] += float(item.get('quantity') or 1)
                hsn_data[hsn_code]['taxable_value'] += taxable
                
                if is_igst:
                    hsn_data[hsn_code]['igst_amount'] += tax_amount
                else:
                    hsn_data[hsn_code]['cgst_amount'] += tax_amount / 2
                    hsn_data[hsn_code]['sgst_amount'] += tax_amount / 2
                
                hsn_data[hsn_code]['total_tax'] += tax_amount
                hsn_data[hsn_code]['total_value'] += taxable + tax_amount
        
        hsn_list = sorted(hsn_data.values(), key=lambda x: x['hsn_code'])
        
        for h in hsn_list:
            h['taxable_value'] = round(h['taxable_value'], 2)
            h['cgst_amount'] = round(h['cgst_amount'], 2)
            h['sgst_amount'] = round(h['sgst_amount'], 2)
            h['igst_amount'] = round(h['igst_amount'], 2)
            h['total_tax'] = round(h['total_tax'], 2)
            h['total_value'] = round(h['total_value'], 2)
        
        summary = {
            'total_taxable': round(sum(h['taxable_value'] for h in hsn_list), 2),
            'total_cgst': round(sum(h['cgst_amount'] for h in hsn_list), 2),
            'total_sgst': round(sum(h['sgst_amount'] for h in hsn_list), 2),
            'total_igst': round(sum(h['igst_amount'] for h in hsn_list), 2),
            'total_tax': round(sum(h['total_tax'] for h in hsn_list), 2),
            'grand_total': round(sum(h['total_value'] for h in hsn_list), 2)
        }
        
        return JSONResponse(content={
            "success": True,
            "hsn_data": hsn_list,
            "summary": summary,
            "filters": {
                "company_id": company_id,
                "date_from": date_from,
                "date_to": date_to
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ===================================================================
# PROCUREMENT REQUEST MULTI-QUOTE WORKFLOW ENDPOINTS (DC_PROCUREMENT_002)
# ===================================================================

@router.post("/procurement/requests")
async def create_procurement_request(
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new procurement request with multiple items
    DC Protocol: Blind bidding - no prices included in request
    """
    try:
        data = await request.json()
        
        company_id = data.get('company_id')
        items = data.get('items', [])
        notes = data.get('notes')
        min_quotes = data.get('min_quotes_required', 2)
        
        if not company_id:
            raise HTTPException(status_code=400, detail="company_id is required")
        
        if not items or len(items) == 0:
            raise HTTPException(status_code=400, detail="At least one item is required")
        
        from app.services.staff_accounts_service import ProcurementRequestService
        
        result = ProcurementRequestService.create_request(
            db, current_user, company_id, items, notes, min_quotes
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Procurement request created successfully",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/requests")
async def get_procurement_requests(
    company_id: Optional[int] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get procurement requests with filters (DC Protocol: company_id filtering)"""
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        
        skip = (page - 1) * limit
        requests, total = ProcurementRequestService.get_requests(
            db, current_user, company_id, status, search, skip, limit
        )
        
        return JSONResponse(content={
            "success": True,
            "data": requests,
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total,
                "pages": (total + limit - 1) // limit
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/requests/{request_id}")
async def get_procurement_request_details(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get full procurement request details with items and quotes"""
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        
        result = ProcurementRequestService.get_request_details(db, current_user, request_id)
        
        return JSONResponse(content={
            "success": True,
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/procurement/requests/{request_id}/status")
async def update_procurement_request_status(
    request_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update procurement request status"""
    try:
        data = await request.json()
        new_status = data.get('status')
        notes = data.get('notes', None)

        if not new_status:
            raise HTTPException(status_code=400, detail="status is required")

        from app.services.staff_accounts_service import ProcurementRequestService

        result = ProcurementRequestService.update_status(
            db, current_user, request_id, new_status, notes=notes
        )

        return JSONResponse(content={
            "success": True,
            "message": f"Status updated to {new_status}",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/requests/{request_id}/whatsapp")
async def send_procurement_request_whatsapp(
    request_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-PROC-WA-001: Send procurement request details to approved vendor via WhatsApp.
    Falls back to returning the phone + message so frontend can open web.whatsapp.com.
    """
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        from app.models.staff_accounts import ProcurementRequest, VendorMaster

        body = await request.json()
        custom_message = body.get('message', '')

        proc_req = db.query(ProcurementRequest).filter(
            ProcurementRequest.id == request_id
        ).first()
        if not proc_req:
            raise HTTPException(status_code=404, detail="Procurement request not found")

        if not proc_req.approved_vendor_id:
            raise HTTPException(status_code=400, detail="No approved vendor on this request")

        vendor = db.query(VendorMaster).filter(
            VendorMaster.id == proc_req.approved_vendor_id
        ).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")

        phone = vendor.phone or vendor.contact_person_1_phone
        if not phone:
            return JSONResponse(content={
                "success": False,
                "message": "Vendor has no phone number — use WhatsApp Web manually",
                "phone": None,
            })

        # Try WA Business API; fall back gracefully
        try:
            from app.services.whatsapp_auto_service import send_auto_whatsapp
            result = send_auto_whatsapp(
                db=db,
                phone=phone,
                template_slug='procurement_request_vendor',
                params={
                    'vendor_name': vendor.vendor_name,
                    'request_number': proc_req.request_number,
                    'company_name': proc_req.company.company_name if proc_req.company else 'MyntReal LLP',
                    'item_count': str(len(proc_req.request_items) if proc_req.request_items else 0),
                },
                created_by_id=current_user.id,
            )
            return JSONResponse(content={
                "success": True,
                "message": f"WhatsApp sent to {phone}",
                "phone": phone,
                "result": str(result),
            })
        except Exception as wa_err:
            # Template may not exist — return phone so frontend opens web.whatsapp.com
            return JSONResponse(content={
                "success": False,
                "message": f"WA template not found ({wa_err}) — opening WhatsApp Web",
                "phone": phone,
            })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/requests/{request_id}/quotes")
async def add_procurement_quote(
    request_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add vendor quote (Proforma Invoice) to procurement request
    Minimum 2 quotes required before approval
    """
    try:
        data = await request.json()
        
        vendor_id = data.get('vendor_id')
        quote_data = data.get('quote_data', {})
        quote_items = data.get('quote_items', [])
        
        if not vendor_id:
            raise HTTPException(status_code=400, detail="vendor_id is required")
        
        if not quote_items or len(quote_items) == 0:
            raise HTTPException(status_code=400, detail="quote_items are required")
        
        from app.services.staff_accounts_service import ProcurementRequestService
        
        result = ProcurementRequestService.add_quote(
            db, current_user, request_id, vendor_id, quote_data, quote_items
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Quote added successfully",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/requests/{request_id}/quotes")
async def get_procurement_quotes(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get all quotes for a procurement request"""
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        
        quotes = ProcurementRequestService.get_quotes(db, current_user, request_id)
        
        return JSONResponse(content={
            "success": True,
            "data": quotes
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/quotes/{quote_id}/approve")
async def approve_procurement_quote(
    quote_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Approve selected quote - validates minimum quotes received
    Only leadership roles can approve
    """
    try:
        data = await request.json()
        approval_remarks = data.get('approval_remarks')
        
        from app.services.staff_accounts_service import ProcurementRequestService
        
        result = ProcurementRequestService.approve_quote(
            db, current_user, quote_id, approval_remarks
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Quote approved successfully",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/requests/{request_id}/generate-po")
async def generate_purchase_order(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Generate Purchase Order from approved quote
    Creates VendorTransactionUpload (Purchase Invoice) entry
    """
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        
        result = ProcurementRequestService.generate_po_from_quote(
            db, current_user, request_id
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Purchase Order generated successfully",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/requests/{request_id}/download")
async def get_procurement_download_data(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get procurement request data for PDF download (BLIND BIDDING - NO PRICES)
    Returns item details without any price information
    """
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        
        result = ProcurementRequestService.get_download_data(db, current_user, request_id)
        
        return JSONResponse(content={
            "success": True,
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/requests/{request_id}/available-vendors")
async def get_available_vendors(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get vendors who haven't submitted quotes yet for this request"""
    try:
        from app.services.staff_accounts_service import ProcurementRequestService
        
        vendors = ProcurementRequestService.get_available_vendors_for_request(
            db, current_user, request_id
        )
        
        return JSONResponse(content={
            "success": True,
            "data": vendors
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ==================== STOCK VALIDATION ENDPOINTS ====================
# DC_STOCK_VALIDATION: Periodic stock validation with VGK Supreme approval

from app.schemas.staff_accounts import (
    StockValidationSessionCreate, StockValidationSessionUpdate,
    StockValidationEntryCreate, StockValidationEntryUpdate,
    StockValidationEntryBulkUpdate, StockValidationApprovalRequest,
    StockValidationSessionResponse, StockValidationSessionListResponse,
    StockValidationEntryResponse, NewStockItemFromValidation,
    AddExistingItemToValidation
)


@router.post("/stock-validation/sessions")
async def create_stock_validation_session(
    data: StockValidationSessionCreate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create a new stock validation session
    DC Protocol: VGK/EA/Accounts can create sessions
    Auto-populates items from current stock if auto_populate_items=True
    """
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        session = StockValidationService.create_session(
            db, current_user, data.model_dump()
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Stock validation session created successfully",
            "data": session
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-validation/sessions")
async def list_stock_validation_sessions(
    company_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    validation_type: Optional[str] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List stock validation sessions with filters"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        result = StockValidationService.list_sessions(
            db, current_user,
            company_id=company_id,
            status=status,
            validation_type=validation_type,
            from_date=from_date,
            to_date=to_date,
            page=page,
            page_size=page_size
        )
        
        return JSONResponse(content={
            "success": True,
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-validation/sessions/{session_id}")
async def get_stock_validation_session(
    session_id: int,
    include_entries: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get stock validation session details with entries"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        session = StockValidationService.get_session(
            db, current_user, session_id, include_entries
        )
        
        return JSONResponse(content={
            "success": True,
            "data": session
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/stock-validation/sessions/{session_id}")
async def update_stock_validation_session(
    session_id: int,
    data: StockValidationSessionUpdate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update stock validation session"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        session = StockValidationService.update_session(
            db, current_user, session_id, data.model_dump(exclude_unset=True)
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Session updated successfully",
            "data": session
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/entries")
async def add_validation_entry(
    session_id: int,
    data: StockValidationEntryCreate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Add a new item to validation session"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        entry = StockValidationService.add_entry(
            db, current_user, session_id, data.model_dump()
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Entry added successfully",
            "data": entry
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/stock-validation/entries/{entry_id}")
async def update_validation_entry(
    entry_id: int,
    data: StockValidationEntryUpdate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update physical count for a validation entry"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        entry = StockValidationService.update_entry(
            db, current_user, entry_id, data.model_dump(exclude_unset=True)
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Entry updated successfully",
            "data": entry
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/bulk-update")
async def bulk_update_entries(
    session_id: int,
    data: StockValidationEntryBulkUpdate = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Bulk update entries from Excel import"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        result = StockValidationService.bulk_update_entries(
            db, current_user, session_id, data.entries
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Updated {result['updated']} entries",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/submit")
async def submit_for_approval(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Submit session for VGK Supreme approval"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        session = StockValidationService.submit_for_approval(
            db, current_user, session_id
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Session submitted for approval",
            "data": session
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/approve")
async def approve_stock_validation(
    session_id: int,
    data: StockValidationApprovalRequest = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Approve or reject stock validation session
    DC Protocol: VGK Supreme only can approve
    On approval, creates ADJUSTMENT entries in stock ledger
    """
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        if data.action == 'approve':
            session = StockValidationService.approve_session(
                db, current_user, session_id, data.approval_notes
            )
            message = "Session approved and stock adjustments processed"
        else:
            session = StockValidationService.reject_session(
                db, current_user, session_id, data.rejection_reason
            )
            message = "Session rejected"
        
        return JSONResponse(content={
            "success": True,
            "message": message,
            "data": session
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-validation/sessions/{session_id}/download")
async def download_validation_template(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Download Excel template with current stock for validation"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        data = StockValidationService.get_download_data(
            db, current_user, session_id
        )
        
        return JSONResponse(content={
            "success": True,
            "data": data
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/upload")
async def upload_validation_data(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Upload Excel with physical counts"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        form = await request.form()
        file = form.get('file')
        
        if not file:
            raise HTTPException(status_code=400, detail="No file uploaded")
        
        result = StockValidationService.process_upload(
            db, current_user, session_id, file
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Processed {result['processed']} entries",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/new-item")
async def create_item_from_validation(
    session_id: int,
    data: NewStockItemFromValidation = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Create new stock item with opening quantity from validation page
    DC Protocol: Creates item + OPENING ledger entry + validation entry
    """
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        result = StockValidationService.create_item_from_validation(
            db, current_user, session_id, data.model_dump()
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Stock item created and added to validation",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock-validation/sessions/{session_id}/add-items")
async def add_existing_items_to_validation(
    session_id: int,
    data: AddExistingItemToValidation = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Add existing stock items to a validation session
    DC Protocol: Fetches current stock position from ledger
    """
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        result = StockValidationService.add_existing_items_to_validation(
            db, current_user, session_id, data.item_ids
        )
        
        return JSONResponse(content={
            "success": True,
            "message": f"Added {result['added']} items to validation",
            "data": result
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock-validation/sessions/{session_id}/audit-log")
async def get_validation_audit_log(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Get audit log for validation session"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        logs = StockValidationService.get_audit_log(
            db, current_user, session_id
        )
        
        return JSONResponse(content={
            "success": True,
            "data": logs
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/stock-validation/sessions/{session_id}")
async def cancel_validation_session(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Cancel a draft validation session"""
    try:
        from app.services.staff_accounts_service import StockValidationService
        
        StockValidationService.cancel_session(
            db, current_user, session_id
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Session cancelled successfully"
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ===========================================================================================
# PURCHASE INTAKE & LIFECYCLE TRACKING (DC_INTAKE_001 - Jan 2026)
# ===========================================================================================

@router.get("/purchase-intake/batches")
async def list_intake_batches(
    company_id: Optional[int] = None,
    status: Optional[str] = None,
    vendor_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List purchase intake batches with filters
    DC Protocol: Company-scoped, role-based access
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        filters = {
            'company_id': company_id,
            'status': status,
            'vendor_id': vendor_id,
            'date_from': date_from,
            'date_to': date_to
        }
        
        batches, total = PurchaseIntakeService.list_batches(
            db, current_user, filters, page, page_size
        )
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "items": batches,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-intake/batches/{batch_id}")
async def get_intake_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get single intake batch with items
    DC Protocol: Full batch data with lifecycle events
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        batch = PurchaseIntakeService.get_batch(db, current_user, batch_id)
        
        return JSONResponse(content={
            "success": True,
            "data": batch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-intake/batches/{batch_id}/receive")
async def record_batch_receipt(
    batch_id: int,
    data: PurchaseIntakeBatchReceive = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record physical receipt of items
    DC Protocol: Updates received_qty, triggers lifecycle event
    Pydantic schema validation enforces WVV rules
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        items = [item.model_dump() for item in data.items]
        notes = data.receipt_notes or ''
        
        batch = PurchaseIntakeService.record_receipt(
            db, current_user, batch_id, items, notes
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Receipt recorded successfully",
            "data": batch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-intake/batches/{batch_id}/qc")
async def submit_qc_results(
    batch_id: int,
    data: PurchaseIntakeBatchQCSubmit = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Submit QC results for intake batch items
    DC Protocol: Standard QC checklist, tracks passed/rejected/pending
    Pydantic schema validation enforces WVV rules
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        items = data.items
        
        batch = PurchaseIntakeService.submit_qc_results(
            db, current_user, batch_id, items
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "QC results submitted",
            "data": batch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-intake/batches/{batch_id}/approve")
async def approve_intake_batch(
    batch_id: int,
    data: PurchaseIntakeBatchApprove = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Approve intake batch - adds QC-passed items to stock
    DC Protocol: VGK/EA/Accounts only, triggers stock_ledger update
    WVV Protocol: Stock ledger ONLY updated here on approval
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        notes = data.approval_notes or ''
        
        batch = PurchaseIntakeService.approve_batch(
            db, current_user, batch_id, notes
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Batch approved and stock updated",
            "data": batch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-intake/batches/{batch_id}/reject")
async def reject_intake_batch(
    batch_id: int,
    data: PurchaseIntakeBatchReject = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Reject entire intake batch
    DC Protocol: VGK/EA/Accounts only
    Pydantic schema enforces minimum rejection reason length
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        reason = data.rejection_reason
        
        batch = PurchaseIntakeService.reject_batch(
            db, current_user, batch_id, reason
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Batch rejected",
            "data": batch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-intake/items/{item_id}/lifecycle")
async def get_item_lifecycle(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get complete lifecycle history for an intake item
    DC Protocol: Immutable event chain with checksums
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        lifecycle = PurchaseIntakeService.get_item_lifecycle(db, current_user, item_id)
        
        return JSONResponse(content={
            "success": True,
            "data": lifecycle
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/purchase-intake/dashboard")
async def get_intake_dashboard(
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get intake dashboard with summary stats
    DC Protocol: Pending, In Progress, QC Required counts
    """
    try:
        from app.services.staff_accounts_service import PurchaseIntakeService
        
        stats = PurchaseIntakeService.get_dashboard_stats(db, current_user, company_id)
        
        return JSONResponse(content={
            "success": True,
            "data": stats
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ===========================================================================================
# SERVICE CENTER ITEM TRACKING (DC_INTAKE_001 - Jan 2026)
# ===========================================================================================

@router.get("/service-center-tracking/receipts")
async def list_service_center_receipts(
    company_id: Optional[int] = None,
    partner_id: Optional[int] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List service center item receipts
    DC Protocol: Items received from customers/vehicles
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        filters = {
            'company_id': company_id,
            'partner_id': partner_id,
            'status': status
        }
        
        receipts, total = ServiceCenterTrackingService.list_receipts(
            db, current_user, filters, page, page_size
        )
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "items": receipts,
                "total": total,
                "page": page,
                "page_size": page_size
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/service-center-tracking/receipts")
async def create_service_receipt(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record item receipt at service center
    DC Protocol: Links to ticket, tracks customer source
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        receipt = ServiceCenterTrackingService.create_receipt(
            db, current_user, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Receipt created",
            "data": receipt
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/service-center-tracking/receipts/{receipt_id}/diagnose")
async def submit_diagnosis(
    receipt_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Submit diagnosis for received item
    DC Protocol: Standard checklist, determines next action
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        receipt = ServiceCenterTrackingService.submit_diagnosis(
            db, current_user, receipt_id, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Diagnosis submitted",
            "data": receipt
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/service-center-tracking/receipts/{receipt_id}/escalate")
async def escalate_to_vendor(
    receipt_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Escalate item to head office/vendor
    DC Protocol: Creates dispatch record, notifies vendor
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        dispatch = ServiceCenterTrackingService.escalate_to_vendor(
            db, current_user, receipt_id, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Item escalated to vendor",
            "data": dispatch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/service-center-tracking/dispatches")
async def list_service_center_dispatches(
    company_id: Optional[int] = None,
    destination_type: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List service center dispatches (to vendor/head office)
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        filters = {
            'company_id': company_id,
            'destination_type': destination_type,
            'status': status
        }
        
        dispatches, total = ServiceCenterTrackingService.list_dispatches(
            db, current_user, filters, page, page_size
        )
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "items": dispatches,
                "total": total,
                "page": page,
                "page_size": page_size
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/service-center-tracking/dispatches/{dispatch_id}/receive")
async def receive_at_destination(
    dispatch_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record receipt at destination (head office/vendor)
    DC Protocol: Closes service center loop
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        dispatch = ServiceCenterTrackingService.record_destination_receipt(
            db, current_user, dispatch_id, data
        )
        
        return JSONResponse(content={
            "success": True,
            "message": "Receipt confirmed at destination",
            "data": dispatch
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ===========================================================================================
# SERVICE TICKET LINKED ITEMS & QUEUE REPORTS (DC_INTAKE_001 - Jan 2026)
# ===========================================================================================

@router.get("/service-center-tracking/by-ticket/{ticket_id}")
async def get_items_by_ticket(
    ticket_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get all service center items linked to a specific service ticket
    DC Protocol: Complete item-wise status for service queue integration
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        items = ServiceCenterTrackingService.get_items_by_ticket(db, current_user, ticket_id)
        
        return JSONResponse(content={
            "success": True,
            "data": items
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/service-center-tracking/receipts/{receipt_id}/detail")
async def get_receipt_detail(
    receipt_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Get detailed receipt view with full item status, history, and linked ticket info
    DC Protocol: Complete item lifecycle visibility
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        detail = ServiceCenterTrackingService.get_receipt_detail(db, current_user, receipt_id)
        
        return JSONResponse(content={
            "success": True,
            "data": detail
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/service-center-tracking/queue-report")
async def get_service_queue_report(
    company_id: Optional[int] = None,
    service_center_id: Optional[int] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Service queue report with item-wise status summary
    DC Protocol: Aggregated view for service management dashboard
    """
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        
        filters = {
            'company_id': company_id,
            'service_center_id': service_center_id,
            'status': status,
            'date_from': date_from,
            'date_to': date_to
        }
        
        report = ServiceCenterTrackingService.get_queue_report(
            db, current_user, filters, page, page_size
        )
        
        return JSONResponse(content={
            "success": True,
            "data": report
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/service-center-tracking/ticket-search")
async def search_tickets_for_dropdown(
    q: str = Query("", description="Search by ticket ID or customer name"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Search service tickets for dropdown (active tickets only)"""
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        results = ServiceCenterTrackingService.search_tickets(db, q)
        return JSONResponse(content={"success": True, "data": results})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/service-center-tracking/vendor-repairs")
async def list_vendor_repairs(
    company_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(30, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List items sent to vendor for repair (ESCALATED receipts)"""
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        filters = {
            'company_id': company_id,
            'status': status
        }
        items, total = ServiceCenterTrackingService.list_vendor_repairs(db, current_user, filters, page, page_size)
        return JSONResponse(content={
            "success": True,
            "data": items,
            "total": total,
            "page": page,
            "page_size": page_size
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/service-center-tracking/receipts/{receipt_id}/issue-replacement")
async def issue_replacement(
    receipt_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Issue replacement for a service receipt and deduct from stock"""
    try:
        from app.services.staff_accounts_service import ServiceCenterTrackingService
        data = await request.json()
        result = ServiceCenterTrackingService.issue_replacement(db, current_user, receipt_id, data)
        return JSONResponse(content={"success": True, "data": result})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ── DC_SERVICE_003: Given Out Tracking ───────────────────────────────────────

@router.get("/service-center-tracking/given-out")
def list_given_out(
    status: Optional[str] = Query(None),
    recipient_type: Optional[str] = Query(None),
    purpose: Optional[str] = Query(None),
    service_center_id: Optional[int] = Query(None),
    ticket_id: Optional[int] = Query(None),
    overdue_only: Optional[bool] = Query(False),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query('given_at'),
    sort_dir: Optional[str] = Query('desc'),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    company_id: Optional[int] = Query(None),
    current_user=Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    try:
        filters = {
            'company_id': company_id or (current_user.company_id if hasattr(current_user, 'company_id') else None),
            'status': status, 'recipient_type': recipient_type, 'purpose': purpose,
            'service_center_id': service_center_id, 'ticket_id': ticket_id,
            'overdue_only': overdue_only, 'date_from': date_from, 'date_to': date_to,
            'search': search, 'sort_by': sort_by, 'sort_dir': sort_dir
        }
        return ServiceCenterGivenOutService.list_given_out(db, current_user, filters, page, page_size)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/service-center-tracking/given-out")
def create_given_out(
    data: dict,
    current_user=Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    import traceback, logging as _log
    try:
        return ServiceCenterGivenOutService.create_given_out(db, current_user, data)
    except HTTPException:
        raise
    except Exception as e:
        _log.getLogger(__name__).error(f"[GIVEN-OUT-500] {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return handle_accounts_error(e)


@router.patch("/service-center-tracking/given-out/{record_id}/status")
def update_given_out_status(
    record_id: int,
    data: dict,
    current_user=Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    try:
        action = data.pop('action', None)
        if not action:
            raise HTTPException(status_code=400, detail="action is required")
        return ServiceCenterGivenOutService.update_status(db, current_user, record_id, action, data)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/service-center-tracking/given-out/invoice-search")
def search_invoices_for_given_out(
    q: str = Query(..., min_length=1),
    company_id: Optional[int] = Query(None),
    current_user=Depends(get_current_staff_user),
    db: Session = Depends(get_db)
):
    try:
        return ServiceCenterGivenOutService.search_invoices(db, current_user, q, company_id)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ── DC_CASHFLOW_001: Cashflow Register CRUD ───────────────────────────────────

@router.get("/cashflow-register")
def list_cashflow_register(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    entity:    Optional[str]  = Query(None, description="all|mnr|mynt|zynova|escrow"),
    page:      int = Query(1, ge=1),
    page_size: int = Query(60, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List cashflow register entries descending by date."""
    try:
        from sqlalchemy import desc as sa_desc
        q = db.query(CashflowRegister)
        if from_date:
            q = q.filter(CashflowRegister.entry_date >= from_date)
        if to_date:
            q = q.filter(CashflowRegister.entry_date <= to_date)
        total = q.count()
        rows = q.order_by(sa_desc(CashflowRegister.entry_date)).offset((page - 1) * page_size).limit(page_size).all()
        data = [r.to_dict() for r in rows]
        # If entity filter requested, zero-out unrelated columns in response (display-only)
        if entity and entity != 'all':
            for row in data:
                if entity == 'mnr':
                    for k in ['mynt_sales_in','mynt_spares_in','mynt_service_in','zynova_in','escrow_in',
                              'mynt_sales_out','mynt_spares_out','mynt_service_out','zynova_out','escrow_out']:
                        row[k] = None
                elif entity == 'mynt':
                    for k in ['mnr_in','zynova_in','escrow_in','mnr_out','zynova_out','escrow_out']:
                        row[k] = None
                elif entity == 'zynova':
                    for k in ['mnr_in','mynt_sales_in','mynt_spares_in','mynt_service_in','escrow_in',
                              'mnr_out','mynt_sales_out','mynt_spares_out','mynt_service_out','escrow_out']:
                        row[k] = None
                elif entity == 'escrow':
                    for k in ['mnr_in','mynt_sales_in','mynt_spares_in','mynt_service_in','zynova_in',
                              'mnr_out','mynt_sales_out','mynt_spares_out','mynt_service_out','zynova_out']:
                        row[k] = None
                # Recompute totals for filtered view (explicit field list to avoid double-counting total_in/total_out)
                _in_fields  = ['mnr_in','mynt_sales_in','mynt_spares_in','mynt_service_in','zynova_in','escrow_in']
                _out_fields = ['mnr_out','mynt_sales_out','mynt_spares_out','mynt_service_out','zynova_out','escrow_out']
                _in  = sum(row[f] for f in _in_fields  if row.get(f) is not None)
                _out = sum(row[f] for f in _out_fields if row.get(f) is not None)
                row['total_in']        = _in
                row['total_out']       = _out
                row['for_the_day']     = _in - _out
                row['closing_balance'] = row['opening_balance'] + (_in - _out)
        return JSONResponse(content={
            "success": True,
            "data": data,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size)
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/cashflow-register")
def create_cashflow_entry(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Create a single daily cashflow register entry."""
    try:
        from datetime import date as date_type
        entry_date_raw = payload.get('entry_date')
        if not entry_date_raw:
            raise HTTPException(status_code=400, detail="entry_date is required")
        if isinstance(entry_date_raw, str):
            from datetime import datetime
            entry_date_val = datetime.strptime(entry_date_raw, '%Y-%m-%d').date()
        else:
            entry_date_val = entry_date_raw

        existing = db.query(CashflowRegister).filter(CashflowRegister.entry_date == entry_date_val).first()
        if existing:
            raise HTTPException(status_code=409, detail=f"Entry for {entry_date_val} already exists. Use edit to update.")

        def _d(key):
            v = payload.get(key, 0)
            return Decimal(str(v)) if v else Decimal('0')

        entry = CashflowRegister(
            entry_date       = entry_date_val,
            opening_balance  = _d('opening_balance'),
            mnr_in           = _d('mnr_in'),
            mynt_sales_in    = _d('mynt_sales_in'),
            mynt_spares_in   = _d('mynt_spares_in'),
            mynt_service_in  = _d('mynt_service_in'),
            zynova_in        = _d('zynova_in'),
            escrow_in        = _d('escrow_in'),
            mnr_out          = _d('mnr_out'),
            mynt_sales_out   = _d('mynt_sales_out'),
            mynt_spares_out  = _d('mynt_spares_out'),
            mynt_service_out = _d('mynt_service_out'),
            zynova_out       = _d('zynova_out'),
            escrow_out       = _d('escrow_out'),
            notes            = payload.get('notes'),
            created_by_id    = current_user.id,
        )
        db.add(entry)
        db.commit()
        db.refresh(entry)
        return JSONResponse(status_code=201, content={"success": True, "data": entry.to_dict()})
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.put("/cashflow-register/{entry_id}")
def update_cashflow_entry(
    entry_id: int = Path(...),
    payload:  dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update an existing cashflow register entry."""
    try:
        entry = db.query(CashflowRegister).filter(CashflowRegister.id == entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Cashflow entry not found")

        def _d(key, current):
            if key in payload:
                v = payload[key]
                return Decimal(str(v)) if v is not None else Decimal('0')
            return current

        if 'entry_date' in payload:
            from datetime import datetime as _dt
            new_date = payload['entry_date']
            if isinstance(new_date, str):
                new_date = _dt.strptime(new_date, '%Y-%m-%d').date()
            conflict = db.query(CashflowRegister).filter(
                CashflowRegister.entry_date == new_date,
                CashflowRegister.id != entry_id
            ).first()
            if conflict:
                raise HTTPException(status_code=409, detail=f"Another entry for {new_date} already exists.")
            entry.entry_date = new_date

        entry.opening_balance  = _d('opening_balance', entry.opening_balance)
        entry.mnr_in           = _d('mnr_in', entry.mnr_in)
        entry.mynt_sales_in    = _d('mynt_sales_in', entry.mynt_sales_in)
        entry.mynt_spares_in   = _d('mynt_spares_in', entry.mynt_spares_in)
        entry.mynt_service_in  = _d('mynt_service_in', entry.mynt_service_in)
        entry.zynova_in        = _d('zynova_in', entry.zynova_in)
        entry.escrow_in        = _d('escrow_in', entry.escrow_in)
        entry.mnr_out          = _d('mnr_out', entry.mnr_out)
        entry.mynt_sales_out   = _d('mynt_sales_out', entry.mynt_sales_out)
        entry.mynt_spares_out  = _d('mynt_spares_out', entry.mynt_spares_out)
        entry.mynt_service_out = _d('mynt_service_out', entry.mynt_service_out)
        entry.zynova_out       = _d('zynova_out', entry.zynova_out)
        entry.escrow_out       = _d('escrow_out', entry.escrow_out)
        if 'notes' in payload:
            entry.notes = payload['notes']

        db.commit()
        db.refresh(entry)
        return JSONResponse(content={"success": True, "data": entry.to_dict()})
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.delete("/cashflow-register/{entry_id}")
def delete_cashflow_entry(
    entry_id: int = Path(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Delete a cashflow register entry."""
    try:
        entry = db.query(CashflowRegister).filter(CashflowRegister.id == entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Cashflow entry not found")
        db.delete(entry)
        db.commit()
        return JSONResponse(content={"success": True, "message": f"Entry for {entry.entry_date} deleted."})
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


# ══════════════════════════════════════════════════════════════════════════════
# DC_ACCT_LEDGER_001: General Ledger (Account Ledger) Endpoints
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/general-ledger/entries")
async def list_general_ledger_entries(
    company_id: Optional[int] = Query(None),
    account_type: Optional[str] = Query(None, description="CASH|BANK|UPI|INCOME|EXPENSE|STOCK"),
    account_name: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    reference_type: Optional[str] = Query(None),
    reference_number: Optional[str] = Query(None),
    particulars: Optional[str] = Query(None),
    source_status: Optional[str] = Query(None, description="Comma-separated voucher statuses: CONFIRMED,MANUAL,TALLY_IMPORT,OPENING_BALANCE"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    List general ledger entries with filters.
    DC_ACCT_LEDGER_001: Accounts/VGK/EA access required.
    DC-SOURCE-STATUS-001: source_status multi-value filter added
    """
    try:
        from app.services.staff_accounts_service import LedgerPostingService
        entries, total, total_debit, total_credit = LedgerPostingService.list_account_ledger(
            db, current_user,
            company_id=company_id, account_type=account_type, account_name=account_name,
            date_from=date_from, date_to=date_to, reference_type=reference_type,
            reference_number=reference_number, particulars=particulars,
            source_status=source_status,
            page=page, page_size=page_size
        )
        # DC_LEDGER_CATEGORY_001: batch-load category names for display
        _gl_sub_ids  = list({e.sub_category_id  for e in entries if getattr(e, 'sub_category_id',  None)})
        _gl_main_ids = list({e.main_category_id for e in entries if getattr(e, 'main_category_id', None)})
        _gl_sub_names: dict  = {}
        _gl_main_names: dict = {}
        if _gl_sub_ids or _gl_main_ids:
            try:
                from app.models.expense_category import ExpenseSubCategory as _ESC_GL, ExpenseMainCategory as _EMC_GL
                if _gl_sub_ids:
                    _gl_sub_names = {s.id: s.name for s in db.query(_ESC_GL).filter(_ESC_GL.id.in_(_gl_sub_ids)).all()}
                if _gl_main_ids:
                    _gl_main_names = {m.id: m.name for m in db.query(_EMC_GL).filter(_EMC_GL.id.in_(_gl_main_ids)).all()}
            except Exception:
                pass

        return JSONResponse(content={
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "totals": {
                "total_debit": float(total_debit),
                "total_credit": float(total_credit),
                "net_balance": float(total_debit - total_credit),
            },
            "entries": [
                {
                    "id": e.id,
                    "company_id": e.company_id,
                    "account_type": e.account_type,
                    "account_name": e.account_name,
                    "transaction_date": e.transaction_date.isoformat() if e.transaction_date else None,
                    "entry_type": e.entry_type,
                    "reference_type": e.reference_type,
                    "reference_id": e.reference_id,
                    "reference_number": e.reference_number,
                    "debit_amount": float(e.debit_amount),
                    "credit_amount": float(e.credit_amount),
                    "running_balance": float(e.running_balance),
                    "narration": e.narration,
                    "voucher_type": e.voucher_type,
                    "particulars": e.particulars,
                    "source_status": getattr(e, 'source_status', None),
                    "main_category_id": getattr(e, 'main_category_id', None),
                    "sub_category_id": getattr(e, 'sub_category_id', None),
                    "main_category_name": _gl_main_names.get(e.main_category_id) if getattr(e, 'main_category_id', None) else None,
                    "sub_category_name": _gl_sub_names.get(e.sub_category_id) if getattr(e, 'sub_category_id', None) else None,
                    "created_at": e.created_at.isoformat() if e.created_at else None,
                }
                for e in entries
            ]
        })
    except Exception as e:
        return handle_accounts_error(e)


def _recompute_account_ledger_balances(db, account_type: str, account_name: str, company_id: int):
    """Recompute running_balance for all account_ledger rows matching type+name+company, sorted by date+id."""
    from app.models.staff_accounts import AccountLedger as _AL
    entries = db.query(_AL).filter(
        _AL.account_type == account_type,
        _AL.account_name == account_name,
        _AL.company_id   == company_id,
    ).order_by(_AL.transaction_date.asc(), _AL.id.asc()).all()
    bal = Decimal('0')
    for e in entries:
        bal = bal + Decimal(str(e.debit_amount or 0)) - Decimal(str(e.credit_amount or 0))
        e.running_balance = bal
    db.flush()


@router.get("/general-ledger/entries/{entry_id}")
async def get_general_ledger_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Fetch a single account_ledger entry by ID for editing.
    If account_type=PARTY, also looks up party_type from party_ledger."""
    from app.models.staff_accounts import AccountLedger as _AL, PartyLedger as _PL
    entry = db.query(_AL).filter(_AL.id == entry_id).first()
    if not entry:
        return JSONResponse(status_code=404, content={"success": False, "message": "Entry not found"})

    party_type = None
    party_id   = None
    if entry.account_type == 'PARTY' and entry.account_name:
        pl = db.query(_PL).filter(
            _PL.company_id == entry.company_id,
            _PL.party_name == entry.account_name,
        ).order_by(_PL.id.desc()).first()
        if pl:
            party_type = pl.party_type
            party_id   = pl.party_id

    from app.models.expense_category import ExpenseMainCategory as _EMC, ExpenseSubCategory as _ESC
    _mid = getattr(entry, 'main_category_id', None)
    _sid = getattr(entry, 'sub_category_id', None)
    _main_cat_name = db.query(_EMC).filter(_EMC.id == _mid).first().name if _mid else None
    _sub_cat_name  = db.query(_ESC).filter(_ESC.id == _sid).first().name if _sid else None

    return JSONResponse(content={
        "success": True,
        "entry": {
            "id": entry.id,
            "company_id": entry.company_id,
            "account_type": entry.account_type,
            "account_name": entry.account_name,
            "transaction_date": entry.transaction_date.isoformat() if entry.transaction_date else None,
            "entry_type": entry.entry_type,
            "reference_type": entry.reference_type,
            "reference_id": entry.reference_id,
            "reference_number": entry.reference_number,
            "debit_amount": float(entry.debit_amount or 0),
            "credit_amount": float(entry.credit_amount or 0),
            "running_balance": float(entry.running_balance or 0),
            "narration": entry.narration,
            "voucher_type": entry.voucher_type,
            "particulars": entry.particulars,
            "party_type": party_type,
            "party_id":   party_id,
            "main_category_id":   _mid,
            "sub_category_id":    _sid,
            "main_category_name": _main_cat_name,
            "sub_category_name":  _sub_cat_name,
        }
    })


@router.patch("/general-ledger/entries/{entry_id}")
async def update_general_ledger_entry(
    entry_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Edit a single account_ledger entry.
    - Updates date, entry_type, amount, account_type, account_name, reference_number,
      voucher_type, narration, particulars.
    - If a counter_entry_id is provided, mirrors amount/type/date changes to the partner entry.
    - Recomputes running balances for all affected account heads.
    """
    from app.models.staff_accounts import AccountLedger as _AL
    try:
        entry = db.query(_AL).filter(_AL.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "message": "Entry not found"})

        old_account_type = entry.account_type
        old_account_name = entry.account_name
        old_company_id   = entry.company_id

        # ── Date ──────────────────────────────────────────────────────────────
        if 'transaction_date' in payload and payload['transaction_date']:
            from datetime import date as _date
            try:
                entry.transaction_date = _date.fromisoformat(payload['transaction_date'])
            except ValueError:
                return JSONResponse(status_code=400, content={"success": False, "message": "Invalid date format"})

        # ── Amount & entry_type ────────────────────────────────────────────────
        new_entry_type = payload.get('entry_type', entry.entry_type)
        if new_entry_type not in ('DEBIT', 'CREDIT'):
            return JSONResponse(status_code=400, content={"success": False, "message": "entry_type must be DEBIT or CREDIT"})
        raw_amt = Decimal(str(payload['amount'])) if 'amount' in payload else \
                  Decimal(str(entry.debit_amount or entry.credit_amount or 0))
        entry.entry_type    = new_entry_type
        entry.debit_amount  = raw_amt if new_entry_type == 'DEBIT'   else Decimal('0')
        entry.credit_amount = raw_amt if new_entry_type == 'CREDIT'  else Decimal('0')

        # ── Account head (move entry) ─────────────────────────────────────────
        new_account_type = payload.get('account_type', '').strip() or entry.account_type
        new_account_name = (payload.get('account_name') or '').strip() or entry.account_name
        entry.account_type = new_account_type
        entry.account_name = new_account_name

        # ── Metadata ──────────────────────────────────────────────────────────
        if 'reference_number' in payload:
            entry.reference_number = payload['reference_number'] or None
        if 'voucher_type' in payload:
            entry.voucher_type = payload['voucher_type'] or None
        if 'particulars' in payload:
            entry.particulars = payload['particulars'] or None
        if 'narration' in payload:
            entry.narration = payload['narration'] or None

        # ── Sync party_type to party_ledger when account_type=PARTY ──────────
        new_party_type = (payload.get('party_type') or '').strip()
        if new_account_type == 'PARTY' and (new_account_name != old_account_name or new_party_type):
            from app.models.staff_accounts import PartyLedger as _PL2
            pl_q = db.query(_PL2).filter(
                _PL2.company_id == entry.company_id,
                _PL2.party_name == old_account_name,
            )
            pl_rows = pl_q.all()
            for pl_row in pl_rows:
                if new_account_name:
                    pl_row.party_name = new_account_name
                if new_party_type:
                    pl_row.party_type = new_party_type
            if pl_rows:
                db.flush()
                _recompute_party_running_balances(
                    db,
                    new_party_type or (pl_rows[0].party_type if pl_rows else ''),
                    pl_rows[0].party_id if pl_rows else 0,
                    new_account_name,
                    entry.company_id
                )

        # ── Mirror changes to counter-entry (double-entry balance) ────────────
        counter_id = payload.get('counter_entry_id')
        if counter_id:
            counter = db.query(_AL).filter(_AL.id == int(counter_id)).first()
            if counter:
                counter.transaction_date = entry.transaction_date
                # Counter always has the opposite entry_type
                counter_type = 'CREDIT' if new_entry_type == 'DEBIT' else 'DEBIT'
                counter.entry_type    = counter_type
                counter.debit_amount  = raw_amt if counter_type == 'DEBIT'  else Decimal('0')
                counter.credit_amount = raw_amt if counter_type == 'CREDIT' else Decimal('0')
                if 'reference_number' in payload:
                    counter.reference_number = payload['reference_number'] or None
                if 'narration' in payload:
                    counter.narration = payload['narration'] or None
                if 'particulars' in payload:
                    counter.particulars = payload['particulars'] or None
                # Rebalance counter's head
                _recompute_account_ledger_balances(db, counter.account_type, counter.account_name, counter.company_id)

        # ── Recompute balances for original head (and new head if moved) ──────
        _recompute_account_ledger_balances(db, old_account_type, old_account_name, old_company_id)
        if entry.account_type != old_account_type or entry.account_name != old_account_name:
            _recompute_account_ledger_balances(db, entry.account_type, entry.account_name, entry.company_id)

        db.commit()
        return JSONResponse(content={"success": True, "message": "Entry updated and balances recomputed"})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.delete("/general-ledger/entries/{entry_id}")
async def delete_general_ledger_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Delete a single account_ledger entry and recompute running balances."""
    from app.models.staff_accounts import AccountLedger as _AL
    try:
        validate_accounts_access(current_user)
        entry = db.query(_AL).filter(_AL.id == entry_id).first()
        if not entry:
            return JSONResponse(status_code=404, content={"success": False, "message": "Entry not found"})
        atype, aname, cid = entry.account_type, entry.account_name, entry.company_id
        db.delete(entry)
        db.flush()
        _recompute_account_ledger_balances(db, atype, aname, cid)
        db.commit()
        return JSONResponse(content={"success": True, "message": "Entry deleted and balances recomputed"})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.post("/general-ledger/recompute-all-balances")
async def recompute_all_account_ledger_balances(
    company_id: Optional[int] = Query(None, description="Limit to one company (omit = all companies)"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-JV-EDIT-AL-RECOMPUTE-001: Admin data-heal endpoint.
    Recomputes stored running_balance for ALL account_ledger entries across all accounts
    (or a single company if company_id is supplied), sorted by date+id ascending.
    Use after historical JV edits made before the recompute fix was deployed, or
    to verify ledger integrity at any time.
    Access: hierarchy_level >= 80 OR role_code in {SUPER_ADMIN, CEO, CTO, FOUNDER}.
    """
    from app.models.staff_accounts import AccountLedger as _AL
    try:
        validate_accounts_access(current_user)
        _allowed_roles = {'SUPER_ADMIN', 'CEO', 'CTO', 'FOUNDER'}
        if current_user.hierarchy_level < 80 and getattr(current_user, 'role_code', '') not in _allowed_roles:
            raise HTTPException(status_code=403, detail='Super-admin access required for balance recompute')
        q = db.query(_AL.company_id, _AL.account_type, _AL.account_name).distinct()
        if company_id:
            q = q.filter(_AL.company_id == company_id)
        combos = q.all()
        healed = 0
        for (_cid, _atype, _aname) in combos:
            _recompute_account_ledger_balances(db, _atype, _aname, _cid)
            healed += 1
        db.commit()
        return JSONResponse(content={
            "success": True,
            "accounts_recomputed": healed,
            "message": f"running_balance recomputed for {healed} account(s)."
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.get("/general-ledger/company-banks")
async def list_all_company_bank_accounts(
    company_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Return all active company bank accounts across all companies (or filtered by company_id).
    Used by GL Account Heads tab to surface bank accounts registered in company settings.
    DC-GL-BANKS-001.
    """
    try:
        from app.models.staff_accounts import CompanyBankAccount as _CBA
        q = db.query(_CBA).filter(_CBA.is_active == True)
        if company_id:
            q = q.filter(_CBA.company_id == company_id)
        rows = q.order_by(_CBA.company_id, _CBA.is_primary.desc(), _CBA.id).all()
        return JSONResponse(content={
            "success": True,
            "banks": [r.to_dict() for r in rows]
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/general-ledger/heads")
async def list_general_ledger_heads(
    company_id: Optional[int] = Query(None),
    account_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Aggregate view: all account heads with total debit/credit/balance.
    DC_ACCT_LEDGER_001.
    """
    try:
        from app.services.staff_accounts_service import LedgerPostingService
        rows = LedgerPostingService.list_account_heads(db, current_user, company_id=company_id, account_type=account_type)
        return JSONResponse(content={
            "success": True,
            "heads": [
                {
                    "account_type": r.account_type,
                    "account_name": r.account_name,
                    "company_id": r.company_id,
                    "total_debit": float(r.total_debit or 0),
                    "total_credit": float(r.total_credit or 0),
                    "balance": float((r.total_debit or 0) - (r.total_credit or 0)),
                    "last_date": r.last_date.isoformat() if r.last_date else None,
                }
                for r in rows
            ]
        })
    except Exception as e:
        return handle_accounts_error(e)


# ══════════════════════════════════════════════════════════════════════════════
# DC-BANK-002b: Party search for journal voucher
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/party-search")
async def party_search(
    q: str = Query(..., min_length=2),
    company_id: Optional[int] = Query(None),
    limit: int = Query(30, le=60),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Search ALL party types for journal voucher autocomplete. DC-BANK-002b.
    NO active/status filter — every person in the system is searchable regardless of status.
    Sources: VGK Members, MNR Members, Partners (all cats), Vendors, Staff, Companies, Manual/External parties.
    DC-PARTY-COMPANY-001: When company_id is provided, vendors are filtered to those whose
    applicable_companies includes that company_id OR the 'ALL' wildcard.
    All other sources (members, staff, companies) remain unfiltered — they are global entities.
    """
    try:
        from app.models.staff_accounts import VendorMaster as _VM, AssociatedCompany as _AC, OfficialPartner as _OP, ManualPartyMaster as _MP
        from app.models.staff import StaffEmployee as _SE
        from app.models.user import User as _User
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        results = []
        pat = f'%{q}%'

        # 1. VGK Members — NO status filter, all VGK team members searchable
        for p in db.query(_OP).filter(
            _OP.partner_name.ilike(pat) | _OP.partner_code.ilike(pat) |
            _OP.phone.ilike(pat) | _OP.contact_person.ilike(pat),
            _OP.category == 'VGK_TEAM'
        ).limit(8).all():
            results.append({
                'id': p.id, 'name': p.partner_name, 'type': 'VGK_MEMBER',
                'sub': p.partner_code,
                'phone': p.phone or p.whatsapp_number or p.contact_person_1_phone
            })

        # 2. MNR Members — all registered platform members
        for u in db.query(_User).filter(
            _User.name.ilike(pat) | _User.id.ilike(pat) | _User.email.ilike(pat)
        ).limit(8).all():
            results.append({
                'id': u.id, 'name': u.name, 'type': 'MNR_MEMBER',
                'sub': str(u.id),
                'phone': getattr(u, 'phone_number', None)
            })

        # 3. Partners — all categories, NO status filter
        _PCAT_LABEL = {
            'DEALER': 'DEALER', 'DISTRIBUTOR': 'DISTRIBUTOR',
            'REAL_DREAM_PARTNER': 'RD_PARTNER', 'SERVICE_CENTER': 'SERVICE_CENTER',
            'VENDOR': 'PARTNER_VENDOR'
        }
        for p in db.query(_OP).filter(
            _OP.partner_name.ilike(pat) | _OP.partner_code.ilike(pat) |
            _OP.contact_person.ilike(pat) | _OP.contact_person_1_name.ilike(pat) |
            _OP.phone.ilike(pat),
            _OP.category.in_(('DEALER', 'DISTRIBUTOR', 'REAL_DREAM_PARTNER', 'SERVICE_CENTER', 'VENDOR'))
        ).limit(10).all():
            results.append({
                'id': p.id, 'name': p.partner_name,
                'type': _PCAT_LABEL.get(p.category, 'PARTNER'),
                'sub': p.partner_code,
                'phone': p.phone or p.whatsapp_number or p.contact_person_1_phone
            })

        # 4. Vendors — DC-PARTY-COMPANY-001: filter by company when company_id provided
        _vendor_q = db.query(_VM).filter(
            _VM.vendor_name.ilike(pat) | _VM.vendor_code.ilike(pat) |
            _VM.contact_person.ilike(pat) | _VM.contact_person_1_name.ilike(pat) |
            _VM.phone.ilike(pat) | _VM.contact_person_1_phone.ilike(pat)
        )
        if company_id:
            _vendor_q = _vendor_q.filter(
                _VM.applicable_companies.contains([company_id]) |
                _VM.applicable_companies.contains(['ALL'])
            )
        for v in _vendor_q.limit(8).all():
            results.append({
                'id': v.id, 'name': v.vendor_name, 'type': 'VENDOR',
                'sub': getattr(v, 'vendor_code', None),
                'phone': getattr(v, 'phone', None) or getattr(v, 'contact_person_1_phone', None)
            })

        # 5. Staff — NO status filter (active + resigned + deactivated all searchable)
        for s in db.query(_SE).filter(
            _SE.full_name.ilike(pat) | _SE.email.ilike(pat) | _SE.emp_code.ilike(pat)
        ).limit(8).all():
            results.append({
                'id': s.id, 'name': s.full_name, 'type': 'STAFF',
                'sub': getattr(s, 'emp_code', None),
                'phone': getattr(s, 'phone', None)
            })

        # 6. Associated Companies
        for c in db.query(_AC).filter(_AC.company_name.ilike(pat)).limit(5).all():
            results.append({'id': c.id, 'name': c.company_name, 'type': 'COMPANY', 'sub': None, 'phone': None})

        # 7. Manual / External parties saved by staff (persisted for future reuse)
        for m in db.query(_MP).filter(
            _MP.name.ilike(pat) | _MP.phone.ilike(pat) | _MP.email.ilike(pat)
        ).order_by(_MP.id.desc()).limit(8).all():
            results.append({
                'id': f'MP:{m.id}', 'name': m.name, 'type': 'EXTERNAL',
                'sub': m.notes or None,
                'phone': m.phone or m.email
            })

        # DC-PARTY-DEDUP-001: Deduplicate by name only (case-insensitive), keep first occurrence
        # Previously deduped by name+type which caused same company to appear multiple times
        seen = set()
        unique = []
        for r in results:
            key = (r['name'] or '').lower()
            if key not in seen:
                seen.add(key)
                unique.append(r)
        return JSONResponse(content={'results': unique[:limit]})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/official-partners")
async def list_official_partners(
    category: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List official partners (dealers, distributors, service centers, vendors) for Parties Master.
    Excludes VGK_TEAM by default unless category=VGK_TEAM is explicitly passed."""
    from app.models.staff_accounts import OfficialPartner as _OP
    from app.services.staff_accounts_service import validate_accounts_access
    validate_accounts_access(current_user)
    try:
        q = db.query(_OP)
        if category:
            q = q.filter(_OP.category == category)
        else:
            q = q.filter(_OP.category != 'VGK_TEAM')
        if search:
            pat = f'%{search}%'
            q = q.filter(
                _OP.partner_name.ilike(pat) | _OP.partner_code.ilike(pat) |
                _OP.phone.ilike(pat) | _OP.gst_number.ilike(pat)
            )
        if is_active is not None:
            q = q.filter(_OP.is_active == is_active)
        partners = q.order_by(_OP.category, _OP.partner_name).all()
        # DC_PARTNER_OB_001: Enrich each partner with OB from AccountLedgerMaster
        from app.models.staff_accounts import AccountLedgerMaster as _ALM_PRT
        _alm_map = {}
        try:
            _codes = [p.partner_code for p in partners if p.partner_code]
            if _codes:
                _alm_rows = db.query(_ALM_PRT).filter(
                    _ALM_PRT.account_code.in_(_codes),
                    _ALM_PRT.account_type == 'PARTY'
                ).all()
                for _r in _alm_rows:
                    if _r.account_code not in _alm_map:
                        _alm_map[_r.account_code] = _r
        except Exception:
            pass
        return JSONResponse(content={
            "success": True,
            "partners": [{
                "id": p.id, "code": p.partner_code, "name": p.partner_name,
                "category": p.category, "phone": p.phone or p.contact_person_1_phone,
                "email": p.email, "gst": p.gst_number, "pan": p.pan_number,
                "city": p.city, "state": p.state, "address": p.address,
                "is_active": p.is_active,
                "opening_balance": float(_alm_map[p.partner_code].opening_balance or 0) if p.partner_code in _alm_map else 0,
                "opening_balance_type": (_alm_map[p.partner_code].opening_balance_type or 'CREDIT') if p.partner_code in _alm_map else 'CREDIT',
                "opening_balance_date": str(_alm_map[p.partner_code].opening_balance_date) if p.partner_code in _alm_map and _alm_map[p.partner_code].opening_balance_date else None
            } for p in partners]
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/official-partners")
async def create_official_partner(
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Create a new official partner. DC-PARTNER-CRUD-001."""
    from app.models.staff_accounts import OfficialPartner as _OP
    from app.services.staff_accounts_service import validate_accounts_access
    validate_accounts_access(current_user)
    try:
        data = await request.json()
        partner_name = (data.get('partner_name') or '').strip()
        if not partner_name:
            return JSONResponse(status_code=400, content={"detail": "Partner name is required"})
        category = (data.get('category') or 'DEALER').strip().upper()
        # Auto-generate partner code if not provided
        partner_code = (data.get('partner_code') or '').strip().upper()
        if not partner_code:
            prefix = {'DEALER': 'DLR', 'DISTRIBUTOR': 'DST', 'SERVICE_CENTER': 'SVC',
                      'VENDOR': 'VND', 'REAL_DREAM_PARTNER': 'RDP'}.get(category, 'PRT')
            import random, string
            suffix = ''.join(random.choices(string.digits, k=4))
            partner_code = f"{prefix}{suffix}"
            while db.query(_OP).filter(_OP.partner_code == partner_code).first():
                suffix = ''.join(random.choices(string.digits, k=4))
                partner_code = f"{prefix}{suffix}"
        # Check uniqueness
        if db.query(_OP).filter(_OP.partner_code == partner_code).first():
            return JSONResponse(status_code=400, content={"detail": f"Partner code '{partner_code}' already exists"})
        p = _OP(
            partner_code=partner_code,
            partner_name=partner_name,
            category=category,
            phone=data.get('phone'),
            email=data.get('email'),
            whatsapp_number=data.get('whatsapp_number'),
            contact_person=data.get('contact_person'),
            contact_person_1_name=data.get('contact_person_1_name'),
            contact_person_1_phone=data.get('contact_person_1_phone'),
            contact_person_1_designation=data.get('contact_person_1_designation'),
            contact_person_2_name=data.get('contact_person_2_name'),
            contact_person_2_phone=data.get('contact_person_2_phone'),
            gst_number=(data.get('gst_number') or '').upper().strip() or None,
            pan_number=(data.get('pan_number') or '').upper().strip() or None,
            address=data.get('address'),
            city=data.get('city'),
            state=data.get('state'),
            pincode=data.get('pincode'),
            bank_name=data.get('bank_name'),
            bank_branch=data.get('bank_branch'),
            account_number=data.get('account_number'),
            ifsc_code=(data.get('ifsc_code') or '').upper().strip() or None,
            payment_terms=data.get('payment_terms', 'ADVANCE'),
            credit_limit=data.get('credit_limit', 0),
            credit_days=data.get('credit_days', 0),
            is_active=data.get('is_active', True),
        )
        db.add(p)
        db.commit()
        db.refresh(p)
        # DC_PARTNER_OB_001: Auto-create AccountLedgerMaster party rows with opening balance
        try:
            from app.models.staff_accounts import AccountLedgerMaster as _ALM_PC, AssociatedCompany as _ACO_PC
            from decimal import Decimal as _D
            from app.services.staff_accounts_service import get_indian_time as _git
            _ob_amt  = _D(str(data.get('opening_balance', 0) or 0))
            _ob_type = (data.get('opening_balance_type') or 'CREDIT').upper()
            _ob_date_raw = data.get('opening_balance_date')
            from datetime import date as _date
            _ob_date = _date.fromisoformat(_ob_date_raw) if _ob_date_raw else None
            _now_pc  = _git()
            _cos = db.query(_ACO_PC).filter(_ACO_PC.is_active == True).all()
            for _co in _cos:
                _ex = db.query(_ALM_PC).filter(
                    _ALM_PC.company_id == _co.id,
                    _ALM_PC.account_type == 'PARTY',
                    _ALM_PC.account_code == p.partner_code
                ).first()
                if not _ex:
                    db.add(_ALM_PC(
                        company_id=_co.id,
                        account_type='PARTY',
                        account_name=p.partner_name.strip(),
                        account_code=p.partner_code,
                        description=f'Auto-created party ledger for Partner: {p.partner_name}',
                        parent_group='Partners & Sundry Creditors',
                        opening_balance=_ob_amt,
                        opening_balance_type=_ob_type,
                        opening_balance_date=_ob_date,
                        opening_balance_posted=True,
                        is_active=True,
                        created_by_id=current_user.id,
                        created_at=_now_pc,
                        updated_at=_now_pc,
                    ))
            db.commit()
        except Exception as _pce:
            import logging as _log
            _log.getLogger(__name__).warning(f"Partner ALM auto-create: {_pce}")
        return JSONResponse(content={
            "success": True,
            "message": f"Partner '{p.partner_name}' created",
            "partner": {"id": p.id, "code": p.partner_code, "name": p.partner_name,
                        "category": p.category, "phone": p.phone, "email": p.email,
                        "gst": p.gst_number, "city": p.city, "state": p.state, "is_active": p.is_active}
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/official-partners/{partner_id}")
async def update_official_partner(
    partner_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update an official partner. DC-PARTNER-CRUD-001."""
    from app.models.staff_accounts import OfficialPartner as _OP
    from app.services.staff_accounts_service import validate_accounts_access
    validate_accounts_access(current_user)
    try:
        p = db.query(_OP).filter(_OP.id == partner_id).first()
        if not p:
            return JSONResponse(status_code=404, content={"detail": "Partner not found"})
        data = await request.json()
        _str_fields = ['partner_name','category','phone','email','whatsapp_number','contact_person',
                       'contact_person_1_name','contact_person_1_phone','contact_person_1_designation',
                       'contact_person_2_name','contact_person_2_phone','address','city','state',
                       'pincode','bank_name','bank_branch','account_number','payment_terms']
        for f in _str_fields:
            if f in data:
                setattr(p, f, data[f])
        if 'gst_number' in data:
            p.gst_number = (data['gst_number'] or '').upper().strip() or None
        if 'pan_number' in data:
            p.pan_number = (data['pan_number'] or '').upper().strip() or None
        if 'ifsc_code' in data:
            p.ifsc_code = (data['ifsc_code'] or '').upper().strip() or None
        if 'is_active' in data:
            p.is_active = bool(data['is_active'])
        if 'credit_limit' in data:
            p.credit_limit = data['credit_limit']
        if 'credit_days' in data:
            p.credit_days = data['credit_days']
        db.commit()
        db.refresh(p)
        # DC_PARTNER_OB_001: Apply opening balance override to ALM rows
        if 'opening_balance' in data:
            try:
                from app.models.staff_accounts import AccountLedgerMaster as _ALM_PU
                from decimal import Decimal as _DP
                from app.services.staff_accounts_service import get_indian_time as _gitu
                _alm_pu_rows = db.query(_ALM_PU).filter(
                    _ALM_PU.account_code == p.partner_code,
                    _ALM_PU.account_type == 'PARTY'
                ).all()
                for _row in _alm_pu_rows:
                    _row.opening_balance = _DP(str(data['opening_balance']))
                    if data.get('opening_balance_type'):
                        _row.opening_balance_type = data['opening_balance_type'].upper()
                    if data.get('opening_balance_date') is not None:
                        from datetime import date as _dateu
                        _row.opening_balance_date = _dateu.fromisoformat(data['opening_balance_date']) if data['opening_balance_date'] else None
                    _row.opening_balance_posted = True
                    _row.updated_at = _gitu()
                db.commit()
            except Exception as _pue:
                import logging as _logu
                _logu.getLogger(__name__).warning(f"Partner OB override: {_pue}")
        return JSONResponse(content={
            "success": True,
            "message": f"Partner '{p.partner_name}' updated",
            "partner": {"id": p.id, "code": p.partner_code, "name": p.partner_name,
                        "category": p.category, "phone": p.phone, "email": p.email,
                        "gst": p.gst_number, "city": p.city, "state": p.state, "is_active": p.is_active}
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/official-partners/{partner_id}")
async def delete_official_partner(
    partner_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Soft-delete (deactivate) an official partner. DC-PARTNER-CRUD-001."""
    from app.models.staff_accounts import OfficialPartner as _OP
    from app.services.staff_accounts_service import validate_accounts_access
    validate_accounts_access(current_user)
    try:
        p = db.query(_OP).filter(_OP.id == partner_id).first()
        if not p:
            return JSONResponse(status_code=404, content={"detail": "Partner not found"})
        p.is_active = False
        db.commit()
        return JSONResponse(content={"success": True, "message": f"Partner '{p.partner_name}' deactivated"})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/staff-list")
async def list_staff_for_parties(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List all staff employees for Parties Master page."""
    from app.models.staff import StaffEmployee as _SE
    from app.services.staff_accounts_service import validate_accounts_access
    validate_accounts_access(current_user)
    try:
        q = db.query(_SE)
        if search:
            pat = f'%{search}%'
            q = q.filter(_SE.full_name.ilike(pat) | _SE.emp_code.ilike(pat) | _SE.phone.ilike(pat))
        if status:
            q = q.filter(_SE.status == status)
        staff = q.order_by(_SE.full_name).all()
        return JSONResponse(content={
            "success": True,
            "staff": [{
                "id": s.id, "code": s.emp_code, "name": s.full_name,
                "role": s.role.role_name if s.role else None,
                "role_code": s.role.role_code if s.role else None,
                "department": s.department.department_name if s.department else None,
                "phone": s.phone, "status": s.status
            } for s in staff]
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/party-search/add-manual")
async def add_manual_party(
    request: Request,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Save an external / manual party to manual_party_master so it appears in future party searches.
    Idempotent: if same name (case-insensitive) already exists, returns existing record.
    DC-BANK-002c.
    """
    try:
        from app.models.staff_accounts import ManualPartyMaster as _MP
        data = await request.json()
        name = (data.get('name') or '').strip()
        phone = (data.get('phone') or '').strip() or None
        email = (data.get('email') or '').strip() or None
        notes = (data.get('notes') or '').strip() or None
        if not name:
            raise HTTPException(status_code=422, detail='name is required')
        # Idempotent: find existing by name (case-insensitive)
        existing = db.query(_MP).filter(_MP.name.ilike(name)).first()
        if existing:
            return JSONResponse(content={
                'id': f'MP:{existing.id}', 'name': existing.name,
                'type': 'EXTERNAL', 'created': False
            })
        mp = _MP(
            name=name, phone=phone, email=email, notes=notes,
            created_by_id=current_user.id
        )
        db.add(mp)
        db.commit()
        db.refresh(mp)
        return JSONResponse(content={
            'id': f'MP:{mp.id}', 'name': mp.name,
            'type': 'EXTERNAL', 'created': True
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ══════════════════════════════════════════════════════════════════════════════
# DC-BANK-002: Open invoice pickers for voucher linking
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/journal-vouchers/open-purchases")
async def list_open_purchases_for_payment(
    company_id: int = Query(...),
    vendor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Open/partial purchase invoices for PAYMENT voucher linking. DC-BANK-002."""
    try:
        from app.models.staff_accounts import VendorTransactionHeader as _VTH, VendorMaster as _VM
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        q = db.query(_VTH).filter(
            _VTH.company_id == company_id,
            _VTH.payment_status.in_(['PENDING', 'PARTIAL_PAID', 'OVERDUE']),
            _VTH.balance_due > 0,
        )
        if vendor_id:
            q = q.filter(_VTH.vendor_id == vendor_id)
        rows = q.order_by(_VTH.transaction_date.desc()).limit(100).all()
        result = []
        vendor_ids = list({r.vendor_id for r in rows})
        vendors = {v.id: v for v in db.query(_VM).filter(_VM.id.in_(vendor_ids)).all()} if vendor_ids else {}
        for r in rows:
            v = vendors.get(r.vendor_id)
            result.append({
                'id': r.id,
                'transaction_number': r.transaction_number,
                'transaction_date': r.transaction_date.isoformat() if r.transaction_date else None,
                'vendor_id': r.vendor_id,
                'vendor_name': v.vendor_name if v else '—',
                'grand_total': float(r.grand_total),
                'amount_paid': float(r.amount_paid),
                'balance_due': float(r.balance_due),
                'payment_status': r.payment_status,
                'due_date': r.due_date.isoformat() if r.due_date else None,
                'label': f"{r.transaction_number} — {v.vendor_name if v else ''} — ₹{float(r.balance_due):,.2f} due",
            })
        return JSONResponse(content={"purchases": result})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/journal-vouchers/open-sales")
async def list_open_sales_for_receipt(
    company_id: int = Query(...),
    party_name: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Open/partial sales invoices for RECEIPT voucher linking. DC-BANK-002."""
    try:
        from app.models.staff_accounts import GeneratedInvoice as _GI
        from app.services.staff_accounts_service import validate_accounts_access
        validate_accounts_access(current_user)
        q = db.query(_GI).filter(
            _GI.company_id == company_id,
            _GI.payment_status.in_(['UNPAID', 'PARTIAL']),
        )
        if party_name:
            q = q.filter(_GI.customer_name.ilike(f'%{party_name}%'))
        rows = q.order_by(_GI.invoice_date.desc()).limit(100).all()
        result = []
        for r in rows:
            paid = float(getattr(r, 'amount_paid', 0) or 0)
            total = float(r.total_amount or 0)
            balance = round(total - paid, 2)
            result.append({
                'id': r.id,
                'invoice_number': r.invoice_number,
                'invoice_date': r.invoice_date.isoformat() if r.invoice_date else None,
                'customer_name': r.customer_name,
                'total_amount': total,
                'amount_paid': paid,
                'balance_due': balance,
                'payment_status': r.payment_status,
                'label': f"{r.invoice_number} — {r.customer_name} — ₹{balance:,.2f} due",
            })
        return JSONResponse(content={"sales": result})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ══════════════════════════════════════════════════════════════════════════════
# DC_JOURNAL_001: Journal / Transfer Voucher Endpoints
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/journal-vouchers")
async def list_journal_vouchers(
    company_id: Optional[int] = Query(None),
    voucher_type: Optional[str] = Query(None),
    party_name: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    reference_number: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import JournalVoucherService
        rows, total, total_amount = JournalVoucherService.list_vouchers(
            db, current_user,
            company_id=company_id, voucher_type=voucher_type, party_name=party_name,
            date_from=date_from, date_to=date_to, status=status,
            reference_number=reference_number,
            page=page, page_size=page_size
        )
        # DC_LEDGER_CATEGORY_001: batch-load category names for display
        _cat_ids = list({r.category_id for r in rows if r.category_id})
        _sub_name_map: dict = {}
        _main_name_map: dict = {}
        _main_cat_id_map: dict = {}
        if _cat_ids:
            try:
                from app.models.expense_category import ExpenseSubCategory as _ESC_L, ExpenseMainCategory as _EMC_L
                _subs = db.query(_ESC_L).filter(_ESC_L.id.in_(_cat_ids)).all()
                _main_ids = list({s.main_category_id for s in _subs if s.main_category_id})
                _mains = db.query(_EMC_L).filter(_EMC_L.id.in_(_main_ids)).all() if _main_ids else []
                _main_name_map = {m.id: m.name for m in _mains}
                for s in _subs:
                    _sub_name_map[s.id] = s.name
                    _main_cat_id_map[s.id] = s.main_category_id
            except Exception:
                pass

        def _jv_row(r):
            _s = _sub_name_map.get(r.category_id) if r.category_id else None
            _m = _main_name_map.get(_main_cat_id_map.get(r.category_id)) if r.category_id else None
            return _jv_dict(r, main_category_name=_m, sub_category_name=_s)

        return JSONResponse(content={
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "total_amount": float(total_amount),
            "vouchers": [_jv_row(r) for r in rows]
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/journal-vouchers")
async def create_journal_voucher(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import JournalVoucherService, validate_voucher_structural_integrity
        from decimal import Decimal as _D
        lines = payload.get('lines')
        v_type = payload.get('voucher_type', 'JOURNAL')

        if lines and isinstance(lines, list) and len(lines) >= 2:
            # DC-SGV-001: structural validation before any DB write (advisory mode)
            validate_voucher_structural_integrity(v_type, lines, strict=False)

            jv = JournalVoucherService.create_compound(
                db, current_user,
                company_id=int(payload['company_id']),
                voucher_date=date.fromisoformat(payload['voucher_date']),
                voucher_type=v_type,
                lines=lines,
                narration=payload.get('narration'),
                payment_mode=payload.get('payment_mode'),
                reference_number=payload.get('reference_number'),
                category_id=int(payload['category_id']) if payload.get('category_id') else None,
                income_category_id=int(payload['income_category_id']) if payload.get('income_category_id') else None,
            )
        else:
            # DC-SGV-001: synthesise lines from simple dr/cr fields for validation
            _amt = payload.get('amount', 0)
            _simple_lines = [
                {
                    'account_type': payload.get('dr_account_type', ''),
                    'account_name': payload.get('dr_account_name', ''),
                    'entry_type': 'DEBIT',
                    'amount': _amt,
                },
                {
                    'account_type': payload.get('cr_account_type', ''),
                    'account_name': payload.get('cr_account_name', ''),
                    'entry_type': 'CREDIT',
                    'amount': _amt,
                },
            ]
            validate_voucher_structural_integrity(v_type, _simple_lines, strict=False)

            jv = JournalVoucherService.create(
                db, current_user,
                company_id=int(payload['company_id']),
                voucher_date=date.fromisoformat(payload['voucher_date']),
                voucher_type=v_type,
                dr_account_type=payload['dr_account_type'],
                dr_account_name=payload['dr_account_name'],
                cr_account_type=payload['cr_account_type'],
                cr_account_name=payload['cr_account_name'],
                amount=_D(str(_amt)),
                narration=payload.get('narration'),
                party_type=payload.get('party_type'),
                party_name=payload.get('party_name'),
                party_id=payload.get('party_id'),
                payment_mode=payload.get('payment_mode'),
                reference_number=payload.get('reference_number'),
                linked_doc_type=payload.get('linked_doc_type'),
                linked_doc_id=int(payload['linked_doc_id']) if payload.get('linked_doc_id') else None,
                category_id=int(payload['category_id']) if payload.get('category_id') else None,
                income_category_id=int(payload['income_category_id']) if payload.get('income_category_id') else None,
            )
        return JSONResponse(content={"success": True, "voucher": _jv_dict(jv)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/journal-vouchers/{jv_id}")
async def update_journal_voucher(
    jv_id: int = Path(...),
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-JV-EDIT-001: Update a POSTED journal voucher (Accounts dept + MR10001)."""
    try:
        from app.services.staff_accounts_service import JournalVoucherService
        from decimal import Decimal as _D
        jv = JournalVoucherService.update(
            db, current_user,
            jv_id=jv_id,
            voucher_date=date.fromisoformat(payload['voucher_date']),
            dr_account_type=payload['dr_account_type'],
            dr_account_name=payload['dr_account_name'],
            cr_account_type=payload['cr_account_type'],
            cr_account_name=payload['cr_account_name'],
            amount=_D(str(payload['amount'])),
            narration=payload.get('narration'),
            party_type=payload.get('party_type'),
            party_name=payload.get('party_name'),
            party_id=int(payload['party_id']) if payload.get('party_id') else None,
            payment_mode=payload.get('payment_mode'),
            reference_number=payload.get('reference_number'),
            category_id=int(payload['category_id']) if payload.get('category_id') else None,
            income_category_id=int(payload['income_category_id']) if payload.get('income_category_id') else None,
        )
        return JSONResponse(content={"success": True, "voucher": _jv_dict(jv)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/journal-vouchers/{jv_id}/cancel")
async def cancel_journal_voucher(
    jv_id: int = Path(...),
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import JournalVoucherService
        snapshot = JournalVoucherService.cancel(db, current_user, jv_id, cancel_reason=payload.get('cancel_reason'))
        return JSONResponse(content={"success": True, "deleted": True, "voucher": snapshot})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


def _jv_dict(jv, main_category_name: str = None, sub_category_name: str = None) -> dict:
    return {
        "id": jv.id,
        "company_id": jv.company_id,
        "voucher_number": jv.voucher_number,
        "voucher_date": jv.voucher_date.isoformat() if jv.voucher_date else None,
        "voucher_type": jv.voucher_type,
        "dr_account_type": jv.dr_account_type,
        "dr_account_name": jv.dr_account_name,
        "cr_account_type": jv.cr_account_type,
        "cr_account_name": jv.cr_account_name,
        "party_type": jv.party_type,
        "party_name": jv.party_name,
        "party_id": jv.party_id,
        "amount": float(jv.amount),
        "narration": jv.narration,
        "payment_mode": jv.payment_mode,
        "reference_number": jv.reference_number,
        "status": jv.status,
        "created_by_id": jv.created_by_id,
        "cancelled_at": jv.cancelled_at.isoformat() if jv.cancelled_at else None,
        "cancel_reason": jv.cancel_reason,
        "created_at": jv.created_at.isoformat() if jv.created_at else None,
        "category_id": jv.category_id if hasattr(jv, 'category_id') else None,
        "income_category_id": jv.income_category_id if hasattr(jv, 'income_category_id') else None,
        "main_category_name": main_category_name,
        "sub_category_name": sub_category_name,
    }


@router.get("/journal-vouchers/{jv_id}")
async def get_journal_voucher_detail(
    jv_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-JV-COMPOUND-001: Return a single voucher with all its lines.
    For compound vouchers: returns lines from journal_voucher_lines.
    For legacy simple vouchers: synthesises two lines from the header DR/CR columns.
    """
    try:
        from app.models.staff_accounts import JournalVoucher as _JV, JournalVoucherLine as _JVL
        from sqlalchemy import text as _dt
        jv = db.query(_JV).filter(_JV.id == jv_id).first()
        if not jv:
            raise HTTPException(status_code=404, detail='Voucher not found')
        validate_accounts_access(current_user)

        lines_rows = db.query(_JVL).filter(_JVL.voucher_id == jv_id).order_by(_JVL.sort_order, _JVL.id).all()
        if lines_rows:
            lines = [
                {
                    "id": l.id,
                    "entry_type": l.entry_type,
                    "account_type": l.account_type,
                    "account_name": l.account_name,
                    "amount": float(l.amount),
                    "party_name": l.party_name,
                    "party_type": l.party_type,
                    "party_id": l.party_id,
                    "line_narration": l.line_narration,
                    "sort_order": l.sort_order,
                }
                for l in lines_rows
            ]
            is_compound = True
        else:
            lines = [
                {
                    "id": None,
                    "entry_type": "DEBIT",
                    "account_type": jv.dr_account_type,
                    "account_name": jv.dr_account_name,
                    "amount": float(jv.amount),
                    "party_name": jv.party_name if jv.dr_account_type == 'PARTY' else None,
                    "party_type": jv.party_type if jv.dr_account_type == 'PARTY' else None,
                    "party_id": jv.party_id if jv.dr_account_type == 'PARTY' else None,
                    "line_narration": None,
                    "sort_order": 0,
                },
                {
                    "id": None,
                    "entry_type": "CREDIT",
                    "account_type": jv.cr_account_type,
                    "account_name": jv.cr_account_name,
                    "amount": float(jv.amount),
                    "party_name": jv.party_name if jv.cr_account_type == 'PARTY' else None,
                    "party_type": jv.party_type if jv.cr_account_type == 'PARTY' else None,
                    "party_id": jv.party_id if jv.cr_account_type == 'PARTY' else None,
                    "line_narration": None,
                    "sort_order": 1,
                },
            ]
            is_compound = False

        company_name = None
        try:
            from app.models.associated_company import AssociatedCompany as _AC
            ac = db.query(_AC).filter(_AC.id == jv.company_id).first()
            company_name = ac.name if ac else None
        except Exception:
            pass

        creator_name = None
        try:
            from app.models.staff_employee import StaffEmployee as _SE
            se = db.query(_SE).filter(_SE.id == jv.created_by_id).first()
            creator_name = se.full_name if se else None
        except Exception:
            pass

        return JSONResponse(content={
            "success": True,
            "voucher": _jv_dict(jv),
            "company_name": company_name,
            "creator_name": creator_name,
            "is_compound": is_compound,
            "lines": lines,
        })
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ══════════════════════════════════════════════════════════════════════════════
# DC_LEDGER_MASTER_001: Chart of Accounts / Ledger Master Endpoints
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/ledger-masters")
async def list_ledger_masters(
    company_id: Optional[int] = Query(None),
    account_type: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import LedgerMasterService
        rows, total = LedgerMasterService.list_masters(
            db, current_user,
            company_id=company_id, account_type=account_type,
            is_active=is_active, search=search,
            page=page, page_size=page_size
        )
        return JSONResponse(content={
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "masters": [_alm_dict(r) for r in rows]
        })
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/ledger-masters")
async def create_ledger_master(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import LedgerMasterService
        from decimal import Decimal as _D
        ob = payload.get('opening_balance')
        ob_date = payload.get('opening_balance_date')
        alm = LedgerMasterService.create(
            db, current_user,
            company_id=int(payload['company_id']),
            account_type=payload['account_type'],
            account_name=payload['account_name'],
            account_code=payload.get('account_code'),
            description=payload.get('description'),
            parent_group=payload.get('parent_group'),
            opening_balance=_D(str(ob)) if ob else _D('0'),
            opening_balance_type=payload.get('opening_balance_type', 'DEBIT'),
            opening_balance_date=date.fromisoformat(ob_date) if ob_date else None,
            account_number=payload.get('account_number'),
            ifsc_code=payload.get('ifsc_code'),
            bank_name=payload.get('bank_name'),
        )
        return JSONResponse(content={"success": True, "master": _alm_dict(alm)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/ledger-masters/{master_id}")
async def update_ledger_master(
    master_id: int = Path(...),
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import LedgerMasterService
        from decimal import Decimal as _D
        ob = payload.get('opening_balance')
        ob_date = payload.get('opening_balance_date')
        alm = LedgerMasterService.update(
            db, current_user, master_id,
            account_name=payload.get('account_name'),
            account_code=payload.get('account_code'),
            description=payload.get('description'),
            parent_group=payload.get('parent_group'),
            opening_balance=_D(str(ob)) if ob is not None else None,
            opening_balance_type=payload.get('opening_balance_type'),
            opening_balance_date=date.fromisoformat(ob_date) if ob_date else None,
            is_active=payload.get('is_active'),
            account_number=payload.get('account_number'),
            ifsc_code=payload.get('ifsc_code'),
            bank_name=payload.get('bank_name'),
        )
        return JSONResponse(content={"success": True, "master": _alm_dict(alm)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/ledger-masters/{master_id}")
async def get_ledger_master(
    master_id: int = Path(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    try:
        from app.services.staff_accounts_service import LedgerMasterService
        alm = LedgerMasterService.get(db, current_user, master_id)
        return JSONResponse(content={"success": True, "master": _alm_dict(alm)})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/ledger-masters/{master_id}")
async def delete_ledger_master(
    master_id: int = Path(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-LEDGER-DELETE-001: Delete a ledger account.
    Restricted to MR10001 and MR10025 (Subash Kumar Kari) only.
    Fails if any account_ledger transactions are linked to this ledger.
    """
    try:
        from app.services.staff_accounts_service import validate_accounts_access
        from app.models.staff_accounts import AccountLedgerMaster as _ALM, AccountLedger as _AL

        validate_accounts_access(current_user)

        _allowed = ('MR10001', 'MR10025')
        if getattr(current_user, 'emp_code', '') not in _allowed:
            raise HTTPException(
                status_code=403,
                detail='Only MR10001 or Subash (MR10025) can delete ledger accounts.'
            )

        alm = db.query(_ALM).filter(_ALM.id == master_id).first()
        if not alm:
            raise HTTPException(status_code=404, detail='Ledger account not found.')

        # Reject if any transactions are linked
        txn_count = db.query(_AL).filter(
            _AL.account_name == alm.account_name,
            _AL.company_id == alm.company_id
        ).count()
        if txn_count > 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f'Cannot delete "{alm.account_name}" — {txn_count} transaction(s) are linked. '
                    'Edit those ledger entries to unlink first, then retry.'
                )
            )

        name = alm.account_name
        db.delete(alm)
        db.commit()
        return JSONResponse(content={"success": True, "message": f'Ledger "{name}" deleted.'})
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


def _alm_dict(alm) -> dict:
    return {
        "id": alm.id,
        "company_id": alm.company_id,
        "account_type": alm.account_type,
        "account_name": alm.account_name,
        "account_code": alm.account_code,
        "description": alm.description,
        "parent_group": alm.parent_group,
        "opening_balance": float(alm.opening_balance or 0),
        "opening_balance_type": alm.opening_balance_type,
        "opening_balance_date": alm.opening_balance_date.isoformat() if alm.opening_balance_date else None,
        "opening_balance_posted": alm.opening_balance_posted,
        "is_active": alm.is_active,
        "account_number": getattr(alm, 'account_number', None),
        "ifsc_code":      getattr(alm, 'ifsc_code', None),
        "bank_name":      getattr(alm, 'bank_name', None),
        "created_at": alm.created_at.isoformat() if alm.created_at else None,
        "updated_at": alm.updated_at.isoformat() if alm.updated_at else None,
    }


# ═══════════════════════════════════════════════════════════════════════════
# STOCK SYSTEM — DC_STOCK_MULTICOMP_001
# T02+T03: Migration, Summary, Ledger, Margin Config CRUD
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/stock/migrate-from-marketplace")
async def migrate_stock_from_marketplace(
    company_id: int = Query(..., description="Company ID that owns the stock (opening balance company)"),
    dry_run: bool = Query(False, description="If true, returns what would be created without writing"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_STOCK_MULTICOMP_001 T02+T03: One-time migration.
    1) Creates StockItemMaster rows from marketplace_spares (288 items).
       - Links via marketplace_sku; merges with any existing stock items.
    2) Inserts OPENING stock_ledger entries for items where available_qty > 0.
    Idempotent: skips items already in stock_item_master; skips items that
    already have an OPENING ledger entry for the given company.
    VGK/EA/Accounts access only.
    """
    try:
        from app.models.marketplace import MarketspareItem
        from app.models.staff_accounts import StockLedger
        from app.services.stock_service import append_stock_ledger
        from decimal import Decimal
        from datetime import date as date_type

        validate_accounts_access(current_user)

        mkt_items = db.query(MarketspareItem).filter(MarketspareItem.is_active == True).all()

        # DC_SYNC_BATCH_001 (Apr 2026): Pre-fetch all lookups in 3 queries instead of N×3
        # This prevents OOM/worker-timeout that killed the Gunicorn worker for 288+ items.
        all_skus = [m.sku for m in mkt_items if m.sku]

        # DC_DEDUP_001: Dual-key duplicate guard —
        # Check both marketplace_sku AND item_code so an item already in stock under
        # any code is never duplicated, even if the marketplace link was removed.
        _existing_by_sku = {
            row.marketplace_sku: row
            for row in db.query(StockItemMaster).filter(
                StockItemMaster.marketplace_sku.in_(all_skus)
            ).all()
        } if all_skus else {}

        # Secondary map: item_code -> StockItemMaster (catches unlinked but code-matching items)
        _existing_by_code = {
            row.item_code.upper(): row
            for row in db.query(StockItemMaster).filter(
                StockItemMaster.item_code.in_(all_skus)
            ).all()
        } if all_skus else {}

        # Set of stock_item_ids that already have at least one image
        _existing_stock_ids = {row.marketplace_sku: row.id for row in _existing_by_sku.values()}
        _item_ids_with_stock = set(_existing_stock_ids.values())
        from app.models.staff_accounts import StockItemImage as _SII_pre
        _items_with_images = set(
            row[0] for row in db.query(_SII_pre.stock_item_id).filter(
                _SII_pre.stock_item_id.in_(_item_ids_with_stock)
            ).distinct().all()
        ) if _item_ids_with_stock else set()

        # Set of item_ids that already have an OPENING ledger entry for this company
        _items_with_opening = set(
            row[0] for row in db.query(StockLedger.item_id).filter(
                StockLedger.company_id == company_id,
                StockLedger.entry_type == 'OPENING',
                StockLedger.item_id.in_(_item_ids_with_stock)
            ).distinct().all()
        ) if _item_ids_with_stock else set()

        created_items = []
        linked_items = []
        skipped_items = []
        opening_created = []
        opening_skipped = []

        today = date_type.today()

        for mkt in mkt_items:
            # ── Step 1: find or create StockItemMaster ──────────────────────────
            # DC_DEDUP_001: Check by marketplace_sku first, then by item_code — never create duplicates
            existing = _existing_by_sku.get(mkt.sku) or _existing_by_code.get((mkt.sku or '').upper())

            if existing:
                stock_item = existing
                # DC_STOCK_MULTICOMP_001: Backfill blank fields from marketplace (preserves manual edits)
                if not dry_run:
                    _updated_fields = []
                    if not existing.brand and getattr(mkt, 'brand', None):
                        existing.brand = mkt.brand
                        _updated_fields.append('brand')
                    if not getattr(existing, 'model_compat', None) and getattr(mkt, 'model_compat', None):
                        existing.model_compat = mkt.model_compat
                        _updated_fields.append('model_compat')
                    if not existing.specification and getattr(mkt, 'specifications', None):
                        existing.specification = mkt.specifications
                        _updated_fields.append('specification')
                    if not existing.colors and getattr(mkt, 'color', None):
                        existing.colors = [mkt.color.strip()]
                        _updated_fields.append('colors')
                    if not existing.description:
                        _desc_parts = [p for p in [
                            getattr(mkt, 'description', None),
                            getattr(mkt, 'speciality', None)
                        ] if p]
                        if _desc_parts:
                            existing.description = ' | '.join(_desc_parts)
                            _updated_fields.append('description')
                    # Backfill category if still default PRODUCT and marketplace has something better
                    _cat_map = {
                        'PRODUCT': 'PRODUCT', 'RAW_MATERIAL': 'RAW_MATERIAL',
                        'CONSUMABLE': 'CONSUMABLE', 'SPARE_PART': 'SPARE_PART', 'ACCESSORY': 'ACCESSORY'
                    }
                    _cat_key = (mkt.category_name or '').upper().replace(' ', '_')
                    _mkt_cat = _cat_map.get(_cat_key)
                    if _mkt_cat and existing.item_category == 'PRODUCT' and _mkt_cat != 'PRODUCT':
                        existing.item_category = _mkt_cat
                        _updated_fields.append('item_category')
                    # Backfill image if none exist
                    _img_url = getattr(mkt, 'image_url', None) or None
                    if _img_url:
                        from app.models.staff_accounts import StockItemImage as _SII
                        _has_img = existing.id in _items_with_images
                        if not _has_img:
                            db.add(_SII(
                                stock_item_id=existing.id,
                                original_path=_img_url,
                                compressed_path=_img_url,
                                thumbnail_path=_img_url,
                                file_name=f"mkt_{mkt.sku}.jpg",
                                mime_type='image/jpeg',
                                is_primary=True,
                                display_order=0,
                                source_type='marketplace',
                                source_url=_img_url,
                                uploaded_by_id=current_user.id,
                            ))
                            _updated_fields.append('image')
                linked_items.append({"id": existing.id, "sku": mkt.sku, "name": mkt.name,
                                     "backfilled": _updated_fields if not dry_run else []})
            else:
                # build a safe item_code from sku (max 30 chars)
                raw_code = (mkt.sku or "").strip()
                item_code = raw_code[:30] if len(raw_code) <= 30 else f"MKT{mkt.id:06d}"

                # check item_code collision (edge case)
                if db.query(StockItemMaster).filter(StockItemMaster.item_code == item_code).first():
                    item_code = f"MKT{mkt.id:06d}"

                if not dry_run:
                    # Map marketplace category to valid StockItemMaster category
                    _cat_map = {
                        'PRODUCT': 'PRODUCT', 'RAW_MATERIAL': 'RAW_MATERIAL',
                        'CONSUMABLE': 'CONSUMABLE', 'SPARE_PART': 'SPARE_PART', 'ACCESSORY': 'ACCESSORY'
                    }
                    _cat_key = (mkt.category_name or 'SPARE_PART').upper().replace(' ', '_')
                    _category = _cat_map.get(_cat_key, 'SPARE_PART')

                    # Build colors list from marketplace color string
                    _colors = [mkt.color.strip()] if getattr(mkt, 'color', None) else None
                    # description = marketplace description + speciality if both exist
                    _desc_parts = [p for p in [
                        getattr(mkt, 'description', None),
                        getattr(mkt, 'speciality', None)
                    ] if p]
                    _desc = ' | '.join(_desc_parts) if _desc_parts else None

                    stock_item = StockItemMaster(
                        item_code=item_code,
                        item_name=(mkt.name or "")[:200],
                        item_category=_category,
                        unit_of_measure='PCS',
                        selling_rate=Decimal(str(mkt.dealer_price or 0)),
                        marketplace_sku=mkt.sku,
                        brand=(getattr(mkt, 'brand', None) or None),
                        model_compat=(getattr(mkt, 'model_compat', None) or None),
                        specification=(getattr(mkt, 'specifications', None) or None),
                        colors=_colors,
                        description=_desc,
                        default_gst_rate=Decimal(str(mkt.gst_percent or 18)),
                        created_by_id=current_user.id,
                    )
                    db.add(stock_item)
                    db.flush()

                    # Create StockItemImage if marketplace has an image URL
                    # No pre-fetch check needed — this is a freshly created item, can't have images yet
                    _img_url = getattr(mkt, 'image_url', None) or None
                    if _img_url:
                        from app.models.staff_accounts import StockItemImage as _SII
                        db.add(_SII(
                                stock_item_id=stock_item.id,
                                original_path=_img_url,
                                compressed_path=_img_url,
                                thumbnail_path=_img_url,
                                file_name=f"mkt_{mkt.sku}.jpg",
                                mime_type='image/jpeg',
                                is_primary=True,
                                display_order=0,
                                source_type='marketplace',
                                source_url=_img_url,
                                uploaded_by_id=current_user.id,
                            ))

                    created_items.append({"id": stock_item.id, "sku": mkt.sku, "name": mkt.name, "item_code": item_code})
                else:
                    skipped_items.append({"sku": mkt.sku, "name": mkt.name, "item_code": item_code, "reason": "dry_run"})
                    continue

            # ── Step 2: opening balance for items with available_qty > 0 ────────
            if (mkt.available_qty or 0) <= 0:
                continue

            already_has_opening = stock_item.id in _items_with_opening

            if already_has_opening:
                opening_skipped.append({"item_id": stock_item.id, "sku": mkt.sku, "reason": "already_exists"})
                continue

            if dry_run:
                opening_skipped.append({"item_id": stock_item.id, "sku": mkt.sku, "reason": "dry_run"})
                continue

            cost = Decimal(str(mkt.proc_ex_tax or mkt.proc_cost or mkt.dealer_price or 0))
            qty = Decimal(str(mkt.available_qty))
            total = (cost * qty).quantize(Decimal("0.01"))

            append_stock_ledger(
                db=db,
                item_id=stock_item.id,
                company_id=company_id,
                entry_type='OPENING',
                quantity_in=qty,
                quantity_out=Decimal('0'),
                unit_rate=cost,
                reference_type='OPENING',
                reference_id=stock_item.id,
                txn_date=today,
                reference_number=f"OB-{stock_item.item_code}",
                narration=f"Opening balance migrated from marketplace_spares (SKU: {mkt.sku})",
                updated_by_id=current_user.id,
            )
            opening_created.append({"item_id": stock_item.id, "sku": mkt.sku, "qty": float(qty), "cost": float(cost)})

        if not dry_run:
            db.commit()

        return JSONResponse(content={
            "success": True,
            "dry_run": dry_run,
            "summary": {
                "marketplace_items_scanned": len(mkt_items),
                "stock_items_created": len(created_items),
                "stock_items_already_linked": len(linked_items),
                "items_backfilled": sum(1 for li in linked_items if li.get('backfilled')),
                "opening_entries_created": len(opening_created),
                "opening_entries_skipped": len(opening_skipped),
            },
            "created_items": created_items,
            "opening_created": opening_created,
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.get("/stock/summary")
async def get_stock_summary(
    company_id: Optional[int] = Query(None, description="Filter by company; omit for all companies"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_STOCK_MULTICOMP_001: Multi-company stock summary.
    Returns per-item balance quantities and values from the stock_ledger.
    """
    try:
        from app.services.stock_service import get_company_stock_summary
        rows = get_company_stock_summary(db, company_id=company_id)
        return JSONResponse(content={"success": True, "data": rows, "total": len(rows)})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock/ledger")
async def get_stock_ledger(
    company_id: Optional[int] = Query(None),
    item_id: Optional[int] = Query(None),
    entry_type: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    to_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_STOCK_MULTICOMP_001: Stock ledger entries with filters.
    Returns paginated ledger rows including inter-company transfer pairs.
    """
    try:
        from datetime import date as date_type
        from sqlalchemy import desc

        query = db.query(StockLedger)
        if company_id:
            query = query.filter(StockLedger.company_id == company_id)
        if item_id:
            query = query.filter(StockLedger.item_id == item_id)
        if entry_type:
            query = query.filter(StockLedger.entry_type == entry_type)
        if from_date:
            query = query.filter(StockLedger.transaction_date >= from_date)
        if to_date:
            query = query.filter(StockLedger.transaction_date <= to_date)

        total = query.count()
        rows = query.order_by(desc(StockLedger.transaction_date), desc(StockLedger.id)).offset(skip).limit(limit).all()

        company_cache: dict = {}
        item_cache: dict = {}

        def _co_name(cid):
            if cid not in company_cache:
                co = db.query(AssociatedCompany).filter(AssociatedCompany.id == cid).first()
                company_cache[cid] = co.company_name if co else str(cid)
            return company_cache[cid]

        def _item_info(iid):
            if iid not in item_cache:
                it = db.query(StockItemMaster).filter(StockItemMaster.id == iid).first()
                item_cache[iid] = {"name": it.item_name if it else str(iid), "code": it.item_code if it else ""}
            return item_cache[iid]

        data = []
        for r in rows:
            ii = _item_info(r.item_id)
            data.append({
                "id": r.id,
                "company_id": r.company_id,
                "company_name": _co_name(r.company_id),
                "item_id": r.item_id,
                "item_name": ii["name"],
                "item_code": ii["code"],
                "transaction_date": r.transaction_date.isoformat() if r.transaction_date else None,
                "entry_type": r.entry_type,
                "reference_type": r.reference_type,
                "reference_id": r.reference_id,
                "reference_number": r.reference_number,
                "quantity_in": float(r.quantity_in or 0),
                "quantity_out": float(r.quantity_out or 0),
                "unit_rate": float(r.unit_rate or 0),
                "total_value": float(r.total_value or 0),
                "balance_qty": float(r.balance_qty or 0),
                "balance_value": float(r.balance_value or 0),
                "narration": r.narration,
                "is_estimate": bool(r.is_estimate) if r.is_estimate is not None else False,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            })

        return JSONResponse(content={"success": True, "data": data, "total": total, "skip": skip, "limit": limit})
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/stock/margin-config")
async def list_margin_configs(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """List all inter-company margin configurations."""
    try:
        from app.models.staff_accounts import InterCompanyMarginConfig
        configs = db.query(InterCompanyMarginConfig).all()
        result = []
        for c in configs:
            result.append({
                "id": c.id,
                "from_company_id": c.from_company_id,
                "to_company_id": c.to_company_id,
                "category_slug": c.category_slug,
                "margin_pct": float(c.margin_pct),
                "is_active": c.is_active,
            })
        return JSONResponse(content={"success": True, "data": result})
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/stock/margin-config")
async def create_margin_config(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Create inter-company margin config. margin_pct in percent (e.g. 6 = 6%)."""
    try:
        from app.models.staff_accounts import InterCompanyMarginConfig
        validate_accounts_access(current_user)
        config = InterCompanyMarginConfig(
            from_company_id=data.get("from_company_id"),
            to_company_id=data.get("to_company_id"),
            category_slug=data.get("category_slug"),
            margin_pct=data.get("margin_pct", 6),
            is_active=data.get("is_active", True),
        )
        db.add(config)
        db.commit()
        db.refresh(config)
        return JSONResponse(content={"success": True, "data": {"id": config.id, "margin_pct": float(config.margin_pct)}})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.put("/stock/margin-config/{config_id}")
async def update_margin_config(
    config_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Update inter-company margin config."""
    try:
        from app.models.staff_accounts import InterCompanyMarginConfig
        validate_accounts_access(current_user)
        config = db.query(InterCompanyMarginConfig).filter(InterCompanyMarginConfig.id == config_id).first()
        if not config:
            return JSONResponse(status_code=404, content={"success": False, "message": "Config not found"})
        for field in ("from_company_id", "to_company_id", "category_slug", "margin_pct", "is_active"):
            if field in data:
                setattr(config, field, data[field])
        db.commit()
        return JSONResponse(content={"success": True, "message": "Updated"})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.delete("/stock/margin-config/{config_id}")
async def delete_margin_config(
    config_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Delete inter-company margin config."""
    try:
        from app.models.staff_accounts import InterCompanyMarginConfig
        validate_accounts_access(current_user)
        config = db.query(InterCompanyMarginConfig).filter(InterCompanyMarginConfig.id == config_id).first()
        if not config:
            return JSONResponse(status_code=404, content={"success": False, "message": "Config not found"})
        db.delete(config)
        db.commit()
        return JSONResponse(content={"success": True, "message": "Deleted"})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


# ─────────────────────────────────────────────────────────────────────────────
# DC_STOCK_ESTIMATE_001: Estimation stock consumption (soft entries)
# Estimate rows are stored in stock_ledger with is_estimate=True.
# They are fully visible in the ledger but excluded from balance calculations.
# Confirm → converts to real movement. Cancel → deletes the soft row (no impact).
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/stock/estimate-consumption")
async def create_estimate_consumption(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Record a SOFT stock deduction for an estimate/quotation.
    This is visible in the ledger (tagged ESTIMATE) but does NOT reduce real balance.
    Required: item_id, company_id, quantity_out, unit_rate
    Optional: reference_number (estimate/job number), narration, specification, color
    """
    try:
        from app.services.stock_service import append_stock_ledger
        from decimal import Decimal

        item_id    = data.get("item_id")
        company_id = data.get("company_id")
        qty_out    = data.get("quantity_out") or data.get("quantity")
        unit_rate  = data.get("unit_rate", 0)

        if not all([item_id, company_id, qty_out]):
            return JSONResponse(status_code=400, content={
                "success": False,
                "message": "item_id, company_id, and quantity_out are required"
            })

        entry = append_stock_ledger(
            db=db,
            item_id=int(item_id),
            company_id=int(company_id),
            entry_type="SERVICE_CONSUMPTION",
            reference_type="SERVICE",
            reference_id=0,                              # no hard FK for estimates
            reference_number=data.get("reference_number"),
            quantity_in=Decimal("0"),
            quantity_out=Decimal(str(qty_out)),
            unit_rate=Decimal(str(unit_rate)),
            narration=data.get("narration") or f"Estimate consumption — {data.get('reference_number', 'draft')}",
            txn_date=None,
            updated_by_id=current_user.id,
            specification=data.get("specification"),
            color=data.get("color"),
            is_estimate=True,
        )
        db.commit()
        return JSONResponse(content={
            "success": True,
            "message": "Estimate stock entry created (soft — does not affect real balance)",
            "data": {
                "id": entry.id,
                "item_id": entry.item_id,
                "company_id": entry.company_id,
                "quantity_out": float(entry.quantity_out),
                "unit_rate": float(entry.unit_rate),
                "is_estimate": True,
                "balance_qty": float(entry.balance_qty),
            }
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.post("/stock/estimate-consumption/{entry_id}/confirm")
async def confirm_estimate_consumption(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Confirm an estimate stock entry — converts it to a real SERVICE_CONSUMPTION.
    Deletes the soft row and appends a new real entry (balance chain is recalculated).
    The new real entry's ID is returned.
    """
    try:
        from app.services.stock_service import confirm_estimate_ledger_entry
        new_entry = confirm_estimate_ledger_entry(
            db=db,
            entry_id=entry_id,
            updated_by_id=current_user.id,
        )
        db.commit()
        return JSONResponse(content={
            "success": True,
            "message": "Estimate confirmed — stock deducted from real balance",
            "data": {
                "new_entry_id": new_entry.id,
                "item_id": new_entry.item_id,
                "company_id": new_entry.company_id,
                "quantity_out": float(new_entry.quantity_out),
                "balance_qty": float(new_entry.balance_qty),
                "is_estimate": False,
            }
        })
    except ValueError as ve:
        return JSONResponse(status_code=404, content={"success": False, "message": str(ve)})
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.delete("/stock/estimate-consumption/{entry_id}")
async def cancel_estimate_consumption(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    Cancel (delete) a pending estimate stock entry.
    Since estimate entries do not affect the real balance, deletion is safe and
    requires no reversal entry — the balance chain is unaffected.
    """
    try:
        row = db.execute(
            text("SELECT id, item_id, company_id, quantity_out, reference_number FROM stock_ledger WHERE id = :eid AND is_estimate = TRUE"),
            {"eid": entry_id}
        ).fetchone()
        if not row:
            return JSONResponse(status_code=404, content={
                "success": False,
                "message": f"No pending estimate entry found with id={entry_id}"
            })
        db.execute(text("DELETE FROM stock_ledger WHERE id = :eid"), {"eid": entry_id})
        db.commit()
        return JSONResponse(content={
            "success": True,
            "message": f"Estimate entry {entry_id} cancelled — no balance impact",
            "data": {
                "deleted_id": entry_id,
                "item_id": row.item_id,
                "reference_number": row.reference_number,
            }
        })
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


# ==================== CONSOLIDATED REPORTS (DC_CONSOLIDATED_001) ====================

@router.get("/consolidated/sales")
async def get_consolidated_sales(
    company_id: int = Query(0, description="Company ID — 0 for all companies"),
    period: str = Query("THIS_FY", description="THIS_MONTH|THIS_QUARTER|THIS_FY|OVERALL|CUSTOM"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    drill_month: Optional[str] = Query(None, description="YYYY-MM for monthly drill-down"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Consolidated Sales — monthly summary + drill-down. DC_CONSOLIDATED_001"""
    from app.services.staff_accounts_service import get_consolidated_sales_report
    try:
        from datetime import datetime
        df = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        dt = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        data = get_consolidated_sales_report(db, current_user, company_id, period, df, dt, drill_month)
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/consolidated/purchases")
async def get_consolidated_purchases(
    company_id: int = Query(0, description="Company ID — 0 for all companies"),
    period: str = Query("THIS_FY", description="THIS_MONTH|THIS_QUARTER|THIS_FY|OVERALL|CUSTOM"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    drill_month: Optional[str] = Query(None, description="YYYY-MM for monthly drill-down"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Consolidated Purchases — monthly summary + drill-down. DC_CONSOLIDATED_001"""
    from app.services.staff_accounts_service import get_consolidated_purchases_report
    try:
        from datetime import datetime
        df = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        dt = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        data = get_consolidated_purchases_report(db, current_user, company_id, period, df, dt, drill_month)
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/consolidated/pl")
async def get_consolidated_pl(
    company_id: int = Query(0, description="Company ID — 0 for all companies"),
    period: str = Query("THIS_FY", description="THIS_MONTH|THIS_QUARTER|THIS_FY|OVERALL|CUSTOM"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    as_on_date: Optional[str] = Query(None, description="Cutoff date — clamps period end to this date"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Consolidated Profit & Loss — Tally-style trading + P&L. DC_CONSOLIDATED_001"""
    from app.services.staff_accounts_service import get_consolidated_pl_report
    try:
        from datetime import datetime
        df = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        dt = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        aod = datetime.strptime(as_on_date, '%Y-%m-%d').date() if as_on_date else None
        data = get_consolidated_pl_report(db, current_user, company_id, period, df, dt, aod)
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/dar")
async def get_dar(
    company_id: int = Query(0, description="Company ID — 0 for all companies"),
    period: str = Query("TODAY", description="TODAY|YESTERDAY|THIS_MONTH|THIS_QUARTER|THIS_FY|OVERALL|CUSTOM"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    mode: str = Query("daily", description="daily | comparison"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DAR — Daily Activity Report. DC_DAR_001 (Task #50, May 2026).
    Returns a 10-business-day snapshot grid for the requested company and
    period end (ending at the period end date, going back 10 calendar weekdays).
    Read-only. RBAC via is_accounts_allowed_employee in the service layer.
    """
    from app.services.staff_accounts_service import (
        get_dar_report, validate_accounts_read_access, log_accounts_audit,
    )
    try:
        validate_accounts_read_access(current_user)
        from datetime import datetime
        df = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        dt = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        data = get_dar_report(db, current_user, company_id, period, df, dt, mode=mode)
        try:
            # Read-only view → audit description only; no old/new value
            # payload (DAR contract requires no value snapshots for views).
            bdates = data.get('business_dates') or []
            bdates_summary = (f"{bdates[0]}..{bdates[-1]} ({len(bdates)} days)"
                              if bdates else "no dates")
            log_accounts_audit(
                db=db,
                employee_id=current_user.id,
                action="VIEW_DAR",
                entity_type="DAR",
                entity_id=company_id or 0,
                description=(
                    f"Viewed DAR view_type=DAR period={period} "
                    f"company_id={company_id} end={data.get('end_date')} "
                    f"business_dates={bdates_summary}"
                ),
            )
        except Exception:
            pass
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/consolidated/balance-sheet")
async def get_consolidated_balance_sheet(
    company_id: int = Query(0, description="Company ID — 0 for all companies"),
    as_on_date: Optional[str] = Query(None, description="As on date YYYY-MM-DD (default today)"),
    period: Optional[str] = Query(None, description="THIS_MONTH|THIS_QUARTER|THIS_FY|OVERALL|CUSTOM — auto-derives as_on_date"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Consolidated Balance Sheet — Tally-style two-column format. DC_CONSOLIDATED_001"""
    from app.services.staff_accounts_service import get_consolidated_balance_sheet_report, _consolidated_period_dates
    try:
        from datetime import datetime, date as _d
        aod = datetime.strptime(as_on_date, '%Y-%m-%d').date() if as_on_date else None
        # If period given (and not Custom) and as_on_date not explicitly set, derive as_on from period end
        if period and period != 'CUSTOM' and not aod:
            _, period_end, _ = _consolidated_period_dates(period, None, None)
            today = _d.today()
            aod = period_end if period_end <= today else today
        data = get_consolidated_balance_sheet_report(db, current_user, company_id, aod, period)
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ---------------------------------------------------------------------------
# DC_NARR_001 — Invoice Narration Log (Sales + Purchase)
# ---------------------------------------------------------------------------

@router.get("/invoice-narration-log")
async def list_invoice_narration_log(
    invoice_type: str = Query(..., regex="^(SALES|PURCHASE)$"),
    invoice_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Return all narration log entries for a given invoice (oldest first)."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Not authorized")
    from app.models.staff_accounts import InvoiceNarrationLog
    entries = (
        db.query(InvoiceNarrationLog)
        .filter(
            InvoiceNarrationLog.invoice_type == invoice_type,
            InvoiceNarrationLog.invoice_id == invoice_id
        )
        .order_by(InvoiceNarrationLog.created_at.asc())
        .all()
    )
    return {"success": True, "entries": [e.to_dict() for e in entries]}


@router.post("/invoice-narration-log")
async def add_invoice_narration_log(
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """Append a narration entry for a Sales or Purchase invoice."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Not authorized")

    invoice_type = (body.get("invoice_type") or "").upper()
    invoice_id = body.get("invoice_id")
    company_id = body.get("company_id")
    narration = (body.get("narration") or "").strip()

    if invoice_type not in ("SALES", "PURCHASE"):
        raise HTTPException(status_code=400, detail="invoice_type must be SALES or PURCHASE")
    if not invoice_id:
        raise HTTPException(status_code=400, detail="invoice_id is required")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if not narration:
        raise HTTPException(status_code=400, detail="narration text is required")

    from app.models.staff_accounts import InvoiceNarrationLog
    entry = InvoiceNarrationLog(
        invoice_type=invoice_type,
        invoice_id=invoice_id,
        company_id=company_id,
        narration=narration,
        created_by_id=current_user.id,
        created_by_name=getattr(current_user, 'full_name', None) or getattr(current_user, 'name', None) or str(current_user.id),
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"success": True, "entry": entry.to_dict()}


# ─────────────────────────────────────────────────────────────────────────────
# DC-CONSOL-SPARE-001: Consolidated Spare Parts Procurement Workbench Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/procurement/marketplace-items")
async def get_marketplace_pending_items(
    company_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-002: Marketplace tab — items from partner orders + service tickets
    that are pending procurement AND have stock shortage (stock < ordered/required qty).
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        items = SpareProcurementService.get_marketplace_pending_items(
            db, current_user, company_id=company_id, search=search
        )
        return {"success": True, "items": items, "total": len(items)}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-items")
async def get_spare_items_consolidated(
    company_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=300),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: All SPARE_PART items with stock + 4-column demand breakdown.
    Demand sources: Manufacturing requirements, Low stock requirements,
    Partner spare requests, Open VGK partner order lines.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        skip = (page - 1) * limit
        items, total = SpareProcurementService.get_spare_items_consolidated(
            db, current_user,
            company_id=company_id, search=search, skip=skip, limit=limit
        )
        return {
            "success": True, "items": items, "total": total,
            "page": page, "pages": (total + limit - 1) // limit
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-items/{item_id}/vendors")
async def get_spare_item_vendors(
    item_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Vendors who supplied this spare item (purchase history)
    + catalog-linked vendors. Returns last rate, last date, contact info.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        vendors = SpareProcurementService.get_item_vendors(db, current_user, item_id)
        return {"success": True, "vendors": vendors}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-vendors/{vendor_id}/items")
async def get_vendor_spare_items(
    vendor_id: int = Path(..., ge=1),
    company_id: Optional[int] = Query(None),
    exclude_ids: Optional[str] = Query(None, description="Comma-separated item IDs to exclude"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Other SPARE_PART items this vendor supplies
    that have reorder need or any demand. Used by vendor panel "Other Items" section.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        exc = [int(x) for x in (exclude_ids or '').split(',') if x.strip().isdigit()]
        items = SpareProcurementService.get_vendor_spare_items(
            db, current_user, vendor_id, company_id=company_id, exclude_item_ids=exc
        )
        return {"success": True, "items": items}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/spare-orders")
async def create_spare_order(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Create or update a DRAFT spare purchase order.
    Body: {company_id, lines:[{vendor_id,item_id,quantity,uom,last_rate,demand_source,demand_qty,notes}], notes?, order_id?}
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        order = SpareProcurementService.save_spare_order(
            db, current_user,
            company_id=data.get('company_id'),
            lines=data.get('lines', []),
            notes=data.get('notes'),
            order_id=data.get('order_id'),
        )
        return {"success": True, "message": f"Order {order.order_number} saved", "order": order.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-orders")
async def list_spare_orders(
    company_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-CONSOL-SPARE-001: List spare purchase orders with filters."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        skip = (page - 1) * limit
        orders, total = SpareProcurementService.list_spare_orders(
            db, current_user, company_id=company_id, status=status, skip=skip, limit=limit
        )
        return {
            "success": True,
            "orders": [o.to_dict(include_lines=False) for o in orders],
            "total": total, "page": page
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-orders/{order_id}")
async def get_spare_order(
    order_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-CONSOL-SPARE-001: Get single spare order with all lines."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        order = SpareProcurementService.get_spare_order(db, current_user, order_id)
        return {"success": True, "order": order.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/procurement/spare-orders/{order_id}/submit")
async def submit_spare_order(
    order_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-CONSOL-SPARE-001: Submit DRAFT → WAITING_APPROVAL."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        order = SpareProcurementService.submit_spare_order(db, current_user, order_id)
        return {"success": True, "message": f"Order {order.order_number} submitted for approval", "order": order.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/procurement/spare-orders/{order_id}/approve")
async def approve_spare_order(
    order_id: int = Path(..., ge=1),
    data: dict = Body(default={}),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Approve order (EA/Accounts/VGK only).
    Creates procurement_request records per vendor — visible in Tab 1.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService, AccountsRBACError
        order = SpareProcurementService.approve_spare_order(
            db, current_user, order_id,
            approval_notes=data.get('approval_notes')
        )
        return {
            "success": True,
            "message": f"Order {order.order_number} approved. {len(order.procurement_req_ids)} procurement request(s) created.",
            "order": order.to_dict(),
            "procurement_req_ids": order.procurement_req_ids,
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/procurement/spare-orders/{order_id}/review-update")
async def review_update_spare_order(
    order_id: int = Path(..., ge=1),
    data: dict = Body(default={}),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Approver updates lines on WAITING_APPROVAL order before final approval.
    Accepts: { lines: [{id, vendor_id, quantity, last_rate}], deleted_line_ids: [int] }
    Status stays WAITING_APPROVAL — no state change.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService, AccountsRBACError
        order = SpareProcurementService.review_update_spare_order(
            db, current_user, order_id,
            lines=data.get('lines', []),
            deleted_line_ids=data.get('deleted_line_ids', []),
        )
        return {"success": True, "order": order.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/procurement/spare-orders/{order_id}/cancel")
async def cancel_spare_order(
    order_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-CONSOL-SPARE-001: Cancel DRAFT or WAITING_APPROVAL order."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        order = SpareProcurementService.cancel_spare_order(db, current_user, order_id)
        return {"success": True, "message": f"Order {order.order_number} cancelled", "order": order.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-orders/{order_id}/receipt/{vendor_id}")
async def download_spare_po_receipt(
    order_id: int = Path(..., ge=1),
    vendor_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Download PO PDF receipt for one vendor from an order.
    Works for APPROVED or WAITING_APPROVAL orders.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        from app.services.spare_po_pdf import generate_spare_po_pdf
        from app.models.staff_accounts import SparePurchaseOrder, VendorMaster, AssociatedCompany
        from fastapi.responses import Response as FastAPIResponse

        order = SpareProcurementService.get_spare_order(db, current_user, order_id)
        vendor = db.query(VendorMaster).filter(VendorMaster.id == vendor_id).first()
        company = db.query(AssociatedCompany).filter(AssociatedCompany.id == order.company_id).first()
        lines = [l for l in order.lines if l.vendor_id == vendor_id]

        if not lines:
            raise HTTPException(status_code=404, detail="No items for this vendor in the order")

        items_payload = [
            {
                'item_code':      l.item_code or '',
                'item_name':      l.item_name or '',
                'qty':            float(l.quantity),
                'uom':            l.unit_of_measure,
                'last_rate':      float(l.last_purchase_rate) if l.last_purchase_rate else 0,
                'estimated_value': float(l.quantity * l.last_purchase_rate) if l.last_purchase_rate else 0,
                'demand_source':  l.demand_source or '',
            }
            for l in lines
        ]

        from datetime import datetime
        pdf_bytes = generate_spare_po_pdf(
            order_number=order.order_number,
            order_date=datetime.now().strftime('%d %b %Y'),
            company_name=company.company_name if company else 'MyntReal LLP',
            company_address=getattr(company, 'address', '') or '',
            company_gstin=getattr(company, 'gst_number', '') or '',
            vendor_name=vendor.vendor_name if vendor else f'Vendor #{vendor_id}',
            vendor_address=vendor.address if vendor else '',
            vendor_city=vendor.city if vendor else '',
            vendor_state=vendor.state if vendor else '',
            vendor_gstin=vendor.gst_number if vendor else '',
            vendor_phone=vendor.phone if vendor else '',
            vendor_contact_person=vendor.contact_person if vendor else '',
            items=items_payload,
            notes=order.notes,
            created_by=order.created_by.full_name if order.created_by else None,
            approved_by=order.approved_by.full_name if order.approved_by else None,
        )

        filename = f"PO_{order.order_number}_{vendor_id}.pdf"
        return FastAPIResponse(
            content=pdf_bytes,
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/procurement/spare-orders/{order_id}/send-whatsapp/{vendor_id}")
async def send_spare_po_whatsapp(
    order_id: int = Path(..., ge=1),
    vendor_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Send PO receipt to vendor via WhatsApp.
    Uses send_auto_whatsapp with document template.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        from app.models.staff_accounts import VendorMaster
        order = SpareProcurementService.get_spare_order(db, current_user, order_id)
        vendor = db.query(VendorMaster).filter(VendorMaster.id == vendor_id).first()
        if not vendor or not vendor.phone:
            raise HTTPException(status_code=400, detail="Vendor has no phone number")

        lines = [l for l in order.lines if l.vendor_id == vendor_id]
        if not lines:
            raise HTTPException(status_code=404, detail="No items for this vendor in the order")

        item_summary = ", ".join(f"{l.item_name or l.item_code} x{float(l.quantity):.0f}" for l in lines[:3])
        if len(lines) > 3:
            item_summary += f" + {len(lines) - 3} more"

        try:
            from app.services.whatsapp_auto_service import send_auto_whatsapp
            result = send_auto_whatsapp(
                db=db,
                phone=vendor.phone,
                template_slug='spare_po_sent',
                params={
                    'vendor_name': vendor.vendor_name,
                    'po_number': order.order_number,
                    'item_summary': item_summary,
                    'company_name': 'MyntReal LLP',
                },
                created_by_id=current_user.id,
            )
            return {"success": True, "message": f"WhatsApp sent to {vendor.phone}", "result": str(result)}
        except Exception as wa_err:
            return {"success": False, "message": f"WA send failed: {wa_err}", "phone": vendor.phone}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/spare-orders/{order_id}/purchase-prefill/{vendor_id}")
async def spare_order_purchase_prefill(
    order_id: int = Path(..., ge=1),
    vendor_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-CONSOL-SPARE-001: Returns VendorTransactionCreate-compatible prefill payload
    for 'Convert to Purchase' from an approved spare order. Frontend stores in
    sessionStorage and vendor-transactions page reads it to pre-fill the create form.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        prefill = SpareProcurementService.get_purchase_prefill(db, current_user, order_id, vendor_id)
        return {"success": True, "prefill": prefill}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/procurement/partner-spare-requests")
async def list_partner_spare_requests(
    company_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-PARTNER-SPARE-001: Staff view of all partner spare requests."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        skip = (page - 1) * limit
        reqs, total = SpareProcurementService.staff_list_partner_spare_requests(
            db, current_user, company_id=company_id, status=status, skip=skip, limit=limit
        )
        return {"success": True, "requests": [r.to_dict() for r in reqs], "total": total, "page": page}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.put("/procurement/partner-spare-requests/{request_id}/acknowledge")
async def acknowledge_partner_spare_request(
    request_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-PARTNER-SPARE-001: Staff acknowledges a partner spare request."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        req = SpareProcurementService.staff_acknowledge_partner_request(db, current_user, request_id)
        return {"success": True, "message": "Request acknowledged", "request": req.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


# ─────────────────────────────────────────────────────────────────────────────
# DC_TRAINING_VIDEOS_001: Training Videos System
# ─────────────────────────────────────────────────────────────────────────────
from datetime import datetime as _dt_tv, timedelta as _td_tv
import pytz as _pytz_tv
from sqlalchemy import text as _tv_text

_TRAINING_EXEMPT_CODES = {'MR10001'}   # VGK4U mentor — only exemption, by design


def _tv_ist_now():
    return _dt_tv.now(_pytz_tv.timezone('Asia/Kolkata')).replace(tzinfo=None)


def _tv_gate_deployed_at(db):
    """Return deployment datetime from dc_migrations, or None if not set."""
    row = db.execute(
        _tv_text("SELECT created_at FROM dc_migrations WHERE key = 'training_gate_deployed_at'")
    ).fetchone()
    return row[0] if row else None


@router.get("/training/videos")
def get_training_videos(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_TRAINING_VIDEOS_001: All active training videos with my completion status."""
    try:
        rows = db.execute(_tv_text("""
            SELECT v.id, v.order_num, v.title, v.youtube_url, v.youtube_video_id, v.is_short,
                   COALESCE(p.is_completed, false) AS is_completed,
                   p.completed_at
              FROM training_videos v
              LEFT JOIN training_video_progress p
                     ON p.video_id = v.id AND p.employee_id = :emp_id
             WHERE v.is_active = true
             ORDER BY v.order_num
        """), {"emp_id": current_user.id}).fetchall()

        videos = [{
            "id":               r.id,
            "order_num":        r.order_num,
            "title":            r.title,
            "youtube_url":      r.youtube_url,
            "youtube_video_id": r.youtube_video_id,
            "is_short":         r.is_short,
            "is_completed":     r.is_completed,
            "completed_at":     r.completed_at.isoformat() if r.completed_at else None,
        } for r in rows]

        total   = len(videos)
        done    = sum(1 for v in videos if v["is_completed"])
        percent = round(done / total * 100) if total else 0
        return {"success": True, "videos": videos,
                "total": total, "completed": done, "percent_done": percent}
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/training/videos/{video_id}/complete")
def mark_training_video_complete(
    video_id: int = Path(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_TRAINING_VIDEOS_001: Mark a training video complete (called by YouTube player ENDED event)."""
    try:
        video = db.execute(
            _tv_text("SELECT id, title FROM training_videos WHERE id = :id AND is_active = true"),
            {"id": video_id}
        ).fetchone()
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")

        now = _tv_ist_now()
        db.execute(_tv_text("""
            INSERT INTO training_video_progress
                (employee_id, video_id, is_completed, completed_at, created_at)
            VALUES (:emp, :vid, true, :now, :now)
            ON CONFLICT (employee_id, video_id)
            DO UPDATE SET is_completed = true,
                          completed_at = CASE WHEN training_video_progress.is_completed
                                              THEN training_video_progress.completed_at
                                              ELSE EXCLUDED.completed_at END
        """), {"emp": current_user.id, "vid": video_id, "now": now})
        db.commit()
        return {"success": True, "message": f"'{video.title}' marked as complete"}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/training/status")
def get_training_status(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC_TRAINING_VIDEOS_001: Is this employee currently gated behind training?
    Called by staff_sidebar.js to decide whether to hard-lock the menu.
    """
    try:
        # DC_FIX: use emp_code (correct column), not employee_code (nonexistent)
        is_exempt = getattr(current_user, 'emp_code', None) in _TRAINING_EXEMPT_CODES

        # Grace period: all employees get 5 days from whichever is LATER —
        # the gate deployment date OR their own hire date.
        deployed_at = _tv_gate_deployed_at(db)
        grace_active = False
        grace_until  = None
        if deployed_at and not is_exempt:
            emp_created = getattr(current_user, 'created_at', None)
            # Anchor = later of deployment date and employee creation date
            anchor = max(deployed_at, emp_created) if emp_created else deployed_at
            grace_until  = anchor + _td_tv(days=5)
            # Use IST naive now (consistent with rest of system)
            grace_active = _tv_ist_now() < grace_until

        # Completion count
        counts = db.execute(_tv_text("""
            SELECT COUNT(v.id)                                          AS total,
                   COUNT(p.id) FILTER (WHERE p.is_completed = true)     AS done
              FROM training_videos v
              LEFT JOIN training_video_progress p
                     ON p.video_id = v.id AND p.employee_id = :emp_id
             WHERE v.is_active = true
        """), {"emp_id": current_user.id}).fetchone()

        total    = int(counts.total or 0)
        done     = int(counts.done  or 0)
        percent  = round(done / total * 100) if total else 0
        all_done = total > 0 and done >= total

        is_gated = not is_exempt and not grace_active and not all_done

        pending = []
        if not all_done:
            rows = db.execute(_tv_text("""
                SELECT v.order_num, v.title
                  FROM training_videos v
                  LEFT JOIN training_video_progress p
                         ON p.video_id = v.id AND p.employee_id = :emp_id
                 WHERE v.is_active = true
                   AND (p.is_completed IS NULL OR p.is_completed = false)
                 ORDER BY v.order_num
            """), {"emp_id": current_user.id}).fetchall()
            pending = [{"order_num": r.order_num, "title": r.title} for r in rows]

        return {
            "success":         True,
            "is_gated":        is_gated,
            "is_exempt":       is_exempt,
            "grace_active":    grace_active,
            "grace_until":     grace_until.isoformat() if grace_until else None,
            "is_complete":     all_done,
            "completed_count": done,
            "total_count":     total,
            "percent_done":    percent,
            "pending_videos":  pending,
        }
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/training/admin")
def get_training_admin_overview(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_TRAINING_VIDEOS_001: Admin — all employees × all videos completion matrix."""
    try:
        if getattr(current_user, 'hierarchy_level', 0) < 80:
            raise HTTPException(status_code=403, detail="Admin access required (hierarchy >= 80)")

        total_videos = int(db.execute(
            _tv_text("SELECT COUNT(*) FROM training_videos WHERE is_active = true")
        ).scalar() or 0)

        deployed_at = _tv_gate_deployed_at(db)

        rows = db.execute(_tv_text("""
            SELECT e.id, e.employee_code, e.full_name, e.staff_type, e.hierarchy_level,
                   e.created_at AS emp_created_at,
                   COUNT(p.id) FILTER (WHERE p.is_completed = true) AS completed,
                   MAX(p.completed_at)                               AS last_completed_at
              FROM staff_employees e
              LEFT JOIN training_video_progress p ON p.employee_id = e.id
             WHERE e.is_active = true
             GROUP BY e.id, e.employee_code, e.full_name, e.staff_type,
                      e.hierarchy_level, e.created_at
             ORDER BY e.full_name
        """)).fetchall()

        employees = []
        for r in rows:
            done      = int(r.completed or 0)
            pct       = round(done / total_videos * 100) if total_videos else 0
            is_exempt = r.employee_code in _TRAINING_EXEMPT_CODES
            grace_ok  = False
            if deployed_at and not is_exempt and r.emp_created_at:
                grace_ok = _dt_tv.utcnow() < deployed_at + _td_tv(days=5)
            employees.append({
                "id":              r.id,
                "employee_code":   r.employee_code,
                "full_name":       r.full_name,
                "staff_type":      r.staff_type,
                "hierarchy_level": r.hierarchy_level,
                "is_exempt":       is_exempt,
                "grace_active":    grace_ok,
                "completed":       done,
                "total":           total_videos,
                "percent_done":    pct,
                "is_complete":     done >= total_videos and total_videos > 0,
                "last_completed_at": r.last_completed_at.isoformat() if r.last_completed_at else None,
            })

        done_count = sum(1 for e in employees if e["is_complete"] or e["is_exempt"])
        return {
            "success":        True,
            "employees":      employees,
            "total_videos":   total_videos,
            "total_staff":    len(employees),
            "fully_complete": done_count,
            "deployed_at":    deployed_at.isoformat() if deployed_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.get("/training/admin/export.csv")
def export_training_progress_csv(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_TRAINING_VIDEOS_001: CSV download of training completion matrix."""
    from fastapi.responses import Response
    import csv, io
    try:
        if getattr(current_user, 'hierarchy_level', 0) < 80:
            raise HTTPException(status_code=403, detail="Admin access required (hierarchy >= 80)")

        videos = db.execute(_tv_text("""
            SELECT id, order_num, title
              FROM training_videos WHERE is_active = true ORDER BY order_num
        """)).fetchall()

        progress = db.execute(_tv_text("""
            SELECT p.employee_id, p.video_id, p.is_completed, p.completed_at
              FROM training_video_progress p
              JOIN training_videos v ON v.id = p.video_id AND v.is_active = true
        """)).fetchall()
        prog_map = {(p.employee_id, p.video_id): (p.is_completed, p.completed_at)
                    for p in progress}

        employees = db.execute(_tv_text("""
            SELECT id, employee_code, full_name, staff_type
              FROM staff_employees WHERE is_active = true ORDER BY full_name
        """)).fetchall()

        buf    = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(
            ["Employee Code", "Full Name", "Staff Type"] +
            [f"{v.order_num}. {v.title}" for v in videos] +
            ["Total Done", "% Complete"]
        )
        for emp in employees:
            row  = [emp.employee_code, emp.full_name, emp.staff_type]
            done = 0
            for v in videos:
                completed, completed_at = prog_map.get((emp.id, v.id), (False, None))
                if completed:
                    row.append(completed_at.strftime('%d/%m/%Y') if completed_at else '✓')
                    done += 1
                else:
                    row.append('')
            pct = round(done / len(videos) * 100) if videos else 0
            row += [done, f"{pct}%"]
            writer.writerow(row)

        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=training_progress.csv"}
        )
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/training/sync")
def trigger_training_sync(
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_TRAINING_VIDEOS_001: Admin — manually re-sync training videos from Google Doc."""
    try:
        if getattr(current_user, 'hierarchy_level', 0) < 80:
            raise HTTPException(status_code=403, detail="Admin access required (hierarchy >= 80)")
        from app.services.training_sync import sync_training_videos_from_gdoc
        result = sync_training_videos_from_gdoc(db)
        return {"success": True, "message": "Sync complete", "result": result}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.post("/purchase-invoices/{invoice_id}/create-intake-batch")
async def create_intake_batch_from_pending(
    invoice_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-001: Create a purchase intake batch for pending line items.
    Triggers the existing intake/QC flow.
    data = { items: [{ line_id, qty_to_receive }] }
    """
    try:
        from app.models.staff_accounts import (
            PurchaseInvoiceUpload, PurchaseInvoiceLineItem,
            PurchaseIntakeBatch, PurchaseIntakeItem
        )
        from app.services.counter_service import get_next_counter

        inv = db.query(PurchaseInvoiceUpload).filter(
            PurchaseInvoiceUpload.id == invoice_id
        ).first()
        if not inv:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if inv.status != 'CONFIRMED':
            raise HTTPException(status_code=400, detail="Only CONFIRMED invoices can create intake batches")

        items_input = data.get('items', [])
        if not items_input:
            raise HTTPException(status_code=400, detail="No items provided")

        batch_number = get_next_counter(db, 'INTAKE', inv.company_id)

        total_ordered = 0.0
        total_ordered_val = 0.0
        intake_items = []

        for item_in in items_input:
            line_id = item_in.get('line_id')
            qty = float(item_in.get('qty_to_receive', 0))
            if qty <= 0:
                continue
            ln = db.query(PurchaseInvoiceLineItem).filter(
                PurchaseInvoiceLineItem.id == line_id,
                PurchaseInvoiceLineItem.upload_id == invoice_id
            ).first()
            if not ln:
                continue
            val = round(qty * float(ln.unit_rate or 0), 2)
            total_ordered += qty
            total_ordered_val += val
            intake_items.append((ln, qty, val))

        if not intake_items:
            raise HTTPException(status_code=400, detail="No valid items with qty > 0")

        batch = PurchaseIntakeBatch(
            batch_number=batch_number,
            company_id=inv.company_id,
            purchase_invoice_id=inv.id,
            vendor_id=None,
            total_ordered_qty=total_ordered,
            total_received_qty=0,
            total_accepted_qty=0,
            total_rejected_qty=0,
            total_pending_qty=total_ordered,
            total_ordered_value=total_ordered_val,
            total_received_value=0,
            total_accepted_value=0,
            total_rejected_value=0,
            intake_status='PENDING_RECEIPT',
            created_by_id=current_user.id,
            updated_by_id=current_user.id,
        )
        db.add(batch)
        db.flush()

        for ln, qty, val in intake_items:
            pii = PurchaseIntakeItem(
                batch_id=batch.id,
                purchase_line_id=ln.id,
                item_id=ln.item_id,
                item_code=ln.item_code,
                item_name=ln.item_description,
                unit_of_measure=ln.unit_of_measure,
                unit_rate=float(ln.unit_rate or 0),
                ordered_qty=qty,
                received_qty=0,
                accepted_qty=0,
                rejected_qty=0,
                ordered_value=val,
                received_value=0,
                accepted_value=0,
                rejected_value=0,
                qc_status='PENDING',
            )
            db.add(pii)

        db.commit()
        return JSONResponse(content={
            "success": True,
            "message": f"Intake batch {batch_number} created with {len(intake_items)} items",
            "batch_id": batch.id,
            "batch_number": batch_number
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.patch("/sales-invoices/{invoice_id}/pending-line-config/{line_id}")
async def update_sales_pending_line_config(
    invoice_id: int,
    line_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-DISPATCH-EXTRA-001: Update pending_qty override for a regular sales invoice line."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesInvoice, SalesInvoiceLineItem, SalesPendingLineConfig
        new_pqty = float(body.get('pending_qty', 0) if body.get('pending_qty') is not None else 0)
        if new_pqty < 0:
            raise HTTPException(status_code=400, detail="pending_qty cannot be negative")
        ln = db.query(SalesInvoiceLineItem).filter(
            SalesInvoiceLineItem.id == line_id,
            SalesInvoiceLineItem.invoice_id == invoice_id
        ).first()
        if not ln:
            raise HTTPException(status_code=404, detail="Line not found")
        if new_pqty > float(ln.quantity or 0):
            raise HTTPException(status_code=400, detail=f"pending_qty cannot exceed invoiced qty ({ln.quantity})")
        cfg = db.query(SalesPendingLineConfig).filter(
            SalesPendingLineConfig.invoice_id == invoice_id,
            SalesPendingLineConfig.invoice_line_id == line_id
        ).first()
        if cfg:
            cfg.pending_qty = new_pqty
        else:
            inv = db.query(SalesInvoice).filter(SalesInvoice.id == invoice_id).first()
            if not inv:
                raise HTTPException(status_code=404, detail="Invoice not found")
            db.add(SalesPendingLineConfig(
                company_id=inv.company_id, invoice_id=invoice_id,
                invoice_line_id=line_id, pending_qty=new_pqty,
                created_by_id=current_user.id
            ))
        db.commit()
        return {"success": True, "pending_qty": new_pqty}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.patch("/sales-invoices/{invoice_id}/pending-extra-items/{extra_id}")
async def update_sales_pending_extra_item(
    invoice_id: int,
    extra_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-DISPATCH-EXTRA-001: Update dispatched_qty and/or pending_qty on a sales extra item."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import SalesPendingExtraItem
        ei = db.query(SalesPendingExtraItem).filter(
            SalesPendingExtraItem.id == extra_id,
            SalesPendingExtraItem.invoice_id == invoice_id
        ).first()
        if not ei:
            raise HTTPException(status_code=404, detail="Extra item not found")

        if 'pending_qty' in body and body['pending_qty'] is not None:
            new_pq = float(body['pending_qty'] if body['pending_qty'] is not None else 0)
            if new_pq < 0:
                raise HTTPException(status_code=400, detail="pending_qty cannot be negative")
            ei.pending_qty = new_pq

        if 'dispatched_qty' in body and body['dispatched_qty'] is not None:
            add_qty = float(body['dispatched_qty'] or 0)
            if add_qty < 0:
                raise HTTPException(status_code=400, detail="dispatched_qty must be >= 0")
            new_disp = float(ei.dispatched_qty or 0) + add_qty
            if new_disp > float(ei.pending_qty or 0):
                raise HTTPException(status_code=400, detail="Total dispatched qty would exceed configured pending qty")
            ei.dispatched_qty = new_disp
            pqty = float(ei.pending_qty or 0)
            if new_disp >= pqty > 0:
                ei.dispatch_status = 'FULLY_DISPATCHED'
            elif new_disp > 0:
                ei.dispatch_status = 'PARTIALLY_DISPATCHED'
            else:
                ei.dispatch_status = 'NOT_DISPATCHED'

        db.commit()
        return {
            "success": True,
            "id": ei.id,
            "pending_qty": float(ei.pending_qty),
            "dispatched_qty": float(ei.dispatched_qty),
            "dispatch_status": ei.dispatch_status
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.patch("/purchase-invoices/{invoice_id}/pending-line-config/{line_id}")
async def update_purchase_pending_line_config(
    invoice_id: int,
    line_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-DISPATCH-EXTRA-001: Update pending_qty override for a regular purchase invoice line."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import PurchaseInvoiceUpload, PurchaseInvoiceLineItem, PurchasePendingLineConfig
        new_pqty = float(body.get('pending_qty', 0) if body.get('pending_qty') is not None else 0)
        if new_pqty < 0:
            raise HTTPException(status_code=400, detail="pending_qty cannot be negative")
        ln = db.query(PurchaseInvoiceLineItem).filter(
            PurchaseInvoiceLineItem.id == line_id,
            PurchaseInvoiceLineItem.upload_id == invoice_id
        ).first()
        if not ln:
            raise HTTPException(status_code=404, detail="Line not found")
        if new_pqty > float(ln.quantity or 0):
            raise HTTPException(status_code=400, detail=f"pending_qty cannot exceed ordered qty ({ln.quantity})")
        cfg = db.query(PurchasePendingLineConfig).filter(
            PurchasePendingLineConfig.invoice_id == invoice_id,
            PurchasePendingLineConfig.invoice_line_id == line_id
        ).first()
        if cfg:
            cfg.pending_qty = new_pqty
        else:
            inv = db.query(PurchaseInvoiceUpload).filter(PurchaseInvoiceUpload.id == invoice_id).first()
            if not inv:
                raise HTTPException(status_code=404, detail="Invoice not found")
            db.add(PurchasePendingLineConfig(
                company_id=inv.company_id, invoice_id=invoice_id,
                invoice_line_id=line_id, pending_qty=new_pqty,
                created_by_id=current_user.id
            ))
        db.commit()
        return {"success": True, "pending_qty": new_pqty}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.patch("/purchase-invoices/{invoice_id}/pending-extra-items/{extra_id}")
async def update_purchase_pending_extra_item(
    invoice_id: int,
    extra_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-DISPATCH-EXTRA-001: Update received_qty and/or pending_qty on a purchase extra item."""
    if not is_accounts_allowed_employee(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        from app.models.staff_accounts import PurchasePendingExtraItem
        ei = db.query(PurchasePendingExtraItem).filter(
            PurchasePendingExtraItem.id == extra_id,
            PurchasePendingExtraItem.invoice_id == invoice_id
        ).first()
        if not ei:
            raise HTTPException(status_code=404, detail="Extra item not found")

        if 'pending_qty' in body and body['pending_qty'] is not None:
            new_pq = float(body['pending_qty'] if body['pending_qty'] is not None else 0)
            if new_pq < 0:
                raise HTTPException(status_code=400, detail="pending_qty cannot be negative")
            ei.pending_qty = new_pq

        if 'received_qty' in body and body['received_qty'] is not None:
            add_qty = float(body['received_qty'] or 0)
            if add_qty < 0:
                raise HTTPException(status_code=400, detail="received_qty must be >= 0")
            new_recv = float(ei.received_qty or 0) + add_qty
            if new_recv > float(ei.pending_qty or 0):
                raise HTTPException(status_code=400, detail="Total received qty would exceed configured pending qty")
            ei.received_qty = new_recv
            pqty = float(ei.pending_qty or 0)
            if new_recv >= pqty > 0:
                ei.receipt_status = 'FULLY_RECEIVED'
            elif new_recv > 0:
                ei.receipt_status = 'PARTIALLY_RECEIVED'
            else:
                ei.receipt_status = 'NOT_RECEIVED'

        db.commit()
        return {
            "success": True,
            "id": ei.id,
            "pending_qty": float(ei.pending_qty),
            "received_qty": float(ei.received_qty),
            "receipt_status": ei.receipt_status
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


@router.post("/sales-invoices/{invoice_id}/record-dispatch")
async def record_sales_dispatch(
    invoice_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """
    DC-DISPATCH-001 + DC-DISPATCH-EXTRA-001: Record partial/full dispatch for a confirmed sales invoice.
    data = { dispatch_date, narration, items: [{ line_id, qty_dispatched }],
             extra_item_dispatches: [{ extra_item_id, qty_dispatched }] }
    Auto-updates dispatch_status on invoice (NOT_DISPATCHED/PARTIALLY_DISPATCHED/FULLY_DISPATCHED).
    """
    try:
        from app.models.staff_accounts import (
            SalesInvoice, SalesInvoiceLineItem, SalesDispatchRecord,
            SalesPendingLineConfig, SalesPendingExtraItem
        )
        from datetime import date as _date

        inv = db.query(SalesInvoice).filter(SalesInvoice.id == invoice_id).first()
        if not inv:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if inv.status != 'CONFIRMED':
            raise HTTPException(status_code=400, detail="Only CONFIRMED invoices can record dispatches")

        items_input = data.get('items', [])
        dispatch_date_str = data.get('dispatch_date') or _date.today().isoformat()
        narration = data.get('narration', '')
        try:
            dispatch_date = _date.fromisoformat(dispatch_date_str)
        except Exception:
            dispatch_date = _date.today()

        # Pre-load line configs for cap validation
        line_configs = {
            cfg.invoice_line_id: float(cfg.pending_qty)
            for cfg in db.query(SalesPendingLineConfig).filter(
                SalesPendingLineConfig.invoice_id == invoice_id
            ).all()
        }

        dispatched_count = 0
        for item_in in items_input:
            line_id = item_in.get('line_id')
            qty = float(item_in.get('qty_dispatched', 0))
            if qty <= 0:
                continue
            ln = db.query(SalesInvoiceLineItem).filter(
                SalesInvoiceLineItem.id == line_id,
                SalesInvoiceLineItem.invoice_id == invoice_id
            ).first()
            if not ln:
                continue

            # DC-DISPATCH-EXTRA-001: validate cap at configured pending qty
            inv_qty = float(ln.quantity or 0)
            cap_qty = line_configs.get(line_id, inv_qty)
            already_disp = float(db.execute(
                _tv_text("SELECT COALESCE(SUM(dispatched_qty),0) FROM sales_dispatch_records WHERE invoice_line_id = :lid"),
                {"lid": line_id}
            ).scalar() or 0)
            remaining = max(cap_qty - already_disp, 0.0)
            qty = min(qty, remaining)
            if qty <= 0:
                continue

            avg_cost_row = db.execute(
                _tv_text("""
                    SELECT CASE WHEN balance_qty > 0 THEN balance_value / balance_qty ELSE 0 END
                    FROM stock_ledger
                    WHERE company_id = :cid AND item_id = :iid
                    ORDER BY transaction_date DESC, id DESC
                    LIMIT 1
                """),
                {"cid": inv.company_id, "iid": ln.item_id}
            ).fetchone() if ln.item_id else None
            avg_cost = float(avg_cost_row[0]) if avg_cost_row and avg_cost_row[0] else 0.0

            rec = SalesDispatchRecord(
                company_id=inv.company_id,
                invoice_id=inv.id,
                invoice_line_id=ln.id,
                item_id=ln.item_id,
                dispatched_qty=qty,
                dispatched_value=round(qty * avg_cost, 2),
                avg_cost=avg_cost,
                dispatch_date=dispatch_date,
                narration=narration,
                dispatched_by_id=current_user.id,
            )
            db.add(rec)
            dispatched_count += 1

        # DC-DISPATCH-EXTRA-001: handle extra item dispatches
        extra_dispatched_count = 0
        for ed in data.get('extra_item_dispatches', []):
            eid = ed.get('extra_item_id')
            eqty = float(ed.get('qty_dispatched', 0) or 0)
            if not eid or eqty <= 0:
                continue
            ei = db.query(SalesPendingExtraItem).filter(
                SalesPendingExtraItem.id == eid,
                SalesPendingExtraItem.invoice_id == invoice_id
            ).first()
            if not ei:
                continue
            ei_pqty = float(ei.pending_qty or 0)
            ei_disp = float(ei.dispatched_qty or 0)
            ei_remain = max(ei_pqty - ei_disp, 0.0)
            eqty = min(eqty, ei_remain)
            if eqty <= 0:
                continue
            ei.dispatched_qty = ei_disp + eqty
            new_disp = ei.dispatched_qty
            if float(new_disp) >= ei_pqty > 0:
                ei.dispatch_status = 'FULLY_DISPATCHED'
            elif float(new_disp) > 0:
                ei.dispatch_status = 'PARTIALLY_DISPATCHED'
            extra_dispatched_count += 1

        db.flush()

        # Recompute invoice dispatch_status using configured pending qty as target
        all_lines = db.query(SalesInvoiceLineItem).filter(
            SalesInvoiceLineItem.invoice_id == invoice_id
        ).all()
        fully_dispatched = True
        any_dispatched = False
        for aln in all_lines:
            total_disp = float(db.execute(
                _tv_text("SELECT COALESCE(SUM(dispatched_qty),0) FROM sales_dispatch_records WHERE invoice_line_id = :lid"),
                {"lid": aln.id}
            ).scalar() or 0)
            target = line_configs.get(aln.id, float(aln.quantity or 0))
            if total_disp > 0:
                any_dispatched = True
            if total_disp < target:
                fully_dispatched = False

        # Also factor in extra items
        extra_items_all = db.query(SalesPendingExtraItem).filter(
            SalesPendingExtraItem.invoice_id == invoice_id
        ).all()
        for ei in extra_items_all:
            if float(ei.dispatched_qty or 0) > 0:
                any_dispatched = True
            if float(ei.dispatched_qty or 0) < float(ei.pending_qty or 0):
                fully_dispatched = False

        if fully_dispatched and any_dispatched:
            inv.dispatch_status = 'FULLY_DISPATCHED'
        elif any_dispatched:
            inv.dispatch_status = 'PARTIALLY_DISPATCHED'
        else:
            inv.dispatch_status = 'NOT_DISPATCHED'

        db.commit()
        total_count = dispatched_count + extra_dispatched_count
        return JSONResponse(content={
            "success": True,
            "message": f"{dispatched_count} line(s) + {extra_dispatched_count} extra item(s) dispatch recorded. Invoice status: {inv.dispatch_status}",
            "dispatch_status": inv.dispatch_status
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return handle_accounts_error(e)


# ══════════════════════════════════════════════════════════════════════════════
# DC-BAL-DASH-001: Executive Balance Dashboard — banks / UPI / cash / receivables
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/balance-dashboard")
async def get_balance_dashboard_endpoint(
    company_id: Optional[int] = Query(None, description="Filter by company (0 = all)"),
    period: str = Query('OVERALL', description="OVERALL|FTD|YESTERDAY|THIS_WEEK|THIS_MONTH|THIS_FY"),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC-BAL-DASH-001: Read-only executive snapshot of all bank, UPI, and cash balances
    with period In/Out/Net and true running balance per account, plus receivables total.
    """
    try:
        from app.services.staff_accounts_service import LedgerPostingService
        result = LedgerPostingService.get_balance_dashboard(
            db, current_user,
            company_id=company_id if company_id else None,
            period=(period or 'OVERALL').upper()
        )
        return JSONResponse(content=result)
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)


@router.delete("/training/progress/{emp_id}/{vid_id}")
def reset_training_progress(
    emp_id: int = Path(...),
    vid_id: int = Path(...),
    db: Session = Depends(get_db),
    current_user: StaffEmployee = Depends(get_current_staff_user)
):
    """DC_TRAINING_VIDEOS_001: Admin — reset one employee's video completion (audit-safe)."""
    try:
        if getattr(current_user, 'hierarchy_level', 0) < 80:
            raise HTTPException(status_code=403, detail="Admin access required (hierarchy >= 80)")
        db.execute(_tv_text("""
            UPDATE training_video_progress
               SET is_completed = false,
                   completed_at = NULL
             WHERE employee_id = :emp AND video_id = :vid
        """), {"emp": emp_id, "vid": vid_id})
        db.commit()
        return {"success": True, "message": "Progress reset"}
    except HTTPException:
        raise
    except Exception as e:
        return handle_accounts_error(e)
