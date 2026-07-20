"""
Partner Authentication API Endpoints (DC Protocol Compliant)
DC_PARTNER_AUTH_001: Separate authentication system for official partners (Dec 2025)
"""

from fastapi import APIRouter, Depends, HTTPException, status, Request, Query, Path, UploadFile, File, Form, Response
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, text, func
from datetime import datetime, timedelta, date
from typing import Optional, List, Dict, Any
from decimal import Decimal
import json
from pydantic import BaseModel, Field, EmailStr

from app.core.database import get_db
from app.core.security import SecurityManager
from app.core.config import settings
from app.models.staff_accounts import OfficialPartner, PartnerCompanySegment, PartnerOrder, PartnerInvoice, PartnerPaymentRecord, VendorMaster, SolarVendorLedger
from app.models.staff import PartnerMenuSettings, StaffMenuMaster, StaffEmployee, StaffDepartment
from app.models.crm import CRMLead
from app.models.kyc_document import KYCDocument
from app.services.universal_upload_service import UniversalUploadService

router = APIRouter(prefix="/partner", tags=["Partner Auth"])


class PartnerLoginRequest(BaseModel):
    """DC_PARTNER_AUTH_001: Partner login request with validation"""
    partner_code: str = Field(..., min_length=1, description="Partner Code (e.g., DLR001)")
    password: str = Field(..., min_length=1)
    company_id: Optional[int] = Field(None, description="Optional company_id for DC Protocol")


class PartnerLoginResponse(BaseModel):
    """DC_PARTNER_AUTH_001: Partner login response"""
    success: bool
    message: str
    access_token: Optional[str] = None
    token_type: str = "bearer"
    expires_in: Optional[int] = None
    partner: Optional[dict] = None


class PartnerProfileResponse(BaseModel):
    """DC_PARTNER_AUTH_001: Partner profile response"""
    success: bool
    partner: dict


def get_indian_time():
    """Get current Indian time"""
    from pytz import timezone
    return datetime.now(timezone('Asia/Kolkata'))


def get_current_partner(request: Request, db: Session = Depends(get_db)) -> OfficialPartner:
    """
    Dependency to get current authenticated partner from JWT token.
    DC_PARTNER_AUTH_001: Validates token and ensures partner exists and is active.
    Falls back to ?Authorization=Bearer... query param for browser-tab PDF opens.
    """
    auth_header = request.headers.get("Authorization")

    # DC_PDF_PARTNER_001: PDF opened in a new tab passes token as a query param
    if not auth_header:
        qp = request.query_params.get("Authorization") or request.query_params.get("authorization")
        if qp:
            auth_header = qp if qp.startswith("Bearer ") else f"Bearer {qp}"

    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Missing or invalid authorization header",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    token = auth_header.split(" ")[1]
    
    try:
        from jose import jwt, JWTError
        from jose.exceptions import ExpiredSignatureError
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        
        user_type = payload.get("user_type")
        if user_type != "partner":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid token type for partner access",
                headers={"WWW-Authenticate": "Bearer"}
            )
        
        partner_id = payload.get("sub")
        if not partner_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid token payload",
                headers={"WWW-Authenticate": "Bearer"}
            )
        
        partner = db.query(OfficialPartner).filter_by(id=int(partner_id)).first()
        if not partner:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Partner not found",
                headers={"WWW-Authenticate": "Bearer"}
            )
        
        # DC_PARTNER_STATUS_001: Status blocking messages (Apr 2026)
        # active=full access; inactive/suspended/locked=login blocked; pause=login blocked(temp); expired=login blocked
        ls = partner.login_status or 'active'
        if ls == 'inactive' or not partner.is_active:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Partner account is inactive. Contact support."
            )
        if ls == 'suspended':
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Partner account is suspended. Contact support."
            )
        if ls == 'locked':
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Partner account is locked due to multiple failed login attempts. Contact support."
            )
        if ls == 'pause':
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Partner account is currently paused. Contact support to reactivate."
            )
        if ls == 'expired':
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Partner account has expired. Please renew your partnership agreement."
            )
        
        return partner
        
    except ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="TOKEN_EXPIRED",
            headers={"WWW-Authenticate": "Bearer"}
        )
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="INVALID_TOKEN",
            headers={"WWW-Authenticate": "Bearer"}
        )


@router.post("/auth/login", response_model=PartnerLoginResponse)
async def partner_login(request: PartnerLoginRequest, db: Session = Depends(get_db)):
    """
    DC_PARTNER_AUTH_001: Partner login endpoint
    - Authenticates partner by partner_code and password
    - Returns JWT token with user_type='partner'
    - DC Protocol: Validates company access via PartnerCompanySegment
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[DC-PARTNER-LOGIN] Attempt for partner_code: {request.partner_code}")
    
    # Find partner by partner_code (case-insensitive)
    partner = db.query(OfficialPartner).filter(
        OfficialPartner.partner_code.ilike(request.partner_code.strip())
    ).first()

    # [DC-PARTNER-CONTACTS-001] Dual-login: if no partner found, try staff credentials
    # Sales/Service dept staff with linked_partner_id can log in via partner portal
    _staff_showroom_login = False
    _staff_role = None  # [DC-ROLE-001] 'sales', 'service', or None
    _login_phone = request.partner_code.strip()
    if not partner:
        try:
            # [DC-PARTNER-CONTACTS-001] Staff log in via MOBILE NUMBER only
            _staff = db.query(StaffEmployee).filter(
                StaffEmployee.phone == _login_phone,
                StaffEmployee.status == 'active',
                StaffEmployee.linked_partner_id != None  # noqa: E711
            ).first()
            if _staff and _staff.linked_partner_id:
                # Verify staff password
                if SecurityManager.verify_password(request.password, _staff.password_hash):
                    # Any staff with a linked_partner_id can access that showroom's portal
                    partner = db.query(OfficialPartner).filter(
                        OfficialPartner.id == _staff.linked_partner_id
                    ).first()
                    if partner:
                        _staff_showroom_login = True
                        # [DC-ROLE-001] Determine role by matching phone against partner contact numbers
                        _sc = (partner.sales_contact_number or '').strip()
                        _svc = (partner.service_contact_number or '').strip()
                        if _sc and _login_phone == _sc:
                            _staff_role = 'sales'
                        elif _svc and _login_phone == _svc:
                            _staff_role = 'service'
                        logger.info(f"[DC-PARTNER-LOGIN] Staff showroom login: {_staff.emp_code} → {partner.partner_code} role={_staff_role}")
        except Exception as _dual_err:
            logger.warning(f"[DC-PARTNER-LOGIN] Dual-login check failed (non-fatal): {_dual_err}")

    # [DC-ROLE-001] Third fallback: staff whose phone is saved as sales/service contact on a partner
    # (staff may not have linked_partner_id set, but their number IS in the partner's contact fields)
    if not partner:
        try:
            from sqlalchemy import or_ as _or
            _contact_partner = db.query(OfficialPartner).filter(
                _or(
                    OfficialPartner.sales_contact_number == _login_phone,
                    OfficialPartner.service_contact_number == _login_phone
                ),
                OfficialPartner.is_active == True  # noqa: E712
            ).first()
            if _contact_partner:
                _contact_staff = db.query(StaffEmployee).filter(
                    StaffEmployee.phone == _login_phone,
                    StaffEmployee.status == 'active'
                ).first()
                if _contact_staff and SecurityManager.verify_password(request.password, _contact_staff.password_hash):
                    partner = _contact_partner
                    _staff_showroom_login = True
                    _sc = (_contact_partner.sales_contact_number or '').strip()
                    _svc = (_contact_partner.service_contact_number or '').strip()
                    if _sc and _login_phone == _sc:
                        _staff_role = 'sales'
                    elif _svc and _login_phone == _svc:
                        _staff_role = 'service'
                    logger.info(f"[DC-ROLE-001] Contact-number login: phone={_login_phone} → {partner.partner_code} role={_staff_role}")
        except Exception as _contact_err:
            logger.warning(f"[DC-ROLE-001] Contact login check failed (non-fatal): {_contact_err}")

    # [DC-SOLAR-VENDOR-001] Fourth fallback: Solar vendor login via vendor_code (Apr 2026)
    # vendor_code is used as both login ID and initial password
    _solar_vendor_login = False
    _solar_vendor_ref_id = None
    if not partner:
        try:
            _vendor = db.query(VendorMaster).filter(
                VendorMaster.vendor_code == request.partner_code.strip().upper(),
                VendorMaster.vendor_type == 'SOLAR'
            ).first()
            if _vendor:
                # Password must match vendor_code (case-insensitive) OR hashed vendor_code
                _pw_input = request.password.strip()
                _pw_plain_ok = (_pw_input.upper() == _vendor.vendor_code.upper())
                # Check if OfficialPartner with category='VENDOR' exists for this vendor
                # Try by legacy_vendor_id first; fall back to partner_code='SV-{vendor_code}'
                _op_existing = db.query(OfficialPartner).filter(
                    OfficialPartner.category == 'VENDOR',
                    OfficialPartner.legacy_vendor_id == _vendor.id
                ).first()
                if not _op_existing:
                    _sv_code = f"SV-{_vendor.vendor_code}"
                    _op_existing = db.query(OfficialPartner).filter(
                        OfficialPartner.partner_code == _sv_code
                    ).first()
                    if _op_existing and _op_existing.legacy_vendor_id is None:
                        # Back-fill the link so future lookups are faster
                        try:
                            _op_existing.legacy_vendor_id = _vendor.id
                            db.flush()
                        except Exception:
                            pass
                if _op_existing:
                    # Verify against stored hash or plain match
                    _hash_ok = (_op_existing.password_hash and SecurityManager.verify_password(_pw_input, _op_existing.password_hash))
                    if _pw_plain_ok or _hash_ok:
                        partner = _op_existing
                        _solar_vendor_login = True
                        _solar_vendor_ref_id = _vendor.id
                        logger.info(f"[DC-SOLAR-VENDOR-001] Existing solar vendor login: {_vendor.vendor_code} → partner_id={partner.id}")
                elif _pw_plain_ok:
                    # Auto-provision OfficialPartner for this solar vendor
                    _pw_hash = SecurityManager.get_password_hash(_vendor.vendor_code)
                    _new_partner = OfficialPartner(
                        partner_code=f"SV-{_vendor.vendor_code}",
                        partner_name=_vendor.vendor_name,
                        category='VENDOR',
                        partner_type='SOLAR',
                        phone=_vendor.phone,
                        email=_vendor.email,
                        contact_person=_vendor.contact_person,
                        gst_number=_vendor.gst_number,
                        pan_number=_vendor.pan_number,
                        address=_vendor.address,
                        city=_vendor.city,
                        state=_vendor.state,
                        pincode=_vendor.pincode,
                        is_active=True,
                        login_status='active',
                        password_hash=_pw_hash,
                        legacy_vendor_id=_vendor.id,
                        failed_login_attempts=0,
                        module_settings={},
                    )
                    db.add(_new_partner)
                    db.flush()
                    partner = _new_partner
                    _solar_vendor_login = True
                    _solar_vendor_ref_id = _vendor.id
                    db.commit()
                    logger.info(f"[DC-SOLAR-VENDOR-001] Auto-provisioned OfficialPartner for vendor: {_vendor.vendor_code} → partner_id={partner.id}")
        except Exception as _sv_err:
            db.rollback()
            logger.warning(f"[DC-SOLAR-VENDOR-001] Solar vendor login check failed (non-fatal): {_sv_err}")

    if not partner:
        logger.warning(f"[DC-PARTNER-LOGIN] Partner not found: {request.partner_code}")
        return PartnerLoginResponse(
            success=False,
            message="Invalid partner code or password"
        )
    
    # DC_PARTNER_STATUS_001: Check login status (Apr 2026)
    _ls = partner.login_status or 'active'
    if not partner.is_active or _ls == 'inactive':
        return PartnerLoginResponse(
            success=False,
            message="Partner account is inactive. Contact support."
        )
    if _ls == 'suspended':
        return PartnerLoginResponse(
            success=False,
            message="Partner account is suspended. Contact support."
        )
    if _ls == 'locked':
        return PartnerLoginResponse(
            success=False,
            message="Partner account is locked. Contact support to unlock."
        )
    if _ls == 'pause':
        return PartnerLoginResponse(
            success=False,
            message="Partner account is currently paused. Contact support to reactivate."
        )
    if _ls == 'expired':
        return PartnerLoginResponse(
            success=False,
            message="Partner account has expired. Please renew your partnership agreement."
        )
    
    # Check if password is set (only for normal partner login — skip for staff showroom login)
    if not _staff_showroom_login and not partner.password_hash:
        return PartnerLoginResponse(
            success=False,
            message="Password not set. Contact admin to set your password."
        )

    # Verify password (skipped for staff showroom login — password already verified above)
    password_valid = _staff_showroom_login or SecurityManager.verify_password(request.password, partner.password_hash)
    logger.info(f"[DC-PARTNER-LOGIN] Password verification for {partner.partner_code}: {password_valid} (showroom={_staff_showroom_login})")

    if not password_valid:
        # Increment failed login attempts
        partner.failed_login_attempts = (partner.failed_login_attempts or 0) + 1
        logger.warning(f"[DC-PARTNER-LOGIN] Failed login for {partner.partner_code}, attempt #{partner.failed_login_attempts}")
        
        # Lock account after 5 failed attempts
        if partner.failed_login_attempts >= 5:
            partner.login_status = 'locked'
            db.commit()
            return PartnerLoginResponse(
                success=False,
                message="Account locked due to multiple failed login attempts. Contact support."
            )
        
        db.commit()
        return PartnerLoginResponse(
            success=False,
            message="Invalid partner code or password"
        )
    
    logger.info(f"[DC-PARTNER-LOGIN] Successful login for {partner.partner_code}")
    
    # DC Protocol: Get partner's company access
    company_segments = db.query(PartnerCompanySegment).filter(
        PartnerCompanySegment.partner_id == partner.id,
        PartnerCompanySegment.is_active == True
    ).all()
    
    company_ids = [cs.company_id for cs in company_segments]
    primary_company_id = None
    for cs in company_segments:
        if cs.is_primary:
            primary_company_id = cs.company_id
            break
    if not primary_company_id and company_ids:
        primary_company_id = company_ids[0]
    
    # Reset failed login attempts on successful login
    partner.failed_login_attempts = 0
    partner.last_login = get_indian_time()
    db.commit()

    # DC Protocol Mar 2026: Record partner login session for analytics (non-blocking)
    try:
        from app.api.v1.endpoints.session_analytics import insert_session_log
        insert_session_log(
            db=db,
            user_type="partner",
            user_id=str(partner.id),
            user_identifier=partner.partner_code,
            display_name=partner.partner_name or partner.partner_code,
            ip_address=None,
            user_agent=None,
            token_expiry_minutes=24 * 60,
        )
    except Exception:
        pass
    
    # DC Protocol (Dec 20, 2025): Auto-sync menu settings on partner login
    # Ensures partner has baseline menu settings even if bulk repair missed them
    try:
        sync_default_menu_settings_for_partner(db, partner.id, primary_company_id or (company_ids[0] if company_ids else None))
    except Exception as sync_err:
        import logging
        logging.getLogger(__name__).warning(f"[DC-PARTNER-LOGIN-SYNC] Menu sync failed for {partner.partner_code}: {sync_err}")
    
    # Create JWT token
    from jose import jwt
    token_data = {
        "sub": str(partner.id),
        "user_type": "partner",
        "partner_code": partner.partner_code,
        "category": partner.category,
        "company_ids": company_ids,
        "primary_company_id": primary_company_id,
        "showroom_login": _staff_showroom_login,  # [DC-PARTNER-CONTACTS-001]
        "staff_role": _staff_role,                 # [DC-ROLE-001] 'sales', 'service', or None
        "solar_vendor_login": _solar_vendor_login, # [DC-SOLAR-VENDOR-001]
        "vendor_ref_id": _solar_vendor_ref_id,     # [DC-SOLAR-VENDOR-001] VendorMaster.id
        "exp": datetime.utcnow() + timedelta(hours=24)
    }
    access_token = jwt.encode(token_data, settings.SECRET_KEY, algorithm=settings.ALGORITHM)

    return PartnerLoginResponse(
        success=True,
        message="Login successful",
        access_token=access_token,
        expires_in=24 * 3600,
        partner={
            "id": partner.id,
            "partner_code": partner.partner_code,
            "partner_name": partner.partner_name,
            "category": partner.category,
            "partner_type": partner.partner_type,
            "email": partner.email,
            "phone": partner.phone,
            "company_ids": company_ids,
            "primary_company_id": primary_company_id,
            "login_status": partner.login_status or 'active',
            "showroom_login": _staff_showroom_login,
            "staff_role": _staff_role,                     # [DC-ROLE-001]
            "solar_vendor_login": _solar_vendor_login,     # [DC-SOLAR-VENDOR-001]
            "vendor_ref_id": _solar_vendor_ref_id,         # [DC-SOLAR-VENDOR-001]
            "module_settings": partner.module_settings or {},
            "sales_contact_number": partner.sales_contact_number,
            "sales_contact_name": partner.sales_contact_name,
            "service_contact_number": partner.service_contact_number,
            "service_contact_name": partner.service_contact_name,
        }
    )


@router.get("/auth/me", response_model=PartnerProfileResponse)
async def get_partner_profile(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_PARTNER_AUTH_001: Get current partner profile
    """
    # Get company access
    company_segments = db.query(PartnerCompanySegment).filter(
        PartnerCompanySegment.partner_id == partner.id,
        PartnerCompanySegment.is_active == True
    ).all()
    
    company_ids = [cs.company_id for cs in company_segments]
    primary_company_id = None
    for cs in company_segments:
        if cs.is_primary:
            primary_company_id = cs.company_id
            break
    
    return PartnerProfileResponse(
        success=True,
        partner={
            "id": partner.id,
            "partner_code": partner.partner_code,
            "partner_name": partner.partner_name,
            "category": partner.category,
            "partner_type": partner.partner_type,
            "email": partner.email,
            "phone": partner.phone,
            "whatsapp_number": partner.whatsapp_number,
            "address": partner.address,
            "city": partner.city,
            "state": partner.state,
            "pincode": partner.pincode,
            "gst_number": partner.gst_number,
            "company_ids": company_ids,
            "primary_company_id": primary_company_id,
            "last_login": partner.last_login.isoformat() if partner.last_login else None,
            "login_status": partner.login_status or 'active',
            "is_active": partner.is_active,
            "logo_path": partner.logo_path or None,
            "logo_url": f"/uploads/{partner.logo_path}" if partner.logo_path else None,
            "sales_contact_name": partner.sales_contact_name,
            "sales_contact_number": partner.sales_contact_number,
            "service_contact_name": partner.service_contact_name,
            "service_contact_number": partner.service_contact_number,
            "module_settings": partner.module_settings or {},
        }
    )


def sync_default_menu_settings_for_partner(db: Session, partner_id: int, company_id: int):
    """
    Auto-sync: Create PartnerMenuSettings for default-visible menus
    DC Protocol: Only creates settings where none exist (respects admin overrides)
    """
    try:
        default_menus = db.query(StaffMenuMaster).filter(
            StaffMenuMaster.company_id == company_id,
            StaffMenuMaster.is_default_visible == True,
            StaffMenuMaster.is_active == True,
            StaffMenuMaster.audience_scope.in_(['partner', 'shared'])
        ).all()
        
        created_count = 0
        for menu in default_menus:
            existing = db.query(PartnerMenuSettings).filter(
                PartnerMenuSettings.partner_id == partner_id,
                PartnerMenuSettings.menu_id == menu.id,
                PartnerMenuSettings.company_id == company_id
            ).first()
            
            if not existing:
                new_setting = PartnerMenuSettings(
                    partner_id=partner_id,
                    menu_id=menu.id,
                    company_id=company_id,
                    can_view=True,
                    can_edit=False,
                    is_overridden=False,
                    set_by_id=None,
                    set_by_code="SYSTEM",
                    set_by_name="Auto-Sync"
                )
                db.add(new_setting)
                created_count += 1
        
        if created_count > 0:
            db.commit()
        return created_count
    except Exception as e:
        db.rollback()
        return 0


@router.get("/auth/my-menus")
async def get_partner_menus(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_PARTNER_AUTH_001: Get partner's accessible menus based on Zero-Default Access Policy
    - Auto-syncs default-visible menus for the partner BEFORE querying
    - Returns menus explicitly granted to the partner via PartnerMenuSettings
    - Respects admin overrides (existing settings are never overwritten)
    - Filters by audience_scope in ('partner', 'shared')
    """
    # DC Protocol: Get company_id from header or partner's primary company
    company_id = request.headers.get("X-Company-ID")
    if not company_id:
        company_segments = db.query(PartnerCompanySegment).filter(
            PartnerCompanySegment.partner_id == partner.id,
            PartnerCompanySegment.is_active == True,
            PartnerCompanySegment.is_primary == True
        ).first()
        if company_segments:
            company_id = company_segments.company_id
    
    if not company_id:
        return {"success": True, "menus": [], "message": "No company access configured"}
    
    # AUTO-SYNC: Create missing settings for default-visible menus BEFORE querying
    # DC Protocol: Only creates settings where none exist (respects admin overrides)
    try:
        partner_companies = set()
        all_segments = db.query(PartnerCompanySegment).filter(
            PartnerCompanySegment.partner_id == partner.id,
            PartnerCompanySegment.is_active == True
        ).all()
        for seg in all_segments:
            if seg.company_id:
                partner_companies.add(int(seg.company_id))
        
        for cid in partner_companies:
            sync_default_menu_settings_for_partner(db, partner.id, cid)
    except Exception as e:
        pass
    
    # Get partner's menu access from PartnerMenuSettings
    partner_menus = db.query(PartnerMenuSettings).filter(
        PartnerMenuSettings.partner_id == partner.id,
        PartnerMenuSettings.company_id == int(company_id),
        PartnerMenuSettings.can_view == True
    ).all()
    
    menu_ids = [pm.menu_id for pm in partner_menus]
    
    # Get menu details for granted menus
    menus = []
    if menu_ids:
        menu_records = db.query(StaffMenuMaster).filter(
            StaffMenuMaster.id.in_(menu_ids),
            StaffMenuMaster.is_active == True,
            StaffMenuMaster.audience_scope.in_(['partner', 'shared'])
        ).order_by(StaffMenuMaster.display_order).all()
        
        for menu in menu_records:
            # Find if partner has edit access
            partner_setting = next((pm for pm in partner_menus if pm.menu_id == menu.id), None)
            menus.append({
                "id": menu.id,
                "menu_name": menu.menu_name,
                "route_path": menu.route_path,
                "icon": menu.menu_icon,
                "parent_id": menu.parent_id,
                "display_order": menu.display_order,
                "category": menu.menu_category,
                "can_view": True,
                "can_edit": partner_setting.can_edit if partner_setting else False
            })
    
    # [DC-PARTNER-CONTACTS-001] Apply module_settings filter
    # [DC-ROLE-001] Role-based filtering: extract staff_role from JWT
    _staff_role = None
    try:
        _auth_hdr = request.headers.get("Authorization") or ""
        if _auth_hdr.startswith("Bearer "):
            from jose import jwt as _jwt
            _payload = _jwt.decode(_auth_hdr.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            _staff_role = _payload.get("staff_role")
    except Exception:
        pass

    # Map module keys → partial route path substring
    _MOD_ROUTE_MAP = {
        "walkins":     "/walkins",
        "leads":       "/my-leads",
        "service":     "/support",
        "marketplace": "/marketplace",
        "stock":       "/stock",
        "sales":       "/sales",
    }
    _ms = partner.module_settings or {}

    # [DC-ROLE-001] Determine the active filter set based on role and format:
    # New format stores {sales_staff:{...}, service_staff:{...}};
    # old flat format {walkins:bool, ...} applies to all (backward compat).
    if _ms:
        if 'sales_staff' in _ms or 'service_staff' in _ms:
            # New role-based format
            if _staff_role == 'sales':
                _active_ms = _ms.get('sales_staff') or {}
            elif _staff_role == 'service':
                _active_ms = _ms.get('service_staff') or {}
            else:
                # Main partner login: use flat residual keys if any, otherwise allow all
                _active_ms = {k: v for k, v in _ms.items() if k not in ('sales_staff', 'service_staff')}
        else:
            # Old flat format — applies to all logins equally (backward compat)
            _active_ms = _ms

        if _active_ms:
            filtered_menus = []
            for m in menus:
                route = m.get("route_path", "") or ""
                blocked = False
                for mod_key, route_frag in _MOD_ROUTE_MAP.items():
                    if route_frag in route and _active_ms.get(mod_key) is False:
                        blocked = True
                        break
                if not blocked:
                    filtered_menus.append(m)
            menus = filtered_menus

    return {
        "success": True,
        "menus": menus,
        "partner_id": partner.id,
        "partner_type": partner.partner_type,
        "staff_role": _staff_role,
        "module_settings": _ms,
        "company_id": int(company_id) if company_id else None
    }


# DC_PARTNER_SELF_SERVICE_001: Partner self-service read endpoints (Apr 2026)

@router.get("/auth/my-orders")
async def get_my_orders(
    request: Request,
    page: int = 1,
    limit: int = 20,
    status: str = None,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """Returns the logged-in partner's own orders (read-only, no staff access required)."""
    try:
        q = db.query(PartnerOrder).filter(PartnerOrder.partner_id == partner.id)
        if status:
            q = q.filter(PartnerOrder.status == status.upper())
        total = q.count()
        orders = q.order_by(PartnerOrder.created_at.desc()).offset((page - 1) * limit).limit(limit).all()
        return {
            "success": True,
            "orders": [o.to_dict() for o in orders],
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if total > 0 else 1
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/auth/my-invoices")
async def get_my_invoices(
    request: Request,
    page: int = 1,
    limit: int = 20,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """Returns the logged-in partner's own invoices."""
    try:
        q = db.query(PartnerInvoice).filter(PartnerInvoice.partner_id == partner.id)
        total = q.count()
        invoices = q.order_by(PartnerInvoice.invoice_date.desc()).offset((page - 1) * limit).limit(limit).all()
        return {
            "success": True,
            "invoices": [i.to_dict() for i in invoices],
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/auth/my-payments")
async def get_my_payments(
    request: Request,
    page: int = 1,
    limit: int = 20,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """Returns the logged-in partner's payment records via their orders."""
    try:
        q = (
            db.query(PartnerPaymentRecord)
            .join(PartnerOrder, PartnerOrder.id == PartnerPaymentRecord.order_id)
            .filter(PartnerOrder.partner_id == partner.id)
        )
        total = q.count()
        payments = q.order_by(PartnerPaymentRecord.created_at.desc()).offset((page - 1) * limit).limit(limit).all()
        return {
            "success": True,
            "payments": [p.to_dict() for p in payments],
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/auth/logout")
async def partner_logout(request: Request, db: Session = Depends(get_db)):
    """
    DC_PARTNER_AUTH_001: Partner logout endpoint
    - Client should clear the token
    DC Protocol Mar 2026: Close session record for analytics (non-blocking)
    """
    try:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
            try:
                from jose import jwt
                from app.core.config import settings
                payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                partner_id = payload.get("sub")
                if partner_id:
                    from app.api.v1.endpoints.session_analytics import close_session_log
                    close_session_log(db=db, user_type="partner", user_id=str(partner_id))
            except Exception:
                pass
    except Exception:
        pass
    return {"success": True, "message": "Logged out successfully"}


@router.post("/auth/change-password")
async def change_partner_password(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_PARTNER_AUTH_001: Change partner password
    """
    body = await request.json()
    current_password = body.get("current_password")
    new_password = body.get("new_password")
    
    if not current_password or not new_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password and new password are required"
        )
    
    if len(new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="New password must be at least 6 characters"
        )
    
    # Verify current password
    if not partner.password_hash or not SecurityManager.verify_password(current_password, partner.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Update password
    partner.password_hash = SecurityManager.get_password_hash(new_password)
    partner.password_changed_at = get_indian_time()
    db.commit()
    
    return {"success": True, "message": "Password changed successfully"}


# =============================================================================
# DC_INTAKE_001: Partner Portal - Vendor Returns (Jan 2026)
# Vendor-facing endpoints for viewing and acknowledging return requests
# =============================================================================

from app.models.staff_accounts import VendorReturnRequest, VendorReturnItem


class VendorAcknowledgeRequest(BaseModel):
    """Request to acknowledge a return request"""
    remarks: Optional[str] = Field(None, max_length=1000, description="Vendor remarks or response")
    accept: bool = Field(..., description="Whether vendor accepts the return request")


@router.get("/returns/my-returns")
async def list_my_returns(
    request: Request,
    status_filter: Optional[str] = Query(None, description="Filter by status"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_INTAKE_001: List return requests addressed to current vendor partner
    Vendors see only their own return requests (vendor_id = partner.id)
    """
    query = db.query(VendorReturnRequest).filter(
        VendorReturnRequest.vendor_id == partner.id
    )
    
    if status_filter:
        query = query.filter(VendorReturnRequest.status == status_filter.upper())
    
    total = query.count()
    returns = query.order_by(VendorReturnRequest.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    results = []
    for r in returns:
        items = db.query(VendorReturnItem).filter(VendorReturnItem.return_request_id == r.id).all()
        results.append({
            "id": r.id,
            "request_number": r.request_number,
            "company_id": r.company_id,
            "request_type": r.request_type,
            "total_items": r.total_items,
            "total_qty": float(r.total_qty) if r.total_qty else 0,
            "total_value": float(r.total_value) if r.total_value else 0,
            "status": r.status,
            "vendor_response_deadline": r.vendor_response_deadline.isoformat() if r.vendor_response_deadline else None,
            "vendor_acknowledged_at": r.vendor_acknowledged_at.isoformat() if r.vendor_acknowledged_at else None,
            "vendor_remarks": r.vendor_remarks,
            "dispatch_date": r.dispatch_date.isoformat() if r.dispatch_date else None,
            "dispatch_courier": r.dispatch_courier,
            "dispatch_tracking_number": r.dispatch_tracking_number,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "items": [
                {
                    "id": item.id,
                    "item_name": item.item_name,
                    "item_code": item.item_code,
                    "qty": float(item.qty) if item.qty else 0,
                    "unit_price": float(item.unit_price) if item.unit_price else 0,
                    "total_price": float(item.total_price) if item.total_price else 0,
                    "reason": item.reason,
                    "status": item.status
                }
                for item in items
            ]
        })
    
    return {
        "success": True,
        "data": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/returns/my-returns/{return_id}")
async def get_my_return_detail(
    request: Request,
    return_id: int = Path(..., gt=0),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_INTAKE_001: Get single return request detail for vendor
    """
    return_req = db.query(VendorReturnRequest).filter(
        VendorReturnRequest.id == return_id,
        VendorReturnRequest.vendor_id == partner.id
    ).first()
    
    if not return_req:
        raise HTTPException(status_code=404, detail="Return request not found")
    
    items = db.query(VendorReturnItem).filter(VendorReturnItem.return_request_id == return_id).all()
    
    return {
        "success": True,
        "data": {
            "id": return_req.id,
            "request_number": return_req.request_number,
            "company_id": return_req.company_id,
            "request_type": return_req.request_type,
            "total_items": return_req.total_items,
            "total_qty": float(return_req.total_qty) if return_req.total_qty else 0,
            "total_value": float(return_req.total_value) if return_req.total_value else 0,
            "status": return_req.status,
            "vendor_response_deadline": return_req.vendor_response_deadline.isoformat() if return_req.vendor_response_deadline else None,
            "vendor_acknowledged_at": return_req.vendor_acknowledged_at.isoformat() if return_req.vendor_acknowledged_at else None,
            "vendor_remarks": return_req.vendor_remarks,
            "dispatch_date": return_req.dispatch_date.isoformat() if return_req.dispatch_date else None,
            "dispatch_courier": return_req.dispatch_courier,
            "dispatch_tracking_number": return_req.dispatch_tracking_number,
            "dispatch_notes": return_req.dispatch_notes,
            "received_by_vendor_at": return_req.received_by_vendor_at.isoformat() if return_req.received_by_vendor_at else None,
            "vendor_received_notes": return_req.vendor_received_notes,
            "credit_note_number": return_req.credit_note_number,
            "credit_note_date": return_req.credit_note_date.isoformat() if return_req.credit_note_date else None,
            "credit_note_amount": float(return_req.credit_note_amount) if return_req.credit_note_amount else None,
            "created_at": return_req.created_at.isoformat() if return_req.created_at else None,
            "items": [
                {
                    "id": item.id,
                    "item_name": item.item_name,
                    "item_code": item.item_code,
                    "qty": float(item.qty) if item.qty else 0,
                    "unit_price": float(item.unit_price) if item.unit_price else 0,
                    "total_price": float(item.total_price) if item.total_price else 0,
                    "reason": item.reason,
                    "qc_remarks": item.qc_remarks,
                    "status": item.status
                }
                for item in items
            ]
        }
    }


@router.post("/returns/my-returns/{return_id}/acknowledge")
async def acknowledge_return_request(
    request: Request,
    return_id: int = Path(..., gt=0),
    ack_data: VendorAcknowledgeRequest = None,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_INTAKE_001: Vendor acknowledges return request
    Updates status to ACKNOWLEDGED and records vendor response
    """
    return_req = db.query(VendorReturnRequest).filter(
        VendorReturnRequest.id == return_id,
        VendorReturnRequest.vendor_id == partner.id
    ).first()
    
    if not return_req:
        raise HTTPException(status_code=404, detail="Return request not found")
    
    if return_req.status not in ['CREATED', 'NOTIFIED']:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot acknowledge return in {return_req.status} status"
        )
    
    return_req.vendor_acknowledged_at = get_indian_time()
    return_req.vendor_remarks = ack_data.remarks if ack_data else None
    return_req.status = 'ACKNOWLEDGED' if (ack_data and ack_data.accept) else 'DISPUTED'
    return_req.updated_at = get_indian_time()
    
    db.commit()
    db.refresh(return_req)
    
    return {
        "success": True,
        "message": f"Return request {return_req.request_number} {'acknowledged' if return_req.status == 'ACKNOWLEDGED' else 'disputed'}",
        "data": {
            "id": return_req.id,
            "request_number": return_req.request_number,
            "status": return_req.status,
            "vendor_acknowledged_at": return_req.vendor_acknowledged_at.isoformat()
        }
    }


@router.post("/returns/my-returns/{return_id}/dispute")
async def dispute_return_request(
    request: Request,
    return_id: int = Path(..., gt=0),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_INTAKE_001: Vendor disputes a return request
    Allows vendor to raise concerns before accepting the return
    """
    body = await request.json()
    dispute_reason = body.get("reason", "")
    
    if not dispute_reason or len(dispute_reason.strip()) < 10:
        raise HTTPException(status_code=400, detail="Dispute reason must be at least 10 characters")
    
    return_req = db.query(VendorReturnRequest).filter(
        VendorReturnRequest.id == return_id,
        VendorReturnRequest.vendor_id == partner.id
    ).first()
    
    if not return_req:
        raise HTTPException(status_code=404, detail="Return request not found")
    
    if return_req.status not in ['CREATED', 'NOTIFIED']:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot dispute return in {return_req.status} status"
        )
    
    return_req.vendor_acknowledged_at = get_indian_time()
    return_req.vendor_remarks = dispute_reason
    return_req.status = 'DISPUTED'
    return_req.updated_at = get_indian_time()
    
    db.commit()
    db.refresh(return_req)
    
    return {
        "success": True,
        "message": f"Return request {return_req.request_number} has been disputed",
        "data": {
            "id": return_req.id,
            "request_number": return_req.request_number,
            "status": return_req.status,
            "vendor_acknowledged_at": return_req.vendor_acknowledged_at.isoformat(),
            "dispute_reason": dispute_reason
        }
    }


@router.post("/returns/my-returns/{return_id}/confirm-receipt")
async def confirm_return_receipt(
    request: Request,
    return_id: int = Path(..., gt=0),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_INTAKE_001: Vendor confirms receipt of returned items
    Updates status to RECEIVED_BY_VENDOR
    """
    body = await request.json()
    notes = body.get("notes", "")
    
    return_req = db.query(VendorReturnRequest).filter(
        VendorReturnRequest.id == return_id,
        VendorReturnRequest.vendor_id == partner.id
    ).first()
    
    if not return_req:
        raise HTTPException(status_code=404, detail="Return request not found")
    
    if return_req.status not in ['DISPATCHED', 'ACKNOWLEDGED']:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot confirm receipt in {return_req.status} status. Items must be dispatched first."
        )
    
    return_req.received_by_vendor_at = get_indian_time()
    return_req.vendor_received_notes = notes
    return_req.status = 'RECEIVED_BY_VENDOR'
    return_req.updated_at = get_indian_time()
    
    db.commit()
    db.refresh(return_req)
    
    return {
        "success": True,
        "message": f"Return receipt confirmed for {return_req.request_number}",
        "data": {
            "id": return_req.id,
            "request_number": return_req.request_number,
            "status": return_req.status,
            "received_by_vendor_at": return_req.received_by_vendor_at.isoformat()
        }
    }


@router.get("/returns/summary")
async def get_returns_summary(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_INTAKE_001: Get summary stats for vendor returns dashboard
    """
    from sqlalchemy import func
    
    base_query = db.query(VendorReturnRequest).filter(
        VendorReturnRequest.vendor_id == partner.id
    )
    
    total = base_query.count()
    pending = base_query.filter(VendorReturnRequest.status.in_(['CREATED', 'NOTIFIED'])).count()
    acknowledged = base_query.filter(VendorReturnRequest.status == 'ACKNOWLEDGED').count()
    dispatched = base_query.filter(VendorReturnRequest.status == 'DISPATCHED').count()
    received = base_query.filter(VendorReturnRequest.status == 'RECEIVED_BY_VENDOR').count()
    resolved = base_query.filter(VendorReturnRequest.status == 'RESOLVED').count()
    
    pending_value = db.query(func.coalesce(func.sum(VendorReturnRequest.total_value), 0)).filter(
        VendorReturnRequest.vendor_id == partner.id,
        VendorReturnRequest.status.in_(['CREATED', 'NOTIFIED', 'ACKNOWLEDGED', 'DISPATCHED'])
    ).scalar() or 0
    
    return {
        "success": True,
        "data": {
            "total": total,
            "pending_action": pending,
            "acknowledged": acknowledged,
            "dispatched": dispatched,
            "received": received,
            "resolved": resolved,
            "pending_value": float(pending_value)
        }
    }


@router.get("/service-tickets")
async def get_partner_service_tickets(
    request: Request,
    status: Optional[str] = Query(None, description="Filter by status"),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_PARTNER_SERVICE_001: Get service tickets assigned to this partner
    Returns list of tickets where the partner is the service center
    """
    from app.models.ticket import ServiceTicket
    
    query = db.query(ServiceTicket).filter(
        ServiceTicket.partner_id == partner.id
    )
    
    if status:
        query = query.filter(ServiceTicket.status == status)
    
    tickets = query.order_by(ServiceTicket.created_at.desc()).limit(100).all()
    
    return {
        "success": True,
        "tickets": [
            {
                "id": t.id,
                "ticket_number": t.ticket_number,
                "customer_name": t.customer_name,
                "customer_mobile": t.customer_mobile,
                "vehicle_number": t.vehicle_number,
                "vehicle_model": t.vehicle_model,
                "issue_type": t.issue_type,
                "issue_description": t.issue_description,
                "status": t.status,
                "priority": t.priority or "normal",
                "created_at": t.created_at.isoformat() if t.created_at else None,
                "updated_at": t.updated_at.isoformat() if t.updated_at else None
            }
            for t in tickets
        ]
    }


# ─── DC Protocol Mar 2026: Partner Profile Edit + KYC ────────────────────────

_PARTNER_KYC_TYPE_MAP = {
    'aadhaar_front': 'aadhar_front',
    'aadhaar_back': 'aadhar_back',
    'aadhar_front': 'aadhar_front',
    'aadhar_back': 'aadhar_back',
    'pan_card': 'pan_card',
    'passport_photo': 'passport_photo',
}
_PARTNER_KYC_REQUIRED = {'aadhar_front', 'aadhar_back', 'pan_card', 'passport_photo'}


class PartnerProfileUpdateRequest(BaseModel):
    email: Optional[str] = None
    whatsapp_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    # [DC-NAME-GENDER] Apr 2026 — split name fields
    name_title: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    gender: Optional[str] = None
    # [DC-PARTNER-CONTACTS-001] Partners can update their own contacts
    sales_contact_number: Optional[str] = None
    sales_contact_name: Optional[str] = None
    service_contact_number: Optional[str] = None
    service_contact_name: Optional[str] = None
    # [DC-PARTNER-MODULES-001] Partner can manage their own staff module access
    module_settings: Optional[Dict[str, Any]] = None


@router.put("/auth/profile")
def partner_update_profile(
    body: PartnerProfileUpdateRequest,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """DC Protocol Mar 2026: Official Partner updates own profile.
    Locked: partner_code, partner_name, phone. Editable: email, whatsapp, address fields."""
    if body.email is not None:
        dup = db.query(OfficialPartner).filter(
            OfficialPartner.email == body.email.strip(),
            OfficialPartner.id != partner.id
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail="Email already in use by another partner.")
        partner.email = body.email.strip() or None
    if body.whatsapp_number is not None:
        partner.whatsapp_number = body.whatsapp_number.strip() or None
    if body.address is not None:
        partner.address = body.address.strip() or None
    if body.city is not None:
        partner.city = body.city.strip() or None
    if body.state is not None:
        partner.state = body.state.strip() or None
    if body.pincode is not None:
        partner.pincode = body.pincode.strip() or None
    # [DC-NAME-GENDER] Save split name fields and rebuild display name
    def _setattr_partner(attr, val):
        try: setattr(partner, attr, val)
        except Exception: pass
    if body.name_title is not None: _setattr_partner('name_title', body.name_title.strip() or None)
    if body.first_name is not None: _setattr_partner('first_name', body.first_name.strip() or None)
    if body.last_name  is not None: _setattr_partner('last_name',  body.last_name.strip()  or None)
    if body.gender     is not None: _setattr_partner('gender',     body.gender.strip()      or None)
    _fn = getattr(partner, 'first_name', None) or ''
    _ln = getattr(partner, 'last_name',  None) or ''
    if _fn and _ln:
        _t = getattr(partner, 'name_title', None) or ''
        partner.partner_name = ' '.join(p for p in [_t, _fn, _ln] if p)
    # [DC-PARTNER-CONTACTS-001] Partner updates their own sales/service contacts
    if body.sales_contact_number is not None:
        partner.sales_contact_number = body.sales_contact_number.strip() or None
    if body.sales_contact_name is not None:
        partner.sales_contact_name = body.sales_contact_name.strip() or None
    if body.service_contact_number is not None:
        partner.service_contact_number = body.service_contact_number.strip() or None
    if body.service_contact_name is not None:
        partner.service_contact_name = body.service_contact_name.strip() or None
    # [DC-PARTNER-MODULES-001] Partner manages their own staff module access
    if body.module_settings is not None:
        _allowed_keys = {'walkins', 'leads', 'service', 'marketplace', 'stock', 'sales'}
        _clean = {}
        for _role in ('sales_staff', 'service_staff'):
            _role_data = body.module_settings.get(_role, {})
            if isinstance(_role_data, dict):
                _clean[_role] = {k: bool(v) for k, v in _role_data.items() if k in _allowed_keys}
        partner.module_settings = _clean
    partner.updated_at = get_indian_time()
    db.commit()
    db.refresh(partner)
    return {"success": True, "message": "Profile updated successfully"}


@router.get("/kyc/status")
def partner_kyc_status(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """DC Protocol Mar 2026: Get KYC documents and overall status for the logged-in Official Partner."""
    docs = db.query(KYCDocument).filter(KYCDocument.partner_id == partner.id).all()
    frontend_map = {'aadhar_front': 'aadhaar_front', 'aadhar_back': 'aadhaar_back',
                    'pan_card': 'pan_card', 'passport_photo': 'passport_photo'}
    return {
        "success": True,
        "kyc_status": partner.kyc_status or 'Not Submitted',
        "documents": [
            {
                "id": d.id,
                "document_type": frontend_map.get(d.document_type, d.document_type),
                "file_name": d.file_name,
                "status": d.status,
                "uploaded_at": d.uploaded_at.isoformat() if d.uploaded_at else None,
                "rejection_reason": d.rejection_reason
            }
            for d in docs
        ]
    }


@router.post("/kyc/upload")
async def partner_kyc_upload(
    document_type: str = Form(...),
    file: UploadFile = File(...),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """DC Protocol Mar 2026: Upload KYC document for Official Partner.
    Supports: aadhaar_front, aadhaar_back, pan_card, passport_photo.
    Blocks re-upload if document is already Approved."""
    if document_type not in _PARTNER_KYC_TYPE_MAP:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid document type. Must be one of: {', '.join(_PARTNER_KYC_TYPE_MAP.keys())}"
        )
    doc_type_db = _PARTNER_KYC_TYPE_MAP[document_type]

    existing = db.query(KYCDocument).filter(
        and_(KYCDocument.partner_id == partner.id, KYCDocument.document_type == doc_type_db)
    ).first()
    if existing:
        was_approved = existing.status == 'Approved'
        doc = existing
        from app.services.object_storage import storage_service
        for path in [existing.file_path, existing.compressed_path]:
            if path:
                try:
                    storage_service.delete_file(path)
                except Exception:
                    pass
        doc.status = 'Pending'
        doc.rejection_reason = None
        record_id = existing.id
        if was_approved:
            fresh_partner = db.query(OfficialPartner).filter(OfficialPartner.id == partner.id).first()
            if fresh_partner and fresh_partner.kyc_status == 'Approved':
                fresh_partner.kyc_status = 'Pending'
    else:
        doc = KYCDocument(
            partner_id=partner.id,
            document_type=doc_type_db,
            file_path='pending_upload',
            file_name='pending',
            original_filename=file.filename,
            file_size=1,
            mime_type='application/octet-stream',
            processing_status='pending',
            status='Pending',
            version=1,
            is_current_version=True
        )
        db.add(doc)
        db.flush()
        record_id = doc.id

    upload_result = await UniversalUploadService.handle_upload(
        file=file,
        table_name='kyc_document',
        record_id=record_id,
        uploaded_by_id=f"partner_{partner.id}",
        uploaded_by_type='partner',
        storage_dir='kyc_documents',
        db=db
    )

    doc.file_path = upload_result['file_path']
    doc.file_name = upload_result['file_name']
    doc.original_filename = upload_result['original_filename']
    doc.file_size = upload_result['file_size']
    doc.mime_type = upload_result['file_type']
    doc.processing_status = 'pending' if upload_result['needs_compression'] else 'completed'
    doc.uploaded_at = get_indian_time()

    db.commit()

    all_docs = db.query(KYCDocument).filter(KYCDocument.partner_id == partner.id).all()
    uploaded_types = {d.document_type for d in all_docs}
    if _PARTNER_KYC_REQUIRED.issubset(uploaded_types):
        fresh = db.query(OfficialPartner).filter(OfficialPartner.id == partner.id).first()
        if fresh and fresh.kyc_status not in ('Pending', 'Approved'):
            fresh.kyc_status = 'Pending'
            db.commit()

    return {
        "success": True,
        "message": f"{document_type.replace('_', ' ').title()} uploaded successfully",
        "status": "Pending verification"
    }


@router.get("/auth/my-partnership")
def partner_my_partnership(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    [DC-PARTNER-TERMS-001] May 2026: Partner views their own partnership terms —
    start/end date, security deposit, reminder days, and document submission status.
    Read-only from partner side; documents are staff-uploaded only.
    """
    from app.models.kyc_document import KYCDocument
    docs = db.query(KYCDocument).filter(KYCDocument.partner_id == partner.id).all()
    doc_map = {d.document_type: d for d in docs}
    frontend_type_map = {'aadhar_front': 'aadhaar_front', 'aadhar_back': 'aadhaar_back',
                         'pan_card': 'pan_card', 'passport_photo': 'passport_photo'}
    kyc_summary = {
        frontend_type_map.get(dt, dt): {
            "status": doc.status,
            "file_name": doc.file_name,
            "uploaded_at": doc.uploaded_at.isoformat() if doc.uploaded_at else None,
            "rejection_reason": doc.rejection_reason,
        }
        for dt, doc in doc_map.items()
    }

    import datetime as _dt
    days_to_expiry = None
    if partner.partner_end_date:
        today = _dt.date.today()
        days_to_expiry = (partner.partner_end_date - today).days

    return {
        "success": True,
        "partnership": {
            "partner_start_date": partner.partner_start_date.isoformat() if partner.partner_start_date else None,
            "partner_end_date":   partner.partner_end_date.isoformat()   if partner.partner_end_date   else None,
            "days_to_expiry":     days_to_expiry,
            "reminder_days_before": partner.reminder_days_before if partner.reminder_days_before is not None else 90,
            "security_deposit":   float(partner.security_deposit) if partner.security_deposit else 0,
            "agreement_document_path":   partner.agreement_document_path,
            "application_document_path": partner.application_document_path,
            "agreement_submitted":   bool(partner.agreement_document_path),
            "application_submitted": bool(partner.application_document_path),
        },
        "kyc_status": partner.kyc_status or 'Not Submitted',
        "kyc_documents": kyc_summary,
    }


@router.post("/logo")
async def partner_upload_logo(
    file: UploadFile = File(...),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """DC_PARTNER_LOGO_001: Partner uploads their own business logo.
    Stored at partner_logos/{partner_id}_{filename}. Used in PDF invoice header."""
    import os, uuid
    ALLOWED_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/webp"}
    content_type = file.content_type or ""
    if content_type not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Only PNG, JPG, or WebP images are allowed")
    MAX_SIZE = 2 * 1024 * 1024  # 2 MB
    file_bytes = await file.read()
    if len(file_bytes) > MAX_SIZE:
        raise HTTPException(status_code=400, detail="Logo file must be under 2 MB")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else "png"
    filename = f"{partner.id}_{uuid.uuid4().hex[:8]}.{ext}"
    logos_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
        "uploads", "partner_logos"
    )
    os.makedirs(logos_dir, exist_ok=True)
    full_path = os.path.join(logos_dir, filename)
    with open(full_path, "wb") as f:
        f.write(file_bytes)
    relative_path = f"partner_logos/{filename}"
    fresh = db.query(OfficialPartner).filter(OfficialPartner.id == partner.id).first()
    if fresh:
        fresh.logo_path = relative_path
        db.commit()
    return {"success": True, "logo_path": relative_path, "logo_url": f"/uploads/{relative_path}"}


@router.post("/admin/partners/{partner_id}/logo")
async def staff_update_partner_logo(
    partner_id: int,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """DC_PARTNER_LOGO_001: Staff updates a partner's business logo from the master config."""
    # Verify staff token
    auth_header = request.headers.get("Authorization", "") if request else ""
    token = auth_header.replace("Bearer ", "").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Staff authentication required")
    try:
        payload = SecurityManager.verify_token(token)
        if not payload or payload.get("type") != "staff_access":
            raise HTTPException(status_code=403, detail="Staff access required")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired staff token")
    import os, uuid
    target = db.query(OfficialPartner).filter(OfficialPartner.id == partner_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Partner not found")
    ALLOWED_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/webp"}
    if (file.content_type or "") not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Only PNG, JPG, or WebP images are allowed")
    file_bytes = await file.read()
    if len(file_bytes) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo file must be under 2 MB")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else "png"
    filename = f"{partner_id}_{uuid.uuid4().hex[:8]}.{ext}"
    logos_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
        "uploads", "partner_logos"
    )
    os.makedirs(logos_dir, exist_ok=True)
    with open(os.path.join(logos_dir, filename), "wb") as f:
        f.write(file_bytes)
    target.logo_path = f"partner_logos/{filename}"
    db.commit()
    return {"success": True, "logo_path": target.logo_path, "logo_url": f"/uploads/{target.logo_path}"}


def _ensure_partner_support_requests_table(db: Session):
    """[DC-PARTNER-CONTACTS-001] Ensure partner_support_requests table with all columns."""
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS partner_support_requests (
            id SERIAL PRIMARY KEY,
            partner_id INTEGER NOT NULL,
            partner_code VARCHAR(20),
            subject VARCHAR(200) NOT NULL,
            category VARCHAR(50) DEFAULT 'other',
            description TEXT NOT NULL,
            status VARCHAR(20) DEFAULT 'OPEN',
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))
    for col_sql in [
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS assign_to VARCHAR(20) DEFAULT 'self'",
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS service_dept_staff_id INTEGER",
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS company_support_requested BOOLEAN DEFAULT FALSE",
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS customer_name VARCHAR(200)",
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS customer_phone VARCHAR(20)",
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS product_name VARCHAR(200)",
        "ALTER TABLE partner_support_requests ADD COLUMN IF NOT EXISTS product_serial VARCHAR(100)",
    ]:
        try:
            db.execute(text(col_sql))
        except Exception:
            pass
    db.commit()


class PartnerSupportRequestIn(BaseModel):
    subject: str = Field(..., min_length=3, max_length=200)
    category: str = Field(default="other", max_length=50)
    description: str = Field(..., min_length=10, max_length=5000)
    # [DC-PARTNER-CONTACTS-001] Service ticket assignment
    assign_to: str = Field(default="self", description="'self' = handled by partner showroom, 'company' = escalate to company")
    service_dept_staff_id: Optional[int] = Field(None, description="ID of service dept staff to notify")
    customer_name: Optional[str] = Field(None, max_length=200)
    customer_phone: Optional[str] = Field(None, max_length=20)
    product_name: Optional[str] = Field(None, max_length=200)
    product_serial: Optional[str] = Field(None, max_length=100)


@router.post("/auth/support-request")
async def submit_partner_support_request(
    payload: PartnerSupportRequestIn,
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    DC_PARTNER_SUPPORT_001: Submit a support/service request from the partner portal.
    - assign_to='self'    → partner showroom handles it (default)
    - assign_to='company' → escalated to company service team
    - service_dept_staff_id → optional staff to notify
    """
    _ensure_partner_support_requests_table(db)

    _company_requested = payload.assign_to == "company"
    result = db.execute(text("""
        INSERT INTO partner_support_requests
            (partner_id, partner_code, subject, category, description, status,
             assign_to, company_support_requested, service_dept_staff_id,
             customer_name, customer_phone, product_name, product_serial,
             created_at, updated_at)
        VALUES
            (:pid, :code, :subject, :category, :description, 'OPEN',
             :assign_to, :company_req, :staff_id,
             :cname, :cphone, :pname, :pserial,
             NOW(), NOW())
        RETURNING id
    """), {
        "pid": partner.id,
        "code": partner.partner_code,
        "subject": payload.subject,
        "category": payload.category,
        "description": payload.description,
        "assign_to": payload.assign_to,
        "company_req": _company_requested,
        "staff_id": payload.service_dept_staff_id,
        "cname": payload.customer_name,
        "cphone": payload.customer_phone,
        "pname": payload.product_name,
        "pserial": payload.product_serial,
    })
    db.commit()
    _new_id = result.fetchone()[0]
    _ref = f"SR-{partner.partner_code}-{_new_id:05d}"

    _msg = (
        "Support request escalated to company team. We'll get back to you shortly."
        if _company_requested else
        f"Service request #{_ref} created for your showroom."
    )
    return {
        "success": True,
        "message": _msg,
        "ticket_reference": _ref,
        "id": _new_id,
        "assign_to": payload.assign_to,
        "company_support_requested": _company_requested,
    }


@router.get("/auth/support-requests")
async def list_partner_support_requests(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """[DC-PARTNER-CONTACTS-001] List service/support requests raised by this partner."""
    _ensure_partner_support_requests_table(db)
    rows = db.execute(text("""
        SELECT id, subject, category, description, status,
               assign_to, company_support_requested, service_dept_staff_id,
               customer_name, customer_phone, product_name, product_serial,
               created_at, updated_at
        FROM partner_support_requests
        WHERE partner_id = :pid
        ORDER BY created_at DESC
        LIMIT 100
    """), {"pid": partner.id}).fetchall()

    keys = ("id","subject","category","description","status","assign_to",
            "company_support_requested","service_dept_staff_id",
            "customer_name","customer_phone","product_name","product_serial",
            "created_at","updated_at")
    return {
        "success": True,
        "requests": [
            {k: (v.isoformat() if hasattr(v, 'isoformat') else v) for k, v in zip(keys, row)}
            for row in rows
        ]
    }


@router.get("/service-dept-staff")
async def list_service_dept_staff(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """[DC-PARTNER-CONTACTS-001] List staff from Sales/Service depts for ticket assignment."""
    try:
        rows = db.execute(text("""
            SELECT e.id, e.full_name, e.phone, e.emp_code, d.name AS dept_name
            FROM staff_employees e
            LEFT JOIN staff_departments d ON e.department_id = d.id
            WHERE e.status = 'active'
              AND e.is_deleted = FALSE
              AND LOWER(d.name) SIMILAR TO '%(sales|service)%'
            ORDER BY e.full_name
            LIMIT 100
        """)).fetchall()
        return {
            "success": True,
            "staff": [
                {"id": r[0], "full_name": r[1], "phone": r[2], "emp_code": r[3], "dept_name": r[4]}
                for r in rows
            ]
        }
    except Exception as e:
        return {"success": True, "staff": []}


@router.post("/auth/support-request/{request_id}/request-company-support")
async def request_company_support(
    request_id: int,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """
    [DC-PARTNER-CONTACTS-001] Partner escalates a self-handled request to company support.
    Even after self-assignment, partner can request company support at any time.
    """
    _ensure_partner_support_requests_table(db)
    db.execute(text("""
        UPDATE partner_support_requests
        SET company_support_requested = TRUE,
            assign_to = 'company',
            updated_at = NOW()
        WHERE id = :rid AND partner_id = :pid
    """), {"rid": request_id, "pid": partner.id})
    db.commit()
    return {"success": True, "message": "Company support has been requested. Our team will follow up shortly."}


@router.get("/auth/updated-leads")
async def get_partner_updated_leads(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("updated_at"),
    sort_dir: Optional[str] = Query("desc"),
):
    """
    DC_PARTNER_UPDATED_LEADS_001 (Apr 2026):
    Returns CRM leads where current partner is assigned as the partner handler
    (associated_partner_id = partner.id). These are leads staff has tagged this
    partner on for responsibility/follow-up.
    """
    _allowed_sort = {"updated_at", "created_at", "name", "status", "priority"}
    if sort_by not in _allowed_sort:
        sort_by = "updated_at"
    sort_dir_sql = "ASC" if (sort_dir or "").lower() == "asc" else "DESC"

    conditions = ["cl.associated_partner_id = :partner_id"]
    params: dict = {"partner_id": partner.id}

    if status:
        conditions.append("cl.status = :status")
        params["status"] = status
    if search:
        conditions.append("(cl.name ILIKE :search OR cl.phone ILIKE :search OR cl.email ILIKE :search)")
        params["search"] = f"%{search}%"

    where = " AND ".join(conditions)

    total = db.execute(
        text(f"SELECT COUNT(*) FROM crm_leads cl WHERE {where}"), params
    ).scalar() or 0

    params["limit"] = per_page
    params["offset"] = (page - 1) * per_page

    rows = db.execute(text(f"""
        SELECT
            cl.id, cl.name, cl.phone, cl.email, cl.status, cl.priority,
            cl.looking_for, cl.city, cl.company_id,
            cl.solar_pipeline_status, cl.ev_b2b_stage,
            cl.created_at, cl.updated_at,
            sc.name AS category_name,
            ac.company_name
        FROM crm_leads cl
        LEFT JOIN signup_categories sc ON sc.id = cl.category_id
        LEFT JOIN associated_companies ac ON ac.id = cl.company_id
        WHERE {where}
        ORDER BY cl.{sort_by} {sort_dir_sql} NULLS LAST
        LIMIT :limit OFFSET :offset
    """), params).fetchall()

    leads = [
        {
            "id": r.id,
            "name": r.name or "",
            "phone": r.phone or "",
            "email": r.email or "",
            "status": r.status or "new",
            "priority": r.priority or "medium",
            "looking_for": r.looking_for or "",
            "city": r.city or "",
            "company_id": r.company_id,
            "company_name": r.company_name or "",
            "category_name": r.category_name or "General",
            "solar_pipeline_status": r.solar_pipeline_status,
            "ev_b2b_stage": r.ev_b2b_stage,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        }
        for r in rows
    ]

    return {
        "success": True,
        "data": leads,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, (total + per_page - 1) // per_page),
    }


# ============================================================
# PARTNER WALK-INS + SALES INVOICES
# DC Protocol: partner_id from JWT, strict data isolation
# ============================================================

from app.models.staff_accounts import (
    SalesInvoice, SalesInvoiceLineItem, SalesInvoicePayment,
    SalesCouponMaster, HSNMaster, StockItemMaster
)
from app.models.crm import CRMLead


# ─────────────────────── Pydantic Schemas ───────────────────

class WalkinCreate(BaseModel):
    visit_date: Optional[str] = None
    visit_time: Optional[str] = None
    customer_name: str
    customer_phone: str
    alternate_phone: Optional[str] = None
    customer_email: Optional[str] = None
    customer_type: str = "WALK_IN"
    is_returning: bool = False
    previous_walkin_id: Optional[int] = None
    registered_with_mnr: bool = False
    registered_with_vgk: bool = False
    is_new_customer: bool = False
    is_existing_customer: bool = False
    is_marketplace_user: bool = False
    is_other_partner_customer: bool = False
    visit_purpose: str = "general"
    category_id: Optional[int] = None
    product_interest: Optional[str] = None
    lead_source: str = "walk_in"
    assigned_to: Optional[str] = None
    visit_outcome: str = "pending"
    follow_up_date: Optional[str] = None
    notes: Optional[str] = None
    company_follow_up: bool = False
    register_as_vgk: bool = True
    vgk_referral_id: Optional[int] = None
    # [DC-PHONE-OTP-001] Phone verification token (required when register_as_vgk=True)
    vgk_phone_verified_token: Optional[str] = None


class WalkinUpdate(BaseModel):
    visit_outcome: Optional[str] = None
    follow_up_date: Optional[str] = None
    notes: Optional[str] = None
    assigned_to: Optional[str] = None
    company_follow_up: Optional[bool] = None
    product_interest: Optional[str] = None
    visit_purpose: Optional[str] = None
    category_id: Optional[int] = None


class InvoiceLineItemIn(BaseModel):
    item_id: Optional[int] = None
    item_code: Optional[str] = None
    item_description: str
    hsn_id: Optional[int] = None
    hsn_code: Optional[str] = None
    quantity: float = 1
    unit_of_measure: str = "PCS"
    unit_rate: float = 0
    mrp: Optional[float] = None
    discount_percent: float = 0
    gst_rate: float = 0
    specification: Optional[str] = None
    color: Optional[str] = None
    warranty_months: int = 0


class SalesInvoiceCreate(BaseModel):
    document_type: str = "tax_invoice"
    invoice_date: str
    customer_type: str = "WALK_IN"
    customer_name: str
    customer_phone: str  # DC_MOBILE_MANDATORY_001: required for all invoices
    alternate_phone: Optional[str] = None
    customer_email: Optional[str] = None
    customer_gstin: Optional[str] = None
    customer_state: Optional[str] = None
    billing_address: Optional[str] = None
    shipping_address: Optional[str] = None
    is_igst: bool = False
    seller_state: Optional[str] = None
    buyer_state: Optional[str] = None
    line_items: List[InvoiceLineItemIn] = []
    terms_conditions: Optional[str] = None
    remarks: Optional[str] = None
    so_number: Optional[str] = None
    is_credit_sale: bool = False
    credit_days: int = 0
    due_date: Optional[str] = None
    payment_mode: Optional[str] = None
    linked_walkin_id: Optional[int] = None
    coupon_code: Optional[str] = None
    manual_discount_amount: float = 0
    manual_discount_note: Optional[str] = None


class PaymentRecordIn(BaseModel):
    amount: float
    payment_mode: str = "CASH"
    payment_date: Optional[str] = None
    reference_number: Optional[str] = None
    notes: Optional[str] = None


# ─────────────────────── Helpers ────────────────────────────

def _get_fy_year(ref_date: date) -> str:
    """Return financial year string e.g. '2526' for FY2025-26"""
    if ref_date.month >= 4:
        return f"{str(ref_date.year)[2:]}{str(ref_date.year + 1)[2:]}"
    return f"{str(ref_date.year - 1)[2:]}{str(ref_date.year)[2:]}"


def _next_partner_invoice_sequence(db: Session, partner_id: int, fy: str) -> int:
    row = db.execute(text(
        "SELECT COALESCE(MAX(fy_sequence),0) + 1 FROM sales_invoices "
        "WHERE partner_id = :pid AND invoice_number LIKE :fy"
    ), {"pid": partner_id, "fy": f"%/{fy}/%"}).fetchone()
    return row[0] if row else 1


def _build_partner_invoice_number(partner_code: str, db: Session, partner_id: int, ref_date: date, doc_type: str) -> tuple:
    fy = _get_fy_year(ref_date)
    seq = _next_partner_invoice_sequence(db, partner_id, fy)
    prefix = "EST" if doc_type in ("estimate", "proforma") else "INV"
    inv_num = f"{partner_code}/{prefix}/{fy}/{seq:03d}"
    return inv_num, seq


def _calc_line_item(item: InvoiceLineItemIn, is_igst: bool) -> dict:
    qty = Decimal(str(item.quantity))
    rate = Decimal(str(item.unit_rate))
    disc_pct = Decimal(str(item.discount_percent))
    gst_rate = Decimal(str(item.gst_rate))

    gross = qty * rate
    disc_amt = (gross * disc_pct / 100).quantize(Decimal("0.01"))
    taxable = (gross - disc_amt).quantize(Decimal("0.01"))

    half_rate = (gst_rate / 2).quantize(Decimal("0.01"))
    if is_igst:
        igst_amt = (taxable * gst_rate / 100).quantize(Decimal("0.01"))
        cgst_amt = sgst_amt = Decimal("0")
        igst_rate = gst_rate
        cgst_rate = sgst_rate = Decimal("0")
    else:
        cgst_amt = (taxable * half_rate / 100).quantize(Decimal("0.01"))
        sgst_amt = cgst_amt
        igst_amt = Decimal("0")
        igst_rate = Decimal("0")
        cgst_rate = sgst_rate = half_rate

    total_tax = (cgst_amt + sgst_amt + igst_amt).quantize(Decimal("0.01"))
    line_total = (taxable + total_tax).quantize(Decimal("0.01"))

    return {
        "gross_amount": float(gross),
        "discount_percent": float(disc_pct),
        "discount_amount": float(disc_amt),
        "taxable_amount": float(taxable),
        "gst_rate": float(gst_rate),
        "cgst_rate": float(cgst_rate),
        "cgst_amount": float(cgst_amt),
        "sgst_rate": float(sgst_rate),
        "sgst_amount": float(sgst_amt),
        "igst_rate": float(igst_rate),
        "igst_amount": float(igst_amt),
        "cess_rate": 0,
        "cess_amount": 0,
        "total_tax": float(total_tax),
        "line_total": float(line_total),
        "taxable_d": taxable,
        "total_tax_d": total_tax,
        "line_total_d": line_total,
    }


def _recalc_invoice_totals(db: Session, invoice: SalesInvoice):
    items = db.query(SalesInvoiceLineItem).filter_by(invoice_id=invoice.id).all()
    subtotal = sum(Decimal(str(it.gross_amount or 0)) for it in items)
    total_disc = sum(Decimal(str(it.discount_amount or 0)) for it in items)
    taxable = sum(Decimal(str(it.taxable_amount or 0)) for it in items)
    total_tax = sum(Decimal(str(it.total_tax or 0)) for it in items)
    cgst = sum(Decimal(str(it.cgst_amount or 0)) for it in items)
    sgst = sum(Decimal(str(it.sgst_amount or 0)) for it in items)
    igst = sum(Decimal(str(it.igst_amount or 0)) for it in items)

    coupon_disc = Decimal(str(invoice.coupon_discount_amount or 0))
    manual_disc = Decimal(str(invoice.manual_discount_amount or 0))
    gross_total = taxable + total_tax
    round_off = round(float(gross_total)) - float(gross_total)
    grand = Decimal(str(round(float(gross_total))))
    net_payable = (grand - coupon_disc - manual_disc).quantize(Decimal("0.01"))

    invoice.subtotal = float(subtotal)
    invoice.total_discount = float(total_disc)
    invoice.taxable_amount = float(taxable)
    invoice.cgst_amount = float(cgst)
    invoice.sgst_amount = float(sgst)
    invoice.igst_amount = float(igst)
    invoice.total_tax = float(total_tax)
    invoice.round_off = round_off
    invoice.grand_total = float(grand)
    invoice.net_payable = float(net_payable)
    invoice.balance_due = float(net_payable - Decimal(str(invoice.amount_received or 0)))


def _invoice_to_dict(inv: SalesInvoice, db: Session) -> dict:
    items = db.query(SalesInvoiceLineItem).filter_by(invoice_id=inv.id).all()
    payments = db.query(SalesInvoicePayment).filter_by(invoice_id=inv.id).all()
    return {
        "id": inv.id,
        "invoice_number": inv.invoice_number,
        "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
        "document_type": inv.document_type,
        "status": inv.status,
        "customer_type": inv.customer_type,
        "customer_name": inv.customer_name,
        "customer_phone": inv.customer_phone,
        "customer_email": inv.customer_email,
        "customer_gstin": inv.customer_gstin,
        "customer_state": inv.customer_state,
        "billing_address": inv.billing_address,
        "shipping_address": inv.shipping_address,
        "is_igst": inv.is_igst,
        "seller_state": inv.seller_state,
        "buyer_state": inv.buyer_state,
        "subtotal": float(inv.subtotal or 0),
        "total_discount": float(inv.total_discount or 0),
        "taxable_amount": float(inv.taxable_amount or 0),
        "cgst_amount": float(inv.cgst_amount or 0),
        "sgst_amount": float(inv.sgst_amount or 0),
        "igst_amount": float(inv.igst_amount or 0),
        "total_tax": float(inv.total_tax or 0),
        "round_off": float(inv.round_off or 0),
        "grand_total": float(inv.grand_total or 0),
        "coupon_code": inv.coupon_code,
        "coupon_discount_pct": float(inv.coupon_discount_pct or 0),
        "coupon_discount_amount": float(inv.coupon_discount_amount or 0),
        "manual_discount_amount": float(inv.manual_discount_amount or 0),
        "manual_discount_note": inv.manual_discount_note,
        "net_payable": float(inv.net_payable or 0),
        "payment_status": inv.payment_status,
        "payment_mode": inv.payment_mode,
        "amount_received": float(inv.amount_received or 0),
        "balance_due": float(inv.balance_due or 0),
        "is_credit_sale": inv.is_credit_sale,
        "credit_days": inv.credit_days,
        "due_date": inv.due_date.isoformat() if inv.due_date else None,
        "terms_conditions": inv.terms_conditions,
        "remarks": inv.remarks,
        "so_number": inv.so_number,
        "linked_walkin_id": inv.linked_walkin_id,
        "partner_id": inv.partner_id,
        "created_at": inv.created_at.isoformat() if inv.created_at else None,
        "updated_at": inv.updated_at.isoformat() if inv.updated_at else None,
        "line_items": [
            {
                "id": it.id,
                "line_number": it.line_number,
                "item_id": it.item_id,
                "item_code": it.item_code,
                "item_description": it.item_description,
                "hsn_code": it.hsn_code,
                "quantity": float(it.quantity or 1),
                "unit_of_measure": it.unit_of_measure,
                "unit_rate": float(it.unit_rate or 0),
                "mrp": float(it.mrp) if it.mrp else None,
                "gross_amount": float(it.gross_amount or 0),
                "discount_percent": float(it.discount_percent or 0),
                "discount_amount": float(it.discount_amount or 0),
                "taxable_amount": float(it.taxable_amount or 0),
                "gst_rate": float(it.gst_rate or 0),
                "cgst_rate": float(it.cgst_rate or 0),
                "cgst_amount": float(it.cgst_amount or 0),
                "sgst_rate": float(it.sgst_rate or 0),
                "sgst_amount": float(it.sgst_amount or 0),
                "igst_rate": float(it.igst_rate or 0),
                "igst_amount": float(it.igst_amount or 0),
                "total_tax": float(it.total_tax or 0),
                "line_total": float(it.line_total or 0),
                "specification": it.specification,
                "color": it.color,
                "warranty_months": it.warranty_months or 0,
            }
            for it in items
        ],
        "payments": [p.to_dict() for p in payments],
    }


# ─────────────────────── VGK Endpoints (Partner-facing) ─────

@router.get("/vgk/check-phone")
async def vgk_check_phone(
    phone: str = Query(...),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_WALKIN_VGK_001: Check if a phone is already a VGK member."""
    member = db.query(OfficialPartner).filter(
        OfficialPartner.phone == phone.strip(),
        OfficialPartner.category == 'VGK_TEAM'
    ).first()
    if member:
        return {"exists": True, "vgk_id": member.id, "partner_code": member.partner_code, "name": member.partner_name}
    return {"exists": False}


@router.get("/vgk/search-referral")
async def vgk_search_referral(
    q: str = Query(""),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_WALKIN_VGK_001: Search VGK members by code / name / mobile for referral/upline picker."""
    if not q.strip():
        return {"success": True, "data": []}
    term = f"%{q.strip()}%"
    members = db.query(OfficialPartner).filter(
        OfficialPartner.category == 'VGK_TEAM',
        or_(
            OfficialPartner.partner_code.ilike(term),
            OfficialPartner.partner_name.ilike(term),
            OfficialPartner.phone.ilike(term),
        )
    ).limit(10).all()
    return {
        "success": True,
        "data": [{"id": m.id, "partner_code": m.partner_code, "name": m.partner_name, "phone": m.phone, "role": m.vgk_role} for m in members]
    }


@router.get("/sales-team/search")
async def sales_team_search(
    q: str = Query(""),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC Protocol (Apr 2026): Search active sales staff by name or emp_code for walk-in Assigned To field.
    No company filter — returns all active staff across all companies."""
    from sqlalchemy import text as sq_text
    q = q.strip()
    if len(q) < 2:
        return {"success": True, "data": []}
    term = f"%{q}%"
    rows = db.execute(sq_text("""
        SELECT emp_code,
               COALESCE(NULLIF(TRIM(full_name), ''), emp_code) AS display_name,
               designation
        FROM staff_employees
        WHERE status = 'active'
          AND (LOWER(COALESCE(full_name,'')) LIKE LOWER(:term) OR LOWER(emp_code) LIKE LOWER(:term))
        ORDER BY display_name
        LIMIT 10
    """), {"term": term}).fetchall()
    return {
        "success": True,
        "data": [
            {"emp_code": r[0], "full_name": r[1], "designation": r[2]}
            for r in rows
        ]
    }


@router.post("/vgk/reset-password/{vgk_id}")
async def vgk_reset_password(
    vgk_id: int = Path(...),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_WALKIN_VGK_001: Reset a VGK member's password to their mobile number.
    Only allowed within 30 days of enrolment, and only by the enrolling partner."""
    from sqlalchemy import text as sq_text
    from datetime import datetime as _dtnow, timedelta

    _ensure_walkin_vgk_cols(db)

    row = db.execute(sq_text(
        "SELECT id, vgk_enrolled_at FROM partner_walkins WHERE partner_id = :pid AND vgk_member_id = :vid ORDER BY id DESC LIMIT 1"
    ), {"pid": partner.id, "vid": vgk_id}).fetchone()

    if not row:
        raise HTTPException(status_code=403, detail="Not authorized — this VGK member was not enrolled by you")

    enrolled_at = row[1]
    if enrolled_at and (_dtnow.now() - enrolled_at).days > 30:
        raise HTTPException(status_code=403, detail="Password reset window expired — only allowed within 30 days of enrolment")

    member = db.query(OfficialPartner).filter(
        OfficialPartner.id == vgk_id,
        OfficialPartner.category == 'VGK_TEAM'
    ).first()
    if not member:
        raise HTTPException(status_code=404, detail="VGK member not found")

    from app.core.security import SecurityManager as _SM2
    member.password_hash = _SM2.get_password_hash(member.phone)
    db.commit()

    return {"success": True, "message": f"Password reset to registered mobile number for {member.partner_code}"}


# ─────────────────────── My Team (Active Sales Team) ────────

@router.get("/my-team/active")
async def get_my_active_team(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """
    DC Protocol (Apr 2026): Returns VGK members enrolled via this partner's walk-ins.
    Used to populate the 'Assigned To' dropdown in the walk-in form.
    Deduped by vgk_member_id, filtered to is_active=True only.
    """
    from sqlalchemy import text as sq_text
    # DC_WALKIN_VGK_FIX_001 (Apr 2026): Removed is_active=TRUE filter — new members start inactive.
    # Show any VGK member enrolled via this partner's walk-ins, regardless of activation status.
    rows = db.execute(sq_text("""
        SELECT DISTINCT op.id, op.partner_code, op.partner_name, op.phone, op.vgk_role
        FROM partner_walkins pw
        JOIN official_partners op ON op.id = pw.vgk_member_id
        WHERE pw.partner_id = :pid
          AND op.category = 'VGK_TEAM'
          AND pw.vgk_member_id IS NOT NULL
        ORDER BY op.partner_name
    """), {"pid": partner.id}).fetchall()

    return {
        "success": True,
        "data": [
            {
                "id": r[0],
                "partner_code": r[1],
                "partner_name": r[2],
                "phone": r[3],
                "role": r[4],
            }
            for r in rows
        ]
    }


# ─────────────────────── Walk-ins Endpoints ─────────────────

@router.post("/walkins/send-otp")
async def walkin_vgk_send_otp(
    req: dict,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """[DC-PHONE-OTP-001] Send WhatsApp OTP to customer phone for walk-in VGK enrollment."""
    phone = (req.get("phone") or "").strip().replace(" ", "")
    if not phone or len(phone) < 10 or not phone.isdigit():
        raise HTTPException(status_code=400, detail="Please provide a valid 10-digit mobile number.")
    from app.utils.phone_otp import generate_and_send_otp
    return generate_and_send_otp(phone=phone, purpose='vgk_walkin', db=db)


@router.post("/walkins/verify-otp")
async def walkin_vgk_verify_otp(
    req: dict,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """[DC-PHONE-OTP-001] Verify OTP and issue phone_verified_token for walk-in VGK enrollment."""
    phone = (req.get("phone") or "").strip().replace(" ", "")
    otp_code = (req.get("otp_code") or "").strip()
    if not phone or not otp_code:
        raise HTTPException(status_code=400, detail="Phone and OTP code are required.")
    from app.utils.phone_otp import verify_otp_and_issue_token
    token = verify_otp_and_issue_token(phone=phone, otp_code=otp_code, purpose='vgk_walkin', db=db)
    return {"success": True, "phone_verified_token": token, "message": "Phone verified successfully."}


@router.post("/walkins")
async def create_walkin(
    data: WalkinCreate,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Record a new walk-in or returning customer visit."""
    from sqlalchemy import text as sq_text
    _ensure_walkin_vgk_cols(db)

    # [DC-PHONE-OTP-001] Validate VGK phone token when enrolling as VGK
    if data.register_as_vgk and data.customer_phone:
        if not data.vgk_phone_verified_token:
            raise HTTPException(
                status_code=400,
                detail="Phone verification required. Please send an OTP to the customer's WhatsApp and verify before enrolling them as a VGK Channel Partner."
            )
        from app.utils.phone_otp import validate_and_consume_token
        validate_and_consume_token(
            phone=data.customer_phone.strip(),
            token=data.vgk_phone_verified_token,
            purpose='vgk_walkin',
            db=db
        )

    # Validate/resolve category_id against partner's company
    company_id_for_cat = partner.company_id or 4
    resolved_category_id = _resolve_walkin_category_id(
        db=db,
        category_id=data.category_id,
        visit_purpose=data.visit_purpose,
        company_id=company_id_for_cat
    )
    data = data.model_copy(update={"category_id": resolved_category_id})

    vdate = date.today()
    if data.visit_date:
        try:
            vdate = date.fromisoformat(data.visit_date)
        except Exception:
            pass

    vtime = None
    if data.visit_time:
        try:
            from datetime import time as dtime
            parts = data.visit_time.split(":")
            vtime = dtime(int(parts[0]), int(parts[1]))
        except Exception:
            pass

    fdate = None
    if data.follow_up_date:
        try:
            fdate = date.fromisoformat(data.follow_up_date)
        except Exception:
            pass

    row = db.execute(sq_text("""
        INSERT INTO partner_walkins (
            partner_id, visit_date, visit_time,
            customer_name, customer_phone, alternate_phone, customer_email,
            customer_type, is_returning, previous_walkin_id,
            registered_with_mnr, registered_with_vgk, is_new_customer, is_existing_customer,
            is_marketplace_user, is_other_partner_customer,
            visit_purpose, category_id, product_interest, lead_source, assigned_to,
            visit_outcome, follow_up_date, notes, company_follow_up,
            created_at, updated_at
        ) VALUES (
            :pid, :vdate, :vtime,
            :cname, :cphone, :aphone, :cemail,
            :ctype, :returning, :prev_id,
            :mnr, :vgk, :new_cust, :existing, :marketplace, :other_partner,
            :purpose, :category_id, :interest, :lsource, :assigned,
            :outcome, :fdate, :notes, :company_fu,
            NOW(), NOW()
        ) RETURNING id
    """), {
        "pid": partner.id,
        "vdate": vdate, "vtime": vtime,
        "cname": data.customer_name, "cphone": data.customer_phone,
        "aphone": data.alternate_phone, "cemail": data.customer_email,
        "ctype": data.customer_type, "returning": data.is_returning,
        "prev_id": data.previous_walkin_id,
        "mnr": data.registered_with_mnr, "vgk": data.registered_with_vgk,
        "new_cust": data.is_new_customer,
        "existing": data.is_existing_customer,
        "marketplace": data.is_marketplace_user,
        "other_partner": data.is_other_partner_customer,
        "purpose": data.visit_purpose, "category_id": data.category_id,
        "interest": data.product_interest,
        "lsource": data.lead_source, "assigned": data.assigned_to,
        "outcome": data.visit_outcome, "fdate": fdate,
        "notes": data.notes, "company_fu": data.company_follow_up,
    })
    walkin_id = row.fetchone()[0]
    db.commit()

    # Push to CRM if company_follow_up is True OR outcome is interested/follow_up
    crm_lead_id = None
    _should_push_crm = data.company_follow_up or (data.visit_outcome in ("interested", "follow_up", "company_follow_up"))
    if _should_push_crm:
        try:
            tags_list = []
            if data.registered_with_mnr:
                tags_list.append("MNR Registered")
            if data.registered_with_vgk:
                tags_list.append("VGK Registered")
            if data.is_existing_customer:
                tags_list.append("Existing Customer")
            tags_list.append(f"Partner: {partner.partner_code}")
            tags_list.append("Walk-in Source")

            company_id = partner.company_id or 4

            # DC Protocol (Apr 2026): Map visit_purpose → signup_categories slug → category_id
            _PURPOSE_SLUG = {
                "solar":        "solar",
                "ev":           "ev-b2c",
                "ev_b2b":       "ev-b2b",
                "ev_spares":    "ev-spares",
                "insurance":    "insurance",
                "real_dreams":  "real-dreams",
                "etc_training": "etc-training",
            }
            _cat_slug = _PURPOSE_SLUG.get(data.visit_purpose or "")
            _category_id = None
            if _cat_slug:
                _cat_row = db.execute(sq_text(
                    "SELECT id FROM signup_categories WHERE company_id = :cid AND slug = :slug LIMIT 1"
                ), {"cid": company_id, "slug": _cat_slug}).fetchone()
                if _cat_row:
                    _category_id = _cat_row[0]

            new_lead = CRMLead(
                company_id=company_id,
                name=data.customer_name,
                phone=data.customer_phone,
                alternate_phone=data.alternate_phone,
                email=data.customer_email,
                category_id=_category_id,
                source="Walk-in",
                source_details=f"Recorded by partner {partner.partner_code} ({partner.partner_name}). Purpose: {data.visit_purpose}. Product interest: {data.product_interest or 'Not specified'}",
                source_ref_type="partner",
                source_ref_id=str(partner.id),
                source_ref_name=partner.partner_name,
                status="new",
                priority="medium",
                handler_type="unassigned",
                associated_partner_id=partner.id,
                looking_for=data.product_interest,
                description=data.notes,
                tags=", ".join(tags_list),
            )
            db.add(new_lead)
            db.flush()
            crm_lead_id = new_lead.id

            # Link CRM lead back to walkin
            db.execute(sq_text(
                "UPDATE partner_walkins SET crm_lead_id = :cid WHERE id = :wid"
            ), {"cid": crm_lead_id, "wid": walkin_id})
            db.commit()
        except Exception as e:
            db.rollback()

    # DC_WALKIN_VGK_001: VGK Enrolment — non-blocking, walk-in save is never rolled back
    vgk_result = None
    if data.register_as_vgk and data.customer_phone:
        try:
            _ensure_walkin_vgk_cols(db)

            # Determine upline: referral_id from form → partner's own VGK → root VGK
            upline_id = data.vgk_referral_id
            if not upline_id:
                partner_vgk = db.query(OfficialPartner).filter(
                    OfficialPartner.phone == partner.phone,
                    OfficialPartner.category == 'VGK_TEAM'
                ).first()
                if partner_vgk:
                    upline_id = partner_vgk.id
                else:
                    root_vgk = db.query(OfficialPartner).filter(
                        OfficialPartner.partner_code == 'VGK07102207'
                    ).first()
                    upline_id = root_vgk.id if root_vgk else None

            vgk_id, was_created, vgk_msg = _get_or_create_vgk_member(
                db=db,
                phone=data.customer_phone,
                name=data.customer_name,
                company_id=partner.company_id or 4,
                parent_partner_id=upline_id,
                partner_ref_id=partner.id
            )

            db.execute(sq_text(
                "UPDATE partner_walkins SET vgk_member_id = :vid, vgk_enrolled_at = NOW() WHERE id = :wid"
            ), {"vid": vgk_id, "wid": walkin_id})
            db.commit()
            # DC Protocol (Apr 2026): Fetch full member data for share modal
            _vgk_member = db.query(OfficialPartner).filter(OfficialPartner.id == vgk_id).first()
            _raw_phone = (_vgk_member.phone if _vgk_member else data.customer_phone) or ""
            vgk_result = {
                "vgk_member_id": vgk_id,
                "was_created": was_created,
                "message": vgk_msg,
                "partner_code": _vgk_member.partner_code if _vgk_member else None,
                "partner_name": _vgk_member.partner_name if _vgk_member else data.customer_name,
                "phone": _raw_phone,
                # auto_password: phone digits only — shown in share modal for new members only
                "auto_password": _raw_phone.replace(" ", "").replace("-", "") if (was_created and _vgk_member) else None,
                # points_balance: fetch live from DB after points entry was committed
                "points_balance": int(_vgk_member.vgk_points_balance or 0) if _vgk_member else (15000 if was_created else 0),
                # advisor info: the partner who recorded this walk-in (shown in share message sign-off)
                "advisor_name": partner.partner_name,
                "advisor_code": partner.partner_code,
            }
        except Exception as _ve:
            try:
                db.rollback()
            except Exception:
                pass
            vgk_result = {"error": str(_ve), "message": "VGK enrolment failed — walk-in saved successfully"}

    return {
        "success": True,
        "id": walkin_id,
        "crm_lead_id": crm_lead_id,
        "vgk": vgk_result,
        "message": "Walk-in recorded successfully" + (" and pushed to CRM" if crm_lead_id else ""),
    }


@router.get("/walkins/summary")
async def walkin_summary(
    date_filter: str = Query("mtd", description="ftd|yesterday|mtd|fty|custom"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Summary stats for walk-ins + estimations for the partner."""
    today = date.today()
    if date_filter == "ftd":
        d_from = d_to = today
    elif date_filter == "yesterday":
        d_from = d_to = today - timedelta(days=1)
    elif date_filter == "mtd":
        d_from = today.replace(day=1)
        d_to = today
    elif date_filter == "fty":
        if today.month >= 4:
            d_from = date(today.year, 4, 1)
        else:
            d_from = date(today.year - 1, 4, 1)
        d_to = today
    else:
        d_from = date.fromisoformat(date_from) if date_from else today.replace(day=1)
        d_to = date.fromisoformat(date_to) if date_to else today

    from sqlalchemy import text as sq_text
    stats = db.execute(sq_text("""
        SELECT
            COUNT(*) FILTER (WHERE visit_date BETWEEN :df AND :dt) as total_walkins,
            COUNT(*) FILTER (WHERE visit_date = :today) as today_walkins,
            COUNT(*) FILTER (WHERE visit_date BETWEEN :df AND :dt AND company_follow_up = true) as company_followups,
            COUNT(*) FILTER (WHERE visit_date BETWEEN :df AND :dt AND visit_outcome = 'converted') as converted,
            COUNT(*) FILTER (WHERE visit_date BETWEEN :df AND :dt AND is_returning = true) as returning_customers
        FROM partner_walkins WHERE partner_id = :pid
    """), {"pid": partner.id, "df": d_from, "dt": d_to, "today": today}).fetchone()

    inv_stats = db.execute(sq_text("""
        SELECT
            COUNT(*) FILTER (WHERE status = 'DRAFT' AND document_type IN ('estimate','proforma')) as estimations_draft,
            COUNT(*) FILTER (WHERE status = 'CONFIRMED' AND document_type IN ('estimate','proforma')) as estimations_confirmed,
            COALESCE(SUM(grand_total) FILTER (WHERE status = 'CONFIRMED' AND document_type IN ('estimate','proforma')), 0) as estimations_value,
            COUNT(*) FILTER (WHERE status = 'CONFIRMED' AND document_type = 'tax_invoice') as invoices_confirmed,
            COALESCE(SUM(grand_total) FILTER (WHERE status = 'CONFIRMED' AND document_type = 'tax_invoice'), 0) as invoices_value
        FROM sales_invoices
        WHERE partner_id = :pid
          AND invoice_date BETWEEN :df AND :dt
    """), {"pid": partner.id, "df": d_from, "dt": d_to}).fetchone()

    return {
        "success": True,
        "period": {"from": d_from.isoformat(), "to": d_to.isoformat(), "filter": date_filter},
        "walkins": {
            "total": int(stats[0] or 0),
            "today": int(stats[1] or 0),
            "company_followups": int(stats[2] or 0),
            "converted": int(stats[3] or 0),
            "returning": int(stats[4] or 0),
        },
        "estimations": {
            "draft": int(inv_stats[0] or 0),
            "confirmed": int(inv_stats[1] or 0),
            "confirmed_value": float(inv_stats[2] or 0),
        },
        "invoices": {
            "confirmed": int(inv_stats[3] or 0),
            "confirmed_value": float(inv_stats[4] or 0),
        },
    }


@router.get("/walkins")
async def list_walkins(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    date_filter: str = Query("mtd"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    purpose: Optional[str] = Query(None),
    outcome: Optional[str] = Query(None),
    customer_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    today = date.today()
    if date_filter == "ftd":
        d_from = d_to = today
    elif date_filter == "yesterday":
        d_from = d_to = today - timedelta(days=1)
    elif date_filter == "mtd":
        d_from = today.replace(day=1); d_to = today
    elif date_filter == "fty":
        d_from = date(today.year if today.month >= 4 else today.year - 1, 4, 1)
        d_to = today
    else:
        d_from = date.fromisoformat(date_from) if date_from else today.replace(day=1)
        d_to = date.fromisoformat(date_to) if date_to else today

    from sqlalchemy import text as sq_text
    conditions = "partner_id = :pid AND visit_date BETWEEN :df AND :dt"
    params: dict = {"pid": partner.id, "df": d_from, "dt": d_to}

    if purpose:
        conditions += " AND visit_purpose = :purpose"
        params["purpose"] = purpose
    if outcome:
        conditions += " AND visit_outcome = :outcome"
        params["outcome"] = outcome
    if customer_type:
        conditions += " AND customer_type = :ctype"
        params["ctype"] = customer_type
    if search:
        conditions += " AND (customer_name ILIKE :s OR customer_phone ILIKE :s)"
        params["s"] = f"%{search}%"

    total = db.execute(sq_text(f"SELECT COUNT(*) FROM partner_walkins WHERE {conditions}"), params).scalar()
    offset = (page - 1) * per_page
    params["limit"] = per_page
    params["offset"] = offset

    _ensure_walkin_vgk_cols(db)
    rows = db.execute(sq_text(f"""
        SELECT id, visit_date, visit_time, customer_name, customer_phone,
               alternate_phone, customer_email, customer_type, is_returning,
               registered_with_mnr, registered_with_vgk, is_new_customer, is_existing_customer,
               visit_purpose, product_interest, visit_outcome, follow_up_date,
               assigned_to, notes, company_follow_up, crm_lead_id, status,
               created_at, vgk_member_id, vgk_enrolled_at
        FROM partner_walkins WHERE {conditions}
        ORDER BY visit_date DESC, visit_time DESC
        LIMIT :limit OFFSET :offset
    """), params).fetchall()

    cols = ["id","visit_date","visit_time","customer_name","customer_phone",
            "alternate_phone","customer_email","customer_type","is_returning",
            "registered_with_mnr","registered_with_vgk","is_new_customer","is_existing_customer",
            "visit_purpose","product_interest","visit_outcome","follow_up_date",
            "assigned_to","notes","company_follow_up","crm_lead_id","status","created_at",
            "vgk_member_id","vgk_enrolled_at"]

    data = []
    for r in rows:
        d = dict(zip(cols, r))
        for k in ["visit_date","follow_up_date"]:
            if d.get(k) and hasattr(d[k], "isoformat"):
                d[k] = d[k].isoformat()
        if d.get("visit_time") and hasattr(d["visit_time"], "isoformat"):
            d["visit_time"] = str(d["visit_time"])
        for k in ["created_at","vgk_enrolled_at"]:
            if d.get(k) and hasattr(d[k], "isoformat"):
                d[k] = d[k].isoformat()
        # compute days_since_enrolment for 30-day reset window
        if d.get("vgk_enrolled_at"):
            from datetime import datetime as _dt2
            try:
                _enr = _dt2.fromisoformat(d["vgk_enrolled_at"]) if isinstance(d["vgk_enrolled_at"], str) else d["vgk_enrolled_at"]
                d["vgk_reset_allowed"] = (_dt2.now() - _enr).days <= 30
            except Exception:
                d["vgk_reset_allowed"] = False
        else:
            d["vgk_reset_allowed"] = False
        data.append(d)

    return {
        "success": True,
        "data": data,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, (total + per_page - 1) // per_page),
    }


@router.get("/walkins/{walkin_id}")
async def get_walkin(
    walkin_id: int,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    from sqlalchemy import text as sq_text
    row = db.execute(sq_text(
        "SELECT * FROM partner_walkins WHERE id = :wid AND partner_id = :pid"
    ), {"wid": walkin_id, "pid": partner.id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Walk-in not found")
    keys = ["id","partner_id","visit_date","visit_time","customer_name","customer_phone",
             "alternate_phone","customer_email","customer_type","is_returning","previous_walkin_id",
             "registered_with_mnr","registered_with_vgk","is_new_customer","is_existing_customer","is_marketplace_user",
             "is_other_partner_customer","visit_purpose","product_interest","lead_source","assigned_to",
             "visit_outcome","follow_up_date","notes","company_follow_up","crm_lead_id","status","created_at","updated_at",
             "vgk_member_id","vgk_enrolled_at","category_id"]
    row_map = row._mapping if hasattr(row, '_mapping') else {}
    d = {}
    for k in keys:
        try:
            v = row_map.get(k)
            if hasattr(v, "isoformat"):
                v = v.isoformat()
            d[k] = v
        except Exception:
            pass
    return {"success": True, "data": d}


@router.put("/walkins/{walkin_id}")
async def update_walkin(
    walkin_id: int,
    data: WalkinUpdate,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    from sqlalchemy import text as sq_text
    _ensure_walkin_vgk_cols(db)
    existing = db.execute(sq_text(
        "SELECT id, company_follow_up, crm_lead_id FROM partner_walkins WHERE id = :wid AND partner_id = :pid"
    ), {"wid": walkin_id, "pid": partner.id}).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Walk-in not found")

    sets = []
    params: dict = {"wid": walkin_id}
    if data.visit_outcome is not None:
        sets.append("visit_outcome = :outcome"); params["outcome"] = data.visit_outcome
    if data.follow_up_date is not None:
        try:
            params["fdate"] = date.fromisoformat(data.follow_up_date)
        except Exception:
            params["fdate"] = None
        sets.append("follow_up_date = :fdate")
    if data.notes is not None:
        sets.append("notes = :notes"); params["notes"] = data.notes
    if data.assigned_to is not None:
        sets.append("assigned_to = :assigned"); params["assigned"] = data.assigned_to
    if data.product_interest is not None:
        sets.append("product_interest = :interest"); params["interest"] = data.product_interest
    if data.visit_purpose is not None:
        sets.append("visit_purpose = :purpose"); params["purpose"] = data.visit_purpose

    if data.category_id is not None:
        _validated_cat_id = _resolve_walkin_category_id(
            db=db,
            category_id=data.category_id,
            visit_purpose=data.visit_purpose or "",
            company_id=partner.company_id or 4
        )
        if _validated_cat_id is not None:
            sets.append("category_id = :category_id"); params["category_id"] = _validated_cat_id
            data = data.model_copy(update={"category_id": _validated_cat_id})

    if data.company_follow_up is not None and data.company_follow_up != existing[1]:
        sets.append("company_follow_up = :cfu"); params["cfu"] = data.company_follow_up

    # Push to CRM if: company_follow_up just set to True, OR outcome changed to interested/follow_up and no CRM lead yet
    _new_outcome = data.visit_outcome if data.visit_outcome is not None else None
    _push_crm_update = (
        (data.company_follow_up and not existing[2]) or
        (_new_outcome in ("interested", "follow_up", "company_follow_up") and not existing[2])
    )
    if _push_crm_update:
        try:
            wlk_row = db.execute(sq_text(
                "SELECT customer_name, customer_phone, product_interest, notes, category_id FROM partner_walkins WHERE id=:wid"
            ), {"wid": walkin_id}).fetchone()
            company_id = partner.company_id or 4
            outcome_label = _new_outcome or "interested"
            _cat_id = data.category_id if data.category_id is not None else (wlk_row[4] if wlk_row else None)
            new_lead = CRMLead(
                company_id=company_id,
                name=wlk_row[0] if wlk_row else "Unknown",
                phone=wlk_row[1] if wlk_row else None,
                source="Walk-in",
                source_details=f"Recorded by partner {partner.partner_code} ({partner.partner_name}). Outcome: {outcome_label}. Product interest: {wlk_row[2] if wlk_row and wlk_row[2] else 'Not specified'}",
                source_ref_type="partner",
                source_ref_id=str(partner.id),
                source_ref_name=partner.partner_name,
                status="new",
                handler_type="unassigned",
                associated_partner_id=partner.id,
                category_id=_cat_id,
                looking_for=wlk_row[2] if wlk_row else None,
                description=wlk_row[3] if wlk_row else None,
                tags=f"Partner: {partner.partner_code}, Walk-in Source, Outcome: {outcome_label}",
            )
            db.add(new_lead)
            db.flush()
            sets.append("crm_lead_id = :cid"); params["cid"] = new_lead.id
        except Exception:
            db.rollback()

    if sets:
        sets.append("updated_at = NOW()")
        db.execute(sq_text(f"UPDATE partner_walkins SET {', '.join(sets)} WHERE id = :wid"), params)
        db.commit()

    return {"success": True, "message": "Walk-in updated"}


# ─────────────── Sales Invoice Endpoints ────────────────────

@router.post("/sales-invoices")
async def create_partner_invoice(
    data: SalesInvoiceCreate,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Create a new partner sales invoice (Draft)."""
    try:
        inv_date = date.fromisoformat(data.invoice_date)
    except Exception:
        inv_date = date.today()

    inv_num, seq = _build_partner_invoice_number(
        partner.partner_code, db, partner.id, inv_date, data.document_type
    )

    # Fetch default terms from staff invoice settings
    terms = data.terms_conditions
    if not terms:
        try:
            from sqlalchemy import text as sq_text
            row = db.execute(sq_text(
                "SELECT terms_conditions FROM sales_invoices WHERE created_by_type='STAFF' AND terms_conditions IS NOT NULL ORDER BY id DESC LIMIT 1"
            )).scalar()
            terms = row
        except Exception:
            terms = None

    due_date_val = None
    if data.due_date:
        try:
            due_date_val = date.fromisoformat(data.due_date)
        except Exception:
            pass

    inv = SalesInvoice(
        invoice_number=inv_num,
        invoice_date=inv_date,
        company_id=partner.company_id or 4,
        document_type=data.document_type,
        status="DRAFT",
        customer_type=data.customer_type,
        customer_name=data.customer_name,
        customer_phone=data.customer_phone,
        customer_email=data.customer_email,
        customer_gstin=data.customer_gstin,
        customer_state=data.customer_state,
        billing_address=data.billing_address,
        shipping_address=data.shipping_address,
        is_igst=data.is_igst,
        seller_state=data.seller_state or partner.state,
        buyer_state=data.buyer_state or data.customer_state,
        terms_conditions=terms,
        remarks=data.remarks,
        so_number=data.so_number,
        is_credit_sale=data.is_credit_sale,
        credit_days=data.credit_days,
        due_date=due_date_val,
        payment_mode=data.payment_mode,
        payment_status="PENDING",
        amount_received=0,
        balance_due=0,
        manual_discount_amount=data.manual_discount_amount or 0,
        manual_discount_note=data.manual_discount_note,
        partner_id=partner.id,
        created_by_type="PARTNER",
        linked_walkin_id=data.linked_walkin_id,
        fy_sequence=seq,
        subtotal=0,
        total_discount=0,
        taxable_amount=0,
        cgst_amount=0,
        sgst_amount=0,
        igst_amount=0,
        total_tax=0,
        round_off=0,
        grand_total=0,
        net_payable=0,
    )
    db.add(inv)
    db.flush()

    # Insert line items
    for i, item in enumerate(data.line_items, 1):
        calc = _calc_line_item(item, data.is_igst)
        li = SalesInvoiceLineItem(
            invoice_id=inv.id,
            line_number=i,
            item_id=item.item_id,
            item_code=item.item_code,
            item_description=item.item_description,
            hsn_id=item.hsn_id,
            hsn_code=item.hsn_code,
            quantity=item.quantity,
            unit_of_measure=item.unit_of_measure,
            unit_rate=item.unit_rate,
            mrp=item.mrp,
            gross_amount=calc["gross_amount"],
            discount_percent=calc["discount_percent"],
            discount_amount=calc["discount_amount"],
            taxable_amount=calc["taxable_amount"],
            gst_rate=calc["gst_rate"],
            cgst_rate=calc["cgst_rate"],
            cgst_amount=calc["cgst_amount"],
            sgst_rate=calc["sgst_rate"],
            sgst_amount=calc["sgst_amount"],
            igst_rate=calc["igst_rate"],
            igst_amount=calc["igst_amount"],
            cess_rate=0,
            cess_amount=0,
            total_tax=calc["total_tax"],
            line_total=calc["line_total"],
            specification=item.specification,
            color=item.color,
            warranty_months=item.warranty_months,
        )
        db.add(li)

    db.flush()

    # Apply coupon if provided
    if data.coupon_code:
        coupon = db.query(SalesCouponMaster).filter(
            SalesCouponMaster.coupon_code == data.coupon_code,
            SalesCouponMaster.is_active == True,
        ).first()
        if coupon:
            inv.coupon_code = coupon.coupon_code
            inv.coupon_discount_pct = float(coupon.discount_percentage)

    _recalc_invoice_totals(db, inv)
    if inv.coupon_code and inv.coupon_discount_pct:
        inv.coupon_discount_amount = round(float(inv.taxable_amount) * float(inv.coupon_discount_pct) / 100, 2)
        _recalc_invoice_totals(db, inv)

    db.commit()
    return {"success": True, "id": inv.id, "invoice_number": inv.invoice_number, "data": _invoice_to_dict(inv, db)}


@router.get("/sales-invoices")
async def list_partner_invoices(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    date_filter: str = Query("mtd"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    payment_status: Optional[str] = Query(None),
    doc_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    today = date.today()
    if date_filter == "ftd":
        d_from = d_to = today
    elif date_filter == "yesterday":
        d_from = d_to = today - timedelta(days=1)
    elif date_filter == "mtd":
        d_from = today.replace(day=1); d_to = today
    elif date_filter == "fty":
        d_from = date(today.year if today.month >= 4 else today.year - 1, 4, 1)
        d_to = today
    else:
        d_from = date.fromisoformat(date_from) if date_from else today.replace(day=1)
        d_to = date.fromisoformat(date_to) if date_to else today

    query = db.query(SalesInvoice).filter(
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.invoice_date >= d_from,
        SalesInvoice.invoice_date <= d_to,
    )
    if status:
        query = query.filter(SalesInvoice.status == status.upper())
    if payment_status:
        query = query.filter(SalesInvoice.payment_status == payment_status.upper())
    if doc_type:
        query = query.filter(SalesInvoice.document_type == doc_type)
    if search:
        query = query.filter(or_(
            SalesInvoice.customer_name.ilike(f"%{search}%"),
            SalesInvoice.invoice_number.ilike(f"%{search}%"),
        ))

    total = query.count()
    invoices = query.order_by(SalesInvoice.invoice_date.desc(), SalesInvoice.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    # Summary stats
    all_q = db.query(SalesInvoice).filter(
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.invoice_date >= d_from,
        SalesInvoice.invoice_date <= d_to,
    )
    total_val = sum(float(i.grand_total or 0) for i in all_q.all())
    confirmed = all_q.filter(SalesInvoice.status == "CONFIRMED").count()
    draft_cnt = db.query(SalesInvoice).filter(SalesInvoice.partner_id == partner.id, SalesInvoice.status == "DRAFT").count()
    pending_pay = sum(float(i.balance_due or 0) for i in all_q.filter(SalesInvoice.payment_status != "PAID").all())

    return {
        "success": True,
        "data": [_invoice_to_dict(inv, db) for inv in invoices],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, (total + per_page - 1) // per_page),
        "summary": {
            "total_value": total_val,
            "confirmed": confirmed,
            "draft": draft_cnt,
            "pending_payment": pending_pay,
        },
    }


@router.get("/sales-invoices/export/gst")
async def export_partner_gst(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    date_filter: str = Query("mtd"),
    columns: Optional[str] = Query(None, description="Comma-separated column names to include"),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Export partner's sales invoices as GST-compliant Excel."""
    import io
    today = date.today()
    if date_filter == "ftd":
        d_from = d_to = today
    elif date_filter == "fty":
        d_from = date(today.year if today.month >= 4 else today.year - 1, 4, 1)
        d_to = today
    elif date_filter == "mtd":
        d_from = today.replace(day=1); d_to = today
    else:
        d_from = date.fromisoformat(date_from) if date_from else today.replace(day=1)
        d_to = date.fromisoformat(date_to) if date_to else today

    invoices = db.query(SalesInvoice).filter(
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "CONFIRMED",
        SalesInvoice.invoice_date >= d_from,
        SalesInvoice.invoice_date <= d_to,
    ).order_by(SalesInvoice.invoice_date).all()

    # Build rows
    all_cols = {
        "Invoice No": lambda i: i.invoice_number,
        "Invoice Date": lambda i: i.invoice_date.isoformat() if i.invoice_date else "",
        "Customer Name": lambda i: i.customer_name,
        "Customer GSTIN": lambda i: i.customer_gstin or "",
        "Customer State": lambda i: i.customer_state or "",
        "Taxable Value": lambda i: float(i.taxable_amount or 0),
        "CGST": lambda i: float(i.cgst_amount or 0),
        "SGST": lambda i: float(i.sgst_amount or 0),
        "IGST": lambda i: float(i.igst_amount or 0),
        "Total Tax": lambda i: float(i.total_tax or 0),
        "Grand Total": lambda i: float(i.grand_total or 0),
        "Payment Mode": lambda i: i.payment_mode or "",
        "Payment Status": lambda i: i.payment_status or "",
    }

    selected = list(all_cols.keys())
    if columns:
        requested = [c.strip() for c in columns.split(",")]
        selected = [c for c in requested if c in all_cols]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GST Report"

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")

        for col_i, col_name in enumerate(selected, 1):
            cell = ws.cell(row=1, column=col_i, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for row_i, inv in enumerate(invoices, 2):
            for col_i, col_name in enumerate(selected, 1):
                ws.cell(row=row_i, column=col_i, value=all_cols[col_name](inv))

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"GST_Export_{partner.partner_code}_{d_from}_{d_to}.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        # Fallback to CSV
        import csv
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=selected)
        writer.writeheader()
        for inv in invoices:
            writer.writerow({col: all_cols[col](inv) for col in selected})
        filename = f"GST_Export_{partner.partner_code}_{d_from}_{d_to}.csv"
        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@router.get("/sales-invoices/{invoice_id}")
async def get_partner_invoice(
    invoice_id: int,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"success": True, "data": _invoice_to_dict(inv, db)}


@router.put("/sales-invoices/{invoice_id}/line-items")
async def update_partner_invoice_line_items(
    invoice_id: int,
    body: dict,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "DRAFT",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or not editable")

    items_data = body.get("line_items", [])
    # Delete existing items
    db.query(SalesInvoiceLineItem).filter_by(invoice_id=inv.id).delete()
    db.flush()

    for i, item_d in enumerate(items_data, 1):
        item = InvoiceLineItemIn(**item_d)
        calc = _calc_line_item(item, inv.is_igst)
        li = SalesInvoiceLineItem(
            invoice_id=inv.id,
            line_number=i,
            item_id=item.item_id,
            item_code=item.item_code,
            item_description=item.item_description,
            hsn_id=item.hsn_id,
            hsn_code=item.hsn_code,
            quantity=item.quantity,
            unit_of_measure=item.unit_of_measure,
            unit_rate=item.unit_rate,
            mrp=item.mrp,
            gross_amount=calc["gross_amount"],
            discount_percent=calc["discount_percent"],
            discount_amount=calc["discount_amount"],
            taxable_amount=calc["taxable_amount"],
            gst_rate=calc["gst_rate"],
            cgst_rate=calc["cgst_rate"],
            cgst_amount=calc["cgst_amount"],
            sgst_rate=calc["sgst_rate"],
            sgst_amount=calc["sgst_amount"],
            igst_rate=calc["igst_rate"],
            igst_amount=calc["igst_amount"],
            cess_rate=0,
            cess_amount=0,
            total_tax=calc["total_tax"],
            line_total=calc["line_total"],
            specification=item.specification,
            color=item.color,
            warranty_months=item.warranty_months,
        )
        db.add(li)

    db.flush()
    _recalc_invoice_totals(db, inv)
    db.commit()
    return {"success": True, "data": _invoice_to_dict(inv, db)}


@router.post("/sales-invoices/{invoice_id}/confirm")
async def confirm_partner_invoice(
    invoice_id: int,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "DRAFT",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or already confirmed")
    if not db.query(SalesInvoiceLineItem).filter_by(invoice_id=inv.id).first():
        raise HTTPException(status_code=400, detail="Cannot confirm invoice with no line items")

    inv.status = "CONFIRMED"
    if inv.document_type in ("estimate", "proforma"):
        inv.document_type = "tax_invoice"

    # DC_PARTNER_STOCK_AUTOSYNC_001: deduct from partner stock on invoice confirm
    try:
        from app.services.partner_stock_service import auto_partner_stock_sync
        _line_items = db.query(SalesInvoiceLineItem).filter_by(invoice_id=inv.id).all()
        _sync_items = [
            {
                "item_name": li.item_description,
                "item_code": li.item_code or "",
                "stock_item_id": li.item_id,
                "qty": float(li.quantity),
                "unit_of_measure": li.unit_of_measure or "PCS",
                "selling_price": float(li.unit_rate) if li.unit_rate else None,
                "hsn_code": li.hsn_code or "",
            }
            for li in _line_items
        ]
        auto_partner_stock_sync(
            db=db,
            partner_id=partner.id,
            items=_sync_items,
            adj_type="SALE_OUT",
            ref_doc_type="PARTNER_INVOICE",
            ref_doc_id=inv.id,
            ref_doc_number=inv.invoice_number,
            reason=f"Auto: Partner sale via invoice {inv.invoice_number}",
            created_by=f"partner:{partner.partner_code}",
        )
    except Exception as _e:
        import logging
        logging.getLogger(__name__).warning(f"[AUTO_STOCK_SYNC] Partner invoice hook skipped: {_e}")

    db.commit()
    return {"success": True, "message": "Invoice confirmed", "invoice_number": inv.invoice_number}


@router.post("/sales-invoices/{invoice_id}/cancel")
async def cancel_partner_invoice(
    invoice_id: int,
    body: dict = {},
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
    ).filter(SalesInvoice.status.in_(["DRAFT", "CONFIRMED"])).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or not cancellable")
    inv.status = "CANCELLED"
    inv.cancellation_reason = body.get("reason", "Cancelled by partner")
    inv.cancelled_at = datetime.utcnow()
    db.commit()
    return {"success": True, "message": "Invoice cancelled"}


@router.post("/sales-invoices/{invoice_id}/payment")
async def record_partner_payment(
    invoice_id: int,
    data: PaymentRecordIn,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "CONFIRMED",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or not confirmed")

    pay_date = date.today()
    if data.payment_date:
        try:
            pay_date = date.fromisoformat(data.payment_date)
        except Exception:
            pass

    payment = SalesInvoicePayment(
        invoice_id=inv.id,
        payment_date=pay_date,
        amount=data.amount,
        payment_mode=data.payment_mode,
        reference_number=data.reference_number,
        notes=data.notes,
    )
    db.add(payment)
    db.flush()

    total_received = float(inv.amount_received or 0) + data.amount
    inv.amount_received = total_received
    balance = float(inv.net_payable or inv.grand_total or 0) - total_received
    inv.balance_due = max(0, balance)
    inv.payment_mode = data.payment_mode

    if balance <= 0.01:
        inv.payment_status = "PAID"
    elif total_received > 0:
        inv.payment_status = "PARTIAL"

    db.commit()
    return {"success": True, "message": "Payment recorded", "payment_status": inv.payment_status, "balance_due": float(inv.balance_due)}


# DC_COUPON_VALIDATE_001: Read-only coupon validation — no DB write, safe for pre-save preview (Apr 2026)
# DC_COUPON_UNIFIED_001: checks sales_coupon_master → marketplace_promo_codes → official_partners (Apr 2026)
@router.get("/coupon-validate")
async def validate_partner_coupon(
    code: str = Query(..., description="Coupon code to validate"),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Validate a coupon code without applying it.
    Unified lookup: sales_coupon_master → marketplace_promo_codes → official_partners (partner codes).
    Same sources as the Marketplace module.
    """
    from app.models.marketplace import MarketplacePromoCode
    from datetime import datetime as dt
    coupon_code = code.strip().upper()
    today = date.today()
    now = dt.utcnow()

    # 1. Check sales_coupon_master (admin-created promo codes)
    sc = db.query(SalesCouponMaster).filter(
        SalesCouponMaster.coupon_code == coupon_code,
        SalesCouponMaster.is_active == True,
    ).filter(
        or_(SalesCouponMaster.valid_from == None, SalesCouponMaster.valid_from <= today),
        or_(SalesCouponMaster.valid_until == None, SalesCouponMaster.valid_until >= today),
    ).first()
    if sc:
        if sc.max_uses and sc.times_used >= sc.max_uses:
            return {"valid": False, "message": "Coupon usage limit reached"}
        return {
            "valid": True,
            "coupon_code": sc.coupon_code,
            "discount_percent": float(sc.discount_percentage),
            "description": sc.description or "",
            "source": "promo",
        }

    # 2. Check marketplace_promo_codes (same pool as Marketplace — active codes)
    mp = db.query(MarketplacePromoCode).filter(
        MarketplacePromoCode.code == coupon_code,
        MarketplacePromoCode.status == "active",
    ).filter(
        or_(MarketplacePromoCode.valid_from == None, MarketplacePromoCode.valid_from <= now),
        or_(MarketplacePromoCode.valid_to == None, MarketplacePromoCode.valid_to >= now),
    ).first()
    if mp:
        if mp.usage_limit and mp.times_used >= mp.usage_limit:
            return {"valid": False, "message": "Coupon usage limit reached"}
        return {
            "valid": True,
            "coupon_code": mp.code,
            "discount_percent": float(mp.default_discount_pct or 0),
            "description": mp.label or "",
            "source": "marketplace",
        }

    # 3. Check official_partners (partner code used as referral discount)
    p = db.query(OfficialPartner).filter(
        OfficialPartner.partner_code == coupon_code,
        OfficialPartner.is_active == True,
    ).first()
    if p:
        from app.services.staff_accounts_service import SalesInvoiceService
        cat = (p.category or p.partner_type or "OTHER").upper()
        # DC_VGK_B2C_DISCOUNT: VGK_TEAM members use B2C commission chart rates
        # Activated (is_paid_activation=True OR vgk_activated_at IS NOT NULL) → 5% (L1 direct)
        # Registered only → 2% (L1 direct)
        if "VGK_TEAM" in cat or (p.partner_code and p.partner_code.upper().startswith("VGK")):
            is_vgk_activated = bool(p.is_paid_activation) or bool(getattr(p, 'vgk_activated_at', None))
            disc_pct = 5.0 if is_vgk_activated else 2.0
            tier_label = "Activated VGK Member" if is_vgk_activated else "Registered VGK Member"
        else:
            disc_pct = float(SalesInvoiceService._PARTNER_CATEGORY_DISCOUNT.get(cat, 10))
            tier_label = f"Partner: {p.partner_name}"
        return {
            "valid": True,
            "coupon_code": p.partner_code,
            "discount_percent": disc_pct,
            "description": tier_label if "VGK" in cat or (p.partner_code or "").startswith("VGK") else f"Partner: {p.partner_name}",
            "source": "partner",
        }

    # 4. DC_MNR_MEMBER_COUPON_001: Accept active MNR member IDs from the user table.
    # Activated members (coupon_status='Activated') → 12%; non-activated → 5%.
    from sqlalchemy import text as _sq_text
    member_row = db.execute(_sq_text(
        'SELECT id, name, account_status, coupon_status FROM "user" WHERE id = :uid LIMIT 1'
    ), {"uid": coupon_code}).fetchone()
    if member_row and (member_row.account_status or "").lower() == "active":
        is_activated = (member_row.coupon_status or "").strip().lower() == "activated"
        disc_pct = 3.0 if is_activated else 2.0
        tier_label = "Activated Member" if is_activated else "Registered Member"
        return {
            "valid": True,
            "coupon_code": member_row.id,
            "discount_percent": disc_pct,
            "description": f"MNR {tier_label}: {member_row.name or member_row.id}",
            "source": "mnr_member",
            "member_activated": is_activated,
        }

    return {"valid": False, "message": "Invalid or expired coupon code"}


@router.post("/sales-invoices/{invoice_id}/coupon")
async def apply_partner_coupon(
    invoice_id: int,
    body: dict,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "DRAFT",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or not editable")

    # DC_COUPON_UNIFIED_001: same 3-source lookup as /coupon-validate
    from app.models.marketplace import MarketplacePromoCode
    from datetime import datetime as dt
    coupon_code = body.get("coupon_code", "").strip().upper()
    today = date.today()
    now = dt.utcnow()

    coupon_pct = None
    resolved_code = coupon_code

    # 1. sales_coupon_master
    sc = db.query(SalesCouponMaster).filter(
        SalesCouponMaster.coupon_code == coupon_code,
        SalesCouponMaster.is_active == True,
    ).filter(
        or_(SalesCouponMaster.valid_from == None, SalesCouponMaster.valid_from <= today),
        or_(SalesCouponMaster.valid_until == None, SalesCouponMaster.valid_until >= today),
    ).first()
    if sc:
        if sc.max_uses and sc.times_used >= sc.max_uses:
            raise HTTPException(status_code=400, detail="Coupon usage limit reached")
        coupon_pct = float(sc.discount_percentage)
        resolved_code = sc.coupon_code

    # 2. marketplace_promo_codes
    if coupon_pct is None:
        mp = db.query(MarketplacePromoCode).filter(
            MarketplacePromoCode.code == coupon_code,
            MarketplacePromoCode.status == "active",
        ).filter(
            or_(MarketplacePromoCode.valid_from == None, MarketplacePromoCode.valid_from <= now),
            or_(MarketplacePromoCode.valid_to == None, MarketplacePromoCode.valid_to >= now),
        ).first()
        if mp:
            if mp.usage_limit and mp.times_used >= mp.usage_limit:
                raise HTTPException(status_code=400, detail="Coupon usage limit reached")
            coupon_pct = float(mp.default_discount_pct or 0)
            resolved_code = mp.code

    # 3. official_partners by partner_code
    if coupon_pct is None:
        p = db.query(OfficialPartner).filter(
            OfficialPartner.partner_code == coupon_code,
            OfficialPartner.is_active == True,
        ).first()
        if p:
            from app.services.staff_accounts_service import SalesInvoiceService
            cat = (p.category or p.partner_type or "OTHER").upper()
            # DC_VGK_B2C_DISCOUNT: VGK_TEAM — Activated→5%, Registered→2%
            if "VGK_TEAM" in cat or (p.partner_code or "").upper().startswith("VGK"):
                is_vgk_activated = bool(p.is_paid_activation) or bool(getattr(p, 'vgk_activated_at', None))
                coupon_pct = 5.0 if is_vgk_activated else 2.0
            else:
                coupon_pct = float(SalesInvoiceService._PARTNER_CATEGORY_DISCOUNT.get(cat, 10))
            resolved_code = p.partner_code

    # 4. DC_MNR_MEMBER_COUPON_001: MNR member IDs — Activated→12%, non-activated→5%
    if coupon_pct is None:
        from sqlalchemy import text as _sq_text
        member_row = db.execute(_sq_text(
            'SELECT id, name, account_status, coupon_status FROM "user" WHERE id = :uid LIMIT 1'
        ), {"uid": coupon_code}).fetchone()
        if member_row and (member_row.account_status or "").lower() == "active":
            is_activated = (member_row.coupon_status or "").strip().lower() == "activated"
            coupon_pct = 3.0 if is_activated else 2.0
            resolved_code = member_row.id

    if coupon_pct is None:
        raise HTTPException(status_code=400, detail="Invalid or expired coupon code")

    disc_amt = round(float(inv.taxable_amount or 0) * coupon_pct / 100, 2)
    inv.coupon_code = resolved_code
    inv.coupon_discount_pct = coupon_pct
    inv.coupon_discount_amount = disc_amt
    _recalc_invoice_totals(db, inv)
    db.commit()
    return {"success": True, "discount_percent": coupon_pct, "discount_amount": disc_amt}


@router.delete("/sales-invoices/{invoice_id}/coupon")
async def remove_partner_coupon(
    invoice_id: int,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "DRAFT",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or not editable")
    inv.coupon_code = None
    inv.coupon_discount_pct = 0
    inv.coupon_discount_amount = 0
    _recalc_invoice_totals(db, inv)
    db.commit()
    return {"success": True, "message": "Coupon removed"}


@router.patch("/sales-invoices/{invoice_id}/discount")
async def set_partner_manual_discount(
    invoice_id: int,
    body: dict,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
        SalesInvoice.status == "DRAFT",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found or not editable")
    inv.manual_discount_amount = body.get("amount", 0)
    inv.manual_discount_note = body.get("note")
    _recalc_invoice_totals(db, inv)
    db.commit()
    return {"success": True, "net_payable": float(inv.net_payable or 0)}


@router.get("/sales-invoices/{invoice_id}/pdf")
async def download_partner_invoice_pdf(
    invoice_id: int,
    mode: str = Query("tax_invoice"),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Generate and stream PDF for partner invoice. DC_PDF_PARTNER_001: seller = partner entity."""
    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.partner_id == partner.id,
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")

    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        company_id = inv.company_id or partner.company_id or 4

        # DC_PDF_PARTNER_001: Build seller override from partner data so PDF header
        # shows the partner as the seller, not MyntReal LLP.
        _addr_parts = [p for p in [partner.address, partner.city, partner.state, partner.pincode] if p]
        seller_override = {
            "seller_name":    partner.partner_name,
            "seller_address": ", ".join(_addr_parts) if _addr_parts else "Address on file",
            "seller_gstin":   f"GSTIN: {partner.gst_number}" if partner.gst_number else None,
            "seller_phone":   partner.phone or "",
            "facilitated_by": "Associated Partners: MNR – Mega Natural Resources",
            "logo_path": partner.logo_path or None,
        }

        pdf_bytes = SalesInvoiceService.generate_pdf(
            db, invoice_id, company_id, mode, seller_override=seller_override
        )
        filename = f"{inv.invoice_number.replace('/', '-')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


# ─────────────── HSN + Stock Item Search (Partner) ──────────

@router.get("/hsn-search")
async def partner_hsn_search(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=50),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    query = db.query(HSNMaster).filter(HSNMaster.is_active == True)
    if search:
        query = query.filter(or_(
            HSNMaster.hsn_code.ilike(f"%{search}%"),
            HSNMaster.description.ilike(f"%{search}%"),
        ))
    total = query.count()
    items = query.order_by(HSNMaster.hsn_code).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "success": True,
        "total": total,
        "data": [
            {
                "id": h.id,
                "hsn_code": h.hsn_code,
                "description": h.description,
                "cgst_rate": float(h.cgst_rate) if h.cgst_rate is not None else 0,
                "sgst_rate": float(h.sgst_rate) if h.sgst_rate is not None else 0,
                "igst_rate": float(h.igst_rate) if h.igst_rate is not None else 0,
                "gst_rate": round((float(h.cgst_rate or 0) + float(h.sgst_rate or 0)), 2),
            }
            for h in items
        ],
    }


@router.get("/stock-items-search")
async def partner_stock_items_search(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=50),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_SEARCH_001: Search stock (own→global) then marketplace_spares. Apr 2026."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)

    def _safe(v):
        try:
            return float(v) if v is not None else None
        except Exception:
            return None

    combined: list = []

    # ── 1. Partner's own stock ────────────────────────────────────────────────
    params: Dict[str, Any] = {"pid": partner.id}
    search_clause = ""
    if search:
        params["s"] = f"%{search}%"
        search_clause = "AND (psi.item_name ILIKE :s OR psi.item_code ILIKE :s OR sim.item_name ILIKE :s OR sim.item_code ILIKE :s)"

    own_total = db.execute(sq_text(f"""
        SELECT COUNT(*) FROM partner_stock_items psi
        LEFT JOIN stock_item_master sim ON sim.id = psi.stock_item_id
        WHERE psi.partner_id = :pid AND psi.is_active = TRUE {search_clause}
    """), params).scalar() or 0

    if own_total > 0:
        params["limit"] = per_page
        params["offset"] = (page - 1) * per_page
        own_rows = db.execute(sq_text(f"""
            SELECT
                psi.id                                                  AS psi_id,
                psi.stock_item_id,
                COALESCE(psi.item_name, sim.item_name)                 AS item_name,
                COALESCE(psi.item_code, sim.item_code)                 AS item_code,
                COALESCE(psi.hsn_code, sim.hsn_code)                   AS hsn_code,
                sim.hsn_id,
                COALESCE(psi.unit_of_measure, sim.unit_of_measure, 'PCS') AS unit_of_measure,
                COALESCE(psi.selling_price, sim.selling_rate)           AS selling_price
            FROM partner_stock_items psi
            LEFT JOIN stock_item_master sim ON sim.id = psi.stock_item_id
            WHERE psi.partner_id = :pid AND psi.is_active = TRUE {search_clause}
            ORDER BY item_name
            LIMIT :limit OFFSET :offset
        """), params).fetchall()
        for r in own_rows:
            combined.append({
                "id": r.stock_item_id or r.psi_id,
                "psi_id": r.psi_id,
                "item_code": r.item_code or "",
                "item_name": r.item_name or "",
                "hsn_code": r.hsn_code,
                "hsn_id": r.hsn_id,
                "unit_of_measure": r.unit_of_measure or "PCS",
                "selling_price": _safe(r.selling_price),
                "mrp": None,
                "source": "stock",
                "source_label": "My Stock",
            })

    # ── 2. Global stock_item_master (if no own stock) ─────────────────────────
    if own_total == 0 and search:
        g_params: Dict[str, Any] = {}
        g_where = "WHERE sim.is_active = TRUE"
        if search:
            g_params["s"] = f"%{search}%"
            g_where += " AND (sim.item_name ILIKE :s OR sim.item_code ILIKE :s)"
        g_params["limit"] = per_page
        g_params["offset"] = (page - 1) * per_page
        global_rows = db.execute(sq_text(f"""
            SELECT
                sim.id              AS stock_item_id,
                sim.item_name,
                sim.item_code,
                sim.hsn_code,
                sim.hsn_id,
                COALESCE(sim.unit_of_measure, 'PCS') AS unit_of_measure,
                sim.selling_rate    AS selling_price
            FROM stock_item_master sim
            {g_where}
            ORDER BY sim.item_name
            LIMIT :limit OFFSET :offset
        """), g_params).fetchall()
        for r in global_rows:
            combined.append({
                "id": r.stock_item_id,
                "psi_id": None,
                "item_code": r.item_code or "",
                "item_name": r.item_name or "",
                "hsn_code": r.hsn_code,
                "hsn_id": r.hsn_id,
                "unit_of_measure": r.unit_of_measure or "PCS",
                "selling_price": _safe(r.selling_price),
                "mrp": None,
                "source": "global_stock",
                "source_label": "Stock",
            })

    # ── 3. Marketplace spares (always appended after stock) ───────────────────
    if search:
        mp_params: Dict[str, Any] = {}
        mp_where = "WHERE ms.is_active = TRUE"
        mp_params["s"] = f"%{search}%"
        mp_where += " AND (ms.name ILIKE :s OR ms.sku ILIKE :s OR ms.category_name ILIKE :s OR ms.brand ILIKE :s)"
        mp_params["limit"] = max(per_page - len(combined), 5)
        mp_rows = db.execute(sq_text(f"""
            SELECT
                ms.id, ms.sku, ms.name, ms.category_name,
                ms.dealer_price, ms.gst_percent, ms.brand
            FROM marketplace_spares ms
            {mp_where}
            ORDER BY ms.name
            LIMIT :limit
        """), mp_params).fetchall()
        for r in mp_rows:
            combined.append({
                "id": f"mp_{r.id}",
                "psi_id": None,
                "item_code": r.sku or "",
                "item_name": r.name or "",
                "hsn_code": None,
                "hsn_id": None,
                "unit_of_measure": "PCS",
                "selling_price": _safe(r.dealer_price),
                "mrp": None,
                "gst_percent": _safe(r.gst_percent),
                "brand": r.brand or "",
                "source": "marketplace",
                "source_label": "Marketplace",
            })

    return {
        "success": True,
        "total": len(combined),
        "data": combined,
    }


# ─────────────── Partner: My Purchases ───────────────────────

@router.get("/purchases")
async def list_partner_purchases(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    payment_status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """
    DC_PARTNER_PURCHASES_001: List all company-raised invoices for this partner.
    Shows PartnerInvoice records (VGK/EA raising invoices for partner orders).
    """
    from sqlalchemy import text as sq_text
    conditions = ["pi.partner_id = :pid"]
    params: Dict[str, Any] = {"pid": partner.id}
    if payment_status:
        conditions.append("pi.payment_status = :ps")
        params["ps"] = payment_status.upper()
    if date_from:
        conditions.append("pi.invoice_date >= :df")
        params["df"] = date_from
    if date_to:
        conditions.append("pi.invoice_date <= :dt")
        params["dt"] = date_to
    if search:
        conditions.append("(pi.invoice_number ILIKE :s OR po.order_number ILIKE :s)")
        params["s"] = f"%{search}%"
    where = " AND ".join(conditions)
    total = db.execute(sq_text(f"""
        SELECT COUNT(*) FROM partner_invoices pi
        LEFT JOIN partner_orders po ON po.id = pi.order_id
        WHERE {where}
    """), params).scalar() or 0
    rows = db.execute(sq_text(f"""
        SELECT pi.id, pi.invoice_number, pi.invoice_date, pi.due_date,
               pi.grand_total, pi.amount_received, pi.balance_due, pi.payment_status,
               pi.subtotal, pi.total_tax, pi.company_id,
               po.order_number, po.id AS order_id,
               ac.company_name, ac.company_code
        FROM partner_invoices pi
        LEFT JOIN partner_orders po ON po.id = pi.order_id
        LEFT JOIN associated_companies ac ON ac.id = pi.company_id
        WHERE {where}
        ORDER BY pi.invoice_date DESC, pi.id DESC
        LIMIT :lim OFFSET :off
    """), {**params, "lim": per_page, "off": (page - 1) * per_page}).fetchall()
    data = []
    for r in rows:
        data.append({
            "id": r.id,
            "invoice_number": r.invoice_number,
            "invoice_date": r.invoice_date.isoformat() if r.invoice_date else None,
            "due_date": r.due_date.isoformat() if r.due_date else None,
            "grand_total": float(r.grand_total or 0),
            "amount_received": float(r.amount_received or 0),
            "balance_due": float(r.balance_due or 0),
            "payment_status": r.payment_status,
            "subtotal": float(r.subtotal or 0),
            "total_tax": float(r.total_tax or 0),
            "company_id": r.company_id,
            "company_name": r.company_name,
            "company_code": r.company_code,
            "order_number": r.order_number,
            "order_id": r.order_id,
        })
    return {"success": True, "total": total, "page": page, "per_page": per_page, "data": data}


@router.get("/purchases/{invoice_id}")
async def get_partner_purchase_detail(
    invoice_id: int = Path(...),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_PURCHASES_001: Get purchase detail with order line items."""
    from sqlalchemy import text as sq_text
    inv = db.execute(sq_text("""
        SELECT pi.id, pi.invoice_number, pi.invoice_date, pi.due_date,
               pi.grand_total, pi.amount_received, pi.balance_due, pi.payment_status,
               pi.subtotal, pi.total_tax, pi.cgst_amount, pi.sgst_amount, pi.igst_amount,
               pi.remarks, pi.irn_number, pi.e_way_bill_number,
               po.order_number, po.id AS order_id, po.status AS order_status,
               ac.company_name, ac.company_code
        FROM partner_invoices pi
        LEFT JOIN partner_orders po ON po.id = pi.order_id
        LEFT JOIN associated_companies ac ON ac.id = pi.company_id
        WHERE pi.id = :iid AND pi.partner_id = :pid
    """), {"iid": invoice_id, "pid": partner.id}).fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail="Purchase not found")
    lines = db.execute(sq_text("""
        SELECT pol.id, pol.quantity, pol.unit_of_measure, pol.unit_rate,
               pol.discount_pct, pol.tax_rate, pol.tax_amount, pol.line_total,
               sim.item_name, sim.item_code, sim.hsn_code
        FROM partner_order_lines pol
        LEFT JOIN stock_item_master sim ON sim.id = pol.item_id
        WHERE pol.order_id = :oid
        ORDER BY pol.id
    """), {"oid": inv.order_id}).fetchall() if inv.order_id else []
    return {
        "success": True,
        "data": {
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
            "due_date": inv.due_date.isoformat() if inv.due_date else None,
            "grand_total": float(inv.grand_total or 0),
            "amount_received": float(inv.amount_received or 0),
            "balance_due": float(inv.balance_due or 0),
            "payment_status": inv.payment_status,
            "subtotal": float(inv.subtotal or 0),
            "total_tax": float(inv.total_tax or 0),
            "cgst_amount": float(inv.cgst_amount or 0),
            "sgst_amount": float(inv.sgst_amount or 0),
            "igst_amount": float(inv.igst_amount or 0),
            "remarks": inv.remarks,
            "irn_number": inv.irn_number,
            "e_way_bill_number": inv.e_way_bill_number,
            "order_number": inv.order_number,
            "order_id": inv.order_id,
            "order_status": inv.order_status,
            "company_name": inv.company_name,
            "company_code": inv.company_code,
            "line_items": [
                {
                    "id": l.id,
                    "item_name": l.item_name,
                    "item_code": l.item_code,
                    "hsn_code": l.hsn_code,
                    "quantity": float(l.quantity or 0),
                    "unit_of_measure": l.unit_of_measure,
                    "unit_rate": float(l.unit_rate or 0),
                    "discount_pct": float(l.discount_pct or 0),
                    "tax_rate": float(l.tax_rate or 0),
                    "tax_amount": float(l.tax_amount or 0),
                    "line_total": float(l.line_total or 0),
                }
                for l in lines
            ],
        },
    }


# ─────────────── Partner: WhatsApp Templates ─────────────────

@router.get("/wa-templates")
async def partner_list_wa_templates(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """Active WhatsApp templates accessible to partners for composing wa.me messages."""
    from app.models.whatsapp import WhatsAppTemplate
    templates = db.query(WhatsAppTemplate).filter(
        WhatsAppTemplate.is_active == True
    ).order_by(WhatsAppTemplate.segment, WhatsAppTemplate.name).all()
    return {"success": True, "templates": [t.to_dict() for t in templates]}


# ─────────────── Partner: My Stock ───────────────────────────

def _resolve_walkin_category_id(db: Session, category_id: Optional[int], visit_purpose: str, company_id: int) -> Optional[int]:
    """
    Task #32: Validate category_id belongs to partner's company.
    If category_id is provided but doesn't belong to company, discard it and
    fall back to slug-based lookup. If still not found, return None.
    """
    from app.models.signup_category import SignupCategory
    if category_id is not None:
        cat = db.query(SignupCategory).filter(
            SignupCategory.id == category_id,
            SignupCategory.company_id == company_id,
            SignupCategory.is_active == True
        ).first()
        if cat:
            return category_id
    if visit_purpose:
        cat = db.query(SignupCategory).filter(
            SignupCategory.slug == visit_purpose,
            SignupCategory.company_id == company_id,
            SignupCategory.is_active == True
        ).first()
        if cat:
            return cat.id
    return None


def _ensure_walkin_vgk_cols(db: Session):
    """DC_WALKIN_VGK_001: Idempotent DDL — add VGK columns to partner_walkins if missing."""
    from sqlalchemy import text as _sq
    try:
        db.execute(_sq("ALTER TABLE partner_walkins ADD COLUMN IF NOT EXISTS vgk_member_id INTEGER REFERENCES official_partners(id)"))
        db.execute(_sq("ALTER TABLE partner_walkins ADD COLUMN IF NOT EXISTS vgk_enrolled_at TIMESTAMP"))
        db.execute(_sq("ALTER TABLE partner_walkins ADD COLUMN IF NOT EXISTS category_id INTEGER REFERENCES signup_categories(id)"))
        db.execute(_sq("ALTER TABLE partner_walkins ADD COLUMN IF NOT EXISTS is_new_customer BOOLEAN DEFAULT FALSE"))
        db.commit()
    except Exception:
        db.rollback()


def _get_or_create_vgk_member(db: Session, phone: str, name: str, company_id: int, parent_partner_id, partner_ref_id=None):
    """
    DC_WALKIN_VGK_001: Find or create a VGK member by phone.
    Returns (vgk_member_id: int, was_created: bool, message: str).
    Dedup: if phone already exists in VGK_TEAM → links instead of creating.
    Password is set to the mobile number (bcrypt).
    """
    import random
    from decimal import Decimal as _Dec
    from app.core.security import SecurityManager as _SM

    existing = db.query(OfficialPartner).filter(
        OfficialPartner.phone == phone,
        OfficialPartner.category == 'VGK_TEAM'
    ).first()
    if existing:
        return existing.id, False, f"Linked to existing VGK member {existing.partner_code}"

    for _ in range(20):
        code = f"VGK0710{random.randint(1000, 9999)}"
        if not db.query(OfficialPartner).filter(OfficialPartner.partner_code == code).first():
            break

    member = OfficialPartner(
        company_id=company_id,
        partner_code=code,
        partner_name=name,
        phone=phone,
        category='VGK_TEAM',
        is_active=False,
        parent_partner_id=parent_partner_id,
        vgk_role='VGK_ASSOCIATE',
        vgk_points_balance=_Dec('0'),
        password_hash=_SM.get_password_hash(phone),
        created_at=datetime.now(),
        updated_at=datetime.now()
    )
    db.add(member)
    db.flush()
    # DC_WALKIN_VGK_FIX_001 (Apr 2026): Commit member BEFORE calling _awpe so that any
    # rollback inside _awpe does not wipe the flushed-but-uncommitted member row.
    # Without this commit the FK update on partner_walkins.vgk_member_id would fail
    # with a constraint violation (referencing a phantom ID), making vgk_result always an error.
    db.commit()

    try:
        from app.services.vgk_commission import add_vgk_points_entry as _awpe
        # 10,000 Welcome Bonus
        _awpe(
            db=db, partner_id=member.id,
            points_credit=_Dec('10000'), points_debit=_Dec('0'),
            reason_code='WELCOME_BONUS', reference_type='partner_walkin',
            reference_id=partner_ref_id,
            notes='Welcome Bonus — registered via partner walk-in',
            created_by=None
        )
        # 5,000 Referral Bonus — walk-ins are always via a partner referral
        _awpe(
            db=db, partner_id=member.id,
            points_credit=_Dec('5000'), points_debit=_Dec('0'),
            reason_code='REFERRAL_BONUS', reference_type='partner_walkin',
            reference_id=partner_ref_id,
            notes='Referral Bonus — joined via partner walk-in registration',
            created_by=None
        )
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

    return member.id, True, f"VGK member created ({code}) — password is mobile number"


def _ensure_stock_tables(db: Session):
    """DC_PARTNER_STOCK_001: Idempotent table creation for partner stock system."""
    from sqlalchemy import text as sq_text
    db.execute(sq_text("""
        CREATE TABLE IF NOT EXISTS partner_stock_items (
            id SERIAL PRIMARY KEY,
            partner_id INTEGER NOT NULL,
            item_type VARCHAR(20) NOT NULL DEFAULT 'catalog',
            stock_item_id INTEGER,
            item_name VARCHAR(200) NOT NULL,
            item_code VARCHAR(100),
            unit_of_measure VARCHAR(20) DEFAULT 'PCS',
            hsn_code VARCHAR(20),
            opening_qty NUMERIC(10,2) DEFAULT 0,
            opening_qty_set_at TIMESTAMP,
            reorder_level NUMERIC(10,2) DEFAULT 0,
            selling_price NUMERIC(10,2),
            is_active BOOLEAN DEFAULT TRUE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))
    db.execute(sq_text("""
        CREATE TABLE IF NOT EXISTS partner_stock_adjustments (
            id SERIAL PRIMARY KEY,
            partner_id INTEGER NOT NULL,
            partner_stock_item_id INTEGER NOT NULL,
            adj_type VARCHAR(30) NOT NULL,
            qty NUMERIC(10,2) NOT NULL,
            reason VARCHAR(200),
            notes TEXT,
            ref_doc_type VARCHAR(50),
            ref_doc_id INTEGER,
            ref_doc_number VARCHAR(100),
            created_by VARCHAR(100),
            created_at TIMESTAMP DEFAULT NOW()
        )
    """))
    db.commit()


@router.get("/stock/items")
async def list_partner_stock_items(
    search: Optional[str] = Query(None),
    item_type: Optional[str] = Query(None),
    low_stock_only: bool = Query(False),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: List partner stock items with live calculated position."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    conditions = ["psi.partner_id = :pid", "psi.is_active = TRUE"]
    params: Dict[str, Any] = {"pid": partner.id}
    if search:
        conditions.append("(psi.item_name ILIKE :s OR psi.item_code ILIKE :s)")
        params["s"] = f"%{search}%"
    if item_type:
        conditions.append("psi.item_type = :itype")
        params["itype"] = item_type
    where = " AND ".join(conditions)
    total = db.execute(sq_text(f"SELECT COUNT(*) FROM partner_stock_items psi WHERE {where}"), params).scalar() or 0
    rows = db.execute(sq_text(f"""
        SELECT
            psi.id, psi.item_type, psi.stock_item_id,
            psi.item_name, psi.item_code, psi.unit_of_measure, psi.hsn_code,
            psi.opening_qty, psi.reorder_level, psi.selling_price, psi.notes,
            psi.created_at,
            COALESCE((
                SELECT SUM(pol.quantity)
                FROM partner_order_lines pol
                JOIN partner_orders po ON po.id = pol.order_id
                WHERE po.partner_id = :pid
                  AND po.status IN ('DELIVERED','COMPLETED','DISPATCHED')
                  AND pol.item_id = psi.stock_item_id
                  AND psi.stock_item_id IS NOT NULL
            ), 0) + COALESCE((
                SELECT SUM(sli.quantity)
                FROM sales_invoice_line_items sli
                JOIN sales_invoices si ON si.id = sli.invoice_id
                WHERE si.partner_id = :pid
                  AND si.status = 'CONFIRMED'
                  AND (si.created_by_type IS NULL OR si.created_by_type != 'PARTNER')
                  AND sli.item_id = psi.stock_item_id
                  AND psi.stock_item_id IS NOT NULL
            ), 0) AS purchases_in,
            COALESCE((
                SELECT SUM(sli.quantity)
                FROM sales_invoice_line_items sli
                JOIN sales_invoices si ON si.id = sli.invoice_id
                WHERE si.partner_id = :pid
                  AND si.status = 'CONFIRMED'
                  AND si.created_by_type = 'PARTNER'
                  AND sli.item_id = psi.stock_item_id
                  AND psi.stock_item_id IS NOT NULL
            ), 0) AS sales_out,
            COALESCE((
                SELECT SUM(adj.qty)
                FROM partner_stock_adjustments adj
                WHERE adj.partner_id = :pid
                  AND adj.partner_stock_item_id = psi.id
            ), 0) AS adj_total
        FROM partner_stock_items psi
        WHERE {where}
        ORDER BY psi.item_name
        LIMIT :lim OFFSET :off
    """), {**params, "lim": per_page, "off": (page - 1) * per_page}).fetchall()

    data = []
    for r in rows:
        current_qty = float(r.opening_qty or 0) + float(r.purchases_in or 0) - float(r.sales_out or 0) + float(r.adj_total or 0)
        is_low = current_qty <= float(r.reorder_level or 0) and float(r.reorder_level or 0) > 0
        if low_stock_only and not is_low:
            continue
        stock_value = current_qty * float(r.selling_price or 0) if r.selling_price else 0
        data.append({
            "id": r.id,
            "item_type": r.item_type,
            "stock_item_id": r.stock_item_id,
            "item_name": r.item_name,
            "item_code": r.item_code,
            "unit_of_measure": r.unit_of_measure,
            "hsn_code": r.hsn_code,
            "opening_qty": float(r.opening_qty or 0),
            "purchases_in": float(r.purchases_in or 0),
            "sales_out": float(r.sales_out or 0),
            "adj_total": float(r.adj_total or 0),
            "current_qty": round(current_qty, 2),
            "reorder_level": float(r.reorder_level or 0),
            "is_low_stock": is_low,
            "selling_price": float(r.selling_price or 0) if r.selling_price else None,
            "stock_value": round(stock_value, 2),
            "notes": r.notes,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    total_stock_value = sum(d["stock_value"] for d in data)
    low_count = sum(1 for d in data if d["is_low_stock"])
    return {
        "success": True,
        "total": total,
        "page": page,
        "per_page": per_page,
        "data": data,
        "summary": {
            "total_items": len(data),
            "total_stock_value": round(total_stock_value, 2),
            "low_stock_count": low_count,
        },
    }


class PartnerStockItemIn(BaseModel):
    item_type: str = Field("catalog", pattern="^(catalog|custom)$")
    stock_item_id: Optional[int] = None
    item_name: str = Field(..., min_length=1, max_length=200)
    item_code: Optional[str] = Field(None, max_length=100)
    unit_of_measure: str = Field("PCS", max_length=20)
    hsn_code: Optional[str] = Field(None, max_length=20)
    opening_qty: float = Field(0.0, ge=0)
    reorder_level: float = Field(0.0, ge=0)
    selling_price: Optional[float] = None
    notes: Optional[str] = None


@router.post("/stock/items")
async def add_partner_stock_item(
    payload: PartnerStockItemIn,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Add a new stock item (catalog item with opening qty, or custom spare)."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    # For catalog items, auto-fill name from stock_item_master if not provided
    item_name = payload.item_name
    item_code = payload.item_code
    hsn_code = payload.hsn_code
    if payload.item_type == "catalog" and payload.stock_item_id:
        sim = db.execute(sq_text("SELECT item_name, item_code, hsn_code FROM stock_item_master WHERE id = :iid"), {"iid": payload.stock_item_id}).fetchone()
        if sim:
            item_name = item_name or sim.item_name
            item_code = item_code or sim.item_code
            hsn_code = hsn_code or sim.hsn_code
    # Check duplicate for catalog items
    if payload.item_type == "catalog" and payload.stock_item_id:
        exists = db.execute(sq_text("SELECT id FROM partner_stock_items WHERE partner_id = :pid AND stock_item_id = :sid AND is_active = TRUE"), {"pid": partner.id, "sid": payload.stock_item_id}).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="This catalog item is already in your stock. Use the update endpoint to change opening qty.")
    result = db.execute(sq_text("""
        INSERT INTO partner_stock_items
            (partner_id, item_type, stock_item_id, item_name, item_code, unit_of_measure, hsn_code,
             opening_qty, opening_qty_set_at, reorder_level, selling_price, notes, created_at, updated_at)
        VALUES (:pid, :itype, :sid, :iname, :icode, :uom, :hsn, :oqty, NOW(), :rlvl, :sp, :notes, NOW(), NOW())
        RETURNING id
    """), {
        "pid": partner.id, "itype": payload.item_type, "sid": payload.stock_item_id,
        "iname": item_name, "icode": item_code, "uom": payload.unit_of_measure,
        "hsn": hsn_code, "oqty": payload.opening_qty, "rlvl": payload.reorder_level,
        "sp": payload.selling_price, "notes": payload.notes,
    })
    new_id = result.fetchone().id
    db.commit()
    return {"success": True, "message": "Stock item added successfully", "id": new_id}


class PartnerStockItemUpdate(BaseModel):
    item_name: Optional[str] = None
    item_code: Optional[str] = None
    unit_of_measure: Optional[str] = None
    hsn_code: Optional[str] = None
    opening_qty: Optional[float] = None
    reorder_level: Optional[float] = None
    selling_price: Optional[float] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None


@router.put("/stock/items/{item_id}")
async def update_partner_stock_item(
    item_id: int = Path(...),
    payload: PartnerStockItemUpdate = ...,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Update a stock item's configuration."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    exists = db.execute(sq_text("SELECT id FROM partner_stock_items WHERE id = :iid AND partner_id = :pid"), {"iid": item_id, "pid": partner.id}).fetchone()
    if not exists:
        raise HTTPException(status_code=404, detail="Stock item not found")
    sets = ["updated_at = NOW()"]
    params: Dict[str, Any] = {"iid": item_id, "pid": partner.id}
    if payload.item_name is not None:
        sets.append("item_name = :iname"); params["iname"] = payload.item_name
    if payload.item_code is not None:
        sets.append("item_code = :icode"); params["icode"] = payload.item_code
    if payload.unit_of_measure is not None:
        sets.append("unit_of_measure = :uom"); params["uom"] = payload.unit_of_measure
    if payload.hsn_code is not None:
        sets.append("hsn_code = :hsn"); params["hsn"] = payload.hsn_code
    if payload.opening_qty is not None:
        sets.append("opening_qty = :oqty, opening_qty_set_at = NOW()"); params["oqty"] = payload.opening_qty
    if payload.reorder_level is not None:
        sets.append("reorder_level = :rlvl"); params["rlvl"] = payload.reorder_level
    if payload.selling_price is not None:
        sets.append("selling_price = :sp"); params["sp"] = payload.selling_price
    if payload.notes is not None:
        sets.append("notes = :notes"); params["notes"] = payload.notes
    if payload.is_active is not None:
        sets.append("is_active = :active"); params["active"] = payload.is_active
    db.execute(sq_text(f"UPDATE partner_stock_items SET {', '.join(sets)} WHERE id = :iid AND partner_id = :pid"), params)
    db.commit()
    return {"success": True, "message": "Stock item updated"}


class PartnerStockAdjIn(BaseModel):
    partner_stock_item_id: int
    adj_type: str = Field(..., description="adj_in | adj_out | damage | return | opening")
    qty: float = Field(..., description="Positive for in, will be made negative for out/damage")
    reason: Optional[str] = None
    notes: Optional[str] = None
    ref_doc_type: Optional[str] = None
    ref_doc_number: Optional[str] = None


@router.post("/stock/adjustments")
async def add_stock_adjustment(
    payload: PartnerStockAdjIn,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Record a manual stock adjustment."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    exists = db.execute(sq_text("SELECT id FROM partner_stock_items WHERE id = :iid AND partner_id = :pid"), {"iid": payload.partner_stock_item_id, "pid": partner.id}).fetchone()
    if not exists:
        raise HTTPException(status_code=404, detail="Stock item not found")
    # Out-type adjustments stored as negative
    qty = abs(payload.qty)
    if payload.adj_type in ("adj_out", "damage"):
        qty = -qty
    db.execute(sq_text("""
        INSERT INTO partner_stock_adjustments
            (partner_id, partner_stock_item_id, adj_type, qty, reason, notes, ref_doc_type, ref_doc_number, created_by, created_at)
        VALUES (:pid, :siid, :atype, :qty, :reason, :notes, :rdt, :rdn, :by, NOW())
    """), {
        "pid": partner.id, "siid": payload.partner_stock_item_id, "atype": payload.adj_type,
        "qty": qty, "reason": payload.reason, "notes": payload.notes,
        "rdt": payload.ref_doc_type, "rdn": payload.ref_doc_number,
        "by": f"partner:{partner.partner_code}",
    })
    db.commit()
    return {"success": True, "message": "Adjustment recorded"}


@router.get("/stock/ledger")
async def get_stock_ledger(
    stock_item_id: Optional[int] = Query(None, description="Filter by partner_stock_items.id"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Full stock ledger — opening + purchases + sales + adjustments with running balance."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    pid = partner.id
    item_filter = "AND psi.id = :siid" if stock_item_id else ""
    item_params: Dict[str, Any] = {"pid": pid}
    if stock_item_id:
        item_params["siid"] = stock_item_id

    rows = db.execute(sq_text(f"""
        WITH ledger AS (
            -- Opening entries
            SELECT
                psi.id AS stock_item_id,
                psi.item_name,
                psi.item_code,
                psi.unit_of_measure,
                'opening' AS entry_type,
                psi.opening_qty_set_at::date AS entry_date,
                psi.opening_qty AS qty,
                'Opening stock set' AS description,
                NULL AS ref_doc,
                psi.id AS sort_key,
                1 AS type_order
            FROM partner_stock_items psi
            WHERE psi.partner_id = :pid AND psi.is_active = TRUE {item_filter}

            UNION ALL

            -- Purchases IN from partner orders
            SELECT
                psi.id AS stock_item_id,
                psi.item_name,
                psi.item_code,
                psi.unit_of_measure,
                'purchase_in' AS entry_type,
                po.updated_at::date AS entry_date,
                pol.quantity AS qty,
                'Purchase from order ' || po.order_number AS description,
                po.order_number AS ref_doc,
                po.id AS sort_key,
                2 AS type_order
            FROM partner_stock_items psi
            JOIN partner_order_lines pol ON pol.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
            JOIN partner_orders po ON po.id = pol.order_id
            WHERE psi.partner_id = :pid AND psi.is_active = TRUE
              AND po.partner_id = :pid
              AND po.status IN ('DELIVERED','COMPLETED','DISPATCHED') {item_filter}

            UNION ALL

            -- Sales OUT from confirmed sales invoices
            SELECT
                psi.id AS stock_item_id,
                psi.item_name,
                psi.item_code,
                psi.unit_of_measure,
                'sales_out' AS entry_type,
                si.invoice_date AS entry_date,
                -sli.quantity AS qty,
                'Sale via invoice ' || si.invoice_number AS description,
                si.invoice_number AS ref_doc,
                si.id AS sort_key,
                3 AS type_order
            FROM partner_stock_items psi
            JOIN sales_invoice_line_items sli ON sli.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
            JOIN sales_invoices si ON si.id = sli.invoice_id
            WHERE psi.partner_id = :pid AND psi.is_active = TRUE
              AND si.partner_id = :pid
              AND si.status = 'CONFIRMED' {item_filter}

            UNION ALL

            -- Manual adjustments
            SELECT
                psi.id AS stock_item_id,
                psi.item_name,
                psi.item_code,
                psi.unit_of_measure,
                adj.adj_type AS entry_type,
                adj.created_at::date AS entry_date,
                adj.qty AS qty,
                COALESCE(adj.reason, adj.adj_type) AS description,
                adj.ref_doc_number AS ref_doc,
                adj.id AS sort_key,
                4 AS type_order
            FROM partner_stock_items psi
            JOIN partner_stock_adjustments adj ON adj.partner_stock_item_id = psi.id
            WHERE psi.partner_id = :pid AND psi.is_active = TRUE {item_filter}
        )
        SELECT *,
               SUM(qty) OVER (PARTITION BY stock_item_id ORDER BY entry_date, type_order, sort_key
                              ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) AS running_balance
        FROM ledger
        ORDER BY stock_item_id, entry_date, type_order, sort_key
        LIMIT :lim OFFSET :off
    """), {**item_params, "lim": per_page, "off": (page - 1) * per_page}).fetchall()

    total = db.execute(sq_text(f"""
        SELECT COUNT(*) FROM (
            SELECT 1 FROM partner_stock_items psi WHERE psi.partner_id = :pid AND psi.is_active = TRUE {item_filter}
            UNION ALL
            SELECT 1 FROM partner_stock_items psi
                JOIN partner_order_lines pol ON pol.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
                JOIN partner_orders po ON po.id = pol.order_id
                WHERE psi.partner_id = :pid AND po.partner_id = :pid AND po.status IN ('DELIVERED','COMPLETED','DISPATCHED') {item_filter}
            UNION ALL
            SELECT 1 FROM partner_stock_items psi
                JOIN sales_invoice_line_items sli ON sli.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
                JOIN sales_invoices si ON si.id = sli.invoice_id
                WHERE psi.partner_id = :pid AND si.partner_id = :pid AND si.status = 'CONFIRMED' {item_filter}
            UNION ALL
            SELECT 1 FROM partner_stock_items psi
                JOIN partner_stock_adjustments adj ON adj.partner_stock_item_id = psi.id
                WHERE psi.partner_id = :pid {item_filter}
        ) sub
    """), item_params).scalar() or 0

    return {
        "success": True,
        "total": int(total),
        "page": page,
        "per_page": per_page,
        "data": [
            {
                "stock_item_id": r.stock_item_id,
                "item_name": r.item_name,
                "item_code": r.item_code,
                "unit_of_measure": r.unit_of_measure,
                "entry_type": r.entry_type,
                "entry_date": r.entry_date.isoformat() if r.entry_date else None,
                "qty": float(r.qty or 0),
                "description": r.description,
                "ref_doc": r.ref_doc,
                "running_balance": round(float(r.running_balance or 0), 2),
            }
            for r in rows
        ],
    }


@router.get("/stock/summary")
async def get_stock_summary(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Summary stats for dashboard and revenue page."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    pid = partner.id
    items = db.execute(sq_text("""
        SELECT
            psi.id, psi.opening_qty, psi.reorder_level, psi.selling_price,
            psi.stock_item_id,
            COALESCE((
                SELECT SUM(pol.quantity)
                FROM partner_order_lines pol JOIN partner_orders po ON po.id = pol.order_id
                WHERE po.partner_id = :pid AND po.status IN ('DELIVERED','COMPLETED','DISPATCHED')
                  AND pol.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
            ), 0) AS purchases_in,
            COALESCE((
                SELECT SUM(sli.quantity)
                FROM sales_invoice_line_items sli JOIN sales_invoices si ON si.id = sli.invoice_id
                WHERE si.partner_id = :pid AND si.status = 'CONFIRMED'
                  AND sli.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
            ), 0) AS sales_out,
            COALESCE((
                SELECT SUM(adj.qty) FROM partner_stock_adjustments adj
                WHERE adj.partner_id = :pid AND adj.partner_stock_item_id = psi.id
            ), 0) AS adj_total
        FROM partner_stock_items psi
        WHERE psi.partner_id = :pid AND psi.is_active = TRUE
    """), {"pid": pid}).fetchall()

    total_items = len(items)
    total_value = 0.0
    low_stock_count = 0
    zero_stock_count = 0
    for r in items:
        cur = float(r.opening_qty or 0) + float(r.purchases_in or 0) - float(r.sales_out or 0) + float(r.adj_total or 0)
        total_value += cur * float(r.selling_price or 0)
        if float(r.reorder_level or 0) > 0 and cur <= float(r.reorder_level or 0):
            low_stock_count += 1
        if cur <= 0:
            zero_stock_count += 1

    return {
        "success": True,
        "summary": {
            "total_items": total_items,
            "total_stock_value": round(total_value, 2),
            "low_stock_count": low_stock_count,
            "zero_stock_count": zero_stock_count,
        },
    }


# ─────────────── Staff: Partner Stock View ───────────────────

@router.get("/staff-view/partner-stock")
async def staff_view_partner_stock(
    request: Request,
    partner_id: Optional[int] = Query(None),
    partner_code: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    low_stock_only: bool = Query(False),
    item_type: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Staff view — showroom-wise stock across all partners."""
    from sqlalchemy import text as sq_text
    _ensure_stock_tables(db)
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Staff authentication required")

    conditions = ["psi.is_active = TRUE"]
    params: Dict[str, Any] = {}
    if partner_id:
        conditions.append("psi.partner_id = :pid")
        params["pid"] = partner_id
    if partner_code:
        conditions.append("op.partner_code ILIKE :pcode")
        params["pcode"] = f"%{partner_code}%"
    if search:
        conditions.append("(psi.item_name ILIKE :s OR psi.item_code ILIKE :s OR op.partner_code ILIKE :s)")
        params["s"] = f"%{search}%"
    if item_type:
        conditions.append("psi.item_type = :itype")
        params["itype"] = item_type
    where = " AND ".join(conditions)

    total = db.execute(sq_text(f"""
        SELECT COUNT(*) FROM partner_stock_items psi
        JOIN official_partners op ON op.id = psi.partner_id
        WHERE {where}
    """), params).scalar() or 0

    rows = db.execute(sq_text(f"""
        SELECT
            psi.id, psi.partner_id, psi.item_type, psi.stock_item_id,
            psi.item_name, psi.item_code, psi.unit_of_measure, psi.hsn_code,
            psi.opening_qty, psi.reorder_level, psi.selling_price,
            op.partner_code, op.partner_name, op.city,
            COALESCE((
                SELECT SUM(pol.quantity)
                FROM partner_order_lines pol JOIN partner_orders po ON po.id = pol.order_id
                WHERE po.partner_id = psi.partner_id
                  AND po.status IN ('DELIVERED','COMPLETED','DISPATCHED')
                  AND pol.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
            ), 0) AS purchases_in,
            COALESCE((
                SELECT SUM(sli.quantity)
                FROM sales_invoice_line_items sli JOIN sales_invoices si ON si.id = sli.invoice_id
                WHERE si.partner_id = psi.partner_id AND si.status = 'CONFIRMED'
                  AND sli.item_id = psi.stock_item_id AND psi.stock_item_id IS NOT NULL
            ), 0) AS sales_out,
            COALESCE((
                SELECT SUM(adj.qty) FROM partner_stock_adjustments adj
                WHERE adj.partner_id = psi.partner_id AND adj.partner_stock_item_id = psi.id
            ), 0) AS adj_total
        FROM partner_stock_items psi
        JOIN official_partners op ON op.id = psi.partner_id
        WHERE {where}
        ORDER BY op.partner_code, psi.item_name
        LIMIT :lim OFFSET :off
    """), {**params, "lim": per_page, "off": (page - 1) * per_page}).fetchall()

    data = []
    for r in rows:
        current_qty = float(r.opening_qty or 0) + float(r.purchases_in or 0) - float(r.sales_out or 0) + float(r.adj_total or 0)
        is_low = current_qty <= float(r.reorder_level or 0) and float(r.reorder_level or 0) > 0
        if low_stock_only and not is_low:
            continue
        stock_value = current_qty * float(r.selling_price or 0) if r.selling_price else 0
        data.append({
            "id": r.id,
            "partner_id": r.partner_id,
            "partner_code": r.partner_code,
            "partner_name": r.partner_name,
            "city": r.city,
            "item_type": r.item_type,
            "stock_item_id": r.stock_item_id,
            "item_name": r.item_name,
            "item_code": r.item_code,
            "unit_of_measure": r.unit_of_measure,
            "hsn_code": r.hsn_code,
            "opening_qty": float(r.opening_qty or 0),
            "purchases_in": float(r.purchases_in or 0),
            "sales_out": float(r.sales_out or 0),
            "adj_total": float(r.adj_total or 0),
            "current_qty": round(current_qty, 2),
            "reorder_level": float(r.reorder_level or 0),
            "is_low_stock": is_low,
            "selling_price": float(r.selling_price or 0) if r.selling_price else None,
            "stock_value": round(stock_value, 2),
        })

    # Summary across filtered set
    total_value = sum(d["stock_value"] for d in data)
    low_count = sum(1 for d in data if d["is_low_stock"])
    partners_with_stock = len(set(d["partner_id"] for d in data))

    return {
        "success": True,
        "total": int(total),
        "page": page,
        "per_page": per_page,
        "data": data,
        "summary": {
            "total_records": len(data),
            "total_stock_value": round(total_value, 2),
            "low_stock_count": low_count,
            "partners_with_stock": partners_with_stock,
        },
    }


@router.get("/staff-view/partners-list")
async def staff_view_partners_list(
    request: Request,
    db: Session = Depends(get_db),
):
    """DC_PARTNER_STOCK_001: Return lightweight list of all partners for staff stock filter."""
    from sqlalchemy import text as sq_text
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Staff authentication required")
    rows = db.execute(sq_text("""
        SELECT id, partner_code, partner_name, city
        FROM official_partners
        WHERE is_active = TRUE
        ORDER BY partner_code
    """)).fetchall()
    return {"success": True, "data": [{"id": r.id, "partner_code": r.partner_code, "partner_name": r.partner_name, "city": r.city} for r in rows]}


# ─────────────── Staff: All Partner Invoices (read-only) ─────

@router.get("/staff-view/partner-invoices")
async def staff_all_partner_invoices(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    partner_code: Optional[str] = Query(None),
    partner_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    payment_status: Optional[str] = Query(None),
    date_filter: str = Query("mtd"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    doc_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Staff endpoint — view all partner-created invoices. Requires staff token."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Staff token required")
    try:
        payload = SecurityManager.verify_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    today = date.today()
    if date_filter == "ftd":
        d_from = d_to = today
    elif date_filter == "yesterday":
        d_from = d_to = today - timedelta(days=1)
    elif date_filter == "mtd":
        d_from = today.replace(day=1); d_to = today
    elif date_filter == "fty":
        d_from = date(today.year if today.month >= 4 else today.year - 1, 4, 1)
        d_to = today
    else:
        d_from = date.fromisoformat(date_from) if date_from else today.replace(day=1)
        d_to = date.fromisoformat(date_to) if date_to else today

    query = db.query(SalesInvoice, OfficialPartner).join(
        OfficialPartner, SalesInvoice.partner_id == OfficialPartner.id
    ).filter(
        SalesInvoice.created_by_type == "PARTNER",
        SalesInvoice.invoice_date >= d_from,
        SalesInvoice.invoice_date <= d_to,
    )

    if partner_id:
        query = query.filter(SalesInvoice.partner_id == partner_id)
    elif partner_code:
        query = query.filter(OfficialPartner.partner_code.ilike(f"%{partner_code}%"))
    if doc_type:
        query = query.filter(SalesInvoice.document_type == doc_type)
    if status:
        query = query.filter(SalesInvoice.status == status.upper())
    if payment_status:
        query = query.filter(SalesInvoice.payment_status == payment_status.upper())
    if search:
        query = query.filter(or_(
            SalesInvoice.customer_name.ilike(f"%{search}%"),
            SalesInvoice.invoice_number.ilike(f"%{search}%"),
        ))

    total = query.count()
    rows = query.order_by(SalesInvoice.invoice_date.desc(), SalesInvoice.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    data = []
    for inv, p in rows:
        d = _invoice_to_dict(inv, db)
        d["partner_name"] = p.partner_name
        d["partner_code"] = p.partner_code
        data.append(d)

    all_rows_q = db.query(SalesInvoice).filter(
        SalesInvoice.created_by_type == "PARTNER",
        SalesInvoice.invoice_date >= d_from,
        SalesInvoice.invoice_date <= d_to,
    )
    all_inv = all_rows_q.all()
    summary = {
        "total_value": sum(float(i.grand_total or 0) for i in all_inv),
        "pending_payment": sum(float(i.balance_due or 0) for i in all_inv if i.payment_status != "PAID"),
        "confirmed": all_rows_q.filter(SalesInvoice.status == "CONFIRMED").count(),
        "draft": all_rows_q.filter(SalesInvoice.status == "DRAFT").count(),
    }

    # Partner dropdown list
    partners_q = db.query(OfficialPartner.id, OfficialPartner.partner_name, OfficialPartner.partner_code).filter(
        OfficialPartner.is_active == True
    ).all()
    partners = [{"id": p.id, "partner_name": p.partner_name, "partner_code": p.partner_code} for p in partners_q]

    return {"success": True, "data": data, "total": total, "page": page, "per_page": per_page,
            "pages": max(1, (total + per_page - 1) // per_page), "summary": summary, "partners": partners}


@router.get("/staff-view/partner-walkins")
async def staff_all_partner_walkins(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    partner_code: Optional[str] = Query(None),
    partner_id: Optional[int] = Query(None),
    date_filter: str = Query("mtd"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    purpose: Optional[str] = Query(None),
    outcome: Optional[str] = Query(None),
    company_follow_up: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Staff endpoint — view all partner walk-ins. Requires staff token."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Staff token required")
    try:
        payload = SecurityManager.verify_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    today = date.today()
    if date_filter == "ftd":
        d_from = d_to = today
    elif date_filter == "yesterday":
        d_from = d_to = today - timedelta(days=1)
    elif date_filter == "mtd":
        d_from = today.replace(day=1); d_to = today
    elif date_filter == "fty":
        d_from = date(today.year if today.month >= 4 else today.year - 1, 4, 1)
        d_to = today
    else:
        d_from = date.fromisoformat(date_from) if date_from else today.replace(day=1)
        d_to = date.fromisoformat(date_to) if date_to else today

    from sqlalchemy import text as sq_text
    conditions = "w.visit_date BETWEEN :df AND :dt"
    params: dict = {"df": d_from, "dt": d_to}

    if partner_id:
        conditions += " AND w.partner_id = :pid"
        params["pid"] = partner_id
    elif partner_code:
        conditions += " AND p.partner_code ILIKE :pc"
        params["pc"] = f"%{partner_code}%"
    if purpose:
        conditions += " AND w.visit_purpose = :purpose"
        params["purpose"] = purpose
    if outcome:
        conditions += " AND w.visit_outcome = :outcome"
        params["outcome"] = outcome
    if company_follow_up in ("true", "false"):
        conditions += " AND w.company_follow_up = :cfu"
        params["cfu"] = company_follow_up == "true"
    if search:
        conditions += " AND (w.customer_name ILIKE :s OR w.customer_phone ILIKE :s)"
        params["s"] = f"%{search}%"

    total = db.execute(sq_text(f"""
        SELECT COUNT(*) FROM partner_walkins w
        JOIN official_partners p ON p.id = w.partner_id
        WHERE {conditions}
    """), params).scalar()

    params["limit"] = per_page
    params["offset"] = (page - 1) * per_page
    rows = db.execute(sq_text(f"""
        SELECT w.id, w.visit_date, w.visit_time, w.customer_name, w.customer_phone,
               w.visit_purpose, w.visit_outcome, w.company_follow_up, w.crm_lead_id,
               w.registered_with_mnr, w.registered_with_vgk, w.is_new_customer, w.is_existing_customer,
               w.notes, w.assigned_to, w.created_at,
               p.partner_code, p.partner_name
        FROM partner_walkins w
        JOIN official_partners p ON p.id = w.partner_id
        WHERE {conditions}
        ORDER BY w.visit_date DESC, w.id DESC
        LIMIT :limit OFFSET :offset
    """), params).fetchall()

    cols = ["id","visit_date","visit_time","customer_name","customer_phone",
            "visit_purpose","visit_outcome","company_follow_up","crm_lead_id",
            "registered_with_mnr","registered_with_vgk","is_new_customer","is_existing_customer",
            "notes","assigned_to","created_at","partner_code","partner_name"]
    data = []
    for r in rows:
        d = dict(zip(cols, r))
        for k in ["visit_date","created_at"]:
            if d.get(k) and hasattr(d[k], "isoformat"):
                d[k] = d[k].isoformat()
        if d.get("visit_time"):
            d["visit_time"] = str(d["visit_time"])
        data.append(d)

    # Summary
    summary_row = db.execute(sq_text(f"""
        SELECT
            COUNT(*) as total,
            COUNT(*) FILTER (WHERE w.company_follow_up=true) as company_fu,
            COUNT(*) FILTER (WHERE w.visit_outcome='converted') as converted,
            COUNT(*) FILTER (WHERE w.visit_date = CURRENT_DATE) as today,
            COUNT(*) FILTER (WHERE w.visit_outcome='interested') as interested,
            COUNT(*) FILTER (WHERE w.visit_outcome='not_interested') as not_interested
        FROM partner_walkins w
        JOIN official_partners p ON p.id = w.partner_id
        WHERE {conditions}
    """), {k: v for k, v in params.items() if k not in ("limit","offset")}).fetchone()

    # Partner list for dropdown
    partners_q = db.query(OfficialPartner.id, OfficialPartner.partner_name, OfficialPartner.partner_code).filter(
        OfficialPartner.is_active == True
    ).all()
    partners = [{"id": p.id, "partner_name": p.partner_name, "partner_code": p.partner_code} for p in partners_q]

    return {
        "success": True,
        "data": data,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, (total + per_page - 1) // per_page),
        "partners": partners,
        "summary": {
            "total": int(summary_row[0] or 0),
            "company_followups": int(summary_row[1] or 0),
            "converted": int(summary_row[2] or 0),
            "today": int(summary_row[3] or 0),
            "interested": int(summary_row[4] or 0),
            "not_interested": int(summary_row[5] or 0),
        },
    }


# ─── Staff Detail Endpoints ───

@router.get("/staff-view/partner-invoices/{invoice_id}")
async def staff_get_partner_invoice(invoice_id: int, request: Request, db: Session = Depends(get_db)):
    """Staff endpoint — view a single partner-created invoice by ID."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Staff token required")
    try:
        payload = SecurityManager.verify_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    inv = db.query(SalesInvoice).filter(
        SalesInvoice.id == invoice_id,
        SalesInvoice.created_by_type == "PARTNER",
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")

    partner = db.query(OfficialPartner).filter(OfficialPartner.id == inv.partner_id).first()
    d = _invoice_to_dict(inv, db)
    if partner:
        d["partner_name"] = partner.partner_name
        d["partner_code"] = partner.partner_code

    return {"success": True, "data": d}


@router.get("/staff-view/partner-walkins/{walkin_id}")
async def staff_get_partner_walkin(walkin_id: int, request: Request, db: Session = Depends(get_db)):
    """Staff endpoint — view a single partner walk-in by ID."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Staff token required")
    try:
        payload = SecurityManager.verify_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    from sqlalchemy import text as sq_text2
    row = db.execute(sq_text2("""
        SELECT w.*, p.partner_name, p.partner_code
        FROM partner_walkins w
        JOIN official_partners p ON p.id = w.partner_id
        WHERE w.id = :wid
    """), {"wid": walkin_id}).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Walk-in not found")

    d = dict(row._mapping)
    for k in ("visit_date", "follow_up_date"):
        if d.get(k) and hasattr(d[k], "isoformat"):
            d[k] = d[k].isoformat()
    if d.get("visit_time"):
        d["visit_time"] = str(d["visit_time"])

    return {"success": True, "data": d}


# ─── [DC-SOLAR-VENDOR-001] Solar Vendor Portal Endpoints ───

def _compute_vendor_status(status: str, solar_pipeline_status: str) -> str:
    """Compute vendor-facing status badge from lead status + solar pipeline status."""
    sps = (solar_pipeline_status or '').lower()
    st  = (status or '').lower()
    _CANCELLED = {'loan_rejected', 'diff_vendor_loan_rejected', 'not_interested', 'cancelled', 'different_vendor'}
    _FINAL     = {'subsidy_cleared', 'bank_loan_completed'}
    _COMPLETED = {'installed', 'net_meter_pending', 'completed'}
    if st in ('lost', 'cancelled') or sps in _CANCELLED:
        return 'Cancelled'
    elif sps in _FINAL:
        return 'Final Invoice Completed'
    elif sps in _COMPLETED:
        return 'Completed'
    else:
        return 'Pending'


@router.get("/solar-vendor/stats")
async def solar_vendor_stats(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db)
):
    """[DC-SOLAR-VENDOR-001] Dashboard stats for a solar vendor partner portal."""
    import logging
    logger = logging.getLogger(__name__)

    # Resolve VendorMaster.id from legacy_vendor_id
    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        # Try from token
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass

    if not vendor_ref_id or partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    leads = db.query(
        CRMLead.id,
        CRMLead.status,
        CRMLead.solar_pipeline_status,
        CRMLead.deal_value_total,
        CRMLead.deal_value_received,
        CRMLead.deal_value_balance,
        CRMLead.kw_size,
    ).filter(
        CRMLead.vendor_id == vendor_ref_id,
    ).all()

    total = len(leads)
    by_status = {"Pending": 0, "Completed": 0, "Final Invoice Completed": 0, "Cancelled": 0}
    by_pipeline: dict = {}
    total_deal_value = 0.0
    total_deal_received = 0.0
    total_deal_balance = 0.0
    deal_value_by_status: dict = {"Pending": 0.0, "Completed": 0.0, "Final Invoice Completed": 0.0, "Cancelled": 0.0}
    deal_value_by_pipeline: dict = {}
    total_kw = 0.0

    for l in leads:
        vs = _compute_vendor_status(l.status, l.solar_pipeline_status)
        by_status[vs] = by_status.get(vs, 0) + 1

        sps = l.solar_pipeline_status or "unknown"
        by_pipeline[sps] = by_pipeline.get(sps, 0) + 1

        dv = float(l.deal_value_total or 0)
        dr = float(l.deal_value_received or 0)
        db_ = float(l.deal_value_balance or 0)
        total_deal_value += dv
        total_deal_received += dr
        total_deal_balance += db_
        deal_value_by_status[vs] = deal_value_by_status.get(vs, 0.0) + dv
        deal_value_by_pipeline[sps] = deal_value_by_pipeline.get(sps, 0.0) + dv

        if l.kw_size:
            try:
                total_kw += float(str(l.kw_size).upper().replace("KW", "").strip())
            except Exception:
                pass

    return {
        "success": True,
        "vendor_ref_id": vendor_ref_id,
        "total_leads": total,
        "by_vendor_status": by_status,
        "by_pipeline": by_pipeline,
        "total_deal_value": round(total_deal_value, 2),
        "total_deal_received": round(total_deal_received, 2),
        "total_deal_balance": round(total_deal_balance, 2),
        "deal_value_by_status": {k: round(v, 2) for k, v in deal_value_by_status.items()},
        "deal_value_by_pipeline": {k: round(v, 2) for k, v in deal_value_by_pipeline.items()},
        "total_kw": round(total_kw, 2),
    }


@router.get("/solar-vendor/leads/{lead_id}/docs")
async def solar_vendor_lead_docs(
    lead_id: int,
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """[DC-SOLAR-VENDOR-001] Get solar documents for a vendor's lead (vendor-accessible)."""
    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass

    if not vendor_ref_id or partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    # Verify this lead belongs to the vendor
    lead_check = db.execute(text(
        "SELECT id FROM crm_leads WHERE id=:lid AND vendor_id=:vid"
    ), {"lid": lead_id, "vid": vendor_ref_id}).fetchone()
    if not lead_check:
        raise HTTPException(status_code=404, detail="Lead not found")

    rows = db.execute(text("""
        SELECT id, doc_category, doc_type, doc_label, doc_number,
               file_name, original_name, file_size, uploaded_by_name, uploaded_at
        FROM crm_lead_solar_documents
        WHERE lead_id = :lid
        ORDER BY doc_category, uploaded_at DESC
    """), {"lid": lead_id}).fetchall()

    docs = []
    for r in rows:
        docs.append({
            "id":               r.id,
            "doc_category":     r.doc_category,
            "doc_type":         r.doc_type,
            "doc_label":        r.doc_label or r.doc_type,
            "doc_number":       r.doc_number or "",
            "file_name":        r.file_name,
            "original_name":    r.original_name,
            "file_size":        r.file_size,
            "uploaded_by_name": r.uploaded_by_name,
            "uploaded_at":      r.uploaded_at.isoformat() if r.uploaded_at else None,
            "view_url":         f"/storage/{r.file_name}" if r.file_name else None,
        })

    return {"success": True, "docs": docs, "lead_id": lead_id}


@router.get("/solar-vendor/leads")
async def solar_vendor_leads(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
    page: int = 1,
    per_page: int = 50,
    vendor_status_filter: Optional[str] = Query(None),
    pipeline_stage: Optional[str] = Query(None),
    pincode: Optional[str] = Query(None),
    loan_bank: Optional[str] = Query(None),
    bank_entry: Optional[str] = Query(None),   # "yes" | "no"
    has_co_applicant: Optional[str] = Query(None),  # "yes" | "no"
    search: Optional[str] = Query(None),
):
    """[DC-SOLAR-VENDOR-001] Read-only solar leads for vendor portal."""
    import logging
    logger = logging.getLogger(__name__)

    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass

    if not vendor_ref_id or partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    from app.models.staff import StaffEmployee as _SE

    # ── SQL-level filters ──────────────────────────────────────────────────
    _CANCELLED_SPS = ('loan_rejected', 'diff_vendor_loan_rejected', 'not_interested', 'cancelled', 'different_vendor')
    _FINAL_SPS     = ('subsidy_cleared', 'bank_loan_completed')
    _COMPLETED_SPS = ('installed', 'net_meter_pending', 'completed')
    _CANCELLED_ST  = ('lost', 'cancelled')

    q = db.query(CRMLead).filter(CRMLead.vendor_id == vendor_ref_id)

    if search:
        _s = f"%{search}%"
        q = q.filter(or_(CRMLead.name.ilike(_s), CRMLead.phone.ilike(_s)))

    if pipeline_stage:
        q = q.filter(CRMLead.solar_pipeline_status == pipeline_stage)

    if pincode:
        q = q.filter(CRMLead.pincode.ilike(f"%{pincode}%"))

    if loan_bank:
        q = q.filter(CRMLead.loan_bank.ilike(f"%{loan_bank}%"))

    if bank_entry == "yes":
        q = q.filter(CRMLead.bank_entry_updated == True)
    elif bank_entry == "no":
        q = q.filter(or_(CRMLead.bank_entry_updated == False, CRMLead.bank_entry_updated.is_(None)))

    if has_co_applicant == "yes":
        q = q.filter(CRMLead.co_applicant_name.isnot(None), CRMLead.co_applicant_name != '')
    elif has_co_applicant == "no":
        q = q.filter(or_(CRMLead.co_applicant_name.is_(None), CRMLead.co_applicant_name == ''))

    # vendor_status_filter translated to SQL conditions
    if vendor_status_filter == 'Cancelled':
        q = q.filter(or_(
            CRMLead.status.in_(_CANCELLED_ST),
            CRMLead.solar_pipeline_status.in_(_CANCELLED_SPS)
        ))
    elif vendor_status_filter == 'Final Invoice Completed':
        q = q.filter(
            ~CRMLead.status.in_(_CANCELLED_ST),
            ~CRMLead.solar_pipeline_status.in_(_CANCELLED_SPS),
            CRMLead.solar_pipeline_status.in_(_FINAL_SPS)
        )
    elif vendor_status_filter == 'Completed':
        q = q.filter(
            ~CRMLead.status.in_(_CANCELLED_ST),
            ~CRMLead.solar_pipeline_status.in_(_CANCELLED_SPS),
            ~CRMLead.solar_pipeline_status.in_(_FINAL_SPS),
            CRMLead.solar_pipeline_status.in_(_COMPLETED_SPS)
        )
    elif vendor_status_filter == 'Pending':
        q = q.filter(
            ~CRMLead.status.in_(_CANCELLED_ST),
            or_(
                CRMLead.solar_pipeline_status.is_(None),
                ~CRMLead.solar_pipeline_status.in_(list(_CANCELLED_SPS) + list(_FINAL_SPS) + list(_COMPLETED_SPS))
            )
        )

    total = q.count()
    leads = q.order_by(CRMLead.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    # Staff enrichment
    staff_ids = {l.telecaller_id for l in leads if l.telecaller_id} | {l.field_staff_id for l in leads if l.field_staff_id}
    staff_map = {}
    if staff_ids:
        for s in db.query(_SE).filter(_SE.id.in_(staff_ids)).all():
            nm = s.full_name or f"{s.first_name or ''} {s.last_name or ''}".strip() or s.emp_code
            staff_map[s.id] = nm

    rows = []
    for l in leads:
        vs = _compute_vendor_status(l.status, l.solar_pipeline_status)
        rows.append({
            "id": l.id,
            "name": l.name,
            "phone": l.phone[-4:].rjust(10, '•') if l.phone else None,
            "status": l.status,
            "solar_pipeline_status": l.solar_pipeline_status,
            "vendor_status": vs,
            "telecaller_name": staff_map.get(l.telecaller_id),
            "field_staff_name": staff_map.get(l.field_staff_id),
            "pincode": l.pincode,
            "loan_bank": l.loan_bank,
            "bank_entry_updated": l.bank_entry_updated,
            "bank_statement_available": l.bank_statement_available,
            "monthly_income": float(l.monthly_income) if l.monthly_income is not None else None,
            "regular_income_available": l.regular_income_available,
            "co_applicant_name": l.co_applicant_name,
            "co_applicant_phone": l.co_applicant_phone,
            "co_applicant_aadhaar": '••••' + str(l.co_applicant_aadhaar)[-4:] if l.co_applicant_aadhaar else None,
            "co_applicant_pan": l.co_applicant_pan,
            "co_applicant_bank_account": '••••' + str(l.co_applicant_bank_account)[-4:] if l.co_applicant_bank_account else None,
            "co_applicant_ifsc": l.co_applicant_ifsc,
            "material_reach_date": l.material_reach_date.isoformat() if l.material_reach_date else None,
            "installation_date": l.installation_date.isoformat() if l.installation_date else None,
            "existing_association": l.existing_association,
            "deal_value_total": float(l.deal_value_total) if l.deal_value_total else None,
            "next_followup_date": l.next_followup_date.isoformat() if l.next_followup_date else None,
            "accepted_date": l.accepted_date.isoformat() if l.accepted_date else None,
            "kw_size": l.kw_size,
            "application_no": l.application_no,
        })

    return {
        "success": True,
        "data": rows,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, (total + per_page - 1) // per_page),
    }


@router.get("/solar-vendor/profile")
async def solar_vendor_profile(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """[DC-SOLAR-VENDOR-001] Full vendor profile including stamps/signatures from VendorMaster."""
    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass

    if not vendor_ref_id or partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    vendor = db.query(VendorMaster).filter(VendorMaster.id == vendor_ref_id).first()
    if not vendor:
        return {"success": True, "data": {}}

    return {
        "success": True,
        "data": {
            "vendor_code": getattr(vendor, 'vendor_code', None),
            "vendor_name": getattr(vendor, 'vendor_name', None),
            "phone": getattr(vendor, 'phone', None),
            "email": getattr(vendor, 'email', None),
            "city": getattr(vendor, 'city', None),
            "state": getattr(vendor, 'state', None),
            "address": getattr(vendor, 'address', None),
            "gst_number": getattr(vendor, 'gst_number', None),
            "pan_number": getattr(vendor, 'pan_number', None),
            "bank_name": getattr(vendor, 'bank_name', None),
            "account_number": getattr(vendor, 'account_number', None),
            "ifsc_code": getattr(vendor, 'ifsc_code', None),
            "vendor_logo_url": getattr(vendor, 'vendor_logo_url', None),
            "stamp_image_url": getattr(vendor, 'stamp_image_url', None),
            "tech_signature_url": getattr(vendor, 'tech_signature_url', None),
            "rep_signature_url": getattr(vendor, 'rep_signature_url', None),
        }
    }


@router.post("/solar-vendor/upload-media")
async def solar_vendor_upload_media(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
    media_type: str = Form(...),
    file: UploadFile = File(...),
):
    """[DC-SOLAR-VENDOR-001] Upload stamp or signature image for the vendor portal."""
    if partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass

    if not vendor_ref_id:
        raise HTTPException(status_code=403, detail="Vendor reference not found")

    _ALLOWED = {'logo', 'stamp', 'tech_sig', 'rep_sig'}
    if media_type not in _ALLOWED:
        raise HTTPException(status_code=400, detail="media_type must be one of: logo, stamp, tech_sig, rep_sig")

    vendor = db.query(VendorMaster).filter(VendorMaster.id == vendor_ref_id).first()
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    # Validate file type — images only
    _fname = (file.filename or '').lower()
    if not any(_fname.endswith(ext) for ext in ('.jpg', '.jpeg', '.png', '.webp')):
        raise HTTPException(status_code=400, detail="Only image files are allowed (JPG, PNG, WEBP)")

    file_data = await file.read()
    if len(file_data) == 0:
        raise HTTPException(status_code=400, detail="File is empty")
    if len(file_data) > 3 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 3MB)")

    import uuid, pathlib, os, mimetypes
    _ext = pathlib.Path(_fname).suffix or '.jpg'
    _storage_key = f"vendor_media/{vendor_ref_id}/{media_type}_{uuid.uuid4().hex[:8]}{_ext}"

    # Try object storage first, fall back to local frontend/storage
    _saved_path = None
    try:
        from app.services.object_storage import storage_service as _ss
        ok = _ss.upload_file(_storage_key, file_data)
        if ok:
            _saved_path = _storage_key
    except Exception:
        pass

    if not _saved_path:
        # Local fallback
        _local_dir = pathlib.Path("frontend/storage") / f"vendor_media/{vendor_ref_id}"
        _local_dir.mkdir(parents=True, exist_ok=True)
        _local_file = _local_dir / f"{media_type}_{uuid.uuid4().hex[:8]}{_ext}"
        with open(_local_file, "wb") as f:
            f.write(file_data)
        _saved_path = str(_local_file).replace("frontend/storage/", "")

    _view_url = f"/storage/{_saved_path}"

    # Update vendor_master field
    _field_map = {'logo': 'vendor_logo_url', 'stamp': 'stamp_image_url', 'tech_sig': 'tech_signature_url', 'rep_sig': 'rep_signature_url'}
    setattr(vendor, _field_map[media_type], _view_url)
    db.commit()

    return {"success": True, "url": _view_url, "media_type": media_type}


@router.get("/solar-vendor/ledger")
async def solar_vendor_ledger_list(
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    direction: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-SOLAR-VENDOR-LEDGER-001: Solar vendor views their own ledger (RECEIVED + RETURNED)."""
    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass
    if not vendor_ref_id or partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    q = db.query(SolarVendorLedger).filter(SolarVendorLedger.solar_vendor_id == int(vendor_ref_id))
    if date_from:
        q = q.filter(SolarVendorLedger.transaction_date >= date_from)
    if date_to:
        q = q.filter(SolarVendorLedger.transaction_date <= date_to)
    if direction:
        q = q.filter(SolarVendorLedger.direction == direction.upper())
    total = q.count()
    rows = q.order_by(SolarVendorLedger.transaction_date.desc(), SolarVendorLedger.id.desc()) \
            .offset((page - 1) * page_size).limit(page_size).all()

    running_balance = 0.0
    result = []
    for r in rows:
        amt = float(r.amount)
        if r.direction == 'RECEIVED':
            running_balance += amt
        else:
            running_balance -= amt
        result.append({
            "id": r.id,
            "transaction_date": r.transaction_date.isoformat() if r.transaction_date else None,
            "entry_number": r.entry_number,
            "customer_name": r.customer_name,
            "amount": amt,
            "direction": r.direction,
            "utr_reference": r.utr_reference,
            "payment_mode": r.payment_mode,
            "notes": r.notes,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    # Summary over all matching rows (not just this page)
    all_q2 = db.query(SolarVendorLedger).filter(SolarVendorLedger.solar_vendor_id == int(vendor_ref_id))
    if date_from:
        all_q2 = all_q2.filter(SolarVendorLedger.transaction_date >= date_from)
    if date_to:
        all_q2 = all_q2.filter(SolarVendorLedger.transaction_date <= date_to)
    all_rows2 = all_q2.all()
    total_received = sum(float(r.amount) for r in all_rows2 if r.direction == 'RECEIVED')
    total_returned = sum(float(r.amount) for r in all_rows2 if r.direction == 'RETURNED')
    return {
        "success": True,
        "rows": result,
        "summary": {
            "total_received": round(total_received, 2),
            "total_returned": round(total_returned, 2),
            "balance": round(total_received - total_returned, 2),
            "count_received": sum(1 for r in all_rows2 if r.direction == 'RECEIVED'),
            "count_returned": sum(1 for r in all_rows2 if r.direction == 'RETURNED'),
        },
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/solar-vendor/ledger/return")
async def solar_vendor_record_return(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-SOLAR-VENDOR-LEDGER-001: Vendor records a payment they are returning to MNR."""
    vendor_ref_id = partner.legacy_vendor_id
    if not vendor_ref_id:
        try:
            from jose import jwt as _jwt
            _auth = request.headers.get("Authorization", "")
            if _auth.startswith("Bearer "):
                _payload = _jwt.decode(_auth.split(" ")[1], settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                vendor_ref_id = _payload.get("vendor_ref_id")
        except Exception:
            pass
    if not vendor_ref_id or partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    from decimal import Decimal
    from datetime import date as _date
    from app.models.base import get_indian_time

    amount_raw = body.get("amount", 0)
    amount = Decimal(str(amount_raw))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be > 0")
    txn_date_str = body.get("transaction_date")
    if not txn_date_str:
        raise HTTPException(status_code=400, detail="transaction_date required")
    txn_date = _date.fromisoformat(txn_date_str)

    row = SolarVendorLedger(
        solar_vendor_id=int(vendor_ref_id),
        income_entry_id=None,
        entry_number=None,
        transaction_date=txn_date,
        customer_name=body.get("customer_name"),
        amount=amount,
        company_id=None,
        direction='RETURNED',
        utr_reference=body.get("utr_reference"),
        payment_mode=body.get("payment_mode"),
        notes=body.get("notes"),
        created_by_id=None,
        created_at=get_indian_time(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"success": True, "message": "Return payment recorded successfully", "id": row.id}


@router.post("/solar-vendor/change-password")
async def solar_vendor_change_password(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """[DC-SOLAR-VENDOR-001] Change password for solar vendor portal account."""
    if partner.category != 'VENDOR':
        raise HTTPException(status_code=403, detail="Not a solar vendor account")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request body")

    old_password = body.get("old_password", "").strip()
    new_password = body.get("new_password", "").strip()

    if not old_password or not new_password:
        raise HTTPException(status_code=400, detail="Both current and new password are required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")

    if not SecurityManager.verify_password(old_password, partner.password_hash or ''):
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    partner.password_hash = SecurityManager.get_password_hash(new_password)
    db.commit()
    return {"success": True, "message": "Password changed successfully"}


# ─────────────────────────────────────────────────────────────────────────────
# DC-PARTNER-SPARE-001: Partner Spare Parts Ordering Endpoints
# Partner can browse spare catalog and submit purchase requests.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/auth/spare-catalog")
async def partner_spare_catalog(
    search: Optional[str] = Query(None),
    skip:   int = Query(0, ge=0),
    limit:  int = Query(100, ge=1, le=200),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """
    DC-PARTNER-SPARE-001: Browse spare parts catalog.
    Returns all SPARE_PART stock items with current stock levels.
    """
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        items, total = SpareProcurementService.partner_get_spare_catalog(
            db,
            partner_id=partner.id,
            company_id=partner.company_id if hasattr(partner, 'company_id') else None,
            search=search,
            skip=skip,
            limit=limit,
        )
        return {"success": True, "items": items, "total": total, "skip": skip, "limit": limit}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/auth/spare-requests")
async def partner_submit_spare_request(
    request: Request,
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """
    DC-PARTNER-SPARE-001: Submit a spare parts request.
    Body: {items: [{item_id, quantity, uom?, notes?}], notes?, company_id}
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request body")

    items = body.get("items", [])
    notes = body.get("notes", "")
    company_id = body.get("company_id") or getattr(partner, 'company_id', None)

    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if not items:
        raise HTTPException(status_code=400, detail="At least one item is required")

    try:
        from app.services.staff_accounts_service import SpareProcurementService, AccountsValidationError
        req = SpareProcurementService.partner_submit_spare_request(
            db,
            partner_id=partner.id,
            company_id=company_id,
            items=items,
            notes=notes,
        )
        return {
            "success": True,
            "message": f"Request {req.request_number} submitted successfully",
            "request": req.to_dict(),
        }
    except AccountsValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/auth/spare-requests")
async def partner_list_spare_requests(
    status: Optional[str] = Query(None),
    skip:   int = Query(0, ge=0),
    limit:  int = Query(50, ge=1, le=100),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-SPARE-001: List own spare part requests with status filter."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService
        reqs, total = SpareProcurementService.partner_list_spare_requests(
            db, partner_id=partner.id, status=status, skip=skip, limit=limit
        )
        return {
            "success": True,
            "requests": [r.to_dict() for r in reqs],
            "total": total,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/auth/spare-requests/{request_id}/cancel")
async def partner_cancel_spare_request(
    request_id: int = Path(..., ge=1),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-SPARE-001: Cancel own SUBMITTED spare request."""
    try:
        from app.services.staff_accounts_service import SpareProcurementService, AccountsValidationError, AccountsNotFoundError
        req = SpareProcurementService.partner_cancel_spare_request(db, partner.id, request_id)
        return {"success": True, "message": "Request cancelled", "request": req.to_dict()}
    except (AccountsValidationError, AccountsNotFoundError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── DC-PARTNER-SFMS-001: SFMS Sales Invoices ───────────────────────────────

@router.get("/sfms-invoices")
async def partner_sfms_invoices(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-SFMS-001: SFMS sales invoices raised by staff for this partner."""
    try:
        from sqlalchemy import text as _st
        conditions = ["si.partner_id = :pid", "si.status != 'CANCELLED'"]
        params: Dict[str, Any] = {"pid": partner.id}
        if search:
            conditions.append("(si.invoice_number ILIKE :s OR si.customer_name ILIKE :s)")
            params["s"] = f"%{search}%"
        if status:
            conditions.append("si.payment_status = :ps")
            params["ps"] = status.upper()
        if date_from:
            conditions.append("si.invoice_date >= :df")
            params["df"] = date_from
        if date_to:
            conditions.append("si.invoice_date <= :dt")
            params["dt"] = date_to
        where = " AND ".join(conditions)
        total = db.execute(_st(f"SELECT COUNT(*) FROM sales_invoices si WHERE {where}"), params).scalar() or 0
        rows = db.execute(_st(f"""
            SELECT si.id, si.invoice_number, si.invoice_date, si.due_date,
                   si.grand_total, si.amount_received, si.balance_due,
                   si.payment_status, si.status, si.dispatch_status,
                   ac.company_name, ac.company_code
            FROM sales_invoices si
            LEFT JOIN associated_companies ac ON ac.id = si.company_id
            WHERE {where}
            ORDER BY si.invoice_date DESC, si.id DESC
            LIMIT :lim OFFSET :off
        """), {**params, "lim": per_page, "off": (page - 1) * per_page}).fetchall()
        return {
            "success": True, "total": total, "page": page, "per_page": per_page,
            "data": [{
                "id": r.id, "invoice_number": r.invoice_number,
                "invoice_date": r.invoice_date.isoformat() if r.invoice_date else None,
                "due_date": r.due_date.isoformat() if r.due_date else None,
                "grand_total": float(r.grand_total or 0),
                "amount_received": float(r.amount_received or 0),
                "balance_due": float(r.balance_due or 0),
                "payment_status": r.payment_status, "status": r.status,
                "dispatch_status": r.dispatch_status,
                "company_name": r.company_name, "company_code": r.company_code,
            } for r in rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sfms-invoices/{invoice_id}")
async def partner_sfms_invoice_detail(
    invoice_id: int = Path(..., ge=1),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-SFMS-001: Detail view — items with qty/UOM only, no unit rates."""
    try:
        from sqlalchemy import text as _st
        row = db.execute(_st("""
            SELECT si.id, si.invoice_number, si.invoice_date, si.due_date,
                   si.grand_total, si.amount_received, si.balance_due,
                   si.payment_status, si.status, si.subtotal, si.total_tax,
                   si.dispatch_status, si.remarks, si.e_way_bill_number,
                   ac.company_name, ac.company_code
            FROM sales_invoices si
            LEFT JOIN associated_companies ac ON ac.id = si.company_id
            WHERE si.id = :iid AND si.partner_id = :pid
        """), {"iid": invoice_id, "pid": partner.id}).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Invoice not found")
        lines = db.execute(_st("""
            SELECT item_description, item_code, hsn_code, quantity, unit_of_measure
            FROM sales_invoice_line_items
            WHERE invoice_id = :iid ORDER BY line_number
        """), {"iid": invoice_id}).fetchall()
        return {
            "success": True,
            "data": {
                "id": row.id, "invoice_number": row.invoice_number,
                "invoice_date": row.invoice_date.isoformat() if row.invoice_date else None,
                "due_date": row.due_date.isoformat() if row.due_date else None,
                "grand_total": float(row.grand_total or 0),
                "amount_received": float(row.amount_received or 0),
                "balance_due": float(row.balance_due or 0),
                "payment_status": row.payment_status, "status": row.status,
                "subtotal": float(row.subtotal or 0),
                "total_tax": float(row.total_tax or 0),
                "dispatch_status": row.dispatch_status,
                "remarks": row.remarks, "e_way_bill_number": row.e_way_bill_number,
                "company_name": row.company_name, "company_code": row.company_code,
                "line_items": [{
                    "item_name": ln.item_description, "item_code": ln.item_code,
                    "hsn_code": ln.hsn_code,
                    "quantity": float(ln.quantity or 0),
                    "unit_of_measure": ln.unit_of_measure,
                } for ln in lines],
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sfms-invoices/{invoice_id}/pdf")
async def download_sfms_invoice_pdf(
    invoice_id: int = Path(..., ge=1),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-SFMS-PDF-001: Download PDF for a company→partner (SFMS) invoice.
    Seller = the company that raised the invoice (MNR/VGK/EA).
    Buyer  = the partner (appears in Bill To from invoice customer fields).
    No seller_override — uses the company's own details as seller.
    """
    from sqlalchemy import text as _st
    row = db.execute(_st("""
        SELECT id, invoice_number, company_id
        FROM sales_invoices
        WHERE id = :iid AND partner_id = :pid
          AND (created_by_type IS NULL OR created_by_type != 'PARTNER')
    """), {"iid": invoice_id, "pid": partner.id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="SFMS invoice not found")
    try:
        from app.services.staff_accounts_service import SalesInvoiceService
        company_id = row.company_id or 1
        pdf_bytes = SalesInvoiceService.generate_pdf(db, invoice_id, company_id)
        filename = f"{row.invoice_number.replace('/', '-')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


# ─── DC-PARTNER-PENDING-001: Pending Dispatch for partner ───────────────────

@router.get("/pending-dispatch")
async def partner_pending_dispatch(
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-PENDING-001: Items pending physical dispatch — no pricing shown."""
    try:
        from sqlalchemy import text as _st
        from app.models.staff_accounts import SalesPendingLineConfig
        invoices = db.execute(_st("""
            SELECT si.id, si.invoice_number, si.invoice_date, si.dispatch_status,
                   ac.company_name, ac.company_code
            FROM sales_invoices si
            LEFT JOIN associated_companies ac ON ac.id = si.company_id
            WHERE si.partner_id = :pid
              AND si.track_physical_dispatch = TRUE
              AND si.dispatch_status IN ('NOT_DISPATCHED', 'PARTIALLY_DISPATCHED')
              AND si.status = 'CONFIRMED'
            ORDER BY si.invoice_date DESC
        """), {"pid": partner.id}).fetchall()

        result = []
        for inv in invoices:
            lines = db.execute(_st("""
                SELECT sili.id, sili.item_description, sili.item_code,
                       sili.unit_of_measure, sili.quantity AS invoiced_qty
                FROM sales_invoice_line_items sili
                WHERE sili.invoice_id = :iid ORDER BY sili.line_number
            """), {"iid": inv.id}).fetchall()
            inv_items = []
            for ln in lines:
                cfg = db.query(SalesPendingLineConfig).filter(
                    SalesPendingLineConfig.invoice_id == inv.id,
                    SalesPendingLineConfig.invoice_line_id == ln.id
                ).first()
                target = float(cfg.pending_qty) if cfg else float(ln.invoiced_qty or 0)
                if target <= 0:
                    continue
                disp = db.execute(_st(
                    "SELECT COALESCE(SUM(dispatched_qty),0) FROM sales_dispatch_records WHERE invoice_line_id=:lid"
                ), {"lid": ln.id}).scalar() or 0
                remaining = max(0.0, target - float(disp))
                if remaining <= 0:
                    continue
                inv_items.append({
                    "item_description": ln.item_description,
                    "item_code": ln.item_code or "",
                    "unit_of_measure": ln.unit_of_measure or "",
                    "invoiced_qty": float(ln.invoiced_qty or 0),
                    "dispatched_qty": float(disp),
                    "remaining_qty": remaining,
                })
            if inv_items:
                result.append({
                    "invoice_id": inv.id,
                    "invoice_number": inv.invoice_number,
                    "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
                    "dispatch_status": inv.dispatch_status,
                    "company_name": inv.company_name,
                    "company_code": inv.company_code,
                    "items": inv_items,
                })
        return {"success": True, "invoices": result, "total_invoices": len(result)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── DC-PARTNER-CUSTSPARE-001: Customer Spares ──────────────────────────────

@router.get("/customer-spares")
async def partner_customer_spares(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    partner: OfficialPartner = Depends(get_current_partner),
    db: Session = Depends(get_db),
):
    """DC-PARTNER-CUSTSPARE-001: Customer-supplied spares from tickets linked to this partner.
    Matches via partner_id FK OR any registered phone on the partner profile."""
    try:
        from sqlalchemy import text as _st
        phones = list({p.strip() for p in [
            partner.phone, partner.whatsapp_number,
            partner.contact_person_1_phone, partner.contact_person_2_phone,
            getattr(partner, 'sales_contact_number', None),
            getattr(partner, 'service_contact_number', None),
        ] if p and str(p).strip()})

        phone_clause = ""
        phone_params: Dict[str, Any] = {}
        if phones:
            placeholders = ", ".join(f":ph{i}" for i in range(len(phones)))
            phone_clause = f"OR st.customer_phone IN ({placeholders})"
            for i, ph in enumerate(phones):
                phone_params[f"ph{i}"] = ph

        base_where = f"""
            sr.spare_source = 'customer'
            AND (st.partner_id = :pid {phone_clause})
        """
        rows = db.execute(_st(f"""
            SELECT sr.id, sr.spare_item_name, sr.spare_item_code,
                   sr.quantity_required, sr.procurement_status,
                   sr.repair_route, sr.vendor_repair_status,
                   sr.sub_ticket_number, sr.requested_at,
                   st.ticket_number, st.customer_name, st.customer_phone,
                   st.product_model AS vehicle_model
            FROM service_ticket_spare_request sr
            JOIN service_ticket st ON st.id = sr.ticket_id
            WHERE {base_where}
            ORDER BY sr.requested_at DESC
            LIMIT :lim OFFSET :off
        """), {"pid": partner.id, **phone_params, "lim": limit, "off": skip}).fetchall()

        total = db.execute(_st(f"""
            SELECT COUNT(*) FROM service_ticket_spare_request sr
            JOIN service_ticket st ON st.id = sr.ticket_id
            WHERE {base_where}
        """), {"pid": partner.id, **phone_params}).scalar() or 0

        def _status_label(r):
            if r.vendor_repair_status:
                return r.vendor_repair_status.replace('_', ' ').title()
            if r.procurement_status:
                return r.procurement_status.replace('_', ' ').title()
            return "Pending"

        return {
            "success": True, "total": total,
            "data": [{
                "id": r.id,
                "spare_item_name": r.spare_item_name,
                "spare_item_code": r.spare_item_code or "",
                "quantity_required": r.quantity_required,
                "status_label": _status_label(r),
                "repair_route": r.repair_route,
                "sub_ticket_number": r.sub_ticket_number,
                "requested_at": r.requested_at.isoformat() if r.requested_at else None,
                "ticket_number": r.ticket_number,
                "customer_name": r.customer_name,
                "customer_phone": r.customer_phone,
                "vehicle_model": r.vehicle_model,
            } for r in rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
